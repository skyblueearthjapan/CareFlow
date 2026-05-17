"""完全置換インポート (バックアップ復元) — スタッフマスタ + 勤務シフト.

通常 import (``importer.py``) との違い:
  * Excel に無い既存 alive スタッフは **soft delete** (関連 shift は物理削除).
  * Excel の **空セル** は NULL で上書き (現在値維持ではなく).
  * `<DELETE>` / `<CLEAR>` フラグは使わない.
  * 全 shift を一度物理削除してから Excel 通りに再投入する.
  * **atomic transaction**: error 1 件でも全 rollback.
"""

from __future__ import annotations

import uuid
from datetime import datetime, time
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.office import Office
from app.models.staff import Staff, StaffShift
from app.schemas.v2.staff_excel import (
    ShiftExcelImportRow,
    StaffExcelChange,
    StaffExcelImportRow,
    StaffExcelReplaceAllSummary,
)
from app.services.staff_excel.schema import (
    ROLE_VALUES,
    SEX_VALUES,
    SHEET_SHIFT,
    SHEET_STAFF,
    SHIFT_COL_INDEX,
    STAFF_COL_INDEX,
    STATUS_VALUES,
    WEEKDAY_LABEL_TO_INT,
)

# ---------------------------------------------------------------------------
# cell helpers (importer の同等関数)
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _read_str(value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _read_uuid(value: Any) -> UUID | None:
    if _is_blank(value):
        return None
    if isinstance(value, UUID):
        return value
    return UUID(str(value).strip())


def _read_bool(value: Any) -> bool | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().upper()
    if s in ("TRUE", "1", "YES", "Y"):
        return True
    if s in ("FALSE", "0", "NO", "N"):
        return False
    raise ValueError(f"bool として読めません: {value!r}")


def _read_hhmm(value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    s = str(value).strip()
    parts = s.split(":")
    if len(parts) < 2:
        raise ValueError(f"HH:MM 形式ではありません: {s!r}")
    h = int(parts[0])
    m = int(parts[1])
    if not (0 <= h <= 23 and 0 <= m <= 59):
        raise ValueError(f"時刻の範囲外: {s!r}")
    return f"{h:02d}:{m:02d}"


def _hhmm_to_time(s: str) -> time:
    h, m = s.split(":")[:2]
    return time(int(h), int(m))


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return all(_is_blank(v) for v in row)


def _serializable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return value.isoformat()
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    return str(value)


# ---------------------------------------------------------------------------
# Staff row parser (replace-all)
# ---------------------------------------------------------------------------


def _parse_staff_row_replace(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_staff: dict[UUID, Staff],
    existing_staff_by_code: dict[str, Staff],
    deleted_staff_by_code: dict[str, Staff],
    offices_by_code: dict[str, Office],
    already_seen_codes: set[str],
    new_code_to_uuid: dict[str, UUID],
    resurrect_code_to_uuid: dict[str, UUID],
) -> tuple[StaffExcelImportRow, dict[str, Any] | None]:
    cells: dict[str, Any] = {}
    for col_key, idx in STAFF_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["staff_id"]
    raw_code = cells["staff_code"]

    try:
        staff_id = _read_uuid(raw_id) if not _is_blank(raw_id) else None
    except ValueError:
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=_read_str(raw_code),
                operation="error",
                error_message=f"staff_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    staff_code = _read_str(raw_code)

    existing_obj: Staff | None = None
    if staff_id is not None:
        existing_obj = existing_staff.get(staff_id)
        if existing_obj is None:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code,
                    operation="error",
                    error_message=f"staff_id が DB に存在しません: {staff_id}",
                ),
                None,
            )

    if existing_obj is None and staff_code is not None:
        existing_obj = existing_staff_by_code.get(staff_code)
        if existing_obj is not None:
            staff_id = existing_obj.id

    # ---- 列パース (空セル → None = NULL) ----
    parsed: dict[str, Any] = {}
    errors: list[str] = []

    for k in ("name", "kana", "note"):
        parsed[k] = _read_str(cells[k])

    parsed["code"] = staff_code

    def _parse_enum(key: str, allowed: tuple[str, ...]) -> None:
        v = cells[key]
        if _is_blank(v):
            parsed[key] = None
            return
        s = _read_str(v)
        if s not in allowed:
            errors.append(f"列「{key}」の値が候補外: {s!r} (許容: {','.join(allowed)})")
            return
        parsed[key] = s

    _parse_enum("sex", SEX_VALUES)
    _parse_enum("status", STATUS_VALUES)
    _parse_enum("role", ROLE_VALUES)

    # is_trainee (NOT NULL bool — 空セルは False)
    raw_trainee = cells["is_trainee"]
    if _is_blank(raw_trainee):
        parsed["is_trainee"] = False
    else:
        try:
            parsed["is_trainee"] = _read_bool(raw_trainee)
        except ValueError as exc:
            errors.append(f"列「is_trainee」が TRUE/FALSE ではありません: {exc}")

    raw_office = cells["office_code"]
    if _is_blank(raw_office):
        parsed["primary_office_id"] = None
    else:
        oc = _read_str(raw_office)
        office = offices_by_code.get(oc) if oc else None
        if office is None:
            errors.append(f"拠点コードが DB に存在しません: {oc!r}")
        else:
            parsed["primary_office_id"] = office.id

    if errors:
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code,
                operation="error",
                error_message=" / ".join(errors),
            ),
            None,
        )

    if staff_code is not None:
        if staff_code in already_seen_codes:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code,
                    operation="error",
                    error_message=f"staff_code が同ファイル内で重複しています: {staff_code!r}",
                ),
                None,
            )
        already_seen_codes.add(staff_code)

    field_map: dict[str, str] = {
        "name": "name",
        "kana": "kana",
        "sex": "sex",
        "status": "status",
        "role": "role",
        "primary_office_id": "primary_office_id",
        "is_trainee": "is_trainee",
        "note": "note",
    }

    if existing_obj is not None:
        # NOT NULL カラムが空セル (None) になっていないか確認
        required_update = ["name", "status", "role"]
        null_required = [k for k in required_update if parsed.get(k) is None]
        if null_required:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=existing_obj.id,
                    staff_code=existing_obj.code,
                    operation="error",
                    error_message=f"必須項目が空セルです (NULL 不可): {', '.join(null_required)}",
                ),
                None,
            )

        changes: list[StaffExcelChange] = []
        update_dict: dict[str, Any] = {"_staff_id": existing_obj.id, "_op": "update"}

        if staff_code is not None and staff_code != existing_obj.code:
            changes.append(
                StaffExcelChange(field="code", old_value=existing_obj.code, new_value=staff_code)
            )
            update_dict["code"] = staff_code

        for parsed_key, orm_attr in field_map.items():
            new_val = parsed.get(parsed_key)
            old_val = getattr(existing_obj, orm_attr, None)
            if old_val == new_val:
                continue
            changes.append(
                StaffExcelChange(
                    field=orm_attr,
                    old_value=_serializable(old_val),
                    new_value=_serializable(new_val),
                )
            )
            update_dict[orm_attr] = new_val

        if not changes:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=existing_obj.id,
                    staff_code=existing_obj.code,
                    operation="noop",
                ),
                None,
            )

        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=existing_obj.id,
                staff_code=existing_obj.code,
                operation="update",
                changes=changes,
            ),
            update_dict,
        )

    # 新規 / 復活
    missing: list[str] = []
    if not staff_code:
        missing.append("staff_code")
    for spec_key in ("name", "status", "role"):
        if parsed.get(spec_key) is None:
            missing.append(spec_key)
    if missing:
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=staff_code,
                operation="error",
                error_message=f"新規作成に必要な項目が不足: {', '.join(missing)}",
            ),
            None,
        )

    resurrect_target = deleted_staff_by_code.get(staff_code) if staff_code else None
    if resurrect_target is not None:
        resurrect_id = resurrect_target.id
        old_deleted_at = resurrect_target.deleted_at
        updates: dict[str, Any] = {}
        changes_for_view: list[StaffExcelChange] = [
            StaffExcelChange(
                field="deleted_at",
                old_value=_serializable(old_deleted_at),
                new_value=None,
            )
        ]
        for parsed_key, orm_attr in field_map.items():
            new_val = parsed.get(parsed_key)
            old_val = getattr(resurrect_target, orm_attr, None)
            if old_val == new_val:
                continue
            changes_for_view.append(
                StaffExcelChange(
                    field=orm_attr,
                    old_value=_serializable(old_val),
                    new_value=_serializable(new_val),
                )
            )
            updates[orm_attr] = new_val
        resurrect_code_to_uuid[staff_code] = resurrect_id  # type: ignore[index]
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=resurrect_id,
                staff_code=staff_code,
                operation="update",
                changes=changes_for_view,
            ),
            {"_op": "resurrect", "_staff_id": resurrect_id, "_updates": updates},
        )

    new_uuid = uuid.uuid4()
    new_code_to_uuid[staff_code] = new_uuid  # type: ignore[index]
    new_dict: dict[str, Any] = {
        "_op": "new",
        "_assigned_id": new_uuid,
        "id": new_uuid,
        "code": staff_code,
    }
    for parsed_key, orm_attr in field_map.items():
        val = parsed.get(parsed_key)
        if val is None:
            continue
        new_dict[orm_attr] = val
    changes_for_view = [
        StaffExcelChange(field=k, old_value=None, new_value=_serializable(new_dict[k]))
        for k in sorted(new_dict.keys())
        if not k.startswith("_") and k not in ("id",)
    ]
    return (
        StaffExcelImportRow(
            row_number=row_number,
            staff_id=new_uuid,
            staff_code=staff_code,
            operation="new",
            changes=changes_for_view,
        ),
        new_dict,
    )


# ---------------------------------------------------------------------------
# Shift row parser (replace-all: 全件 INSERT のみ)
# ---------------------------------------------------------------------------


def _parse_shift_row_replace(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_staff: dict[UUID, Staff],
    new_code_to_uuid: dict[str, UUID],
    existing_code_to_id: dict[str, UUID],
    resurrect_staff_ids: set[UUID],
    pending_new_keys: set[tuple[UUID, int]],
) -> tuple[ShiftExcelImportRow, dict[str, Any] | None]:
    cells: dict[str, Any] = {}
    for col_key, idx in SHIFT_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["staff_id"]
    raw_code = cells["staff_code"]

    staff_id: UUID | None = None
    try:
        if not _is_blank(raw_id):
            staff_id = _read_uuid(raw_id)
    except ValueError:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=_read_str(raw_code),
                operation="error",
                error_message=f"staff_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    shift_code = _read_str(raw_code)
    if staff_id is None:
        if not shift_code:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=None,
                    operation="error",
                    error_message="staff_id / staff_code のどちらも空です",
                ),
                None,
            )
        if shift_code in existing_code_to_id:
            staff_id = existing_code_to_id[shift_code]
        elif shift_code in new_code_to_uuid:
            staff_id = new_code_to_uuid[shift_code]
        else:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=shift_code,
                    operation="error",
                    error_message=f"staff_code が DB / 同ファイル内に存在しません: {shift_code!r}",
                ),
                None,
            )

    is_pending_new_staff = staff_id in set(new_code_to_uuid.values())
    is_resurrect_staff = staff_id in resurrect_staff_ids
    is_pending = is_pending_new_staff or is_resurrect_staff
    staff_obj: Staff | None = existing_staff.get(staff_id) if not is_pending else None
    if staff_obj is None and not is_pending:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=shift_code,
                operation="error",
                error_message=f"staff_id が DB に存在しません: {staff_id}",
            ),
            None,
        )

    staff_code_for_view = staff_obj.code if staff_obj is not None else shift_code

    raw_weekday = cells["weekday"]
    if _is_blank(raw_weekday):
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message="weekday が空です",
            ),
            None,
        )
    wd_str = _read_str(raw_weekday)
    weekday = WEEKDAY_LABEL_TO_INT.get(wd_str or "")
    if weekday is None:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message=f"weekday の値が候補外: {wd_str!r} (許容: 月火水木金土日)",
            ),
            None,
        )

    raw_is_on = cells["is_on"]
    if _is_blank(raw_is_on):
        # 完全置換セマンティクス: 空セル = 既定 (True)
        is_on = True
    else:
        try:
            parsed_on = _read_bool(raw_is_on)
        except ValueError as exc:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code_for_view,
                    weekday=weekday,
                    operation="error",
                    error_message=f"列「勤務」が TRUE/FALSE ではありません: {exc}",
                ),
                None,
            )
        is_on = parsed_on if parsed_on is not None else True

    try:
        start_str = _read_hhmm(cells["start_time"])
    except (ValueError, TypeError) as exc:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message=f"start_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )
    try:
        end_str = _read_hhmm(cells["end_time"])
    except (ValueError, TypeError) as exc:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message=f"end_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )

    if is_on and (start_str is None or end_str is None):
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message="勤務=TRUE のときは開始時刻と終了時刻が必須です",
            ),
            None,
        )

    key = (staff_id, weekday)
    if key in pending_new_keys:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message="同ファイル内で (staff_id, weekday) が重複しています",
            ),
            None,
        )
    pending_new_keys.add(key)

    new_start = _hhmm_to_time(start_str) if start_str is not None else None
    new_end = _hhmm_to_time(end_str) if end_str is not None else None

    new_dict: dict[str, Any] = {
        "_op": "new",
        "staff_id": staff_id,
        "weekday": weekday,
        "is_on": is_on,
        "start_time": new_start,
        "end_time": new_end,
    }
    return (
        ShiftExcelImportRow(
            row_number=row_number,
            staff_id=staff_id,
            staff_code=staff_code_for_view,
            weekday=weekday,
            operation="new",
            changes=[
                StaffExcelChange(field="is_on", old_value=None, new_value=is_on),
                StaffExcelChange(field="start_time", old_value=None, new_value=start_str),
                StaffExcelChange(field="end_time", old_value=None, new_value=end_str),
            ],
        ),
        new_dict,
    )


# ---------------------------------------------------------------------------
# public entrypoints
# ---------------------------------------------------------------------------


async def parse_and_diff_replace_all(
    db: AsyncSession,
    *,
    file_bytes: bytes,
) -> tuple[
    list[StaffExcelImportRow],
    list[ShiftExcelImportRow],
    StaffExcelReplaceAllSummary,
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],  # staff_to_soft_delete_preview
]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if SHEET_STAFF not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_STAFF}」が見つかりません")
    if SHEET_SHIFT not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_SHIFT}」が見つかりません")

    rows = (await db.scalars(select(Staff).where(Staff.deleted_at.is_(None)))).all()
    existing_staff: dict[UUID, Staff] = {s.id: s for s in rows}
    existing_staff_by_code: dict[str, Staff] = {s.code: s for s in rows if s.code}
    existing_code_to_id: dict[str, UUID] = {s.code: s.id for s in rows if s.code}

    deleted_rows = (
        await db.scalars(
            select(Staff).where(Staff.deleted_at.is_not(None), Staff.code.is_not(None))
        )
    ).all()
    deleted_staff_by_code: dict[str, Staff] = {s.code: s for s in deleted_rows}

    offices_rows = (
        await db.scalars(
            select(Office).where(Office.deleted_at.is_(None), Office.code.is_not(None))
        )
    ).all()
    offices_by_code: dict[str, Office] = {str(o.code): o for o in offices_rows}

    # ---- staff sheet ----
    ws_s = wb[SHEET_STAFF]
    staff_rows: list[StaffExcelImportRow] = []
    staff_ops: list[dict[str, Any]] = []
    already_seen_codes: set[str] = set()
    new_code_to_uuid: dict[str, UUID] = {}
    resurrect_code_to_uuid: dict[str, UUID] = {}
    excel_staff_ids_seen: set[UUID] = set()

    for r_idx, row in enumerate(ws_s.iter_rows(min_row=2, values_only=True), start=2):
        if _row_is_empty(row):
            continue
        diff_row, op = _parse_staff_row_replace(
            r_idx,
            row,
            existing_staff=existing_staff,
            existing_staff_by_code=existing_staff_by_code,
            deleted_staff_by_code=deleted_staff_by_code,
            offices_by_code=offices_by_code,
            already_seen_codes=already_seen_codes,
            new_code_to_uuid=new_code_to_uuid,
            resurrect_code_to_uuid=resurrect_code_to_uuid,
        )
        staff_rows.append(diff_row)
        if op is not None:
            staff_ops.append(op)
        if diff_row.operation in ("update", "noop") and diff_row.staff_id is not None:
            excel_staff_ids_seen.add(diff_row.staff_id)

    staff_to_soft_delete: list[Staff] = [
        s for sid, s in existing_staff.items() if sid not in excel_staff_ids_seen
    ]
    staff_to_soft_delete_preview: list[dict[str, Any]] = [
        {"staff_id": str(s.id), "staff_code": s.code, "name": s.name} for s in staff_to_soft_delete
    ]

    existing_code_to_id_with_resurrect: dict[str, UUID] = {
        **{c: i for c, i in existing_code_to_id.items() if i in excel_staff_ids_seen},
        **resurrect_code_to_uuid,
    }

    # ---- shift sheet ----
    ws_f = wb[SHEET_SHIFT]
    shift_rows: list[ShiftExcelImportRow] = []
    shift_ops: list[dict[str, Any]] = []
    pending_new_keys: set[tuple[UUID, int]] = set()
    for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
        if _row_is_empty(row):
            continue
        diff_row, op = _parse_shift_row_replace(
            r_idx,
            row,
            existing_staff=existing_staff,
            new_code_to_uuid=new_code_to_uuid,
            existing_code_to_id=existing_code_to_id_with_resurrect,
            resurrect_staff_ids=set(resurrect_code_to_uuid.values()),
            pending_new_keys=pending_new_keys,
        )
        shift_rows.append(diff_row)
        if op is not None:
            shift_ops.append(op)

    # 既存 shift count (replace count) — alive staff の shift のみ count
    shift_count = (
        await db.execute(
            select(func.count())
            .select_from(StaffShift)
            .where(StaffShift.staff_id.in_(select(Staff.id).where(Staff.deleted_at.is_(None))))
        )
    ).scalar_one()

    summary = StaffExcelReplaceAllSummary()
    for r in staff_rows:
        match r.operation:
            case "new":
                summary.staff_to_create += 1
            case "update":
                summary.staff_to_update += 1
            case "error":
                summary.staff_error += 1
    summary.staff_to_soft_delete = len(staff_to_soft_delete)
    for r in shift_rows:
        match r.operation:
            case "new":
                summary.shift_to_create += 1
            case "error":
                summary.shift_error += 1
    summary.shift_to_replace = int(shift_count or 0)

    return (
        staff_rows,
        shift_rows,
        summary,
        staff_ops,
        shift_ops,
        staff_to_soft_delete_preview,
    )


async def apply_replace_all(
    db: AsyncSession,
    *,
    staff_ops: list[dict[str, Any]],
    shift_ops: list[dict[str, Any]],
    staff_ids_to_soft_delete: list[UUID],
) -> None:
    """完全置換を DB に反映 (atomic).

    順序:
      1. 削除対象 staff の shift を物理削除
      2. 削除対象 staff を soft delete
      3. 残り alive staff の shift を全件物理削除
      4. staff_ops apply
      5. shift_ops INSERT
    """
    if staff_ids_to_soft_delete:
        shifts_for_delete = (
            await db.scalars(
                select(StaffShift).where(StaffShift.staff_id.in_(staff_ids_to_soft_delete))
            )
        ).all()
        for sh in shifts_for_delete:
            await db.delete(sh)
        await db.flush()
        for sid in staff_ids_to_soft_delete:
            staff_obj = await db.get(Staff, sid)
            if staff_obj is not None:
                staff_obj.deleted_at = func.now()
        await db.flush()

    alive_staff_ids = (await db.scalars(select(Staff.id).where(Staff.deleted_at.is_(None)))).all()
    if alive_staff_ids:
        rest_shifts = (
            await db.scalars(
                select(StaffShift).where(StaffShift.staff_id.in_(list(alive_staff_ids)))
            )
        ).all()
        for sh in rest_shifts:
            await db.delete(sh)
        await db.flush()

    staff_by_id: dict[UUID, Staff] = {}
    for op in staff_ops:
        op_type = op.get("_op")
        if op_type == "new":
            data = {k: v for k, v in op.items() if not k.startswith("_")}
            db.add(Staff(**data))
        elif op_type == "resurrect":
            sid = op["_staff_id"]
            if sid not in staff_by_id:
                staff_by_id[sid] = await db.get(Staff, sid)
            staff_obj = staff_by_id[sid]
            if staff_obj is None:
                continue
            staff_obj.deleted_at = None
            for k, v in op["_updates"].items():
                setattr(staff_obj, k, v)
        elif op_type == "update":
            sid = op["_staff_id"]
            if sid not in staff_by_id:
                staff_by_id[sid] = await db.get(Staff, sid)
            staff_obj = staff_by_id[sid]
            if staff_obj is None:
                continue
            for k, v in op.items():
                if k.startswith("_"):
                    continue
                setattr(staff_obj, k, v)
    await db.flush()

    for op in shift_ops:
        if op.get("_op") != "new":
            continue
        data = {k: v for k, v in op.items() if not k.startswith("_")}
        db.add(StaffShift(**data))
    await db.flush()
