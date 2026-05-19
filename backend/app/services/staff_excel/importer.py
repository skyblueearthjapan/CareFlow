"""Excel パース + 差分計算 + DB 反映.

エントリポイント:
  * ``parse_and_diff`` — Excel bytes → (スタッフ行差分, 勤務シフト行差分, summary).
    プレビュー (dry_run=True) でも apply (dry_run=False) でも事前にこの関数で
    差分計算する.
  * ``apply_changes`` — ``parse_and_diff`` の戻り値を受け取り、DB に書き戻す.
    1 transaction で INSERT / UPDATE / DELETE をまとめて実行 (partial commit).
    error 行は ``parse_and_diff`` の時点で ``op=None`` として除外されているので
    ``apply_changes`` には積まれず、自動的に skip される.

差分の表現:
  * ``StaffExcelImportRow`` — staff sheet の 1 行に対応.
  * ``ShiftExcelImportRow`` — shift sheet の 1 行に対応.
  * いずれも ``operation`` が "new" / "update" / "delete" / "noop" / "error".

# UUID 自動発番 (patient_excel パターン踏襲)

``staff_id`` が空で ``staff_code`` が記入されている行は「新規スタッフ候補」と
扱う。``parse_and_diff`` の中で ``staff_code → 仮 UUID`` のマップを構築し、
勤務シフトシートで同じ ``staff_code`` を参照されたら同じ仮 UUID を割り当てる。
apply 時に ``Staff(id=仮 UUID, ...)`` を INSERT すれば一貫した参照が貼れる。
"""

from __future__ import annotations

import logging
import uuid
from datetime import datetime, time
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.office import Office
from app.models.staff import Staff, StaffSecondaryOffice, StaffShift, StaffWeeklyOverride
from app.schemas.v2.staff_excel import (
    OverrideExcelImportRow,
    ShiftExcelImportRow,
    StaffExcelChange,
    StaffExcelImportRow,
    StaffExcelImportSummary,
)
from app.services.staff_excel.schema import (
    OVERRIDE_COL_INDEX,
    OVERRIDE_TYPE_VALUES,
    ROLE_VALUES,
    SEX_VALUES,
    SHEET_OVERRIDE,
    SHEET_SHIFT,
    SHEET_STAFF,
    SHIFT_COL_INDEX,
    STAFF_COL_INDEX,
    STATUS_VALUES,
    WEEKDAY_LABEL_TO_INT,
    is_magic_clear,
    is_magic_delete,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# cell value helpers
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    """空セルの判定. openpyxl は空セルを None で返すが、空文字 / 空白文字も空扱い."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _read_str(value: Any) -> str | None:
    """セル値を str に正規化する (空セルは None)."""
    if _is_blank(value):
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _read_uuid(value: Any) -> UUID | None:
    if _is_blank(value):
        return None
    if isinstance(value, UUID):
        return value
    return UUID(str(value).strip())


def _read_bool(value: Any) -> bool | None:
    """TRUE/FALSE (case-insensitive) を bool に正規化. 失敗時は ValueError."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().upper()
    if s in ("TRUE", "1", "YES", "Y"):
        return True
    if s in ("FALSE", "0", "NO", "N"):
        return False
    raise ValueError(f"bool として読めません: {value!r}")


def _read_int(value: Any) -> int | None:
    """セル値を int に正規化. 失敗時は ValueError. Phase E-7 (gap P1) Override 用."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        raise ValueError(f"int 型として読めません: {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"int 型として読めません (小数): {value!r}")
    s = str(value).strip()
    return int(s)


def _read_hhmm(value: Any) -> str | None:
    """セル値を "HH:MM" 文字列として正規化. 失敗時は ValueError."""
    if _is_blank(value):
        return None
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    s = str(value).strip()
    # ``HH:MM`` または ``HH:MM:SS``
    parts = s.split(":")
    if len(parts) < 2:
        raise ValueError(f"HH:MM 形式ではありません: {s!r}")
    h = int(parts[0])
    m = int(parts[1])
    if not (0 <= h <= 23 and 0 <= m <= 59):
        raise ValueError(f"時刻の範囲外: {s!r}")
    return f"{h:02d}:{m:02d}"


def _hhmm_to_time(s: str) -> time:
    h, m = s.split(":")[:2]
    return time(int(h), int(m))


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return all(_is_blank(v) for v in row)


# ---------------------------------------------------------------------------
# lookup builders
# ---------------------------------------------------------------------------


async def _load_offices_by_code(db: AsyncSession) -> dict[str, Office]:
    """code → Office (deleted_at IS NULL のみ)."""
    rows = (
        await db.scalars(
            select(Office).where(
                Office.deleted_at.is_(None),
                Office.code.is_not(None),
            )
        )
    ).all()
    return {str(o.code): o for o in rows}


async def _load_staff_by_id(db: AsyncSession) -> dict[UUID, Staff]:
    rows = (await db.scalars(select(Staff).where(Staff.deleted_at.is_(None)))).all()
    return {s.id: s for s in rows}


async def _load_staff_codes(db: AsyncSession) -> set[str]:
    """既存 (alive) の staff.code 集合 (重複チェック用)."""
    rows = (
        await db.execute(
            select(Staff.code).where(Staff.deleted_at.is_(None), Staff.code.is_not(None))
        )
    ).all()
    return {r[0] for r in rows}


async def _load_deleted_staff_by_code(db: AsyncSession) -> dict[str, Staff]:
    """soft-deleted スタッフの code → Staff マップ (resurrection 用).

    DB は ``code`` 列に UNIQUE 制約があるため、同じ code を再 INSERT すると
    UniqueViolation で 409 になる。新規候補行 (staff_id 空 + staff_code 入り) が
    DB に soft-deleted で存在する code なら、INSERT ではなく既存行の
    deleted_at を NULL に戻して内容を上書きする (= 復活).
    """
    rows = (
        await db.scalars(
            select(Staff).where(
                Staff.deleted_at.is_not(None),
                Staff.code.is_not(None),
            )
        )
    ).all()
    return {s.code: s for s in rows}


async def _load_shifts_by_key(
    db: AsyncSession,
    alive_staff_ids: set[UUID],
) -> dict[tuple[UUID, int], StaffShift]:
    """(staff_id, weekday) → StaffShift.

    alive な staff の shift のみ load する.
    """
    if not alive_staff_ids:
        return {}
    rows = (
        await db.scalars(select(StaffShift).where(StaffShift.staff_id.in_(alive_staff_ids)))
    ).all()
    return {(s.staff_id, s.weekday): s for s in rows}


async def _load_secondary_offices_by_staff(
    db: AsyncSession,
    alive_staff_ids: set[UUID],
) -> dict[UUID, list[UUID]]:
    """Phase E-7 (gap P0-2): Staff.id → [Office.id, ...] のマップ.

    Excel 往復で secondary_offices relationship を扱うため、
    既存値を bulk SELECT で取得する.
    """
    if not alive_staff_ids:
        return {}
    # Phase E-7 (gap LOW#5): order_by office_id で deterministic な順序を保つ.
    # round-trip 順序の安定化のため.
    rows = (
        await db.scalars(
            select(StaffSecondaryOffice)
            .where(StaffSecondaryOffice.staff_id.in_(alive_staff_ids))
            .order_by(StaffSecondaryOffice.office_id)
        )
    ).all()
    result: dict[UUID, list[UUID]] = {}
    for r in rows:
        result.setdefault(r.staff_id, []).append(r.office_id)
    return result


async def _load_overrides_by_key(
    db: AsyncSession,
    alive_staff_ids: set[UUID],
) -> dict[tuple[UUID, int, int, int], StaffWeeklyOverride]:
    """Phase E-7 (gap P1): (staff_id, iso_year, iso_week, weekday) → StaffWeeklyOverride."""
    if not alive_staff_ids:
        return {}
    rows = (
        await db.scalars(
            select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id.in_(alive_staff_ids))
        )
    ).all()
    return {(o.staff_id, o.iso_year, o.iso_week, o.weekday): o for o in rows}


# ---------------------------------------------------------------------------
# Staff sheet parser
# ---------------------------------------------------------------------------


def _parse_staff_row(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_staff: dict[UUID, Staff],
    existing_staff_by_code: dict[str, Staff],
    deleted_staff_by_code: dict[str, Staff],
    offices_by_code: dict[str, Office],
    existing_codes: set[str],
    already_seen_codes: set[str],
    new_code_to_uuid: dict[str, UUID],
    resurrect_code_to_uuid: dict[str, UUID],
    existing_secondary_offices: dict[UUID, list[UUID]] | None = None,
) -> tuple[StaffExcelImportRow, dict[str, Any] | None]:
    """1 行を差分行に変換する.

    戻り値の第 2 要素は apply 時に使う「DB 用 dict」. operation が "error" / "noop" /
    "delete" のときは None or 部分的に意味のあるもの.
    """
    cells: dict[str, Any] = {}
    for col_key, idx in STAFF_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    # 共通: staff_id / staff_code を読む.
    raw_id = cells["staff_id"]
    raw_code = cells["staff_code"]
    raw_delete = cells["delete_flag"]

    # staff_id がある場合は UUID パース.
    try:
        staff_id = _read_uuid(raw_id) if not _is_blank(raw_id) else None
    except ValueError:
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=_read_str(raw_code),
                operation="error",
                error_message=f"staff_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    staff_code = _read_str(raw_code)

    # 既存スタッフの検索.
    existing_obj: Staff | None = None
    if staff_id is not None:
        existing_obj = existing_staff.get(staff_id)
        if existing_obj is None:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code,
                    operation="error",
                    error_message=f"staff_id が DB に存在しません: {staff_id}",
                ),
                None,
            )

    # 削除フラグ
    if is_magic_delete(raw_delete):
        # 1) staff_id があれば既存通り (上で resolution 済み)
        if existing_obj is not None:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=existing_obj.id,
                    staff_code=existing_obj.code,
                    operation="delete",
                ),
                {"_staff_id": existing_obj.id, "_op": "delete"},
            )
        # 2) staff_id 空でも staff_code があれば code で解決
        if staff_code is not None:
            resolved = existing_staff_by_code.get(staff_code)
            if resolved is None:
                # 該当 code が DB に居ない → idempotent な noop 扱い (再 import で
                # error にしないため). export → 1 行削除 → 再 import のようなケースを
                # 想定。
                return (
                    StaffExcelImportRow(
                        row_number=row_number,
                        staff_id=None,
                        staff_code=staff_code,
                        operation="noop",
                    ),
                    None,
                )
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=resolved.id,
                    staff_code=resolved.code,
                    operation="delete",
                ),
                {"_staff_id": resolved.id, "_op": "delete"},
            )
        # 3) 両方空 → error
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=None,
                operation="error",
                error_message=("削除フラグが指定されましたが staff_id / staff_code がありません"),
            ),
            None,
        )

    # 各列をパース.
    parsed: dict[str, Any] = {}
    errors: list[str] = []

    # name / kana / note: 文字列系
    for k in ("name", "kana", "note"):
        v = cells[k]
        if _is_blank(v):
            parsed[k] = None
        elif is_magic_clear(v):
            parsed[k] = ("CLEAR", None)
        else:
            parsed[k] = ("SET", _read_str(v))

    # staff_code (現在値と diff したいので raw を保持)
    if _is_blank(raw_code):
        parsed["code"] = None
    else:
        parsed["code"] = ("SET", staff_code)

    # enum 系
    def _parse_enum(key: str, allowed: tuple[str, ...]) -> None:
        v = cells[key]
        if _is_blank(v):
            parsed[key] = None
            return
        if is_magic_clear(v):
            parsed[key] = ("CLEAR", None)
            return
        s = _read_str(v)
        if s not in allowed:
            errors.append(f"列「{key}」の値が候補外: {s!r} (許容: {','.join(allowed)})")
            return
        parsed[key] = ("SET", s)

    _parse_enum("sex", SEX_VALUES)
    _parse_enum("status", STATUS_VALUES)
    _parse_enum("role", ROLE_VALUES)

    # is_trainee (bool)
    raw_trainee = cells["is_trainee"]
    if _is_blank(raw_trainee):
        parsed["is_trainee"] = None
    elif is_magic_clear(raw_trainee):
        # is_trainee は NOT NULL なので CLEAR は False と等価扱い
        parsed["is_trainee"] = ("SET", False)
    else:
        try:
            b = _read_bool(raw_trainee)
        except ValueError as exc:
            errors.append(f"列「is_trainee」が TRUE/FALSE ではありません: {exc}")
        else:
            parsed["is_trainee"] = ("SET", b)

    # 拠点コード → primary_office_id
    raw_office = cells["office_code"]
    if _is_blank(raw_office):
        parsed["primary_office_id"] = None
    elif is_magic_clear(raw_office):
        parsed["primary_office_id"] = ("CLEAR", None)
    else:
        oc = _read_str(raw_office)
        office = offices_by_code.get(oc) if oc else None
        if office is None:
            errors.append(f"拠点コードが DB に存在しません: {oc!r}")
        else:
            parsed["primary_office_id"] = ("SET", office.id)

    # Phase E-7 (gap P0-2): secondary_office_codes (comma-separated) → list[Office.id].
    # 空セル: 触らない (更新時) / 関連無し (新規時).
    # <CLEAR>: 関連解除.
    # 不明 code は warning + skip (= error にせず PFV course_template と同じ
    # ベストエフォート方針, バックアップ運用優先).
    raw_secondary = cells["secondary_office_codes"]
    secondary_warnings: list[str] = []
    if _is_blank(raw_secondary):
        parsed["secondary_office_ids"] = None
    elif is_magic_clear(raw_secondary):
        parsed["secondary_office_ids"] = ("SET", [])
    else:
        sec_str = _read_str(raw_secondary) or ""
        sec_codes = [c.strip() for c in sec_str.split(",") if c.strip()]
        sec_ids: list[UUID] = []
        seen_ids: set[UUID] = set()
        for sc in sec_codes:
            sec_office = offices_by_code.get(sc)
            if sec_office is None:
                secondary_warnings.append(sc)
                continue
            if sec_office.id in seen_ids:
                continue
            seen_ids.add(sec_office.id)
            sec_ids.append(sec_office.id)
        parsed["secondary_office_ids"] = ("SET", sec_ids)
        if secondary_warnings:
            logger.warning(
                "staff_excel import row=%d: secondary_office 不明コード %r を skip しました",
                row_number,
                secondary_warnings,
            )

    if errors:
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code,
                operation="error",
                error_message=" / ".join(errors),
            ),
            None,
        )

    # ---- 操作判定 ----
    if existing_obj is None:
        # 新規候補
        missing: list[str] = []
        # patient_excel と同じく「必須項目チェック」
        if not staff_code:
            missing.append("staff_code")
        for spec_key in ("name", "status", "role"):
            v = parsed.get(spec_key)
            if v is None or (isinstance(v, tuple) and v[0] == "CLEAR"):
                missing.append(spec_key)
        if missing:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=staff_code,
                    operation="error",
                    error_message=f"新規作成に必要な項目が不足: {', '.join(missing)}",
                ),
                None,
            )
        # 重複チェック (DB の他レコード + 同ファイル内の他行).
        if staff_code in existing_codes:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=staff_code,
                    operation="error",
                    error_message=f"staff_code が既に存在します: {staff_code!r}",
                ),
                None,
            )
        if staff_code in already_seen_codes:
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=staff_code,
                    operation="error",
                    error_message=(f"staff_code が同ファイル内で重複しています: {staff_code!r}"),
                ),
                None,
            )
        already_seen_codes.add(staff_code)  # type: ignore[arg-type]

        # ---- resurrection 判定 ----
        # DB に同じ code が soft-deleted で残っていれば INSERT すると UNIQUE 制約違反.
        # 新規 INSERT ではなく既存行の deleted_at を NULL に戻して内容を上書きする.
        # staff.id は元の UUID をそのまま使う (再発番しない). shifts は staff
        # soft-delete 時に cascade で物理削除済みなので、resurrection 時点での
        # 残骸処理は不要.
        resurrect_target = deleted_staff_by_code.get(staff_code)
        if resurrect_target is not None:
            resurrect_id = resurrect_target.id
            old_deleted_at = resurrect_target.deleted_at
            updates: dict[str, Any] = {}
            changes_for_view: list[StaffExcelChange] = []
            changes_for_view.append(
                StaffExcelChange(
                    field="deleted_at",
                    old_value=_serializable(old_deleted_at),
                    new_value=None,
                )
            )
            field_map_resurrect: dict[str, str] = {
                "name": "name",
                "kana": "kana",
                "sex": "sex",
                "status": "status",
                "role": "role",
                "primary_office_id": "primary_office_id",
                "is_trainee": "is_trainee",
                "note": "note",
            }
            for parsed_key, orm_attr in field_map_resurrect.items():
                v = parsed.get(parsed_key)
                if v is None:
                    continue  # 空セル = 旧値維持
                tag, val = v if isinstance(v, tuple) else ("SET", v)
                new_val: Any = None if tag == "CLEAR" else val
                old_val = getattr(resurrect_target, orm_attr, None)
                if old_val == new_val:
                    continue
                changes_for_view.append(
                    StaffExcelChange(
                        field=orm_attr,
                        old_value=_serializable(old_val),
                        new_value=_serializable(new_val),
                    )
                )
                updates[orm_attr] = new_val
            # Phase E-7 (gap P0-2): secondary_offices は relationship なので別 key で.
            sec_op = parsed.get("secondary_office_ids")
            secondary_op_dict: dict[str, Any] = {}
            if sec_op is not None:
                _, new_sec_ids = sec_op
                old_sec_ids = (
                    existing_secondary_offices.get(resurrect_id, [])
                    if existing_secondary_offices is not None
                    else []
                )
                if sorted(old_sec_ids) != sorted(new_sec_ids):
                    changes_for_view.append(
                        StaffExcelChange(
                            field="secondary_offices",
                            old_value=[str(i) for i in old_sec_ids],
                            new_value=[str(i) for i in new_sec_ids],
                        )
                    )
                    secondary_op_dict["_secondary_office_ids"] = new_sec_ids
            # shift シートが staff_code で参照できるよう、復活 UUID を登録.
            resurrect_code_to_uuid[staff_code] = resurrect_id
            return (
                StaffExcelImportRow(
                    row_number=row_number,
                    staff_id=resurrect_id,
                    staff_code=staff_code,
                    operation="update",
                    changes=changes_for_view,
                ),
                {
                    "_op": "resurrect",
                    "_staff_id": resurrect_id,
                    "_updates": updates,
                    **secondary_op_dict,
                },
            )

        # 仮 UUID を発番 (apply 時にそのまま id 列として INSERT).
        # shift シート側で staff_code 経由で参照されたら同じ UUID を使う.
        new_uuid = uuid.uuid4()
        new_code_to_uuid[staff_code] = new_uuid

        new_dict: dict[str, Any] = {
            "_op": "new",
            "_assigned_id": new_uuid,
            "id": new_uuid,
            "code": staff_code,
        }
        # 必須以外も含めて値をフラット化.
        for k, v in parsed.items():
            if k in ("code",):
                continue
            # Phase E-7 (gap P0-2): secondary_office_ids は relationship なので
            # Staff(**data) には渡さず、apply 時に StaffSecondaryOffice を別途
            # INSERT する. ここでは _secondary_office_ids として外出し.
            if k == "secondary_office_ids":
                if v is None:
                    continue
                _, sec_ids = v
                if sec_ids:
                    new_dict["_secondary_office_ids"] = sec_ids
                continue
            if v is None:
                # 空セル: 触らない (新規時は default に任せる).
                continue
            if isinstance(v, tuple):
                tag, val = v
                # CLEAR は新規時は単に「None で INSERT」と等価.
                new_dict[k] = val
        changes_for_view = [
            StaffExcelChange(field=k, old_value=None, new_value=_serializable(new_dict[k]))
            for k in sorted(new_dict.keys())
            if not k.startswith("_") and k not in ("code", "id")
        ]
        # secondary_offices を新規時に表示用 changes に含める.
        if "_secondary_office_ids" in new_dict:
            changes_for_view.append(
                StaffExcelChange(
                    field="secondary_offices",
                    old_value=None,
                    new_value=[str(i) for i in new_dict["_secondary_office_ids"]],
                )
            )
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=new_uuid,
                staff_code=staff_code,
                operation="new",
                changes=changes_for_view,
            ),
            new_dict,
        )

    # 既存スタッフの update / noop.
    changes: list[StaffExcelChange] = []
    update_dict: dict[str, Any] = {"_staff_id": existing_obj.id, "_op": "update"}

    # staff_code は通常変更しない方針だが、明示的に書いてあって既存値と違えば update.
    if parsed.get("code") is not None:
        _, new_code = parsed["code"]
        if new_code != existing_obj.code:
            # 重複チェック: DB の他レコード + 同ファイル内の他行 (自分以外).
            if new_code in (existing_codes | already_seen_codes) - {existing_obj.code}:
                return (
                    StaffExcelImportRow(
                        row_number=row_number,
                        staff_id=existing_obj.id,
                        staff_code=staff_code,
                        operation="error",
                        error_message=f"staff_code が既に存在します: {new_code!r}",
                    ),
                    None,
                )
            already_seen_codes.add(new_code)
            changes.append(
                StaffExcelChange(
                    field="code",
                    old_value=existing_obj.code,
                    new_value=new_code,
                )
            )
            update_dict["code"] = new_code

    # 他フィールド.
    field_map: dict[str, str] = {
        "name": "name",
        "kana": "kana",
        "sex": "sex",
        "status": "status",
        "role": "role",
        "primary_office_id": "primary_office_id",
        "is_trainee": "is_trainee",
        "note": "note",
    }
    for parsed_key, orm_attr in field_map.items():
        v = parsed.get(parsed_key)
        if v is None:
            continue  # 空セル = 触らない
        tag, val = v if isinstance(v, tuple) else ("SET", v)
        if tag == "CLEAR":
            new_val: Any = None
        else:
            new_val = val
        old_val = getattr(existing_obj, orm_attr, None)
        if old_val == new_val:
            continue
        changes.append(
            StaffExcelChange(
                field=orm_attr,
                old_value=_serializable(old_val),
                new_value=_serializable(new_val),
            )
        )
        update_dict[orm_attr] = new_val

    # Phase E-7 (gap P0-2): secondary_offices (relationship) の差分.
    sec_op = parsed.get("secondary_office_ids")
    if sec_op is not None:
        _, new_sec_ids = sec_op
        old_sec_ids = (
            existing_secondary_offices.get(existing_obj.id, [])
            if existing_secondary_offices is not None
            else []
        )
        if sorted(old_sec_ids) != sorted(new_sec_ids):
            changes.append(
                StaffExcelChange(
                    field="secondary_offices",
                    old_value=[str(i) for i in old_sec_ids],
                    new_value=[str(i) for i in new_sec_ids],
                )
            )
            update_dict["_secondary_office_ids"] = new_sec_ids

    if not changes:
        return (
            StaffExcelImportRow(
                row_number=row_number,
                staff_id=existing_obj.id,
                staff_code=existing_obj.code,
                operation="noop",
            ),
            None,
        )

    return (
        StaffExcelImportRow(
            row_number=row_number,
            staff_id=existing_obj.id,
            staff_code=existing_obj.code,
            operation="update",
            changes=changes,
        ),
        update_dict,
    )


def _serializable(value: Any) -> Any:
    """Pydantic Response 用に安全な型へ変換 (UUID/Decimal/time など)."""
    if value is None:
        return None
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return value.isoformat()
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    return str(value)


# ---------------------------------------------------------------------------
# Shift sheet parser
# ---------------------------------------------------------------------------


def _parse_shift_row(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_staff: dict[UUID, Staff],
    existing_shifts: dict[tuple[UUID, int], StaffShift],
    pending_new_keys: set[tuple[UUID, int]],
    new_code_to_uuid: dict[str, UUID],
    existing_code_to_id: dict[str, UUID],
    resurrect_staff_ids: set[UUID] | None = None,
) -> tuple[ShiftExcelImportRow, dict[str, Any] | None]:
    cells: dict[str, Any] = {}
    for col_key, idx in SHIFT_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["staff_id"]
    raw_code = cells["staff_code"]
    raw_delete = cells["delete_flag"]

    # staff_id の解決: id 列があればそれを優先. なければ code から解決
    # (DB 既存 staff + 同 import 内で新規発番された staff の両方を見る).
    staff_id: UUID | None = None
    try:
        if not _is_blank(raw_id):
            staff_id = _read_uuid(raw_id)
    except ValueError:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=_read_str(raw_code),
                operation="error",
                error_message=f"staff_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    shift_code = _read_str(raw_code)
    if staff_id is None:
        # code リンク
        if not shift_code:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=None,
                    operation="error",
                    error_message="staff_id / staff_code のどちらも空です",
                ),
                None,
            )
        # 1) 既存 alive staff から探す
        if shift_code in existing_code_to_id:
            staff_id = existing_code_to_id[shift_code]
        # 2) 同 import 内で新規発番された staff から探す
        elif shift_code in new_code_to_uuid:
            staff_id = new_code_to_uuid[shift_code]
        else:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=shift_code,
                    operation="error",
                    error_message=f"staff_code が DB / 同ファイル内に存在しません: {shift_code!r}",
                ),
                None,
            )

    # staff_id が new_code_to_uuid 由来の場合、existing_staff には居ない (apply 時に作る).
    # resurrect_staff_ids も同様 (soft-deleted のため _load_staff_by_id では拾わない).
    is_pending_new_staff = staff_id in {v for v in new_code_to_uuid.values()}
    is_resurrect_staff = (resurrect_staff_ids is not None) and (staff_id in resurrect_staff_ids)
    is_pending = is_pending_new_staff or is_resurrect_staff
    staff_obj: Staff | None = existing_staff.get(staff_id) if not is_pending else None
    if staff_obj is None and not is_pending:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=shift_code,
                operation="error",
                error_message=f"staff_id が DB に存在しません: {staff_id}",
            ),
            None,
        )

    staff_code_for_view = staff_obj.code if staff_obj is not None else shift_code

    # weekday
    raw_weekday = cells["weekday"]
    if _is_blank(raw_weekday):
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message="weekday が空です",
            ),
            None,
        )
    wd_str = _read_str(raw_weekday)
    weekday = WEEKDAY_LABEL_TO_INT.get(wd_str or "")
    if weekday is None:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message=f"weekday の値が候補外: {wd_str!r} (許容: 月火水木金土日)",
            ),
            None,
        )

    key = (staff_id, weekday)
    existing_shift = existing_shifts.get(key)

    # 削除フラグ (shift は物理削除)
    if is_magic_delete(raw_delete):
        if existing_shift is None:
            # 存在しないものを delete しようとした → noop 扱い
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code_for_view,
                    weekday=weekday,
                    operation="noop",
                ),
                None,
            )
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="delete",
            ),
            {"_shift_key": key, "_op": "delete"},
        )

    # is_on
    raw_is_on = cells["is_on"]
    if _is_blank(raw_is_on):
        # 既存があれば現在値を継承、無ければ True
        is_on = existing_shift.is_on if existing_shift is not None else True
    else:
        try:
            parsed_on = _read_bool(raw_is_on)
        except ValueError as exc:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code_for_view,
                    weekday=weekday,
                    operation="error",
                    error_message=f"列「勤務」が TRUE/FALSE ではありません: {exc}",
                ),
                None,
            )
        is_on = parsed_on if parsed_on is not None else True

    # start_time / end_time
    try:
        start_str = _read_hhmm(cells["start_time"])
    except (ValueError, TypeError) as exc:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message=f"start_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )
    try:
        end_str = _read_hhmm(cells["end_time"])
    except (ValueError, TypeError) as exc:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message=f"end_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )

    # 「空セル = 変更しない」契約: 既存 shift の上書きパスでは blank cell を
    # 既存値で fallback してから validation / diff を行う.
    # (新規 shift の場合は existing_shift が None なので fallback は発生しない)
    effective_start_str = start_str
    effective_end_str = end_str
    if existing_shift is not None:
        if effective_start_str is None and existing_shift.start_time is not None:
            effective_start_str = (
                f"{existing_shift.start_time.hour:02d}:{existing_shift.start_time.minute:02d}"
            )
        if effective_end_str is None and existing_shift.end_time is not None:
            effective_end_str = (
                f"{existing_shift.end_time.hour:02d}:{existing_shift.end_time.minute:02d}"
            )

    # 勤務=TRUE 時は start / end 必須.
    if is_on:
        if effective_start_str is None or effective_end_str is None:
            return (
                ShiftExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code_for_view,
                    weekday=weekday,
                    operation="error",
                    error_message="勤務=TRUE のときは開始時刻と終了時刻が必須です",
                ),
                None,
            )

    # ---- ファイル内 (staff_id, weekday) 重複検査 ----
    if key in pending_new_keys:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="error",
                error_message="同ファイル内で (staff_id, weekday) が重複しています",
            ),
            None,
        )

    new_start = _hhmm_to_time(effective_start_str) if effective_start_str is not None else None
    new_end = _hhmm_to_time(effective_end_str) if effective_end_str is not None else None

    if existing_shift is None:
        pending_new_keys.add(key)
        new_dict: dict[str, Any] = {
            "_op": "new",
            "staff_id": staff_id,
            "weekday": weekday,
            "is_on": is_on,
            "start_time": new_start,
            "end_time": new_end,
        }
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="new",
                changes=[
                    StaffExcelChange(field="is_on", old_value=None, new_value=is_on),
                    StaffExcelChange(field="start_time", old_value=None, new_value=start_str),
                    StaffExcelChange(field="end_time", old_value=None, new_value=end_str),
                ],
            ),
            new_dict,
        )

    # update / noop. blank cell は既存値を継承する契約なので effective 値ベースで diff.
    changes: list[StaffExcelChange] = []
    update_dict: dict[str, Any] = {"_shift_key": key, "_op": "update"}
    if existing_shift.is_on != is_on:
        changes.append(
            StaffExcelChange(
                field="is_on",
                old_value=existing_shift.is_on,
                new_value=is_on,
            )
        )
        update_dict["is_on"] = is_on
    if existing_shift.start_time != new_start:
        changes.append(
            StaffExcelChange(
                field="start_time",
                old_value=_serializable(existing_shift.start_time),
                new_value=effective_start_str,
            )
        )
        update_dict["start_time"] = new_start
    if existing_shift.end_time != new_end:
        changes.append(
            StaffExcelChange(
                field="end_time",
                old_value=_serializable(existing_shift.end_time),
                new_value=effective_end_str,
            )
        )
        update_dict["end_time"] = new_end
    if not changes:
        return (
            ShiftExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                weekday=weekday,
                operation="noop",
            ),
            None,
        )
    return (
        ShiftExcelImportRow(
            row_number=row_number,
            staff_id=staff_id,
            staff_code=staff_code_for_view,
            weekday=weekday,
            operation="update",
            changes=changes,
        ),
        update_dict,
    )


# ---------------------------------------------------------------------------
# Override sheet parser (Phase E-7 gap P1)
# ---------------------------------------------------------------------------


def _parse_override_row(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_staff: dict[UUID, Staff],
    existing_overrides: dict[tuple[UUID, int, int, int], StaffWeeklyOverride],
    pending_new_keys: set[tuple[UUID, int, int, int]],
    new_code_to_uuid: dict[str, UUID],
    existing_code_to_id: dict[str, UUID],
    resurrect_staff_ids: set[UUID] | None = None,
) -> tuple[OverrideExcelImportRow, dict[str, Any] | None]:
    """Phase E-7 (gap P1): StaffWeeklyOverride 行のパース.

    operation: new / update / delete / noop / error.
    UNIQUE key: (staff_id, iso_year, iso_week, weekday).
    """
    cells: dict[str, Any] = {}
    for col_key, idx in OVERRIDE_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["staff_id"]
    raw_code = cells["staff_code"]
    raw_delete = cells["delete_flag"]

    # staff_id (任意; 空なら staff_code で解決)
    staff_id: UUID | None = None
    try:
        if not _is_blank(raw_id):
            staff_id = _read_uuid(raw_id)
    except ValueError:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=None,
                staff_code=_read_str(raw_code),
                operation="error",
                error_message=f"staff_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    staff_code = _read_str(raw_code)
    if staff_id is None:
        if not staff_code:
            return (
                OverrideExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=None,
                    operation="error",
                    error_message="staff_id / staff_code のどちらも空です",
                ),
                None,
            )
        if staff_code in existing_code_to_id:
            staff_id = existing_code_to_id[staff_code]
        elif staff_code in new_code_to_uuid:
            staff_id = new_code_to_uuid[staff_code]
        else:
            # 不明な staff_code は warning + skip (PFV course_template と同じ
            # ベストエフォート方針). row は error として扱い、apply からは除外.
            logger.warning(
                "staff_excel override row=%d: staff_code 不明 %r → skip",
                row_number,
                staff_code,
            )
            return (
                OverrideExcelImportRow(
                    row_number=row_number,
                    staff_id=None,
                    staff_code=staff_code,
                    operation="error",
                    error_message=f"staff_code が DB / 同ファイル内に存在しません: {staff_code!r}",
                ),
                None,
            )

    is_pending_new_staff = staff_id in set(new_code_to_uuid.values())
    is_resurrect_staff = resurrect_staff_ids is not None and staff_id in resurrect_staff_ids
    is_pending = is_pending_new_staff or is_resurrect_staff
    staff_obj: Staff | None = existing_staff.get(staff_id) if not is_pending else None
    if staff_obj is None and not is_pending:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code,
                operation="error",
                error_message=f"staff_id が DB に存在しません: {staff_id}",
            ),
            None,
        )
    staff_code_for_view = staff_obj.code if staff_obj is not None else staff_code

    # iso_year / iso_week
    try:
        iso_year = _read_int(cells["iso_year"])
    except (ValueError, TypeError):
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message=f"iso_year が整数ではありません: {cells['iso_year']!r}",
            ),
            None,
        )
    try:
        iso_week = _read_int(cells["iso_week"])
    except (ValueError, TypeError):
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message=f"iso_week が整数ではありません: {cells['iso_week']!r}",
            ),
            None,
        )
    if iso_year is None or iso_week is None:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                operation="error",
                error_message="iso_year / iso_week は必須です",
            ),
            None,
        )
    if iso_week < 1 or iso_week > 53:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                operation="error",
                error_message=f"iso_week は 1..53 のみ: {iso_week}",
            ),
            None,
        )

    # weekday
    raw_weekday = cells["weekday"]
    if _is_blank(raw_weekday):
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                operation="error",
                error_message="weekday が空です",
            ),
            None,
        )
    wd_str = _read_str(raw_weekday)
    weekday = WEEKDAY_LABEL_TO_INT.get(wd_str or "")
    if weekday is None:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                operation="error",
                error_message=f"weekday の値が候補外: {wd_str!r} (許容: 月火水木金土日)",
            ),
            None,
        )

    key = (staff_id, iso_year, iso_week, weekday)
    existing_ov = existing_overrides.get(key)

    # 削除フラグ
    if is_magic_delete(raw_delete):
        if existing_ov is None:
            return (
                OverrideExcelImportRow(
                    row_number=row_number,
                    staff_id=staff_id,
                    staff_code=staff_code_for_view,
                    iso_year=iso_year,
                    iso_week=iso_week,
                    weekday=weekday,
                    operation="noop",
                ),
                None,
            )
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="delete",
            ),
            {"_op": "delete", "_override_id": existing_ov.id},
        )

    # override_type
    raw_ot = cells["override_type"]
    if _is_blank(raw_ot):
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="error",
                error_message="override_type が空です",
            ),
            None,
        )
    override_type = _read_str(raw_ot)
    if override_type not in OVERRIDE_TYPE_VALUES:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="error",
                error_message=(
                    f"override_type の値が候補外: {override_type!r} "
                    f"(許容: {','.join(OVERRIDE_TYPE_VALUES)})"
                ),
            ),
            None,
        )

    # start_time / end_time
    try:
        start_str = _read_hhmm(cells["start_time"])
    except (ValueError, TypeError) as exc:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="error",
                error_message=f"start_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )
    try:
        end_str = _read_hhmm(cells["end_time"])
    except (ValueError, TypeError) as exc:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="error",
                error_message=f"end_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )

    # custom_time は start/end 必須
    if override_type == "custom_time" and (start_str is None or end_str is None):
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="error",
                error_message="override_type=custom_time のときは start_time / end_time が必須です",
            ),
            None,
        )

    reason = _read_str(cells["reason"])

    new_start = _hhmm_to_time(start_str) if start_str is not None else None
    new_end = _hhmm_to_time(end_str) if end_str is not None else None

    # 同ファイル内の (staff_id, iso_year, iso_week, weekday) 重複検査
    if key in pending_new_keys:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="error",
                error_message=(
                    "同ファイル内で (staff_id, iso_year, iso_week, weekday) が重複しています"
                ),
            ),
            None,
        )

    if existing_ov is None:
        pending_new_keys.add(key)
        new_dict: dict[str, Any] = {
            "_op": "new",
            "staff_id": staff_id,
            "iso_year": iso_year,
            "iso_week": iso_week,
            "weekday": weekday,
            "override_type": override_type,
            "start_time": new_start,
            "end_time": new_end,
            "reason": reason,
        }
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="new",
                changes=[
                    StaffExcelChange(
                        field="override_type", old_value=None, new_value=override_type
                    ),
                    StaffExcelChange(field="start_time", old_value=None, new_value=start_str),
                    StaffExcelChange(field="end_time", old_value=None, new_value=end_str),
                ],
            ),
            new_dict,
        )

    # update / noop
    changes: list[StaffExcelChange] = []
    update_dict: dict[str, Any] = {"_op": "update", "_override_id": existing_ov.id}
    if existing_ov.override_type != override_type:
        changes.append(
            StaffExcelChange(
                field="override_type",
                old_value=existing_ov.override_type,
                new_value=override_type,
            )
        )
        update_dict["override_type"] = override_type
    if existing_ov.start_time != new_start:
        changes.append(
            StaffExcelChange(
                field="start_time",
                old_value=_serializable(existing_ov.start_time),
                new_value=start_str,
            )
        )
        update_dict["start_time"] = new_start
    if existing_ov.end_time != new_end:
        changes.append(
            StaffExcelChange(
                field="end_time",
                old_value=_serializable(existing_ov.end_time),
                new_value=end_str,
            )
        )
        update_dict["end_time"] = new_end
    if existing_ov.reason != reason:
        changes.append(
            StaffExcelChange(
                field="reason",
                old_value=existing_ov.reason,
                new_value=reason,
            )
        )
        update_dict["reason"] = reason
    if not changes:
        return (
            OverrideExcelImportRow(
                row_number=row_number,
                staff_id=staff_id,
                staff_code=staff_code_for_view,
                iso_year=iso_year,
                iso_week=iso_week,
                weekday=weekday,
                operation="noop",
            ),
            None,
        )
    return (
        OverrideExcelImportRow(
            row_number=row_number,
            staff_id=staff_id,
            staff_code=staff_code_for_view,
            iso_year=iso_year,
            iso_week=iso_week,
            weekday=weekday,
            operation="update",
            changes=changes,
        ),
        update_dict,
    )


# ---------------------------------------------------------------------------
# public entrypoints
# ---------------------------------------------------------------------------


async def parse_and_diff(
    db: AsyncSession,
    *,
    file_bytes: bytes,
) -> tuple[
    list[StaffExcelImportRow],
    list[ShiftExcelImportRow],
    StaffExcelImportSummary,
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[OverrideExcelImportRow],
    list[dict[str, Any]],
]:
    """Excel をパースして差分行 + summary + apply 用 op list を返す.

    戻り値:
      - staff_rows: API レスポンス用の per-row 結果
      - shift_rows: API レスポンス用の per-row 結果
      - summary: 集計
      - staff_ops: apply 用の DB op list (内部利用)
      - shift_ops: apply 用の DB op list (内部利用)
      - override_rows: Phase E-7 (gap P1) 勤務例外 per-row 結果
      - override_ops: 勤務例外 apply 用 op list
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    if SHEET_STAFF not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_STAFF}」が見つかりません")
    if SHEET_SHIFT not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_SHIFT}」が見つかりません")
    # Phase E-7 (gap P1): SHEET_OVERRIDE は古い Excel との互換のため任意.

    existing_staff = await _load_staff_by_id(db)
    existing_codes = await _load_staff_codes(db)
    offices_by_code = await _load_offices_by_code(db)
    existing_shifts = await _load_shifts_by_key(db, set(existing_staff.keys()))
    # Phase E-7 (gap P0-2 / P1): 既存 secondary_offices, override の lookup を bulk 取得.
    existing_secondary_offices = await _load_secondary_offices_by_staff(
        db, set(existing_staff.keys())
    )
    existing_overrides = await _load_overrides_by_key(db, set(existing_staff.keys()))

    # alive staff の code → id (shift シート側で code リンクに使う)
    existing_code_to_id: dict[str, UUID] = {s.code: s.id for s in existing_staff.values() if s.code}
    # alive staff の code → Staff (削除パスで staff_code リンク削除に使う)
    existing_staff_by_code: dict[str, Staff] = {
        s.code: s for s in existing_staff.values() if s.code
    }
    # resurrection 用: soft-deleted な staff_code → Staff.
    deleted_staff_by_code = await _load_deleted_staff_by_code(db)

    # ---- スタッフシート ----
    ws_s = wb[SHEET_STAFF]
    staff_rows: list[StaffExcelImportRow] = []
    staff_ops: list[dict[str, Any]] = []
    already_seen_codes: set[str] = set()
    # 新規 staff の code → 仮 UUID マップ. shift シート側の code リンクに使う.
    new_code_to_uuid: dict[str, UUID] = {}
    # 復活 staff の code → 既存 UUID マップ. shift シート側の code リンクに使う.
    resurrect_code_to_uuid: dict[str, UUID] = {}
    # row_number は 1-indexed; ヘッダーが 1 行目, データは 2 行目から.
    for r_idx, row in enumerate(ws_s.iter_rows(min_row=2, values_only=True), start=2):
        if _row_is_empty(row):
            continue
        diff_row, op = _parse_staff_row(
            r_idx,
            row,
            existing_staff=existing_staff,
            existing_staff_by_code=existing_staff_by_code,
            deleted_staff_by_code=deleted_staff_by_code,
            offices_by_code=offices_by_code,
            existing_codes=existing_codes,
            already_seen_codes=already_seen_codes,
            new_code_to_uuid=new_code_to_uuid,
            resurrect_code_to_uuid=resurrect_code_to_uuid,
            existing_secondary_offices=existing_secondary_offices,
        )
        staff_rows.append(diff_row)
        if op is not None:
            staff_ops.append(op)

    # 復活 staff も shift シートから code リンクで参照できるよう、既存 code → id
    # マップに追加する (resurrect 行で更新済み staff の UUID は既存値そのまま).
    existing_code_to_id_with_resurrect: dict[str, UUID] = {
        **existing_code_to_id,
        **resurrect_code_to_uuid,
    }

    # ---- 勤務シフトシート ----
    ws_f = wb[SHEET_SHIFT]
    shift_rows: list[ShiftExcelImportRow] = []
    shift_ops: list[dict[str, Any]] = []
    pending_new_keys: set[tuple[UUID, int]] = set()
    for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
        if _row_is_empty(row):
            continue
        diff_row, op = _parse_shift_row(
            r_idx,
            row,
            existing_staff=existing_staff,
            existing_shifts=existing_shifts,
            pending_new_keys=pending_new_keys,
            new_code_to_uuid=new_code_to_uuid,
            existing_code_to_id=existing_code_to_id_with_resurrect,
            resurrect_staff_ids=set(resurrect_code_to_uuid.values()),
        )
        shift_rows.append(diff_row)
        if op is not None:
            shift_ops.append(op)

    # ---- 勤務例外シート (Phase E-7 gap P1) ----
    override_rows: list[OverrideExcelImportRow] = []
    override_ops: list[dict[str, Any]] = []
    if SHEET_OVERRIDE in wb.sheetnames:
        ws_o = wb[SHEET_OVERRIDE]
        pending_override_keys: set[tuple[UUID, int, int, int]] = set()
        for r_idx, row in enumerate(ws_o.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            diff_row, op = _parse_override_row(
                r_idx,
                row,
                existing_staff=existing_staff,
                existing_overrides=existing_overrides,
                pending_new_keys=pending_override_keys,
                new_code_to_uuid=new_code_to_uuid,
                existing_code_to_id=existing_code_to_id_with_resurrect,
                resurrect_staff_ids=set(resurrect_code_to_uuid.values()),
            )
            override_rows.append(diff_row)
            if op is not None:
                override_ops.append(op)

    # ---- summary ----
    summary = StaffExcelImportSummary()
    for r in staff_rows:
        match r.operation:
            case "new":
                summary.staff_new += 1
            case "update":
                summary.staff_update += 1
            case "delete":
                summary.staff_delete += 1
            case "error":
                summary.staff_error += 1
            case "noop":
                summary.staff_noop += 1
    for r in shift_rows:
        match r.operation:
            case "new":
                summary.shift_new += 1
            case "update":
                summary.shift_update += 1
            case "delete":
                summary.shift_delete += 1
            case "error":
                summary.shift_error += 1
            case "noop":
                summary.shift_noop += 1
    for r in override_rows:
        match r.operation:
            case "new":
                summary.override_new += 1
            case "update":
                summary.override_update += 1
            case "delete":
                summary.override_delete += 1
            case "error":
                summary.override_error += 1
            case "noop":
                summary.override_noop += 1

    return staff_rows, shift_rows, summary, staff_ops, shift_ops, override_rows, override_ops


async def apply_changes(
    db: AsyncSession,
    *,
    staff_ops: list[dict[str, Any]],
    shift_ops: list[dict[str, Any]],
    override_ops: list[dict[str, Any]] | None = None,
) -> None:
    """差分を DB に反映する (partial commit).

    ``staff_ops`` / ``shift_ops`` / ``override_ops`` は ``parse_and_diff`` の時点で
    error 行が除外された有効な op (new / update / delete) のみを含む. error 行は
    ``op=None`` として既に skip されているので、ここでは触れない.
    SQLAlchemy 例外は呼び出し側にバブルアップする (rollback は呼び出し側で実施).

    Phase E-7 (gap P0-2): 各 staff op が ``_secondary_office_ids`` を含む場合、
    StaffSecondaryOffice の clear → re-add を実行する.
    """
    # 1) Staff: new → insert / update → 既存行を更新 / delete → soft delete + 関連 shift 物理削除 /
    #    resurrect → deleted_at=NULL に戻して内容上書き
    staff_by_id: dict[UUID, Staff] = {}
    # Phase E-7 (gap P0-2): secondary_offices の差分を後で apply するため記録.
    secondary_office_updates: list[tuple[UUID, list[UUID]]] = []
    for op in staff_ops:
        op_type = op.get("_op")
        if op_type == "new":
            data = {
                k: v for k, v in op.items() if not k.startswith("_") and k != "secondary_office_ids"
            }
            new_staff = Staff(**data)
            db.add(new_staff)
            if "_secondary_office_ids" in op:
                secondary_office_updates.append((op["_assigned_id"], op["_secondary_office_ids"]))
        elif op_type == "resurrect":
            sid = op["_staff_id"]
            if sid not in staff_by_id:
                staff_by_id[sid] = await db.get(Staff, sid)
            staff_obj = staff_by_id[sid]
            if staff_obj is None:
                continue
            staff_obj.deleted_at = None
            for k, v in op["_updates"].items():
                setattr(staff_obj, k, v)
            if "_secondary_office_ids" in op:
                secondary_office_updates.append((sid, op["_secondary_office_ids"]))
        elif op_type == "update":
            sid = op["_staff_id"]
            if sid not in staff_by_id:
                staff_by_id[sid] = await db.get(Staff, sid)
            staff_obj = staff_by_id[sid]
            if staff_obj is None:
                continue
            for k, v in op.items():
                if k.startswith("_"):
                    continue
                setattr(staff_obj, k, v)
            if "_secondary_office_ids" in op:
                secondary_office_updates.append((sid, op["_secondary_office_ids"]))
        elif op_type == "delete":
            sid = op["_staff_id"]
            if sid not in staff_by_id:
                staff_by_id[sid] = await db.get(Staff, sid)
            staff_obj = staff_by_id[sid]
            if staff_obj is not None:
                staff_obj.deleted_at = func.now()
                # 関連 shift を物理削除 (orphan 防止).
                shifts_to_cleanup = (
                    await db.scalars(select(StaffShift).where(StaffShift.staff_id == sid))
                ).all()
                for sh in shifts_to_cleanup:
                    await db.delete(sh)
                # secondary_offices も physical delete (cascade で消えるが明示).
                sec_to_cleanup = (
                    await db.scalars(
                        select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == sid)
                    )
                ).all()
                for so in sec_to_cleanup:
                    await db.delete(so)
                # 関連 override も物理削除 (cascade で消えるが明示).
                ovs_to_cleanup = (
                    await db.scalars(
                        select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid)
                    )
                ).all()
                for ov in ovs_to_cleanup:
                    await db.delete(ov)
    # 中間 flush で staff row を確定 (shift new で staff_id を再参照する可能性).
    await db.flush()

    # Phase E-7 (gap P0-2): secondary_offices を clear → re-add で完全置換.
    for sid, new_ids in secondary_office_updates:
        existing_sec_rows = (
            await db.scalars(
                select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == sid)
            )
        ).all()
        for r in existing_sec_rows:
            await db.delete(r)
        await db.flush()
        for off_id in new_ids:
            db.add(StaffSecondaryOffice(staff_id=sid, office_id=off_id))
    if secondary_office_updates:
        await db.flush()

    # 2) Shift: new → insert / update → 既存行を更新 / delete → 物理削除
    for op in shift_ops:
        op_type = op.get("_op")
        if op_type == "new":
            data = {k: v for k, v in op.items() if not k.startswith("_")}
            db.add(StaffShift(**data))
        elif op_type == "update":
            sid, wd = op["_shift_key"]
            obj = await db.scalar(
                select(StaffShift).where(
                    StaffShift.staff_id == sid,
                    StaffShift.weekday == wd,
                )
            )
            if obj is None:
                continue
            for k, v in op.items():
                if k.startswith("_"):
                    continue
                setattr(obj, k, v)
        elif op_type == "delete":
            sid, wd = op["_shift_key"]
            obj = await db.scalar(
                select(StaffShift).where(
                    StaffShift.staff_id == sid,
                    StaffShift.weekday == wd,
                )
            )
            if obj is not None:
                await db.delete(obj)
    await db.flush()

    # 3) Override (Phase E-7 gap P1).
    if override_ops:
        for op in override_ops:
            op_type = op.get("_op")
            if op_type == "new":
                data = {k: v for k, v in op.items() if not k.startswith("_")}
                db.add(StaffWeeklyOverride(**data))
            elif op_type == "update":
                ov_id = op["_override_id"]
                obj = await db.get(StaffWeeklyOverride, ov_id)
                if obj is None:
                    continue
                for k, v in op.items():
                    if k.startswith("_"):
                        continue
                    setattr(obj, k, v)
            elif op_type == "delete":
                ov_id = op["_override_id"]
                obj = await db.get(StaffWeeklyOverride, ov_id)
                if obj is not None:
                    await db.delete(obj)
        await db.flush()
