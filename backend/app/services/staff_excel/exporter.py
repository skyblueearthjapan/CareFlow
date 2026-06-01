"""Excel 出力 (テンプレート / 全件エクスポート).

openpyxl で 2 シート構成のワークブックを組み立てる:

  1. スタッフマスタシート
  2. 勤務シフトシート

呼び出し側 (API endpoint) は ``build_workbook`` の戻り値である ``Workbook``
を ``save`` するか ``BytesIO`` にダンプしてレスポンスする.
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import time
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.models.office import Office
from app.models.staff import Staff, StaffShift, StaffWeeklyOverride
from app.services.staff_excel.schema import (
    HEADER_FILL_COLOR,
    HEADER_FONT_COLOR,
    ID_COLUMN_FILL_COLOR,
    OVERRIDE_EDIT_COL_INDEX,
    OVERRIDE_EDIT_COLUMNS,
    OVERRIDE_EDIT_WEEKDAY_KEY_TO_INT,
    SAMPLE_ROW_FILL_COLOR,
    SAMPLE_ROW_MARKER,
    SHEET_OVERRIDE_EDIT,
    SHEET_SHIFT_EDIT,
    SHEET_STAFF,
    SHIFT_EDIT_COL_INDEX,
    SHIFT_EDIT_COLUMNS,
    SHIFT_EDIT_WEEKDAY_KEY_TO_INT,
    STAFF_COL_INDEX,
    STAFF_COLUMNS,
    format_override_cell,
)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _set_header_row(ws: Worksheet, columns: list[dict[str, object]]) -> None:
    """1 行目にヘッダーを書き、太字 + 背景色 + freeze を設定."""
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_def["header"]))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        width = col_def.get("width", 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = int(width)  # type: ignore[arg-type]
    ws.freeze_panes = "A2"


def _attach_dropdowns(
    ws: Worksheet,
    columns: list[dict[str, object]],
    *,
    max_data_rows: int = 1000,
) -> None:
    """各列の dropdown を ``max_data_rows`` 行分まで設定."""
    last_row = 1 + max_data_rows  # ヘッダー行を除外
    for col_idx, col_def in enumerate(columns, start=1):
        dropdown = col_def.get("dropdown")
        if dropdown is None:
            continue
        # openpyxl の DataValidation list 値は "値1,値2,..." をダブルクォートで囲む.
        # 値内にカンマが含まれない前提 (今回は全て安全な値).
        formula = '"' + ",".join(str(v) for v in dropdown) + '"'  # type: ignore[arg-type]
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "リストにない値です"
        dv.errorTitle = "入力エラー"
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)


def _shade_id_column_data_rows(
    ws: Worksheet,
    col_key: str,
    columns: list[dict[str, object]],
    *,
    data_row_count: int,
    start_row: int = 2,
) -> None:
    """staff_id 列の **データ行のみ** を薄いグレー背景でハイライト.

    無条件で固定行数 (例: 1000) を塗ると ``ws.max_row`` が 1000 を返してしまい、
    テンプレート判定 (空 = max_row==1) ができなくなる。データ行が 0 件のときは
    何もしない。``start_row`` は記入例行などの先頭固定行を除外するために使う.
    """
    if data_row_count <= 0:
        return
    target_index = None
    for i, col_def in enumerate(columns, start=1):
        if col_def["key"] == col_key:
            target_index = i
            break
    if target_index is None:
        return
    fill = PatternFill("solid", fgColor=ID_COLUMN_FILL_COLOR)
    col_letter = get_column_letter(target_index)
    for row in range(start_row, start_row + data_row_count):
        ws[f"{col_letter}{row}"].fill = fill


def _write_staff_row(
    ws: Worksheet,
    row_idx: int,
    staff: Staff,
    *,
    office_code_by_id: dict[UUID, str],
    secondary_office_ids_by_staff: dict[UUID, list[UUID]] | None = None,
) -> None:
    """1 行分のスタッフデータを書き込む."""
    code = ""
    if staff.primary_office_id is not None:
        code = office_code_by_id.get(staff.primary_office_id, "")
    # Phase E-7 (gap P0-2): secondary_offices → comma-separated コード列.
    # 解決できない id は silent skip (export 時点で alive offices のみ load される前提).
    # Phase E-7 (gap LOW#5): コード辞書順に sort して deterministic にする.
    sec_codes: list[str] = []
    if secondary_office_ids_by_staff is not None:
        for off_id in secondary_office_ids_by_staff.get(staff.id, []):
            oc = office_code_by_id.get(off_id)
            if oc:
                sec_codes.append(oc)
    sec_codes.sort()
    sec_codes_str = ",".join(sec_codes) if sec_codes else None
    values: dict[str, object | None] = {
        "staff_id": str(staff.id),
        "staff_code": staff.code,
        "name": staff.name,
        "kana": staff.kana,
        "sex": staff.sex,
        "status": staff.status,
        "role": staff.role,
        "office_code": code or None,
        "secondary_office_codes": sec_codes_str,
        "is_trainee": "TRUE" if staff.is_trainee else "FALSE",
        "note": staff.note,
        "delete_flag": None,
    }
    for col_key, col_idx in STAFF_COL_INDEX.items():
        ws.cell(row=row_idx, column=col_idx + 1, value=values.get(col_key))


def _hhmm(t: time | None) -> str | None:
    if t is None:
        return None
    return f"{t.hour:02d}:{t.minute:02d}"


def _write_shift_edit_row(
    ws: Worksheet,
    row_idx: int,
    staff: Staff,
    shifts_for_staff: list[StaffShift],
) -> None:
    """Phase G-56: 1 スタッフ分の週次シフトを 1 行に書き込む (編集用シート).

    各曜日について開始/終了を書く. is_on=False (休) もしくは shift 未登録の曜日は空欄.
    """
    by_weekday: dict[int, StaffShift] = {sh.weekday: sh for sh in shifts_for_staff}
    values: dict[str, object | None] = {
        "staff_id": str(staff.id),
        "staff_code": staff.code,
        "staff_name": staff.name,
        "delete_flag": None,
    }
    for wd_key, wd_int in SHIFT_EDIT_WEEKDAY_KEY_TO_INT.items():
        sh = by_weekday.get(wd_int)
        if sh is not None and sh.is_on and sh.start_time is not None and sh.end_time is not None:
            values[f"{wd_key}_start"] = _hhmm(sh.start_time)
            values[f"{wd_key}_end"] = _hhmm(sh.end_time)
        else:
            # 休 (is_on=False) もしくは shift 未登録 → 空欄.
            values[f"{wd_key}_start"] = None
            values[f"{wd_key}_end"] = None
    for col_key, col_idx in SHIFT_EDIT_COL_INDEX.items():
        ws.cell(row=row_idx, column=col_idx + 1, value=values.get(col_key))


def _write_override_edit_row(
    ws: Worksheet,
    row_idx: int,
    staff: Staff,
    iso_year: int,
    iso_week: int,
    overrides_for_week: list[StaffWeeklyOverride],
) -> None:
    """Phase G-56: 1 (staff, iso_year, iso_week) 分の例外を 1 行に書き込む.

    各曜日セルに "休" / "HH:MM-HH:MM" を書く. 例外なしの曜日は空欄. 理由は代表値
    (週内で最初に見つかった非空の reason).
    """
    by_weekday: dict[int, StaffWeeklyOverride] = {ov.weekday: ov for ov in overrides_for_week}
    reason: str | None = None
    for ov in overrides_for_week:
        if ov.reason:
            reason = ov.reason
            break
    values: dict[str, object | None] = {
        "staff_id": str(staff.id),
        "staff_code": staff.code,
        "staff_name": staff.name,
        "iso_year": iso_year,
        "iso_week": iso_week,
        "reason": reason,
        "delete_flag": None,
    }
    for wd_key, wd_int in OVERRIDE_EDIT_WEEKDAY_KEY_TO_INT.items():
        ov = by_weekday.get(wd_int)
        if ov is None:
            values[wd_key] = None
        else:
            values[wd_key] = format_override_cell(
                ov.override_type, _hhmm(ov.start_time), _hhmm(ov.end_time)
            )
    for col_key, col_idx in OVERRIDE_EDIT_COL_INDEX.items():
        ws.cell(row=row_idx, column=col_idx + 1, value=values.get(col_key))


def _write_override_sample_row(ws: Worksheet, row_idx: int) -> None:
    """Phase G-58.1: 勤務例外（編集用）シートに「記入例 (サンプル)」行を書く.

    空シートで記入方法が分からない問題への対策。staff_code を ``SAMPLE_ROW_MARKER``
    (「（記入例）」) にすることで importer / replace_all が is_sample_row 判定で
    **完全に無視** する (DB 変更ゼロ・エラー無し). 行全体をグレー背景 + 斜体にして
    「編集対象でない」雰囲気を出す.
    """
    values: dict[str, object | None] = {
        "staff_id": None,
        "staff_code": SAMPLE_ROW_MARKER,
        "staff_name": f"↓この行は取り込まれません{SAMPLE_ROW_MARKER}",
        "iso_year": 2026,
        "iso_week": 23,
        "mon": "休",
        "tue": "10:00-15:00",
        "reason": "（例）火曜は通院のため時間変更、月曜は休み",
    }
    sample_font = Font(italic=True, color="FF808080")
    sample_fill = PatternFill("solid", fgColor=SAMPLE_ROW_FILL_COLOR)
    for col_key, col_idx in OVERRIDE_EDIT_COL_INDEX.items():
        cell = ws.cell(row=row_idx, column=col_idx + 1, value=values.get(col_key))
        cell.font = sample_font
        cell.fill = sample_fill


# ---------------------------------------------------------------------------
# public entrypoints
# ---------------------------------------------------------------------------


def build_workbook(
    *,
    staff_list: Sequence[Staff],
    shifts: Sequence[StaffShift],
    offices: Sequence[Office],
    secondary_office_ids_by_staff: dict[UUID, list[UUID]] | None = None,
    overrides: Sequence[StaffWeeklyOverride] | None = None,
) -> Workbook:
    """テンプレート + データ込みのワークブックを構築する.

    ``staff_list`` を空 list で呼べばテンプレート (ヘッダー + dropdown のみ) になる.

    Phase E-7 (gap P0-2, P1):
      * ``secondary_office_ids_by_staff`` : Staff.id → [Office.id, ...] のマップ.
        staff.secondary_offices relationship を呼び出し側で precompute して渡す
        (lazy-load を async session で叩けないため).
      * ``overrides`` : StaffWeeklyOverride 一覧. 「勤務例外」シートに書き出す.
        None もしくは空 list の場合はヘッダーのみの空シートが出力される.
    """
    wb = Workbook()
    # default のシートを 1 枚目として使う
    ws_s: Worksheet = wb.active  # type: ignore[assignment]
    ws_s.title = SHEET_STAFF

    _set_header_row(ws_s, STAFF_COLUMNS)
    _attach_dropdowns(ws_s, STAFF_COLUMNS)

    office_code_by_id: dict[UUID, str] = {
        office.id: (office.code or "")
        for office in offices
        if office.code  # コード未設定の拠点はスキップ
    }
    for i, staff in enumerate(staff_list, start=2):
        _write_staff_row(
            ws_s,
            i,
            staff,
            office_code_by_id=office_code_by_id,
            secondary_office_ids_by_staff=secondary_office_ids_by_staff,
        )
    _shade_id_column_data_rows(ws_s, "staff_id", STAFF_COLUMNS, data_row_count=len(staff_list))

    staff_lookup: dict[UUID, Staff] = {s.id: s for s in staff_list}

    # 勤務シフト（編集用）シート — Phase G-56: 1 スタッフ 1 行.
    ws_f: Worksheet = wb.create_sheet(title=SHEET_SHIFT_EDIT)
    _set_header_row(ws_f, SHIFT_EDIT_COLUMNS)
    _attach_dropdowns(ws_f, SHIFT_EDIT_COLUMNS)

    # staff_id → その staff の shift list. 全スタッフを 1 行ずつ出力する (shift 無しの
    # staff も行を出す = 週次シフトの正本).
    shifts_by_staff: dict[UUID, list[StaffShift]] = {}
    for sh in shifts:
        shifts_by_staff.setdefault(sh.staff_id, []).append(sh)
    for i, staff in enumerate(staff_list, start=2):
        _write_shift_edit_row(ws_f, i, staff, shifts_by_staff.get(staff.id, []))
    _shade_id_column_data_rows(ws_f, "staff_id", SHIFT_EDIT_COLUMNS, data_row_count=len(staff_list))

    # 勤務例外（編集用）シート — Phase G-56: 1 スタッフ × 1 週 = 1 行.
    ws_o: Worksheet = wb.create_sheet(title=SHEET_OVERRIDE_EDIT)
    _set_header_row(ws_o, OVERRIDE_EDIT_COLUMNS)
    _attach_dropdowns(ws_o, OVERRIDE_EDIT_COLUMNS)
    # Phase G-58.1: ヘッダー直下 (row 2) に記入例 (サンプル) 行を常に書く.
    # 実データの有無に関わらず先頭に置き、import 時は is_sample_row で skip される.
    _write_override_sample_row(ws_o, 2)
    override_list = list(overrides or [])
    # (staff_id, iso_year, iso_week) でグルーピング. round-trip 安定のため
    # staff_list 順 → (year, week) 昇順で並べる.
    grouped: dict[tuple[UUID, int, int], list[StaffWeeklyOverride]] = {}
    for ov in override_list:
        grouped.setdefault((ov.staff_id, ov.iso_year, ov.iso_week), []).append(ov)
    staff_order: dict[UUID, int] = {s.id: idx for idx, s in enumerate(staff_list)}
    ordered_keys = sorted(
        grouped.keys(),
        key=lambda k: (staff_order.get(k[0], len(staff_list)), k[1], k[2]),
    )
    # 記入例行が row 2 なので実データは row 3 から.
    row_idx = 3
    for staff_id, iso_year, iso_week in ordered_keys:
        staff = staff_lookup.get(staff_id)
        if staff is None:
            continue
        _write_override_edit_row(
            ws_o, row_idx, staff, iso_year, iso_week, grouped[(staff_id, iso_year, iso_week)]
        )
        row_idx += 1
    # データ行のグレー装飾は row 3 以降のみ (記入例行は独自装飾済み).
    _shade_id_column_data_rows(
        ws_o, "staff_id", OVERRIDE_EDIT_COLUMNS, data_row_count=row_idx - 3, start_row=3
    )

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
