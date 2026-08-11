"""Excel 出力 (テンプレート / 全件エクスポート).

openpyxl で 2 シート構成のワークブックを組み立てる:

  1. 患者マスタシート
  2. 固定訪問パターンシート (PFV)

呼び出し側 (API endpoint) は ``build_workbook`` の戻り値である ``Workbook``
を ``save`` するか ``BytesIO`` にダンプしてレスポンスする.
"""

from __future__ import annotations

import logging
from collections.abc import Sequence
from datetime import time
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.course_template import CourseTemplate
from app.models.office import Office
from app.models.patient import Patient
from app.models.patient_fixed_visit import PatientFixedVisit
from app.models.patient_ng_staff import PatientNgStaff
from app.models.staff import Staff
from app.services.patient_excel.schema import (
    COMMENT_AUTHOR,
    DEFAULT_TIME_TYPE,
    GRID_BLOCK_HEIGHT,
    GRID_COLS_PER_DAY,
    GRID_COURSE_ORDER,
    GRID_DAY_SUBHEADERS,
    GRID_OFFICE_GAP,
    GRID_OFFICE_ORDER,
    GRID_TIME_SLOTS,
    GRID_WEEKDAY_FULL_LABELS,
    HEADER_FILL_COLOR,
    HEADER_FONT_COLOR,
    ID_COLUMN_FILL_COLOR,
    ID_COLUMN_FONT_COLOR,
    INSURANCE_EN_TO_JA,
    LATLNG_COMMENT_TEXT,
    OFFICE_CODE_COMMENT_TEXT,
    PATIENT_COL_INDEX,
    PATIENT_COLUMNS,
    PATIENT_ID_COMMENT_TEXT,
    PFV_EDIT_COL_INDEX,
    PFV_EDIT_COLUMNS,
    PFV_EDIT_PATIENT_ID_COMMENT_TEXT,
    PFV_EDIT_WEEKDAY_KEY_TO_INT,
    PFV_EDIT_WEEKDAY_KEYS,
    PFV_GRID_A1_COMMENT_TEXT,
    SEX_EN_TO_JA,
    SEX_RESTRICTION_EN_TO_JA,
    SHEET_PATIENTS,
    SHEET_PFV_EDIT,
    SHEET_PFV_GRID,
    STATUS_EN_TO_JA,
    TIME_TYPE_GREYOUT_VALUES,
    VISIT_FREQUENCY_EN_TO_JA,
    build_office_code_short_maps,
    course_token,
    weekdays_en_to_yesno_cells,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _set_header_row(ws: Worksheet, columns: list[dict[str, object]]) -> None:
    """1 行目にヘッダーを書き、太字 + 背景色 + freeze を設定."""
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_def["header"]))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        width = col_def.get("width", 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = int(width)  # type: ignore[arg-type]
    ws.freeze_panes = "A2"


def _attach_dropdowns(
    ws: Worksheet,
    columns: list[dict[str, object]],
    *,
    max_data_rows: int = 1000,
) -> None:
    """各列の dropdown を ``max_data_rows`` 行分まで設定."""
    last_row = 1 + max_data_rows  # ヘッダー行を除外
    for col_idx, col_def in enumerate(columns, start=1):
        dropdown = col_def.get("dropdown")
        if dropdown is None:
            continue
        # openpyxl の DataValidation list 値は "値1,値2,..." をダブルクォートで囲む.
        # 値内にカンマが含まれない前提 (今回は全て安全な値).
        formula = '"' + ",".join(str(v) for v in dropdown) + '"'  # type: ignore[arg-type]
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "リストにない値です"
        dv.errorTitle = "入力エラー"
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)


def _column_index(col_key: str, columns: list[dict[str, object]]) -> int | None:
    """1-indexed のカラム位置を返す (見つからなければ None)."""
    for i, col_def in enumerate(columns, start=1):
        if col_def["key"] == col_key:
            return i
    return None


def _shade_id_column_data_rows(
    ws: Worksheet,
    col_key: str,
    columns: list[dict[str, object]],
    *,
    data_row_count: int,
) -> None:
    """patient_id 列の **データ行のみ** に「触らない雰囲気」装飾を適用.

    薄いグレー背景 + イタリック + やや薄い色のフォントで、ユーザーに
    「ここは編集不要 (新規登録時は空欄)」を視覚的に伝える。

    無条件で固定行数 (例: 1000) を塗ると ``ws.max_row`` が 1000 を返してしまい、
    テンプレート判定 (空 = max_row==1) ができなくなる。データ行が 0 件のときは
    何もしない。
    """
    if data_row_count <= 0:
        return
    target_index = _column_index(col_key, columns)
    if target_index is None:
        return
    fill = PatternFill("solid", fgColor=ID_COLUMN_FILL_COLOR)
    font = Font(italic=True, color=ID_COLUMN_FONT_COLOR)
    col_letter = get_column_letter(target_index)
    for row in range(2, 2 + data_row_count):
        cell = ws[f"{col_letter}{row}"]
        cell.fill = fill
        cell.font = font


def _attach_header_comment(
    ws: Worksheet,
    col_key: str,
    columns: list[dict[str, object]],
    *,
    text: str,
) -> None:
    """ヘッダー (1 行目) の指定カラムにコメント (ホバーで表示) を付ける."""
    target_index = _column_index(col_key, columns)
    if target_index is None:
        return
    col_letter = get_column_letter(target_index)
    cell = ws[f"{col_letter}1"]
    # openpyxl の Comment(text, author) は幅/高さを明示しないとデフォルトが小さい
    comment = Comment(text, COMMENT_AUTHOR)
    comment.width = 280
    comment.height = 80
    cell.comment = comment


def _attach_time_greyout_conditional_format(
    ws: Worksheet,
    columns: list[dict[str, object]],
    *,
    max_data_rows: int = 1000,
) -> None:
    """Phase G-49: 時間タイプが 午前/午後/終日 のとき希望開始/終了時刻をグレーアウト.

    マクロを使わず openpyxl の ``FormulaRule`` (条件付き書式) で、時間タイプ列を
    参照する数式 (例 ``OR($S2="午前",$S2="午後",$S2="終日")``) を時刻 2 列に適用する.
    列レターは実際の列順から算出する (列追加/並び替えに追従).
    """
    tt_idx = _column_index("weekly_time_type", columns)
    start_idx = _column_index("preferred_start", columns)
    end_idx = _column_index("preferred_end", columns)
    if tt_idx is None or start_idx is None or end_idx is None:
        return
    tt_letter = get_column_letter(tt_idx)
    last_row = 1 + max_data_rows
    # 数式は適用範囲の先頭行 (2 行目) 基準で書く. openpyxl が各行に相対展開する.
    # 時間タイプ列は絶対列 ($), 行は相対 (行番号なし → 評価行) にする.
    conditions = ",".join(f'${tt_letter}2="{v}"' for v in TIME_TYPE_GREYOUT_VALUES)
    formula = f"OR({conditions})"
    grey_fill = PatternFill("solid", fgColor=ID_COLUMN_FILL_COLOR)
    grey_font = Font(italic=True, color=ID_COLUMN_FONT_COLOR)
    for col_idx in (start_idx, end_idx):
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{last_row}"
        rule = FormulaRule(formula=[formula], fill=grey_fill, font=grey_font)
        ws.conditional_formatting.add(rng, rule)


async def load_ng_staff_codes_by_patient(
    db: AsyncSession,
    patient_ids: Sequence[UUID],
) -> dict[UUID, str]:
    """患者 ID → NG スタッフコードのカンマ区切り文字列 (sorted) を一括ロードする.

    N+1 を避けるため patient_ng_staff × staff を 1 クエリで JOIN して引く.

    * 退職済み (staff.deleted_at IS NOT NULL) スタッフも含める — 既存 NG 行は
      退職後も残るため、export で落とすと round-trip で暗黙解除されてしまう.
    * ``staff.code`` が NULL / 空のスタッフは **skip** する (staff_id (UUID 文字列)
      へのフォールバックはしない). import 側は staff_code でしか解決できないため、
      UUID を書き出しても再取込で「不明コード」となり無意味だから. code 無しスタッフは
      運用上存在しない想定なので、発生した場合は warning ログで気付けるようにする.
    """
    if not patient_ids:
        return {}
    rows = (
        await db.execute(
            select(PatientNgStaff.patient_id, Staff.code)
            .join(Staff, Staff.id == PatientNgStaff.staff_id)
            .where(PatientNgStaff.patient_id.in_(list(patient_ids)))
        )
    ).all()
    codes_by_patient: dict[UUID, list[str]] = {}
    missing_code_count = 0
    for patient_id, staff_code in rows:
        if not staff_code:
            missing_code_count += 1
            continue
        codes_by_patient.setdefault(patient_id, []).append(str(staff_code))
    if missing_code_count:
        logger.warning(
            "patient_excel export: staff_code 未設定のスタッフを NG 列から %d 件 skip しました.",
            missing_code_count,
        )
    return {pid: ",".join(sorted(codes)) for pid, codes in codes_by_patient.items() if codes}


def _write_patient_row(
    ws: Worksheet,
    row_idx: int,
    patient: Patient,
    *,
    office_code_by_id: dict[UUID, str],
    ng_staff_codes_by_patient: dict[UUID, str] | None = None,
) -> None:
    """1 行分の患者データを書き込む.

    primary_office_id → 拠点コード (INAGE/TSUGA) の lookup は呼び出し側で構築済みの
    ``office_code_by_id`` を使う.

    NG スタッフ列は ``load_ng_staff_codes_by_patient`` で precompute 済みの
    ``ng_staff_codes_by_patient`` (patient_id → "S-001,S-014") をそのまま書く.
    """
    code = ""
    if patient.primary_office_id is not None:
        code = office_code_by_id.get(patient.primary_office_id, "")

    # Phase G-48: enum 系は日本語ラベルで書き出す (DB 英語値 → 日本語).
    # 想定外の DB 値はそのまま書き出す (import 側で英語値受理にフォールバックさせる).
    sex_ja = SEX_EN_TO_JA.get(patient.sex, patient.sex) if patient.sex else None
    status_ja = STATUS_EN_TO_JA.get(patient.status, patient.status) if patient.status else None
    insurance_ja = (
        INSURANCE_EN_TO_JA.get(patient.insurance, patient.insurance) if patient.insurance else None
    )
    # 性別制限: DB NULL → 「なし」を明示出力 (dropdown に合わせ運用者に意図を示す).
    sex_restriction_ja = (
        SEX_RESTRICTION_EN_TO_JA.get(patient.sex_restriction, patient.sex_restriction)
        if patient.sex_restriction
        else "なし"
    )

    # Phase G-48: weekly_pattern (希望訪問パターン) を統合シートの横一列に展開.
    wp = patient.weekly_pattern or {}
    if not isinstance(wp, dict):
        wp = {}
    vf_raw = wp.get("visit_frequency")
    visit_frequency_ja = (
        VISIT_FREQUENCY_EN_TO_JA.get(vf_raw, vf_raw) if isinstance(vf_raw, str) else None
    )
    # Phase G-49/G-50: 希望曜日を 7 列 (月..日) の 〇/× に展開.
    pref_weekdays = wp.get("preferred_weekdays")
    weekday_cells = weekdays_en_to_yesno_cells(
        pref_weekdays if isinstance(pref_weekdays, list) else None
    )

    values: dict[str, object | None] = {
        "patient_id": str(patient.id),
        "patient_code": patient.code,
        "name": patient.name,
        "kana": patient.kana,
        "sex": sex_ja,
        "status": status_ja,
        "insurance": insurance_ja,
        "address": patient.address,
        # Numeric → 文字列化を避けて数値のまま書く. None はそのまま空セル.
        "lat": float(patient.lat) if patient.lat is not None else None,
        "lng": float(patient.lng) if patient.lng is not None else None,
        "office_code": code or None,
        "sex_restriction": sex_restriction_ja,
        # Phase E-7 (gap P0-1): requires_multiple_staff を「はい/いいえ」で書き出す.
        # NOT NULL bool 列なのでデフォルト False で必ず値が入る.
        "requires_multiple_staff": "はい" if patient.requires_multiple_staff else "いいえ",
        # ---- Phase G-48: weekly_pattern 統合列 ----
        "frequency_per_week": wp.get("frequency_per_week"),
        "visit_frequency": visit_frequency_ja,
        "visit_weeks": wp.get("visit_weeks"),
        # Phase G-49: 希望曜日 7 列 (はい/いいえ).
        **weekday_cells,
        "service_minutes": wp.get("service_minutes"),
        "weekly_time_type": wp.get("time_type"),
        "preferred_start": wp.get("preferred_start"),
        "preferred_end": wp.get("preferred_end"),
        "note": patient.note,
        "delete_flag": None,
        # NG スタッフ: 無ければ空セル (= import 側で「維持」).
        "ng_staff_codes": ((ng_staff_codes_by_patient or {}).get(patient.id) or None),
    }
    for col_key, col_idx in PATIENT_COL_INDEX.items():
        ws.cell(row=row_idx, column=col_idx + 1, value=values.get(col_key))


def _hhmm(t: time | None) -> str | None:
    if t is None:
        return None
    return f"{t.hour:02d}:{t.minute:02d}"


def _write_pfv_edit_patient_row(
    ws: Worksheet,
    row_idx: int,
    patient: Patient,
    pfvs: list[PatientFixedVisit],
    *,
    course_template_by_id: dict[UUID, CourseTemplate],
    office_code_by_id: dict[UUID, str],
    code_to_short: dict[str, str] | None = None,
    crossoffice_warnings: list[str] | None = None,
) -> None:
    """Phase G-51: 患者 1 行 (月〜土 各曜日 時刻 + コース) を「編集用」シートに書く.

    ``pfvs`` はこの患者の **normal mode** PFV のみ (呼び出し側でフィルタ済).
    各曜日 (0=月..5=土) に start_time → ``{曜日}_時刻``、course_template → ``{曜日}_コース``
    (拠点付きトークン) を書く. duration / time_type は患者代表値 (全曜日同一前提;
    異なる場合は最初の行の値を採用しログ警告). slot_index=1 は現データ 0 だが、
    存在したらログ警告し slot_index=0 の行を優先採用 (将来対応).
    """
    # weekday → 採用する PFV (slot_index=0 を優先).
    by_weekday: dict[int, PatientFixedVisit] = {}
    for pfv in pfvs:
        if pfv.slot_index != 0:
            logger.warning(
                "PFV slot_index!=0 を「編集用」シートで検出 (patient=%s weekday=%d slot=%d): "
                "1 行形式では slot_index=0 のみ採用します (将来対応)",
                patient.id,
                pfv.weekday,
                pfv.slot_index,
            )
            continue
        existing = by_weekday.get(pfv.weekday)
        if existing is not None:
            logger.warning(
                "PFV 同一曜日に複数行 (patient=%s weekday=%d): 1 行目を採用します",
                patient.id,
                pfv.weekday,
            )
            continue
        by_weekday[pfv.weekday] = pfv

    # 代表 duration / time_type を最初の曜日 (昇順) から導く.
    rep_duration: int | None = None
    durations_seen: set[int] = set()
    for wd in sorted(by_weekday):
        durations_seen.add(by_weekday[wd].duration_min)
        if rep_duration is None:
            rep_duration = by_weekday[wd].duration_min
    if len(durations_seen) > 1:
        logger.warning(
            "PFV duration が曜日で不一致 (patient=%s durations=%s): 最初の曜日の値 %s を採用",
            patient.id,
            sorted(durations_seen),
            rep_duration,
        )
    rep_tt = _resolve_time_type(patient, sorted(by_weekday)[0]) if by_weekday else None

    values: dict[str, object | None] = {
        "patient_id": str(patient.id),
        "patient_code": patient.code,
        "patient_name": patient.name,
        "service_minutes": rep_duration,
        "time_type": rep_tt or DEFAULT_TIME_TYPE,
        "delete_flag": None,
    }

    for wd_key in PFV_EDIT_WEEKDAY_KEYS:
        wd_int = PFV_EDIT_WEEKDAY_KEY_TO_INT[wd_key]
        pfv = by_weekday.get(wd_int)
        if pfv is None:
            values[f"{wd_key}_time"] = None
            values[f"{wd_key}_course"] = None
            continue
        values[f"{wd_key}_time"] = _hhmm(pfv.start_time)
        # course_template → 拠点付きトークン. クロスオフィス参照は空欄 + warning.
        token: str | None = None
        if pfv.course_template_id is not None:
            ct = course_template_by_id.get(pfv.course_template_id)
            if ct is not None and ct.label:
                office_code = office_code_by_id.get(ct.office_id)
                if office_code:
                    token = course_token(office_code, ct.label, code_to_short)
                if token is None:
                    msg = (
                        f"PFV course_template の拠点コードが解決できません "
                        f"(patient={patient.id} weekday={wd_int} template_office={ct.office_id} "
                        f"label={ct.label}) — コースを空欄化"
                    )
                    logger.warning(msg)
                    if crossoffice_warnings is not None:
                        crossoffice_warnings.append(msg)
        values[f"{wd_key}_course"] = token

    for col_key, col_idx in PFV_EDIT_COL_INDEX.items():
        ws.cell(row=row_idx, column=col_idx + 1, value=values.get(col_key))


def _resolve_time_type(patient: Patient, weekday: int) -> str | None:
    """patient.weekly_pattern.entries[].weekday と一致するエントリの time_type を返す.

    weekly_pattern が辞書で entries が無い場合は patient.weekly_pattern.time_type を返す.
    どちらも無い場合は ``None``.
    """
    wp = patient.weekly_pattern
    if not isinstance(wp, dict):
        return None
    short = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")[weekday]
    entries = wp.get("entries") or []
    if isinstance(entries, list):
        for entry in entries:
            if isinstance(entry, dict) and entry.get("weekday") == short:
                tt = entry.get("time_type")
                if isinstance(tt, str):
                    return tt
    tt_root = wp.get("time_type")
    return tt_root if isinstance(tt_root, str) else None


def _course_token_dropdown_values(
    offices: Sequence[Office],
    course_templates: Sequence[CourseTemplate],
) -> list[str]:
    """course_templates (office×label) から拠点付きコーストークンの dropdown 値を作る.

    例 ["稲A", "稲B", ..., "稲M", "津A", "津M"]. Phase G-57: 並びは
    GRID_OFFICE_ORDER (稲毛→都賀) → GRID_COURSE_ORDER (A,B,C,D,E,M) 順.
    単純な sorted() だと Unicode 順で「津」が「稲」より先に来てしまうため、
    グリッドのブロック順と同じ規則で明示的に並べる (= 稲毛 A から始まる).
    """
    office_code_by_id: dict[UUID, str] = {o.id: o.code for o in offices if o.code}
    # 0059: 拠点マスタ (offices.short_label) 駆動で短縮名を解決する.
    code_to_short, _ = build_office_code_short_maps((o.code, o.short_label) for o in offices)
    by_office: dict[str, set[str]] = {}
    for ct in course_templates:
        code = office_code_by_id.get(ct.office_id)
        if not code or not ct.label:
            continue
        by_office.setdefault(code, set()).add(ct.label)
    ordered_offices = [c for c in GRID_OFFICE_ORDER if c in by_office]
    ordered_offices += sorted(c for c in by_office if c not in GRID_OFFICE_ORDER)
    out: list[str] = []
    for code in ordered_offices:
        labels = by_office[code]
        ordered_labels = [lbl for lbl in GRID_COURSE_ORDER if lbl in labels]
        ordered_labels += sorted(lbl for lbl in labels if lbl not in GRID_COURSE_ORDER)
        for lbl in ordered_labels:
            token = course_token(code, lbl, code_to_short)
            if token is not None:
                out.append(token)
    return out


def _attach_course_dropdowns_to_edit_sheet(
    ws: Worksheet,
    tokens: list[str],
    *,
    max_data_rows: int = 1000,
) -> None:
    """「編集用」シートの各 ``{曜日}_course`` 列に拠点付きコース dropdown を設定."""
    if not tokens:
        return
    last_row = 1 + max_data_rows
    formula = '"' + ",".join(tokens) + '"'
    for wd_key in PFV_EDIT_WEEKDAY_KEYS:
        col_idx = PFV_EDIT_COL_INDEX[f"{wd_key}_course"] + 1
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "リストにないコースです"
        dv.errorTitle = "入力エラー"
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)


def _patient_condition_label(patient: Patient | None) -> str | None:
    """グリッドの「条件」列に出す患者条件 (性別制限) を返す.

    sex_restriction が設定されていれば「女性のみ / 男性のみ」を返す. 無ければ None.
    """
    if patient is None or not patient.sex_restriction:
        return None
    return SEX_RESTRICTION_EN_TO_JA.get(patient.sex_restriction, patient.sex_restriction)


def _patient_multi_label(patient: Patient | None) -> str | None:
    """グリッドの「複数」列に出すラベルを返す (requires_multiple_staff → "複数")."""
    if patient is not None and patient.requires_multiple_staff:
        return "複数"
    return None


def _slot_index_for_time(t: time | None) -> int | None:
    """start_time を 30 分グリッドのスロット index (0..17) に丸める.

    GRID_TIME_SLOTS (09:30〜18:00 を 30 分刻み) のうち最近接スロットを返す.
    範囲外 (09:30 未満 / 18:00 超過) は端のスロットにクランプする.
    """
    if t is None:
        return None
    minutes = t.hour * 60 + t.minute
    best_idx: int | None = None
    best_diff: int | None = None
    for idx, (h, m) in enumerate(GRID_TIME_SLOTS):
        diff = abs(minutes - (h * 60 + m))
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_idx = idx
    return best_idx


def _build_grid_sheet(
    ws: Worksheet,
    *,
    patients: Sequence[Patient],
    pfvs: Sequence[PatientFixedVisit],
    offices: Sequence[Office],
    course_templates: Sequence[CourseTemplate],
) -> None:
    """Phase G-55: 静的グリッド集計シート (閲覧専用スナップショット) を描画.

    参照元「スケジュール枠組み（仮）」と同形式: コースごとの縦ブロックを
    稲毛(INAGE)→都賀(TSUGA)、各拠点内 A,B,C,D,E,M 順に並べる. 1 ブロック = 6 日分
    (月〜土), 1 日 = 5 列 [時間帯, 氏名, 住所, 複数, 条件]. 日ブロックは
    col 1,6,11,16,21,26 開始.
      row1: 各日 [日合計(全コース合計), 曜日名]  (日合計は各拠点の先頭ブロックのみ)
      row2: 各日 [コース件数(office×course×曜日), コース文字]
      row3: サブ見出し [時間帯, 氏名, 住所, 複数, 条件] ×6
      row4〜row21: 09:30〜18:00 を 30 分刻み (18 スロット) の時刻 + その枠の患者

    A1 = 月曜の日合計 (= 稲毛先頭ブロック row1 col1). A1 には閲覧専用コメントを添える.
    importer はこのシート (シート名 SHEET_PFV_GRID) を一切読まない.
    """
    patient_by_id: dict[UUID, Patient] = {p.id: p for p in patients}
    office_code_by_id: dict[UUID, str] = {o.id: o.code for o in offices if o.code}
    # Phase G-57: ブロックが稲毛/都賀どちらか判別できるよう、拠点名を表示する.
    # office_code (INAGE/TSUGA) → 拠点名 (稲毛/都賀). name 未設定なら code を使う.
    office_name_by_code: dict[str, str] = {o.code: (o.name or o.code) for o in offices if o.code}

    num_days = len(GRID_WEEKDAY_FULL_LABELS)  # 6 (月〜土)

    # ブロック順序を組み立てる: (office_code, course_label, course_template_id?).
    # 拠点 = 稲毛→都賀 (GRID_OFFICE_ORDER), 各拠点内 = A,B,C,D,E,M (GRID_COURSE_ORDER) で
    # 存在する course_template のみ. 順序外の office / label は末尾に安定追加する.
    by_office: dict[str, dict[str, UUID]] = {}
    for ct in course_templates:
        code = office_code_by_id.get(ct.office_id)
        if not code or not ct.label:
            continue
        by_office.setdefault(code, {})[ct.label] = ct.id

    ordered_offices = [c for c in GRID_OFFICE_ORDER if c in by_office]
    ordered_offices += sorted(c for c in by_office if c not in GRID_OFFICE_ORDER)

    # (office_code, label, course_template_id) のブロック並び + 各拠点の先頭ブロック判定.
    blocks: list[tuple[str, str, UUID]] = []
    office_first_block_index: dict[str, int] = {}
    for code in ordered_offices:
        labels = by_office[code]
        ordered_labels = [lbl for lbl in GRID_COURSE_ORDER if lbl in labels]
        ordered_labels += sorted(lbl for lbl in labels if lbl not in GRID_COURSE_ORDER)
        for lbl in ordered_labels:
            if code not in office_first_block_index:
                office_first_block_index[code] = len(blocks)
            blocks.append((code, lbl, labels[lbl]))

    # 集計: (course_template_id, weekday, slot_index) → list[patient].
    # 日合計 (全コース): (weekday) → 患者数 (normal PFV 件数, 月〜土のみ).
    cell_patients: dict[tuple[UUID, int, int], list[Patient]] = {}
    course_count: dict[tuple[UUID, int], int] = {}
    day_total: dict[int, int] = dict.fromkeys(range(num_days), 0)
    for pfv in pfvs:
        if pfv.mode != "normal":
            continue
        if pfv.weekday not in range(num_days):
            continue  # 月〜土のみ (日曜は対象外)
        day_total[pfv.weekday] += 1
        if pfv.course_template_id is None:
            continue
        course_count[(pfv.course_template_id, pfv.weekday)] = (
            course_count.get((pfv.course_template_id, pfv.weekday), 0) + 1
        )
        slot = _slot_index_for_time(pfv.start_time)
        if slot is None:
            continue
        p = patient_by_id.get(pfv.patient_id)
        if p is not None:
            cell_patients.setdefault((pfv.course_template_id, pfv.weekday, slot), []).append(p)

    # ---- styling ----
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    daytotal_font = Font(bold=True, color="FF9C0006")  # 赤系 (日合計を強調)
    dayname_font = Font(bold=True, color=HEADER_FONT_COLOR)
    dayname_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    course_font = Font(bold=True)
    course_fill = PatternFill("solid", fgColor=ID_COLUMN_FILL_COLOR)
    subheader_font = Font(bold=True, color="FF333333")
    subheader_fill = PatternFill("solid", fgColor="FFDDEBF7")  # 薄い青
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    time_align = Alignment(horizontal="center", vertical="top")

    # 列幅: 時間帯=9, 氏名=20, 住所=34, 複数=6, 条件=10 を各日繰り返す.
    day_widths = (9, 20, 34, 6, 10)
    for d in range(num_days):
        for k, w in enumerate(day_widths):
            col = d * GRID_COLS_PER_DAY + 1 + k
            ws.column_dimensions[get_column_letter(col)].width = w

    def _day_start_col(day_idx: int) -> int:
        return day_idx * GRID_COLS_PER_DAY + 1  # 1,6,11,16,21,26

    # ---- 各ブロックを描画 ----
    block_start_row = 1
    prev_office: str | None = None
    for b_index, (code, label, ct_id) in enumerate(blocks):
        # 拠点グループが変わったら追加の空き行を 1 行入れる (参照元同様).
        if prev_office is not None and code != prev_office:
            block_start_row += GRID_OFFICE_GAP
        prev_office = code

        is_office_first = office_first_block_index.get(code) == b_index
        r1 = block_start_row  # 曜日名 + 日合計
        r2 = block_start_row + 1  # コース件数 + コース文字
        r3 = block_start_row + 2  # サブ見出し
        slot_row0 = block_start_row + 3  # スロット先頭行

        for d in range(num_days):
            c0 = _day_start_col(d)  # 時間帯列
            # row1: [日合計, 曜日名]. 日合計は各拠点の先頭ブロックのみ.
            if is_office_first:
                dt = ws.cell(row=r1, column=c0, value=day_total.get(d, 0))
                dt.font = daytotal_font
                dt.alignment = center
            dn = ws.cell(row=r1, column=c0 + 1, value=GRID_WEEKDAY_FULL_LABELS[d])
            dn.font = dayname_font
            dn.fill = dayname_fill
            dn.alignment = center
            # row2: [コース件数, コース文字].
            cc = ws.cell(row=r2, column=c0, value=course_count.get((ct_id, d), 0))
            cc.font = course_font
            cc.alignment = center
            lc = ws.cell(row=r2, column=c0 + 1, value=label)
            lc.font = course_font
            lc.fill = course_fill
            lc.alignment = center
            # Phase G-57: コース行の住所列(c0+2)に拠点名(稲毛/都賀)を表示し、
            # 稲毛-A / 都賀-A の判別を可能にする.
            ofc = ws.cell(row=r2, column=c0 + 2, value=office_name_by_code.get(code, code))
            ofc.font = course_font
            ofc.fill = course_fill
            ofc.alignment = center
            # row3: サブ見出し [時間帯, 氏名, 住所, 複数, 条件].
            for k, sub in enumerate(GRID_DAY_SUBHEADERS):
                sh = ws.cell(row=r3, column=c0 + k, value=sub)
                sh.font = subheader_font
                sh.fill = subheader_fill
                sh.alignment = center
                sh.border = border
            # row4〜: 30 分スロット.
            for s, (h, m) in enumerate(GRID_TIME_SLOTS):
                rr = slot_row0 + s
                tc = ws.cell(row=rr, column=c0, value=f"{h:02d}:{m:02d}")
                tc.alignment = time_align
                tc.border = border
                ps = cell_patients.get((ct_id, d, s), [])
                names = "\n".join(p.name for p in ps if p.name)
                addrs = "\n".join(p.address or "" for p in ps)
                multi = "\n".join(_patient_multi_label(p) or "" for p in ps)
                cond = "\n".join(_patient_condition_label(p) or "" for p in ps)
                nc = ws.cell(row=rr, column=c0 + 1, value=names or None)
                nc.alignment = wrap
                nc.border = border
                ac = ws.cell(row=rr, column=c0 + 2, value=addrs.strip("\n") or None)
                ac.alignment = wrap
                ac.border = border
                mc = ws.cell(row=rr, column=c0 + 3, value=multi.strip("\n") or None)
                mc.alignment = wrap
                mc.border = border
                oc = ws.cell(row=rr, column=c0 + 4, value=cond.strip("\n") or None)
                oc.alignment = wrap
                oc.border = border

        block_start_row += GRID_BLOCK_HEIGHT + 1  # ブロック間 1 行の空き

    # A1 = 月曜 (day 0) の日合計. ブロックが 1 つも無い場合は 0.
    a1 = ws.cell(row=1, column=1, value=day_total.get(0, 0))
    a1.font = daytotal_font
    a1.alignment = center
    a1_comment = Comment(PFV_GRID_A1_COMMENT_TEXT, COMMENT_AUTHOR)
    a1_comment.width = 320
    a1_comment.height = 110
    a1.comment = a1_comment

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# public entrypoints
# ---------------------------------------------------------------------------


def build_workbook(
    *,
    patients: Sequence[Patient],
    pfvs: Sequence[PatientFixedVisit],
    offices: Sequence[Office],
    course_templates: Sequence[CourseTemplate],
    crossoffice_warnings_out: list[str] | None = None,
    ng_staff_codes_by_patient: dict[UUID, str] | None = None,
) -> Workbook:
    """テンプレート + データ込みのワークブックを構築する.

    ``patients`` を空 list で呼べばテンプレート (ヘッダー + dropdown のみ) になる.

    Phase E-7 (gap P2): ``crossoffice_warnings_out`` が渡された場合、PFV のクロス
    オフィス course_template 参照 (患者拠点と template 拠点が一致しない参照) の
    warning メッセージを append する. 呼び出し側 (API endpoint) は件数を response
    header (X-Excel-Crossoffice-Warnings-Count) として返し、operator が export 結果
    から気付けるようにする.
    """
    wb = Workbook()
    # default のシートを 1 枚目として使う
    ws_p: Worksheet = wb.active  # type: ignore[assignment]
    ws_p.title = SHEET_PATIENTS

    _set_header_row(ws_p, PATIENT_COLUMNS)
    _attach_dropdowns(ws_p, PATIENT_COLUMNS)
    _attach_header_comment(ws_p, "patient_id", PATIENT_COLUMNS, text=PATIENT_ID_COMMENT_TEXT)
    # Phase G-48: 緯度/経度は住所からの自動算出 (派生列). 編集不要コメントを付与.
    _attach_header_comment(ws_p, "lat", PATIENT_COLUMNS, text=LATLNG_COMMENT_TEXT)
    _attach_header_comment(ws_p, "lng", PATIENT_COLUMNS, text=LATLNG_COMMENT_TEXT)
    # Phase G-49: 拠点コードは住所から自動割当 (派生列). 編集不要コメントを付与.
    _attach_header_comment(ws_p, "office_code", PATIENT_COLUMNS, text=OFFICE_CODE_COMMENT_TEXT)
    # Phase G-49: 時間タイプ 午前/午後/終日 のとき時刻 2 列を条件付き書式でグレーアウト.
    _attach_time_greyout_conditional_format(ws_p, PATIENT_COLUMNS)

    office_code_by_id: dict[UUID, str] = {
        office.id: (office.code or "")
        for office in offices
        if office.code  # コード未設定の拠点はスキップ
    }
    # 0059: 拠点マスタ (offices.short_label) 駆動の code→短縮名 map (PFV 編集用シート).
    code_to_short, _ = build_office_code_short_maps(
        (office.code, office.short_label) for office in offices
    )
    for i, patient in enumerate(patients, start=2):
        _write_patient_row(
            ws_p,
            i,
            patient,
            office_code_by_id=office_code_by_id,
            ng_staff_codes_by_patient=ng_staff_codes_by_patient,
        )
    # Phase G-48/G-49: patient_id / 緯度 / 経度 / 拠点コード の派生列はグレー塗りで
    # 「触らない雰囲気」.
    _shade_id_column_data_rows(ws_p, "patient_id", PATIENT_COLUMNS, data_row_count=len(patients))
    _shade_id_column_data_rows(ws_p, "lat", PATIENT_COLUMNS, data_row_count=len(patients))
    _shade_id_column_data_rows(ws_p, "lng", PATIENT_COLUMNS, data_row_count=len(patients))
    _shade_id_column_data_rows(ws_p, "office_code", PATIENT_COLUMNS, data_row_count=len(patients))

    # ---- 固定訪問パターン（編集用）シート (Phase G-51: 患者 1 行) ----
    ws_f: Worksheet = wb.create_sheet(title=SHEET_PFV_EDIT)
    _set_header_row(ws_f, PFV_EDIT_COLUMNS)
    _attach_dropdowns(ws_f, PFV_EDIT_COLUMNS)
    _attach_header_comment(
        ws_f, "patient_id", PFV_EDIT_COLUMNS, text=PFV_EDIT_PATIENT_ID_COMMENT_TEXT
    )
    # コース dropdown は course_templates から動的生成 (拠点付きトークン).
    course_tokens = _course_token_dropdown_values(offices, course_templates)
    _attach_course_dropdowns_to_edit_sheet(ws_f, course_tokens)

    course_template_by_id: dict[UUID, CourseTemplate] = {ct.id: ct for ct in course_templates}
    # Phase E-7 (gap P2): クロスオフィス参照 warning 収集用.
    crossoffice_warnings: list[str] = (
        crossoffice_warnings_out if crossoffice_warnings_out is not None else []
    )

    # normal mode PFV を患者ごとに集約. 行 = normal PFV を持つ患者のみ (確定週次固定枠の正本).
    normal_pfvs_by_patient: dict[UUID, list[PatientFixedVisit]] = {}
    for pfv in pfvs:
        if pfv.mode != "normal":
            continue
        normal_pfvs_by_patient.setdefault(pfv.patient_id, []).append(pfv)

    patient_by_id: dict[UUID, Patient] = {p.id: p for p in patients}
    # patient.code 順で安定出力 (export 順序の決定性).
    ordered_pids = sorted(
        normal_pfvs_by_patient.keys(),
        key=lambda pid: (patient_by_id[pid].code or "") if pid in patient_by_id else "",
    )
    row_idx = 2
    edit_row_count = 0
    for pid in ordered_pids:
        patient = patient_by_id.get(pid)
        if patient is None:
            continue  # alive patient のみ (export は alive のみ渡される想定)
        _write_pfv_edit_patient_row(
            ws_f,
            row_idx,
            patient,
            normal_pfvs_by_patient[pid],
            course_template_by_id=course_template_by_id,
            office_code_by_id=office_code_by_id,
            code_to_short=code_to_short,
            crossoffice_warnings=crossoffice_warnings,
        )
        row_idx += 1
        edit_row_count += 1
    _shade_id_column_data_rows(ws_f, "patient_id", PFV_EDIT_COLUMNS, data_row_count=edit_row_count)

    if crossoffice_warnings:
        logger.warning(
            "patient_excel export: 解決できない course_template 拠点参照を %d 件 空欄化しました.",
            len(crossoffice_warnings),
        )

    # ---- グリッド集計（静的・閲覧専用）シート (Phase G-51) ----
    # importer はこのシートを一切読まない (シート名で識別して無視).
    ws_grid: Worksheet = wb.create_sheet(title=SHEET_PFV_GRID)
    _build_grid_sheet(
        ws_grid,
        patients=patients,
        pfvs=pfvs,
        offices=offices,
        course_templates=course_templates,
    )

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
