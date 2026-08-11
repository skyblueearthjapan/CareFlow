"""Excel パース + 差分計算 + DB 反映.

エントリポイント:
  * ``parse_and_diff`` — Excel bytes → (患者行差分, PFV 行差分, summary).
    プレビュー (dry_run=True) でも apply (dry_run=False) でも事前にこの関数で
    差分計算する.
  * ``apply_changes`` — ``parse_and_diff`` の戻り値を受け取り、DB に書き戻す.
    1 transaction で INSERT / UPDATE / DELETE をまとめて実行 (partial commit).
    error 行は ``parse_and_diff`` の時点で ``op=None`` として除外されているので
    ``apply_changes`` には積まれず、自動的に skip される.

差分の表現:
  * ``PatientExcelImportRow`` — patient sheet の 1 行に対応.
  * ``PfvExcelImportRow`` — pfv sheet の 1 行に対応.
  * いずれも ``operation`` が "new" / "update" / "delete" / "noop" / "error".
"""

from __future__ import annotations

import logging
import uuid
from collections.abc import Sequence
from datetime import datetime, time
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.course_template import CourseTemplate
from app.models.office import Office
from app.models.patient import Patient
from app.models.patient_fixed_visit import PatientFixedVisit
from app.models.patient_ng_staff import PatientNgStaff
from app.models.patient_same_address_link import PatientSameAddressLink
from app.models.staff import Staff
from app.schemas.v2.patient_excel import (
    PatientExcelChange,
    PatientExcelImportRow,
    PatientExcelImportSummary,
    PfvExcelImportRow,
)
from app.services.office_assigner import PreloadedOfficeCities, load_office_cities_index
from app.services.patient_excel.schema import (
    DEFAULT_TIME_TYPE,
    INSURANCE_JA_TO_EN,
    INSURANCE_VALUES,
    LEGACY_SHEET_PFV,
    PATIENT_COL_INDEX,
    PFV_COL_INDEX,
    PFV_EDIT_COL_INDEX,
    PFV_EDIT_WEEKDAY_KEY_TO_INT,
    PFV_EDIT_WEEKDAY_KEYS,
    PFV_MODE_JA_TO_EN,
    PFV_MODE_VALUES,
    SEX_JA_TO_EN,
    SEX_RESTRICTION_JA_TO_EN,
    SEX_RESTRICTION_JA_VALUES,
    SEX_RESTRICTION_NONE_JA,
    SEX_RESTRICTION_VALUES,
    SEX_VALUES,
    SHEET_PATIENTS,
    SHEET_PFV,
    SHEET_PFV_EDIT,
    SHEET_WEEKLY,
    STATUS_JA_TO_EN,
    STATUS_VALUES,
    TIME_TYPE_VALUES,
    VISIT_FREQUENCY_JA_TO_EN,
    VISIT_FREQUENCY_VALUES,
    WEEKDAY_EN_LIST,
    WEEKDAY_LABEL_TO_INT,
    WEEKLY_COL_INDEX,
    build_office_code_short_maps,
    is_magic_clear,
    is_magic_delete,
    parse_course_token,
    weekday_yesno_cells_to_en,
    weekdays_cell_to_en,
)

logger = logging.getLogger(__name__)


# Phase G-48: 日本語ラベル → DB 英語 enum 値への変換マップ.
# importer は日本語ラベル / 英語正準値の双方を受理する (後方互換).
#   * key を日本語ラベルとした dict に英語正準値を自己写像で追加することで、
#     1 つの dict で「日本語 or 英語」両受理を実現する.
def _bidir_enum_map(ja_to_en: dict[str, str], en_values: tuple[str, ...]) -> dict[str, str]:
    """日本語→英語 map に英語値の自己写像を足して「双方向受理」 map を作る."""
    merged = dict(ja_to_en)
    for en in en_values:
        merged.setdefault(en, en)
    return merged


SEX_ANY_TO_EN: dict[str, str] = _bidir_enum_map(SEX_JA_TO_EN, SEX_VALUES)
STATUS_ANY_TO_EN: dict[str, str] = _bidir_enum_map(STATUS_JA_TO_EN, STATUS_VALUES)
INSURANCE_ANY_TO_EN: dict[str, str] = _bidir_enum_map(INSURANCE_JA_TO_EN, INSURANCE_VALUES)
SEX_RESTRICTION_ANY_TO_EN: dict[str, str] = _bidir_enum_map(
    SEX_RESTRICTION_JA_TO_EN, SEX_RESTRICTION_VALUES
)
PFV_MODE_ANY_TO_EN: dict[str, str] = _bidir_enum_map(PFV_MODE_JA_TO_EN, PFV_MODE_VALUES)
# 訪問頻度: 日本語ラベル + 英語正準値に加え、旧バグで日本語が DB に保存され得る
# ケースを吸収するため、英語値も自己写像で受理する.
VISIT_FREQUENCY_ANY_TO_EN: dict[str, str] = _bidir_enum_map(
    VISIT_FREQUENCY_JA_TO_EN, VISIT_FREQUENCY_VALUES
)

# ---------------------------------------------------------------------------
# cell value helpers
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    """空セルの判定. openpyxl は空セルを None で返すが、空文字 / 空白文字も空扱い."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _read_str(value: Any) -> str | None:
    """セル値を str に正規化する (空セルは None)."""
    if _is_blank(value):
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _read_int(value: Any) -> int | None:
    """セル値を int に正規化. 失敗時は ValueError."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        # Excel で bool が出ることはほぼ無いが念のため
        raise ValueError(f"int 型として読めません: {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"int 型として読めません (小数): {value!r}")
    s = str(value).strip()
    return int(s)


def _read_float(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return float(str(value).strip())


def _read_uuid(value: Any) -> UUID | None:
    if _is_blank(value):
        return None
    if isinstance(value, UUID):
        return value
    return UUID(str(value).strip())


def _read_bool(value: Any) -> bool | None:
    """TRUE/FALSE (case-insensitive) を bool に正規化. 失敗時は ValueError.

    Phase E-7 (gap P0-1): requires_multiple_staff 用. staff_excel/importer の
    同名関数と完全に同じ仕様.
    """
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    raw = str(value).strip()
    # Phase G-48: 日本語ラベル「はい/いいえ」も受理.
    if raw in ("はい",):
        return True
    if raw in ("いいえ",):
        return False
    s = raw.upper()
    if s in ("TRUE", "1", "YES", "Y"):
        return True
    if s in ("FALSE", "0", "NO", "N"):
        return False
    raise ValueError(f"bool として読めません: {value!r}")


def _read_hhmm(value: Any) -> str | None:
    """セル値を "HH:MM" 文字列として正規化. 失敗時は ValueError."""
    if _is_blank(value):
        return None
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    s = str(value).strip()
    # ``HH:MM`` または ``HH:MM:SS``
    parts = s.split(":")
    if len(parts) < 2:
        raise ValueError(f"HH:MM 形式ではありません: {s!r}")
    h = int(parts[0])
    m = int(parts[1])
    if not (0 <= h <= 23 and 0 <= m <= 59):
        raise ValueError(f"時刻の範囲外: {s!r}")
    return f"{h:02d}:{m:02d}"


def _hhmm_to_time(s: str) -> time:
    h, m = s.split(":")[:2]
    return time(int(h), int(m))


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return all(_is_blank(v) for v in row)


# ---------------------------------------------------------------------------
# lookup builders
# ---------------------------------------------------------------------------


async def _load_offices_by_code(db: AsyncSession) -> dict[str, Office]:
    """code → Office (deleted_at IS NULL のみ)."""
    rows = (
        await db.scalars(
            select(Office).where(
                Office.deleted_at.is_(None),
                Office.code.is_not(None),
            )
        )
    ).all()
    return {str(o.code): o for o in rows}


async def _load_course_templates_by_label(
    db: AsyncSession,
) -> dict[tuple[UUID, str], CourseTemplate]:
    """(office_id, label) → CourseTemplate (deleted_at IS NULL のみ).

    label は label そのもの (大文字小文字区別あり) で使う.
    """
    rows = (
        await db.scalars(select(CourseTemplate).where(CourseTemplate.deleted_at.is_(None)))
    ).all()
    return {(ct.office_id, ct.label): ct for ct in rows}


async def _load_patients_by_id(db: AsyncSession) -> dict[UUID, Patient]:
    rows = (await db.scalars(select(Patient).where(Patient.deleted_at.is_(None)))).all()
    return {p.id: p for p in rows}


async def _load_patient_codes(db: AsyncSession) -> set[str]:
    """既存 (alive) の patient.code 集合 (重複チェック用)."""
    rows = (
        await db.execute(
            select(Patient.code).where(Patient.deleted_at.is_(None), Patient.code.is_not(None))
        )
    ).all()
    return {r[0] for r in rows}


async def _load_deleted_patients_by_code(db: AsyncSession) -> dict[str, Patient]:
    """soft-deleted 患者の code → Patient マップ (resurrection 用).

    DB は ``code`` 列に UNIQUE 制約があるため、同じ code を再 INSERT すると
    UniqueViolation で 409 になる。新規候補行 (patient_id 空 + patient_code 入り) が
    DB に soft-deleted で存在する code なら、INSERT ではなく
    既存行の deleted_at を NULL に戻して内容を上書きする (= 復活).
    """
    rows = (
        await db.scalars(
            select(Patient).where(
                Patient.deleted_at.is_not(None),
                Patient.code.is_not(None),
            )
        )
    ).all()
    return {p.code: p for p in rows}


async def _load_pfvs_by_key(
    db: AsyncSession,
    alive_patient_ids: set[UUID],
) -> dict[tuple[UUID, str, int, int], PatientFixedVisit]:
    """(patient_id, mode, weekday, slot_index) → PatientFixedVisit.

    alive な患者の PFV のみ load する (全 PFV bulk SELECT を避けるため).
    """
    if not alive_patient_ids:
        return {}
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id.in_(alive_patient_ids))
        )
    ).all()
    return {(p.patient_id, p.mode, p.weekday, p.slot_index): p for p in rows}


# ---------------------------------------------------------------------------
# NG スタッフ (patient_ng_staff 中間テーブル) — 共通ヘルパー
#
# セマンティクスは staff_excel の secondary_office_codes 列に揃える:
#   空セル = 維持 / <CLEAR> = 全解除 / カンマ区切り = その集合に一致させる.
# replace_all.py もこれらのヘルパーを import して同じ挙動を共有する.
# ---------------------------------------------------------------------------


async def _load_staff_by_code(db: AsyncSession) -> dict[str, Staff]:
    """staff.code → Staff (退職済みも含む).

    退職済み (deleted_at IS NOT NULL) を含めるのは、既存 NG 行が退職後も残るため.
    export → そのまま import の round-trip で退職者の既存行を「不明コード」扱いに
    してしまわないよう、存在判定には退職者も載せる. 退職者の **新規** 指定だけを
    後段 (:func:`_resolve_ng_staff_ids`) で skip する.
    """
    rows = (await db.scalars(select(Staff).where(Staff.code.is_not(None)))).all()
    return {str(s.code): s for s in rows}


async def _load_ng_staff_by_patient(db: AsyncSession) -> dict[UUID, set[UUID]]:
    """patient_id → NG スタッフ staff_id 集合 (全患者分を 1 クエリ)."""
    rows = (await db.execute(select(PatientNgStaff.patient_id, PatientNgStaff.staff_id))).all()
    out: dict[UUID, set[UUID]] = {}
    for patient_id, staff_id in rows:
        out.setdefault(patient_id, set()).add(staff_id)
    return out


def _parse_ng_staff_cell(
    raw: Any,
    *,
    staff_by_code: dict[str, Staff],
    row_number: int,
) -> list[tuple[UUID, bool]] | None:
    """NG スタッフ列のセルを ``[(staff_id, staff_is_deleted), ...]`` に解決する.

    戻り値 ``None`` = 「この列は触らない」(空セル / 列そのものが無い旧ファイル).
    空 list = 全解除 (``<CLEAR>`` もしくは区切り文字のみ).

    ``<CLEAR>`` は通常 import / 完全置換 import の **どちらでも** 有効にする.
    完全置換でも NG 列は「空セル = 維持」の merge 例外扱いなので、明示的な
    全解除手段が他に無いため.

    不明コードは error にせず warning + skip (= secondary_office_codes と同じ
    ベストエフォート方針. バックアップ復元運用で 1 セルのタイプミスが全体を
    落とさないようにするため).
    """
    if _is_blank(raw):
        return None
    if is_magic_clear(raw):
        return []
    raw_str = _read_str(raw) or ""
    codes = [c.strip() for c in raw_str.split(",") if c.strip()]
    resolved: list[tuple[UUID, bool]] = []
    unknown: list[str] = []
    for code in codes:
        staff = staff_by_code.get(code)
        if staff is None:
            unknown.append(code)
            continue
        resolved.append((staff.id, staff.deleted_at is not None))
    if unknown:
        logger.warning(
            "patient_excel import row=%d: NG スタッフの不明コード %r を skip しました",
            row_number,
            unknown,
        )
    return resolved


def _resolve_ng_staff_ids(
    resolved: list[tuple[UUID, bool]],
    current_ids: set[UUID],
    *,
    row_number: int,
) -> list[UUID]:
    """解決済み候補を「実際に設定する staff_id list」へ絞り込む (重複除去 + 順序保持).

    退職済みスタッフの **新規** 指定は warning + skip (CRUD API が 422 で弾くのと
    整合させる). 既に NG 行が存在するスタッフはそのまま維持する (export した
    退職者付き行をそのまま import して暗黙解除されないようにするため).
    """
    out: list[UUID] = []
    seen: set[UUID] = set()
    skipped_retired: list[str] = []
    for staff_id, is_deleted in resolved:
        if staff_id in seen:
            continue
        if is_deleted and staff_id not in current_ids:
            skipped_retired.append(str(staff_id))
            continue
        seen.add(staff_id)
        out.append(staff_id)
    if skipped_retired:
        logger.warning(
            "patient_excel import row=%d: 退職済みスタッフ %r の NG 新規指定を skip しました",
            row_number,
            skipped_retired,
        )
    return out


def _ng_codes_for_view(
    staff_ids: Sequence[UUID],
    staff_code_by_id: dict[UUID, str],
) -> list[str]:
    """diff 表示用: staff_id list → sorted な staff_code list (code 無しは UUID 文字列)."""
    return sorted(staff_code_by_id.get(sid) or str(sid) for sid in staff_ids)


def _build_ng_staff_change(
    resolved: list[tuple[UUID, bool]] | None,
    current_ids: set[UUID],
    *,
    row_number: int,
    staff_code_by_id: dict[UUID, str],
) -> tuple[PatientExcelChange, list[UUID]] | None:
    """NG スタッフの差分 (表示用 change + 目標 staff_id list) を返す. 差分無しなら None."""
    if resolved is None:
        return None  # 空セル = 維持
    target_ids = _resolve_ng_staff_ids(resolved, current_ids, row_number=row_number)
    if set(target_ids) == current_ids:
        return None
    change = PatientExcelChange(
        field="ng_staff",
        old_value=_ng_codes_for_view(sorted(current_ids, key=str), staff_code_by_id),
        new_value=_ng_codes_for_view(target_ids, staff_code_by_id),
    )
    return change, target_ids


async def apply_ng_staff_updates(
    db: AsyncSession,
    updates: Sequence[tuple[UUID, list[UUID]]],
) -> None:
    """``patient_ng_staff`` を目標集合に一致させる (差分 INSERT / DELETE).

    既存行は note / decided_by_user_id ごと維持する (Excel は note を扱わないため、
    集合に残るスタッフの note を潰さない). 新規行は note=NULL /
    decided_by_user_id=NULL で INSERT する (Excel 経由の設定者は記録しない).
    """
    for patient_id, target_ids in updates:
        current_ids = set(
            (
                await db.scalars(
                    select(PatientNgStaff.staff_id).where(PatientNgStaff.patient_id == patient_id)
                )
            ).all()
        )
        target_set = set(target_ids)
        to_delete = current_ids - target_set
        if to_delete:
            await db.execute(
                delete(PatientNgStaff).where(
                    PatientNgStaff.patient_id == patient_id,
                    PatientNgStaff.staff_id.in_(list(to_delete)),
                )
            )
        for staff_id in target_ids:
            if staff_id in current_ids:
                continue  # 既存行は note ごと維持
            db.add(PatientNgStaff(patient_id=patient_id, staff_id=staff_id))
    if updates:
        await db.flush()


async def cleanup_soft_deleted_patient_links(
    db: AsyncSession,
    patient_ids: Sequence[UUID],
) -> None:
    """soft-delete された患者の中間テーブル行を物理削除する.

    ``patients.deleted_at`` を立てるだけの soft delete では FK ``ON DELETE CASCADE``
    が発火しないため、``api/v1/patients.py::delete_patient`` と同じ流儀でアプリ層から
    明示 DELETE する (同じ code で患者が復活したときに古い紐付けが蘇るのを防ぐ):

      * :class:`PatientSameAddressLink` — patient_a_id / patient_b_id の **両側**.
      * :class:`PatientNgStaff` — patient_id 側.
    """
    if not patient_ids:
        return
    ids = list(patient_ids)
    await db.execute(
        delete(PatientSameAddressLink).where(
            or_(
                PatientSameAddressLink.patient_a_id.in_(ids),
                PatientSameAddressLink.patient_b_id.in_(ids),
            )
        )
    )
    await db.execute(delete(PatientNgStaff).where(PatientNgStaff.patient_id.in_(ids)))


# ---------------------------------------------------------------------------
# Patient sheet parser
# ---------------------------------------------------------------------------


def _parse_patient_row(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    existing_patients_by_code: dict[str, Patient],
    deleted_patients_by_code: dict[str, Patient],
    offices_by_code: dict[str, Office],
    existing_codes: set[str],
    already_seen_codes: set[str],
    office_index: PreloadedOfficeCities | None = None,
    staff_by_code: dict[str, Staff] | None = None,
    ng_staff_by_patient: dict[UUID, set[UUID]] | None = None,
    staff_code_by_id: dict[UUID, str] | None = None,
) -> tuple[PatientExcelImportRow, dict[str, Any] | None]:
    """1 行を差分行に変換する.

    戻り値の第 2 要素は apply 時に使う「DB 用 dict」. operation が "error" / "noop" /
    "delete" のときは None or 部分的に意味のあるもの.
    """
    staff_by_code = staff_by_code if staff_by_code is not None else {}
    ng_staff_by_patient = ng_staff_by_patient if ng_staff_by_patient is not None else {}
    staff_code_by_id = staff_code_by_id if staff_code_by_id is not None else {}
    cells: dict[str, Any] = {}
    for col_key, idx in PATIENT_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    # 共通: patient_id / patient_code を読む.
    raw_id = cells["patient_id"]
    raw_code = cells["patient_code"]
    raw_delete = cells["delete_flag"]

    # patient_id がある場合は UUID パース.
    try:
        patient_id = _read_uuid(raw_id) if not _is_blank(raw_id) else None
    except ValueError:
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=None,
                patient_code=_read_str(raw_code),
                operation="error",
                error_message=f"patient_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    patient_code = _read_str(raw_code)

    # 既存患者の検索.
    existing_patient: Patient | None = None
    if patient_id is not None:
        existing_patient = existing_patients.get(patient_id)
        if existing_patient is None:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=f"patient_id が DB に存在しません: {patient_id}",
                ),
                None,
            )

    # Phase G-48 (最重要): patient_id 空でも patient_code が既存 (alive) 患者に
    # 一致すれば UPDATE 突合する. これによりユーザーは export (code 入り) を直接
    # 編集して再アップするだけで更新できる (UUID 列が空 / 消えていても可).
    # soft-deleted 患者の code は existing_patients_by_code に含まれないため、
    # ここでは一致せず、後段の resurrection ロジックへ正しくフォールバックする.
    if existing_patient is None and patient_code is not None:
        existing_patient = existing_patients_by_code.get(patient_code)
        if existing_patient is not None:
            patient_id = existing_patient.id

    # 削除フラグ
    if is_magic_delete(raw_delete):
        # 1) patient_id があれば既存通り (上で resolution 済み)
        if existing_patient is not None:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=existing_patient.id,
                    patient_code=existing_patient.code,
                    operation="delete",
                ),
                {"_patient_id": existing_patient.id, "_op": "delete"},
            )
        # 2) patient_id 空でも patient_code があれば code で解決
        if patient_code is not None:
            resolved = existing_patients_by_code.get(patient_code)
            if resolved is None:
                # 該当 code が DB に居ない → idempotent な noop 扱い (再 import で
                # error にしないため). export → 1 行削除 → 再 import のようなケースを
                # 想定。
                return (
                    PatientExcelImportRow(
                        row_number=row_number,
                        patient_id=None,
                        patient_code=patient_code,
                        operation="noop",
                    ),
                    None,
                )
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=resolved.id,
                    patient_code=resolved.code,
                    operation="delete",
                ),
                {"_patient_id": resolved.id, "_op": "delete"},
            )
        # 3) 両方空 → error
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=None,
                patient_code=None,
                operation="error",
                error_message=(
                    "削除フラグが指定されましたが patient_id / patient_code がありません"
                ),
            ),
            None,
        )

    # 各列をパース.
    parsed: dict[str, Any] = {}
    errors: list[str] = []

    # name / kana / address / note: 文字列系
    for k in ("name", "kana", "address", "note"):
        v = cells[k]
        if _is_blank(v):
            parsed[k] = None
        elif is_magic_clear(v):
            parsed[k] = ("CLEAR", None)
        else:
            parsed[k] = ("SET", _read_str(v))

    # patient_code (現在値と diff したいので raw を保持)
    if _is_blank(raw_code):
        parsed["code"] = None
    else:
        parsed["code"] = ("SET", patient_code)

    # enum 系
    # Phase G-48: 日本語ラベル / 英語正準値の双方を受理し、DB 英語値へ正規化する.
    def _parse_enum(key: str, mapping: dict[str, str]) -> None:
        v = cells[key]
        if _is_blank(v):
            parsed[key] = None
            return
        if is_magic_clear(v):
            parsed[key] = ("CLEAR", None)
            return
        s = _read_str(v)
        en = mapping.get(s or "")
        if en is None:
            allowed = ",".join(dict.fromkeys(mapping.keys()))
            errors.append(f"列「{key}」の値が候補外: {s!r} (許容: {allowed})")
            return
        parsed[key] = ("SET", en)

    _parse_enum("sex", SEX_ANY_TO_EN)
    _parse_enum("status", STATUS_ANY_TO_EN)
    _parse_enum("insurance", INSURANCE_ANY_TO_EN)

    # sex_restriction: 「なし」(空相当) は DB NULL に倒す. CLEAR / 「なし」 → CLEAR.
    # それ以外は日本語/英語ラベルを英語値に正規化.
    raw_sr = cells["sex_restriction"]
    if _is_blank(raw_sr):
        parsed["sex_restriction"] = None
    elif is_magic_clear(raw_sr):
        parsed["sex_restriction"] = ("CLEAR", None)
    else:
        sr_s = _read_str(raw_sr)
        if sr_s == SEX_RESTRICTION_NONE_JA:
            # 「なし」 = 制限解除 (DB NULL). round-trip 安定性のため CLEAR 扱い.
            parsed["sex_restriction"] = ("CLEAR", None)
        else:
            sr_en = SEX_RESTRICTION_ANY_TO_EN.get(sr_s or "")
            if sr_en is None:
                allowed = ",".join(SEX_RESTRICTION_JA_VALUES)
                errors.append(f"列「sex_restriction」の値が候補外: {sr_s!r} (許容: {allowed})")
            else:
                parsed["sex_restriction"] = ("SET", sr_en)

    # lat / lng (float, 範囲チェック)
    for k, lo, hi in (("lat", -90.0, 90.0), ("lng", -180.0, 180.0)):
        v = cells[k]
        if _is_blank(v):
            parsed[k] = None
            continue
        if is_magic_clear(v):
            parsed[k] = ("CLEAR", None)
            continue
        try:
            f = _read_float(v)
        except (ValueError, TypeError):
            errors.append(f"列「{k}」が数値ではありません: {v!r}")
            continue
        if f is None or f < lo or f > hi:
            errors.append(f"列「{k}」が範囲外 [{lo}, {hi}]: {f!r}")
            continue
        parsed[k] = ("SET", f)

    # 拠点コード → primary_office_id
    # Phase G-49: コードが入っていれば従来どおり尊重 (手動上書き可). 空欄のときは
    # 「住所から自動割当」を試みる. ただし auto-assign は「新規 / クリア時」のみ適用し、
    # 既存患者の空欄は従来どおり「触らない」(下流の操作判定で None = 維持) とする.
    # round-trip 安定: export は常にコードを出力するため import はコードありで尊重 → noop.
    raw_office = cells["office_code"]
    auto_office_id: UUID | None = None  # 空欄時に住所から解決した office_id (新規用).
    if _is_blank(raw_office):
        parsed["primary_office_id"] = None
        if office_index is not None:
            addr_for_office = _read_str(cells.get("address"))
            auto_office_id = office_index.resolve_office_id(addr_for_office)
    elif is_magic_clear(raw_office):
        parsed["primary_office_id"] = ("CLEAR", None)
    else:
        oc = _read_str(raw_office)
        office = offices_by_code.get(oc) if oc else None
        if office is None:
            errors.append(f"拠点コードが DB に存在しません: {oc!r}")
        else:
            parsed["primary_office_id"] = ("SET", office.id)

    # Phase E-7 (gap P0-1): requires_multiple_staff. NOT NULL bool.
    # 空セル = 触らない (更新時) / default False (新規時). CLEAR は False と等価.
    raw_req_multi = cells["requires_multiple_staff"]
    if _is_blank(raw_req_multi):
        parsed["requires_multiple_staff"] = None
    elif is_magic_clear(raw_req_multi):
        parsed["requires_multiple_staff"] = ("SET", False)
    else:
        try:
            rm = _read_bool(raw_req_multi)
        except ValueError as exc:
            errors.append(f"列「requires_multiple_staff」が TRUE/FALSE ではありません: {exc}")
        else:
            parsed["requires_multiple_staff"] = ("SET", rm)

    # NG スタッフ (patient_ng_staff). relationship なので ``parsed`` には入れず
    # (新規行の flatten ループが Patient(**data) に渡してしまうため) 別変数で持つ.
    # None = 空セル / 列なし = 維持.
    ng_resolved = _parse_ng_staff_cell(
        cells.get("ng_staff_codes"),
        staff_by_code=staff_by_code,
        row_number=row_number,
    )

    if errors:
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient_code,
                operation="error",
                error_message=" / ".join(errors),
            ),
            None,
        )

    # ---- 操作判定 ----
    if existing_patient is None:
        # 新規候補
        # 必須項目チェック.
        missing: list[str] = []
        required_map: dict[str, str] = {
            "patient_code": "patient_code",
            "name": "name",
            "sex": "sex",
            "status": "status",
            "address": "address",
        }
        for spec_key, parsed_key in required_map.items():
            if spec_key == "patient_code":
                if not patient_code:
                    missing.append("patient_code")
                continue
            v = parsed.get(parsed_key)
            if v is None or (isinstance(v, tuple) and v[0] == "CLEAR"):
                missing.append(spec_key)
        if missing:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=patient_code,
                    operation="error",
                    error_message=f"新規作成に必要な項目が不足: {', '.join(missing)}",
                ),
                None,
            )
        # 重複チェック (DB の他レコード + 同ファイル内の他行).
        if patient_code in existing_codes:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=patient_code,
                    operation="error",
                    error_message=f"patient_code が既に存在します: {patient_code!r}",
                ),
                None,
            )
        if patient_code in already_seen_codes:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=patient_code,
                    operation="error",
                    error_message=(
                        f"patient_code が同ファイル内で重複しています: {patient_code!r}"
                    ),
                ),
                None,
            )
        already_seen_codes.add(patient_code)  # type: ignore[arg-type]

        # ---- resurrection 判定 ----
        # DB に同じ code が soft-deleted で残っていれば INSERT すると UNIQUE 制約違反.
        # 新規 INSERT ではなく既存行の deleted_at を NULL に戻して内容を上書きする.
        # patient.id は元の UUID をそのまま使う (再発番しない) ため、外部参照
        # (audit 等) の継続性が保たれる. PFV / shifts は patient soft-delete 時に
        # cascade で物理削除済みなので、resurrection 時点での残骸処理は不要.
        resurrect_target = deleted_patients_by_code.get(patient_code)
        if resurrect_target is not None:
            resurrect_id = resurrect_target.id
            old_deleted_at = resurrect_target.deleted_at
            updates: dict[str, Any] = {}
            changes_for_view: list[PatientExcelChange] = []
            # 復活 = deleted_at を NULL に戻す
            changes_for_view.append(
                PatientExcelChange(
                    field="deleted_at",
                    old_value=_serializable(old_deleted_at),
                    new_value=None,
                )
            )
            # patient_code は必ず一致 (lookup key) なので changes に出さない.
            # 他フィールド: 「空セル = 旧値維持」(update と同じ挙動).
            field_map_resurrect: dict[str, str] = {
                "name": "name",
                "kana": "kana",
                "sex": "sex",
                "status": "status",
                "insurance": "insurance",
                "address": "address",
                "lat": "lat",
                "lng": "lng",
                "primary_office_id": "primary_office_id",
                "sex_restriction": "sex_restriction",
                # Phase E-7 (gap P0-1): NOT NULL bool 列.
                "requires_multiple_staff": "requires_multiple_staff",
                "note": "note",
            }
            for parsed_key, orm_attr in field_map_resurrect.items():
                v = parsed.get(parsed_key)
                if v is None:
                    continue  # 空セル = 旧値維持
                tag, val = v if isinstance(v, tuple) else ("SET", v)
                new_val: Any = None if tag == "CLEAR" else val
                old_val = getattr(resurrect_target, orm_attr, None)
                if old_val is not None and isinstance(new_val, float):
                    try:
                        if float(old_val) == float(new_val):
                            continue
                    except (TypeError, ValueError):
                        pass
                if old_val == new_val:
                    continue
                changes_for_view.append(
                    PatientExcelChange(
                        field=orm_attr,
                        old_value=_serializable(old_val),
                        new_value=_serializable(new_val),
                    )
                )
                updates[orm_attr] = new_val
            # Phase G-49: 拠点コード空欄での復活も住所から自動割当 (新規同様).
            # 明示コードがあれば上の field_map_resurrect で既に反映済.
            if (
                auto_office_id is not None
                and "primary_office_id" not in updates
                and resurrect_target.primary_office_id != auto_office_id
            ):
                changes_for_view.append(
                    PatientExcelChange(
                        field="primary_office_id",
                        old_value=_serializable(resurrect_target.primary_office_id),
                        new_value=_serializable(auto_office_id),
                    )
                )
                updates["primary_office_id"] = auto_office_id
            # NG スタッフ: relationship なので ``_`` 前置きキーで op dict に外出しする.
            ng_op_dict: dict[str, Any] = {}
            ng_change = _build_ng_staff_change(
                ng_resolved,
                ng_staff_by_patient.get(resurrect_id, set()),
                row_number=row_number,
                staff_code_by_id=staff_code_by_id,
            )
            if ng_change is not None:
                change, target_ids = ng_change
                changes_for_view.append(change)
                ng_op_dict["_ng_staff_ids"] = target_ids
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=resurrect_id,
                    patient_code=patient_code,
                    operation="update",
                    changes=changes_for_view,
                ),
                {
                    "_op": "resurrect",
                    "_patient_id": resurrect_id,
                    "_updates": updates,
                    **ng_op_dict,
                },
            )

        # 仮 UUID をここで発番する。PFV シート側で patient_code 経由でこの新規患者を
        # 参照できるよう、apply_changes は同じ UUID で Patient を INSERT する。
        new_patient_id = uuid.uuid4()
        new_dict: dict[str, Any] = {
            "_op": "new",
            "_new_patient_id": new_patient_id,
            "id": new_patient_id,
            "code": patient_code,
            # API レスポンス schema (PatientV2Read.special_week_active: list[dict] = [])
            # は NOT NULL の list を要求する。DB DEFAULT は NULL なので、INSERT 時に
            # 明示的に [] を埋めないと GET /api/v1/patients が Pydantic validation で
            # 500 になる。Excel importer は通常の POST /patients を通らずに直接 ORM
            # INSERT するので、ここで明示的に補完する必要がある。
            "special_week_active": [],
        }
        # 必須以外も含めて値をフラット化.
        for k, v in parsed.items():
            if k in ("code",):
                continue
            if v is None:
                # 空セル: 触らない (新規時は default に任せる).
                continue
            if isinstance(v, tuple):
                tag, val = v
                # CLEAR は新規時は単に「None で INSERT」と等価.
                new_dict[k] = val
        # Phase G-49: 拠点コード空欄の新規患者は住所から自動割当 (UI 挙動と一致).
        # 明示コードがあれば上の loop で既にセット済 (auto では上書きしない).
        if auto_office_id is not None and "primary_office_id" not in new_dict:
            new_dict["primary_office_id"] = auto_office_id
        changes_for_view = [
            PatientExcelChange(field=k, old_value=None, new_value=new_dict[k])
            for k in sorted(new_dict.keys())
            if not k.startswith("_") and k not in ("code", "id")
        ]
        # NG スタッフ: Patient(**data) には渡せないので ``_ng_staff_ids`` で外出し.
        # 新規患者の現状集合は空 → 退職済みスタッフの指定は必ず「新規指定」= skip.
        ng_change = _build_ng_staff_change(
            ng_resolved,
            set(),
            row_number=row_number,
            staff_code_by_id=staff_code_by_id,
        )
        if ng_change is not None:
            change, target_ids = ng_change
            changes_for_view.append(change)
            new_dict["_ng_staff_ids"] = target_ids
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=new_patient_id,
                patient_code=patient_code,
                operation="new",
                changes=changes_for_view,
            ),
            new_dict,
        )

    # 既存患者の update / noop.
    changes: list[PatientExcelChange] = []
    update_dict: dict[str, Any] = {"_patient_id": existing_patient.id, "_op": "update"}

    # patient_code は通常変更しない方針だが、明示的に書いてあって既存値と違えば update.
    if parsed.get("code") is not None:
        _, new_code = parsed["code"]
        if new_code != existing_patient.code:
            # 重複チェック: DB の他レコード + 同ファイル内の他行 (自分以外).
            if new_code in (existing_codes | already_seen_codes) - {existing_patient.code}:
                return (
                    PatientExcelImportRow(
                        row_number=row_number,
                        patient_id=existing_patient.id,
                        patient_code=patient_code,
                        operation="error",
                        error_message=f"patient_code が既に存在します: {new_code!r}",
                    ),
                    None,
                )
            already_seen_codes.add(new_code)
            changes.append(
                PatientExcelChange(
                    field="code",
                    old_value=existing_patient.code,
                    new_value=new_code,
                )
            )
            update_dict["code"] = new_code

    # 他フィールド.
    field_map: dict[str, str] = {
        "name": "name",
        "kana": "kana",
        "sex": "sex",
        "status": "status",
        "insurance": "insurance",
        "address": "address",
        "lat": "lat",
        "lng": "lng",
        "primary_office_id": "primary_office_id",
        "sex_restriction": "sex_restriction",
        # Phase E-7 (gap P0-1): NOT NULL bool 列.
        "requires_multiple_staff": "requires_multiple_staff",
        "note": "note",
    }
    for parsed_key, orm_attr in field_map.items():
        v = parsed.get(parsed_key)
        if v is None:
            continue  # 空セル = 触らない
        tag, val = v if isinstance(v, tuple) else ("SET", v)
        if tag == "CLEAR":
            new_val: Any = None
        else:
            new_val = val
        old_val = getattr(existing_patient, orm_attr, None)
        # Numeric (Decimal) と float の比較対策
        if old_val is not None and isinstance(new_val, float):
            try:
                if float(old_val) == float(new_val):
                    continue
            except (TypeError, ValueError):
                pass
        if old_val == new_val:
            continue
        changes.append(
            PatientExcelChange(
                field=orm_attr,
                old_value=_serializable(old_val),
                new_value=_serializable(new_val),
            )
        )
        update_dict[orm_attr] = new_val

    # NG スタッフ (relationship): ``_ng_staff_ids`` で op dict に外出し.
    ng_change = _build_ng_staff_change(
        ng_resolved,
        ng_staff_by_patient.get(existing_patient.id, set()),
        row_number=row_number,
        staff_code_by_id=staff_code_by_id,
    )
    if ng_change is not None:
        change, target_ids = ng_change
        changes.append(change)
        update_dict["_ng_staff_ids"] = target_ids

    if not changes:
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=existing_patient.id,
                patient_code=existing_patient.code,
                operation="noop",
            ),
            None,
        )

    return (
        PatientExcelImportRow(
            row_number=row_number,
            patient_id=existing_patient.id,
            patient_code=existing_patient.code,
            operation="update",
            changes=changes,
        ),
        update_dict,
    )


def _serializable(value: Any) -> Any:
    """Pydantic Response 用に安全な型へ変換 (UUID/Decimal/time など)."""
    if value is None:
        return None
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    # Decimal は float 化, time/datetime は ISO 文字列.
    try:
        # Decimal
        return float(value)
    except (TypeError, ValueError):
        pass
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


# ---------------------------------------------------------------------------
# PFV sheet parser
# ---------------------------------------------------------------------------


class _PendingNewPatient:
    """同 import 内で新規追加される患者の最小限の代理オブジェクト。

    PFV 行のパーサーが ``patient.code`` / ``patient.primary_office_id`` を参照する
    ため、Patient ORM と同じ属性面を持つ軽量オブジェクトを差し込む。
    """

    __slots__ = ("id", "code", "primary_office_id")

    def __init__(
        self,
        *,
        patient_id: UUID,
        code: str,
        primary_office_id: UUID | None,
    ) -> None:
        self.id = patient_id
        self.code = code
        self.primary_office_id = primary_office_id


def _parse_pfv_row(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    pending_new_patients: dict[UUID, _PendingNewPatient],
    patient_code_to_id: dict[str, UUID],
    course_templates: dict[tuple[UUID, str], CourseTemplate],
    existing_pfvs: dict[tuple[UUID, str, int, int], PatientFixedVisit],
    pending_new_keys: set[tuple[UUID, str, int, int]],
    offices_by_code: dict[str, Office] | None = None,
) -> tuple[PfvExcelImportRow, dict[str, Any] | None]:
    cells: dict[str, Any] = {}
    for col_key, idx in PFV_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["patient_id"]
    raw_code = cells["patient_code"]
    raw_delete = cells["delete_flag"]

    # patient_id (任意; 空なら patient_code で解決を試みる)
    try:
        patient_id = _read_uuid(raw_id)
    except ValueError:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=None,
                operation="error",
                error_message=f"patient_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    patient_code = _read_str(raw_code)

    # patient_id 空 → patient_code で解決
    if patient_id is None:
        if patient_code is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=None,
                    operation="error",
                    error_message="patient_id または patient_code が必須です",
                ),
                None,
            )
        resolved_id = patient_code_to_id.get(patient_code)
        if resolved_id is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=patient_code,
                    operation="error",
                    error_message=(
                        f"patient_code が DB にも import 内にも存在しません: {patient_code!r}"
                    ),
                ),
                None,
            )
        patient_id = resolved_id
    elif patient_code is not None:
        # 両方記入されている → 整合性チェック。患者を取り違えてデータ破壊する事故を防ぐ.
        expected_id_from_code = patient_code_to_id.get(patient_code)
        if expected_id_from_code is not None and expected_id_from_code != patient_id:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=(
                        f"patient_id と patient_code が異なる患者を指しています "
                        f"(id={patient_id}, code={patient_code} → 期待 id={expected_id_from_code})"
                    ),
                ),
                None,
            )

    # 既存患者 → ORM patient を取得 / 新規候補 → pseudo patient を組み立て
    patient = existing_patients.get(patient_id)
    if patient is None:
        pending = pending_new_patients.get(patient_id)
        if pending is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=f"patient_id が DB に存在しません: {patient_id}",
                ),
                None,
            )
        # 同 import 内で新規追加される患者。Patient ORM はまだ無いので
        # _parse_pfv_row 内で使う最小限の属性を pseudo オブジェクトで賄う。
        patient = pending  # type: ignore[assignment]

    # weekday / slot_index / mode
    raw_weekday = cells["weekday"]
    if _is_blank(raw_weekday):
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                operation="error",
                error_message="weekday が空です",
            ),
            None,
        )
    wd_str = _read_str(raw_weekday)
    weekday = WEEKDAY_LABEL_TO_INT.get(wd_str or "")
    if weekday is None:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                operation="error",
                error_message=f"weekday の値が候補外: {wd_str!r} (許容: 月火水木金土日)",
            ),
            None,
        )

    raw_slot = cells["slot_index"]
    try:
        slot_index = _read_int(raw_slot)
    except (ValueError, TypeError):
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                operation="error",
                error_message=f"slot_index が整数ではありません: {raw_slot!r}",
            ),
            None,
        )
    if slot_index is None:
        slot_index = 0
    if slot_index < 0 or slot_index > 1:
        # NOTE: 仕様書には 0-9 と書いてあるが DB CHECK 制約は 0-1. DB に合わせる.
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                operation="error",
                error_message=f"slot_index は 0 または 1 のみ: {slot_index}",
            ),
            None,
        )

    raw_mode = cells["mode"]
    if _is_blank(raw_mode):
        mode = "normal"
    else:
        # Phase G-48: モードは日本語 (通常/特別) / 英語 (normal/special) 双方受理.
        mode_str = _read_str(raw_mode)
        mode_en = PFV_MODE_ANY_TO_EN.get(mode_str or "")
        if mode_en is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"mode の値が候補外: {mode_str!r}",
                ),
                None,
            )
        mode = mode_en

    key = (patient_id, mode, weekday, slot_index)
    existing_pfv = existing_pfvs.get(key)

    # 削除フラグ (PFV は物理削除)
    if is_magic_delete(raw_delete):
        if existing_pfv is None:
            # 存在しないものを delete しようとした → noop 扱い (warn でなく許容)
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="noop",
                ),
                None,
            )
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="delete",
            ),
            {"_pfv_id": existing_pfv.id, "_op": "delete"},
        )

    # time_type
    # 空セルは default ("時間帯") で吸収する (E-4: round-trip 運用で 0 エラーを担保).
    # patient.weekly_pattern にエントリが無いと export 時に空で出てくるため、
    # 単に default で埋めて success 扱いする。time_type は PFV テーブルに保存先が
    # 無いため (importer.py 1178-1181 の note 参照) UI 用の参考情報のみ。
    raw_tt = cells["time_type"]
    if _is_blank(raw_tt):
        time_type: str | None = DEFAULT_TIME_TYPE
    else:
        time_type = _read_str(raw_tt)
        if time_type not in TIME_TYPE_VALUES:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"time_type の値が候補外: {time_type!r}",
                ),
                None,
            )

    # start_time
    try:
        start_str = _read_hhmm(cells["start_time"])
    except (ValueError, TypeError) as exc:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message=f"start_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )
    if start_str is None:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message="start_time が空です",
            ),
            None,
        )

    # end_time / duration_min: どちらかから duration_min を導出.
    raw_end = cells["end_time"]
    raw_duration = cells["duration_min"]
    duration_min: int | None = None
    if not _is_blank(raw_duration):
        try:
            duration_min = _read_int(raw_duration)
        except (ValueError, TypeError) as exc:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"duration_min が整数ではありません: {exc}",
                ),
                None,
            )
    if duration_min is None and not _is_blank(raw_end):
        try:
            end_str = _read_hhmm(raw_end)
        except (ValueError, TypeError) as exc:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"end_time が HH:MM 形式ではありません: {exc}",
                ),
                None,
            )
        if end_str is not None:
            sh, sm = (int(x) for x in start_str.split(":"))
            eh, em = (int(x) for x in end_str.split(":"))
            duration_min = (eh * 60 + em) - (sh * 60 + sm)
    if duration_min is None:
        duration_min = 30  # default
    if duration_min < 1 or duration_min > 480:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message=f"duration_min が範囲外 [1, 480]: {duration_min}",
            ),
            None,
        )

    # Phase E-5 (項目 ⑥B): sub_office_code → sub_office_id を先に解決.
    # 解決できない code (拠点に存在しない) は course_template と同じく
    # 「sub_office_id=None として PFV 自体は保持」のベストエフォート方針.
    raw_sub_office = cells.get("sub_office_code")
    sub_office_id: UUID | None = None
    if not _is_blank(raw_sub_office) and offices_by_code is not None:
        oc = _read_str(raw_sub_office)
        if oc:
            office = offices_by_code.get(oc)
            if office is not None:
                sub_office_id = office.id

    # course_template_code → course_template_id
    # E-4: 「拠点に存在しない code」は error にせず course_template_id=None として
    # 取り込み、PFV 自体は保持する (round-trip / バックアップ運用優先).
    # patient.primary_office 未設定 or template が当該拠点に居ない場合のいずれも
    # 「ベストエフォートで PFV を保存し、course_template は剥がす」挙動.
    # Phase E-5: sub_office_id が指定された場合は course_template の解決先 office も
    # sub_office を優先する (= サブ拠点の course を sub_office_id 経由で指せる).
    raw_ct = cells["course_template_code"]
    course_template_id: UUID | None = None
    if not _is_blank(raw_ct):
        ct_label = _read_str(raw_ct)
        # 優先順位: sub_office (Phase E-5) → primary_office (既存).
        # sub_office で見つからない場合は primary_office にフォールバック.
        resolved = False
        if sub_office_id is not None:
            ct = course_templates.get((sub_office_id, ct_label or ""))
            if ct is not None:
                course_template_id = ct.id
                resolved = True
        if not resolved and patient.primary_office_id is not None:
            ct = course_templates.get((patient.primary_office_id, ct_label or ""))
            if ct is not None:
                course_template_id = ct.id
        # 解決できなかった場合は course_template_id を None のまま継続 (info 扱い).
        # round-trip 運用で多発するため、この経路では warning ログを出さない.

    # ---- ファイル内 (patient_id, mode, weekday, slot_index) 重複検査 ----
    if key in pending_new_keys:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message=(
                    "同ファイル内で (patient_id, mode, weekday, slot_index) が重複しています"
                ),
            ),
            None,
        )

    new_start = _hhmm_to_time(start_str)
    if existing_pfv is None:
        pending_new_keys.add(key)
        new_dict: dict[str, Any] = {
            "_op": "new",
            "patient_id": patient_id,
            "mode": mode,
            "weekday": weekday,
            "slot_index": slot_index,
            "start_time": new_start,
            "duration_min": duration_min,
            "course_template_id": course_template_id,
            # Phase E-5 (項目 ⑥B): サブ拠点 ID. 解決できなかった場合は None.
            "sub_office_id": sub_office_id,
        }
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="new",
                changes=[
                    PatientExcelChange(field="start_time", old_value=None, new_value=start_str),
                    PatientExcelChange(
                        field="duration_min", old_value=None, new_value=duration_min
                    ),
                ],
            ),
            new_dict,
        )

    # update / noop
    changes: list[PatientExcelChange] = []
    update_dict: dict[str, Any] = {"_pfv_id": existing_pfv.id, "_op": "update"}
    if existing_pfv.start_time != new_start:
        changes.append(
            PatientExcelChange(
                field="start_time",
                old_value=_serializable(existing_pfv.start_time),
                new_value=start_str,
            )
        )
        update_dict["start_time"] = new_start
    if existing_pfv.duration_min != duration_min:
        changes.append(
            PatientExcelChange(
                field="duration_min",
                old_value=existing_pfv.duration_min,
                new_value=duration_min,
            )
        )
        update_dict["duration_min"] = duration_min
    if existing_pfv.course_template_id != course_template_id:
        changes.append(
            PatientExcelChange(
                field="course_template_id",
                old_value=_serializable(existing_pfv.course_template_id),
                new_value=_serializable(course_template_id),
            )
        )
        update_dict["course_template_id"] = course_template_id
    # Phase E-5 (項目 ⑥B): sub_office_id の差分.
    if existing_pfv.sub_office_id != sub_office_id:
        changes.append(
            PatientExcelChange(
                field="sub_office_id",
                old_value=_serializable(existing_pfv.sub_office_id),
                new_value=_serializable(sub_office_id),
            )
        )
        update_dict["sub_office_id"] = sub_office_id
    # time_type は patients.weekly_pattern 側の情報. PFV テーブル自体には保存先が無い.
    # 仕様書「time_type → patients.weekly_pattern にも反映」は WeeklyPattern の構造
    # 上、PFV 行 1 件から単純に書き換えると他曜日の設定を壊しうるため、本実装では
    # 表示用のみとし、import 時には patients.weekly_pattern を変更しない (TODO).
    if not changes:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="noop",
            ),
            None,
        )
    return (
        PfvExcelImportRow(
            row_number=row_number,
            patient_id=patient_id,
            patient_code=patient.code,
            weekday=weekday,
            slot_index=slot_index,
            operation="update",
            changes=changes,
        ),
        update_dict,
    )


# ---------------------------------------------------------------------------
# Phase G-51: 「固定訪問パターン（編集用）」シート (患者 1 行) パーサ
# ---------------------------------------------------------------------------


def _resolve_pfv_course(
    token: str | None,
    *,
    patient: Patient | _PendingNewPatient,
    offices_by_code: dict[str, Office],
    course_templates: dict[tuple[UUID, str], CourseTemplate],
) -> UUID | None:
    """拠点付きコーストークン (例 "稲A" / クロス拠点 "津A") → course_template_id.

    トークンを (office_code, label) に分解し、その拠点の course_template を解決する.
    トークンに拠点が含まれるためクロス拠点コースもそのまま course_template_id で表現できる
    (sub_office_id は本シートでは扱わない: 既存 PFV の sub_office_id を保持し round-trip 安定).
    解決できない場合は None (= best-effort: コース未解決でも PFV は保持).
    """
    if token is None:
        return None
    # 0059: 拠点マスタ (offices.short_label) 駆動で短縮名を解決する.
    _, short_to_code = build_office_code_short_maps(
        (o.code, o.short_label) for o in offices_by_code.values()
    )
    parsed = parse_course_token(token, short_to_code)
    if parsed is None:
        return None
    office_code, label = parsed
    office = offices_by_code.get(office_code)
    if office is None:
        return None
    ct = course_templates.get((office.id, label))
    if ct is None:
        return None
    return ct.id


def _parse_pfv_edit_patient(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    pending_new_patients: dict[UUID, _PendingNewPatient],
    patient_code_to_id: dict[str, UUID],
    course_templates: dict[tuple[UUID, str], CourseTemplate],
    existing_pfvs: dict[tuple[UUID, str, int, int], PatientFixedVisit],
    offices_by_code: dict[str, Office],
) -> tuple[list[PfvExcelImportRow], list[dict[str, Any]]]:
    """「固定訪問パターン（編集用）」1 行 (患者 1 行) → PFV 差分行 + apply op の list.

    per-patient replace セマンティクス (normal mode のみ対象):
      - 各曜日 (月..土) の ``{曜日}_時刻`` が入っていれば PFV(weekday, slot_index=0,
        mode=normal) を upsert. ``{曜日}_コース`` を (office, label) → course_template に解決.
      - 行が示す曜日集合に既存 normal PFV を **一致させる**: 行に無い曜日の既存 normal
        PFV は削除、ある曜日は upsert (空欄曜日 = 訪問なし).
      - special mode / slot_index!=0 の既存 PFV は触らない (本シート対象外で保持).
      - 行に出てこない患者の PFV は触らない (incremental; 呼び出し側でループ).
    """
    cells: dict[str, Any] = {
        col_key: (row[idx] if idx < len(row) else None)
        for col_key, idx in PFV_EDIT_COL_INDEX.items()
    }

    raw_id = cells["patient_id"]
    raw_code = cells["patient_code"]

    try:
        patient_id = _read_uuid(raw_id)
    except ValueError:
        return (
            [
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=_read_str(raw_code),
                    operation="error",
                    error_message=f"patient_id が UUID 形式ではありません: {raw_id!r}",
                )
            ],
            [],
        )

    patient_code = _read_str(raw_code)

    # patient 解決 (patient_id 空 → patient_code).
    if patient_id is None:
        if patient_code is None:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=None,
                        patient_code=None,
                        operation="error",
                        error_message="patient_id または patient_code が必須です",
                    )
                ],
                [],
            )
        resolved_id = patient_code_to_id.get(patient_code)
        if resolved_id is None:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=None,
                        patient_code=patient_code,
                        operation="error",
                        error_message=(
                            f"patient_code が DB にも import 内にも存在しません: {patient_code!r}"
                        ),
                    )
                ],
                [],
            )
        patient_id = resolved_id
    elif patient_code is not None:
        expected = patient_code_to_id.get(patient_code)
        if expected is not None and expected != patient_id:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=patient_id,
                        patient_code=patient_code,
                        operation="error",
                        error_message=(
                            f"patient_id と patient_code が異なる患者を指しています "
                            f"(id={patient_id}, code={patient_code} → 期待 id={expected})"
                        ),
                    )
                ],
                [],
            )

    patient: Patient | _PendingNewPatient | None = existing_patients.get(patient_id)
    if patient is None:
        patient = pending_new_patients.get(patient_id)
        if patient is None:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=patient_id,
                        patient_code=patient_code,
                        operation="error",
                        error_message=f"patient_id が DB に存在しません: {patient_id}",
                    )
                ],
                [],
            )

    # サービス時間 (患者共通 duration) / 時間タイプ.
    raw_sm = cells["service_minutes"]
    try:
        service_minutes = _read_int(raw_sm)
    except (ValueError, TypeError):
        service_minutes = None
    if service_minutes is not None and (service_minutes < 1 or service_minutes > 480):
        service_minutes = None  # 範囲外は default に倒す (壊さない方針)

    # 各曜日のセル → 「行が示す曜日 → (start_time, course_template_id, sub_office_id, duration)」.
    row_pfvs: dict[int, dict[str, Any]] = {}
    errors: list[str] = []
    for wd_key in PFV_EDIT_WEEKDAY_KEYS:
        wd_int = PFV_EDIT_WEEKDAY_KEY_TO_INT[wd_key]
        raw_time = cells.get(f"{wd_key}_time")
        if _is_blank(raw_time):
            continue  # この曜日は訪問なし.
        try:
            start_str = _read_hhmm(raw_time)
        except (ValueError, TypeError) as exc:
            errors.append(f"{wd_key} の時刻が HH:MM 形式ではありません: {exc}")
            continue
        if start_str is None:
            continue
        ct_id = _resolve_pfv_course(
            _read_str(cells.get(f"{wd_key}_course")),
            patient=patient,
            offices_by_code=offices_by_code,
            course_templates=course_templates,
        )
        duration = service_minutes if service_minutes is not None else 30
        row_pfvs[wd_int] = {
            "start_time": _hhmm_to_time(start_str),
            "start_str": start_str,
            "course_template_id": ct_id,
            "duration_min": duration,
        }

    if errors:
        return (
            [
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    operation="error",
                    error_message=" / ".join(errors),
                )
            ],
            [],
        )

    # 既存 normal PFV (slot_index=0) を曜日でひく.
    existing_by_weekday: dict[int, PatientFixedVisit] = {}
    for (pid, mode, wd, slot), pfv in existing_pfvs.items():
        if pid == patient_id and mode == "normal" and slot == 0:
            existing_by_weekday[wd] = pfv

    diff_rows: list[PfvExcelImportRow] = []
    ops: list[dict[str, Any]] = []

    # 行が示す曜日: upsert. 既存にあれば update/noop, 無ければ new.
    for wd_int in sorted(row_pfvs):
        spec = row_pfvs[wd_int]
        existing_pfv = existing_by_weekday.get(wd_int)
        if existing_pfv is None:
            ops.append(
                {
                    "_op": "new",
                    "patient_id": patient_id,
                    "mode": "normal",
                    "weekday": wd_int,
                    "slot_index": 0,
                    "start_time": spec["start_time"],
                    "duration_min": spec["duration_min"],
                    "course_template_id": spec["course_template_id"],
                    # sub_office_id は本シート対象外 (model default None). 既存 PFV の
                    # sub_office_id は update でも触らない (round-trip 安定).
                }
            )
            diff_rows.append(
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=wd_int,
                    slot_index=0,
                    operation="new",
                    changes=[
                        PatientExcelChange(
                            field="start_time", old_value=None, new_value=spec["start_str"]
                        ),
                        PatientExcelChange(
                            field="duration_min",
                            old_value=None,
                            new_value=spec["duration_min"],
                        ),
                    ],
                )
            )
            continue
        # update / noop.
        changes: list[PatientExcelChange] = []
        update_dict: dict[str, Any] = {"_pfv_id": existing_pfv.id, "_op": "update"}
        if existing_pfv.start_time != spec["start_time"]:
            changes.append(
                PatientExcelChange(
                    field="start_time",
                    old_value=_serializable(existing_pfv.start_time),
                    new_value=spec["start_str"],
                )
            )
            update_dict["start_time"] = spec["start_time"]
        if existing_pfv.duration_min != spec["duration_min"]:
            changes.append(
                PatientExcelChange(
                    field="duration_min",
                    old_value=existing_pfv.duration_min,
                    new_value=spec["duration_min"],
                )
            )
            update_dict["duration_min"] = spec["duration_min"]
        if existing_pfv.course_template_id != spec["course_template_id"]:
            changes.append(
                PatientExcelChange(
                    field="course_template_id",
                    old_value=_serializable(existing_pfv.course_template_id),
                    new_value=_serializable(spec["course_template_id"]),
                )
            )
            update_dict["course_template_id"] = spec["course_template_id"]
        # sub_office_id は本シートでは触らない (既存値を保持; round-trip 安定 + 誤削除防止).
        if changes:
            ops.append(update_dict)
            diff_rows.append(
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=wd_int,
                    slot_index=0,
                    operation="update",
                    changes=changes,
                )
            )
        else:
            diff_rows.append(
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=wd_int,
                    slot_index=0,
                    operation="noop",
                )
            )

    # 行に無い曜日の既存 normal PFV は削除 (空欄曜日 = 訪問なし).
    # ただし削除スコープは「このシートが実際に管理する曜日」(PFV_EDIT_WEEKDAY_KEYS =
    # 月..土) に限定する. 本シート (カルテ / 固定訪問パターン（編集用）) は月..土しか
    # 入出力しないため、管理対象外の曜日 (日曜=6) の既存 normal slot0 PFV は
    # 空欄=削除と誤解せず必ず温存する (CRITICAL: 無編集 round-trip で日曜 PFV が
    # 黙殺削除されるのを防ぐ). special / slot_index!=0 / manual は existing_by_weekday
    # の構築段階で既に除外済み.
    managed_weekdays = set(PFV_EDIT_WEEKDAY_KEY_TO_INT.values())
    for wd_int, existing_pfv in existing_by_weekday.items():
        if wd_int in row_pfvs:
            continue
        if wd_int not in managed_weekdays:
            continue  # 管理対象外曜日 (日曜) は削除しない (温存).
        ops.append({"_pfv_id": existing_pfv.id, "_op": "delete"})
        diff_rows.append(
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=wd_int,
                slot_index=0,
                operation="delete",
            )
        )

    return diff_rows, ops


# ---------------------------------------------------------------------------
# public entrypoints
# ---------------------------------------------------------------------------


def _build_weekly_pattern(
    *,
    frequency_per_week: Any,
    visit_frequency: Any,
    visit_weeks: Any,
    preferred_weekdays_en: list[str],
    service_minutes: Any,
    time_type: Any,
    preferred_start: Any,
    preferred_end: Any,
) -> dict[str, Any]:
    """各 weekly フィールド (生セル値) → weekly_pattern dict (WeeklyPatternV2 準拠).

    空 / 候補外の値は dict に積まない (= round-trip 安定性のため「維持」セマンティクス).
    visit_frequency は日本語/英語いずれも受理し DB 英語正準値 (every 等) へ正規化.
    preferred_weekdays_en は呼び出し側で英 list 化済みのものを受け取る.
    """
    wp: dict[str, Any] = {}

    freq = _read_int(frequency_per_week)
    if freq is not None and 1 <= freq <= 7:
        wp["frequency_per_week"] = freq

    vf = _read_str(visit_frequency)
    if vf:
        vf_en = VISIT_FREQUENCY_ANY_TO_EN.get(vf)
        if vf_en is not None:
            wp["visit_frequency"] = vf_en

    vw = _read_str(visit_weeks)
    if vw:
        wp["visit_weeks"] = vw

    if preferred_weekdays_en:
        wp["preferred_weekdays"] = preferred_weekdays_en

    sm = _read_int(service_minutes)
    if sm is not None and 0 <= sm <= 300:
        wp["service_minutes"] = sm

    tt = _read_str(time_type)
    if tt and tt in TIME_TYPE_VALUES:
        wp["time_type"] = tt

    ps = _read_hhmm(preferred_start)
    if ps:
        wp["preferred_start"] = ps
    pe = _read_hhmm(preferred_end)
    if pe:
        wp["preferred_end"] = pe

    return wp


# Phase G-48 hotfix: 統合「患者マスタ」シートが扱う weekly_pattern の管理キー (8 個).
# exporter/_build_weekly_pattern が読み書きするのはこの 8 キーのみ. ``entries`` /
# ``staff_count`` 等の細粒度キー (WeeklyPatternV2 は extra="allow") は本パイプラインの
# 管理対象外で、merge 時に必ず保持しなければならない (消すとスケジューラの訪問時刻 /
# 2 名体制が静かに壊れる).
MANAGED_WEEKLY_KEYS: tuple[str, ...] = (
    "frequency_per_week",
    "visit_frequency",
    "visit_weeks",
    "preferred_weekdays",
    "service_minutes",
    "time_type",
    "preferred_start",
    "preferred_end",
)


def _merge_weekly_pattern(
    existing_wp: dict[str, Any] | None,
    wp: dict[str, Any] | None,
) -> dict[str, Any] | None:
    """既存 weekly_pattern のコピーに、Excel 由来の管理 8 キーのみ上書きしたものを返す.

    Phase G-48 hotfix (CRITICAL データ損失修正):
    - ``existing_wp`` のコピーから開始 → ``entries`` / ``staff_count`` 等の **管理外キーを保持**.
    - ``wp`` (= ``_build_weekly_pattern`` の出力、非空セルの管理キーのみ) を上書き.
    - ``wp`` に無い管理キーは既存値のまま (= 空セル → 「維持」semantics).

    これにより blank=keep を保ったまま、丸ごと置換による管理外キーの消失を防ぐ.

    Phase G-48 hotfix-2 (HIGH クラッシュ修正):
    - ``wp`` が None / 空 dict のとき weekly_pattern を **None にクリア** (返り値 None).
      旧 3 シート構成の「希望訪問パターン」シートで ``<DELETE>`` 行を import すると
      呼び出し側が ``wp=None`` を渡すため、明示クリアとして扱う (旧挙動の維持).
      統合シート経路では全空行は merge を呼ばず blank=keep なので、この guard が
      維持 semantics を壊すことはない (merge は非空 wp でのみ到達).
    """
    if not wp:  # None or {} => 旧シート <DELETE> のクリア意味論
        return None
    base: dict[str, Any] = dict(existing_wp) if isinstance(existing_wp, dict) else {}
    for key in MANAGED_WEEKLY_KEYS:
        if key in wp:
            base[key] = wp[key]
    return base


def _parse_weekly_from_patient_cells(cells: dict[str, Any]) -> dict[str, Any] | None:
    """Phase G-48: 統合「患者マスタ」シート 1 行の cells → weekly_pattern dict.

    cells は PATIENT_COL_INDEX のキーで引ける dict (``_parse_patient_row`` 内で構築済).
    Phase G-49: 希望曜日は 7 列 (pref_wd_mon..pref_wd_sun) の はい/いいえ. 全列いいえ/
    空なら preferred_weekdays は wp に積まない (= 既存維持 / blank=keep).

    後方互換: 旧 1 セル形式 ("月,水,金") のキー ``preferred_weekdays`` が cells に
    残っている場合 (旧 export を新スキーマ index で読んだ場合は通常存在しないが、
    呼び出し側が補完した場合に備える) はそちらも併せて解釈する.
    全 weekly フィールドが空なら None (= 既存維持). <DELETE> は呼び出し側
    (delete_flag) で処理するためここでは扱わない.
    """
    # Phase G-49: 7 列 はい/いいえ → 英 list.
    weekdays_en = weekday_yesno_cells_to_en(cells)
    # 後方互換: 旧 1 セル形式が残っていれば併合 (どちらかに値があれば拾う).
    legacy_cell = _read_str(cells.get("preferred_weekdays"))
    if legacy_cell:
        legacy_en = weekdays_cell_to_en(legacy_cell)
        merged = set(weekdays_en) | set(legacy_en)
        weekdays_en = [en for en in WEEKDAY_EN_LIST if en in merged]
    wp = _build_weekly_pattern(
        frequency_per_week=cells.get("frequency_per_week"),
        visit_frequency=cells.get("visit_frequency"),
        visit_weeks=cells.get("visit_weeks"),
        preferred_weekdays_en=weekdays_en,
        service_minutes=cells.get("service_minutes"),
        time_type=cells.get("weekly_time_type"),
        preferred_start=cells.get("preferred_start"),
        preferred_end=cells.get("preferred_end"),
    )
    return wp or None


def _parse_weekly_row(
    r_idx: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    existing_patients_by_code: dict[str, Patient],
    patient_code_to_id: dict[str, UUID],
    pending_new_patients: dict[UUID, _PendingNewPatient],
) -> tuple[UUID, dict[str, Any]] | None:
    """【後方互換専用】旧「希望訪問パターン」独立シート 1 行 → (patient_id, weekly dict).

    Phase E-8 で導入した 3 シート構成 (曜日 7 列 TRUE/FALSE) の旧 export ファイルを
    import するためのパーサ. Phase G-48 で weekly_pattern は患者マスタへ統合された
    ため新規 export はこのシートを出力しないが、旧ファイル受理のため温存する.

    return None: 行が無効 (patient 解決失敗、または全セル空).
    """

    def _g(key: str) -> Any:
        idx = WEEKLY_COL_INDEX.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    # patient 解決: patient_id → existing → patient_code → 既存 or pending_new
    pid_raw = _g("patient_id")
    code_raw = _g("patient_code")
    target_pid: UUID | None = None

    pid_uuid = _read_uuid(pid_raw)
    if pid_uuid is not None:
        if pid_uuid in existing_patients or pid_uuid in pending_new_patients:
            target_pid = pid_uuid
    if target_pid is None:
        code_str = _read_str(code_raw)
        if code_str:
            target_pid = patient_code_to_id.get(code_str)

    if target_pid is None:
        return None  # patient 解決不可 = skip (warning は呼び出し側で必要なら出す)

    # 希望曜日 (旧 7 列、TRUE → list に追加)
    weekday_keys = [
        ("wd_mon", "Mon"),
        ("wd_tue", "Tue"),
        ("wd_wed", "Wed"),
        ("wd_thu", "Thu"),
        ("wd_fri", "Fri"),
        ("wd_sat", "Sat"),
        ("wd_sun", "Sun"),
    ]
    preferred_weekdays: list[str] = []
    for col_key, wd_en in weekday_keys:
        val = _read_bool(_g(col_key))
        if val:  # TRUE のみ拾う (FALSE/空 = この曜日希望なし)
            preferred_weekdays.append(wd_en)

    wp = _build_weekly_pattern(
        frequency_per_week=_g("frequency_per_week"),
        visit_frequency=_g("visit_frequency"),
        visit_weeks=_g("visit_weeks"),
        preferred_weekdays_en=preferred_weekdays,
        service_minutes=_g("service_minutes"),
        time_type=_g("time_type"),
        preferred_start=_g("preferred_start"),
        preferred_end=_g("preferred_end"),
    )

    # 削除フラグ: <DELETE> なら weekly_pattern = None (= clear)
    if is_magic_delete(_g("delete_flag")):
        return target_pid, {}  # 呼び出し側で「空 dict = clear」と解釈

    # 全セル空なら skip (= 「維持」セマンティクス)
    if not wp:
        return None

    return target_pid, wp


async def parse_and_diff(
    db: AsyncSession,
    *,
    file_bytes: bytes,
) -> tuple[
    list[PatientExcelImportRow],
    list[PfvExcelImportRow],
    PatientExcelImportSummary,
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    """Excel をパースして差分行 + summary + apply 用 op list を返す.

    戻り値:
      - patient_rows: API レスポンス用の per-row 結果
      - pfv_rows: API レスポンス用の per-row 結果
      - summary: 集計
      - patient_ops: apply 用の DB op list (内部利用)
      - pfv_ops: apply 用の DB op list (内部利用)
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    if SHEET_PATIENTS not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_PATIENTS}」が見つかりません")
    # Phase G-51: PFV シートの形式を判定する.
    #   * 新「固定訪問パターン（編集用）」(患者 1 行) があれば最優先で使う.
    #   * 無ければ旧 per-visit「固定訪問パターン」/「固定訪問スケジュール」(旧名) に
    #     フォールバック (過去 export 済 xlsx も import 可能にするため).
    #   * 静的「グリッド集計（静的・閲覧専用）」シートは一切読まない (シート名で識別して無視).
    use_pfv_edit_sheet = SHEET_PFV_EDIT in wb.sheetnames
    pfv_sheet_name = (
        SHEET_PFV
        if SHEET_PFV in wb.sheetnames
        else LEGACY_SHEET_PFV
        if LEGACY_SHEET_PFV in wb.sheetnames
        else None
    )
    if not use_pfv_edit_sheet and pfv_sheet_name is None:
        raise ValueError(f"シート「{SHEET_PFV_EDIT}」が見つかりません")

    existing_patients = await _load_patients_by_id(db)
    existing_codes = await _load_patient_codes(db)
    offices_by_code = await _load_offices_by_code(db)
    course_templates = await _load_course_templates_by_label(db)
    existing_pfvs = await _load_pfvs_by_key(db, set(existing_patients.keys()))
    # 削除パスで patient_code → Patient を解決するため (patient_id 列が空でも
    # patient_code リンクで削除できるようにする).
    existing_patients_by_code: dict[str, Patient] = {
        p.code: p for p in existing_patients.values() if p.code
    }
    # resurrection 用: soft-deleted な patient_code → Patient.
    deleted_patients_by_code = await _load_deleted_patients_by_code(db)
    # Phase G-49: 拠点コード空欄の住所→拠点 自動割当用. cities/office_cities を
    # 一度だけロードし、行ごとは同期照合 (N+1 回避).
    office_index = await load_office_cities_index(db)
    # NG スタッフ列用: staff_code → Staff (退職者含む) と 既存 NG 集合を bulk load.
    staff_by_code = await _load_staff_by_code(db)
    staff_code_by_id: dict[UUID, str] = {s.id: str(s.code) for s in staff_by_code.values()}
    ng_staff_by_patient = await _load_ng_staff_by_patient(db)

    # ---- 患者シート ----
    ws_p = wb[SHEET_PATIENTS]
    patient_rows: list[PatientExcelImportRow] = []
    patient_ops: list[dict[str, Any]] = []
    already_seen_codes: set[str] = set()
    # Phase G-48: 統合シートから読み取った weekly_pattern. patient_id →
    # weekly_pattern dict (空 dict = clear). 後段で patient_ops にマージする.
    weekly_from_patient_sheet: dict[UUID, dict[str, Any] | None] = {}
    # row_number は 1-indexed; ヘッダーが 1 行目, データは 2 行目から.
    for r_idx, row in enumerate(ws_p.iter_rows(min_row=2, values_only=True), start=2):
        if _row_is_empty(row):
            continue
        diff_row, op = _parse_patient_row(
            r_idx,
            row,
            existing_patients=existing_patients,
            existing_patients_by_code=existing_patients_by_code,
            deleted_patients_by_code=deleted_patients_by_code,
            offices_by_code=offices_by_code,
            existing_codes=existing_codes,
            already_seen_codes=already_seen_codes,
            office_index=office_index,
            staff_by_code=staff_by_code,
            ng_staff_by_patient=ng_staff_by_patient,
            staff_code_by_id=staff_code_by_id,
        )
        patient_rows.append(diff_row)
        if op is not None:
            patient_ops.append(op)
        # Phase G-48: 統合シートの weekly_pattern を抽出.
        # error 行 / patient 未解決行 (patient_id None) は対象外.
        # delete 行は後段の merge で delete op に統合しないので除外.
        if diff_row.operation in ("new", "update", "noop") and diff_row.patient_id is not None:
            row_cells: dict[str, Any] = {
                col_key: (row[idx] if idx < len(row) else None)
                for col_key, idx in PATIENT_COL_INDEX.items()
            }
            weekly = _parse_weekly_from_patient_cells(row_cells)
            if weekly is not None:
                weekly_from_patient_sheet[diff_row.patient_id] = weekly

    # PFV シートが patient_code 経由で新規患者をリンクできるよう、
    # patient_code → patient_id の lookup を構築する。新規患者は
    # _parse_patient_row が ``_new_patient_id`` (uuid.uuid4()) を発番済み。
    pending_new_patients: dict[UUID, _PendingNewPatient] = {}
    patient_code_to_id: dict[str, UUID] = {}
    # 既存患者 (alive) は code → id を載せる
    for existing in existing_patients.values():
        if existing.code:
            patient_code_to_id[existing.code] = existing.id
    # 同 import 内の新規患者は仮 UUID を載せる
    for op in patient_ops:
        if op.get("_op") != "new":
            continue
        new_id: UUID | None = op.get("_new_patient_id")
        new_code: str | None = op.get("code")
        if new_id is None or new_code is None:
            continue
        patient_code_to_id[new_code] = new_id
        pending_new_patients[new_id] = _PendingNewPatient(
            patient_id=new_id,
            code=new_code,
            primary_office_id=op.get("primary_office_id"),
        )
    # resurrection 対象患者も PFV から code リンクで参照できるようにする.
    # primary_office_id は更新後の値があればそれを、無ければ既存 (旧) 値を使う.
    deleted_by_id: dict[UUID, Patient] = {p.id: p for p in deleted_patients_by_code.values()}
    for op in patient_ops:
        if op.get("_op") != "resurrect":
            continue
        resurrect_pid: UUID = op["_patient_id"]
        original = deleted_by_id.get(resurrect_pid)
        if original is None or not original.code:
            continue
        updated_office = op["_updates"].get("primary_office_id", original.primary_office_id)
        patient_code_to_id[original.code] = resurrect_pid
        pending_new_patients[resurrect_pid] = _PendingNewPatient(
            patient_id=resurrect_pid,
            code=original.code,
            primary_office_id=updated_office,
        )

    # ---- PFV シート ----
    pfv_rows: list[PfvExcelImportRow] = []
    pfv_ops: list[dict[str, Any]] = []
    if use_pfv_edit_sheet:
        # Phase G-51: 新「固定訪問パターン（編集用）」(患者 1 行) を per-patient replace で処理.
        ws_f = wb[SHEET_PFV_EDIT]
        seen_edit_pids: set[UUID] = set()
        for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            diff_rows_edit, ops_edit = _parse_pfv_edit_patient(
                r_idx,
                row,
                existing_patients=existing_patients,
                pending_new_patients=pending_new_patients,
                patient_code_to_id=patient_code_to_id,
                course_templates=course_templates,
                existing_pfvs=existing_pfvs,
                offices_by_code=offices_by_code,
            )
            # 同一患者が複数行に出てきた場合の二重処理を防ぐ (1 行目を採用).
            for dr in diff_rows_edit:
                if (
                    dr.operation != "error"
                    and dr.patient_id is not None
                    and dr.patient_id in seen_edit_pids
                ):
                    # 既に処理済 patient の重複行: error 化して誤削除を防ぐ.
                    pfv_rows.append(
                        PfvExcelImportRow(
                            row_number=r_idx,
                            patient_id=dr.patient_id,
                            patient_code=dr.patient_code,
                            operation="error",
                            error_message="同ファイル内で同一患者が複数行に出現しています",
                        )
                    )
                    break
            else:
                pfv_rows.extend(diff_rows_edit)
                pfv_ops.extend(ops_edit)
                for dr in diff_rows_edit:
                    if dr.operation != "error" and dr.patient_id is not None:
                        seen_edit_pids.add(dr.patient_id)
    else:
        # 旧 per-visit シート (後方互換 fallback).
        assert pfv_sheet_name is not None
        ws_f = wb[pfv_sheet_name]
        pending_new_keys: set[tuple[UUID, str, int, int]] = set()
        for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            diff_row, op = _parse_pfv_row(
                r_idx,
                row,
                existing_patients=existing_patients,
                pending_new_patients=pending_new_patients,
                patient_code_to_id=patient_code_to_id,
                course_templates=course_templates,
                existing_pfvs=existing_pfvs,
                pending_new_keys=pending_new_keys,
                # Phase E-5: sub_office_code 解決用 (患者シートと共有).
                offices_by_code=offices_by_code,
            )
            pfv_rows.append(diff_row)
            if op is not None:
                pfv_ops.append(op)

    # ---- weekly_pattern マージ ----
    # weekly_pattern dict (空 dict = clear) を該当 patient op に統合する共通処理.
    # 既存値と同じ場合は skip (= noop 維持、round-trip 安定化).
    def _merge_weekly_updates(weekly_pattern_updates: dict[UUID, dict[str, Any] | None]) -> None:
        ops_by_pid: dict[UUID, dict[str, Any]] = {}
        for op in patient_ops:
            op_pid = op.get("_patient_id") or op.get("_new_patient_id")
            if op_pid is not None:
                ops_by_pid[op_pid] = op

        for pid, wp in weekly_pattern_updates.items():
            existing_patient = existing_patients.get(pid)
            existing_wp = existing_patient.weekly_pattern if existing_patient else None
            existing_op = ops_by_pid.get(pid)
            # 既存 op が weekly_pattern を既にセット済 (= 同 import で先に統合済) なら
            # それを merge 基準にする (統合シート → 旧シートの順で上書きされ得るため).
            if existing_op is not None and "weekly_pattern" in existing_op:
                current_wp = existing_op["weekly_pattern"]
            elif existing_op is not None and existing_op.get("_op") == "resurrect":
                current_wp = existing_op.get("_updates", {}).get("weekly_pattern", existing_wp)
            else:
                current_wp = existing_wp
            # Phase G-48 hotfix: 丸ごと置換ではなく merge. current_wp の管理外キー
            # (entries/staff_count 等) を保持しつつ、管理 8 キーのみ上書きする.
            # noop 判定も merge 後の dict で行う (entries 持ち患者の無編集 round-trip を
            # 真の noop にするため).
            merged_wp = _merge_weekly_pattern(current_wp, wp)
            if merged_wp == current_wp:
                continue  # 差分なし → DB 更新不要

            if existing_op is None:
                # No existing op (= noop だった patient or weekly only update)
                # → 新規 update op を patient_ops に追加.
                new_op = {
                    "_op": "update",
                    "_patient_id": pid,
                    "weekly_pattern": merged_wp,
                }
                patient_ops.append(new_op)
                ops_by_pid[pid] = new_op
                # patient_rows 側の operation も "update" に上書き (noop → update)
                for pr in patient_rows:
                    if pr.patient_id == pid and pr.operation == "noop":
                        pr.operation = "update"
                        pr.changes.append(
                            PatientExcelChange(
                                field="weekly_pattern",
                                old_value=_serializable(existing_wp),
                                new_value="(希望訪問パターン更新)",
                            )
                        )
                        break
            else:
                op_type = existing_op.get("_op")
                if op_type == "delete":
                    # delete op に weekly_pattern を統合しない (削除されるため)
                    continue
                if op_type == "resurrect":
                    existing_op.setdefault("_updates", {})["weekly_pattern"] = merged_wp
                else:  # new / update
                    existing_op["weekly_pattern"] = merged_wp

    # Phase G-48: 統合「患者マスタ」シートから読み取った weekly_pattern をマージ.
    _merge_weekly_updates(weekly_from_patient_sheet)

    # ---- 【後方互換】旧 独立「希望訪問パターン」シート (Phase E-8 / 3 シート構成) ----
    # SHEET_WEEKLY は optional. 旧 Excel ファイルとの後方互換のためのみ読み込む.
    # 統合シートより後にマージすることで、旧シートに明示された値が優先される.
    if SHEET_WEEKLY in wb.sheetnames:
        ws_w = wb[SHEET_WEEKLY]
        legacy_weekly_updates: dict[UUID, dict[str, Any] | None] = {}
        for r_idx, row in enumerate(ws_w.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            result = _parse_weekly_row(
                r_idx,
                row,
                existing_patients=existing_patients,
                existing_patients_by_code=existing_patients_by_code,
                patient_code_to_id=patient_code_to_id,
                pending_new_patients=pending_new_patients,
            )
            if result is None:
                continue
            pid, wp = result
            legacy_weekly_updates[pid] = wp if wp else None
        _merge_weekly_updates(legacy_weekly_updates)

    # ---- summary ----
    summary = PatientExcelImportSummary()
    for r in patient_rows:
        match r.operation:
            case "new":
                summary.patients_new += 1
            case "update":
                summary.patients_update += 1
            case "delete":
                summary.patients_delete += 1
            case "error":
                summary.patients_error += 1
            case "noop":
                summary.patients_noop += 1
    for r in pfv_rows:
        match r.operation:
            case "new":
                summary.pfv_new += 1
            case "update":
                summary.pfv_update += 1
            case "delete":
                summary.pfv_delete += 1
            case "error":
                summary.pfv_error += 1
            case "noop":
                summary.pfv_noop += 1

    return patient_rows, pfv_rows, summary, patient_ops, pfv_ops


async def apply_changes(
    db: AsyncSession,
    *,
    patient_ops: list[dict[str, Any]],
    pfv_ops: list[dict[str, Any]],
) -> None:
    """差分を DB に反映する (partial commit).

    ``patient_ops`` / ``pfv_ops`` は ``parse_and_diff`` の時点で error 行が除外
    された有効な op (new / update / delete) のみを含む. error 行は ``op=None``
    として既に skip されているので、ここでは触れない.
    SQLAlchemy 例外は呼び出し側にバブルアップする (rollback は呼び出し側で実施).
    """
    # 1) Patient: new → insert / update → 既存行を更新 / delete → soft delete /
    #    resurrect → deleted_at=NULL に戻して内容上書き
    patients_by_id: dict[UUID, Patient] = {}
    # NG スタッフ (中間テーブル) の差分は patient flush 後にまとめて適用する.
    ng_staff_updates: list[tuple[UUID, list[UUID]]] = []
    # soft delete した患者 (中間テーブル掃除の対象).
    soft_deleted_patient_ids: list[UUID] = []
    for op in patient_ops:
        op_type = op.get("_op")
        if op_type == "new":
            data = {k: v for k, v in op.items() if not k.startswith("_")}
            new_patient = Patient(**data)
            db.add(new_patient)
            if "_ng_staff_ids" in op:
                ng_staff_updates.append((op["_new_patient_id"], op["_ng_staff_ids"]))
        elif op_type == "resurrect":
            pid = op["_patient_id"]
            if pid not in patients_by_id:
                patients_by_id[pid] = await db.get(Patient, pid)
            patient = patients_by_id[pid]
            if patient is None:
                continue
            patient.deleted_at = None
            for k, v in op["_updates"].items():
                setattr(patient, k, v)
            if "_ng_staff_ids" in op:
                ng_staff_updates.append((pid, op["_ng_staff_ids"]))
        elif op_type == "update":
            pid = op["_patient_id"]
            if pid not in patients_by_id:
                patients_by_id[pid] = await db.get(Patient, pid)
            patient = patients_by_id[pid]
            if patient is None:
                continue
            for k, v in op.items():
                if k.startswith("_"):
                    continue
                setattr(patient, k, v)
            if "_ng_staff_ids" in op:
                ng_staff_updates.append((pid, op["_ng_staff_ids"]))
        elif op_type == "delete":
            pid = op["_patient_id"]
            if pid not in patients_by_id:
                patients_by_id[pid] = await db.get(Patient, pid)
            patient = patients_by_id[pid]
            if patient is not None:
                patient.deleted_at = func.now()
                soft_deleted_patient_ids.append(pid)
                # 関連 PFV を物理削除 (soft-delete 患者に orphan PFV が残ると
                # スケジュール生成側でゴースト枠として現れるのを防ぐ).
                pfvs_to_cleanup = (
                    await db.scalars(
                        select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid)
                    )
                ).all()
                for pfv in pfvs_to_cleanup:
                    await db.delete(pfv)
    # soft delete では FK CASCADE が発火しないため、中間テーブル (同住所リンク /
    # NG スタッフ) をアプリ層で明示削除する (delete_patient endpoint と同じ流儀).
    await cleanup_soft_deleted_patient_links(db, soft_deleted_patient_ids)
    # 中間 flush で patient row を確定 (PFV new で patient_id を再参照する可能性).
    await db.flush()

    # 1-b) NG スタッフ (patient_ng_staff): 新規患者行の INSERT 後に差分適用.
    await apply_ng_staff_updates(db, ng_staff_updates)

    # 2) PFV: new → insert / update → 既存行を更新 / delete → 物理削除
    pfvs_by_id: dict[UUID, PatientFixedVisit] = {}
    for op in pfv_ops:
        op_type = op.get("_op")
        if op_type == "new":
            data = {k: v for k, v in op.items() if not k.startswith("_")}
            db.add(PatientFixedVisit(**data))
        elif op_type == "update":
            pfv_id = op["_pfv_id"]
            if pfv_id not in pfvs_by_id:
                pfvs_by_id[pfv_id] = await db.get(PatientFixedVisit, pfv_id)
            pfv = pfvs_by_id[pfv_id]
            if pfv is None:
                continue
            for k, v in op.items():
                if k.startswith("_"):
                    continue
                setattr(pfv, k, v)
        elif op_type == "delete":
            pfv_id = op["_pfv_id"]
            obj = await db.get(PatientFixedVisit, pfv_id)
            if obj is not None:
                await db.delete(obj)
    await db.flush()
