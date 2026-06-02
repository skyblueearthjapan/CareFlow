"""患者個別カルテ (A4・1 患者 1 ファイル) の Excel 入出力.

カルテは 1 シート (A4 縦) の固定セルレイアウトに患者 1 名の基本情報・訪問条件・
固定訪問スケジュール (週) を配置する. 既存の患者マスタ Excel パイプライン
(``schema`` / ``importer`` / ``exporter``) を最大限再利用し、**round-trip 0 差分**
(DB の実患者 → ``build_karte_workbook`` → ``parse_karte_workbook`` → diff が全 noop)
を不変条件とする.

公開 API:
  * ``build_karte_workbook``      — 1 患者のカルテ Workbook (値入り + dropdown + 体裁).
  * ``build_blank_karte_template`` — 空白カルテ (値なし・dropdown と体裁のみ).
  * ``parse_karte_workbook``       — カルテを読み、**既存 importer が解釈できる標準
    2 シート Workbook の bytes** に変換する. API はこの bytes を既存 ``parse_and_diff``
    に渡すことで、患者 upsert / weekly merge / PFV 患者単位 replace を流用する.

設計方針 (reuse):
  カルテ固有のセル⇔項目マップは本モジュールに閉じ込め、差分計算・DB 反映は
  既存パイプラインに完全に乗せる. ``parse_karte_workbook`` は固定セルを読み、
  「患者マスタ」シート 1 行 + 「固定訪問パターン（編集用）」シート 1 行 (= G-55
  患者単位 replace セマンティクス) に組み替えた標準 Workbook を生成する.
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import time
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.models.course_template import CourseTemplate
from app.models.office import Office
from app.models.patient import Patient
from app.models.patient_fixed_visit import PatientFixedVisit
from app.services.patient_excel.exporter import _course_token_dropdown_values
from app.services.patient_excel.schema import (
    BOOL_JA_VALUES,
    FREQUENCY_PER_WEEK_VALUES,
    HEADER_FONT_COLOR,
    HHMM_VALUES,
    ID_COLUMN_FILL_COLOR,
    ID_COLUMN_FONT_COLOR,
    INSURANCE_EN_TO_JA,
    OFFICE_CODE_VALUES,
    PATIENT_COL_INDEX,
    PATIENT_COLUMNS,
    PATIENT_WEEKDAY_COL_KEYS,
    PFV_EDIT_COL_INDEX,
    PFV_EDIT_COLUMNS,
    PFV_EDIT_WEEKDAY_KEY_TO_INT,
    PFV_EDIT_WEEKDAY_KEYS,
    SERVICE_MINUTES_VALUES,
    SEX_EN_TO_JA,
    SEX_JA_VALUES,
    SEX_RESTRICTION_EN_TO_JA,
    SEX_RESTRICTION_JA_VALUES,
    SHEET_PATIENTS,
    SHEET_PFV_EDIT,
    STATUS_EN_TO_JA,
    STATUS_JA_VALUES,
    TIME_TYPE_GREYOUT_VALUES,
    TIME_TYPE_VALUES,
    VISIT_FREQUENCY_EN_TO_JA,
    VISIT_FREQUENCY_JA_VALUES,
    WEEKDAY_LABELS,
    WEEKDAY_MARK_VALUES,
    course_token,
    weekdays_en_to_yesno_cells,
)

# ---------------------------------------------------------------------------
# シート名 / レイアウト定数
# ---------------------------------------------------------------------------

SHEET_KARTE: str = "訪問看護スケジューリング用カルテ"
KARTE_TITLE: str = "訪問看護スケジューリング用カルテ"

# 拠点コード → 自動表示用ラベル (B6). 「{office_label}（自動）」.
OFFICE_AUTO_SUFFIX: str = "（自動）"

# 印刷範囲 / 体裁.
KARTE_PRINT_AREA: str = "A1:H25"
KARTE_FIT_SCALE: int = 91  # fitToPage 91%

# 月..日 の 7 列を B..H に割り当てる (固定訪問スケジュール / 希望曜日).
# B=2 .. H=8. WEEKDAY_LABELS (月火水木金土日) と同順.
WEEKDAY_COL_LETTERS: tuple[str, ...] = ("B", "C", "D", "E", "F", "G", "H")

# 固定訪問の「休み」表示マーカー (時刻/コース両方).
REST_MARK: str = "―"

# 希望曜日 7 セル (B11:H11). 希望曜日は日曜も有効 (weekly path は 7 日対応).
PREF_WEEKDAY_CELLS: tuple[str, ...] = tuple(f"{c}11" for c in WEEKDAY_COL_LETTERS)
# 固定訪問の開始時刻 7 セル (B18:H18) とコース 7 セル (B19:H19).
FIXED_TIME_CELLS: tuple[str, ...] = tuple(f"{c}18" for c in WEEKDAY_COL_LETTERS)
FIXED_COURSE_CELLS: tuple[str, ...] = tuple(f"{c}19" for c in WEEKDAY_COL_LETTERS)

# 固定訪問スケジュールの月..土 列 (B..G). 日曜 (H 列) は固定訪問の取込対象外
# (PFV 編集パイプライン = PFV_EDIT_WEEKDAY_KEYS は月..土のみ). dropdown は月..土だけ.
FIXED_TIME_CELLS_MON_SAT: tuple[str, ...] = tuple(f"{c}18" for c in WEEKDAY_COL_LETTERS[:6])
FIXED_COURSE_CELLS_MON_SAT: tuple[str, ...] = tuple(f"{c}19" for c in WEEKDAY_COL_LETTERS[:6])
# 固定訪問の日曜セル (時刻/コース). グレー塗り + 斜体・dropdown 無しで「取込対象外」を示す.
FIXED_SUNDAY_CELLS: tuple[str, ...] = (
    f"{WEEKDAY_COL_LETTERS[6]}18",
    f"{WEEKDAY_COL_LETTERS[6]}19",
)
# 固定訪問の日曜が取込対象外である旨の注記. A16 見出し (A16:H16 merge) に併記する.
FIXED_SUNDAY_NOTE_CELL: str = "A16"
FIXED_SUNDAY_NOTE_TEXT: str = "※固定訪問の日曜は取込対象外"
# 固定訪問スケジュール見出し (注記を併記した最終文字列).
FIXED_SCHEDULE_HEADING: str = f"■ 固定訪問スケジュール（週）　{FIXED_SUNDAY_NOTE_TEXT}"


# ---------------------------------------------------------------------------
# 体裁 helpers
# ---------------------------------------------------------------------------


def _apply_print_setup(ws: Worksheet) -> None:
    """A4 縦 / 印刷範囲 A1:H25 / fitToPage 91% を設定."""
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.scale = KARTE_FIT_SCALE
    ws.print_area = KARTE_PRINT_AREA


def _style_layout(ws: Worksheet) -> None:
    """カルテの静的体裁 (タイトル / セクション見出し / 静的見出し / ラベル / 列幅 / merge)."""
    title_font = Font(bold=True, size=16, color=HEADER_FONT_COLOR)
    title_fill = PatternFill("solid", fgColor="FF4472C4")
    section_font = Font(bold=True, size=12, color="FF222222")
    section_fill = PatternFill("solid", fgColor="FFD9D9D9")
    label_font = Font(bold=True, color="FF444444")
    static_font = Font(bold=True, color="FF333333")
    static_fill = PatternFill("solid", fgColor="FFDDEBF7")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 列幅 (A=ラベル, B..H=値). A4 縦に収まるよう抑え目.
    widths = {"A": 12, "B": 12, "C": 8, "D": 12, "E": 8, "F": 12, "G": 10, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- merge (サンプル準拠. ただし B13:G13 は解除し 3 セル運用) ----
    for rng in ("B2:C2", "A4:H4", "D6:G6", "A8:H8", "A16:H16", "A21:H21", "A22:H25"):
        ws.merge_cells(rng)

    # ---- タイトル (A1) ----
    a1 = ws["A1"]
    a1.value = KARTE_TITLE
    a1.font = title_font
    a1.fill = title_fill
    a1.alignment = left

    # ---- コードラベル F1 ----
    ws["F1"].value = "コード:"
    ws["F1"].font = label_font
    ws["F1"].alignment = Alignment(horizontal="right", vertical="center")

    # ---- 氏名 / フリガナ (行 2) ----
    ws["A2"].value = "氏名:"
    ws["D2"].value = "フリガナ:"
    for ref in ("A2", "D2"):
        ws[ref].font = label_font
        ws[ref].alignment = left

    # ---- セクション見出し ----
    for ref, text in (
        ("A4", "■ 基本情報"),
        ("A8", "■ 訪問条件"),
        ("A16", FIXED_SCHEDULE_HEADING),
        ("A21", "■ 備考 / ヒアリングメモ"),
    ):
        ws[ref].value = text
        ws[ref].font = section_font
        ws[ref].fill = section_fill
        ws[ref].alignment = left

    # ---- 基本情報ラベル (行 5/6) ----
    for ref, text in (
        ("A5", "性別:"),
        ("C5", "ステータス:"),
        ("E5", "保険:"),
        ("A6", "拠点:"),
        ("C6", "住所:"),
    ):
        ws[ref].value = text
        ws[ref].font = label_font
        ws[ref].alignment = left

    # ---- 訪問条件ラベル (行 9/10/11/12/13/14) ----
    for ref, text in (
        ("A9", "週訪問回数:"),
        ("C9", "訪問頻度:"),
        ("A10", "希望曜日:"),
        ("A11", "希望"),
        ("A12", "サービス時間:"),
        ("C12", "時間タイプ:"),
        ("A13", "希望時刻:"),
        ("A14", "性別制限:"),
        ("C14", "複数スタッフ:"),
    ):
        ws[ref].value = text
        ws[ref].font = label_font
        ws[ref].alignment = left

    # 希望時刻の区切り「〜」(C13 静的).
    ws["C13"].value = "〜"
    ws["C13"].alignment = center

    # ---- 静的曜日見出し (行 10: 希望曜日見出し / 行 17: 固定訪問見出し) ----
    for col, label in zip(WEEKDAY_COL_LETTERS, WEEKDAY_LABELS, strict=True):
        for r in (10, 17):
            cell = ws[f"{col}{r}"]
            cell.value = label
            cell.font = static_font
            cell.fill = static_fill
            cell.alignment = center

    # ---- 固定訪問スケジュールの行ラベル (A18 時刻 / A19 コース) ----
    for ref, text in (("A18", "時刻"), ("A19", "コース")):
        ws[ref].value = text
        ws[ref].font = label_font
        ws[ref].alignment = left

    # 値セル (希望曜日 / 固定訪問時刻・コース) は中央寄せ.
    for ref in (
        *PREF_WEEKDAY_CELLS,
        *FIXED_TIME_CELLS,
        *FIXED_COURSE_CELLS,
        "B5",
        "D5",
        "F5",
        "B6",
        "B9",
        "D9",
        "B12",
        "D12",
        "B13",
        "D13",
        "B14",
        "D14",
    ):
        ws[ref].alignment = center

    # 備考は左上寄せ + 折り返し.
    ws["A22"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _add_list_validation(ws: Worksheet, cell: str, values: Sequence[str]) -> None:
    """単一セルに list DataValidation を付ける."""
    formula = '"' + ",".join(str(v) for v in values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "リストにない値です"
    dv.errorTitle = "入力エラー"
    dv.add(cell)
    ws.add_data_validation(dv)


def _attach_dropdowns(ws: Worksheet, course_tokens: Sequence[str]) -> None:
    """カルテの全プルダウンを付与 (確定仕様のセル割当)."""
    _add_list_validation(ws, "B5", SEX_JA_VALUES)  # 性別
    _add_list_validation(ws, "D5", STATUS_JA_VALUES)  # ステータス
    _add_list_validation(ws, "F5", INSURANCE_JA_SHORT)  # 保険 (医療/介護)
    _add_list_validation(ws, "B6", OFFICE_LABEL_VALUES)  # 拠点 (稲毛/都賀)
    _add_list_validation(ws, "B9", FREQUENCY_PER_WEEK_VALUES)  # 週回数 1-7
    _add_list_validation(ws, "D9", VISIT_FREQUENCY_JA_VALUES)  # 頻度
    for cell in PREF_WEEKDAY_CELLS:  # 希望曜日 〇/×
        _add_list_validation(ws, cell, WEEKDAY_MARK_VALUES)
    _add_list_validation(ws, "B12", SERVICE_MINUTES_JA_VALUES)  # サービス時間 「N分」
    _add_list_validation(ws, "D12", TIME_TYPE_VALUES)  # 時間タイプ
    _add_list_validation(ws, "B13", HHMM_VALUES)  # 希望開始
    _add_list_validation(ws, "D13", HHMM_VALUES)  # 希望終了
    _add_list_validation(ws, "B14", SEX_RESTRICTION_JA_VALUES)  # 性別制限
    _add_list_validation(ws, "D14", BOOL_JA_VALUES)  # 複数スタッフ
    # 固定訪問時刻 (HH:MM + 「―」) / コース (動的トークン + 「―」).
    # dropdown は月..土のみ. 日曜 (H 列) は固定訪問の取込対象外のため dropdown を付けない
    # (グレー塗り + 斜体で視覚的に「非対応」を明示し、parse 側も日曜セルを読まない).
    time_values = (*HHMM_VALUES, REST_MARK)
    course_values = (*course_tokens, REST_MARK)
    for cell in FIXED_TIME_CELLS_MON_SAT:
        _add_list_validation(ws, cell, time_values)
    for cell in FIXED_COURSE_CELLS_MON_SAT:
        _add_list_validation(ws, cell, course_values)


def _attach_time_greyout(ws: Worksheet) -> None:
    """時間タイプ (D12) が 午前/午後/終日 のとき 希望時刻 (B13/D13) をグレーアウト.

    既存 exporter の条件付き書式 (FormulaRule) の作法を流用. 時間タイプセル D12 を
    絶対参照し、希望開始/終了の 2 セルに適用する.
    """
    conditions = ",".join(f'$D$12="{v}"' for v in TIME_TYPE_GREYOUT_VALUES)
    formula = f"OR({conditions})"
    grey_fill = PatternFill("solid", fgColor=ID_COLUMN_FILL_COLOR)
    grey_font = Font(italic=True, color=ID_COLUMN_FONT_COLOR)
    for cell in ("B13", "D13"):
        rule = FormulaRule(formula=[formula], fill=grey_fill, font=grey_font)
        ws.conditional_formatting.add(cell, rule)


def _style_sunday_fixed_unsupported(ws: Worksheet) -> None:
    """固定訪問の日曜セル (H18 時刻 / H19 コース) を「取込対象外」として体裁付与.

    固定訪問パイプライン (PFV_EDIT_WEEKDAY_KEYS) は月..土のみを扱うため、カルテの
    日曜固定訪問セルはグレー塗り + 斜体にし、dropdown を付けない (非対応の明示).
    注記 (FIXED_SUNDAY_NOTE_TEXT) は A16 見出し (A16:H16 merge) に併記済み.
    希望曜日 (B11:H11) の日曜は従来どおり有効であり、ここでは触らない
    (weekly path は 7 日対応).
    """
    grey_fill = PatternFill("solid", fgColor=ID_COLUMN_FILL_COLOR)
    grey_font = Font(italic=True, color=ID_COLUMN_FONT_COLOR)
    for cell in FIXED_SUNDAY_CELLS:
        ws[cell].fill = grey_fill
        ws[cell].font = grey_font


# 保険は短縮表記 (医療/介護) で統一 (取込は「医療保険/介護保険」も後方互換受理).
INSURANCE_EN_TO_SHORT_JA: dict[str, str] = {"medical": "医療", "care": "介護"}
INSURANCE_JA_SHORT: tuple[str, ...] = ("医療", "介護")
# 短縮ラベル → importer 正準ラベル (INSURANCE_JA_VALUES). import 時の橋渡し.
INSURANCE_SHORT_JA_TO_FULL: dict[str, str] = {
    "医療": INSURANCE_EN_TO_JA["medical"],
    "介護": INSURANCE_EN_TO_JA["care"],
}
# サービス時間は「N分」表記の dropdown (取込時は「分」除去).
SERVICE_MINUTES_JA_VALUES: tuple[str, ...] = tuple(f"{v}分" for v in SERVICE_MINUTES_VALUES)
# 拠点 dropdown は表示ラベル (稲毛/都賀).
OFFICE_CODE_TO_LABEL: dict[str, str] = {"INAGE": "稲毛", "TSUGA": "都賀"}
OFFICE_LABEL_TO_CODE: dict[str, str] = {v: k for k, v in OFFICE_CODE_TO_LABEL.items()}
OFFICE_LABEL_VALUES: tuple[str, ...] = tuple(OFFICE_CODE_TO_LABEL.values())


# ---------------------------------------------------------------------------
# build_karte_workbook / build_blank_karte_template
# ---------------------------------------------------------------------------


def _minutes_to_service_label(minutes: int | None) -> str | None:
    """service_minutes (int) → 「60分」表記. None はそのまま None."""
    if minutes is None:
        return None
    return f"{minutes}分"


def _hhmm(t: time | None) -> str | None:
    if t is None:
        return None
    return f"{t.hour:02d}:{t.minute:02d}"


def build_karte_workbook(
    *,
    patient: Patient,
    fixed_visits: Sequence[PatientFixedVisit],
    offices: Sequence[Office],
    course_templates: Sequence[CourseTemplate],
    office_label: str | None = None,
    office_code: str | None = None,
) -> Workbook:
    """1 患者のカルテ Workbook を構築する (A4 縦・固定セル).

    ``fixed_visits`` はこの患者の PFV (全 mode/slot). normal/slot_index=0 のみを
    曜日別の固定訪問スケジュールに展開する (special / slot1 は本シート対象外で保持).
    ``office_label`` / ``office_code`` が省略された場合は ``patient.primary_office_id``
    から ``offices`` を引いて解決する. 拠点セル (B6) は「{office_label}（自動）」表示.
    """
    office_by_id: dict[UUID, Office] = {o.id: o for o in offices}
    office_code_by_id: dict[UUID, str] = {o.id: o.code for o in offices if o.code}

    # 拠点ラベル / コードの解決.
    resolved_office: Office | None = None
    if patient.primary_office_id is not None:
        resolved_office = office_by_id.get(patient.primary_office_id)
    if office_code is None and resolved_office is not None:
        office_code = resolved_office.code
    if office_label is None:
        if resolved_office is not None:
            # 表示は拠点名優先. 無ければ code → ラベル. 最終 fallback は code.
            office_label = resolved_office.name or OFFICE_CODE_TO_LABEL.get(
                resolved_office.code or "", resolved_office.code or ""
            )
        elif office_code:
            office_label = OFFICE_CODE_TO_LABEL.get(office_code, office_code)

    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = SHEET_KARTE

    _style_layout(ws)
    course_tokens = _course_token_dropdown_values(offices, course_templates)
    _attach_dropdowns(ws, course_tokens)
    _attach_time_greyout(ws)
    _apply_print_setup(ws)

    # ---- 値の書き込み ----
    # G1: 患者コード (突合キー).
    ws["G1"].value = patient.code
    # B2: 氏名 / F2: フリガナ.
    ws["B2"].value = patient.name
    ws["F2"].value = patient.kana
    # B5: 性別 / D5: ステータス / F5: 保険.
    ws["B5"].value = SEX_EN_TO_JA.get(patient.sex, patient.sex) if patient.sex else None
    ws["D5"].value = STATUS_EN_TO_JA.get(patient.status, patient.status) if patient.status else None
    ws["F5"].value = (
        INSURANCE_EN_TO_SHORT_JA.get(patient.insurance, patient.insurance)
        if patient.insurance
        else None
    )
    # B6: 拠点 (自動表示). D6: 住所.
    ws["B6"].value = f"{office_label}{OFFICE_AUTO_SUFFIX}" if office_label else None
    ws["D6"].value = patient.address

    # ---- 訪問条件 (weekly_pattern) ----
    wp: dict[str, Any] = patient.weekly_pattern if isinstance(patient.weekly_pattern, dict) else {}
    ws["B9"].value = wp.get("frequency_per_week")
    vf = wp.get("visit_frequency")
    ws["D9"].value = VISIT_FREQUENCY_EN_TO_JA.get(vf, vf) if isinstance(vf, str) else None

    # 希望曜日 B11:H11 (〇/×).
    pref_weekdays = wp.get("preferred_weekdays")
    yesno = weekdays_en_to_yesno_cells(pref_weekdays if isinstance(pref_weekdays, list) else None)
    # yesno は pref_wd_mon..sun キー. PATIENT_WEEKDAY_COL_KEYS は月..日順 = B..H.
    for col, key in zip(WEEKDAY_COL_LETTERS, PATIENT_WEEKDAY_COL_KEYS, strict=True):
        ws[f"{col}11"].value = yesno[key]

    # ---- 固定訪問スケジュール (週): normal / slot_index=0 のみ ----
    # B12 (サービス時間) より先に PFV 代表 duration を導く. サービス時間は
    # 固定訪問パターン（編集用）シートの duration 源泉でもあり、PFV の duration_min と
    # 一致していないと round-trip で偽の update (duration 差分) が出る.
    course_template_by_id: dict[UUID, CourseTemplate] = {ct.id: ct for ct in course_templates}
    by_weekday: dict[int, PatientFixedVisit] = {}
    for pfv in fixed_visits:
        if pfv.mode != "normal" or pfv.slot_index != 0:
            continue
        by_weekday.setdefault(pfv.weekday, pfv)
    # 代表 duration: 最初の曜日 (昇順) の PFV duration_min (exporter と同じ作法).
    rep_duration: int | None = None
    for wd in sorted(by_weekday):
        rep_duration = by_weekday[wd].duration_min
        break

    # サービス時間 / 時間タイプ / 希望時刻 / 性別制限 / 複数スタッフ.
    # service_minutes は weekly_pattern を優先し、無ければ PFV 代表 duration を採用
    # (round-trip 安定: 固定訪問 duration と一致させる).
    sm = wp.get("service_minutes")
    sm_value = sm if isinstance(sm, int) else rep_duration
    ws["B12"].value = _minutes_to_service_label(sm_value)
    tt = wp.get("time_type")
    ws["D12"].value = tt if isinstance(tt, str) and tt in TIME_TYPE_VALUES else None
    ws["B13"].value = wp.get("preferred_start")
    ws["D13"].value = wp.get("preferred_end")
    ws["B14"].value = (
        SEX_RESTRICTION_EN_TO_JA.get(patient.sex_restriction, patient.sex_restriction)
        if patient.sex_restriction
        else "なし"
    )
    ws["D14"].value = "はい" if patient.requires_multiple_staff else "いいえ"

    # 固定訪問スケジュールは月..土のみ展開する (日曜=weekday 6 は取込対象外で、
    # H18/H19 はグレー塗り・dropdown 無し. 既存の日曜 normal slot0 PFV は import で
    # 温存されるが、カルテ上では値を書かず「非対応」表示に統一する).
    for col, weekday in zip(WEEKDAY_COL_LETTERS[:6], range(6), strict=True):
        pfv = by_weekday.get(weekday)
        if pfv is None:
            ws[f"{col}18"].value = REST_MARK
            ws[f"{col}19"].value = REST_MARK
            continue
        ws[f"{col}18"].value = _hhmm(pfv.start_time)
        token: str | None = None
        if pfv.course_template_id is not None:
            ct = course_template_by_id.get(pfv.course_template_id)
            if ct is not None and ct.label:
                oc = office_code_by_id.get(ct.office_id)
                if oc:
                    token = course_token(oc, ct.label)
        ws[f"{col}19"].value = token if token is not None else REST_MARK

    # 固定訪問の日曜セルを「取込対象外」として体裁付与 (グレー + 斜体 + 注記).
    _style_sunday_fixed_unsupported(ws)

    # ---- 備考 ----
    ws["A22"].value = patient.note

    return wb


def build_blank_karte_template(
    *,
    offices: Sequence[Office],
    course_templates: Sequence[CourseTemplate],
) -> Workbook:
    """空白カルテ (値なし・dropdown と体裁のみ). 新規ヒアリング記入用.

    固定訪問スケジュールは空欄 (= 記入されなければ訪問なし). 「―」は明示しない
    (空白テンプレなので dropdown 候補にのみ存在).
    """
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = SHEET_KARTE

    _style_layout(ws)
    course_tokens = _course_token_dropdown_values(offices, course_templates)
    _attach_dropdowns(ws, course_tokens)
    _attach_time_greyout(ws)
    # 固定訪問の日曜セルを「取込対象外」として体裁付与 (グレー + 斜体 + 注記).
    _style_sunday_fixed_unsupported(ws)
    _apply_print_setup(ws)
    return wb


# ---------------------------------------------------------------------------
# parse_karte_workbook — カルテ → 標準 2 シート Workbook bytes
# ---------------------------------------------------------------------------


def _read_cell_str(ws: Worksheet, ref: str) -> str | None:
    """セル値を str に正規化 (空/空白は None). time/数値も文字列化."""
    v = ws[ref].value
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, time):
        return f"{v.hour:02d}:{v.minute:02d}"
    return str(v).strip() or None


def _strip_office_auto_suffix(label: str | None) -> str | None:
    """「{ラベル}（自動）」suffix を除去. 空欄は None."""
    if label is None:
        return None
    s = label.strip()
    if s.endswith(OFFICE_AUTO_SUFFIX):
        s = s[: -len(OFFICE_AUTO_SUFFIX)].strip()
    return s or None


def _office_label_to_code(label: str | None) -> str | None:
    """拠点ラベル (稲毛/都賀) → office_code (INAGE/TSUGA). 解決不能は None (= 住所自動割当)."""
    if label is None:
        return None
    # 「（自動）」除去後のラベルで照合.
    s = _strip_office_auto_suffix(label)
    if s is None:
        return None
    if s in OFFICE_LABEL_TO_CODE:
        return OFFICE_LABEL_TO_CODE[s]
    # 後方互換: コードそのものが書かれていた場合.
    if s in OFFICE_CODE_VALUES:
        return s
    return None


def _normalize_insurance(label: str | None) -> str | None:
    """保険の短縮表記 (医療/介護) → importer が受理する正準ラベル (医療保険/介護保険).

    既存 importer の ``INSURANCE_ANY_TO_EN`` は「医療保険/介護保険/medical/care」のみ
    受理するため、カルテの短縮表記をここで正準ラベルへ橋渡しする (importer 非破壊).
    既に正準ラベル / 英語値 / その他はそのまま通す (importer 側で判定).
    """
    if label is None:
        return None
    s = label.strip()
    return INSURANCE_SHORT_JA_TO_FULL.get(s, s)


def _service_label_to_minutes(label: str | None) -> str | None:
    """「60分」→ "60". 「分」除去で数値文字列化. 数字以外は None.

    返り値は文字列 (importer の _read_int が解釈). 純数字でなければ None.
    """
    if label is None:
        return None
    s = label.strip().removesuffix("分").strip()
    return s if s.isdigit() else None


def _normalize_fixed_value(value: str | None) -> str | None:
    """固定訪問セルの値. 「―」/空欄 → None (= その曜日 訪問なし)."""
    if value is None:
        return None
    if value.strip() in (REST_MARK, "-", "ー", "─"):
        return None
    return value.strip() or None


def parse_karte_workbook(wb_or_bytes: Workbook | bytes) -> bytes:
    """カルテ Workbook (or bytes) を読み、**既存 importer 互換の標準 2 シート
    Workbook の bytes** に変換する.

    生成シート:
      * 「患者マスタ」: 1 行 (患者 upsert; weekly_pattern 統合列含む).
      * 「固定訪問パターン（編集用）」: 1 行 (= G-55 患者単位 replace; 各曜日 時刻+コース,
        空欄曜日 = normal slot0 削除).

    この bytes を既存 ``parse_and_diff`` に渡すことで、患者 upsert / weekly merge /
    PFV 患者単位 replace / dry_run プレビュー / ImportPreviewModal 互換応答を流用する.
    """
    if isinstance(wb_or_bytes, bytes):
        src = load_workbook(BytesIO(wb_or_bytes), data_only=True)
    else:
        src = wb_or_bytes
    if SHEET_KARTE in src.sheetnames:
        ws = src[SHEET_KARTE]
    else:
        # シート名が想定外でも最初のシートをカルテとして読む (柔軟性).
        ws = src[src.sheetnames[0]]

    # ---- 固定セル読み取り ----
    patient_code = _read_cell_str(ws, "G1")  # 空欄 → 新規
    name = _read_cell_str(ws, "B2")
    kana = _read_cell_str(ws, "F2")
    sex = _read_cell_str(ws, "B5")
    status = _read_cell_str(ws, "D5")
    insurance = _read_cell_str(ws, "F5")
    office_label = _read_cell_str(ws, "B6")
    address = _read_cell_str(ws, "D6")
    frequency_per_week = _read_cell_str(ws, "B9")
    visit_frequency = _read_cell_str(ws, "D9")
    service_label = _read_cell_str(ws, "B12")
    time_type = _read_cell_str(ws, "D12")
    preferred_start = _read_cell_str(ws, "B13")
    preferred_end = _read_cell_str(ws, "D13")
    sex_restriction = _read_cell_str(ws, "B14")
    requires_multiple_staff = _read_cell_str(ws, "D14")
    note = _read_cell_str(ws, "A22")

    office_code = _office_label_to_code(office_label)  # None → 住所自動割当
    service_minutes = _service_label_to_minutes(service_label)
    insurance = _normalize_insurance(insurance)  # 短縮「医療/介護」→ importer 受理形へ

    # 希望曜日 B11:H11 → pref_wd_* セル (〇/× をそのまま渡す; importer の _is_yes が解釈).
    pref_marks: dict[str, str | None] = {}
    for key, col in zip(PATIENT_WEEKDAY_COL_KEYS, WEEKDAY_COL_LETTERS, strict=True):
        pref_marks[key] = _read_cell_str(ws, f"{col}11")

    # ---- 標準 Workbook を組み立てる ----
    out = Workbook()
    ws_p: Worksheet = out.active  # type: ignore[assignment]
    ws_p.title = SHEET_PATIENTS

    # 患者マスタ ヘッダー (1 行目).
    for col_idx, col_def in enumerate(PATIENT_COLUMNS, start=1):
        ws_p.cell(row=1, column=col_idx, value=str(col_def["header"]))

    patient_values: dict[str, object | None] = {
        "patient_code": patient_code,
        "name": name,
        "kana": kana,
        "sex": sex,
        "status": status,
        "insurance": insurance,
        "address": address,
        # office_code 空欄 → importer が住所から自動割当.
        "office_code": office_code,
        "sex_restriction": sex_restriction,
        "requires_multiple_staff": requires_multiple_staff,
        "frequency_per_week": frequency_per_week,
        "visit_frequency": visit_frequency,
        **pref_marks,
        "service_minutes": service_minutes,
        "weekly_time_type": time_type,
        "preferred_start": preferred_start,
        "preferred_end": preferred_end,
        "note": note,
    }
    for col_key, idx in PATIENT_COL_INDEX.items():
        v = patient_values.get(col_key)
        if v is not None:
            ws_p.cell(row=2, column=idx + 1, value=v)

    # ---- 固定訪問パターン（編集用）シート ----
    ws_f: Worksheet = out.create_sheet(title=SHEET_PFV_EDIT)
    for col_idx, col_def in enumerate(PFV_EDIT_COLUMNS, start=1):
        ws_f.cell(row=1, column=col_idx, value=str(col_def["header"]))

    # 固定訪問: 月..土 (PFV_EDIT_WEEKDAY_KEYS) の各曜日 時刻+コース.
    # カルテは月..日 (7 列) だが PFV 編集シートは月..土 (6 列). 日曜の固定訪問は
    # 既存 PFV 編集パイプライン (G-55) が扱わないため本シートでは扱わない
    # (round-trip: 既存データも normal/slot0 は月..土に限られる運用).
    # MEDIUM (G-47 review): PFV の duration_min は B12 (weekly のサービス時間) を
    # 正本として正規化する. 編集用シートには各曜日の duration 列が無く、
    # importer (_parse_pfv_edit_patient) は service_minutes を全曜日 PFV の duration_min
    # に適用する. これは意図的挙動で、カルテ上は「サービス時間」1 値が固定訪問の
    # duration を兼ねる (本番確認で weekly service_minutes と PFV duration の不整合は 0 件).
    edit_values: dict[str, object | None] = {
        "patient_code": patient_code,
        "service_minutes": service_minutes,
        "time_type": time_type,
    }
    # WEEKDAY_COL_LETTERS は月..日. PFV_EDIT_WEEKDAY_KEYS は月..土 (index 0..5).
    for wd_key in PFV_EDIT_WEEKDAY_KEYS:
        wd_int = PFV_EDIT_WEEKDAY_KEY_TO_INT[wd_key]
        col = WEEKDAY_COL_LETTERS[wd_int]
        time_val = _normalize_fixed_value(_read_cell_str(ws, f"{col}18"))
        course_val = _normalize_fixed_value(_read_cell_str(ws, f"{col}19"))
        edit_values[f"{wd_key}_time"] = time_val
        edit_values[f"{wd_key}_course"] = course_val

    for col_key, idx in PFV_EDIT_COL_INDEX.items():
        v = edit_values.get(col_key)
        if v is not None:
            ws_f.cell(row=2, column=idx + 1, value=v)

    buf = BytesIO()
    out.save(buf)
    return buf.getvalue()
