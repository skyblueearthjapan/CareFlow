"""完全置換インポート (バックアップ復元) — 患者マスタ + PFV.

通常 import (``importer.py``) との違い:

  * Excel に無い既存 alive 患者は **soft delete** (関連 PFV は物理削除).
  * Excel の **空セル** は NULL で上書き (現在値維持ではなく).
  * `<DELETE>` / `<CLEAR>` フラグは使わない (Excel に無い = 削除、空セル = NULL).
  * 全 PFV を一度物理削除してから Excel 通りに再投入する (= シート完全置換).
  * **atomic transaction**: error 1 件でも全 rollback (partial commit ではない).

エントリポイント:
  * ``parse_and_diff_replace_all`` — Excel bytes → 差分行 + summary + apply 用 op list.
  * ``apply_replace_all`` — DB に書き戻す (1 transaction, all-or-nothing).
"""

from __future__ import annotations

import uuid
from datetime import datetime, time
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.course_template import CourseTemplate
from app.models.office import Office
from app.models.patient import Patient
from app.models.patient_fixed_visit import PatientFixedVisit
from app.models.staff import Staff
from app.schemas.v2.patient_excel import (
    PatientExcelChange,
    PatientExcelImportRow,
    PatientExcelReplaceAllSummary,
    PfvExcelImportRow,
)
from app.services.office_assigner import PreloadedOfficeCities, load_office_cities_index

# weekly_pattern parse ヘルパーは importer.py 側に実装済. 完全置換でも同じ
# セマンティクスを使うため import して再利用.
#   * _parse_weekly_row              : 旧 独立シート (後方互換).
#   * _parse_weekly_from_patient_cells: Phase G-48 統合シート.
#   * 各種 ANY_TO_EN map              : 日本語/英語双方向受理.
from app.services.patient_excel.importer import (
    INSURANCE_ANY_TO_EN,
    PFV_MODE_ANY_TO_EN,
    SEX_ANY_TO_EN,
    SEX_RESTRICTION_ANY_TO_EN,
    STATUS_ANY_TO_EN,
    _build_ng_staff_change,
    _load_ng_staff_by_patient,
    _load_staff_by_code,
    _merge_weekly_pattern,
    _parse_ng_staff_cell,
    _parse_weekly_from_patient_cells,
    apply_ng_staff_updates,
    cleanup_soft_deleted_patient_links,
)
from app.services.patient_excel.importer import (
    _parse_weekly_row as _parse_weekly_row_replace,  # noqa: E402
)
from app.services.patient_excel.schema import (
    DEFAULT_TIME_TYPE,
    LEGACY_SHEET_PFV,
    PATIENT_COL_INDEX,
    PFV_COL_INDEX,
    PFV_EDIT_COL_INDEX,
    PFV_EDIT_WEEKDAY_KEY_TO_INT,
    PFV_EDIT_WEEKDAY_KEYS,
    SEX_RESTRICTION_JA_VALUES,
    SEX_RESTRICTION_NONE_JA,
    SHEET_PATIENTS,
    SHEET_PFV,
    SHEET_PFV_EDIT,
    SHEET_WEEKLY,
    TIME_TYPE_VALUES,
    WEEKDAY_LABEL_TO_INT,
    build_office_code_short_maps,
    parse_course_token,
)

# ---------------------------------------------------------------------------
# cell value helpers (importer.py の同名関数を完全置換用に最小限再実装)
# 完全置換セマンティクスでは <DELETE> / <CLEAR> magic word を無視するため、
# importer の関数をそのまま流用するわけにいかない.
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _read_str(value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _read_int(value: Any) -> int | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        raise ValueError(f"int 型として読めません: {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"int 型として読めません (小数): {value!r}")
    s = str(value).strip()
    return int(s)


def _read_float(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return float(str(value).strip())


def _read_uuid(value: Any) -> UUID | None:
    if _is_blank(value):
        return None
    if isinstance(value, UUID):
        return value
    return UUID(str(value).strip())


def _read_bool(value: Any) -> bool | None:
    """TRUE/FALSE → bool. Phase E-7 (gap P0-1) requires_multiple_staff 用."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    raw = str(value).strip()
    # Phase G-48: 日本語ラベル「はい/いいえ」も受理.
    if raw == "はい":
        return True
    if raw == "いいえ":
        return False
    s = raw.upper()
    if s in ("TRUE", "1", "YES", "Y"):
        return True
    if s in ("FALSE", "0", "NO", "N"):
        return False
    raise ValueError(f"bool として読めません: {value!r}")


def _read_hhmm(value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    s = str(value).strip()
    parts = s.split(":")
    if len(parts) < 2:
        raise ValueError(f"HH:MM 形式ではありません: {s!r}")
    h = int(parts[0])
    m = int(parts[1])
    if not (0 <= h <= 23 and 0 <= m <= 59):
        raise ValueError(f"時刻の範囲外: {s!r}")
    return f"{h:02d}:{m:02d}"


def _hhmm_to_time(s: str) -> time:
    h, m = s.split(":")[:2]
    return time(int(h), int(m))


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return all(_is_blank(v) for v in row)


def _serializable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


# ---------------------------------------------------------------------------
# Patient row parser (replace-all セマンティクス)
# ---------------------------------------------------------------------------


def _parse_patient_row_replace(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    existing_patients_by_code: dict[str, Patient],
    deleted_patients_by_code: dict[str, Patient],
    offices_by_code: dict[str, Office],
    already_seen_codes: set[str],
    office_index: PreloadedOfficeCities | None = None,
    staff_by_code: dict[str, Staff] | None = None,
    ng_staff_by_patient: dict[UUID, set[UUID]] | None = None,
    staff_code_by_id: dict[UUID, str] | None = None,
) -> tuple[PatientExcelImportRow, dict[str, Any] | None]:
    """1 行をパースして差分行 + apply 用 op を返す.

    完全置換セマンティクス:
      - <DELETE> / <CLEAR> マーカーは無視 (空セル = NULL の方が強い).
      - 既存 patient_id にマッチしたら全列を Excel 値で上書き (blank → NULL).
      - patient_id 空 + patient_code が DB の alive 患者と一致 → 同じく上書き
        (UI 経由で id を消した行も identity を保つ).
      - patient_id 空 + patient_code が DB に無い (or soft-deleted) → 新規 or 復活.

    **NG スタッフ列だけは weekly_pattern と同じ「merge 例外」**:
      - 空セル = 維持 (NULL 上書きしない). 列そのものが無い旧ファイルも同じく維持.
      - 値がある場合のみその集合に一致させる. ``<CLEAR>`` で明示的に全解除.
      理由: 完全置換はバックアップ復元用途で使われるため、NG 列を持たない/埋めて
      いない古いバックアップを取り込むと NG 設定が全消しになる事故が起きる.
      NG は「このスタッフを絶対に当ててはいけない」という安全側の情報なので、
      暗黙の消失は許容できない (weekly_pattern の全消し事故防止と同じ判断).
    """
    staff_by_code = staff_by_code if staff_by_code is not None else {}
    ng_staff_by_patient = ng_staff_by_patient if ng_staff_by_patient is not None else {}
    staff_code_by_id = staff_code_by_id if staff_code_by_id is not None else {}
    cells: dict[str, Any] = {}
    for col_key, idx in PATIENT_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["patient_id"]
    raw_code = cells["patient_code"]

    # patient_id のパース
    try:
        patient_id = _read_uuid(raw_id) if not _is_blank(raw_id) else None
    except ValueError:
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=None,
                patient_code=_read_str(raw_code),
                operation="error",
                error_message=f"patient_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    patient_code = _read_str(raw_code)

    # patient_id 指定がある場合は alive 既存患者を resolve
    existing_patient: Patient | None = None
    if patient_id is not None:
        existing_patient = existing_patients.get(patient_id)
        if existing_patient is None:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=f"patient_id が DB に存在しません: {patient_id}",
                ),
                None,
            )

    # patient_id 空 + patient_code 指定 → alive 既存 patient を code で resolve
    if existing_patient is None and patient_code is not None:
        existing_patient = existing_patients_by_code.get(patient_code)
        if existing_patient is not None:
            patient_id = existing_patient.id

    # ---- 列パース (空セルは None = NULL に上書き) ----
    parsed: dict[str, Any] = {}
    errors: list[str] = []

    # 文字列系
    for k in ("name", "kana", "address", "note"):
        parsed[k] = _read_str(cells[k])

    # patient_code (新規時は必須).
    parsed["code"] = patient_code

    # enum 系: 空はそのまま None、値があれば日本語/英語双方を受理し英語値へ正規化.
    def _parse_enum(key: str, mapping: dict[str, str]) -> None:
        v = cells[key]
        if _is_blank(v):
            parsed[key] = None
            return
        s = _read_str(v)
        en = mapping.get(s or "")
        if en is None:
            allowed = ",".join(dict.fromkeys(mapping.keys()))
            errors.append(f"列「{key}」の値が候補外: {s!r} (許容: {allowed})")
            return
        parsed[key] = en

    _parse_enum("sex", SEX_ANY_TO_EN)
    _parse_enum("status", STATUS_ANY_TO_EN)
    _parse_enum("insurance", INSURANCE_ANY_TO_EN)

    # sex_restriction: 「なし」/ 空セルは DB NULL (制限解除). それ以外は正規化.
    raw_sr = cells["sex_restriction"]
    if _is_blank(raw_sr) or _read_str(raw_sr) == SEX_RESTRICTION_NONE_JA:
        parsed["sex_restriction"] = None
    else:
        sr_s = _read_str(raw_sr)
        sr_en = SEX_RESTRICTION_ANY_TO_EN.get(sr_s or "")
        if sr_en is None:
            allowed = ",".join(SEX_RESTRICTION_JA_VALUES)
            errors.append(f"列「sex_restriction」の値が候補外: {sr_s!r} (許容: {allowed})")
        else:
            parsed["sex_restriction"] = sr_en

    # lat / lng
    for k, lo, hi in (("lat", -90.0, 90.0), ("lng", -180.0, 180.0)):
        v = cells[k]
        if _is_blank(v):
            parsed[k] = None
            continue
        try:
            f = _read_float(v)
        except (ValueError, TypeError):
            errors.append(f"列「{k}」が数値ではありません: {v!r}")
            continue
        if f is None or f < lo or f > hi:
            errors.append(f"列「{k}」が範囲外 [{lo}, {hi}]: {f!r}")
            continue
        parsed[k] = f

    # 拠点コード → primary_office_id
    # Phase G-49: コードが入っていれば尊重 (手動上書き). 空欄なら新規/復活時のみ
    # 住所から自動割当を試みる (UI 挙動と一致). 既存患者の空欄は従来どおり NULL 上書き
    # (完全置換セマンティクス). これにより export (コードあり) → import = noop が保たれ、
    # 既存患者の round-trip 安定性を壊さない.
    raw_office = cells["office_code"]
    auto_office_id: UUID | None = None  # 空欄時に住所から解決 (新規/復活でのみ採用).
    if _is_blank(raw_office):
        parsed["primary_office_id"] = None
        if office_index is not None:
            auto_office_id = office_index.resolve_office_id(_read_str(cells.get("address")))
    else:
        oc = _read_str(raw_office)
        office = offices_by_code.get(oc) if oc else None
        if office is None:
            errors.append(f"拠点コードが DB に存在しません: {oc!r}")
        else:
            parsed["primary_office_id"] = office.id

    # Phase E-7 (gap P0-1): requires_multiple_staff. NOT NULL bool 列.
    # 完全置換セマンティクスでは空セル = False で確定上書き.
    raw_req_multi = cells["requires_multiple_staff"]
    if _is_blank(raw_req_multi):
        parsed["requires_multiple_staff"] = False
    else:
        try:
            parsed["requires_multiple_staff"] = _read_bool(raw_req_multi)
        except ValueError as exc:
            errors.append(f"列「requires_multiple_staff」が TRUE/FALSE ではありません: {exc}")

    # NG スタッフ (merge 例外列). None = 空セル / 列なし = 維持.
    # relationship なので ``parsed`` には入れず別変数で持つ.
    ng_resolved = _parse_ng_staff_cell(
        cells.get("ng_staff_codes"),
        staff_by_code=staff_by_code,
        row_number=row_number,
    )

    if errors:
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient_code,
                operation="error",
                error_message=" / ".join(errors),
            ),
            None,
        )

    # ---- code 重複チェック (同ファイル内) ----
    if patient_code is not None:
        if patient_code in already_seen_codes:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=(
                        f"patient_code が同ファイル内で重複しています: {patient_code!r}"
                    ),
                ),
                None,
            )
        already_seen_codes.add(patient_code)

    # ---- 操作判定 ----
    field_map: dict[str, str] = {
        "name": "name",
        "kana": "kana",
        "sex": "sex",
        "status": "status",
        "insurance": "insurance",
        "address": "address",
        "lat": "lat",
        "lng": "lng",
        "primary_office_id": "primary_office_id",
        "sex_restriction": "sex_restriction",
        # Phase E-7 (gap P0-1): NOT NULL bool 列. 完全置換では空セル=False.
        "requires_multiple_staff": "requires_multiple_staff",
        "note": "note",
    }

    if existing_patient is not None:
        # update: 全列を Excel 値で上書き (blank → NULL).
        # patient_code は通常変更しない方針だが、明示的に書いてあって既存値と違えば update.

        # NOT NULL カラムが空セル (None) になっていないか確認
        required_update = ["name", "status"]
        null_required = [k for k in required_update if parsed.get(k) is None]
        if null_required:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=existing_patient.id,
                    patient_code=existing_patient.code,
                    operation="error",
                    error_message=f"必須項目が空セルです (NULL 不可): {', '.join(null_required)}",
                ),
                None,
            )

        changes: list[PatientExcelChange] = []
        update_dict: dict[str, Any] = {"_patient_id": existing_patient.id, "_op": "update"}

        if patient_code is not None and patient_code != existing_patient.code:
            changes.append(
                PatientExcelChange(
                    field="code",
                    old_value=existing_patient.code,
                    new_value=patient_code,
                )
            )
            update_dict["code"] = patient_code

        for parsed_key, orm_attr in field_map.items():
            new_val = parsed.get(parsed_key)
            old_val = getattr(existing_patient, orm_attr, None)
            # Numeric/float 比較対策
            if old_val is not None and isinstance(new_val, float):
                try:
                    if float(old_val) == float(new_val):
                        continue
                except (TypeError, ValueError):
                    pass
            if old_val == new_val:
                continue
            changes.append(
                PatientExcelChange(
                    field=orm_attr,
                    old_value=_serializable(old_val),
                    new_value=_serializable(new_val),
                )
            )
            update_dict[orm_attr] = new_val

        # NG スタッフ (merge 例外): 値がある場合のみ集合一致.
        ng_change = _build_ng_staff_change(
            ng_resolved,
            ng_staff_by_patient.get(existing_patient.id, set()),
            row_number=row_number,
            staff_code_by_id=staff_code_by_id,
        )
        if ng_change is not None:
            change, target_ids = ng_change
            changes.append(change)
            update_dict["_ng_staff_ids"] = target_ids

        if not changes:
            return (
                PatientExcelImportRow(
                    row_number=row_number,
                    patient_id=existing_patient.id,
                    patient_code=existing_patient.code,
                    operation="noop",
                ),
                None,
            )

        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=existing_patient.id,
                patient_code=existing_patient.code,
                operation="update",
                changes=changes,
            ),
            update_dict,
        )

    # ---- 新規 or 復活 ----
    # 必須チェック
    missing: list[str] = []
    if not patient_code:
        missing.append("patient_code")
    for spec_key in ("name", "sex", "status", "address"):
        if parsed.get(spec_key) is None:
            missing.append(spec_key)
    if missing:
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=None,
                patient_code=patient_code,
                operation="error",
                error_message=f"新規作成に必要な項目が不足: {', '.join(missing)}",
            ),
            None,
        )

    # Phase G-49: ここに到達するのは新規/復活のみ (既存 update は上で return 済).
    # 拠点コード空欄なら住所から自動割当した値を採用する (UI 挙動と一致).
    if parsed.get("primary_office_id") is None and auto_office_id is not None:
        parsed["primary_office_id"] = auto_office_id

    # 復活 (soft-deleted 同じ code が DB に居る)
    resurrect_target = deleted_patients_by_code.get(patient_code) if patient_code else None
    if resurrect_target is not None:
        resurrect_id = resurrect_target.id
        old_deleted_at = resurrect_target.deleted_at
        updates: dict[str, Any] = {}
        changes_for_view: list[PatientExcelChange] = [
            PatientExcelChange(
                field="deleted_at",
                old_value=_serializable(old_deleted_at),
                new_value=None,
            )
        ]
        for parsed_key, orm_attr in field_map.items():
            new_val = parsed.get(parsed_key)
            old_val = getattr(resurrect_target, orm_attr, None)
            if old_val is not None and isinstance(new_val, float):
                try:
                    if float(old_val) == float(new_val):
                        continue
                except (TypeError, ValueError):
                    pass
            if old_val == new_val:
                continue
            changes_for_view.append(
                PatientExcelChange(
                    field=orm_attr,
                    old_value=_serializable(old_val),
                    new_value=_serializable(new_val),
                )
            )
            updates[orm_attr] = new_val
        ng_op_dict: dict[str, Any] = {}
        ng_change = _build_ng_staff_change(
            ng_resolved,
            ng_staff_by_patient.get(resurrect_id, set()),
            row_number=row_number,
            staff_code_by_id=staff_code_by_id,
        )
        if ng_change is not None:
            change, target_ids = ng_change
            changes_for_view.append(change)
            ng_op_dict["_ng_staff_ids"] = target_ids
        return (
            PatientExcelImportRow(
                row_number=row_number,
                patient_id=resurrect_id,
                patient_code=patient_code,
                operation="update",
                changes=changes_for_view,
            ),
            {
                "_op": "resurrect",
                "_patient_id": resurrect_id,
                "_updates": updates,
                **ng_op_dict,
            },
        )

    # 純粋新規. importer と同じく仮 UUID 発番 + special_week_active=[] を補完.
    new_patient_id = uuid.uuid4()
    new_dict: dict[str, Any] = {
        "_op": "new",
        "_new_patient_id": new_patient_id,
        "id": new_patient_id,
        "code": patient_code,
        "special_week_active": [],
    }
    for parsed_key, orm_attr in field_map.items():
        val = parsed.get(parsed_key)
        if val is None:
            continue  # default に任せる
        new_dict[orm_attr] = val

    changes_for_view = [
        PatientExcelChange(field=k, old_value=None, new_value=_serializable(new_dict[k]))
        for k in sorted(new_dict.keys())
        if not k.startswith("_") and k not in ("id",)
    ]
    # NG スタッフ: Patient(**data) には渡せないので ``_ng_staff_ids`` で外出し.
    ng_change = _build_ng_staff_change(
        ng_resolved,
        set(),
        row_number=row_number,
        staff_code_by_id=staff_code_by_id,
    )
    if ng_change is not None:
        change, target_ids = ng_change
        changes_for_view.append(change)
        new_dict["_ng_staff_ids"] = target_ids
    return (
        PatientExcelImportRow(
            row_number=row_number,
            patient_id=new_patient_id,
            patient_code=patient_code,
            operation="new",
            changes=changes_for_view,
        ),
        new_dict,
    )


# ---------------------------------------------------------------------------
# PFV row parser (replace-all セマンティクス: 全件物理削除 → 再投入)
# ---------------------------------------------------------------------------


class _PendingNewPatient:
    """importer._PendingNewPatient と同等. PFV パース時の patient 参照に使う."""

    __slots__ = ("id", "code", "primary_office_id")

    def __init__(
        self, *, patient_id: UUID, code: str | None, primary_office_id: UUID | None
    ) -> None:
        self.id = patient_id
        self.code = code
        self.primary_office_id = primary_office_id


def _parse_pfv_row_replace(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    pending_new_patients: dict[UUID, _PendingNewPatient],
    patient_code_to_id: dict[str, UUID],
    course_templates: dict[tuple[UUID, str], CourseTemplate],
    pending_new_keys: set[tuple[UUID, str, int, int]],
    offices_by_code: dict[str, Office] | None = None,
) -> tuple[PfvExcelImportRow, dict[str, Any] | None]:
    """1 PFV 行をパース. 完全置換では update / delete は無く、new のみ.

    既存 PFV は全件物理削除されてから Excel の各行を INSERT するため、
    op_type は "new" のみ (or "error").
    """
    cells: dict[str, Any] = {}
    for col_key, idx in PFV_COL_INDEX.items():
        cells[col_key] = row[idx] if idx < len(row) else None

    raw_id = cells["patient_id"]
    raw_code = cells["patient_code"]

    try:
        patient_id = _read_uuid(raw_id)
    except ValueError:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=None,
                operation="error",
                error_message=f"patient_id が UUID 形式ではありません: {raw_id!r}",
            ),
            None,
        )

    patient_code = _read_str(raw_code)

    # patient_id 空 → patient_code で解決
    if patient_id is None:
        if patient_code is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=None,
                    operation="error",
                    error_message="patient_id または patient_code が必須です",
                ),
                None,
            )
        resolved_id = patient_code_to_id.get(patient_code)
        if resolved_id is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=patient_code,
                    operation="error",
                    error_message=(
                        f"patient_code が DB にも import 内にも存在しません: {patient_code!r}"
                    ),
                ),
                None,
            )
        patient_id = resolved_id
    elif patient_code is not None:
        expected_id_from_code = patient_code_to_id.get(patient_code)
        if expected_id_from_code is not None and expected_id_from_code != patient_id:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=(
                        f"patient_id と patient_code が異なる患者を指しています "
                        f"(id={patient_id}, code={patient_code} → 期待 id={expected_id_from_code})"
                    ),
                ),
                None,
            )

    patient = existing_patients.get(patient_id)
    if patient is None:
        pending = pending_new_patients.get(patient_id)
        if pending is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient_code,
                    operation="error",
                    error_message=f"patient_id が DB に存在しません: {patient_id}",
                ),
                None,
            )
        patient = pending  # type: ignore[assignment]

    # weekday
    raw_weekday = cells["weekday"]
    if _is_blank(raw_weekday):
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                operation="error",
                error_message="weekday が空です",
            ),
            None,
        )
    wd_str = _read_str(raw_weekday)
    weekday = WEEKDAY_LABEL_TO_INT.get(wd_str or "")
    if weekday is None:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                operation="error",
                error_message=f"weekday の値が候補外: {wd_str!r} (許容: 月火水木金土日)",
            ),
            None,
        )

    raw_slot = cells["slot_index"]
    try:
        slot_index = _read_int(raw_slot)
    except (ValueError, TypeError):
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                operation="error",
                error_message=f"slot_index が整数ではありません: {raw_slot!r}",
            ),
            None,
        )
    if slot_index is None:
        slot_index = 0
    if slot_index < 0 or slot_index > 1:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                operation="error",
                error_message=f"slot_index は 0 または 1 のみ: {slot_index}",
            ),
            None,
        )

    raw_mode = cells["mode"]
    if _is_blank(raw_mode):
        mode = "normal"
    else:
        # Phase G-48: モードは日本語 (通常/特別) / 英語 (normal/special) 双方受理.
        mode_str = _read_str(raw_mode)
        mode_en = PFV_MODE_ANY_TO_EN.get(mode_str or "")
        if mode_en is None:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"mode の値が候補外: {mode_str!r}",
                ),
                None,
            )
        mode = mode_en

    key = (patient_id, mode, weekday, slot_index)

    # time_type
    # E-4: 空セルは default で吸収 (round-trip 運用 0 エラー). 詳細は importer.py の
    # 同 block コメント参照.
    raw_tt = cells["time_type"]
    if _is_blank(raw_tt):
        time_type: str | None = DEFAULT_TIME_TYPE
    else:
        time_type = _read_str(raw_tt)
        if time_type not in TIME_TYPE_VALUES:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"time_type の値が候補外: {time_type!r}",
                ),
                None,
            )

    # start_time
    try:
        start_str = _read_hhmm(cells["start_time"])
    except (ValueError, TypeError) as exc:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message=f"start_time が HH:MM 形式ではありません: {exc}",
            ),
            None,
        )
    if start_str is None:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message="start_time が空です",
            ),
            None,
        )

    # end_time / duration_min
    raw_end = cells["end_time"]
    raw_duration = cells["duration_min"]
    duration_min: int | None = None
    if not _is_blank(raw_duration):
        try:
            duration_min = _read_int(raw_duration)
        except (ValueError, TypeError) as exc:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"duration_min が整数ではありません: {exc}",
                ),
                None,
            )
    if duration_min is None and not _is_blank(raw_end):
        try:
            end_str = _read_hhmm(raw_end)
        except (ValueError, TypeError) as exc:
            return (
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    weekday=weekday,
                    slot_index=slot_index,
                    operation="error",
                    error_message=f"end_time が HH:MM 形式ではありません: {exc}",
                ),
                None,
            )
        if end_str is not None:
            sh, sm = (int(x) for x in start_str.split(":"))
            eh, em = (int(x) for x in end_str.split(":"))
            duration_min = (eh * 60 + em) - (sh * 60 + sm)
    if duration_min is None:
        duration_min = 30
    if duration_min < 1 or duration_min > 480:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message=f"duration_min が範囲外 [1, 480]: {duration_min}",
            ),
            None,
        )

    # Phase E-5 (項目 ⑥B): sub_office_code → sub_office_id を先に解決.
    raw_sub_office = cells.get("sub_office_code")
    sub_office_id: UUID | None = None
    if not _is_blank(raw_sub_office) and offices_by_code is not None:
        oc = _read_str(raw_sub_office)
        if oc:
            office = offices_by_code.get(oc)
            if office is not None:
                sub_office_id = office.id

    # course_template_code
    # E-4: 「拠点に存在しない code」は error にせず course_template_id=None として
    # 取り込み、PFV 自体は保持する (round-trip / バックアップ運用優先).
    # importer.py の同 block コメント参照.
    # Phase E-5: sub_office_id が指定された場合は sub_office を優先して解決.
    raw_ct = cells["course_template_code"]
    course_template_id: UUID | None = None
    if not _is_blank(raw_ct):
        ct_label = _read_str(raw_ct)
        resolved = False
        if sub_office_id is not None:
            ct = course_templates.get((sub_office_id, ct_label or ""))
            if ct is not None:
                course_template_id = ct.id
                resolved = True
        if not resolved and patient.primary_office_id is not None:
            ct = course_templates.get((patient.primary_office_id, ct_label or ""))
            if ct is not None:
                course_template_id = ct.id

    if key in pending_new_keys:
        return (
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=weekday,
                slot_index=slot_index,
                operation="error",
                error_message=(
                    "同ファイル内で (patient_id, mode, weekday, slot_index) が重複しています"
                ),
            ),
            None,
        )
    pending_new_keys.add(key)

    new_start = _hhmm_to_time(start_str)
    new_dict: dict[str, Any] = {
        "_op": "new",
        "patient_id": patient_id,
        "mode": mode,
        "weekday": weekday,
        "slot_index": slot_index,
        "start_time": new_start,
        "duration_min": duration_min,
        "course_template_id": course_template_id,
        # Phase E-5 (項目 ⑥B): サブ拠点 ID.
        "sub_office_id": sub_office_id,
    }
    return (
        PfvExcelImportRow(
            row_number=row_number,
            patient_id=patient_id,
            patient_code=patient.code,
            weekday=weekday,
            slot_index=slot_index,
            operation="new",
            changes=[
                PatientExcelChange(field="start_time", old_value=None, new_value=start_str),
                PatientExcelChange(field="duration_min", old_value=None, new_value=duration_min),
            ],
        ),
        new_dict,
    )


def _resolve_pfv_course_replace(
    token: str | None,
    *,
    offices_by_code: dict[str, Office],
    course_templates: dict[tuple[UUID, str], CourseTemplate],
) -> UUID | None:
    """拠点付きコーストークン → course_template_id (replace 用).

    importer._resolve_pfv_course と同じセマンティクス (拠点付きトークンで
    クロス拠点もそのまま course_template_id で表現. sub_office_id は扱わない).
    """
    if token is None:
        return None
    # 0059: 拠点マスタ (offices.short_label) 駆動で短縮名を解決する.
    _, short_to_code = build_office_code_short_maps(
        (o.code, o.short_label) for o in offices_by_code.values()
    )
    parsed = parse_course_token(token, short_to_code)
    if parsed is None:
        return None
    office_code, label = parsed
    office = offices_by_code.get(office_code)
    if office is None:
        return None
    ct = course_templates.get((office.id, label))
    if ct is None:
        return None
    return ct.id


def _parse_pfv_edit_patient_replace(
    row_number: int,
    row: tuple[Any, ...],
    *,
    existing_patients: dict[UUID, Patient],
    pending_new_patients: dict[UUID, _PendingNewPatient],
    patient_code_to_id: dict[str, UUID],
    course_templates: dict[tuple[UUID, str], CourseTemplate],
    offices_by_code: dict[str, Office],
) -> tuple[list[PfvExcelImportRow], list[dict[str, Any]]]:
    """Phase G-51: 「固定訪問パターン（編集用）」1 行 → new PFV op の list (replace-all).

    完全置換では既存 PFV は全件物理削除済なので、行が示す各曜日 (時刻入り) について
    normal mode / slot_index=0 の new op を生成するだけ (update/delete は無い).
    """
    cells: dict[str, Any] = {
        col_key: (row[idx] if idx < len(row) else None)
        for col_key, idx in PFV_EDIT_COL_INDEX.items()
    }
    raw_id = cells["patient_id"]
    raw_code = cells["patient_code"]

    try:
        patient_id = _read_uuid(raw_id)
    except ValueError:
        return (
            [
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=None,
                    patient_code=_read_str(raw_code),
                    operation="error",
                    error_message=f"patient_id が UUID 形式ではありません: {raw_id!r}",
                )
            ],
            [],
        )

    patient_code = _read_str(raw_code)
    if patient_id is None:
        if patient_code is None:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=None,
                        patient_code=None,
                        operation="error",
                        error_message="patient_id または patient_code が必須です",
                    )
                ],
                [],
            )
        resolved_id = patient_code_to_id.get(patient_code)
        if resolved_id is None:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=None,
                        patient_code=patient_code,
                        operation="error",
                        error_message=(
                            f"patient_code が DB にも import 内にも存在しません: {patient_code!r}"
                        ),
                    )
                ],
                [],
            )
        patient_id = resolved_id
    elif patient_code is not None:
        expected = patient_code_to_id.get(patient_code)
        if expected is not None and expected != patient_id:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=patient_id,
                        patient_code=patient_code,
                        operation="error",
                        error_message=(
                            f"patient_id と patient_code が異なる患者を指しています "
                            f"(id={patient_id}, code={patient_code} → 期待 id={expected})"
                        ),
                    )
                ],
                [],
            )

    patient: Patient | _PendingNewPatient | None = existing_patients.get(patient_id)
    if patient is None:
        patient = pending_new_patients.get(patient_id)
        if patient is None:
            return (
                [
                    PfvExcelImportRow(
                        row_number=row_number,
                        patient_id=patient_id,
                        patient_code=patient_code,
                        operation="error",
                        error_message=f"patient_id が DB に存在しません: {patient_id}",
                    )
                ],
                [],
            )

    raw_sm = cells["service_minutes"]
    try:
        service_minutes = _read_int(raw_sm)
    except (ValueError, TypeError):
        service_minutes = None
    if service_minutes is not None and (service_minutes < 1 or service_minutes > 480):
        service_minutes = None

    diff_rows: list[PfvExcelImportRow] = []
    ops: list[dict[str, Any]] = []
    errors: list[str] = []
    for wd_key in PFV_EDIT_WEEKDAY_KEYS:
        wd_int = PFV_EDIT_WEEKDAY_KEY_TO_INT[wd_key]
        raw_time = cells.get(f"{wd_key}_time")
        if _is_blank(raw_time):
            continue
        try:
            start_str = _read_hhmm(raw_time)
        except (ValueError, TypeError) as exc:
            errors.append(f"{wd_key} の時刻が HH:MM 形式ではありません: {exc}")
            continue
        if start_str is None:
            continue
        ct_id = _resolve_pfv_course_replace(
            _read_str(cells.get(f"{wd_key}_course")),
            offices_by_code=offices_by_code,
            course_templates=course_templates,
        )
        duration = service_minutes if service_minutes is not None else 30
        ops.append(
            {
                "_op": "new",
                "patient_id": patient_id,
                "mode": "normal",
                "weekday": wd_int,
                "slot_index": 0,
                "start_time": _hhmm_to_time(start_str),
                "duration_min": duration,
                "course_template_id": ct_id,
                # sub_office_id は本シート対象外 (model default None).
            }
        )
        diff_rows.append(
            PfvExcelImportRow(
                row_number=row_number,
                patient_id=patient_id,
                patient_code=patient.code,
                weekday=wd_int,
                slot_index=0,
                operation="new",
                changes=[
                    PatientExcelChange(field="start_time", old_value=None, new_value=start_str),
                    PatientExcelChange(field="duration_min", old_value=None, new_value=duration),
                ],
            )
        )
    if errors:
        return (
            [
                PfvExcelImportRow(
                    row_number=row_number,
                    patient_id=patient_id,
                    patient_code=patient.code,
                    operation="error",
                    error_message=" / ".join(errors),
                )
            ],
            [],
        )
    return diff_rows, ops


# ---------------------------------------------------------------------------
# public entrypoints
# ---------------------------------------------------------------------------


async def parse_and_diff_replace_all(
    db: AsyncSession,
    *,
    file_bytes: bytes,
) -> tuple[
    list[PatientExcelImportRow],
    list[PfvExcelImportRow],
    PatientExcelReplaceAllSummary,
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],  # patients_to_soft_delete (UI preview)
]:
    """Excel をパースして「完全置換」用の差分行 + summary + apply op list を返す.

    戻り値の最後の要素 ``patients_to_soft_delete_preview`` は
    Excel に無い alive 既存患者の {patient_id, patient_code, name} 一覧.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    if SHEET_PATIENTS not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_PATIENTS}」が見つかりません")
    # Phase G-51: 新「固定訪問パターン（編集用）」(患者 1 行) を最優先. 無ければ旧
    # per-visit シート (固定訪問パターン / 固定訪問スケジュール) に fallback.
    # 静的「グリッド集計（静的・閲覧専用）」シートは読まない (無視).
    use_pfv_edit_sheet = SHEET_PFV_EDIT in wb.sheetnames
    pfv_sheet_name = (
        SHEET_PFV
        if SHEET_PFV in wb.sheetnames
        else LEGACY_SHEET_PFV
        if LEGACY_SHEET_PFV in wb.sheetnames
        else None
    )
    if not use_pfv_edit_sheet and pfv_sheet_name is None:
        raise ValueError(f"シート「{SHEET_PFV_EDIT}」が見つかりません")

    # ---- DB lookup ----
    rows = (await db.scalars(select(Patient).where(Patient.deleted_at.is_(None)))).all()
    existing_patients: dict[UUID, Patient] = {p.id: p for p in rows}
    existing_patients_by_code: dict[str, Patient] = {p.code: p for p in rows if p.code}

    deleted_rows = (
        await db.scalars(
            select(Patient).where(Patient.deleted_at.is_not(None), Patient.code.is_not(None))
        )
    ).all()
    deleted_patients_by_code: dict[str, Patient] = {p.code: p for p in deleted_rows}

    offices_rows = (
        await db.scalars(
            select(Office).where(Office.deleted_at.is_(None), Office.code.is_not(None))
        )
    ).all()
    offices_by_code: dict[str, Office] = {str(o.code): o for o in offices_rows}

    ct_rows = (
        await db.scalars(select(CourseTemplate).where(CourseTemplate.deleted_at.is_(None)))
    ).all()
    course_templates: dict[tuple[UUID, str], CourseTemplate] = {
        (ct.office_id, ct.label): ct for ct in ct_rows
    }

    # Phase G-49: 拠点コード空欄の住所→拠点 自動割当用. cities/office_cities を
    # 一度だけロードし、行ごとは同期照合 (N+1 回避).
    office_index = await load_office_cities_index(db)
    # NG スタッフ列用 (merge 例外): staff_code → Staff (退職者含む) / 既存 NG 集合.
    staff_by_code = await _load_staff_by_code(db)
    staff_code_by_id: dict[UUID, str] = {s.id: str(s.code) for s in staff_by_code.values()}
    ng_staff_by_patient = await _load_ng_staff_by_patient(db)

    # ---- patient sheet ----
    ws_p = wb[SHEET_PATIENTS]
    patient_rows: list[PatientExcelImportRow] = []
    patient_ops: list[dict[str, Any]] = []
    already_seen_codes: set[str] = set()
    # Excel に出てきた既存 patient_id の集合 (Excel に居る = 削除しない).
    excel_patient_ids_seen: set[UUID] = set()
    # Phase G-48: 統合シートから読み取った weekly_pattern (patient_id → dict|None).
    weekly_from_patient_sheet: dict[UUID, dict[str, Any] | None] = {}

    for r_idx, row in enumerate(ws_p.iter_rows(min_row=2, values_only=True), start=2):
        if _row_is_empty(row):
            continue
        diff_row, op = _parse_patient_row_replace(
            r_idx,
            row,
            existing_patients=existing_patients,
            existing_patients_by_code=existing_patients_by_code,
            deleted_patients_by_code=deleted_patients_by_code,
            offices_by_code=offices_by_code,
            already_seen_codes=already_seen_codes,
            office_index=office_index,
            staff_by_code=staff_by_code,
            ng_staff_by_patient=ng_staff_by_patient,
            staff_code_by_id=staff_code_by_id,
        )
        patient_rows.append(diff_row)
        if op is not None:
            patient_ops.append(op)
        # alive 既存患者が Excel に居る → 削除リストから除外
        if diff_row.operation in ("update", "noop") and diff_row.patient_id is not None:
            excel_patient_ids_seen.add(diff_row.patient_id)
        # Phase G-48: 統合シートの weekly_pattern を抽出 (new/update/noop 行のみ).
        if diff_row.operation in ("new", "update", "noop") and diff_row.patient_id is not None:
            row_cells: dict[str, Any] = {
                col_key: (row[idx] if idx < len(row) else None)
                for col_key, idx in PATIENT_COL_INDEX.items()
            }
            weekly = _parse_weekly_from_patient_cells(row_cells)
            if weekly is not None:
                weekly_from_patient_sheet[diff_row.patient_id] = weekly

    # ---- Excel に無い alive 既存患者 = soft delete 対象 ----
    patients_to_soft_delete: list[Patient] = [
        p for pid, p in existing_patients.items() if pid not in excel_patient_ids_seen
    ]
    patients_to_soft_delete_preview: list[dict[str, Any]] = [
        {"patient_id": str(p.id), "patient_code": p.code, "name": p.name}
        for p in patients_to_soft_delete
    ]

    # ---- PFV ref maps ----
    pending_new_patients: dict[UUID, _PendingNewPatient] = {}
    patient_code_to_id: dict[str, UUID] = {}
    for existing in existing_patients.values():
        if existing.code and existing.id in excel_patient_ids_seen:
            patient_code_to_id[existing.code] = existing.id
    for op in patient_ops:
        if op.get("_op") != "new":
            continue
        new_id: UUID | None = op.get("_new_patient_id")
        new_code: str | None = op.get("code")
        if new_id is None or new_code is None:
            continue
        patient_code_to_id[new_code] = new_id
        pending_new_patients[new_id] = _PendingNewPatient(
            patient_id=new_id,
            code=new_code,
            primary_office_id=op.get("primary_office_id"),
        )
    # resurrection 対象の patient も PFV から参照できるよう登録
    deleted_by_id = {p.id: p for p in deleted_patients_by_code.values()}
    for op in patient_ops:
        if op.get("_op") != "resurrect":
            continue
        resurrect_pid: UUID = op["_patient_id"]
        original = deleted_by_id.get(resurrect_pid)
        if original is None or not original.code:
            continue
        updated_office = op["_updates"].get("primary_office_id", original.primary_office_id)
        patient_code_to_id[original.code] = resurrect_pid
        pending_new_patients[resurrect_pid] = _PendingNewPatient(
            patient_id=resurrect_pid,
            code=original.code,
            primary_office_id=updated_office,
        )

    # ---- PFV sheet ----
    pfv_rows: list[PfvExcelImportRow] = []
    pfv_ops: list[dict[str, Any]] = []
    if use_pfv_edit_sheet:
        # Phase G-51: 新「編集用」(患者 1 行) → 各曜日 new op (replace は全件再投入).
        ws_f = wb[SHEET_PFV_EDIT]
        for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            diff_rows_edit, ops_edit = _parse_pfv_edit_patient_replace(
                r_idx,
                row,
                existing_patients=existing_patients,
                pending_new_patients=pending_new_patients,
                patient_code_to_id=patient_code_to_id,
                course_templates=course_templates,
                offices_by_code=offices_by_code,
            )
            pfv_rows.extend(diff_rows_edit)
            pfv_ops.extend(ops_edit)
    else:
        assert pfv_sheet_name is not None
        ws_f = wb[pfv_sheet_name]
        pending_new_keys: set[tuple[UUID, str, int, int]] = set()
        for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            diff_row, op = _parse_pfv_row_replace(
                r_idx,
                row,
                existing_patients=existing_patients,
                pending_new_patients=pending_new_patients,
                patient_code_to_id=patient_code_to_id,
                course_templates=course_templates,
                pending_new_keys=pending_new_keys,
                # Phase E-5: sub_office_code 解決用.
                offices_by_code=offices_by_code,
            )
            pfv_rows.append(diff_row)
            if op is not None:
                pfv_ops.append(op)

    # ---- weekly_pattern マージ ----
    # Phase G-48 hotfix: 丸ごと置換ではなく merge. 該当行がある patient の
    # weekly_pattern を、管理外キー (entries/staff_count 等) を保持したまま管理 8 キー
    # のみ上書きする. 行が無い patient の weekly_pattern は維持 (既存値を破壊しない).
    # バックアップ復元経路 (replace_all) でも entries 等の喪失を防ぐ.
    def _merge_weekly_updates(weekly_pattern_updates: dict[UUID, dict[str, Any] | None]) -> None:
        ops_by_pid: dict[UUID, dict[str, Any]] = {}
        for op in patient_ops:
            op_pid = op.get("_patient_id") or op.get("_new_patient_id")
            if op_pid is not None:
                ops_by_pid[op_pid] = op

        for pid, wp in weekly_pattern_updates.items():
            existing_patient = existing_patients.get(pid)
            existing_wp = existing_patient.weekly_pattern if existing_patient else None
            existing_op = ops_by_pid.get(pid)
            # 同 import 内で先に統合済の値があればそれを merge 基準にする.
            if existing_op is not None and "weekly_pattern" in existing_op:
                current_wp = existing_op["weekly_pattern"]
            elif existing_op is not None and existing_op.get("_op") == "resurrect":
                current_wp = existing_op.get("_updates", {}).get("weekly_pattern", existing_wp)
            else:
                current_wp = existing_wp
            # merge (管理外キー保持) + noop 判定を merge 後の dict で行う.
            merged_wp = _merge_weekly_pattern(current_wp, wp)
            if merged_wp == current_wp:
                continue

            if existing_op is None:
                # noop だった patient → 新規 update op を追加
                new_op = {
                    "_op": "update",
                    "_patient_id": pid,
                    "weekly_pattern": merged_wp,
                }
                patient_ops.append(new_op)
                ops_by_pid[pid] = new_op
                # patient_rows の noop → update に上書き
                for pr in patient_rows:
                    if pr.patient_id == pid and pr.operation == "noop":
                        pr.operation = "update"
                        pr.changes.append(
                            PatientExcelChange(
                                field="weekly_pattern",
                                old_value=_serializable(existing_wp),
                                new_value="(希望訪問パターン更新)",
                            )
                        )
                        break
            else:
                op_type = existing_op.get("_op")
                if op_type == "resurrect":
                    existing_op.setdefault("_updates", {})["weekly_pattern"] = merged_wp
                elif op_type in ("new", "update"):
                    existing_op["weekly_pattern"] = merged_wp

    # Phase G-48: 統合「患者マスタ」シートから読み取った weekly_pattern をマージ.
    _merge_weekly_updates(weekly_from_patient_sheet)

    # ---- 【後方互換】旧 独立「希望訪問パターン」シート ----
    # SHEET_WEEKLY が存在する場合のみ読み込み (旧 3 シート構成ファイル受理用).
    # 統合シートより後にマージ → 旧シートの明示値が優先.
    if SHEET_WEEKLY in wb.sheetnames:
        ws_w = wb[SHEET_WEEKLY]
        legacy_weekly_updates: dict[UUID, dict[str, Any] | None] = {}
        for r_idx, row in enumerate(ws_w.iter_rows(min_row=2, values_only=True), start=2):
            if _row_is_empty(row):
                continue
            result = _parse_weekly_row_replace(
                r_idx,
                row,
                existing_patients=existing_patients,
                existing_patients_by_code=existing_patients_by_code,
                patient_code_to_id=patient_code_to_id,
                pending_new_patients=pending_new_patients,
            )
            if result is None:
                continue
            pid, wp = result
            legacy_weekly_updates[pid] = wp if wp else None
        _merge_weekly_updates(legacy_weekly_updates)

    # ---- 既存 PFV 件数 (全件再投入なので "replace count" として表示) ----
    # alive patients の PFV のみ count (soft-deleted patients の PFV は除外)
    pfv_count_row = (
        await db.execute(
            select(func.count(PatientFixedVisit.id)).where(
                PatientFixedVisit.patient_id.in_(
                    select(Patient.id).where(Patient.deleted_at.is_(None))
                )
            )
        )
    ).scalar_one()
    pfv_to_replace = int(pfv_count_row or 0)

    # ---- summary ----
    summary = PatientExcelReplaceAllSummary()
    for r in patient_rows:
        match r.operation:
            case "new":
                summary.patients_to_create += 1
            case "update":
                summary.patients_to_update += 1
            case "error":
                summary.patients_error += 1
    summary.patients_to_soft_delete = len(patients_to_soft_delete)
    for r in pfv_rows:
        match r.operation:
            case "new":
                summary.pfv_to_create += 1
            case "error":
                summary.pfv_error += 1
    summary.pfv_to_replace = pfv_to_replace

    return (
        patient_rows,
        pfv_rows,
        summary,
        patient_ops,
        pfv_ops,
        patients_to_soft_delete_preview,
    )


async def apply_replace_all(
    db: AsyncSession,
    *,
    patient_ops: list[dict[str, Any]],
    pfv_ops: list[dict[str, Any]],
    patient_ids_to_soft_delete: list[UUID],
) -> None:
    """完全置換を DB に反映する (1 transaction, atomic).

    順序:
      1. 削除対象 patient の関連 PFV を全件物理削除 (orphan 防止 / FK 整合)
      2. 削除対象 patient を soft delete (deleted_at=now) +
         中間テーブル (同住所リンク / NG スタッフ) を物理削除
         (soft delete では FK CASCADE が発火しないため — delete_patient と同じ流儀)
      3. 残り (alive) 全 patient の PFV を全件物理削除 (replace 仕様)
      4. patient_ops を順に apply (new / update / resurrect)
      5. NG スタッフ (patient_ng_staff) の差分を適用
      6. pfv_ops を順に INSERT

    例外は呼び出し側にバブルアップ (rollback は呼び出し側で実施).
    """
    # 1) 削除対象 patient の PFV を物理削除
    if patient_ids_to_soft_delete:
        pfvs_for_delete = (
            await db.scalars(
                select(PatientFixedVisit).where(
                    PatientFixedVisit.patient_id.in_(patient_ids_to_soft_delete)
                )
            )
        ).all()
        for pfv in pfvs_for_delete:
            await db.delete(pfv)
        await db.flush()
        # 2) 削除対象 patient を soft delete + 中間テーブル掃除
        for pid in patient_ids_to_soft_delete:
            patient = await db.get(Patient, pid)
            if patient is not None:
                patient.deleted_at = func.now()
        await cleanup_soft_deleted_patient_links(db, patient_ids_to_soft_delete)
        await db.flush()

    # 3) 残り (alive) の patient の PFV を全件物理削除. Excel から再投入する.
    #    削除対象は既に上で消したので、ここでは「alive な patient のもの」のみ.
    alive_patient_ids = (
        await db.scalars(select(Patient.id).where(Patient.deleted_at.is_(None)))
    ).all()
    if alive_patient_ids:
        rest_pfvs = (
            await db.scalars(
                select(PatientFixedVisit).where(
                    PatientFixedVisit.patient_id.in_(list(alive_patient_ids))
                )
            )
        ).all()
        for pfv in rest_pfvs:
            await db.delete(pfv)
        await db.flush()

    # 4) Patient ops
    patients_by_id: dict[UUID, Patient] = {}
    ng_staff_updates: list[tuple[UUID, list[UUID]]] = []
    for op in patient_ops:
        op_type = op.get("_op")
        if op_type == "new":
            data = {k: v for k, v in op.items() if not k.startswith("_")}
            new_patient = Patient(**data)
            db.add(new_patient)
            if "_ng_staff_ids" in op:
                ng_staff_updates.append((op["_new_patient_id"], op["_ng_staff_ids"]))
        elif op_type == "resurrect":
            pid = op["_patient_id"]
            if pid not in patients_by_id:
                patients_by_id[pid] = await db.get(Patient, pid)
            patient = patients_by_id[pid]
            if patient is None:
                continue
            patient.deleted_at = None
            for k, v in op["_updates"].items():
                setattr(patient, k, v)
            if "_ng_staff_ids" in op:
                ng_staff_updates.append((pid, op["_ng_staff_ids"]))
        elif op_type == "update":
            pid = op["_patient_id"]
            if pid not in patients_by_id:
                patients_by_id[pid] = await db.get(Patient, pid)
            patient = patients_by_id[pid]
            if patient is None:
                continue
            for k, v in op.items():
                if k.startswith("_"):
                    continue
                setattr(patient, k, v)
            if "_ng_staff_ids" in op:
                ng_staff_updates.append((pid, op["_ng_staff_ids"]))
    await db.flush()

    # 5) NG スタッフ (patient_ng_staff): 新規患者 INSERT 後に差分適用.
    await apply_ng_staff_updates(db, ng_staff_updates)

    # 6) PFV: 新規 INSERT のみ
    for op in pfv_ops:
        if op.get("_op") != "new":
            continue
        data = {k: v for k, v in op.items() if not k.startswith("_")}
        db.add(PatientFixedVisit(**data))
    await db.flush()
