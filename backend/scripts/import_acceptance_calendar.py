"""「受け入れカレンダー」シート取込スクリプト (W15 Phase 2 B-2).

スケジュール手動.xlsx の「受け入れカレンダー」シートをパースし、
acceptance_calendar テーブルへ bulk upsert する。

シート構造 (0-indexed):
  ブロック 1 (稲毛拠点): Row 2-20
    - B2 (col 1) : タイトル「稲毛区・花見川区・美浜区にお住まいの方」
    - Row 3  : 曜日ヘッダ (C=月,E=火,G=水,I=木,K=金,M=土,O=日)
    - Row 4-18: 時間帯行 (偶数行のみ: 4,6,8,...18)
                B 列 = "HH:MM-", C/E/G/I/K/M/O = ○/△/×/定休日
  ブロック 2 (都賀拠点): Row 22-40 (同構造)

Usage:
    python -m scripts.import_acceptance_calendar --xlsx PATH [--dry-run | --apply]
    python -m scripts.import_acceptance_calendar --xlsx PATH --apply
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
import uuid
from dataclasses import dataclass
from datetime import time
from pathlib import Path
from typing import Any

import openpyxl

# ── sys.path: allow `python -m scripts.…` or direct `python scripts/….py` ──
_SCRIPTS_DIR = Path(__file__).resolve().parent
_BACKEND_ROOT = _SCRIPTS_DIR.parent
for _p in (_SCRIPTS_DIR, _BACKEND_ROOT):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from sqlalchemy import delete, select  # noqa: E402

from app.db.session import dispose_engine, get_session_factory  # noqa: E402
from app.models.acceptance_calendar import AcceptanceCalendar  # noqa: E402
from app.models.office import Office  # noqa: E402

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SHEET_NAME = "受け入れカレンダー"

# ○/△/× → DB status
_STATUS_MAP: dict[str, str] = {
    "○": "available",
    "△": "consult",
    "×": "unavailable",
}

# Japanese weekday column → (weekday_int, col_index_0based)
# Row 3 / Row 23: C(2), E(4), G(6), I(8), K(10), M(12), O(14)
_WEEKDAY_COL_MAP: list[tuple[int, int]] = [
    (0, 2),  # 月 → col C
    (1, 4),  # 火 → col E
    (2, 6),  # 水 → col G
    (3, 8),  # 木 → col I
    (4, 10),  # 金 → col K
    (5, 12),  # 土 → col M
    (6, 14),  # 日 → col O
]

# Block definitions: (block_title_row_0indexed, data_start_row, data_end_row, office_like)
# Rows are 1-indexed in xlsx; we use 0-indexed offsets here for clarity.
# Block 1: title row=2(1-idx), header row=3, data rows=4,6,8,10,12,14,16,18
# Block 2: title row=22, header row=23, data rows=24,26,28,30,32,34,36,38
_BLOCK_SPECS = [
    {
        "title_row": 2,  # 1-indexed xlsx row
        "header_row": 3,
        "data_rows": list(range(4, 19, 2)),  # 4,6,8,10,12,14,16,18
        "office_like": "稲毛",
        "area_label": "稲毛区・花見川区・美浜区",
    },
    {
        "title_row": 22,
        "header_row": 23,
        "data_rows": list(range(24, 39, 2)),  # 24,26,28,30,32,34,36,38
        "office_like": "都賀",
        "area_label": "中央区・若葉区",
    },
]

# time_slot col index (col B = index 1)
_TIME_COL = 1

# Regex to extract HH:MM from "HH:MM-"
_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})")


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class ParsedEntry:
    weekday: int
    time_slot: time
    status: str  # 'available' | 'consult' | 'unavailable'


@dataclass
class BlockResult:
    office_like: str
    area_label: str
    entries: list[ParsedEntry]
    counts: dict[str, int]  # {'available': n, 'consult': n, 'unavailable': n}


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def _parse_time_slot(raw: Any) -> time | None:
    """Parse "HH:MM-" cell → datetime.time, or None on failure."""
    if raw is None:
        return None
    s = str(raw).strip()
    m = _TIME_RE.match(s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    try:
        return time(h, mi, 0)
    except ValueError:
        return None


def _parse_status(raw: Any) -> str | None:
    """Map ○/△/× → status string. 定休日 → None (skip row). blank → 'unavailable'."""
    if raw is None:
        return "unavailable"
    s = str(raw).strip()
    if s == "定休日":
        return None  # sentinel: skip this cell
    if s in _STATUS_MAP:
        return _STATUS_MAP[s]
    if s == "":
        return "unavailable"
    return None  # unknown value → skip


def parse_sheet(xlsx_path: Path) -> list[BlockResult]:
    """Open xlsx and parse all blocks. Returns list of BlockResult."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise KeyError(f"シート '{SHEET_NAME}' が {xlsx_path.name} に見つかりません")

        # Load all rows indexed by 1-based row number (sheet is small, ~40 rows × 16 cols)
        ws2 = wb[SHEET_NAME]
        rows_by_num: dict[int, tuple[Any, ...]] = {}
        for r_idx, row_tuple in enumerate(
            ws2.iter_rows(min_row=1, max_row=45, values_only=True), start=1
        ):
            rows_by_num[r_idx] = tuple(row_tuple) if row_tuple else ()

    finally:
        wb.close()

    results: list[BlockResult] = []
    for spec in _BLOCK_SPECS:
        entries: list[ParsedEntry] = []
        counts: dict[str, int] = {"available": 0, "consult": 0, "unavailable": 0}

        for data_row_num in spec["data_rows"]:
            row = rows_by_num.get(data_row_num, ())
            if not row:
                continue

            # Parse time slot from col B (index 1)
            ts_raw = row[_TIME_COL] if len(row) > _TIME_COL else None
            time_slot = _parse_time_slot(ts_raw)
            if time_slot is None:
                continue  # not a data row

            for weekday, col_idx in _WEEKDAY_COL_MAP:
                cell_val = row[col_idx] if len(row) > col_idx else None
                status = _parse_status(cell_val)
                if status is None:
                    # 定休日 or unrecognised → do not create an entry
                    continue
                entries.append(ParsedEntry(weekday=weekday, time_slot=time_slot, status=status))
                counts[status] += 1

        results.append(
            BlockResult(
                office_like=spec["office_like"],
                area_label=spec["area_label"],
                entries=entries,
                counts=counts,
            )
        )

    return results


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------


async def _resolve_office(session: Any, like_pattern: str, area_label: str) -> Office:
    """Resolve an Office by LIKE match on name. ERROR if not found."""
    result = await session.execute(
        select(Office).where(
            Office.name.like(f"%{like_pattern}%"),
            Office.deleted_at.is_(None),
        )
    )
    offices = result.scalars().all()
    if not offices:
        raise SystemExit(
            f"[ERROR] エリア「{area_label}」の拠点が見つかりません "
            f"(offices.name LIKE '%{like_pattern}%')。"
            " DBに該当拠点を登録してください。"
        )
    if len(offices) > 1:
        names = ", ".join(o.name for o in offices)
        raise SystemExit(
            f"[ERROR] エリア「{area_label}」で複数の拠点が見つかりました: {names}。"
            " like パターンを絞り込んでください。"
        )
    return offices[0]


async def _upsert_block(
    session: Any,
    office: Office,
    block: BlockResult,
    counters: dict[str, int],
) -> None:
    """Delete existing entries for office_id, then insert parsed entries."""
    # Count existing rows to distinguish NEW vs UPDATE
    existing = await session.execute(
        select(AcceptanceCalendar).where(AcceptanceCalendar.office_id == office.id)
    )
    existing_count = len(existing.scalars().all())

    # Delete all existing for this office (idempotent pattern from PUT endpoint)
    await session.execute(
        delete(AcceptanceCalendar).where(AcceptanceCalendar.office_id == office.id)
    )

    # Bulk insert
    for entry in block.entries:
        session.add(
            AcceptanceCalendar(
                id=uuid.uuid4(),
                office_id=office.id,
                weekday=entry.weekday,
                time_slot=entry.time_slot,
                status=entry.status,
                notes=None,
            )
        )

    new_count = len(block.entries)
    if existing_count == 0:
        counters["new"] += new_count
    else:
        counters["update"] += new_count


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------


def _fmt_block_summary(block: BlockResult, office_id: Any) -> str:
    c = block.counts
    total = sum(c.values())
    return (
        f"- {block.office_like}拠点: "
        f"○{c['available']}, △{c['consult']}, ×{c['unavailable']} (計 {total} entries)"
    )


def print_report(
    blocks: list[BlockResult],
    offices: list[Office] | None,
    counters: dict[str, int],
    warnings: int,
    dry_run: bool,
) -> None:
    mode = "[dry-run]" if dry_run else "[apply]"
    print(f"=== 受け入れカレンダー 取込結果 {mode} ===")
    print()
    print("[拠点認識]")
    for i, block in enumerate(blocks):
        if offices:
            oid = offices[i].id
            oname = offices[i].name
            print(f"- {block.office_like}拠点 (id: {oid}): エリア「{block.area_label}」 [{oname}]")
        else:
            print(f"- {block.office_like}拠点: エリア「{block.area_label}」")
    print()
    print("[acceptance_calendar upsert]")
    for i, block in enumerate(blocks):
        oid = offices[i].id if offices else "N/A"
        print(_fmt_block_summary(block, oid))
    print()
    print("[Summary]")
    print(f"- acceptance_calendar: NEW={counters['new']}, UPDATE={counters['update']}")
    print(f"- 警告: {warnings} 件")


# ---------------------------------------------------------------------------
# Main import logic
# ---------------------------------------------------------------------------


async def import_acceptance_calendar(xlsx: Path, dry_run: bool) -> None:
    """Parse xlsx and upsert to DB (or just print in dry-run mode)."""
    blocks = parse_sheet(xlsx)

    if dry_run:
        print_report(
            blocks, offices=None, counters={"new": 0, "update": 0}, warnings=0, dry_run=True
        )
        # Show parsed entry counts per block
        for block in blocks:
            total = sum(block.counts.values())
            print(f"  [{block.office_like}拠点] parsed {total} entries")
        return

    factory = get_session_factory()
    counters: dict[str, int] = {"new": 0, "update": 0}
    resolved_offices: list[Office] = []

    async with factory() as session:
        # Resolve all offices first (error-fast)
        for block in blocks:
            office = await _resolve_office(session, block.office_like, block.area_label)
            resolved_offices.append(office)

        # Upsert each block
        for office, block in zip(resolved_offices, blocks, strict=True):
            await _upsert_block(session, office, block, counters)

        await session.commit()

    print_report(blocks, resolved_offices, counters, warnings=0, dry_run=False)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="「受け入れカレンダー」シートを acceptance_calendar テーブルへ取込"
    )
    p.add_argument("--xlsx", type=Path, required=True, help="path to source xlsx file")
    mode = p.add_mutually_exclusive_group()
    mode.add_argument(
        "--dry-run", action="store_true", help="パース結果を表示するのみ (DB 更新なし)"
    )
    mode.add_argument("--apply", action="store_true", help="DB に書き込む")
    return p


async def _main(xlsx: Path, dry_run: bool) -> None:
    try:
        await import_acceptance_calendar(xlsx, dry_run)
    finally:
        if not dry_run:
            await dispose_engine()


def main() -> int:
    args = build_parser().parse_args()
    if not args.dry_run and not args.apply:
        print("[ERROR] --dry-run または --apply を指定してください", file=sys.stderr)
        return 1
    asyncio.run(_main(args.xlsx, args.dry_run))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
