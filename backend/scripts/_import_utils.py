"""Shared helpers for Excel-driven initial import scripts (W3-B).

Centralises ``argparse`` defaults, ``openpyxl`` row iteration, value coercion
(sex/area/weekday list parsing, "選択なし" → None) and the small CLI summary
that every importer prints. Intentionally lightweight: no DB code lives here,
each importer owns its own SQLAlchemy session usage.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

# Ensure backend/ on sys.path when invoked via `python scripts/foo.py`.
_BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))


_BLANK_TOKENS = {"", "選択なし", "空白", "なし", "None", "none", "null"}

_SEX_MAP = {
    "男性": "male",
    "男": "male",
    "M": "male",
    "male": "male",
    "女性": "female",
    "女": "female",
    "F": "female",
    "female": "female",
}

_WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_WEEKDAY_JP = {
    "月": "Mon", "火": "Tue", "水": "Wed", "木": "Thu",
    "金": "Fri", "土": "Sat", "日": "Sun",
}


def build_parser(description: str) -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=description)
    p.add_argument("xlsx", type=Path, help="path to source xlsx file")
    p.add_argument("--dry-run", action="store_true", help="print counts only")
    return p


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() in _BLANK_TOKENS
    return False


def clean_str(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_sex(value: Any) -> str | None:
    s = clean_str(value)
    if s is None:
        return None
    return _SEX_MAP.get(s, None)


def parse_weekdays(value: Any) -> list[str]:
    """Accept 'Mon, Wed, Fri' or '月,水,金' → ['Mon','Wed','Fri']."""
    s = clean_str(value)
    if s is None:
        return []
    out: list[str] = []
    for token in s.replace("、", ",").replace(" ", "").split(","):
        if not token:
            continue
        if token in _WEEKDAY_NAMES:
            out.append(token)
        elif token in _WEEKDAY_JP:
            out.append(_WEEKDAY_JP[token])
    return out


def parse_areas(value: Any) -> list[str]:
    s = clean_str(value)
    if s is None:
        return []
    return [t.strip() for t in s.replace("、", ",").split(",") if t.strip()]


def parse_int(value: Any) -> int | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    # Strip trailing 人/件 etc.
    digits = "".join(ch for ch in s if ch.isdigit() or ch == "-")
    if not digits or digits == "-":
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def parse_float(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def parse_bool(value: Any, *, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    s = str(value).strip().lower()
    if s in {"true", "1", "yes", "y", "有効", "あり", "継続", "○"}:
        return True
    if s in {"false", "0", "no", "n", "無効", "なし", "✕", "×"}:
        return False
    return default


def parse_time(value: Any) -> time | None:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    s = clean_str(value)
    if s is None:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = clean_str(value)
    if s is None:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    s = clean_str(value)
    if s is None:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_id_list(value: Any) -> list[str]:
    """Comma-separated external ID list (e.g. 'S001,S003' or '選択なし')."""
    s = clean_str(value)
    if s is None:
        return []
    return [t.strip() for t in s.replace("、", ",").split(",") if t.strip()]


def header_index(headers: Iterable[Any]) -> dict[str, int]:
    """Build {normalized_header_text: 0-based-col-index}. Strips/None-skips."""
    out: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        if key:
            out[key] = idx
    return out


def cell(row: tuple[Any, ...], idx_map: dict[str, int], header: str) -> Any:
    idx = idx_map.get(header)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def iter_rows(xlsx_path: Path, sheet_name: str) -> tuple[dict[str, int], list[tuple[Any, ...]]]:
    """Open *xlsx_path*, return (header_index_map, list of data rows)."""
    import openpyxl  # local import: cheap dep, keep top-of-file clean

    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"sheet '{sheet_name}' missing in {xlsx_path.name}")
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {}, []
        idx_map = header_index(header_row)
        data: list[tuple[Any, ...]] = []
        for row in rows_iter:
            if not any(c is not None and c != "" for c in row):
                continue
            data.append(tuple(row))
        return idx_map, data
    finally:
        wb.close()


def print_summary(label: str, *, created: int, updated: int, skipped: int,
                  failed: int, errors: list[str] | None = None) -> None:
    print(
        f"[{label}] created={created} updated={updated} "
        f"skipped={skipped} failed={failed}"
    )
    if errors:
        for line in errors[:20]:
            print(f"  - {line}")
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more")


__all__ = [
    "build_parser",
    "cell",
    "clean_str",
    "header_index",
    "is_blank",
    "iter_rows",
    "normalize_sex",
    "parse_areas",
    "parse_bool",
    "parse_date",
    "parse_datetime",
    "parse_float",
    "parse_id_list",
    "parse_int",
    "parse_time",
    "parse_weekdays",
    "print_summary",
]
