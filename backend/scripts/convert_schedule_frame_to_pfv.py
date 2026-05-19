"""新スプレッドシート「スケジュール枠組み (仮)」シートを CareFlow PFV 形式に変換.

入力:
  - .claude/.../mcp-claude_ai_Google_Drive-read_file_content-1779169466214.txt
    (User 提供スプレッドシート 1P5aH... の生データ)

処理:
  1. ファイル先頭 (L0-L77) の「元データ」シートから患者マスタを構築
  2. L335 以降の複数テーブル (各コース × 6 曜日 × 時間軸) を解析
  3. 患者名 → patient_code 解決
  4. PFV 行生成 (CareFlow 完全置換テンプレート形式)
  5. 既存の _user_master_patient.xlsx を上書き (新患者マスタ + 新 PFV)

出力:
  - backend/scripts/_user_master_patient.xlsx (上書き更新)
  - backend/scripts/_pfv_extraction_log.txt (抽出された PFV の詳細ログ)
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

SOURCE_FILE = Path(
    r"C:\Users\imaizumi.LINEWORKS-NET\.claude\projects"
    r"\C--Users-imaizumi-LINEWORKS-NET-Documents-CareFlow"
    r"\8175abfe-2af3-49cd-9fe8-2d8008ae0777\tool-results"
    r"\mcp-claude_ai_Google_Drive-read_file_content-1779169466214.txt"
)
SCRIPT_DIR = Path(__file__).parent
OUTPUT_XLSX = SCRIPT_DIR / "_user_master_patient.xlsx"
PFV_LOG = SCRIPT_DIR / "_pfv_extraction_log.txt"

WEEKDAY_JP_MAP = {"月曜日": 0, "火曜日": 1, "水曜日": 2, "木曜日": 3, "金曜日": 4, "土曜日": 5}
WEEKDAY_JP_LIST = ["月", "火", "水", "木", "金", "土", "日"]

STATUS_MAP = {
    "稼働": "active",
    "休止": "suspended",
    "入院": "admitted",
    "未開始": "pending",
    "未契約": "cancelled",
    "解約": "cancelled",
}
SEX_MAP = {"女性": "female", "男性": "male"}
SEX_RESTRICTION_MAP = {"女性のみ": "female_only", "男性のみ": "male_only"}
TIME_TYPE_PASS = {"固定", "時間帯", "午前", "午後", "終日"}

# CourseTemplate.code mapping (CareFlow 内で User の A/B/C/D/E がどう対応するか)
# 暫定: そのまま A-E を CourseTemplate.code として使う (CareFlow テンプレートが
# A/B/C/D/E を持っている前提)。実際は CareFlow の course_templates table に
# 該当 code が存在するか要確認。存在しない場合は import 時 fallback で None になる。
COURSE_CODE_MAP = {"A": "A", "B": "B", "C": "C", "D": "D", "E": "E"}


def unwrap_json(raw: str) -> str:
    if raw.startswith("﻿"):
        raw = raw[1:]
    raw = raw.lstrip()
    if raw.startswith('{"fileContent":"'):
        raw = raw[len('{"fileContent":"') :]
        if raw.rstrip().endswith('"}'):
            raw = raw.rstrip()[:-2]
    raw = raw.replace("\\\\_", "_").replace("\\\\n", "\n").replace('\\\\"', '"')
    raw = raw.replace("\\n", "\n").replace('\\"', '"').replace("\\_", "_")
    return raw


def parse_time_hhmm(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if not m:
        return ""
    return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"


def parse_weekdays_en(raw: str) -> list[int]:
    """'Mon, Wed, Fri' → [0, 2, 4]."""
    raw = (raw or "").strip()
    if not raw or raw in ("選択なし", "空欄", "なし"):
        return []
    wd_map = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6}
    out: list[int] = []
    for tok in re.split(r"[,，、\s]+", raw):
        tok = tok.strip()
        if tok in wd_map:
            out.append(wd_map[tok])
    return sorted(set(out))


def is_separator_line(ln: str) -> bool:
    cells = [c.strip() for c in ln.strip("|").split("|")]
    return all(c in (":-:", "---", ":---", "---:", "") for c in cells if c is not None)


def parse_table(lines: list[str], header_idx: int) -> tuple[list[str], list[dict[str, str]]]:
    """Markdown table from header_idx を parse. (header_cells, data_rows) を返す."""
    if header_idx >= len(lines):
        return [], []
    header = lines[header_idx]
    header_cells = [c.strip() for c in header.strip("|").split("|")]
    rows: list[dict[str, str]] = []
    i = header_idx + 1
    # skip separator
    while i < len(lines) and is_separator_line(lines[i]):
        i += 1
    while i < len(lines):
        ln = lines[i]
        if not ln.strip().startswith("|"):
            break
        if is_separator_line(ln):
            i += 1
            continue
        cells = [c.strip() for c in ln.strip("|").split("|")]
        if len(cells) < len(header_cells):
            cells = cells + [""] * (len(header_cells) - len(cells))
        elif len(cells) > len(header_cells):
            cells = cells[: len(header_cells)]
        rows.append(dict(zip(header_cells, cells, strict=False)))
        i += 1
    return header_cells, rows


def build_patient_master(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    """元データ rows → patient_code → patient info dict.

    Also builds name → patient_code reverse map.
    """
    by_code: dict[str, dict[str, str]] = {}
    for row in rows:
        code = row.get("patient_id", "").strip()
        if not code:
            continue
        by_code[code] = {
            "code": code,
            "name": row.get("患者名", "").strip(),
            "kana": row.get("フリガナ", "").strip(),
            "status": STATUS_MAP.get(row.get("稼働状況", "").strip(), "active"),
            "sex": SEX_MAP.get(row.get("性別", "").strip(), ""),
            "address": row.get("住所", "").strip(),
            "lat": row.get("緯度", "").strip(),
            "lng": row.get("経度", "").strip(),
            "sex_restriction": SEX_RESTRICTION_MAP.get(row.get("性別制限", "").strip(), ""),
            "multi_staff": (
                "TRUE" if row.get("必要スタッフ数", "").strip() in ("2人", "2名", "2") else "FALSE"
            ),
            "note": row.get("備考", "").strip(),
            "insurance": ("care" if row.get("保険区分", "").strip() == "介護保険" else "medical"),
            "service_min": _parse_int(row.get("サービス時間", ""), 35),
            "time_type": (
                row.get("時間タイプ", "").strip()
                if row.get("時間タイプ", "").strip() in TIME_TYPE_PASS
                else ""
            ),
            "pref_start": parse_time_hhmm(row.get("希望時間帯（開始）", "")),
            "pref_end": parse_time_hhmm(row.get("希望時間帯（終了）", "")),
            "weekdays_en": ", ".join(
                ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][wd]
                for wd in parse_weekdays_en(row.get("希望曜日（複数可）", ""))
            ),
        }
    return by_code


def _parse_int(s: str, default: int = 35) -> int:
    s = (s or "").strip()
    try:
        return int(s)
    except (ValueError, TypeError):
        return default


def resolve_office_from_address(address: str) -> str:
    if not address:
        return "INAGE"
    if "若葉区都賀" in address or ("都賀" in address and "稲毛" not in address):
        return "TSUGA"
    return "INAGE"


def parse_schedule_frame(
    lines: list[str], start_line: int, end_line: int, *, name_to_code: dict[str, str]
) -> list[tuple[str, int, str, str]]:
    """L335-L460 の複数 PFV テーブルを解析.

    各テーブルは:
      - ヘッダー: `| col0/曜日/拠点/時間帯 | 曜日/コース/氏名 | 曜日/コース/住所 | 曜日/コース/複数 | 曜日/コース/条件 | (繰り返し 6 曜日分) |`
      - データ行: 時刻 | (各曜日コース セル) ...

    返り値: list of (patient_code, weekday, start_time HH:MM, course_code) tuples.
    """
    pfv_records: list[tuple[str, int, str, str]] = []
    i = start_line
    while i < end_line:
        ln = lines[i] if i < len(lines) else ""
        if not ln.strip().startswith("|") or is_separator_line(ln):
            i += 1
            continue
        cells = [c.strip() for c in ln.strip("|").split("|")]
        # ヘッダー候補: col 0 が `数字/拠点/時間帯` 形式 or `数字/時間帯` 形式
        if not re.match(r"^[\d\s\w/]*[時間帯]", cells[0]) and "時間帯" not in cells[0]:
            # データ行 (= 9:30 等の時刻) でもなく header でもない → skip
            i += 1
            continue
        # ここから 1 つのテーブル開始
        header_cells = cells
        col_meta: list[tuple[str, str, str] | None] = []  # (weekday_jp, course, field) or None
        for c in header_cells:
            parts = c.split("/")
            if len(parts) == 3:
                a, b, field = parts
                # 「月曜日/A/氏名」or「4/都賀/時間帯」
                if a in WEEKDAY_JP_MAP:
                    col_meta.append((a, b, field))
                else:
                    col_meta.append(None)  # 時間帯列など特殊
            elif len(parts) == 2:
                # `4/時間帯` 等
                col_meta.append(None)
            else:
                col_meta.append(None)

        # データ行を読む (次の table or 空行まで)
        j = i + 1
        while j < end_line and j < len(lines):
            data_ln = lines[j]
            if not data_ln.strip().startswith("|"):
                break
            if is_separator_line(data_ln):
                j += 1
                continue
            data_cells = [c.strip() for c in data_ln.strip("|").split("|")]
            # 次のヘッダー行か (= col 0 が時刻でない and 時間帯 含む)
            if "時間帯" in data_cells[0]:
                break
            # 時刻行か? col 0 が「9:30」「10:00」等
            time_match = re.match(r"^(\d{1,2}):(\d{2})$", data_cells[0])
            if not time_match:
                j += 1
                continue
            time_str = f"{int(time_match.group(1)):02d}:{int(time_match.group(2)):02d}"
            # 各「氏名」列をスキャン
            for col_idx, meta in enumerate(col_meta):
                if meta is None:
                    continue
                weekday_jp, course, field = meta
                if field != "氏名":
                    continue
                if col_idx >= len(data_cells):
                    continue
                name_cell = data_cells[col_idx].strip()
                if not name_cell or name_cell in ("空", "0", "-"):
                    continue
                # 名前 → patient_code 解決 (全角/半角 space 正規化)
                name_norm = name_cell.replace(" ", "　").strip()
                pcode = name_to_code.get(name_norm)
                if pcode is None:
                    name_norm2 = name_cell.replace("　", " ").strip()
                    pcode = name_to_code.get(name_norm2)
                if pcode is None:
                    continue  # 未解決 (skip)
                weekday = WEEKDAY_JP_MAP[weekday_jp]
                course_code = COURSE_CODE_MAP.get(course, "")
                pfv_records.append((pcode, weekday, time_str, course_code))
            j += 1
        i = j
    return pfv_records


def main() -> None:
    raw = unwrap_json(SOURCE_FILE.read_text(encoding="utf-8"))
    lines = raw.split("\n")
    print(f"Total lines: {len(lines)}")

    # 1) 元データシート (L0-L77) を patient master として parse
    _, motodata_rows = parse_table(lines, 0)
    patient_master = build_patient_master(motodata_rows)
    print(f"Patient master from 元データ: {len(patient_master)} patients")

    # 2) name → patient_code 逆 map (全角/半角 space 正規化込み)
    name_to_code: dict[str, str] = {}
    for code, info in patient_master.items():
        name = info["name"]
        if name:
            name_to_code[name] = code
            name_to_code[name.replace("　", " ")] = code
            name_to_code[name.replace(" ", "　")] = code

    # 3) スケジュール枠組み (仮) を parse (L335-L460)
    pfv_records = parse_schedule_frame(lines, 335, 461, name_to_code=name_to_code)
    # 重複排除 (= 同じ patient + 同曜日 + 同時刻 + 同コース は 1 行に)
    pfv_records_unique = sorted(set(pfv_records))
    print(f"PFV records extracted: {len(pfv_records)} (unique: {len(pfv_records_unique)})")

    # ログ出力
    with PFV_LOG.open("w", encoding="utf-8") as f:
        f.write("# PFV extraction log\n\n")
        f.write(f"Total records: {len(pfv_records_unique)}\n\n")
        f.write("| patient_code | weekday | start_time | course |\n")
        f.write("|---|---|---|---|\n")
        for pcode, wd, ts, cc in pfv_records_unique:
            f.write(f"| {pcode} | {WEEKDAY_JP_LIST[wd]} | {ts} | {cc} |\n")
    print(f"PFV log → {PFV_LOG}")

    # 4) Excel 出力
    wb = Workbook()
    ws_p: Worksheet = wb.active
    ws_p.title = "患者マスタ"
    patient_headers = [
        "患者ID",
        "患者コード",
        "患者名",
        "フリガナ",
        "性別",
        "ステータス",
        "保険区分",
        "住所",
        "緯度",
        "経度",
        "拠点コード",
        "性別制限",
        "複数スタッフ必須",
        "備考",
        "削除フラグ",
    ]
    ws_p.append(patient_headers)
    for code in sorted(patient_master.keys()):
        info = patient_master[code]
        office_code = resolve_office_from_address(info["address"])
        ws_p.append(
            [
                "",
                code,
                info["name"],
                info["kana"],
                info["sex"],
                info["status"],
                info["insurance"],
                info["address"],
                info["lat"],
                info["lng"],
                office_code,
                info["sex_restriction"],
                info["multi_staff"],
                info["note"],
                "",
            ]
        )

    ws_f = wb.create_sheet("固定訪問スケジュール")
    # CareFlow PFV_COLUMNS 完全準拠 (Phase E-7 後の 13 列):
    pfv_headers = [
        "patient_id (※新規時は空欄、code 入力可)",
        "patient_code",
        "患者名",
        "曜日",
        "slot_index",
        "モード",
        "時間タイプ",
        "開始時刻",
        "終了時刻",
        "duration_min",
        "course_template_code",
        "sub_office_code",
        "(削除フラグ)",
    ]
    ws_f.append(pfv_headers)
    # 同 patient × 同 weekday の 2 件目に slot_index=1 を割り当て (CareFlow Wave 37 仕様: 1 日 2 回訪問対応)
    slot_index_counter: dict[tuple[str, int], int] = {}
    for pcode, wd, ts, cc in pfv_records_unique:
        info = patient_master.get(pcode)
        if info is None:
            continue
        duration = info["service_min"]
        # slot_index を patient × weekday 単位でインクリメント (0, 1, 2...)
        key = (pcode, wd)
        slot_idx = slot_index_counter.get(key, 0)
        slot_index_counter[key] = slot_idx + 1
        # end_time = start_time + duration
        start_h, start_m = map(int, ts.split(":"))
        end_total = start_h * 60 + start_m + duration
        end_h, end_m = divmod(end_total, 60)
        end_str = f"{end_h % 24:02d}:{end_m:02d}"
        # time_type は patient の time_type を継承 (固定/時間帯/午前/午後/終日)
        time_type = info["time_type"] or "時間帯"  # default 時間帯
        ws_f.append(
            [
                "",  # patient_id (空 = code で resolve)
                pcode,  # patient_code
                info["name"],  # 患者名
                WEEKDAY_JP_LIST[wd],  # 曜日
                slot_idx,  # slot_index (0, 1, ...)
                "normal",  # モード
                time_type,  # 時間タイプ
                ts,  # 開始時刻
                end_str,  # 終了時刻
                duration,  # duration_min
                cc,  # course_template_code
                "",  # sub_office_code
                "",  # 削除フラグ
            ]
        )

    # Phase E-8: 「希望訪問パターン」シート (1 patient = 1 行).
    # User シートの 時間タイプ / 希望時間帯 / 希望曜日 / サービス時間 等を
    # patient.weekly_pattern に展開する.
    ws_w = wb.create_sheet("希望訪問パターン")
    weekly_headers = [
        "patient_id (※新規時は空欄)",
        "patient_code",
        "患者名",
        "週訪問回数",
        "訪問頻度",
        "訪問週 (例 '1,3')",
        "希望曜日_月",
        "希望曜日_火",
        "希望曜日_水",
        "希望曜日_木",
        "希望曜日_金",
        "希望曜日_土",
        "希望曜日_日",
        "サービス時間 (分)",
        "時間タイプ",
        "希望開始時刻",
        "希望終了時刻",
        "(削除フラグ)",
    ]
    ws_w.append(weekly_headers)
    for code in sorted(patient_master.keys()):
        info = patient_master[code]
        # 希望曜日 list (英) を 7 列 TRUE/FALSE に展開
        weekdays_en_set = set((info["weekdays_en"] or "").replace(",", " ").split())
        wd_flags = [
            "TRUE" if wd in weekdays_en_set else "FALSE"
            for wd in ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")
        ]
        ws_w.append(
            [
                "",  # patient_id (空 = code で resolve)
                code,  # patient_code
                info["name"],  # 患者名 (見やすさ用)
                None,  # 週訪問回数 (User シートには「週訪問回数」あるが、CareFlow 側で activeなら未設定で OK)
                None,  # 訪問頻度
                None,  # 訪問週
                *wd_flags,  # 希望曜日 月-日 (7 列)
                info["service_min"],  # サービス時間
                info["time_type"],  # 時間タイプ
                info["pref_start"],  # 希望開始時刻
                info["pref_end"],  # 希望終了時刻
                "",  # 削除フラグ
            ]
        )

    wb.save(OUTPUT_XLSX)
    print("\n=== 完了 ===")
    print(f"Output: {OUTPUT_XLSX}")
    print(f"Patient master: {len(patient_master)} patients")
    print(f"PFV records: {len(pfv_records_unique)}")
    print(f"Weekly pattern records: {len(patient_master)} (1 per patient)")


if __name__ == "__main__":
    main()
