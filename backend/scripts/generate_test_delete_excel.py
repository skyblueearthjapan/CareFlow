"""テストデータ一括削除用 Excel 生成スクリプト.

Usage:
    cd backend
    python scripts/generate_test_delete_excel.py

出力:
    ../.omc/test_data/test_patients_DELETE.xlsx
    ../.omc/test_data/test_staff_DELETE.xlsx

patient_code / staff_code 列のみ記入し、(削除フラグ) 列に <DELETE> を書いた
最小 Excel を生成する。Backend exporter と同じヘッダー構造を使うため、
インポート機能がそのまま解釈できる。

DB セッション不要。openpyxl で直接構築する。
"""

from __future__ import annotations

import sys
from pathlib import Path

# backend/ をパスに追加して app.* を import できるようにする
SCRIPT_DIR = Path(__file__).parent
BACKEND_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from app.services.patient_excel.schema import (  # noqa: E402
    HEADER_FILL_COLOR as P_HEADER_FILL,
)
from app.services.patient_excel.schema import (  # noqa: E402
    HEADER_FONT_COLOR as P_HEADER_FONT,
)
from app.services.patient_excel.schema import (  # noqa: E402
    MAGIC_DELETE,
    PATIENT_COL_INDEX,
    PATIENT_COLUMNS,
    PFV_COLUMNS,
    SHEET_PATIENTS,
    SHEET_PFV,
)
from app.services.staff_excel.schema import (  # noqa: E402
    HEADER_FILL_COLOR as S_HEADER_FILL,
)
from app.services.staff_excel.schema import (  # noqa: E402
    HEADER_FONT_COLOR as S_HEADER_FONT,
)
from app.services.staff_excel.schema import (  # noqa: E402
    MAGIC_DELETE as S_MAGIC_DELETE,
)
from app.services.staff_excel.schema import (  # noqa: E402
    SHEET_SHIFT,
    SHEET_STAFF,
    SHIFT_COLUMNS,
    STAFF_COL_INDEX,
    STAFF_COLUMNS,
)

# ---------------------------------------------------------------------------
# 出力先
# ---------------------------------------------------------------------------

OUTPUT_DIR = BACKEND_DIR.parent / ".omc" / "test_data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

PATIENTS_DELETE_PATH = OUTPUT_DIR / "test_patients_DELETE.xlsx"
STAFF_DELETE_PATH = OUTPUT_DIR / "test_staff_DELETE.xlsx"

# ---------------------------------------------------------------------------
# テストコード一覧
# ---------------------------------------------------------------------------

PATIENT_CODES = [f"P-TEST-{i:03d}" for i in range(1, 11)]  # P-TEST-001 〜 P-TEST-010
STAFF_CODES = [f"S-TEST-{i:02d}" for i in range(1, 6)]  # S-TEST-01 〜 S-TEST-05


# ---------------------------------------------------------------------------
# ヘルパー: exporter と同じヘッダー行 + dropdown を構築
# ---------------------------------------------------------------------------


def _set_header_row(
    ws: Worksheet,
    columns: list[dict[str, object]],
    header_fill_color: str,
    header_font_color: str,
) -> None:
    header_font = Font(bold=True, color=header_font_color)
    header_fill = PatternFill("solid", fgColor=header_fill_color)
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_def["header"]))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        width = col_def.get("width", 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = int(width)  # type: ignore[arg-type]
    ws.freeze_panes = "A2"


def _attach_dropdowns(
    ws: Worksheet,
    columns: list[dict[str, object]],
    *,
    max_data_rows: int = 1000,
) -> None:
    last_row = 1 + max_data_rows
    for col_idx, col_def in enumerate(columns, start=1):
        dropdown = col_def.get("dropdown")
        if dropdown is None:
            continue
        formula = '"' + ",".join(str(v) for v in dropdown) + '"'  # type: ignore[arg-type]
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "リストにない値です"
        dv.errorTitle = "入力エラー"
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# 患者削除 Excel 構築
# ---------------------------------------------------------------------------


def build_patient_delete_workbook() -> Workbook:
    """患者シート + PFV シート (空) の 2 シート構成.

    患者シートには patient_code 列と (削除フラグ) 列のみ記入する。
    PFV シートはヘッダーのみ (空行なし)。
    """
    wb = Workbook()
    ws_p: Worksheet = wb.active  # type: ignore[assignment]
    ws_p.title = SHEET_PATIENTS

    _set_header_row(ws_p, PATIENT_COLUMNS, P_HEADER_FILL, P_HEADER_FONT)
    _attach_dropdowns(ws_p, PATIENT_COLUMNS)

    # patient_code 列インデックス (0-based → +1 で 1-based)
    code_col = PATIENT_COL_INDEX["patient_code"] + 1
    delete_col = PATIENT_COL_INDEX["delete_flag"] + 1

    for i, code in enumerate(PATIENT_CODES, start=2):
        ws_p.cell(row=i, column=code_col, value=code)
        ws_p.cell(row=i, column=delete_col, value=MAGIC_DELETE)

    # PFV シート: ヘッダーのみ (空シート)
    ws_f: Worksheet = wb.create_sheet(title=SHEET_PFV)
    _set_header_row(ws_f, PFV_COLUMNS, P_HEADER_FILL, P_HEADER_FONT)
    _attach_dropdowns(ws_f, PFV_COLUMNS)

    return wb


# ---------------------------------------------------------------------------
# スタッフ削除 Excel 構築
# ---------------------------------------------------------------------------


def build_staff_delete_workbook() -> Workbook:
    """スタッフシート + 勤務シフトシート (空) の 2 シート構成.

    スタッフシートには staff_code 列と (削除フラグ) 列のみ記入する。
    勤務シフトシートはヘッダーのみ (空行なし)。
    """
    wb = Workbook()
    ws_s: Worksheet = wb.active  # type: ignore[assignment]
    ws_s.title = SHEET_STAFF

    _set_header_row(ws_s, STAFF_COLUMNS, S_HEADER_FILL, S_HEADER_FONT)
    _attach_dropdowns(ws_s, STAFF_COLUMNS)

    code_col = STAFF_COL_INDEX["staff_code"] + 1
    delete_col = STAFF_COL_INDEX["delete_flag"] + 1

    for i, code in enumerate(STAFF_CODES, start=2):
        ws_s.cell(row=i, column=code_col, value=code)
        ws_s.cell(row=i, column=delete_col, value=S_MAGIC_DELETE)

    # 勤務シフトシート: ヘッダーのみ (空シート)
    ws_f: Worksheet = wb.create_sheet(title=SHEET_SHIFT)
    _set_header_row(ws_f, SHIFT_COLUMNS, S_HEADER_FILL, S_HEADER_FONT)
    _attach_dropdowns(ws_f, SHIFT_COLUMNS)

    return wb


# ---------------------------------------------------------------------------
# 検証ヘルパー
# ---------------------------------------------------------------------------


def _verify_patient_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    ws = wb[SHEET_PATIENTS]

    code_col = PATIENT_COL_INDEX["patient_code"] + 1
    delete_col = PATIENT_COL_INDEX["delete_flag"] + 1
    patient_id_col = PATIENT_COL_INDEX["patient_id"] + 1

    codes_found = []
    delete_flags = []
    id_values = []

    for row in range(2, 2 + len(PATIENT_CODES)):
        codes_found.append(ws.cell(row=row, column=code_col).value)
        delete_flags.append(ws.cell(row=row, column=delete_col).value)
        id_values.append(ws.cell(row=row, column=patient_id_col).value)

    assert codes_found == PATIENT_CODES, f"patient_code mismatch: {codes_found}"
    assert all(v == MAGIC_DELETE for v in delete_flags), f"delete_flag mismatch: {delete_flags}"
    assert all(v is None for v in id_values), f"patient_id should be empty: {id_values}"

    # 他の列 (patient_id と code と delete_flag 以外) が空か確認
    total_cols = len(PATIENT_COLUMNS)
    skip_cols = {patient_id_col, code_col, delete_col}
    for row in range(2, 2 + len(PATIENT_CODES)):
        for col in range(1, total_cols + 1):
            if col in skip_cols:
                continue
            val = ws.cell(row=row, column=col).value
            assert val is None, f"Row {row} col {col} should be empty but got: {val!r}"

    print(
        f"  [検証 OK] 患者シート: {len(PATIENT_CODES)} 行、patient_code + <DELETE> のみ記入、他列空"
    )


def _verify_staff_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    ws = wb[SHEET_STAFF]

    code_col = STAFF_COL_INDEX["staff_code"] + 1
    delete_col = STAFF_COL_INDEX["delete_flag"] + 1
    staff_id_col = STAFF_COL_INDEX["staff_id"] + 1

    codes_found = []
    delete_flags = []
    id_values = []

    for row in range(2, 2 + len(STAFF_CODES)):
        codes_found.append(ws.cell(row=row, column=code_col).value)
        delete_flags.append(ws.cell(row=row, column=delete_col).value)
        id_values.append(ws.cell(row=row, column=staff_id_col).value)

    assert codes_found == STAFF_CODES, f"staff_code mismatch: {codes_found}"
    assert all(v == S_MAGIC_DELETE for v in delete_flags), f"delete_flag mismatch: {delete_flags}"
    assert all(v is None for v in id_values), f"staff_id should be empty: {id_values}"

    total_cols = len(STAFF_COLUMNS)
    skip_cols = {staff_id_col, code_col, delete_col}
    for row in range(2, 2 + len(STAFF_CODES)):
        for col in range(1, total_cols + 1):
            if col in skip_cols:
                continue
            val = ws.cell(row=row, column=col).value
            assert val is None, f"Row {row} col {col} should be empty but got: {val!r}"

    print(
        f"  [検証 OK] スタッフシート: {len(STAFF_CODES)} 行、staff_code + <DELETE> のみ記入、他列空"
    )


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main() -> None:
    print("テストデータ削除用 Excel を生成中...")

    # --- 患者 ---
    patient_wb = build_patient_delete_workbook()
    patient_wb.save(str(PATIENTS_DELETE_PATH))
    size_p = PATIENTS_DELETE_PATH.stat().st_size
    print(f"  [保存] {PATIENTS_DELETE_PATH}")
    print(f"         サイズ: {size_p:,} bytes")

    _verify_patient_workbook(PATIENTS_DELETE_PATH)

    # --- スタッフ ---
    staff_wb = build_staff_delete_workbook()
    staff_wb.save(str(STAFF_DELETE_PATH))
    size_s = STAFF_DELETE_PATH.stat().st_size
    print(f"  [保存] {STAFF_DELETE_PATH}")
    print(f"         サイズ: {size_s:,} bytes")

    _verify_staff_workbook(STAFF_DELETE_PATH)

    print()
    print("=== 生成完了 ===")
    print(f"患者削除ファイル : {PATIENTS_DELETE_PATH.name}  ({size_p:,} bytes)")
    print("  - P-TEST-001 〜 P-TEST-010 (10 名) に <DELETE> 付き")
    print(f"スタッフ削除ファイル: {STAFF_DELETE_PATH.name}  ({size_s:,} bytes)")
    print("  - S-TEST-01 〜 S-TEST-05 (5 名) に <DELETE> 付き")
    print()
    print("ファイルを /patients または /staff の Excel インポートでアップロードしてください。")


if __name__ == "__main__":
    main()
