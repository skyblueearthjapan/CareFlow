"""Phase F-3-C: User「スケジュール枠組み（仮）」 vs DB PFV 76 件 diff.

入力:
  - User: Sampledata/2026.05.19段階最新ユーザーシート/スケジュール手動 のコピー.xlsx
    シート: スケジュール枠組み（仮）
  - DB: backend/scripts/_current_db_export.xlsx (= 既存) or 再 export
    シート: 固定訪問パターン

仕様:
  - User シートを ブロック単位 (course = A/B/C/D/E/M) で parse
  - 各ブロック: R1=曜日 header / R2=訪問件数+course / R3=内側 header / R4以降=data
  - 「氏名」セル: 中点ペア「牧・河野」など は SEPARATORS で split
  - 「空」「-」「x」 等は skip
  - 各 patient name を 86 名から resolve (= DB 患者マスタの 86 名 name → code map)

出力: backend/scripts/_diff_pfv.md
  1. PFV 追加候補 (= User にあるが DB に無い)
  2. PFV 削除候補 (= DB にあるが User に無い)
  3. 名前 resolve 失敗
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
DB_XLSX = SCRIPT_DIR / "_current_db_export.xlsx"  # 古い (76 patients)
DB_XLSX_NEW = SCRIPT_DIR / "_final_db_export.xlsx"  # post all fixes (86 + 135 PFV)
USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)
OUTPUT_MD = SCRIPT_DIR / "_diff_pfv.md"

SEPARATORS = ["・", "／", "/", "、", ",", " と ", "&"]
SKIP_VALUES = {"空", "-", "x", "×", "0", "ー"}

# User シートの「元セル」 → [resolved name, ...] の対応表.
# 中点ペア: 同住所 2 名同時刻.
# 親子: 同苗字 2 名同時刻 (= 同住所家族).
# surname-only: 文脈から特定.
PAIR_RESOLVE: dict[str, list[str]] = {
    "槇・河野": ["槇恵", "河野天歩"],  # P089 + P086
    "藤田・渡辺": ["藤田みゆき", "渡辺春香"],  # P074 + P067
    "藤田・河野": ["藤田みゆき", "河野天歩"],  # P074 + P086
    "槇・藤田": ["槇恵", "藤田みゆき"],  # P089 + P074
    "馬場・小俣": ["馬場聖也", "小俣生実和"],  # P050 + P014
    "帆足・海老澤": ["帆足尋光", "海老澤均尚"],  # P069 + P079
    "安永・菅原": ["安永愛菜", "菅原茉結"],  # P065 + P092
    # 親子: 同苗字 2 名 (= 同住所家族)
    "川上親子": ["川上萌黄", "川上礼子"],  # P021 + P080
    "遠藤親子": ["遠藤孝子", "遠藤哲也"],  # P008 + P009
    "増島親子": ["増島英子", "増島由里彩"],  # P058 + P059
}

# 単一氏名の表記揺れ / surname only → 確定 name (DB) への対応.
ALIAS_RESOLVE: dict[str, str] = {
    "青柳あい": "青栁あい",  # 柳 vs 栁 → P001
    "川上もえぎ": "川上萌黄",  # ひらがな vs 漢字 → P021
    "久松由香里": "久松由伽里",  # 香 vs 伽 → P051
    "園田": "園田貞子",  # surname only (only 1 P036)
    "海老澤": "海老澤均尚",  # surname only (only 1 P079)
    "槇": "槇恵",  # surname only (only 1 P089)
    "藤田": "藤田みゆき",  # surname only (複数あるが文脈最頻 P074)
}

WEEKDAY_JP_TO_EN = {
    "月曜日": "mon",
    "火曜日": "tue",
    "水曜日": "wed",
    "木曜日": "thu",
    "金曜日": "fri",
    "土曜日": "sat",
    "日曜日": "sun",
}
# DB Excel export は曜日を日本語短縮 (月/火/...) で書く
WEEKDAY_SHORT_TO_EN = {
    "月": "mon",
    "火": "tue",
    "水": "wed",
    "木": "thu",
    "金": "fri",
    "土": "sat",
    "日": "sun",
}


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return unicodedata.normalize("NFKC", s) if s else ""


def norm_name(v) -> str:
    return norm(v).replace("　", "").replace(" ", "")


def norm_time(v) -> str:
    if v is None or v == "":
        return ""
    s = str(v).strip()
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h, m = int(parts[0]), int(parts[1])
            return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    return s


def split_names(cell_raw: str) -> list[str]:
    s = norm(cell_raw)
    if not s or s in SKIP_VALUES:
        return []
    # 1. ペア / 親子 の事前対応表 (中点 split 前にチェック).
    if s in PAIR_RESOLVE:
        return PAIR_RESOLVE[s]
    # 2. 中点 split
    for sep in SEPARATORS:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip() and p.strip() not in SKIP_VALUES]
            if len(parts) >= 2:
                return parts
    return [s]


def resolve_name(name: str, name_to_code: dict[str, str]) -> str | None:
    """name を patient_code に resolve. 失敗時 None.

    1. norm_name で完全一致 → code
    2. ALIAS_RESOLVE で表記揺れ正規化 → norm_name 再試行
    """
    nm_key = norm_name(name)
    if nm_key in name_to_code:
        return name_to_code[nm_key]
    aliased = ALIAS_RESOLVE.get(name) or ALIAS_RESOLVE.get(nm_key)
    if aliased:
        ali_key = norm_name(aliased)
        if ali_key in name_to_code:
            return name_to_code[ali_key]
    return None


def hidx(header: list, prefix: str) -> int | None:
    for i, c in enumerate(header):
        if c and str(c).startswith(prefix):
            return i
    return None


def load_db_patient_codes(db_xlsx: Path) -> dict[str, str]:
    """name (空白除去) → patient_code."""
    wb = load_workbook(db_xlsx, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header = None
    code_idx = name_idx = None
    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            code_idx = hidx(header, "patient_code")
            name_idx = hidx(header, "患者名")
            continue
        if not row or code_idx is None:
            continue
        code = norm(row[code_idx])
        name = norm(row[name_idx]) if name_idx is not None else ""
        if code.startswith("P") and name:
            out[norm_name(name)] = code
    wb.close()
    return out


def load_db_pfv(db_xlsx: Path, code_to_name: dict[str, str]) -> list[dict]:
    """DB の PFV を (weekday_en, course, time_hhmm, patient_code, duration_min) で抽出."""
    wb = load_workbook(db_xlsx, data_only=True, read_only=True)
    ws = wb["固定訪問パターン"]
    header = None
    h: dict[str, int] = {}
    out: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            for key in [
                "patient_code",
                "曜日",
                "モード",
                "時間タイプ",
                "開始時刻",
                "終了時刻",
                "duration_min",
                "course_template_code",
            ]:
                idx = hidx(header, key)
                if idx is not None:
                    h[key] = idx
            continue
        if not row or "patient_code" not in h:
            continue
        code = norm(row[h["patient_code"]])
        if not code.startswith("P"):
            continue
        wd_raw = norm(row[h["曜日"]])
        # DB は日本語短縮 (月/火) または英語 (mon/tue) どちらでも来る可能性
        wd = WEEKDAY_SHORT_TO_EN.get(wd_raw, wd_raw.lower())
        course = norm(row[h["course_template_code"]])
        st = norm_time(row[h["開始時刻"]])
        et = norm_time(row[h["終了時刻"]])
        dur = (
            row[h["duration_min"]]
            if h.get("duration_min") is not None and h["duration_min"] < len(row)
            else None
        )
        tt = norm(row[h["時間タイプ"]]) if "時間タイプ" in h else ""
        out.append(
            {
                "patient_code": code,
                "weekday": wd,
                "course": course,
                "start": st,
                "end": et,
                "duration": int(dur) if dur is not None else None,
                "time_type": tt,
                "name": code_to_name.get(code, ""),
            }
        )
    wb.close()
    return out


def parse_user_schedule(name_to_code: dict[str, str]) -> tuple[list[dict], list[dict]]:
    """User シート「スケジュール枠組み（仮）」 を parse.

    戻り値: (resolved_pfv, unresolved_findings)
      resolved_pfv: [{weekday, course, start, patient_code, name, raw_cell, is_pair}]
      unresolved_findings: [{weekday, course, start, raw_cell, name, status}]
    """
    wb = load_workbook(USER_XLSX, data_only=True)
    ws = wb["スケジュール枠組み（仮）"]

    resolved: list[dict] = []
    unresolved: list[dict] = []
    max_row = ws.max_row
    ri = 1
    while ri <= max_row:
        row = [ws.cell(row=ri, column=ci).value for ci in range(1, ws.max_column + 1)]
        if not row or not row[1]:
            ri += 1
            continue
        # ブロック header の検出: R1 のいずれかセルが「月曜日」「火曜日」etc
        weekday_cells = [(i, v) for i, v in enumerate(row) if norm(v) in WEEKDAY_JP_TO_EN]
        if not weekday_cells:
            ri += 1
            continue
        # R2: course
        if ri + 1 > max_row:
            break
        row2 = [ws.cell(row=ri + 1, column=ci).value for ci in range(1, ws.max_column + 1)]
        # 各 weekday block の course を取得 (R2: col_idx は氏名列、course は col_idx 自体)
        weekday_blocks: list[tuple[str, str, int]] = []
        for col_idx, wd_jp in weekday_cells:
            wd_en = WEEKDAY_JP_TO_EN[norm(wd_jp)]
            course_val = row2[col_idx] if col_idx < len(row2) else None
            course = norm(course_val) if course_val else ""
            # course が「6.0」等の数字なら違う列との混同。「A/B/C/D/E/M」 期待
            if not re.match(r"^[A-Z]$", course):
                # 隣の列が course の可能性 (R2 で  '6.0', 'A' なので col_idx に 6.0、+1 に A)
                # ↑ もう col_idx + 1 で取れているはず.
                # それでも違う場合は course を空にする (= 後で skip)
                course = ""
            # block の data 開始列: col_idx = 「曜日」 という merge cell の先頭、
            # 内側 column は (時間帯, 氏名, 住所, 複数, 条件) で col_idx 自体が時間帯, +1 が氏名
            weekday_blocks.append((wd_en, course, col_idx))

        # R3: 内側 header (skip)
        # R4 以降の data row
        data_ri = ri + 3
        while data_ri <= max_row:
            data_row = [ws.cell(row=data_ri, column=ci).value for ci in range(1, ws.max_column + 1)]
            # 全セル空 → block 終わり
            if all(v is None or norm(v) == "" for v in data_row):
                break
            # 別の block header に到達? (col_idx に weekday)
            check_weekday = any(norm(v) in WEEKDAY_JP_TO_EN for v in data_row)
            if check_weekday:
                break
            # 各 block 内処理
            for wd_en, course, col_idx in weekday_blocks:
                # R1 「月曜日」 セル col_idx は merge 内の代表列で、実際の block
                # 内列は (時間帯, 氏名, 住所, 複数, 条件) = (col_idx-1, col_idx, col_idx+1, ...)
                time_col = col_idx - 1
                name_col = col_idx
                time_v = data_row[time_col] if 0 <= time_col < len(data_row) else None
                name_v = data_row[name_col] if 0 <= name_col < len(data_row) else None
                t_str = norm_time(time_v)
                if not t_str:
                    continue
                names = split_names(name_v)
                if not names:
                    continue
                is_pair = len(names) > 1
                for nm in names:
                    code = resolve_name(nm, name_to_code)
                    if code:
                        resolved.append(
                            {
                                "weekday": wd_en,
                                "course": course,
                                "start": t_str,
                                "patient_code": code,
                                "name": nm,
                                "raw_cell": str(name_v).strip() if name_v else "",
                                "is_pair": is_pair,
                            }
                        )
                    else:
                        unresolved.append(
                            {
                                "weekday": wd_en,
                                "course": course,
                                "start": t_str,
                                "raw_cell": str(name_v).strip() if name_v else "",
                                "name": nm,
                                "status": "unresolved",
                            }
                        )
            data_ri += 1
        ri = data_ri + 1  # 次のブロックを探す
    wb.close()
    return resolved, unresolved


def main() -> None:
    # post-import DB 使用 (86 patients)
    db_xlsx = DB_XLSX_NEW if DB_XLSX_NEW.exists() else DB_XLSX
    print(f"DB xlsx: {db_xlsx}")

    name_to_code = load_db_patient_codes(db_xlsx)
    code_to_name = {v: k for k, v in name_to_code.items()}
    print(f"DB patients (86 expected): {len(name_to_code)}")

    db_pfv = load_db_pfv(db_xlsx, code_to_name)
    print(f"DB PFV: {len(db_pfv)}")

    resolved, unresolved = parse_user_schedule(name_to_code)
    print(f"User シート抽出: resolved={len(resolved)}, unresolved={len(unresolved)}")

    # キー化: (patient_code, weekday, course)  ← duplicate may occur for same patient with multiple PFV
    user_keys: dict[tuple[str, str, str], list[dict]] = {}
    for r in resolved:
        k = (r["patient_code"], r["weekday"], r["course"])
        user_keys.setdefault(k, []).append(r)
    db_keys: dict[tuple[str, str, str], list[dict]] = {}
    for r in db_pfv:
        k = (r["patient_code"], r["weekday"], r["course"])
        db_keys.setdefault(k, []).append(r)

    user_only_keys = sorted(set(user_keys) - set(db_keys))
    db_only_keys = sorted(set(db_keys) - set(user_keys))
    common_keys = sorted(set(user_keys) & set(db_keys))

    # 共通でも 時刻 / duration がずれているもの
    time_mismatch: list[tuple] = []
    for k in common_keys:
        user_starts = sorted({r["start"] for r in user_keys[k]})
        db_starts = sorted({r["start"] for r in db_keys[k]})
        if user_starts != db_starts:
            time_mismatch.append((k, user_starts, db_starts))

    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase F-3-C: PFV diff (User 枠組み vs DB 86 名後 PFV 76 件)\n\n")
        f.write(f"- User: `{USER_XLSX.name}` シート「スケジュール枠組み（仮）」\n")
        f.write(f"- DB: `{db_xlsx.name}` シート「固定訪問パターン」\n")
        f.write("- 突合キー: (patient_code, weekday, course)\n\n")
        f.write("## サマリ\n\n")
        f.write(f"- User シート抽出 (resolved): **{len(resolved)}** 件 (中点ペア split 後)\n")
        f.write(f"- User シート 未解決 (name 不一致): **{len(unresolved)}** 件\n")
        f.write(f"- DB PFV: **{len(db_pfv)}** 件\n")
        f.write(f"- 共通 key: **{len(common_keys)}**\n")
        f.write(f"- User のみ (PFV 追加候補): **{len(user_only_keys)}**\n")
        f.write(f"- DB のみ (PFV 削除候補): **{len(db_only_keys)}**\n")
        f.write(f"- 共通で時刻ずれ: **{len(time_mismatch)}**\n\n")

        # 1. User のみ
        f.write("## 1. PFV 追加候補 (User シートにあるが DB に無い)\n\n")
        if not user_only_keys:
            f.write("_該当なし_\n\n")
        else:
            f.write("| code | 氏名 | 曜日 | コース | 時刻 | 元セル |\n")
            f.write("|---|---|---|---|---|---|\n")
            for k in user_only_keys:
                code, wd, course = k
                name = code_to_name.get(code, "")
                for r in user_keys[k]:
                    pair_mark = " (ペア)" if r["is_pair"] else ""
                    f.write(
                        f"| {code} | {name}{pair_mark} | {wd} | {course} | {r['start']} | `{r['raw_cell']}` |\n"
                    )
            f.write("\n")

        # 2. DB のみ
        f.write("## 2. PFV 削除候補 (DB にあるが User シートに無い)\n\n")
        if not db_only_keys:
            f.write("_該当なし_\n\n")
        else:
            f.write("| code | 氏名 | 曜日 | コース | 時刻 | duration |\n")
            f.write("|---|---|---|---|---|---|\n")
            for k in db_only_keys:
                code, wd, course = k
                for r in db_keys[k]:
                    f.write(
                        f"| {code} | {r['name']} | {wd} | {course} | {r['start']} | {r['duration']}分 |\n"
                    )
            f.write("\n")

        # 3. 時刻ずれ
        f.write("## 3. 共通 key で時刻ずれ\n\n")
        if not time_mismatch:
            f.write("_該当なし_\n\n")
        else:
            f.write("| code | 氏名 | 曜日 | コース | User 時刻 | DB 時刻 |\n")
            f.write("|---|---|---|---|---|---|\n")
            for k, us, ds in time_mismatch:
                code, wd, course = k
                name = code_to_name.get(code, "")
                f.write(
                    f"| {code} | {name} | {wd} | {course} | {','.join(us)} | {','.join(ds)} |\n"
                )
            f.write("\n")

        # 4. unresolved
        if unresolved:
            f.write("## 4. 名前 resolve 失敗 (User シートに居るが DB マスタに無い)\n\n")
            f.write("| 曜日 | コース | 時刻 | 元セル | 試行 name |\n")
            f.write("|---|---|---|---|---|\n")
            seen = set()
            for u in unresolved:
                key = (u["raw_cell"], u["name"])
                if key in seen:
                    continue
                seen.add(key)
                f.write(
                    f"| {u['weekday']} | {u['course']} | {u['start']} | `{u['raw_cell']}` | {u['name']} |\n"
                )
            f.write("\n")

    print("\n=== 出力 ===")
    print(f"{OUTPUT_MD}")


if __name__ == "__main__":
    main()
