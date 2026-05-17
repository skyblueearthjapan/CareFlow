"""
generate_edit_test_excel.py
----------------------------
既存エクスポート済み Excel を「一目で編集が分かる」テスト用ファイルに加工して別名保存する。
import 後に「ちゃんと更新として反映された」ことを確認するためのデータ生成スクリプト。

使用方法:
    python backend/scripts/generate_edit_test_excel.py

入力:
    .omc/test_data/patients_master_2026-05-17.xlsx
    .omc/test_data/staff_master_2026-05-17.xlsx

出力:
    .omc/test_data/patients_master_2026-05-17_EDITED.xlsx
    .omc/test_data/staff_master_2026-05-17_EDITED.xlsx
"""

from __future__ import annotations

import os
import sys
from datetime import datetime

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# パス設定
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
TEST_DATA_DIR = os.path.join(REPO_ROOT, ".omc", "test_data")

PATIENTS_IN = os.path.join(TEST_DATA_DIR, "patients_master_2026-05-17.xlsx")
PATIENTS_OUT = os.path.join(TEST_DATA_DIR, "patients_master_2026-05-17_EDITED.xlsx")
STAFF_IN = os.path.join(TEST_DATA_DIR, "staff_master_2026-05-17.xlsx")
STAFF_OUT = os.path.join(TEST_DATA_DIR, "staff_master_2026-05-17_EDITED.xlsx")

EDIT_LABEL = "[編集テスト 2026-05-17] 備考更新"


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------
def col_index(ws, col_name: str) -> int:
    """ヘッダー行からカラム名に一致する列インデックス (1-based) を返す。
    完全一致で見つからない場合は先頭一致を試みる。"""
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None and col_name in str(h):
            return c
    raise ValueError(
        f"列が見つかりません: {col_name!r}  (ヘッダー: {[ws.cell(1, c).value for c in range(1, ws.max_column + 1)]})"
    )


def shift_time(time_str: str, minutes: int) -> str:
    """'HH:MM' 文字列を minutes 分ずらして返す。"""
    if time_str is None:
        return time_str
    h, m = map(int, time_str.split(":"))
    total = h * 60 + m + minutes
    total = total % (24 * 60)  # 24h 超えは wrap
    return f"{total // 60:02d}:{total % 60:02d}"


# ---------------------------------------------------------------------------
# 編集サマリ収集用
# ---------------------------------------------------------------------------
summary_lines: list[str] = []


def log_change(section: str, row: int, label: str, old_val, new_val):
    line = f"  Row {row}: {label} / {repr(old_val)} → {repr(new_val)}"
    print(f"[変更] {section}  {line.strip()}")
    summary_lines.append((section, line))


# ---------------------------------------------------------------------------
# patients_master 編集
# ---------------------------------------------------------------------------
def edit_patients():
    print(f"\n{'=' * 60}")
    print("patients_master 編集開始")
    print(f"  入力: {PATIENTS_IN}")
    print(f"  出力: {PATIENTS_OUT}")
    print(f"{'=' * 60}")

    wb = openpyxl.load_workbook(PATIENTS_IN)

    # ---- シート1: 患者マスタ ----
    ws_master = wb["患者マスタ"]
    col_biko = col_index(ws_master, "備考")
    col_address = col_index(ws_master, "住所")
    col_status = col_index(ws_master, "ステータス")
    col_name = col_index(ws_master, "患者名")

    # Row 2: 備考のみ上書き
    r = 2
    name_r2 = ws_master.cell(r, col_name).value
    old_biko = ws_master.cell(r, col_biko).value
    ws_master.cell(r, col_biko).value = EDIT_LABEL
    log_change("[patients] 患者マスタ", r, f"{name_r2} / 備考", old_biko, EDIT_LABEL)

    # Row 3: 備考上書き + 住所末尾 append
    r = 3
    name_r3 = ws_master.cell(r, col_name).value
    old_biko = ws_master.cell(r, col_biko).value
    ws_master.cell(r, col_biko).value = EDIT_LABEL
    log_change("[patients] 患者マスタ", r, f"{name_r3} / 備考", old_biko, EDIT_LABEL)

    old_addr = ws_master.cell(r, col_address).value
    new_addr = (old_addr or "") + " (テスト編集)"
    ws_master.cell(r, col_address).value = new_addr
    log_change("[patients] 患者マスタ", r, f"{name_r3} / 住所", old_addr, new_addr)

    # Row 4: 備考上書き + ステータス → suspended
    r = 4
    name_r4 = ws_master.cell(r, col_name).value
    old_biko = ws_master.cell(r, col_biko).value
    ws_master.cell(r, col_biko).value = EDIT_LABEL
    log_change("[patients] 患者マスタ", r, f"{name_r4} / 備考", old_biko, EDIT_LABEL)

    old_status = ws_master.cell(r, col_status).value
    new_status = "suspended"
    ws_master.cell(r, col_status).value = new_status
    log_change("[patients] 患者マスタ", r, f"{name_r4} / ステータス", old_status, new_status)

    # ---- シート2: 固定訪問スケジュール ----
    ws_sched = wb["固定訪問スケジュール"]
    col_code_s = col_index(ws_sched, "patient_code")
    col_name_s = col_index(ws_sched, "患者名")
    col_youbi = col_index(ws_sched, "曜日")
    col_start = col_index(ws_sched, "開始時刻")
    col_end = col_index(ws_sched, "終了時刻")
    col_duration = col_index(ws_sched, "duration_min")

    # P001 の行を全部収集
    p001_rows = []
    for r in range(2, ws_sched.max_row + 1):
        if ws_sched.cell(r, col_code_s).value == "P001":
            p001_rows.append(r)

    if not p001_rows:
        print("[WARN] 固定訪問スケジュールに P001 の行が見つかりません")
    else:
        # 最初の行: 開始時刻を 30 分後ろにずらす
        first_row = p001_rows[0]
        p001_name = ws_sched.cell(first_row, col_name_s).value
        p001_youbi_first = ws_sched.cell(first_row, col_youbi).value
        old_start = ws_sched.cell(first_row, col_start).value
        old_end = ws_sched.cell(first_row, col_end).value
        new_start = shift_time(old_start, 30)
        new_end = shift_time(old_end, 30)
        ws_sched.cell(first_row, col_start).value = new_start
        ws_sched.cell(first_row, col_end).value = new_end
        log_change(
            "[patients] 固定訪問スケジュール",
            first_row,
            f"{p001_name} {p001_youbi_first}曜 / 開始時刻",
            old_start,
            new_start,
        )
        log_change(
            "[patients] 固定訪問スケジュール",
            first_row,
            f"{p001_name} {p001_youbi_first}曜 / 終了時刻",
            old_end,
            new_end,
        )

        # 最後の行: duration_min を 30 → 45
        last_row = p001_rows[-1]
        p001_youbi_last = ws_sched.cell(last_row, col_youbi).value
        old_dur = ws_sched.cell(last_row, col_duration).value
        new_dur = 45
        ws_sched.cell(last_row, col_duration).value = new_dur
        log_change(
            "[patients] 固定訪問スケジュール",
            last_row,
            f"{p001_name} {p001_youbi_last}曜 / duration_min",
            old_dur,
            new_dur,
        )

    wb.save(PATIENTS_OUT)
    print(f"\n[保存] {PATIENTS_OUT}")
    size = os.path.getsize(PATIENTS_OUT)
    print(f"[サイズ] {size:,} bytes")


# ---------------------------------------------------------------------------
# staff_master 編集
# ---------------------------------------------------------------------------
def edit_staff():
    print(f"\n{'=' * 60}")
    print("staff_master 編集開始")
    print(f"  入力: {STAFF_IN}")
    print(f"  出力: {STAFF_OUT}")
    print(f"{'=' * 60}")

    wb = openpyxl.load_workbook(STAFF_IN)

    # ---- シート1: スタッフマスタ ----
    ws_master = wb["スタッフマスタ"]
    col_biko = col_index(ws_master, "備考")
    col_kana = col_index(ws_master, "フリガナ")
    col_sname = col_index(ws_master, "スタッフ名")

    # Row 2: 備考のみ上書き
    r = 2
    name_r2 = ws_master.cell(r, col_sname).value
    old_biko = ws_master.cell(r, col_biko).value
    ws_master.cell(r, col_biko).value = EDIT_LABEL
    log_change("[staff] スタッフマスタ", r, f"{name_r2} / 備考", old_biko, EDIT_LABEL)

    # Row 3: 備考上書き + フリガナ末尾 append
    r = 3
    name_r3 = ws_master.cell(r, col_sname).value
    old_biko = ws_master.cell(r, col_biko).value
    ws_master.cell(r, col_biko).value = EDIT_LABEL
    log_change("[staff] スタッフマスタ", r, f"{name_r3} / 備考", old_biko, EDIT_LABEL)

    old_kana = ws_master.cell(r, col_kana).value
    new_kana = (old_kana or "") + " (テスト)"
    ws_master.cell(r, col_kana).value = new_kana
    log_change("[staff] スタッフマスタ", r, f"{name_r3} / フリガナ", old_kana, new_kana)

    # ---- シート2: 勤務シフト ----
    ws_shift = wb["勤務シフト"]
    col_code_sh = col_index(ws_shift, "staff_code")
    col_sname_sh = col_index(ws_shift, "staff_name")
    col_youbi_sh = col_index(ws_shift, "曜日")
    col_kinshu = col_index(ws_shift, "勤務")
    col_end_sh = col_index(ws_shift, "終了時刻")
    col_index(ws_shift, "開始時刻")

    # S001 の行を全部収集
    s001_rows_by_youbi: dict[str, int] = {}
    for r in range(2, ws_shift.max_row + 1):
        if ws_shift.cell(r, col_code_sh).value == "S001":
            youbi = ws_shift.cell(r, col_youbi_sh).value
            s001_rows_by_youbi[youbi] = r

    if not s001_rows_by_youbi:
        print("[WARN] 勤務シフトに S001 の行が見つかりません")
    else:
        s001_name = ws_shift.cell(list(s001_rows_by_youbi.values())[0], col_sname_sh).value

        # 月曜: 終了時刻を 30 分繰り上げ
        if "月" in s001_rows_by_youbi:
            r = s001_rows_by_youbi["月"]
            old_end = ws_shift.cell(r, col_end_sh).value
            new_end = shift_time(old_end, -30)
            ws_shift.cell(r, col_end_sh).value = new_end
            log_change(
                "[staff] 勤務シフト",
                r,
                f"{s001_name} 月曜 / 終了時刻",
                old_end,
                new_end,
            )
        else:
            print("[WARN] S001 の月曜行が見つかりません")

        # 金曜: 勤務 → FALSE
        if "金" in s001_rows_by_youbi:
            r = s001_rows_by_youbi["金"]
            old_kinshu = ws_shift.cell(r, col_kinshu).value
            new_kinshu = "FALSE"
            ws_shift.cell(r, col_kinshu).value = new_kinshu
            log_change(
                "[staff] 勤務シフト",
                r,
                f"{s001_name} 金曜 / 勤務",
                old_kinshu,
                new_kinshu,
            )
        else:
            print("[WARN] S001 の金曜行が見つかりません")

    wb.save(STAFF_OUT)
    print(f"\n[保存] {STAFF_OUT}")
    size = os.path.getsize(STAFF_OUT)
    print(f"[サイズ] {size:,} bytes")


# ---------------------------------------------------------------------------
# 検証: 編集箇所が正しく保存されているか再読み込みして確認
# ---------------------------------------------------------------------------
def verify():
    print(f"\n{'=' * 60}")
    print("検証: 出力ファイルを再読み込みして編集箇所を確認")
    print(f"{'=' * 60}")

    # --- patients ---
    wb_p = openpyxl.load_workbook(PATIENTS_OUT)
    ws_pm = wb_p["患者マスタ"]
    col_biko_p = col_index(ws_pm, "備考")
    col_address_p = col_index(ws_pm, "住所")
    col_status_p = col_index(ws_pm, "ステータス")
    col_index(ws_pm, "患者名")

    errors = []

    def check(label, actual, expected):
        if actual == expected:
            print(f"  [OK] {label}: {repr(actual)}")
        else:
            msg = f"  [NG] {label}: expected={repr(expected)}, actual={repr(actual)}"
            print(msg)
            errors.append(msg)

    print("\n[patients] 患者マスタ")
    check("Row2 備考", ws_pm.cell(2, col_biko_p).value, EDIT_LABEL)
    check("Row3 備考", ws_pm.cell(3, col_biko_p).value, EDIT_LABEL)
    addr_r3 = ws_pm.cell(3, col_address_p).value
    print(f"  [OK] Row3 住所末尾: {repr(addr_r3[-10:] if addr_r3 else None)}")
    check("Row4 備考", ws_pm.cell(4, col_biko_p).value, EDIT_LABEL)
    check("Row4 ステータス", ws_pm.cell(4, col_status_p).value, "suspended")

    # 編集対象外 row5 が変わっていないか (備考は None のまま)
    row5_biko = ws_pm.cell(5, col_biko_p).value
    if row5_biko != EDIT_LABEL:
        print(f"  [OK] Row5(非編集) 備考変更なし: {repr(row5_biko)}")
    else:
        errors.append("Row5(非編集行) の備考が誤って変更されている")

    ws_ps = wb_p["固定訪問スケジュール"]
    col_code_ps = col_index(ws_ps, "patient_code")
    col_start_ps = col_index(ws_ps, "開始時刻")
    col_dur_ps = col_index(ws_ps, "duration_min")

    p001_rows_v = [
        r for r in range(2, ws_ps.max_row + 1) if ws_ps.cell(r, col_code_ps).value == "P001"
    ]
    print(f"\n[patients] 固定訪問スケジュール (P001 rows: {p001_rows_v})")
    if p001_rows_v:
        fr = p001_rows_v[0]
        lr = p001_rows_v[-1]
        print(f"  [OK] Row{fr} 開始時刻: {ws_ps.cell(fr, col_start_ps).value}")
        print(f"  [OK] Row{lr} duration_min: {ws_ps.cell(lr, col_dur_ps).value}")
        if ws_ps.cell(lr, col_dur_ps).value != 45:
            errors.append(f"Row{lr} duration_min が 45 になっていない")

    # --- staff ---
    wb_s = openpyxl.load_workbook(STAFF_OUT)
    ws_sm = wb_s["スタッフマスタ"]
    col_biko_s = col_index(ws_sm, "備考")
    col_kana_s = col_index(ws_sm, "フリガナ")
    col_index(ws_sm, "スタッフ名")

    print("\n[staff] スタッフマスタ")
    check("Row2 備考", ws_sm.cell(2, col_biko_s).value, EDIT_LABEL)
    check("Row3 備考", ws_sm.cell(3, col_biko_s).value, EDIT_LABEL)
    kana_r3 = ws_sm.cell(3, col_kana_s).value
    print(f"  [OK] Row3 フリガナ末尾: {repr(kana_r3[-8:] if kana_r3 else None)}")

    ws_ss = wb_s["勤務シフト"]
    col_code_ss = col_index(ws_ss, "staff_code")
    col_youbi_ss = col_index(ws_ss, "曜日")
    col_kinshu_s = col_index(ws_ss, "勤務")
    col_end_ss = col_index(ws_ss, "終了時刻")

    s001_by_youbi_v = {}
    for r in range(2, ws_ss.max_row + 1):
        if ws_ss.cell(r, col_code_ss).value == "S001":
            s001_by_youbi_v[ws_ss.cell(r, col_youbi_ss).value] = r

    print(f"\n[staff] 勤務シフト (S001 rows by youbi: {s001_by_youbi_v})")
    if "月" in s001_by_youbi_v:
        r = s001_by_youbi_v["月"]
        print(f"  [OK] Row{r} 月曜 終了時刻: {ws_ss.cell(r, col_end_ss).value}")
    if "金" in s001_by_youbi_v:
        r = s001_by_youbi_v["金"]
        actual_kinshu = ws_ss.cell(r, col_kinshu_s).value
        check(f"Row{r} 金曜 勤務", actual_kinshu, "FALSE")

    print()
    if errors:
        print(f"[検証結果] NG {len(errors)} 件")
        for e in errors:
            print(f"  {e}")
    else:
        print("[検証結果] 全 OK - 編集が正しく保存されました")


# ---------------------------------------------------------------------------
# サマリ表示
# ---------------------------------------------------------------------------
def print_summary():
    print(f"\n{'=' * 60}")
    print("=== 編集対象一覧 ===")
    current_section = None
    for section, line in summary_lines:
        if section != current_section:
            print(f"\n{section}")
            current_section = section
        print(line)
    print()


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"生成開始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 入力ファイル存在確認
    for path in [PATIENTS_IN, STAFF_IN]:
        if not os.path.exists(path):
            print(f"[ERROR] ファイルが見つかりません: {path}")
            sys.exit(1)

    edit_patients()
    edit_staff()
    print_summary()
    verify()

    print(f"\n生成完了: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
