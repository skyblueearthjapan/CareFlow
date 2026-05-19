"""File 2 訪問看護よりより様 の患者マスタ件数を確認."""

from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SAMPLE_DIR = Path(
    r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata\2026.05.19段階最新ユーザーシート"
)

# File 2: 訪問看護：よりより様/Ver.1 (1).xlsx
FILE2_CANDIDATES = list(SAMPLE_DIR.glob("訪問看護*.xlsx"))


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return unicodedata.normalize("NFKC", s) if s else ""


def main() -> None:
    for f in FILE2_CANDIDATES:
        print(f"FILE: {f.name}")
        wb = load_workbook(f, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if "患者" not in sheet_name:
                continue
            ws = wb[sheet_name]
            count = 0
            active_count = 0
            for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
                if ri == 1:
                    continue
                if not row or row[0] is None:
                    continue
                code = norm(row[0])
                if code.startswith("P"):
                    count += 1
                    status = norm(row[3]) if len(row) > 3 else ""
                    if status in ("稼働", "active"):
                        active_count += 1
            print(f"  シート「{sheet_name}」: {count} 行 (稼働: {active_count})")
        wb.close()


if __name__ == "__main__":
    main()
