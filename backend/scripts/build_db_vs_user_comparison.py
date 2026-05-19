"""Phase F-2: 現 DB の PFV を 「曜日 × コース × 時刻」 マトリクスに組み直して
User の「スケジュール枠組み (仮)」 と比較レポートを生成する.

入力:
  - backend/scripts/_current_db_export.xlsx (= 今 export した DB の状態)
  - _unresolved_patterns.md (= F-1 で出した抜けリスト)

出力:
  - backend/scripts/_db_vs_user_comparison.md
    各曜日 × コース 単位で、現 DB に入っている PFV を時刻順に並べる.
    User が手元のシートと突合できる形式.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
EXPORT_XLSX = SCRIPT_DIR / "_current_db_export.xlsx"
OUTPUT_MD = SCRIPT_DIR / "_db_vs_user_comparison.md"

WEEKDAY_EN_TO_JP = {
    "mon": "月",
    "tue": "火",
    "wed": "水",
    "thu": "木",
    "fri": "金",
    "sat": "土",
    "sun": "日",
}
WEEKDAY_ORDER = ["月", "火", "水", "木", "金", "土", "日"]


def main() -> None:
    wb = load_workbook(EXPORT_XLSX, data_only=True)
    print(f"sheets: {wb.sheetnames}")

    def hidx(header: list, prefix: str) -> int:
        """Find column index by header prefix (header 内 cell に注釈混入対応)."""
        for i, c in enumerate(header):
            if c and str(c).startswith(prefix):
                return i
        raise KeyError(prefix)

    # 患者マスタ
    ws_p = wb["患者マスタ"]
    header_p = [c.value for c in ws_p[1]]
    code_idx = hidx(header_p, "patient_code")
    name_idx = hidx(header_p, "患者名")
    try:
        office_idx = hidx(header_p, "拠点コード")
    except KeyError:
        office_idx = None
    pid_to_meta: dict[str, tuple[str, str, str]] = {}
    code_to_name: dict[str, str] = {}
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pid = str(row[0])
        code = str(row[code_idx]) if row[code_idx] else "?"
        name = str(row[name_idx]) if row[name_idx] else ""
        office = str(row[office_idx]) if office_idx is not None and row[office_idx] else ""
        pid_to_meta[pid] = (code, name, office)
        code_to_name[code] = name
    print(f"patients: {len(pid_to_meta)}")

    # PFV
    ws_f = wb["固定訪問パターン"]
    header_f = [c.value for c in ws_f[1]]
    h = {
        "patient_id": hidx(header_f, "patient_id"),
        "patient_code": hidx(header_f, "patient_code"),
        "患者名": hidx(header_f, "患者名"),
        "曜日": hidx(header_f, "曜日"),
        "slot_index": hidx(header_f, "slot_index"),
        "モード": hidx(header_f, "モード"),
        "時間タイプ": hidx(header_f, "時間タイプ"),
        "開始時刻": hidx(header_f, "開始時刻"),
        "終了時刻": hidx(header_f, "終了時刻"),
        "duration_min": hidx(header_f, "duration_min"),
        "course_template_code": hidx(header_f, "course_template_code"),
        "sub_office_code": hidx(header_f, "sub_office_code"),
    }

    # PFV rows
    # 期待 columns: patient_id / patient_code / 患者名 / 曜日 / slot_index / モード / 時間タイプ / 開始時刻 / 終了時刻 / duration_min / course_template_code / sub_office_code / 削除フラグ
    pfvs: list[dict] = []
    for row in ws_f.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rec = {
            "patient_id": str(row[h["patient_id"]]),
            "patient_code": str(row[h["patient_code"]]) if row[h["patient_code"]] else "",
            "patient_name": str(row[h["患者名"]]) if row[h["患者名"]] else "",
            "weekday": str(row[h["曜日"]]) if row[h["曜日"]] else "",
            "slot_index": row[h["slot_index"]],
            "mode": str(row[h["モード"]]) if row[h["モード"]] else "",
            "time_type": str(row[h["時間タイプ"]]) if row[h["時間タイプ"]] else "",
            "start_time": str(row[h["開始時刻"]]) if row[h["開始時刻"]] else "",
            "end_time": str(row[h["終了時刻"]]) if row[h["終了時刻"]] else "",
            "duration_min": row[h["duration_min"]],
            "course_template_code": str(row[h["course_template_code"]])
            if row[h["course_template_code"]]
            else "",
            "sub_office_code": str(row[h["sub_office_code"]]) if row[h["sub_office_code"]] else "",
        }
        pfvs.append(rec)
    print(f"PFV records: {len(pfvs)}")

    # 曜日 × コース で grouping
    buckets: dict[tuple[str, str], list[dict]] = {}
    for r in pfvs:
        wd = WEEKDAY_EN_TO_JP.get(r["weekday"], r["weekday"])
        course = r["course_template_code"] or "(未指定)"
        buckets.setdefault((wd, course), []).append(r)

    # 時刻順に並べる
    for v in buckets.values():
        v.sort(key=lambda x: (x["start_time"] or "", x.get("slot_index") or 0))

    # markdown 出力
    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase F-2: 現 DB と User シートの並列比較\n\n")
        f.write(f"- 患者数 (DB): {len(pid_to_meta)}\n")
        f.write(f"- PFV 件数 (DB): {len(pfvs)}\n\n")
        f.write("各 「曜日 × コース」 ごとに、DB に入っている固定枠を時刻順で並べた.\n")
        f.write(
            "User の「スケジュール枠組み (仮)」シートと突合して、抜け / 余分 / 時刻ずれ を確認.\n\n"
        )

        # 曜日順 → コース順
        all_keys = sorted(
            buckets.keys(),
            key=lambda x: (WEEKDAY_ORDER.index(x[0]) if x[0] in WEEKDAY_ORDER else 99, x[1]),
        )
        prev_wd = None
        for wd, course in all_keys:
            if wd != prev_wd:
                f.write(f"\n## {wd}曜日\n\n")
                prev_wd = wd
            f.write(f"### {wd}/{course} ({len(buckets[(wd, course)])} 件)\n\n")
            f.write("| 時刻 | duration | mode | time_type | patient_code | 患者名 |\n")
            f.write("|------|---------|------|-----------|--------------|--------|\n")
            for r in buckets[(wd, course)]:
                f.write(
                    f"| {r['start_time']}–{r['end_time']} | "
                    f"{r['duration_min']}分 | {r['mode']} | {r['time_type']} | "
                    f"{r['patient_code']} | {r['patient_name']} |\n"
                )
            f.write("\n")

    print("\n=== 出力 ===")
    print(f"{OUTPUT_MD}")


if __name__ == "__main__":
    main()
