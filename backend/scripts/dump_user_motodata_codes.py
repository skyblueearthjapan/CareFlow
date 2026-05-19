"""User「元データ」シート全 patient_code リスト + DB との突合."""

from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)
DB_XLSX = Path(__file__).parent / "_current_db_export.xlsx"


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return unicodedata.normalize("NFKC", s) if s else ""


def main() -> None:
    wb = load_workbook(USER_XLSX, data_only=True, read_only=True)
    ws = wb["元データ"]
    user_codes = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            continue
        if not row or row[0] is None:
            continue
        code = norm(row[0])
        name = norm(row[1]) if len(row) > 1 else ""
        if code.startswith("P"):
            user_codes.append((code, name))
    wb.close()
    print(f"User 元データ patients: {len(user_codes)}")

    # DB alive codes
    wb2 = load_workbook(DB_XLSX, data_only=True, read_only=True)
    ws_p = wb2["患者マスタ"]
    header2 = None
    code_idx = name_idx = None
    db_codes = []
    for row in ws_p.iter_rows(values_only=True):
        if header2 is None:
            header2 = list(row)
            for i, h in enumerate(header2):
                if h and str(h).startswith("patient_code"):
                    code_idx = i
                if h and str(h).startswith("患者名"):
                    name_idx = i
            continue
        if not row or code_idx is None:
            continue
        code = norm(row[code_idx])
        name = norm(row[name_idx]) if name_idx is not None else ""
        if code.startswith("P"):
            db_codes.append((code, name))
    wb2.close()
    print(f"DB alive patients: {len(db_codes)}")

    user_set = {c for c, _ in user_codes}
    db_set = {c for c, _ in db_codes}
    print(f"\nUser only (codes): {sorted(user_set - db_set)}")
    print(f"DB only (codes): {sorted(db_set - user_set)}")

    print("\n--- User 元データ 全 code ---")
    for c, n in sorted(user_codes):
        print(f"  {c}: {n}")
    print("\n--- DB alive 全 code ---")
    for c, n in sorted(db_codes):
        in_user = "✓" if c in user_set else "✗"
        print(f"  {c}: {n} [in_user={in_user}]")


if __name__ == "__main__":
    main()
