"""Phase F-3-B-2: 0519 シート → CareFlow Excel テンプレ変換.

入力: Sampledata/2026.05.19段階最新ユーザーシート/スケジュール手動 のコピー.xlsx
出力: backend/scripts/_user_0519_weekly_update.xlsx

User 確定方針:
  「0519 シート 86 名 (= File 2 マスタ 86 名) が正 → 10 名追加」
  → 9 名は soft deleted で残っているので resurrect (deleted_at NULL に戻す).
  → 1 名 (今井) は完全新規 INSERT.

新規 10 名の code 採番 + 拠点割当 (User 確定値):
  P084 朝倉 美夢 (花見川区, INAGE) [resurrect: 既存 soft deleted を復活]
  P085 岡村 敏子 (若葉区, TSUGA) [resurrect]
  P086 河野 天歩 (稲毛区, INAGE) [resurrect]
  P087 清水 政憲 (四街道市, TSUGA) [resurrect]
  P088 中尾 要太 (花見川区, INAGE) [resurrect]
  P089 槇 恵 (稲毛区, INAGE) [resurrect]
  P090 幸 千秋 (美浜区, INAGE) [resurrect]
  P091 渡邉 愛 (中央区, TSUGA) [resurrect]
  P092 菅原 茉結 (若葉区, TSUGA) [resurrect]
  P093 今井 康敦 (若葉区, TSUGA) [真の新規 INSERT]

シート構成:
  1. 患者マスタ: **新規 10 名のみ** (resurrect 9 + new 1).
  2. 固定訪問パターン (PFV): **空** (今回 PFV は触らない).
  3. 希望訪問パターン: **86 名** (= 既存 76 + 新規 10).

突合: 0519 シート患者名 → DB 患者マスタ patient_code (空白除去比較).
"""

from __future__ import annotations

import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Append backend/app sys.path so scripts can import.
BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from app.services.patient_excel.schema import (  # noqa: E402
    PATIENT_COLUMNS,
    PFV_COLUMNS,
    SHEET_PATIENTS,
    SHEET_PFV,
    SHEET_WEEKLY,
    WEEKLY_COLUMNS,
)

USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)
DB_XLSX = Path(__file__).parent / "_current_db_export.xlsx"
OUT_XLSX = Path(__file__).parent / "_user_0519_weekly_update.xlsx"

USER_SHEET = "松岡作業中_マスタヒアリング_0519"

# 新規 10 名 確定マッピング: name (空白除去) → (code, office_code).
# P084-P092 は soft deleted で DB に残っているので resurrect される.
# P093 今井のみ完全新規 INSERT.
NEW_PATIENT_ASSIGNMENT: dict[str, tuple[str, str]] = {
    "朝倉美夢": ("P084", "INAGE"),
    "岡村敏子": ("P085", "TSUGA"),
    "河野天歩": ("P086", "INAGE"),
    "清水政憲": ("P087", "TSUGA"),
    "中尾要太": ("P088", "INAGE"),
    "槇恵": ("P089", "INAGE"),
    "幸千秋": ("P090", "INAGE"),
    "渡邉愛": ("P091", "TSUGA"),
    "菅原茉結": ("P092", "TSUGA"),
    "今井康敦": ("P093", "TSUGA"),
}

# 値変換
SEX_MAP = {"女性": "female", "男性": "male"}
STATUS_MAP = {
    "稼働": "active",
    "休止": "suspended",
    "入院": "admitted",
    "未開始": "pending",
    "解約": "cancelled",
    "未契約": "pending",
}
# User 時間タイプ → DB time_type
TIME_TYPE_MAP = {
    "固定": "固定",
    "時間帯": "時間帯",
    "午前": "午前",
    "午後": "午後",
    "終日": "終日",
}


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return unicodedata.normalize("NFKC", s) if s else ""


def norm_name(v) -> str:
    return norm(v).replace("　", "").replace(" ", "")


def norm_num(v) -> str:
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return str(v).strip()


def norm_time(v) -> str:
    if v is None or v == "":
        return ""
    s = str(v).strip()
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h, m = int(parts[0]), int(parts[1])
            return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    return s


def parse_weekdays(v) -> set[str]:
    """'Mon, Wed, Fri' → {'Mon','Wed','Fri'}."""
    s = norm(v)
    if not s:
        return set()
    return {p.strip() for p in s.replace(";", ",").split(",") if p.strip()}


def hidx(header: list, prefix: str) -> int | None:
    for i, c in enumerate(header):
        if c and str(c).startswith(prefix):
            return i
    return None


def load_db_patient_codes() -> dict[str, str]:
    """DB 患者マスタから (name 空白除去 → patient_code) マップ."""
    wb = load_workbook(DB_XLSX, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header = None
    code_idx = name_idx = None
    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            code_idx = hidx(header, "patient_code")
            name_idx = hidx(header, "患者名")
            continue
        if not row or code_idx is None:
            continue
        code = norm(row[code_idx])
        name = norm(row[name_idx]) if name_idx is not None else ""
        if code.startswith("P") and name:
            out[norm_name(name)] = code
    wb.close()
    return out


def load_0519_rows() -> list[dict]:
    wb = load_workbook(USER_XLSX, data_only=True, read_only=True)
    ws = wb[USER_SHEET]
    header = None
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [norm(c) for c in row]
            continue
        if not row or all(c is None or c == "" for c in row):
            continue
        rec = {h: row[i] for i, h in enumerate(header) if h and i < len(row)}
        if not rec.get(norm("患者名")):
            continue
        rows.append(rec)
    wb.close()
    return rows


def _write_header(ws: Worksheet, cols: list[dict]) -> None:
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col["header"])


def _col_idx(cols: list[dict], key: str) -> int:
    for i, c in enumerate(cols, 1):
        if c["key"] == key:
            return i
    raise KeyError(key)


def build_workbook(rows_0519: list[dict], name_to_code: dict[str, str]) -> Workbook:
    wb = Workbook()
    ws_p: Worksheet = wb.active  # type: ignore[assignment]
    ws_p.title = SHEET_PATIENTS
    _write_header(ws_p, PATIENT_COLUMNS)

    # 患者マスタ: 新規 10 名のみ (resurrect 9 + new 1)
    p_row = 2
    for r in rows_0519:
        name_str = norm(r.get(norm("患者名")))
        if not name_str:
            continue
        nm = norm_name(name_str)
        if nm in name_to_code:
            continue  # 既存 alive → 触らない
        if nm not in NEW_PATIENT_ASSIGNMENT:
            print(f"WARNING: 新規候補だが ASSIGNMENT に無い: {name_str!r}", file=sys.stderr)
            continue
        code, office_code = NEW_PATIENT_ASSIGNMENT[nm]
        sex_jp = norm(r.get(norm("性別")))
        status_jp = norm(r.get(norm("稼働状況")))
        ws_p.cell(row=p_row, column=_col_idx(PATIENT_COLUMNS, "patient_id"), value="")
        ws_p.cell(row=p_row, column=_col_idx(PATIENT_COLUMNS, "patient_code"), value=code)
        ws_p.cell(row=p_row, column=_col_idx(PATIENT_COLUMNS, "name"), value=name_str)
        ws_p.cell(
            row=p_row, column=_col_idx(PATIENT_COLUMNS, "kana"), value=norm(r.get(norm("フリガナ")))
        )
        ws_p.cell(
            row=p_row, column=_col_idx(PATIENT_COLUMNS, "sex"), value=SEX_MAP.get(sex_jp, sex_jp)
        )
        ws_p.cell(
            row=p_row,
            column=_col_idx(PATIENT_COLUMNS, "status"),
            value=STATUS_MAP.get(status_jp, status_jp),
        )
        ws_p.cell(
            row=p_row, column=_col_idx(PATIENT_COLUMNS, "address"), value=norm(r.get(norm("住所")))
        )
        ws_p.cell(row=p_row, column=_col_idx(PATIENT_COLUMNS, "office_code"), value=office_code)
        p_row += 1
    print(f"患者マスタ シート: 新規 {p_row - 2} 行")

    # 固定訪問パターン: header のみ (空)
    ws_f: Worksheet = wb.create_sheet(title=SHEET_PFV)
    _write_header(ws_f, PFV_COLUMNS)
    print("PFV シート: header のみ (空)")

    # 希望訪問パターン: 全 86 名
    ws_w: Worksheet = wb.create_sheet(title=SHEET_WEEKLY)
    _write_header(ws_w, WEEKLY_COLUMNS)
    w_row = 2
    for r in rows_0519:
        name_str = norm(r.get(norm("患者名")))
        if not name_str:
            continue
        nm = norm_name(name_str)
        if nm in name_to_code:
            code = name_to_code[nm]
            patient_id_cell = ""  # 既存
        elif nm in NEW_PATIENT_ASSIGNMENT:
            code, _ = NEW_PATIENT_ASSIGNMENT[nm]
            patient_id_cell = ""  # 新規 (patient_code で resolve される)
        else:
            print(f"WARNING: weekly skip (unmapped): {name_str!r}", file=sys.stderr)
            continue

        freq = norm_num(r.get(norm("週訪問回数")))
        tt_jp = norm(r.get(norm("時間タイプ")))
        svc = norm_num(r.get(norm("サービス時間")))
        ps = norm_time(r.get(norm("希望時間帯(開始)")))
        pe = norm_time(r.get(norm("希望時間帯(終了)")))
        wd_set = parse_weekdays(r.get(norm("希望曜日(複数可)")))

        ws_w.cell(row=w_row, column=_col_idx(WEEKLY_COLUMNS, "patient_id"), value=patient_id_cell)
        ws_w.cell(row=w_row, column=_col_idx(WEEKLY_COLUMNS, "patient_code"), value=code)
        ws_w.cell(row=w_row, column=_col_idx(WEEKLY_COLUMNS, "patient_name"), value=name_str)
        ws_w.cell(
            row=w_row,
            column=_col_idx(WEEKLY_COLUMNS, "frequency_per_week"),
            value=int(freq) if freq.isdigit() else freq,
        )
        # visit_frequency / visit_weeks は今回未指定
        for wd_key, en in [
            ("wd_mon", "Mon"),
            ("wd_tue", "Tue"),
            ("wd_wed", "Wed"),
            ("wd_thu", "Thu"),
            ("wd_fri", "Fri"),
            ("wd_sat", "Sat"),
            ("wd_sun", "Sun"),
        ]:
            ws_w.cell(
                row=w_row,
                column=_col_idx(WEEKLY_COLUMNS, wd_key),
                value="TRUE" if en in wd_set else "FALSE",
            )
        if svc:
            ws_w.cell(
                row=w_row,
                column=_col_idx(WEEKLY_COLUMNS, "service_minutes"),
                value=int(svc) if svc.isdigit() else svc,
            )
        ws_w.cell(
            row=w_row,
            column=_col_idx(WEEKLY_COLUMNS, "time_type"),
            value=TIME_TYPE_MAP.get(tt_jp, tt_jp),
        )
        if ps:
            ws_w.cell(row=w_row, column=_col_idx(WEEKLY_COLUMNS, "preferred_start"), value=ps)
        if pe:
            ws_w.cell(row=w_row, column=_col_idx(WEEKLY_COLUMNS, "preferred_end"), value=pe)
        w_row += 1

    print(f"希望訪問パターン シート: {w_row - 2} 行")
    return wb


def main() -> None:
    name_to_code = load_db_patient_codes()
    rows = load_0519_rows()
    print(f"DB existing patients: {len(name_to_code)}")
    print(f"0519 rows: {len(rows)}")

    wb = build_workbook(rows, name_to_code)
    wb.save(OUT_XLSX)
    print("\n=== 出力 ===")
    print(f"{OUT_XLSX}  ({OUT_XLSX.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
