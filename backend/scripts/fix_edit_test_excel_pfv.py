"""
fix_edit_test_excel_pfv.py
PFV シート（固定訪問パターン）の不整合 11 件を補完して _FIXED.xlsx を出力する。

修正内容:
  - time_type が空 → 「固定」で埋める (3 件)
  - course_template_code を A に修正 (8 件)
"""

import sys
from pathlib import Path

import openpyxl

INPUT = Path(
    r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\.omc\test_data"
    r"\patients_master_2026-05-17_EDITED.xlsx"
)
OUTPUT = Path(
    r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\.omc\test_data"
    r"\patients_master_2026-05-17_EDITED_FIXED.xlsx"
)

# 修正仕様 (Excel 行番号, patient_code, 列名, 旧値, 新値)
FIXES = [
    # time_type が空 → 固定 (3 件)
    (9, "P044", "time_type", None, "固定"),
    (43, "P030", "time_type", None, "固定"),
    (76, "P060", "time_type", None, "固定"),
    # course_template_code → A (8 件)
    (2, "P039", "course_template_code", "C", "A"),
    (27, "P048", "course_template_code", "B", "A"),
    (40, "P087", "course_template_code", "B", "A"),
    (42, "P027", "course_template_code", "D", "A"),
    (59, "P062", "course_template_code", "D", "A"),
    (85, "P041", "course_template_code", "D", "A"),
    (93, "P092", "course_template_code", "B", "A"),
    (136, "P078", "course_template_code", "D", "A"),
]


def find_col(ws, keyword: str) -> int | None:
    """ヘッダー行から keyword を含む列番号 (1-indexed) を返す。"""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and keyword in str(val):
            return col
    return None


def main() -> int:
    print(f"入力: {INPUT}")
    print(f"出力: {OUTPUT}")
    print()

    wb = openpyxl.load_workbook(INPUT)

    # PFV シート = 2 シート目
    if len(wb.sheetnames) < 2:
        print("ERROR: シートが 2 枚以上ありません", file=sys.stderr)
        return 1

    ws = wb.worksheets[1]
    print(f"PFV シート: '{ws.title}'  (max_row={ws.max_row})")
    print()

    # 列位置を特定
    # time_type 列: ヘッダーに「時間タイプ」または「time_type」を含む
    col_time = find_col(ws, "時間タイプ") or find_col(ws, "time_type")
    # course_template_code 列: ヘッダーに「course_template_code」またはコーステンプレを含む
    col_course = find_col(ws, "course_template_code") or find_col(ws, "コーステンプレ")

    if col_time is None:
        print("ERROR: time_type 列が見つかりません", file=sys.stderr)
        return 1
    if col_course is None:
        print("ERROR: course_template_code 列が見つかりません", file=sys.stderr)
        return 1

    # patient_code 列
    col_pcode = find_col(ws, "patient_code")
    if col_pcode is None:
        print("ERROR: patient_code 列が見つかりません", file=sys.stderr)
        return 1

    print(f"  patient_code 列: {col_pcode}")
    print(f"  time_type 列:    {col_time}")
    print(f"  course_template_code 列: {col_course}")
    print()

    col_map = {
        "time_type": col_time,
        "course_template_code": col_course,
    }

    errors = 0
    applied = 0
    print(f"{'Row':>5}  {'patient_code':<14} {'列名':<25} {'旧値':<8} {'新値':<8}")
    print("-" * 65)

    for row, expected_pcode, col_name, old_val, new_val in FIXES:
        col_idx = col_map[col_name]
        actual_pcode = ws.cell(row=row, column=col_pcode).value
        actual_val = ws.cell(row=row, column=col_idx).value

        # patient_code の照合
        if actual_pcode != expected_pcode:
            print(
                f"  ERROR Row {row}: patient_code が一致しない "
                f"(期待={expected_pcode}, 実際={actual_pcode})",
                file=sys.stderr,
            )
            errors += 1
            continue

        # 旧値チェック (time_type=None の場合は None と照合)
        if col_name == "time_type" and old_val is None:
            if actual_val is not None:
                print(
                    f"  WARN  Row {row}: {col_name} が None でない ({actual_val!r}) — 上書きします"
                )
        elif old_val is not None and actual_val != old_val:
            print(
                f"  WARN  Row {row}: {col_name} が期待値 '{old_val}' でない "
                f"({actual_val!r}) — 上書きします"
            )

        ws.cell(row=row, column=col_idx).value = new_val
        print(f"{row:>5}  {actual_pcode:<14} {col_name:<25} {str(actual_val or ''):<8} → {new_val}")
        applied += 1

    print()
    print(f"適用: {applied} 件 / エラー: {errors} 件")

    if errors > 0:
        print("ERROR があったため保存をスキップします", file=sys.stderr)
        return 1

    wb.save(OUTPUT)
    print(f"\n保存完了: {OUTPUT}  ({OUTPUT.stat().st_size:,} bytes)")

    # ---- 再読み込み検証 ----
    print("\n--- 再読み込み検証 ---")
    wb2 = openpyxl.load_workbook(OUTPUT)
    ws2 = wb2.worksheets[1]

    all_ok = True
    print(f"{'Row':>5}  {'patient_code':<14} {'列名':<25} {'期待値':<8} {'実際値':<8} {'結果'}")
    print("-" * 70)

    for row, expected_pcode, col_name, _old_val, new_val in FIXES:
        col_idx = col_map[col_name]
        actual = ws2.cell(row=row, column=col_idx).value
        ok = actual == new_val
        mark = "OK" if ok else "NG"
        if not ok:
            all_ok = False
        print(
            f"{row:>5}  {expected_pcode:<14} {col_name:<25} "
            f"{new_val:<8} {str(actual or ''):<8} {mark}"
        )

    print()
    if all_ok:
        print("検証: 全 11 件 OK")
    else:
        print("検証: NG あり — 再確認してください", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
