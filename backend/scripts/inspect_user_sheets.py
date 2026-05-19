"""User の最新シート 2 ファイルのシート構成と先頭数行を確認."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

SAMPLE_DIR = Path(
    r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata\2026.05.19段階最新ユーザーシート"
)
FILES = sorted(SAMPLE_DIR.glob("*.xlsx"))


def main() -> None:
    for f in FILES:
        if not f.exists():
            print(f"NOT FOUND: {f}")
            continue
        print("=" * 80)
        print(f"FILE: {f.name}")
        print(f"SIZE: {f.stat().st_size:,} bytes")
        try:
            wb = load_workbook(f, data_only=True, read_only=True)
        except Exception as e:
            print(f"  ERROR loading: {type(e).__name__}: {e}")
            continue
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  [Sheet: {sheet_name}]  dimensions={ws.max_row}r × {ws.max_column}c")
            # 先頭 3 行を表示
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
                cells = [str(c)[:30] if c is not None else "" for c in row[:12]]
                print(f"    R{ri}: {cells}")
        wb.close()


if __name__ == "__main__":
    main()
