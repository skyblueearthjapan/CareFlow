"""W15 Phase2 B-1: 「スケジュール枠組み（仮）」シート取込スクリプト.

## 概要

Excel ファイルのシート「スケジュール枠組み（仮）」をパースし、
``course_templates`` と ``patient_fixed_visits`` を upsert する。

## 使い方

::

    # 内容確認のみ (DB 書き込みなし)
    python -m scripts.import_schedule_template --xlsx PATH --dry-run

    # 実際に upsert 実行
    python -m scripts.import_schedule_template --xlsx PATH --apply

--dry-run と --apply は排他必須。

## シート構造

拠点別 × コース (A〜E) × 曜日 (月〜土) × 時間帯 のマトリクス。

各コースブロック (約 22 行):
- Row N:   曜日ヘッダ (col2=月, col7=火, col12=水, col17=木, col22=金, col27=土)
- Row N+1: 件数 (col1,6,11,16,21,26) + コースラベル (col2,7,12,17,22,27)
           件数欄に「稲毛」「都賀」= 当該拠点ではこの曜日は運用なし (capacity=0)
- Row N+2: フィールドヘッダ (時間帯/氏名/住所/複数/条件)
- Row N+3〜: データ行 (time|name|address|companion|condition)

第 2 拠点 (都賀) では、曜日ヘッダ行 (Row N) にも件数数値が入り、
コースラベル行 (Row N+1) の件数欄が「都賀」マーカーになる。
その場合、件数は曜日ヘッダ行側から取得する。

## 注意

- 本スクリプトは Phase 7 で本番 VPS 上で実行予定。
- xlsx は read-only。元ファイルは変更しない。
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import time
from pathlib import Path
from typing import Any, Literal

import openpyxl
import sqlalchemy as sa

# ---------------------------------------------------------------------------
# sys.path 設定
# ---------------------------------------------------------------------------
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

_BACKEND_ROOT = _SCRIPTS_DIR.parent
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))

from sqlalchemy import select  # noqa: E402

from app.db.session import dispose_engine, get_session_factory  # noqa: E402
from app.models.course_template import CourseTemplate  # noqa: E402
from app.models.office import Office  # noqa: E402
from app.models.patient import Patient  # noqa: E402
from app.models.patient_fixed_visit import PatientFixedVisit  # noqa: E402

# ---------------------------------------------------------------------------
# ロガー
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
SHEET_NAME = "スケジュール枠組み（仮）"

# 曜日ヘッダ行の識別トークン
_WEEKDAY_COL2_TOKEN = "月曜日"

# コースラベル列 (0-indexed columns): B(2), G(7), L(12), Q(17), V(22), AA(27)
# ※ openpyxl は 1-indexed なので実際は col 2,7,12,17,22,27
_CAP_COLS = [1, 6, 11, 16, 21, 26]  # A, F, K, P, U, Z (1-indexed)
_LABEL_COLS = [2, 7, 12, 17, 22, 27]  # B, G, L, Q, V, AA (1-indexed)

# 5 列ずつ繰り返す各曜日セクション: (time_col, name_col, addr_col, companion_col, cond_col)
_DAY_COLUMNS = [
    (1, 2, 3, 4, 5),  # 月曜日
    (6, 7, 8, 9, 10),  # 火曜日
    (11, 12, 13, 14, 15),  # 水曜日
    (16, 17, 18, 19, 20),  # 木曜日
    (21, 22, 23, 24, 25),  # 金曜日
    (26, 27, 28, 29, 30),  # 土曜日
]
WEEKDAY_INDEX = {0: "mon", 1: "tue", 2: "wed", 3: "thu", 4: "fri", 5: "sat"}

# 拠点判定用キーワード (住所の区名)
_OFFICE1_DISTRICTS = frozenset(["稲毛区", "花見川区", "美浜区"])
_OFFICE2_DISTRICTS = frozenset(["中央区", "若葉区"])

# 「運用なし」マーカー文字列
_NOT_AVAIL_TOKENS = frozenset(["稲毛", "都賀"])

# 「空きスロット」マーカー
_EMPTY_SLOT_TOKENS = frozenset(["空"])

# 既定 duration (30 分)
DEFAULT_DURATION_MIN = 30


# ---------------------------------------------------------------------------
# データクラス
# ---------------------------------------------------------------------------


@dataclass
class VisitSlot:
    """1 訪問スロット (データ行の 1 セル群)."""

    weekday_idx: int  # 0=Mon … 5=Sat
    start_time: time
    name: str
    address: str | None
    companion: bool  # 「複数」フラグ
    condition: str | None  # 「条件」テキスト


@dataclass
class CourseBlock:
    """1 コースブロックのパース結果."""

    label: str  # "A" 〜 "E"
    capacities: dict[str, int]  # {"mon": 6, "tue": 5, ...}
    visits: list[VisitSlot]  # スロット一覧
    notes_set: set[str]  # 条件テキストの集合 (course_template.notes 用)


@dataclass
class OfficeSection:
    """1 拠点セクションのパース結果."""

    office_hint: str | None  # "稲毛" / "都賀" / None (住所推定)
    course_blocks: list[CourseBlock]
    addresses: list[str]  # 住所一覧 (拠点推定用)


# ---------------------------------------------------------------------------
# 集計
# ---------------------------------------------------------------------------


@dataclass
class MatchedVisit:
    patient_id: object  # uuid.UUID
    patient_name: str
    weekday_idx: int
    start_time: time
    label: str
    office_hint: str | None


@dataclass
class UnmatchedVisit:
    name: str
    address: str | None
    weekday_idx: int
    start_time: time
    label: str
    office_hint: str | None
    reason: str


@dataclass
class ImportSummary:
    office_infos: list[tuple[str, object, list[str]]] = field(default_factory=list)
    # (office_name, office_id, [labels])

    ct_new: int = 0
    ct_update: int = 0
    pfv_new: int = 0
    pfv_update: int = 0

    matched: list[MatchedVisit] = field(default_factory=list)
    unmatched: list[UnmatchedVisit] = field(default_factory=list)
    companion_warnings: list[str] = field(default_factory=list)
    condition_notes: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Excel パーサー
# ---------------------------------------------------------------------------


def _cell_str(ws: Any, row: int, col: int) -> str | None:
    """セル値を文字列として返す (None or 空文字なら None)."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _cell_val(ws: Any, row: int, col: int) -> Any:
    """セル生値を返す."""
    return ws.cell(row=row, column=col).value


def _parse_cap(v: Any) -> int:
    """件数セル値を整数に変換。'稲毛'/'都賀' など非数値は 0。"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s in _NOT_AVAIL_TOKENS:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _parse_time(v: Any) -> time | None:
    """セル値を datetime.time に変換。"""
    if isinstance(v, time):
        return v
    if v is None:
        return None
    s = str(v).strip()
    # "09:30:00" 形式
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            return time(int(parts[0]), int(parts[1]))
        except ValueError:
            pass
    return None


def _is_weekday_header_row(ws: Any, row: int) -> bool:
    """行が曜日ヘッダ行 (col2 == '月曜日') かどうか判定。"""
    v = _cell_str(ws, row, 2)
    return v == _WEEKDAY_COL2_TOKEN


def _is_block_row(ws: Any, row: int) -> bool:
    """行がコースラベル行 (col2 が単一英字 A-E) かどうか判定。"""
    v = ws.cell(row=row, column=2).value
    return isinstance(v, str) and v.strip() in "ABCDE" and len(v.strip()) == 1


def _parse_course_block(ws: Any, weekday_row: int, block_row: int) -> CourseBlock:
    """曜日ヘッダ行 + コースラベル行 から CourseBlock を生成。

    - weekday_row: 曜日ヘッダ行 (col2='月曜日')
    - block_row  : 件数 + コースラベル行 (col2='A'〜'E')
    """
    label = str(ws.cell(row=block_row, column=2).value).strip()

    capacities: dict[str, int] = {}
    for day_idx, (cap_col, _label_col) in enumerate(zip(_CAP_COLS, _LABEL_COLS, strict=True)):
        # コースラベル行の件数欄が非数値 (稲毛/都賀) のとき、
        # 曜日ヘッダ行の同位置に実際の件数が入っていることがある (第2拠点形式)。
        block_cap_val = _cell_val(ws, block_row, cap_col)
        if isinstance(block_cap_val, (int, float)):
            cap = int(block_cap_val)
        else:
            # 第2拠点: 曜日ヘッダ行から取得
            header_cap_val = _cell_val(ws, weekday_row, cap_col)
            cap = _parse_cap(header_cap_val)

        capacities[WEEKDAY_INDEX[day_idx]] = cap

    # データ行: field_header = block_row + 1, data starts at block_row + 2
    data_start_row = block_row + 2

    visits: list[VisitSlot] = []
    notes_set: set[str] = set()

    # データ行の末尾を探す: 次の空白行か次の曜日ヘッダ行まで
    row = data_start_row
    while row <= ws.max_row:
        # 終端判定: 6 列すべて None → ブロック終端
        all_none = all(ws.cell(row=row, column=tc).value is None for tc, *_ in _DAY_COLUMNS)
        if all_none:
            break
        # 次の曜日ヘッダ行に達したら終了
        if _is_weekday_header_row(ws, row):
            break

        for day_idx, (tc, nc, ac, cc, cond_c) in enumerate(_DAY_COLUMNS):
            time_val = _cell_val(ws, row, tc)
            start = _parse_time(time_val)
            if start is None:
                continue

            name_raw = _cell_str(ws, row, nc)
            if name_raw is None or name_raw in _EMPTY_SLOT_TOKENS:
                continue

            address = _cell_str(ws, row, ac)
            companion_raw = _cell_str(ws, row, cc)
            companion = companion_raw is not None and companion_raw not in {"", "なし"}
            condition = _cell_str(ws, row, cond_c)

            if condition:
                notes_set.add(condition)

            visits.append(
                VisitSlot(
                    weekday_idx=day_idx,
                    start_time=start,
                    name=name_raw,
                    address=address,
                    companion=companion,
                    condition=condition,
                )
            )

        row += 1

    return CourseBlock(
        label=label,
        capacities=capacities,
        visits=visits,
        notes_set=notes_set,
    )


def _has_double_blank_before(ws: Any, target_row: int) -> bool:
    """target_row の直前に 2 行連続の空行があるか判定 (拠点セクション区切り検出用)."""

    def _row_is_blank(r: int) -> bool:
        return all(ws.cell(row=r, column=c).value is None for c in range(1, 31))

    # target_row-1 と target_row-2 がどちらも空行
    return target_row >= 3 and _row_is_blank(target_row - 1) and _row_is_blank(target_row - 2)


def parse_sheet(xlsx_path: Path) -> list[OfficeSection]:
    """xlsx を開いてシートをパースし、OfficeSection のリストを返す。"""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_NAME}」が見つかりません。シート一覧: {wb.sheetnames}")
    # read_only モードではランダムアクセス不可なので通常モードで再読込
    wb.close()
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[SHEET_NAME]

    # 全行をスキャンしてブロック開始行 (コースラベル行) を特定
    block_rows: list[int] = []
    for row_idx in range(1, ws.max_row + 1):
        if _is_block_row(ws, row_idx):
            block_rows.append(row_idx)

    if not block_rows:
        raise ValueError("コースブロックが見つかりません。シート構造を確認してください。")

    # 各ブロックに対応する曜日ヘッダ行を特定 (block_row の 1〜2 行前)
    sections: list[OfficeSection] = []
    current_section_blocks: list[CourseBlock] = []
    current_office_hint: str | None = None
    current_addresses: list[str] = []

    prev_block_row: int = 0

    for br in block_rows:
        # 曜日ヘッダ行を探す: br-1 か br-2
        wday_row: int = br - 1
        for candidate in [br - 1, br - 2]:
            if candidate >= 1 and _is_weekday_header_row(ws, candidate):
                wday_row = candidate
                break

        # 新しいセクション開始の判定:
        # 曜日ヘッダ行の直前に 2 連続空行がある場合 = 新しい拠点セクション
        # (コース間区切りは 1 空行のみ; 拠点間は 2 連続空行)
        is_new_section = prev_block_row > 0 and _has_double_blank_before(ws, wday_row)

        if is_new_section and current_section_blocks:
            sections.append(
                OfficeSection(
                    office_hint=current_office_hint,
                    course_blocks=current_section_blocks,
                    addresses=list(current_addresses),
                )
            )
            current_section_blocks = []
            current_office_hint = None
            current_addresses = []

        block = _parse_course_block(ws, wday_row, br)
        current_section_blocks.append(block)

        # 住所を蓄積 (拠点推定用)
        for slot in block.visits:
            if slot.address and "県" in slot.address:
                current_addresses.append(slot.address)

        # 拠点ヒント: コースラベル行の件数欄が '都賀' ならヒント = '都賀'
        block_cap_col1 = _cell_val(ws, br, _CAP_COLS[0])
        if isinstance(block_cap_col1, str) and block_cap_col1.strip() == "都賀":
            current_office_hint = "都賀"

        prev_block_row = br

    if current_section_blocks:
        sections.append(
            OfficeSection(
                office_hint=current_office_hint,
                course_blocks=current_section_blocks,
                addresses=list(current_addresses),
            )
        )

    wb.close()
    return sections


# ---------------------------------------------------------------------------
# 拠点推定
# ---------------------------------------------------------------------------


def infer_office_name(section: OfficeSection) -> str:
    """住所の区名頻度と office_hint から拠点名を推定。"""
    if section.office_hint == "都賀":
        return "都賀"

    district_counts: dict[str, int] = {}
    for addr in section.addresses:
        for d in _OFFICE1_DISTRICTS | _OFFICE2_DISTRICTS:
            if d in addr:
                district_counts[d] = district_counts.get(d, 0) + 1

    office1_score = sum(district_counts.get(d, 0) for d in _OFFICE1_DISTRICTS)
    office2_score = sum(district_counts.get(d, 0) for d in _OFFICE2_DISTRICTS)

    if office2_score > office1_score:
        return "都賀"
    return "宮野木"


# ---------------------------------------------------------------------------
# DB アクセス
# ---------------------------------------------------------------------------


async def _resolve_office_for_section(
    session: Any,
    section_kind: Literal["main", "tsuga"],
    *,
    apply: bool,
) -> object:
    """セクションキーから offices を解決する.

    main (宮野木): name に「稲毛」「宮野木」「本店」のいずれかを含む
    tsuga (都賀): name に「都賀」を含む

    複数パターンを順に試行し、最初にヒットした office を採用する。
    apply=True 時に未登録なら具体的なパターンを示して SystemExit(1) を raise する。
    """
    if section_kind == "main":
        patterns = ["%稲毛%", "%宮野木%", "%本店%"]
        section_label = "宮野木 (main)"
    elif section_kind == "tsuga":
        patterns = ["%都賀%"]
        section_label = "都賀 (tsuga)"
    else:
        return None

    for pattern in patterns:
        result = await session.scalars(
            select(Office)
            .where(
                Office.name.ilike(pattern),
                Office.deleted_at.is_(None),
            )
            .order_by(Office.created_at.asc())
        )
        office = result.first()
        if office is not None:
            return office

    # 未登録
    patterns_str = " / ".join(f"「{p}」" for p in patterns)
    if not apply:
        logger.warning(
            "拠点 %s が DB に存在しません (検索 pattern: %s) (--dry-run のため INSERT しません)",
            section_label,
            patterns_str,
        )
        return None

    # 拠点が未登録の場合は自動 INSERT せずに終了する。
    # 自動 INSERT した拠点は address/lat/lng が NULL となり、
    # アロケーションエンジンが座標を必要とする際に失敗するため、事前登録を強制する。
    print(
        f"[ERROR] 拠点 {section_label} が offices テーブルに未登録です。\n"
        f"        {patterns_str} のいずれかを name に含む office を"
        f" offices テーブルに事前登録してから再実行してください。",
        file=sys.stderr,
    )
    raise SystemExit(1)


async def _upsert_course_template(
    session: Any,
    office_id: object,
    block: CourseBlock,
    *,
    apply: bool,
) -> tuple[str, bool, CourseTemplate | None]:
    """CourseTemplate を upsert し ('new'|'update', changed, template) を返す。

    W22 Phase A: 取込済みの ``CourseTemplate`` を呼び出し側に返し、
    後段で ``patient_fixed_visits.course_template_id`` を埋めるために使用する。
    dry-run 時 / 既存なし時は ``None`` を返す可能性がある (apply=True の
    新規作成では INSERT 直後の row を返す)。
    """
    stmt = select(CourseTemplate).where(
        CourseTemplate.office_id == office_id,
        CourseTemplate.label == block.label,
        CourseTemplate.deleted_at.is_(None),
    )
    result = await session.scalars(stmt)
    existing = result.first()

    notes = "; ".join(sorted(block.notes_set)) if block.notes_set else None

    if existing is None:
        if apply:
            ct = CourseTemplate(
                office_id=office_id,
                label=block.label,
                capacity_mon=block.capacities.get("mon", 0),
                capacity_tue=block.capacities.get("tue", 0),
                capacity_wed=block.capacities.get("wed", 0),
                capacity_thu=block.capacities.get("thu", 0),
                capacity_fri=block.capacities.get("fri", 0),
                capacity_sat=block.capacities.get("sat", 0),
                notes=notes,
            )
            session.add(ct)
            await session.flush()
            return "new", True, ct
        # dry-run: template は INSERT されないため None を返す.
        return "new", True, None

    # UPDATE
    changed = False
    updates = {
        "capacity_mon": block.capacities.get("mon", 0),
        "capacity_tue": block.capacities.get("tue", 0),
        "capacity_wed": block.capacities.get("wed", 0),
        "capacity_thu": block.capacities.get("thu", 0),
        "capacity_fri": block.capacities.get("fri", 0),
        "capacity_sat": block.capacities.get("sat", 0),
        "notes": notes,
    }
    for attr, val in updates.items():
        if getattr(existing, attr) != val:
            changed = True
            if apply:
                setattr(existing, attr, val)

    return "update", changed, existing


# 漢字違い (読み同じ) でフリガナ列も無いケースの明示エイリアス。
# シート「氏名」表記 → patient.name (DB に存在する正式表記).
# 2026-05-15 imaizumi 確認済みの個別マッピング。
# 追加が必要になったらここに足す (シート側で表記揃える方が望ましい)。
_NAME_ALIASES: dict[str, str] = {
    "川上もえぎ": "川上　萌黄",  # P021
    "久松由香里": "久松　由伽里",  # P051
}


async def _find_patient(
    session: Any,
    name: str,
    address: str | None,
) -> Patient | None:
    """患者を氏名+住所で検索 (name完全一致 → エイリアス → 空白除去 → 住所部分一致)。"""
    # 1. name 完全一致
    stmt = select(Patient).where(
        Patient.name == name,
        Patient.deleted_at.is_(None),
    )
    result = await session.scalars(stmt)
    patients = list(result.all())
    if len(patients) == 1:
        return patients[0]
    if len(patients) > 1 and address:
        # 複数ヒット → 住所で絞り込む
        for p in patients:
            if p.address and address and _addr_match(p.address, address):
                return p
        return patients[0]  # 絞り込めなければ先頭

    # 1.5. 明示エイリアス (漢字違いの個別対応)
    aliased = _NAME_ALIASES.get(name)
    if aliased:
        stmt_alias = select(Patient).where(
            Patient.name == aliased,
            Patient.deleted_at.is_(None),
        )
        p_alias = (await session.scalars(stmt_alias)).first()
        if p_alias:
            return p_alias

    # 2. シート側スペースを除去した name で再検索
    name_compact = name.replace("　", "").replace(" ", "")
    if name_compact != name:
        stmt2 = select(Patient).where(
            Patient.name == name_compact,
            Patient.deleted_at.is_(None),
        )
        result2 = await session.scalars(stmt2)
        p2 = result2.first()
        if p2:
            return p2

    # 2.5. DB 側 name の空白を除去した形がシート name と一致
    #      (「遠藤孝子」 ←→ DB「遠藤　孝子」のような姓名スペース無しフル表記の救済)
    stmt_db_compact = select(Patient).where(
        Patient.deleted_at.is_(None),
        sa.func.replace(sa.func.replace(Patient.name, "　", ""), " ", "") == name,
    )
    p_dc = (await session.scalars(stmt_db_compact)).first()
    if p_dc:
        return p_dc

    # 3. 住所部分一致 (address が十分に長い場合のみ)
    if address and len(address) >= 10:
        stmt3 = select(Patient).where(
            Patient.deleted_at.is_(None),
            Patient.address.contains(address[:15]),  # type: ignore[attr-defined]
        )
        result3 = await session.scalars(stmt3)
        addr_candidates = list(result3.all())
        if len(addr_candidates) == 1:
            return addr_candidates[0]

    return None


def _addr_match(db_addr: str, sheet_addr: str) -> bool:
    """住所の先頭 20 文字が一致するか判定。"""
    return db_addr[:20] == sheet_addr[:20]


# シート「氏名」セルの複合表記分割パターン
# 例: 「槇・河野」「藤田、渡辺」「川上,親子」を ["槇","河野"] / ["藤田","渡辺"] / ["川上","親子"] に分割
_COMPOUND_SEP_RE = re.compile(r"[・、,]")

# 末尾 suffix: 「○○親子」「○○双子」 → 苗字単独に正規化
_GROUP_SUFFIX_RE = re.compile(r"(親子|双子|兄弟|姉妹|姉弟|兄妹|夫妻|夫婦)$")


def _split_compound_name(name: str) -> tuple[list[str], bool]:
    """シート上の氏名表記を patient.name の苗字相当 token に分解。

    Returns:
        (tokens, expand_to_all_with_same_surname)
        - tokens: 分解された苗字相当の token 一覧
        - expand_to_all_with_same_surname: True なら「同苗字+同住所」の全 patient
          に展開する (「○○親子」「○○双子」のような表記が含まれる場合)
    """
    raw_parts = [p.strip() for p in _COMPOUND_SEP_RE.split(name) if p.strip()]
    tokens: list[str] = []
    expand = False
    for part in raw_parts:
        m = _GROUP_SUFFIX_RE.search(part)
        if m:
            expand = True
            stripped = _GROUP_SUFFIX_RE.sub("", part).strip()
            if stripped:
                tokens.append(stripped)
        else:
            tokens.append(part)
    return tokens, expand


async def _find_patients_by_surname(
    session: Any,
    surname: str,
    sheet_address: str | None,
) -> list[Patient]:
    """苗字 (氏名の先頭トークン) 一致で patient 候補を取得し、住所で絞り込む。

    patient.name は「青栁　あい」のように姓と名の間に全角/半角スペースが入る
    前提。surname がそれらの先頭トークンと一致するレコードを返す。
    """
    stmt = select(Patient).where(
        Patient.deleted_at.is_(None),
        sa.or_(
            Patient.name.startswith(surname + "　"),  # 全角スペース
            Patient.name.startswith(surname + " "),
            Patient.name == surname,  # 苗字単独 (まれ)
        ),
    )
    candidates = list((await session.scalars(stmt)).all())
    if len(candidates) <= 1 or not sheet_address:
        return candidates

    # 住所部分一致で絞り込み (シート住所は施設略称「よりより宮野木」等の場合あり)
    narrowed: list[Patient] = []
    for p in candidates:
        if p.address and _addr_overlap(p.address, sheet_address):
            narrowed.append(p)
    return narrowed if narrowed else candidates


def _addr_overlap(db_addr: str, sheet_addr: str) -> bool:
    """db_addr と sheet_addr の重なりがあるか緩めに判定。

    施設略称 (「よりより宮野木」「グループホームCOCOL」等) と DB 上のフル住所を
    マッチさせるため、`sheet_addr` から非ノイズ文字 4 文字以上の部分文字列が
    db_addr に含まれるかでチェック。
    """
    if not sheet_addr:
        return False
    s = sheet_addr.strip()
    if len(s) >= 4 and s in db_addr:
        return True
    # 4-gram 一致をいくつか拾えるか
    grams_db = {db_addr[i : i + 4] for i in range(len(db_addr) - 3)} if len(db_addr) >= 4 else set()
    hits = sum(1 for i in range(max(0, len(s) - 3)) if s[i : i + 4] in grams_db)
    return hits >= 2


async def _find_patients_multi(
    session: Any,
    name: str,
    address: str | None,
) -> list[Patient]:
    """複合表記対応の patient 探索。

    1) まず単一 patient とみなして既存 `_find_patient` を試行
    2) ヒットしなければ複合表記として分割し、各 token を苗字検索
    3) 「○○親子」等の suffix がある場合は同苗字+同住所の全 patient に展開
    """
    single = await _find_patient(session, name, address)
    if single is not None:
        return [single]

    tokens, expand = _split_compound_name(name)
    if not tokens:
        return []

    if expand and len(tokens) == 1:
        # 「遠藤親子」「前川双子」のようなケース。住所マッチで複数 patient 取得
        return await _find_patients_by_surname(session, tokens[0], address)

    if len(tokens) == 1:
        # 苗字単独 (「園田」のみ等)。住所マッチで絞った 1 名想定
        return await _find_patients_by_surname(session, tokens[0], address)

    # 複合表記 (例: 「槇・河野」「藤田・渡辺」)
    collected: list[Patient] = []
    seen_ids: set[Any] = set()
    for token in tokens:
        cands = await _find_patients_by_surname(session, token, address)
        if not cands:
            continue
        chosen = cands[0]  # 住所絞り込み済の先頭を採用
        if chosen.id not in seen_ids:
            collected.append(chosen)
            seen_ids.add(chosen.id)
    return collected


async def _upsert_patient_fixed_visit(
    session: Any,
    patient_id: object,
    weekday_idx: int,
    start_time: time,
    *,
    apply: bool,
    course_template_id: object | None = None,
) -> tuple[str, bool]:
    """PatientFixedVisit を upsert し ('new'|'update', changed) を返す。

    W22 Phase A: ``course_template_id`` (Excel のコースセクションに対応する
    course_templates.id) を保存する。``None`` の場合は既存値を変更しない。
    """
    # NOTE: PatientFixedVisit モデルに deleted_at 列は存在しないため、
    # deleted_at IS NULL フィルタは省略している。
    # 論理削除が将来追加された場合は .deleted_at.is_(None) を追加すること。
    stmt = select(PatientFixedVisit).where(
        PatientFixedVisit.patient_id == patient_id,
        PatientFixedVisit.mode == "normal",
        PatientFixedVisit.weekday == weekday_idx,
    )
    result = await session.scalars(stmt)
    existing = result.first()

    if existing is None:
        if apply:
            pfv = PatientFixedVisit(
                patient_id=patient_id,
                mode="normal",
                weekday=weekday_idx,
                start_time=start_time,
                duration_min=DEFAULT_DURATION_MIN,
                course_template_id=course_template_id,
            )
            session.add(pfv)
            await session.flush()
        return "new", True

    start_time_changed = existing.start_time != start_time
    # W22 Phase A: course_template_id が指定されており既存値と異なれば差分扱い.
    ct_changed = (
        course_template_id is not None and existing.course_template_id != course_template_id
    )

    if apply:
        if start_time_changed:
            existing.start_time = start_time
        if ct_changed:
            existing.course_template_id = course_template_id

    changed = start_time_changed or ct_changed
    return "update", changed


# ---------------------------------------------------------------------------
# メイン処理
# ---------------------------------------------------------------------------


async def run_import(
    xlsx_path: Path,
    *,
    apply: bool,
) -> ImportSummary:
    """xlsx をパース→DB upsert して ImportSummary を返す。"""
    summary = ImportSummary()

    # --- パース ---
    sections = parse_sheet(xlsx_path)

    # --- DB セッション ---
    factory = get_session_factory()
    async with factory() as session:
        for section in sections:
            office_name = infer_office_name(section)

            section_kind: Literal["main", "tsuga"] = "tsuga" if office_name == "都賀" else "main"
            office = await _resolve_office_for_section(session, section_kind, apply=apply)
            if office is None:
                # dry-run で拠点未登録 → スキップ
                if section_kind == "main":
                    patterns_hint = "%稲毛% / %宮野木% / %本店%"
                else:
                    patterns_hint = "%都賀%"
                summary.errors.append(
                    f"拠点「{office_name}」が DB に未登録 (検索 pattern: {patterns_hint})。"
                    f"offices テーブルに事前登録してから --apply で再実行してください。"
                )
                summary.office_infos.append(
                    (office_name, None, [b.label for b in section.course_blocks])
                )
                continue

            office_id = office.id
            labels = [b.label for b in section.course_blocks]
            summary.office_infos.append((office_name, office_id, labels))

            for block in section.course_blocks:
                # --- CourseTemplate upsert ---
                ct_status, _, ct_obj = await _upsert_course_template(
                    session, office_id, block, apply=apply
                )
                if ct_status == "new":
                    summary.ct_new += 1
                else:
                    summary.ct_update += 1

                # W22 Phase A: 当該 block の course_template.id を取得.
                # apply=True かつ既存 / 新規 CT があれば id を渡す。
                # apply=False (dry-run) で既存なし時は ct_obj=None なので、
                # course_template_id は付かない (= 旧挙動と等価).
                ct_id = ct_obj.id if ct_obj is not None else None

                # --- PatientFixedVisit upsert ---
                for slot in block.visits:
                    if slot.companion:
                        summary.companion_warnings.append(
                            f"複数訪問: {slot.name} ({office_name}-{block.label} {['月', '火', '水', '木', '金', '土'][slot.weekday_idx]})"
                        )

                    if slot.condition:
                        summary.condition_notes.append(f"{slot.name}: {slot.condition}")

                    patients = await _find_patients_multi(session, slot.name, slot.address)
                    if not patients:
                        summary.unmatched.append(
                            UnmatchedVisit(
                                name=slot.name,
                                address=slot.address,
                                weekday_idx=slot.weekday_idx,
                                start_time=slot.start_time,
                                label=block.label,
                                office_hint=section.office_hint,
                                reason="患者マスタに見つかりません",
                            )
                        )
                        continue

                    for patient in patients:
                        pfv_status, _ = await _upsert_patient_fixed_visit(
                            session,
                            patient.id,
                            slot.weekday_idx,
                            slot.start_time,
                            apply=apply,
                            course_template_id=ct_id,
                        )
                        summary.matched.append(
                            MatchedVisit(
                                patient_id=patient.id,
                                patient_name=patient.name,
                                weekday_idx=slot.weekday_idx,
                                start_time=slot.start_time,
                                label=block.label,
                                office_hint=section.office_hint,
                            )
                        )
                        if pfv_status == "new":
                            summary.pfv_new += 1
                        else:
                            summary.pfv_update += 1

        if apply:
            await session.commit()
        else:
            await session.rollback()

    return summary


# ---------------------------------------------------------------------------
# 出力フォーマット
# ---------------------------------------------------------------------------


def _print_summary(summary: ImportSummary, *, apply: bool) -> None:
    mode_label = "APPLY" if apply else "DRY-RUN"
    print(f"=== スケジュール枠組み 取込結果 ({mode_label}) ===")
    print()

    print("[拠点認識]")
    for office_name, office_id, labels in summary.office_infos:
        id_str = str(office_id) if office_id else "未登録"
        labels_str = ", ".join(labels)
        print(f"- {office_name}拠点 (id: {id_str}): {labels_str} ({len(labels)} コース)")
    print()

    print("[course_templates upsert]")
    print(f"- NEW={summary.ct_new}, UPDATE={summary.ct_update}")
    print()

    print("[patient_fixed_visits upsert]")
    print(f"- 突合成功: {len(summary.matched)} 件")
    print(f"- 突合失敗: {len(summary.unmatched)} 件")
    if summary.unmatched:
        print("  失敗詳細:")
        for u in summary.unmatched[:20]:
            wd = ["月", "火", "水", "木", "金", "土"][u.weekday_idx]
            addr_str = u.address[:30] if u.address else "(住所なし)"
            print(f"    - {u.name} ({wd} {u.start_time.strftime('%H:%M')}) addr={addr_str}")
        if len(summary.unmatched) > 20:
            print(f"    ... 他 {len(summary.unmatched) - 20} 件")
    print()

    print("[notes]")
    print(f"- 複数訪問: {len(summary.companion_warnings)} 件")
    for w in summary.companion_warnings:
        print(f"  WARNING: {w}")
    print(f"- 条件付き: {len(summary.condition_notes)} 件")
    for c in summary.condition_notes:
        print(f"  NOTE: {c}")
    print()

    print("[Summary]")
    print(f"- course_templates: NEW={summary.ct_new}, UPDATE={summary.ct_update}")
    print(f"- patient_fixed_visits: NEW={summary.pfv_new}, UPDATE={summary.pfv_update}")
    print(f"- 警告: {len(summary.unmatched)} 件")

    if summary.errors:
        print()
        print("[エラー]", file=sys.stderr)
        for e in summary.errors:
            print(f"  ERROR: {e}", file=sys.stderr)

    if not apply:
        print()
        print("※ --dry-run モード: DB への書き込みは行いませんでした。")
        print("   実際に適用するには --apply を指定してください。")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "「スケジュール枠組み（仮）」シートから course_templates / patient_fixed_visits を取込む。\n"
            "--dry-run か --apply のどちらか一方を必ず指定すること。"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--xlsx",
        type=Path,
        required=True,
        help="取込元 xlsx ファイルのパス (必須)",
    )
    grp = p.add_mutually_exclusive_group(required=True)
    grp.add_argument(
        "--dry-run",
        dest="dry_run",
        action="store_true",
        help="内容と差分を表示するのみ。DB に書き込まない。",
    )
    grp.add_argument(
        "--apply",
        dest="apply",
        action="store_true",
        help="実際に upsert を実行して commit する。",
    )
    return p


async def _main(xlsx_path: Path, apply: bool) -> int:
    logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")
    try:
        summary = await run_import(xlsx_path, apply=apply)
    finally:
        await dispose_engine()

    _print_summary(summary, apply=apply)
    return 1 if summary.errors else 0


def main() -> int:
    args = _build_parser().parse_args()
    return asyncio.run(_main(xlsx_path=args.xlsx, apply=args.apply))


if __name__ == "__main__":
    raise SystemExit(main())
