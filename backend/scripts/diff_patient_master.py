"""Phase F-3-A: User シート「元データ」 vs 現 DB の patient マスタ diff.

入力:
  - User: Sampledata/2026.05.19段階最新ユーザーシート/スケジュール手動 のコピー.xlsx
    シート: 元データ (列: patient_id, 患者名, フリガナ, 稼働状況, 性別, 住所,
              緯度, 経度, エリア, 週訪問回数, 訪問頻度, 訪問週)
  - DB: backend/scripts/_current_db_export.xlsx
    シート: 患者マスタ (CareFlow Excel テンプレート列)

出力: backend/scripts/_diff_patient_master.md

突合キー: patient_code (P001, P002, ...).

レポート構成:
  1. 共通 code で属性が異なる行 (= 修正候補)
  2. User のみ (DB に無い code) = 新規追加候補
  3. DB のみ (User にない code) = 削除候補
  4. サマリ件数表
"""

from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
DB_XLSX = SCRIPT_DIR / "_current_db_export.xlsx"
USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)
OUTPUT_MD = SCRIPT_DIR / "_diff_patient_master.md"


def norm(v) -> str:
    """文字列正規化: NFKC + strip + 全角空白統一."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    return s


# DB は CareFlow の enum 値 (female/male/active/inactive) で保存. User シートは日本語.
# 同義扱いマッピングで偽 diff を消す.
SEX_SYNONYM = {
    "女性": "female",
    "男性": "male",
    "female": "female",
    "male": "male",
    "F": "female",
    "M": "male",
}
STATUS_SYNONYM = {
    "稼働": "active",
    "休止": "inactive",
    "active": "active",
    "inactive": "inactive",
}


def norm_sex(v) -> str:
    return SEX_SYNONYM.get(norm(v), norm(v))


def norm_status(v) -> str:
    return STATUS_SYNONYM.get(norm(v), norm(v))


def norm_int(v) -> str:
    """数値正規化: 3.0 → '3' 等."""
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return str(v).strip()


def load_user_patients() -> dict[str, dict]:
    """User シート「元データ」 から patient_code → dict を返す."""
    wb = load_workbook(USER_XLSX, data_only=True, read_only=True)
    ws = wb["元データ"]
    header = None
    rows: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [norm(c) for c in row]
            continue
        if not row or row[0] is None:
            continue
        code = norm(row[0])
        if not code or not code.startswith("P"):
            continue
        rec = {}
        for i, h in enumerate(header):
            if h and i < len(row):
                rec[h] = row[i]
        rows[code] = rec
    wb.close()
    return rows


def load_db_patients() -> dict[str, dict]:
    """DB export「患者マスタ」から patient_code → dict を返す."""
    wb = load_workbook(DB_XLSX, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header_raw = None
    rows: dict[str, dict] = {}

    def hidx(header: list, prefix: str) -> int | None:
        for i, c in enumerate(header):
            if c and str(c).startswith(prefix):
                return i
        return None

    code_idx = None
    for row in ws.iter_rows(values_only=True):
        if header_raw is None:
            header_raw = list(row)
            code_idx = hidx(header_raw, "patient_code")
            continue
        if not row or code_idx is None:
            continue
        code = norm(row[code_idx])
        if not code or not code.startswith("P"):
            continue
        rec = {}
        for i, h in enumerate(header_raw):
            if h:
                key = norm(h).split(" ")[0].split("(")[0]
                rec[key] = row[i] if i < len(row) else None
        rows[code] = rec
    wb.close()
    return rows


# DB 側 Excel template に「エリア」「週訪問回数」列は存在しない (Phase E-8 で
# weekly_pattern シートに移動 / area は CareFlow patient table 自体に無い).
# → これらは「DB 列が無い」ため diff を出しても User 側で対応不能なので除外し、
#   informational として末尾に集計だけ出す.

# 比較対象フィールド: (User 列名, DB 列名, 正規化関数, 表示ラベル)
COMPARE_FIELDS = [
    ("患者名", "患者名", norm, "氏名"),
    ("フリガナ", "フリガナ", norm, "フリガナ"),
    ("性別", "性別", norm_sex, "性別"),
    ("住所", "住所", norm, "住所"),
    ("稼働状況", "ステータス", norm_status, "稼働状況"),
]


def diff_attrs(user_rec: dict, db_rec: dict) -> list[tuple[str, str, str]]:
    """共通 code の patient で属性 diff. (label, user_value, db_value) のリスト."""
    out = []
    for user_key, db_key, normalize, label in COMPARE_FIELDS:
        u = normalize(user_rec.get(user_key))
        d = normalize(db_rec.get(db_key))
        if u != d:
            out.append((label, u, d))
    return out


def main() -> None:
    print(f"Loading user: {USER_XLSX}")
    user = load_user_patients()
    print(f"  User patients: {len(user)}")
    print(f"Loading DB: {DB_XLSX}")
    db = load_db_patients()
    print(f"  DB patients: {len(db)}")

    user_codes = set(user)
    db_codes = set(db)

    common = sorted(user_codes & db_codes)
    user_only = sorted(user_codes - db_codes)
    db_only = sorted(db_codes - user_codes)

    # 共通 code で属性 diff
    attr_diffs: list[tuple[str, list[tuple[str, str, str]]]] = []
    for code in common:
        d = diff_attrs(user[code], db[code])
        if d:
            attr_diffs.append((code, d))

    print(f"\nCommon: {len(common)}, with attr diff: {len(attr_diffs)}")
    print(f"User-only: {len(user_only)}")
    print(f"DB-only: {len(db_only)}")

    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase F-3-A: 患者マスタ diff (User シート vs 現 DB)\n\n")
        f.write(
            "- User: `Sampledata/2026.05.19段階最新ユーザーシート/スケジュール手動 のコピー.xlsx` シート「元データ」\n"
        )
        f.write("- DB: `_current_db_export.xlsx` シート「患者マスタ」\n")
        f.write("- 突合キー: patient_code\n\n")

        # サマリ
        f.write("## サマリ\n\n")
        f.write(f"- User シート患者数: **{len(user)}**\n")
        f.write(f"- DB 患者数: **{len(db)}**\n")
        f.write(f"- 共通 code: {len(common)} (うち属性差分あり: **{len(attr_diffs)}**)\n")
        f.write(f"- User のみ (新規候補): **{len(user_only)}**\n")
        f.write(f"- DB のみ (削除候補): **{len(db_only)}**\n\n")

        # 1. User のみ (新規候補)
        f.write("## 1. User シートのみに存在 (= DB への新規追加候補)\n\n")
        if not user_only:
            f.write("_該当なし_\n\n")
        else:
            f.write("| patient_code | 氏名 | フリガナ | 性別 | 住所 | エリア | 週訪問回数 |\n")
            f.write("|---|---|---|---|---|---|---|\n")
            for code in user_only:
                r = user[code]
                f.write(
                    f"| {code} | {r.get('患者名', '')} | {r.get('フリガナ', '')} | "
                    f"{r.get('性別', '')} | {r.get('住所', '')} | "
                    f"{r.get('エリア', '')} | {norm_int(r.get('週訪問回数'))} |\n"
                )
            f.write("\n")

        # 2. DB のみ (削除候補)
        f.write("## 2. 現 DB のみに存在 (= User シートに無い = 削除候補)\n\n")
        if not db_only:
            f.write("_該当なし_\n\n")
        else:
            f.write("| patient_code | 氏名 | フリガナ | 性別 | 住所 | エリア | 週訪問回数 |\n")
            f.write("|---|---|---|---|---|---|---|\n")
            for code in db_only:
                r = db[code]
                f.write(
                    f"| {code} | {r.get('患者名', '')} | {r.get('フリガナ', '')} | "
                    f"{r.get('性別', '')} | {r.get('住所', '')} | "
                    f"{r.get('エリア', '')} | {norm_int(r.get('週訪問回数'))} |\n"
                )
            f.write("\n")

        # 3. 共通で属性差分
        f.write("## 3. 共通 patient_code で属性差分あり (= 修正候補)\n\n")
        if not attr_diffs:
            f.write("_該当なし_\n\n")
        else:
            f.write("| patient_code | 氏名 (User) | 差分項目 | User 値 | DB 値 |\n")
            f.write("|---|---|---|---|---|\n")
            for code, diffs in attr_diffs:
                user_name = norm(user[code].get("患者名"))
                for label, u, d in diffs:
                    f.write(f"| {code} | {user_name} | {label} | {u} | {d} |\n")
            f.write("\n")

    print("\n=== 出力 ===")
    print(f"{OUTPUT_MD}")


if __name__ == "__main__":
    main()
