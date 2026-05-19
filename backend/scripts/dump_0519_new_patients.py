"""Phase F-3-B-1: 0519 シートから新規 10 名候補の付帯情報を抽出.

DB に名前一致しない 10 名:
  朝倉 美夢 / 岡村 敏子 / 河野 天歩 / 清水 政憲 / 中尾 要太 /
  槇 恵 / 幸 千秋 / 渡邉 愛 / 菅原 茉結 / 今井 康敦

→ 住所・性別・所属・フリガナ等 新規追加に必要な情報を一覧化.
"""

from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)
DB_XLSX = Path(__file__).parent / "_current_db_export.xlsx"
OUTPUT_MD = Path(__file__).parent / "_dump_0519_new_patients.md"


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    return unicodedata.normalize("NFKC", s)


def norm_name(v) -> str:
    return norm(v).replace("　", "").replace(" ", "")


def hidx(header: list, prefix: str) -> int | None:
    for i, c in enumerate(header):
        if c and str(c).startswith(prefix):
            return i
    return None


def main() -> None:
    # DB 患者名 set
    wb_db = load_workbook(DB_XLSX, data_only=True, read_only=True)
    ws_p = wb_db["患者マスタ"]
    header = None
    db_names: set[str] = set()
    for row in ws_p.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            code_idx = hidx(header, "patient_code")
            name_idx = hidx(header, "患者名")
            continue
        if not row or row[code_idx] is None:
            continue
        nm = norm_name(row[name_idx])
        if nm:
            db_names.add(nm)
    wb_db.close()

    # 0519 シート全 dump
    wb_user = load_workbook(USER_XLSX, data_only=True, read_only=True)
    ws = wb_user["松岡作業中_マスタヒアリング_0519"]
    user_header = None
    new_patients: list[dict] = []
    all_rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if user_header is None:
            user_header = [norm(c) for c in row]
            continue
        if not row or all(c is None or c == "" for c in row):
            continue
        rec = {h: row[i] for i, h in enumerate(user_header) if h and i < len(row)}
        nm = rec.get(norm("患者名"))
        if not nm:
            continue
        all_rows.append(rec)
        if norm_name(nm) not in db_names:
            new_patients.append(rec)
    wb_user.close()

    print(f"User 0519 total: {len(all_rows)}")
    print(f"New patients (not in DB): {len(new_patients)}")

    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase F-3-B-1: 新規 10 名候補 (0519 シート vs DB 患者名)\n\n")
        f.write(f"- 元シート: `{USER_XLSX.name}` 「松岡作業中_マスタヒアリング_0519」\n")
        f.write(f"- 全行数: {len(all_rows)}\n")
        f.write(f"- DB に名前無し (新規候補): **{len(new_patients)}**\n\n")
        f.write("## 新規追加候補 詳細\n\n")
        f.write(
            "| 患者名 | フリガナ | 所属 | 性別 | 住所 | 稼働状況 | 週訪問回数 | 希望曜日 | サービス時間 | 時間タイプ | 希望開始 | 希望終了 |\n"
        )
        f.write("|---|---|---|---|---|---|---|---|---|---|---|---|\n")

        def _fmt(rec: dict, k: str) -> str:
            v = rec.get(norm(k))
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        for r in new_patients:
            f.write(
                f"| {_fmt(r, '患者名')} | {_fmt(r, 'フリガナ')} | {_fmt(r, '所属')} | "
                f"{_fmt(r, '性別')} | {_fmt(r, '住所')} | {_fmt(r, '稼働状況')} | "
                f"{_fmt(r, '週訪問回数')} | {_fmt(r, '希望曜日(複数可)')} | "
                f"{_fmt(r, 'サービス時間')} | {_fmt(r, '時間タイプ')} | "
                f"{_fmt(r, '希望時間帯(開始)')} | {_fmt(r, '希望時間帯(終了)')} |\n"
            )

        # 既存 DB と patient_code resolve 不能な行 (= 何かしらの理由で名前一致しないが
        # DB には居る可能性も。チェック用に全 86 行を出してもいい)

    print("\n=== 出力 ===")
    print(f"{OUTPUT_MD}")


if __name__ == "__main__":
    main()
