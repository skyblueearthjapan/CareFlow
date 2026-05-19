"""User の Google Sheets マスタを CareFlow Excel 完全置換テンプレート形式に変換.

入力:
  - backend/scripts/_user_patient_master.md (患者マスタ markdown table)
  - backend/scripts/_user_staff_master.md (スタッフマスタ markdown table)

出力:
  - backend/scripts/_user_master_patient.xlsx (CareFlow 患者 Excel)
  - backend/scripts/_user_master_staff.xlsx (CareFlow スタッフ Excel)

仕様 (User 確認済):
  - 対象: 患者 + スタッフ マスタのみ (運用シート無視)
  - 未対応 field (エリア/指定スタッフ/スキル/曜日NG/継続希望等) は破棄
  - 反映方法: 完全置換 (= Phase E-4 / E-7 の replace_all import)
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

SCRIPT_DIR = Path(__file__).parent
PATIENT_INPUT = SCRIPT_DIR / "_user_patient_master.md"
STAFF_INPUT = SCRIPT_DIR / "_user_staff_master.md"
PATIENT_OUTPUT = SCRIPT_DIR / "_user_master_patient.xlsx"
STAFF_OUTPUT = SCRIPT_DIR / "_user_master_staff.xlsx"

# ---------------------------------------------------------------------------
# 値マッピング
# ---------------------------------------------------------------------------

# User 稼働状況 → CareFlow status enum
STATUS_MAP = {
    "稼働": "active",
    "休止": "suspended",
    "入院": "admitted",
    "未開始": "pending",
    "未契約": "cancelled",
}

# 性別: User 「女性/男性」→ CareFlow 「female/male」
SEX_MAP = {"女性": "female", "男性": "male"}

# 性別制限: そのまま受理 (CareFlow Excel は「女性のみ/男性のみ」を受理)
SEX_RESTRICTION_PASSTHROUGH = {"女性のみ", "男性のみ"}


# 必要スタッフ数 → requires_multiple_staff (TRUE/FALSE)
def parse_multi_staff(raw: str) -> str:
    raw = (raw or "").strip()
    if raw in ("2人", "2名", "2"):
        return "TRUE"
    return "FALSE"


# 時間タイプ: 「固定/時間帯/午前/午後/終日」そのまま (CareFlow と一致)
TIME_TYPE_PASSTHROUGH = {"固定", "時間帯", "午前", "午後", "終日"}

# 曜日: User "Mon, Wed, Fri" → CareFlow PFV weekday 0-6
WEEKDAY_MAP = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6}
WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]


# 拠点判定 (住所文字列から)
def resolve_office_from_address(address: str) -> str:
    """住所文字列から拠点コードを判定 (INAGE/TSUGA)."""
    if not address:
        return "INAGE"
    if "若葉区都賀" in address or "都賀" in address:
        # ただし「都賀」が住所の他部分にあると誤検出するリスク。
        # User シートで都賀 office に該当するのは S005 のみと判明済み.
        return "TSUGA"
    return "INAGE"


# ---------------------------------------------------------------------------
# Markdown table parser
# ---------------------------------------------------------------------------


def parse_markdown_table(text: str) -> list[dict[str, str]]:
    """| A | B | ... | 形式の markdown table を dict 配列に変換.

    最初の行はヘッダー、2 行目は区切り (skip)、3 行目以降がデータ.
    各セル文字列は trim 済.
    """
    # BOM 除去
    if text.startswith("﻿"):
        text = text[1:]
    # JSON wrapper の処理: ファイルは `{"fileContent":"..."}` 形式
    text = text.lstrip()
    if text.startswith('{"fileContent":"'):
        text = text[len('{"fileContent":"') :]
        if text.rstrip().endswith('"}'):
            text = text.rstrip()[:-2]
    # PowerShell 保存時に backslash が double-escape されているケース
    # (file 内に literal `\\_` が入っている = Python 文字列としては 2 char "\\")
    text = text.replace("\\\\_", "_").replace("\\\\n", "\n").replace('\\\\"', '"')
    # その他 escape sequences
    text = text.replace("\\n", "\n").replace('\\"', '"').replace("\\_", "_").replace("\\*", "*")

    # Find table-like lines (start with `|`, contains at least 2 `|`)
    lines = [ln.strip() for ln in text.split("\n") if "|" in ln and ln.strip().startswith("|")]
    if len(lines) < 3:
        return []

    # Find header: first non-separator line; data lines = everything after it (excluding separators)
    def is_separator_line(ln: str) -> bool:
        cells = [c.strip() for c in ln.strip("|").split("|")]
        return all(c in (":-:", "---", ":---", "---:", "") for c in cells if c is not None)

    header_idx = None
    for i, ln in enumerate(lines):
        if not is_separator_line(ln):
            header_idx = i
            break
    if header_idx is None:
        return []
    header_cells = [c.strip() for c in lines[header_idx].strip("|").split("|")]
    rows: list[dict[str, str]] = []
    for ln in lines[header_idx + 1 :]:
        if is_separator_line(ln):
            continue
        cells = [c.strip() for c in ln.strip("|").split("|")]
        # 列数が合わない行は skip (defensive)
        if len(cells) < len(header_cells):
            cells = cells + [""] * (len(header_cells) - len(cells))
        elif len(cells) > len(header_cells):
            cells = cells[: len(header_cells)]
        rows.append(dict(zip(header_cells, cells, strict=False)))
    return rows


# ---------------------------------------------------------------------------
# 患者マスタ 変換
# ---------------------------------------------------------------------------


def parse_time_hhmm(raw: str) -> str:
    """User '13:00:00' or '13:00' → CareFlow 'HH:MM'."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    m = re.match(r"^(\d{1,2}):(\d{2})", raw)
    if not m:
        return ""
    h = int(m.group(1))
    mi = int(m.group(2))
    return f"{h:02d}:{mi:02d}"


def parse_weekdays(raw: str) -> list[int]:
    """'Mon, Wed, Fri' → [0, 2, 4]."""
    raw = (raw or "").strip()
    if not raw or raw in ("選択なし", "空欄", "なし"):
        return []
    result: list[int] = []
    for token in re.split(r"[,，、\s]+", raw):
        token = token.strip()
        if token in WEEKDAY_MAP:
            result.append(WEEKDAY_MAP[token])
    return sorted(set(result))


def build_patient_workbook(rows: list[dict[str, str]]) -> Workbook:
    """User の患者 markdown rows から CareFlow Excel (患者 + PFV シート) を生成.

    CareFlow Excel schema (Phase E-7 後):
      Sheet "患者マスタ" (PATIENT_COLUMNS):
        - 患者ID (= UUID, 新規時は空欄)
        - 患者コード (= P001)
        - 患者名 / フリガナ / 性別 / ステータス / 保険区分 / 住所 / 緯度 / 経度 /
          拠点コード / 性別制限 / 複数スタッフ必須 / 備考 / 削除フラグ

      Sheet "固定訪問枠" (PFV_COLUMNS):
        - 患者ID / 患者コード / モード / 曜日 / 開始時刻 / duration_min /
          slot_index / course_template_code / sub_office_code / 削除フラグ
    """
    wb = Workbook()
    # 1 つ目のシート = 患者マスタ
    ws_p: Worksheet = wb.active
    ws_p.title = "患者マスタ"
    patient_headers = [
        "患者ID",
        "患者コード",
        "患者名",
        "フリガナ",
        "性別",
        "ステータス",
        "保険区分",
        "住所",
        "緯度",
        "経度",
        "拠点コード",
        "性別制限",
        "複数スタッフ必須",
        "備考",
        "削除フラグ",
    ]
    ws_p.append(patient_headers)

    # 2 つ目のシート = 固定訪問枠
    ws_f = wb.create_sheet("固定訪問枠")
    pfv_headers = [
        "患者ID",
        "患者コード",
        "モード",
        "曜日",
        "開始時刻",
        "duration_min",
        "slot_index",
        "course_template_code",
        "sub_office_code",
        "削除フラグ",
    ]
    ws_f.append(pfv_headers)

    for row in rows:
        # 患者基本情報
        patient_code = row.get("patient_id", "").strip()
        if not patient_code:
            continue  # skip empty

        name = row.get("患者名", "").strip()
        kana = row.get("フリガナ", "").strip()
        sex = SEX_MAP.get(row.get("性別", "").strip(), "")
        status = STATUS_MAP.get(row.get("稼働状況", "").strip(), "")
        address = row.get("住所", "").strip()
        lat = row.get("緯度", "").strip()
        lng = row.get("経度", "").strip()
        sex_restriction_raw = row.get("性別制限", "").strip()
        sex_restriction = (
            sex_restriction_raw if sex_restriction_raw in SEX_RESTRICTION_PASSTHROUGH else ""
        )
        multi_staff = parse_multi_staff(row.get("必要スタッフ数", ""))
        note = row.get("備考", "").strip()
        # 保険区分: User 「医療保険」→ CareFlow "medical" (default)
        insurance_raw = row.get("保険区分", "").strip()
        if insurance_raw == "介護保険":
            insurance = "care"
        else:
            insurance = "medical"
        # 拠点: 住所から推定 (全員 INAGE がデフォルト、都賀のみ判定)
        office_code = resolve_office_from_address(address)

        ws_p.append(
            [
                "",  # 患者ID (新規/UUID 未指定)
                patient_code,
                name,
                kana,
                sex,
                status,
                insurance,
                address,
                lat,
                lng,
                office_code,
                sex_restriction,
                multi_staff,
                note,
                "",  # 削除フラグ
            ]
        )

        # PFV 行: 希望曜日 + 時間タイプ + 時間 + サービス時間 から展開
        weekdays = parse_weekdays(row.get("希望曜日（複数可）", ""))
        time_type = row.get("時間タイプ", "").strip()
        if time_type not in TIME_TYPE_PASSTHROUGH:
            time_type = ""
        pref_start = parse_time_hhmm(row.get("希望時間帯（開始）", ""))
        pref_end = parse_time_hhmm(row.get("希望時間帯（終了）", ""))  # noqa: F841
        service_min_raw = row.get("サービス時間", "").strip()
        try:
            service_min = int(service_min_raw)
        except (ValueError, TypeError):
            service_min = 35  # default

        # 固定 / 時間帯 の場合のみ PFV 行を生成 (午前/午後/終日 は preferred のみで PFV 不要)
        # ただし、CareFlow では time_type は patient.weekly_pattern.time_type に
        # 入る (= 患者基本情報の一部) で、PFV にはあくまで「固定枠の予定」が入る.
        # User シートでは PFV と weekly_pattern が混在しているため、簡略化:
        #   - 希望曜日が指定されている (= 週訪問予定) なら PFV を生成
        #   - PFV の start_time は preferred_start (時間帯/固定どちらでも)
        if weekdays and pref_start:
            for wd in weekdays:
                ws_f.append(
                    [
                        "",  # 患者ID
                        patient_code,
                        "normal",
                        WEEKDAY_JP[wd],
                        pref_start,
                        service_min,
                        0,  # slot_index
                        "",  # course_template_code (空 = Layer 1 fallback)
                        "",  # sub_office_code
                        "",  # 削除フラグ
                    ]
                )

    return wb


# ---------------------------------------------------------------------------
# スタッフマスタ 変換
# ---------------------------------------------------------------------------


def build_staff_workbook(rows: list[dict[str, str]]) -> Workbook:
    """User の staff markdown rows から CareFlow Excel (Staff + Shift + Override) を生成."""
    wb = Workbook()
    ws_s: Worksheet = wb.active
    ws_s.title = "スタッフマスタ"
    staff_headers = [
        "スタッフID",
        "スタッフコード",
        "スタッフ名",
        "フリガナ",
        "性別",
        "ステータス",
        "ロール",
        "拠点コード",
        "新人フラグ",
        "サブ拠点コード (カンマ区切り, 解除は <CLEAR>)",
        "備考",
        "削除フラグ",
    ]
    ws_s.append(staff_headers)

    ws_h = wb.create_sheet("勤務シフト")
    shift_headers = [
        "スタッフコード",
        "曜日",
        "勤務",
        "開始時刻",
        "終了時刻",
    ]
    ws_h.append(shift_headers)

    # Override シートは空のまま (今回は対象外)
    ws_o = wb.create_sheet("勤務例外")
    override_headers = [
        "スタッフコード",
        "iso_year",
        "iso_week",
        "曜日",
        "override_type",
        "開始時刻",
        "終了時刻",
        "理由",
        "削除フラグ",
    ]
    ws_o.append(override_headers)

    for row in rows:
        staff_code = row.get("staff_id", "").strip()
        if not staff_code:
            continue
        name = row.get("スタッフ名", "").strip()
        sex = SEX_MAP.get(row.get("性別", "").strip(), "")
        address = row.get("拠点住所", "").strip()
        office_code = resolve_office_from_address(address)
        note = row.get("備考", "").strip()
        shift_start = parse_time_hhmm(row.get("シフト開始", ""))
        shift_end = parse_time_hhmm(row.get("シフト終了", ""))
        weekdays_raw = row.get("勤務曜日", "").strip()
        on_weekdays = parse_weekdays(weekdays_raw)

        ws_s.append(
            [
                "",  # スタッフID (UUID、新規)
                staff_code,
                name,
                "",  # フリガナ (User シートに無し)
                sex,
                "active",  # User シートに無し、全員 active 想定
                "staff",  # ロール: 一般 staff (manager は別途設定)
                office_code,
                "FALSE",  # 新人フラグ: false default
                "",  # サブ拠点コード: 空 (= 関連解除しない、維持)
                note,
                "",  # 削除フラグ
            ]
        )

        # シフト行: 月-土 (週 6 日) × 勤務曜日に含まれていれば TRUE, それ以外 FALSE
        for wd in range(6):  # 月-土
            is_on = "TRUE" if wd in on_weekdays else "FALSE"
            ws_h.append(
                [
                    staff_code,
                    WEEKDAY_JP[wd],
                    is_on,
                    shift_start if is_on == "TRUE" else "",
                    shift_end if is_on == "TRUE" else "",
                ]
            )

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    print("=== User マスタ → CareFlow Excel 変換 ===")

    # 患者マスタ
    patient_text = PATIENT_INPUT.read_text(encoding="utf-8")
    patient_rows = parse_markdown_table(patient_text)
    print(f"Patient rows parsed: {len(patient_rows)}")
    if patient_rows:
        print(f"  First row weekday: {patient_rows[0].get('希望曜日（複数可）', 'MISSING')!r}")
        print(f"  First row service_min: {patient_rows[0].get('サービス時間', 'MISSING')!r}")
    patient_wb = build_patient_workbook(patient_rows)
    patient_wb.save(PATIENT_OUTPUT)
    print(f"  → {PATIENT_OUTPUT}")
    # PFV 行数を数える
    pfv_sheet = patient_wb["固定訪問枠"]
    print(f"  PFV rows: {pfv_sheet.max_row - 1}")

    # スタッフマスタ
    staff_text = STAFF_INPUT.read_text(encoding="utf-8")
    staff_rows = parse_markdown_table(staff_text)
    print(f"Staff rows parsed: {len(staff_rows)}")
    staff_wb = build_staff_workbook(staff_rows)
    staff_wb.save(STAFF_OUTPUT)
    print(f"  → {STAFF_OUTPUT}")
    shift_sheet = staff_wb["勤務シフト"]
    print(f"  Shift rows: {shift_sheet.max_row - 1}")

    print("\n=== 完了 ===")
    print(f"Patient master: {len(patient_rows)} patients")
    print(f"Staff master: {len(staff_rows)} staffs")


if __name__ == "__main__":
    main()
