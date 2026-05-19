"""スケジュール枠組み（仮）シート構造を inspect."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)


def main() -> None:
    wb = load_workbook(USER_XLSX, data_only=True)
    ws = wb["スケジュール枠組み（仮）"]
    print(f"max_row={ws.max_row}  max_col={ws.max_column}")
    print(f"merged ranges: {len(ws.merged_cells.ranges)}")
    print()
    for ri in range(1, min(ws.max_row + 1, 60)):
        row = [ws.cell(row=ri, column=ci).value for ci in range(1, min(ws.max_column + 1, 36))]
        cells_compact = []
        for v in row:
            if v is None:
                cells_compact.append("")
            else:
                cells_compact.append(str(v)[:14])
        print(f"R{ri:3d}: {cells_compact}")


if __name__ == "__main__":
    main()
