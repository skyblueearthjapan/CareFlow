"""Phase G-1: User File 2 (Ver.1) 「患者マスタ」 vs 現 DB の 全カラム徹底 diff.

User 指摘:
  - 井川 雄太 が DB で「埼玉県」 ⇔ File 2 で「千葉県」
  - 必要スタッフ数は全員 1 人 (= DB の requires_multiple_staff=true 患者は全員間違い)

入力:
  - User File 2: Sampledata/2026.05.19段階最新ユーザーシート/訪問看護:よりより様.../Ver.1 (1).xlsx
    シート: 患者マスタ (86 行)
  - DB: backend/scripts/_g1_db_export.xlsx
    シート: 患者マスタ (86 行)

突合キー: patient_code

比較対象 (全 column):
  - 患者名 / フリガナ / 性別 / 住所 / エリア / 週訪問回数 / 稼働状況
  - 緯度 / 経度 (= 住所が変わると変動するため要注意)
  - File 2 にあって DB に無い項目 + その逆

出力:
  - backend/scripts/_diff_g1_full.md
    1. 住所違い (= 都道府県 mismatch がある場合は特に強調)
    2. 全カラム属性差分
    3. requires_multiple_staff = true な患者の一覧 (全員 false にすべき疑い)
    4. patient_code 不一致 (User のみ / DB のみ)
"""

from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
DB_XLSX = SCRIPT_DIR / "_g1_db_export.xlsx"
USER_FILE2 = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
)
OUTPUT_MD = SCRIPT_DIR / "_diff_g1_full.md"


def find_file2() -> Path:
    candidates = list(USER_FILE2.glob("訪問看護*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"File 2 not found in {USER_FILE2}")
    return candidates[0]


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    return unicodedata.normalize("NFKC", s)


def norm_name(v) -> str:
    return norm(v).replace("　", "").replace(" ", "")


def norm_int(v) -> str:
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return str(v).strip()


def norm_latlng(v) -> str:
    """緯度・経度を小数点 2 桁丸めで比較 (User シートは精度落ちあり)."""
    if v is None or v == "":
        return ""
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return str(v).strip()


SEX_SYNONYM = {
    "女性": "female",
    "男性": "male",
    "female": "female",
    "male": "male",
    "F": "female",
    "M": "male",
}
STATUS_SYNONYM = {
    "稼働": "active",
    "休止": "suspended",
    "入院": "admitted",
    "未開始": "pending",
    "解約": "cancelled",
    "未契約": "pending",
    "active": "active",
    "suspended": "suspended",
    "admitted": "admitted",
    "pending": "pending",
    "cancelled": "cancelled",
}


def norm_sex(v) -> str:
    return SEX_SYNONYM.get(norm(v), norm(v))


def norm_status(v) -> str:
    return STATUS_SYNONYM.get(norm(v), norm(v))


def addr_prefecture(addr: str) -> str:
    """住所先頭の都道府県を抽出. 「千葉県」「東京都」 等."""
    s = norm(addr)
    if not s:
        return ""
    # 「..県」「..都」「..府」「..道」 のいずれかで終わる先頭部分
    for end in ("県", "都", "府", "道"):
        idx = s.find(end)
        if idx >= 0 and idx <= 4:  # 都道府県名は 2-4 文字
            return s[: idx + 1]
    return ""


def load_user_file2(xlsx: Path) -> dict[str, dict]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header = None
    out: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [norm(c) for c in row]
            continue
        if not row or row[0] is None:
            continue
        code = norm(row[0])
        if not code.startswith("P"):
            continue
        rec = {h: row[i] for i, h in enumerate(header) if h and i < len(row)}
        out[code] = rec
    wb.close()
    return out


def load_db(xlsx: Path) -> dict[str, dict]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header_raw = None
    code_idx = None
    out: dict[str, dict] = {}

    def hidx(h, prefix):
        for i, c in enumerate(h):
            if c and str(c).startswith(prefix):
                return i
        return None

    for row in ws.iter_rows(values_only=True):
        if header_raw is None:
            header_raw = list(row)
            code_idx = hidx(header_raw, "patient_code")
            continue
        if not row or code_idx is None:
            continue
        code = norm(row[code_idx])
        if not code.startswith("P"):
            continue
        rec = {}
        for i, h in enumerate(header_raw):
            if h:
                key = norm(h).split(" ")[0].split("(")[0]
                rec[key] = row[i] if i < len(row) else None
        out[code] = rec
    wb.close()
    return out


# 比較対象: (User 列名, DB 列名, normalize, label)
COMPARE_FIELDS = [
    ("患者名", "患者名", norm, "氏名"),
    ("フリガナ", "フリガナ", norm, "フリガナ"),
    ("性別", "性別", norm_sex, "性別"),
    ("住所", "住所", norm, "住所"),
    ("緯度", "緯度", norm_latlng, "緯度"),
    ("経度", "経度", norm_latlng, "経度"),
    ("稼働状況", "ステータス", norm_status, "稼働状況"),
]


def main() -> None:
    file2 = find_file2()
    print(f"File 2: {file2.name}")
    user = load_user_file2(file2)
    db = load_db(DB_XLSX)
    print(f"User patients: {len(user)}, DB patients: {len(db)}")

    user_codes = set(user)
    db_codes = set(db)
    common = sorted(user_codes & db_codes)
    user_only = sorted(user_codes - db_codes)
    db_only = sorted(db_codes - user_codes)

    # 全カラム diff (= 比較対象に少しでも違いがあれば list)
    attr_diffs: list[dict] = []
    pref_diffs: list[dict] = []  # 都道府県違い (= 住所が県レベルで違う)
    for code in common:
        u, d = user[code], db[code]
        local = []
        for user_key, db_key, normalize, label in COMPARE_FIELDS:
            uv = normalize(u.get(user_key))
            dv = normalize(d.get(db_key))
            if uv != dv:
                local.append((label, uv, dv))
        if local:
            attr_diffs.append({"code": code, "name": norm(u.get("患者名")), "diffs": local})

        # 都道府県違い
        u_pref = addr_prefecture(u.get("住所", ""))
        d_pref = addr_prefecture(d.get("住所", ""))
        if u_pref and d_pref and u_pref != d_pref:
            pref_diffs.append(
                {
                    "code": code,
                    "name": norm(u.get("患者名")),
                    "user_pref": u_pref,
                    "db_pref": d_pref,
                    "user_addr": norm(u.get("住所")),
                    "db_addr": norm(d.get("住所")),
                }
            )

    # 複数スタッフ必須 (= true な DB 患者を抽出)
    multi_staff: list[dict] = []
    for code in sorted(db):
        d = db[code]
        v = d.get("複数スタッフ必須")
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in ("TRUE", "1", "YES"):
            multi_staff.append(
                {
                    "code": code,
                    "name": norm(d.get("患者名")),
                }
            )

    print(f"\nCommon: {len(common)}")
    print(f"User only: {len(user_only)}")
    print(f"DB only: {len(db_only)}")
    print(f"全カラム attr diff: {len(attr_diffs)}")
    print(f"都道府県違い: {len(pref_diffs)}")
    print(f"複数スタッフ必須 = TRUE な DB 患者: {len(multi_staff)}")

    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase G-1: 患者マスタ 全カラム徹底 diff (User File 2 vs 現 DB)\n\n")
        f.write(f"- User: `{file2.name}` シート「患者マスタ」\n")
        f.write("- DB: `_g1_db_export.xlsx` シート「患者マスタ」 (現本番状態)\n")
        f.write("- 突合キー: patient_code\n\n")

        f.write("## サマリ\n\n")
        f.write(f"- User シート患者数: **{len(user)}**\n")
        f.write(f"- DB 患者数: **{len(db)}**\n")
        f.write(f"- 共通 code: {len(common)}\n")
        f.write(f"- 全カラム属性差分: **{len(attr_diffs)}**\n")
        f.write(f"- ⚠️ 都道府県違い: **{len(pref_diffs)}**\n")
        f.write(
            f"- ⚠️ 複数スタッフ必須 = TRUE な患者 (= 全員 FALSE にすべき): **{len(multi_staff)}**\n"
        )
        f.write(f"- User のみ (DB に無い code): {len(user_only)}\n")
        f.write(f"- DB のみ (User に無い code): {len(db_only)}\n\n")

        # 1. 都道府県違い (最重要)
        f.write("## 1. ⚠️ 住所の都道府県違い (要修正)\n\n")
        if not pref_diffs:
            f.write("_該当なし_\n\n")
        else:
            f.write("| code | 氏名 | User 住所 | DB 住所 |\n")
            f.write("|---|---|---|---|\n")
            for d in pref_diffs:
                f.write(f"| {d['code']} | {d['name']} | {d['user_addr']} | {d['db_addr']} |\n")
            f.write("\n")

        # 2. 複数スタッフ必須 = TRUE (全員 FALSE にすべき)
        f.write("## 2. ⚠️ 複数スタッフ必須 = TRUE な患者 (User: 全員 1 人 = FALSE が正)\n\n")
        if not multi_staff:
            f.write("_該当なし_\n\n")
        else:
            f.write("| code | 氏名 |\n")
            f.write("|---|---|\n")
            for d in multi_staff:
                f.write(f"| {d['code']} | {d['name']} |\n")
            f.write("\n")

        # 3. 全カラム属性差分
        f.write("## 3. 全カラム属性差分 (修正候補)\n\n")
        if not attr_diffs:
            f.write("_該当なし — 全件一致_\n\n")
        else:
            f.write("| code | 氏名 | 差分項目 | User 値 | DB 値 |\n")
            f.write("|---|---|---|---|---|\n")
            for d in attr_diffs:
                for label, uv, dv in d["diffs"]:
                    # 長い文字列は省略
                    uv_s = uv if len(uv) < 60 else uv[:57] + "..."
                    dv_s = dv if len(dv) < 60 else dv[:57] + "..."
                    f.write(f"| {d['code']} | {d['name']} | {label} | {uv_s} | {dv_s} |\n")
            f.write("\n")

        # 4. User のみ (新規候補) / DB のみ (削除候補)
        f.write("## 4. patient_code の差分\n\n")
        f.write(f"### 4-a. User File 2 のみに存在 (DB に無い、 新規候補) — {len(user_only)} 件\n\n")
        if user_only:
            f.write("| code | 氏名 | 住所 | 稼働状況 |\n")
            f.write("|---|---|---|---|\n")
            for code in user_only:
                r = user[code]
                f.write(
                    f"| {code} | {norm(r.get('患者名'))} | {norm(r.get('住所'))} | {norm(r.get('稼働状況'))} |\n"
                )
            f.write("\n")
        else:
            f.write("_該当なし_\n\n")

        f.write(f"### 4-b. DB のみに存在 (User File 2 に無い、 削除候補) — {len(db_only)} 件\n\n")
        if db_only:
            f.write("| code | 氏名 | 住所 | ステータス |\n")
            f.write("|---|---|---|---|\n")
            for code in db_only:
                r = db[code]
                f.write(
                    f"| {code} | {norm(r.get('患者名'))} | {norm(r.get('住所'))} | {norm(r.get('ステータス'))} |\n"
                )
            f.write("\n")
        else:
            f.write("_該当なし_\n\n")

    print("\n=== 出力 ===")
    print(f"{OUTPUT_MD}")


if __name__ == "__main__":
    main()
