"""カルテ装飾済みテンプレート資産 (karte_template.xlsx) の生成スクリプト.

Phase G-60: 実患者カルテ出力がサンプル `Sampledata/カルテ表／Sampleフォーマット.xlsx`
と見た目が別物になっていた問題への対処. サンプルを「装飾済みテンプレート資産」として
同梱し、build 関数はこの資産をロードして値だけ流し込む方式に切り替える.

本スクリプトはサンプルを openpyxl で読み、

  * 値セル (コードが書き込む対象) のみクリア
  * 希望時刻 B13:G13 の結合を解除し C13 に静的「〜」を配置
  * シート名を本番カルテのシート名に統一

した上で `app/services/patient_excel/templates/karte_template.xlsx` に保存する.
ラベル / 見出し / 曜日見出し / 行高 / 列幅 / 罫線 / 塗り / フォント / 結合 /
ページ設定 (A4 縦 / 印刷範囲 A1:H25 / fitToPage) は全て維持する.

成果物 .xlsx をコミットするが、本スクリプトも再生成可能性のため残す.

実行: cd backend && python -m scripts.build_karte_template
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# カルテ本番シート名 (karte.SHEET_KARTE と一致させる).
SHEET_KARTE = "訪問看護スケジューリング用カルテ"

# サンプル (装飾の正本).
SAMPLE_PATH = (
    Path(__file__).resolve().parents[2] / "Sampledata" / "カルテ表／Sampleフォーマット.xlsx"
)
# 生成先資産.
TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "app"
    / "services"
    / "patient_excel"
    / "templates"
    / "karte_template.xlsx"
)

# コードが値を流し込む対象セル (= クリア対象). ラベル/見出しは残す.
#   G1 患者コード / B2 氏名 / F2 フリガナ
#   B5 性別 / D5 ステータス / F5 保険
#   B6 拠点 / D6 住所
#   B9 週回数 / D9 頻度
#   B11:H11 希望曜日 〇×
#   B12 サービス時間 / D12 時間タイプ
#   B13 希望開始 / D13 希望終了 (C13 は静的「〜」)
#   B18:H18 固定訪問時刻 / B19:H19 固定訪問コース
#   A22 備考
VALUE_CELLS: tuple[str, ...] = (
    "G1",
    "B2",
    "F2",
    "B5",
    "D5",
    "F5",
    "B6",
    "D6",
    "B9",
    "D9",
    *(f"{c}11" for c in "BCDEFGH"),
    "B12",
    "D12",
    "B13",
    "D13",
    *(f"{c}18" for c in "BCDEFGH"),
    *(f"{c}19" for c in "BCDEFGH"),
    "A22",
)

# 希望時刻の分割: サンプルは B13:G13 結合 ("09:30〜10:30"). これを解除し
# B13=開始 / C13=静的「〜」 / D13=終了 の 3 セル運用にする.
PREF_TIME_MERGE = "B13:G13"
PREF_TIME_SEP_CELL = "C13"
PREF_TIME_SEP_TEXT = "〜"


def build_template() -> Path:
    wb = load_workbook(SAMPLE_PATH)
    ws = wb[wb.sheetnames[0]]
    ws.title = SHEET_KARTE

    # 希望時刻の結合解除 (装飾は各セルに残る). フレーム (A13 left=thick / H13 right=thick)
    # は B13:G13 の結合とは独立なので解除しても保持される.
    if PREF_TIME_MERGE in {str(r) for r in ws.merged_cells.ranges}:
        ws.unmerge_cells(PREF_TIME_MERGE)

    # 値セルをクリア (style は openpyxl では value とは独立なので維持される).
    for ref in VALUE_CELLS:
        ws[ref].value = None

    # C13 に静的「〜」を配置し中央寄せ.
    sep = ws[PREF_TIME_SEP_CELL]
    sep.value = PREF_TIME_SEP_TEXT
    sep.alignment = Alignment(horizontal="center", vertical="center")

    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


if __name__ == "__main__":
    path = build_template()
    print(f"karte template written: {path}")
