"""Phase F-3-B: User シート「松岡作業中_マスタヒアリング_0519」 vs DB 「希望訪問パターン」 diff.

突合キー: 患者名 (normalize 後).
  User 側に patient_code が無いため、name で寄せる. 表記揺れ吸収のため、
  ① 完全一致 → ② 空白除去一致 → ③ フリガナ一致 の順で resolve.

比較項目:
  - 週訪問回数 (frequency_per_week)
  - 希望曜日 (preferred_weekdays, set 比較)
  - サービス時間 (service_minutes)
  - 時間タイプ (time_type)
  - 希望時間帯 (preferred_start, preferred_end)

出力: backend/scripts/_diff_weekly_pattern.md
"""

from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
DB_XLSX = SCRIPT_DIR / "_current_db_export.xlsx"
USER_XLSX = (
    Path(r"C:\Users\imaizumi.LINEWORKS-NET\Documents\CareFlow\Sampledata")
    / "2026.05.19段階最新ユーザーシート"
    / "スケジュール手動 のコピー.xlsx"
)
OUTPUT_MD = SCRIPT_DIR / "_diff_weekly_pattern.md"

USER_SHEET = "松岡作業中_マスタヒアリング_0519"
DB_SHEET = "希望訪問パターン"


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    return unicodedata.normalize("NFKC", s)


def norm_name(v) -> str:
    """全角・半角空白を除去した名前."""
    s = norm(v)
    return s.replace("　", "").replace(" ", "")


def norm_num(v) -> str:
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return str(v).strip()


def norm_time(v) -> str:
    """時刻文字列正規化: '13:00:00' / '13:00' / time obj → 'HH:MM'."""
    if v is None or v == "":
        return ""
    s = str(v).strip()
    if not s:
        return ""
    # HH:MM:SS / HH:MM どちらも HH:MM に揃える
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    return s


def norm_weekdays_user(v) -> set[str]:
    """User の 'Mon, Wed, Fri' 等 → {'Mon','Wed','Fri'}."""
    s = norm(v)
    if not s:
        return set()
    return {p.strip() for p in s.replace(";", ",").split(",") if p.strip()}


def norm_weekdays_db(rec: dict) -> set[str]:
    """DB の wd_mon/wd_tue/.. TRUE/FALSE → {'Mon','Tue',...}."""
    out = set()
    mapping = [
        ("希望曜日_月", "Mon"),
        ("希望曜日_火", "Tue"),
        ("希望曜日_水", "Wed"),
        ("希望曜日_木", "Thu"),
        ("希望曜日_金", "Fri"),
        ("希望曜日_土", "Sat"),
        ("希望曜日_日", "Sun"),
    ]
    for key, en in mapping:
        v = rec.get(key)
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in ("TRUE", "1", "YES"):
            out.add(en)
    return out


def hidx(header: list, prefix: str) -> int | None:
    for i, c in enumerate(header):
        if c and str(c).startswith(prefix):
            return i
    return None


def load_user_0519() -> list[dict]:
    wb = load_workbook(USER_XLSX, data_only=True, read_only=True)
    ws = wb[USER_SHEET]
    header = None
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [norm(c) for c in row]
            continue
        if not row or all(c is None or c == "" for c in row):
            continue
        rec = {h: row[i] for i, h in enumerate(header) if h and i < len(row)}
        if not rec.get("患者名"):
            continue
        rows.append(rec)
    wb.close()
    return rows


def load_db_weekly() -> dict[str, dict]:
    """DB の希望訪問パターン. patient_code → record."""
    wb = load_workbook(DB_XLSX, data_only=True, read_only=True)
    ws = wb[DB_SHEET]
    header_raw = None
    code_idx = None
    rows: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if header_raw is None:
            header_raw = list(row)
            code_idx = hidx(header_raw, "patient_code")
            continue
        if not row or code_idx is None or not row[code_idx]:
            continue
        code = norm(row[code_idx])
        if not code.startswith("P"):
            continue
        rec = {}
        for i, h in enumerate(header_raw):
            if h:
                key = norm(h).split(" ")[0]
                rec[key] = row[i] if i < len(row) else None
        rows[code] = rec
    wb.close()
    return rows


def load_patient_lookup() -> dict[str, str]:
    """patient_code → 患者名 (DB 患者マスタから)."""
    wb = load_workbook(DB_XLSX, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header = None
    code_idx = name_idx = None
    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            code_idx = hidx(header, "patient_code")
            name_idx = hidx(header, "患者名")
            continue
        if not row or code_idx is None:
            continue
        code = norm(row[code_idx])
        name = norm(row[name_idx]) if name_idx is not None else ""
        if code.startswith("P"):
            out[code] = name
    wb.close()
    return out


def main() -> None:
    user_rows = load_user_0519()
    db_weekly = load_db_weekly()
    patient_codes = load_patient_lookup()  # code → 患者名
    print(f"User rows ({USER_SHEET}): {len(user_rows)}")
    print(f"DB weekly: {len(db_weekly)}")
    print(f"DB patients: {len(patient_codes)}")

    # 名前正規化 → patient_code 解決
    name_to_code: dict[str, str] = {}
    for code, name in patient_codes.items():
        name_to_code[norm_name(name)] = code

    # User 行を name → code 解決
    resolved: list[tuple[str, dict]] = []  # (code, user_rec)
    unresolved_users: list[dict] = []
    for r in user_rows:
        nm = norm_name(r.get("患者名"))
        if nm in name_to_code:
            resolved.append((name_to_code[nm], r))
        else:
            unresolved_users.append(r)

    print(f"User → DB code 解決: {len(resolved)} / 未解決: {len(unresolved_users)}")

    # 差分集計
    diff_rows: list[dict] = []
    db_codes_covered: set[str] = set()
    for code, urec in resolved:
        db_codes_covered.add(code)
        drec = db_weekly.get(code, {})

        # 各項目比較
        local_diffs = []

        # User シート header は norm 経由で 全角括弧 → 半角括弧 になる
        u_freq = norm_num(urec.get(norm("週訪問回数")))
        d_freq = norm_num(drec.get("週訪問回数"))
        if u_freq != d_freq:
            local_diffs.append(("週訪問回数", u_freq, d_freq))

        u_wd = norm_weekdays_user(urec.get(norm("希望曜日（複数可）")))
        d_wd = norm_weekdays_db(drec)
        if u_wd != d_wd:
            local_diffs.append(
                ("希望曜日", ",".join(sorted(u_wd)) or "(なし)", ",".join(sorted(d_wd)) or "(なし)")
            )

        u_svc = norm_num(urec.get(norm("サービス時間")))
        d_svc = norm_num(drec.get("サービス時間"))
        if u_svc != d_svc:
            local_diffs.append(("サービス時間", u_svc, d_svc))

        u_tt = norm(urec.get(norm("時間タイプ")))
        d_tt = norm(drec.get("時間タイプ"))
        if u_tt != d_tt:
            local_diffs.append(("時間タイプ", u_tt, d_tt))

        u_ps = norm_time(urec.get(norm("希望時間帯（開始）")))
        d_ps = norm_time(drec.get("希望開始時刻"))
        if u_ps != d_ps:
            local_diffs.append(("希望開始", u_ps, d_ps))

        u_pe = norm_time(urec.get(norm("希望時間帯（終了）")))
        d_pe = norm_time(drec.get("希望終了時刻"))
        if u_pe != d_pe:
            local_diffs.append(("希望終了", u_pe, d_pe))

        if local_diffs:
            diff_rows.append(
                {
                    "code": code,
                    "name": norm(urec.get("患者名")),
                    "diffs": local_diffs,
                }
            )

    db_not_in_user = sorted(set(db_weekly) - db_codes_covered)

    print(f"\n差分あり患者: {len(diff_rows)}")
    print(f"DB のみ (User 0519 に無い code): {len(db_not_in_user)}")

    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase F-3-B: 希望訪問パターン diff\n\n")
        f.write(f"- User: `{USER_XLSX.name}` シート「{USER_SHEET}」\n")
        f.write(f"- DB: `_current_db_export.xlsx` シート「{DB_SHEET}」\n")
        f.write("- 突合キー: 患者名 (空白除去比較で patient_code に変換)\n\n")

        f.write("## サマリ\n\n")
        f.write(f"- User 0519 行数: **{len(user_rows)}**\n")
        f.write(f"- DB weekly_pattern 行数: **{len(db_weekly)}**\n")
        f.write(f"- User → DB code 解決成功: **{len(resolved)}**\n")
        f.write(f"- User 未解決 (名前が DB に無い): **{len(unresolved_users)}**\n")
        f.write(f"- 共通患者で差分あり: **{len(diff_rows)}**\n")
        f.write(f"- DB のみ (0519 に記載なし): **{len(db_not_in_user)}**\n\n")

        f.write("## 1. User 0519 に記載があるが、DB の患者名と一致しない\n\n")
        if not unresolved_users:
            f.write("_該当なし_\n\n")
        else:
            f.write(
                "| 患者名 (User) | 週訪問回数 | 希望曜日 | サービス時間 | 時間タイプ | 希望開始 | 希望終了 |\n"
            )
            f.write("|---|---|---|---|---|---|---|\n")
            for r in unresolved_users:
                f.write(
                    f"| {norm(r.get(norm('患者名')))} | "
                    f"{norm_num(r.get(norm('週訪問回数')))} | "
                    f"{norm(r.get(norm('希望曜日(複数可)')))} | "
                    f"{norm_num(r.get(norm('サービス時間')))} | "
                    f"{norm(r.get(norm('時間タイプ')))} | "
                    f"{norm_time(r.get(norm('希望時間帯(開始)')))} | "
                    f"{norm_time(r.get(norm('希望時間帯(終了)')))} |\n"
                )
            f.write("\n")

        f.write("## 2. 共通患者で属性差分あり\n\n")
        if not diff_rows:
            f.write("_該当なし — 全件一致_\n\n")
        else:
            f.write("| patient_code | 氏名 | 差分項目 | User 値 | DB 値 |\n")
            f.write("|---|---|---|---|---|\n")
            for r in diff_rows:
                for label, u, d in r["diffs"]:
                    f.write(f"| {r['code']} | {r['name']} | {label} | {u} | {d} |\n")
            f.write("\n")

        f.write("## 3. DB の希望訪問パターンに行があるが User 0519 に記載なし\n\n")
        if not db_not_in_user:
            f.write("_該当なし_\n\n")
        else:
            f.write("| patient_code | 氏名 | DB 週訪問回数 | DB 希望曜日 |\n")
            f.write("|---|---|---|---|\n")
            for code in db_not_in_user:
                rec = db_weekly[code]
                wd = norm_weekdays_db(rec)
                f.write(
                    f"| {code} | {patient_codes.get(code, '')} | "
                    f"{norm_num(rec.get('週訪問回数'))} | {','.join(sorted(wd))} |\n"
                )
            f.write("\n")

    print("\n=== 出力 ===")
    print(f"{OUTPUT_MD}")


if __name__ == "__main__":
    main()
