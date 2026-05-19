"""Phase F-3-E: PFV 追加候補 59 件 → CareFlow Excel テンプレ変換.

入力:
  - _diff_pfv.md の logic を再実行して PFV 追加候補リストを取得
    (diff_pfv.py の関数を呼ぶ)
  - DB の現状: _after_course_fix_db_export.xlsx (= 86 patients + 76 PFV)
    特に「希望訪問パターン」シートから duration_min, time_type を取得

出力: backend/scripts/_pfv_additions.xlsx

シート構成:
  1. 患者マスタ: 空 (= 患者は触らない)
  2. 固定訪問スケジュール: 59 件 (追加)
  3. 希望訪問パターン: 空 (= weekly_pattern は触らない)

PFV 各列の決定方法:
  - patient_id: 空 (patient_code で resolve)
  - patient_code: 解決済み code
  - 患者名: 表示用 (任意)
  - 曜日: 日本語短縮 (月/火/水/...)
  - slot_index: 同 patient × 同 weekday 内で 0 から連番 (DB に既存があれば次の slot)
  - モード: "normal"
  - 時間タイプ: patient.weekly_pattern.time_type を使う (なければ default "終日")
  - 開始時刻: User シートの時刻
  - 終了時刻: 開始 + duration
  - duration_min: patient.weekly_pattern.service_minutes (なければ 35)
  - course_template_code: User シートの course
  - sub_office_code: 空
"""

from __future__ import annotations

import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

# diff_pfv.py の関数を再利用
sys.path.insert(0, str(Path(__file__).parent))
from diff_pfv import (  # noqa: E402
    load_db_patient_codes,
    load_db_pfv,
    norm,
    parse_user_schedule,
)

from app.services.patient_excel.schema import (  # noqa: E402
    PATIENT_COLUMNS,
    PFV_COLUMNS,
    SHEET_PATIENTS,
    SHEET_PFV,
    SHEET_WEEKLY,
    WEEKLY_COLUMNS,
)

SCRIPT_DIR = Path(__file__).parent
DB_XLSX = SCRIPT_DIR / "_after_course_fix_db_export.xlsx"
OUT_XLSX = SCRIPT_DIR / "_pfv_additions.xlsx"

WEEKDAY_EN_TO_JP_SHORT = {
    "mon": "月",
    "tue": "火",
    "wed": "水",
    "thu": "木",
    "fri": "金",
    "sat": "土",
    "sun": "日",
}


def _hidx(header: list, prefix: str) -> int | None:
    for i, c in enumerate(header):
        if c and str(c).startswith(prefix):
            return i
    return None


def load_weekly_pattern_db(db_xlsx: Path) -> dict[str, dict]:
    """patient_code → weekly_pattern dict (service_minutes / time_type)."""
    wb = load_workbook(db_xlsx, data_only=True, read_only=True)
    ws = wb[SHEET_WEEKLY]
    header = None
    h: dict[str, int] = {}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            for k in ["patient_code", "サービス時間", "時間タイプ"]:
                idx = _hidx(header, k)
                if idx is not None:
                    h[k] = idx
            continue
        if not row or "patient_code" not in h:
            continue
        code = norm(row[h["patient_code"]])
        if not code.startswith("P"):
            continue
        svc = row[h.get("サービス時間", -1)] if h.get("サービス時間") is not None else None
        tt = row[h.get("時間タイプ", -1)] if h.get("時間タイプ") is not None else None
        out[code] = {
            "service_minutes": int(svc) if svc and str(svc).strip() else 35,
            "time_type": norm(tt) if tt else "終日",
        }
    wb.close()
    return out


def time_add(start_hhmm: str, duration_min: int) -> str:
    """'HH:MM' + duration → 'HH:MM'."""
    dt = datetime.strptime(start_hhmm, "%H:%M")
    dt2 = dt + timedelta(minutes=duration_min)
    return dt2.strftime("%H:%M")


def _col_idx(cols: list[dict], key: str) -> int:
    for i, c in enumerate(cols, 1):
        if c["key"] == key:
            return i
    raise KeyError(key)


def _write_header(ws: Worksheet, cols: list[dict]) -> None:
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col["header"])


def main() -> None:
    print(f"DB xlsx: {DB_XLSX}")
    name_to_code = load_db_patient_codes(DB_XLSX)
    code_to_name = {v: k for k, v in name_to_code.items()}
    db_pfv = load_db_pfv(DB_XLSX, code_to_name)
    print(f"DB patients: {len(name_to_code)}, PFV: {len(db_pfv)}")

    # User schedule 再 parse (fuzzy match 込み)
    resolved, unresolved = parse_user_schedule(name_to_code)
    print(f"User resolved: {len(resolved)}, unresolved: {len(unresolved)}")

    if unresolved:
        print("WARNING: unresolved があります:", file=sys.stderr)
        for u in unresolved[:5]:
            print(f"  {u}", file=sys.stderr)

    # キー化して User のみを抽出
    user_keys: dict[tuple, list[dict]] = {}
    for r in resolved:
        k = (r["patient_code"], r["weekday"], r["course"], r["start"])
        user_keys.setdefault(k, []).append(r)
    db_keys: set[tuple] = set()
    for r in db_pfv:
        # 完全 key (start も含む) で
        db_keys.add((r["patient_code"], r["weekday"], r["course"], r["start"]))

    # 追加候補 = User key が DB に無い
    additions: list[dict] = []
    for k, rs in user_keys.items():
        if k not in db_keys:
            additions.append(rs[0])  # 同 key 重複時は最初の 1 件
    # ソート: patient_code, weekday, course, start
    weekday_order = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
    additions.sort(
        key=lambda x: (
            x["patient_code"],
            weekday_order.get(x["weekday"], 99),
            x["course"],
            x["start"],
        )
    )

    print(f"PFV 追加候補: {len(additions)}")

    # DB の既存 (patient_code, weekday) ごとの max slot_index を取得
    next_slot: dict[tuple[str, str], int] = {}
    for r in db_pfv:
        k = (r["patient_code"], r["weekday"])
        # DB 側 slot_index を取らないので、count + 1 で代用 (= 同 weekday 内の件数)
        next_slot[k] = next_slot.get(k, 0) + 1  # = 既存件数

    # weekly_pattern 取得
    weekly_db = load_weekly_pattern_db(DB_XLSX)
    print(f"weekly_pattern loaded: {len(weekly_db)}")

    # Build workbook
    wb = Workbook()
    ws_p: Worksheet = wb.active  # type: ignore
    ws_p.title = SHEET_PATIENTS
    _write_header(ws_p, PATIENT_COLUMNS)
    print("患者マスタシート: header のみ (患者は触らない)")

    ws_f: Worksheet = wb.create_sheet(title=SHEET_PFV)
    _write_header(ws_f, PFV_COLUMNS)
    row_idx = 2
    for r in additions:
        code = r["patient_code"]
        wd_en = r["weekday"]
        wd_jp = WEEKDAY_EN_TO_JP_SHORT.get(wd_en, wd_en)
        course = r["course"]
        start = r["start"]
        wp = weekly_db.get(code, {"service_minutes": 35, "time_type": "終日"})
        dur = wp["service_minutes"]
        tt = wp["time_type"]
        end = time_add(start, dur)
        # slot_index (DB CHECK 制約: 0 or 1 のみ. importer も >1 で reject)
        k = (code, wd_en)
        cur_slot = next_slot.get(k, 0)
        if cur_slot > 1:
            print(
                f"WARNING: slot_index overflow {code}/{wd_en}={cur_slot} (max=1), skipping",
                file=sys.stderr,
            )
            continue
        next_slot[k] = cur_slot + 1

        name = code_to_name.get(code, "")

        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "patient_id"), value="")
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "patient_code"), value=code)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "patient_name"), value=name)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "weekday"), value=wd_jp)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "slot_index"), value=cur_slot)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "mode"), value="normal")
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "time_type"), value=tt)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "start_time"), value=start)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "end_time"), value=end)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "duration_min"), value=dur)
        ws_f.cell(row=row_idx, column=_col_idx(PFV_COLUMNS, "course_template_code"), value=course)
        row_idx += 1
    print(f"PFV シート: {row_idx - 2} 行")

    ws_w: Worksheet = wb.create_sheet(title=SHEET_WEEKLY)
    _write_header(ws_w, WEEKLY_COLUMNS)
    print("希望訪問パターンシート: header のみ (weekly_pattern は触らない)")

    wb.save(OUT_XLSX)
    print("\n=== 出力 ===")
    print(f"{OUT_XLSX} ({OUT_XLSX.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
