"""Phase G-8: User シート枠組み vs DB の course_template_id 突合 + 更新 SQL 生成.

User 「スケジュール枠組み（仮）」 シートでは「都賀A」コースに INAGE 拠点の患者
(石塚/田中/山﨑/岩船/兼行/野口/髙山/松戸 etc.) が大量に含まれている。
しかし現 DB では PFV.course_template_id が INAGE-A (patient.primary_office_id ベース)
を指しているため、reset-to-fixed 実行後にこれらの患者が INAGE-A コーステーブルに
表示されてしまい、User の「都賀A」と乖離している。

本スクリプトは:
  1. User シートから (patient_code, weekday, start_time, office, course_label) を抽出
     (= diff_pfv.parse_user_schedule を再利用 / office_code 追加版)
  2. DB エクスポート (_final_db_export.xlsx) から (patient_code, weekday, start_time,
     course_label) と patient.primary_office_code を読み込み、現状の course_template
     (office_code, label) を再構成
  3. キー (patient_code, weekday, start_time) で突合し、User と DB が違うものを
     「変更要」としてリストアップ。TSUGA に存在しない label (B/C/D/E) は TSUGA-A に
     集約 + warning
  4. _diff_pfv_course_assignment.md (report) と _update_pfv_course_assignment.sql
     (UPDATE 文) + _rollback_pfv_course_assignment.sql (revert UPDATE 文) を出力

入力:
  - User: Sampledata/2026.05.19段階最新ユーザーシート/スケジュール手動 のコピー.xlsx
    シート: スケジュール枠組み（仮）
  - DB: backend/scripts/_final_db_export.xlsx

出力:
  - backend/scripts/_diff_pfv_course_assignment.md
  - backend/scripts/_update_pfv_course_assignment.sql
  - backend/scripts/_rollback_pfv_course_assignment.sql

DB 直接変更はしない (dry-run). SQL を確認後、別途 VPS で psql 実行する想定.

⚠️ 非冪等: 本 script は **未適用の DB** に対して 1 回だけ実行する想定.
  - 比較ロジックは `db_office = patient.primary_office_code` 由来のため、
    一部適用後の DB (PFV.course_template_id が既に TSUGA-A に変わっている行が
    混在) を再 export して再 run すると、「現 course」を primary_office から
    再構成する関係で **誤検出 / 件数ズレ / 部分的に逆方向 UPDATE 生成**
    が発生し得る.
  - 再 run が必要な場合は事前に rollback SQL or pg_dump で完全に戻すこと.
"""

from __future__ import annotations

from pathlib import Path

from diff_pfv import (
    DB_XLSX,
    DB_XLSX_NEW,
    hidx,
    load_db_patient_codes,
    norm,
    norm_time,
    parse_user_schedule,
)
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
OUTPUT_MD = SCRIPT_DIR / "_diff_pfv_course_assignment.md"
OUTPUT_SQL = SCRIPT_DIR / "_update_pfv_course_assignment.sql"
OUTPUT_ROLLBACK_SQL = SCRIPT_DIR / "_rollback_pfv_course_assignment.sql"

# course_template (office_code, label) → course_template_id.
# VPS 本番 DB から取得済み (Phase G-8 開始時点、 ハードコード).
# INAGE: A/B/C/D/E/M / TSUGA: A/M のみ.
COURSE_TEMPLATE_IDS: dict[tuple[str, str], str] = {
    ("INAGE", "A"): "a4f2276e-dd8c-410d-9319-9eb7cbe8e084",
    ("INAGE", "B"): "c687fb66-d351-473e-b408-b95993e11546",
    ("INAGE", "C"): "e0b5d1aa-e75a-4aca-b6d8-99dd5985e8df",
    ("INAGE", "D"): "b4d33c20-9297-4c13-bbf1-7de110a689da",
    ("INAGE", "E"): "566eca75-86b2-4f90-8aa9-47309826b694",
    ("INAGE", "M"): "8ead9177-01eb-4358-8b50-5d716c614384",
    ("TSUGA", "A"): "d1bba8ad-408d-42db-84a8-6ffb54b83bdc",
    ("TSUGA", "M"): "de73343c-9957-4527-b16f-bed4cba3f529",
}

WEEKDAY_SHORT_TO_EN = {
    "月": "mon",
    "火": "tue",
    "水": "wed",
    "木": "thu",
    "金": "fri",
    "土": "sat",
    "日": "sun",
}
WEEKDAY_EN_TO_INT = {
    "mon": 0,
    "tue": 1,
    "wed": 2,
    "thu": 3,
    "fri": 4,
    "sat": 5,
    "sun": 6,
}


def load_patient_master(db_xlsx: Path) -> dict[str, dict]:
    """patient_code → {patient_id (UUID str), name, primary_office_code} の dict."""
    wb = load_workbook(db_xlsx, data_only=True, read_only=True)
    ws = wb["患者マスタ"]
    header = None
    h: dict[str, int] = {}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            for key in ["patient_id", "patient_code", "患者名", "拠点コード"]:
                idx = hidx(header, key)
                if idx is not None:
                    h[key] = idx
            continue
        if not row or "patient_code" not in h:
            continue
        code = norm(row[h["patient_code"]])
        if not code.startswith("P"):
            continue
        out[code] = {
            "patient_id": norm(row[h["patient_id"]]) if "patient_id" in h else "",
            "name": norm(row[h["患者名"]]) if "患者名" in h else "",
            "primary_office_code": (norm(row[h["拠点コード"]]) if "拠点コード" in h else ""),
        }
    wb.close()
    return out


def load_db_pfv_full(db_xlsx: Path) -> list[dict]:
    """DB PFV を {patient_code, weekday_en, start, course_label, patient_id_str} で抽出.

    diff_pfv.load_db_pfv と違い、patient_id (UUID) と元 row の重複検出に必要な
    最小キーのみ。end/duration は対象外.
    """
    wb = load_workbook(db_xlsx, data_only=True, read_only=True)
    ws = wb["固定訪問パターン"]
    header = None
    h: dict[str, int] = {}
    out: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            for key in [
                "patient_id",
                "patient_code",
                "曜日",
                "開始時刻",
                "course_template_code",
            ]:
                idx = hidx(header, key)
                if idx is not None:
                    h[key] = idx
            continue
        if not row or "patient_code" not in h:
            continue
        code = norm(row[h["patient_code"]])
        if not code.startswith("P"):
            continue
        wd_raw = norm(row[h["曜日"]])
        wd = WEEKDAY_SHORT_TO_EN.get(wd_raw, wd_raw.lower())
        start = norm_time(row[h["開始時刻"]])
        label = norm(row[h["course_template_code"]])
        pid = norm(row[h["patient_id"]]) if "patient_id" in h else ""
        out.append(
            {
                "patient_code": code,
                "patient_id": pid,
                "weekday": wd,
                "start": start,
                "course_label": label,
            }
        )
    wb.close()
    return out


def resolve_user_course_for_tsuga(user_office: str, user_course: str) -> tuple[str, str, list[str]]:
    """User の (office, course_label) を course_template が実在する組合せに正規化.

    TSUGA は A/M のみ存在 → B/C/D/E は TSUGA-A に集約 + warning を返す.
    """
    warnings: list[str] = []
    if user_office == "TSUGA" and user_course not in ("A", "M"):
        warnings.append(
            f"TSUGA-{user_course} template が無いため TSUGA-A に集約 (User シート上は TSUGA-{user_course})"
        )
        return "TSUGA", "A", warnings
    return user_office, user_course, warnings


def main() -> None:
    db_xlsx = DB_XLSX_NEW if DB_XLSX_NEW.exists() else DB_XLSX
    print(f"DB xlsx: {db_xlsx}")

    # 患者マスタ
    pm = load_patient_master(db_xlsx)
    print(f"DB patients: {len(pm)}")
    name_to_code = load_db_patient_codes(db_xlsx)

    # DB PFV
    db_pfv = load_db_pfv_full(db_xlsx)
    print(f"DB PFV: {len(db_pfv)}")

    # User シート抽出 (parse_user_schedule は diff_pfv.py 拡張版を利用)
    user_resolved, unresolved = parse_user_schedule(name_to_code)
    print(f"User: resolved={len(user_resolved)}, unresolved={len(unresolved)}")

    # キー (patient_code, weekday, start) → user entry (office, course)
    # 中点ペア / 親子 でも resolve 後は patient_code が分離されるため key 衝突は基本的に無いが
    # 念のため list で保持し最初のものを採用 (= 重複時 warning)
    user_by_key: dict[tuple[str, str, str], list[dict]] = {}
    for r in user_resolved:
        k = (r["patient_code"], r["weekday"], r["start"])
        user_by_key.setdefault(k, []).append(r)

    # DB 側も同様にキー化
    db_by_key: dict[tuple[str, str, str], list[dict]] = {}
    for r in db_pfv:
        k = (r["patient_code"], r["weekday"], r["start"])
        db_by_key.setdefault(k, []).append(r)

    # 突合
    changes: list[dict] = []
    skipped_no_db: list[tuple[str, str, str]] = []
    skipped_no_user: list[tuple[str, str, str]] = []
    match_count = 0
    cross_office_inage_to_tsuga = 0  # INAGE 拠点 → User=TSUGA (= 今回の主目的)
    cross_office_tsuga_to_inage = 0  # TSUGA 拠点 → User=INAGE (= 想定外)
    tsuga_fallback_count = 0  # TSUGA B/C/D/E → TSUGA-A に集約された件数

    for k, urows in user_by_key.items():
        patient_code, wd, start = k
        if k not in db_by_key:
            skipped_no_db.append(k)
            continue
        ur = urows[0]
        dr = db_by_key[k][0]
        pinfo = pm.get(patient_code, {})
        primary_office = pinfo.get("primary_office_code", "")
        patient_name = pinfo.get("name", "")

        user_office, user_course, warns = resolve_user_course_for_tsuga(ur["office"], ur["course"])
        if warns:
            tsuga_fallback_count += 1

        # DB 側の現コース: patient.primary_office + course_template_code(letter)
        # 拠点コードは patient マスタの「拠点コード」列 = "INAGE"/"TSUGA"
        db_office = primary_office
        db_course = dr["course_label"]
        current_key = (db_office, db_course)
        new_key = (user_office, user_course)

        if current_key == new_key:
            match_count += 1
            continue

        # cross-office 判定
        if primary_office == "INAGE" and user_office == "TSUGA":
            cross_office_inage_to_tsuga += 1
        elif primary_office == "TSUGA" and user_office == "INAGE":
            cross_office_tsuga_to_inage += 1

        new_template_id = COURSE_TEMPLATE_IDS.get(new_key)
        current_template_id = COURSE_TEMPLATE_IDS.get(current_key)
        notes: list[str] = []
        notes.extend(warns)
        if new_template_id is None:
            notes.append(f"WARNING: new template_id 不明 ({new_key})")
        if current_template_id is None:
            notes.append(f"WARNING: 現 template_id 不明 ({current_key})")
        if primary_office == "TSUGA" and user_office == "INAGE":
            notes.append("WARNING: 想定外の逆方向 cross-office (患者=TSUGA, User=INAGE)")

        changes.append(
            {
                "patient_code": patient_code,
                "patient_id": dr["patient_id"],
                "patient_name": patient_name,
                "primary_office": primary_office,
                "weekday": wd,
                "weekday_int": WEEKDAY_EN_TO_INT.get(wd, -1),
                "start": start,
                "current_office": db_office,
                "current_label": db_course,
                "current_template_id": current_template_id,
                "new_office": user_office,
                "new_label": user_course,
                "new_template_id": new_template_id,
                "raw_cell": ur["raw_cell"],
                "is_pair": ur["is_pair"],
                "user_office_raw": ur["office"],
                "user_course_raw": ur["course"],
                "notes": notes,
            }
        )

    for k in db_by_key:
        if k not in user_by_key:
            skipped_no_user.append(k)

    # ===== 出力: Markdown report =====
    total = len(user_by_key)
    with OUTPUT_MD.open("w", encoding="utf-8") as f:
        f.write("# Phase G-8: PFV course_template_id 一括変更 dry-run\n\n")
        f.write(
            f"- User: `{Path('スケジュール手動 のコピー.xlsx').name}` シート「スケジュール枠組み（仮）」\n"
        )
        f.write(f"- DB: `{db_xlsx.name}` シート「固定訪問パターン」「患者マスタ」\n")
        f.write("- 突合キー: (patient_code, weekday, start_time)\n\n")

        f.write("## サマリ\n\n")
        f.write(f"- User キー総数: **{total}** 件\n")
        f.write(f"- DB と一致 (変更不要): **{match_count}** 件\n")
        f.write(f"- 変更要: **{len(changes)}** 件\n")
        f.write(
            f"  - うち INAGE 拠点 → User=TSUGA (今回の主目的 cross-office): **{cross_office_inage_to_tsuga}** 件\n"
        )
        f.write(
            f"  - うち TSUGA 拠点 → User=INAGE (想定外逆方向 cross-office, 要確認): **{cross_office_tsuga_to_inage}** 件\n"
        )
        f.write(
            f"  - うち TSUGA-B/C/D/E → TSUGA-A 集約 (template 不在): **{tsuga_fallback_count}** 件\n"
        )
        f.write(
            f"- User のみ (DB に無い key, 既存 diff_pfv.md と整合): **{len(skipped_no_db)}** 件\n"
        )
        f.write(
            f"- DB のみ (User に無い key, 既存 diff_pfv.md と整合): **{len(skipped_no_user)}** 件\n\n"
        )

        f.write("## 変更要リスト\n\n")
        if not changes:
            f.write("_該当なし_\n\n")
        else:
            f.write(
                "| patient_code | name | p_office | weekday | start | 現 course (DB) | 新 course (User) | 備考 |\n"
            )
            f.write("|---|---|---|---|---|---|---|---|\n")
            for c in sorted(
                changes,
                key=lambda x: (x["patient_code"], x["weekday_int"], x["start"]),
            ):
                cur = f"{c['current_office']}-{c['current_label']}"
                new = f"{c['new_office']}-{c['new_label']}"
                note = "; ".join(c["notes"]) if c["notes"] else ""
                f.write(
                    f"| {c['patient_code']} | {c['patient_name']} | {c['primary_office']} | "
                    f"{c['weekday']} | {c['start']} | {cur} | {new} | {note} |\n"
                )
            f.write("\n")

        if skipped_no_db:
            f.write("## (参考) User にあるが DB に無い key — 既存 diff_pfv で報告済み\n\n")
            f.write("| patient_code | weekday | start |\n|---|---|---|\n")
            for k in skipped_no_db:
                f.write(f"| {k[0]} | {k[1]} | {k[2]} |\n")
            f.write("\n")
        if skipped_no_user:
            f.write("## (参考) DB にあるが User に無い key — 既存 diff_pfv で報告済み\n\n")
            f.write("| patient_code | weekday | start |\n|---|---|---|\n")
            for k in skipped_no_user:
                f.write(f"| {k[0]} | {k[1]} | {k[2]} |\n")
            f.write("\n")

    # 適用対象 (new_template_id が解決済み) の changes のみ
    applicable = [c for c in changes if c["new_template_id"] is not None]
    applicable_sorted = sorted(
        applicable, key=lambda x: (x["patient_code"], x["weekday_int"], x["start"])
    )

    def _start_sql(start: str) -> str:
        """HH:MM → HH:MM:SS に揃える (DB の time 型に合わせる)."""
        return start if start.count(":") == 2 else f"{start}:00"

    # ===== 出力: forward UPDATE SQL =====
    with OUTPUT_SQL.open("w", encoding="utf-8") as f:
        f.write("-- Phase G-8: PFV course_template_id 一括更新 (dry-run 生成)\n")
        f.write(
            "-- User 「スケジュール枠組み（仮）」 シート通りに、 各 PFV.course_template_id を変更.\n"
        )
        f.write("-- patient.primary_office_id は変更しない (cross-office assignment を許容).\n")
        f.write(
            "-- course_template_id は VPS 本番 DB から取得済み:\n"
            "--   INAGE: A=a4f2276e..., B=c687fb66..., C=e0b5d1aa..., D=b4d33c20..., E=566eca75..., M=8ead9177...\n"
            "--   TSUGA: A=d1bba8ad..., M=de73343c...\n"
            "-- TSUGA-B/C/D/E は template が存在しないため TSUGA-A に集約 (上記 report の備考列参照).\n\n"
        )
        f.write("-- ============================================================\n")
        f.write("-- 本 SQL は dry-run 用. 本番適用手順:\n")
        f.write("--   1. psql に paste して BEGIN; → 各 UPDATE 確認 → verification 実行\n")
        f.write(
            "--   2. 件数が期待通り (= "
            + str(len(applicable_sorted))
            + " 件) なら -- COMMIT のコメントアウトを外し ROLLBACK を消す\n"
        )
        f.write("--   3. 適用後問題発生時は _rollback_pfv_course_assignment.sql を実行\n")
        f.write("--   4. それでもダメなら pg_dump 復元: /tmp/carelink_pre_g8_20260520_110942.sql\n")
        f.write("-- ============================================================\n")
        f.write("-- ⚠️ WHERE 句には mode='normal' AND slot_index=0 を含める:\n")
        f.write(
            "--   - mode: 同 (patient_id, weekday, start_time) で special mode 行が存在し得る\n"
        )
        f.write("--   - slot_index: UNIQUE 制約は (patient_id, mode, weekday, slot_index).\n")
        f.write(
            "--     Phase G-1 で全 86 名 requires_multiple_staff=false が確定済み → slot_index=0 のみ存在\n"
        )
        f.write("--     (NOT NULL カラムなので 0 で限定)\n")
        f.write("-- ============================================================\n\n")
        f.write("BEGIN;\n\n")
        for c in applicable_sorted:
            cur = f"{c['current_office']}-{c['current_label']}"
            new = f"{c['new_office']}-{c['new_label']}"
            pair_mark = " (ペア)" if c["is_pair"] else ""
            note_inline = " -- " + " | ".join(c["notes"]) if c["notes"] else ""
            f.write(
                f"-- {c['patient_code']} {c['patient_name']}{pair_mark} "
                f"{c['weekday']} {c['start']} {cur} -> {new}{note_inline}\n"
            )
            wd_int = c["weekday_int"]
            start_sql = _start_sql(c["start"])
            f.write("-- EXPECT: 1 row affected\n")
            f.write(
                f"UPDATE patient_fixed_visits SET course_template_id = '{c['new_template_id']}'"
                f" WHERE patient_id = '{c['patient_id']}'"
                f" AND mode = 'normal'"
                f" AND slot_index = 0"
                f" AND weekday = {wd_int}"
                f" AND start_time = '{start_sql}';\n\n"
            )

        # SKIP 行 (new_template_id 不明) を末尾にコメントとして残す
        skipped = [c for c in changes if c["new_template_id"] is None]
        if skipped:
            f.write("-- ===== SKIP (new template_id 不明 / 適用対象外) =====\n")
            for c in sorted(
                skipped, key=lambda x: (x["patient_code"], x["weekday_int"], x["start"])
            ):
                cur = f"{c['current_office']}-{c['current_label']}"
                new = f"{c['new_office']}-{c['new_label']}"
                note_inline = " | " + " | ".join(c["notes"]) if c["notes"] else ""
                f.write(
                    f"-- SKIP {c['patient_code']} {c['patient_name']} "
                    f"{c['weekday']} {c['start']} {cur} -> {new}{note_inline}\n"
                )
            f.write("\n")

        # ===== verification queries =====
        f.write("-- ============================================================\n")
        f.write(
            f"-- 適用件数確認 ({len(applicable_sorted)} 件のはず. COMMIT する前に必ず実行して件数を目視確認):\n"
        )
        f.write("-- ============================================================\n")
        f.write("SELECT 'changes_total' AS metric, COUNT(*) AS count\n")
        f.write("  FROM patient_fixed_visits\n")
        f.write(" WHERE mode = 'normal'\n")
        f.write("   AND slot_index = 0\n")
        f.write("   AND (patient_id, weekday, start_time) IN (\n")
        for i, c in enumerate(applicable_sorted):
            wd_int = c["weekday_int"]
            start_sql = _start_sql(c["start"])
            sep = "," if i < len(applicable_sorted) - 1 else ""
            f.write(
                f"      ('{c['patient_id']}', {wd_int}, '{start_sql}'){sep}"
                f"  -- {c['patient_code']} {c['patient_name']} {c['weekday']}\n"
            )
        f.write("   );\n\n")

        f.write("-- 拠点別の variation 確認 (TSUGA-A 件数が増えているか):\n")
        f.write("SELECT\n")
        f.write("  o.code AS office,\n")
        f.write("  ct.label AS course,\n")
        f.write("  COUNT(*) AS pfv_count\n")
        f.write("FROM patient_fixed_visits pfv\n")
        f.write("JOIN course_templates ct ON pfv.course_template_id = ct.id\n")
        f.write("JOIN offices o ON ct.office_id = o.id\n")
        f.write("WHERE pfv.mode = 'normal'\n")
        f.write("GROUP BY 1, 2\n")
        f.write("ORDER BY 1, 2;\n\n")

        f.write("-- (参考) 全体 course_template_id 分布:\n")
        f.write(
            "-- SELECT course_template_id, COUNT(*) FROM patient_fixed_visits WHERE mode='normal' GROUP BY 1;\n\n"
        )

        f.write("-- 問題なければ COMMIT, 問題あれば ROLLBACK\n")
        f.write("-- COMMIT;\n")
        f.write("ROLLBACK;\n")

    # ===== 出力: rollback SQL =====
    with OUTPUT_ROLLBACK_SQL.open("w", encoding="utf-8") as f:
        f.write("-- Phase G-8: PFV course_template_id ロールバック SQL (dry-run 生成)\n")
        f.write("-- ============================================================\n")
        f.write("-- 本 SQL は、 _update_pfv_course_assignment.sql を本番 COMMIT した後に\n")
        f.write("-- 問題が発生した場合、 psql で実行することで restore 可能.\n")
        f.write("-- 各 UPDATE は new_template_id → current_template_id (適用前の値) に戻す.\n")
        f.write("-- ============================================================\n")
        f.write("-- ⚠️ 注意:\n")
        f.write(
            "--   - 本 rollback SQL は、 forward UPDATE が COMMIT された直後の DB 状態を前提とする.\n"
        )
        f.write(
            "--     forward UPDATE 後に手動編集が入ると、 rollback で意図しない上書きが起こる可能性あり.\n"
        )
        f.write(
            "--   - そのため rollback は forward COMMIT 直後 (= まだ運用が動いていない時) に限り推奨.\n"
        )
        f.write("--   - 大きく状況が変わっている場合は pg_dump restore を優先:\n")
        f.write("--       /tmp/carelink_pre_g8_20260520_110942.sql\n")
        f.write("-- ============================================================\n")
        f.write("-- WHERE 句には forward と同じく mode='normal' AND slot_index=0 を含む.\n")
        f.write("-- ============================================================\n\n")
        f.write("BEGIN;\n\n")
        # rollback 対象: current_template_id が解決済みのもの (= 元の値が分かるもの)
        rollback_targets = [c for c in applicable_sorted if c["current_template_id"] is not None]
        for c in rollback_targets:
            cur = f"{c['current_office']}-{c['current_label']}"
            new = f"{c['new_office']}-{c['new_label']}"
            pair_mark = " (ペア)" if c["is_pair"] else ""
            f.write(
                f"-- {c['patient_code']} {c['patient_name']}{pair_mark} "
                f"{c['weekday']} {c['start']} REVERT {new} -> {cur}\n"
            )
            wd_int = c["weekday_int"]
            start_sql = _start_sql(c["start"])
            f.write("-- EXPECT: 1 row affected\n")
            f.write(
                f"UPDATE patient_fixed_visits SET course_template_id = '{c['current_template_id']}'"
                f" WHERE patient_id = '{c['patient_id']}'"
                f" AND mode = 'normal'"
                f" AND slot_index = 0"
                f" AND weekday = {wd_int}"
                f" AND start_time = '{start_sql}';\n\n"
            )

        # current_template_id 不明な行があれば WARNING として残す
        unknown_current = [c for c in applicable_sorted if c["current_template_id"] is None]
        if unknown_current:
            f.write("-- ===== WARNING: 元 template_id 不明のため rollback 不能 =====\n")
            for c in unknown_current:
                cur = f"{c['current_office']}-{c['current_label']}"
                f.write(
                    f"-- SKIP {c['patient_code']} {c['patient_name']} "
                    f"{c['weekday']} {c['start']} (current={cur} の template_id 不明)\n"
                )
            f.write("\n")

        f.write("-- ============================================================\n")
        f.write(f"-- rollback 件数確認 ({len(rollback_targets)} 件のはず):\n")
        f.write("-- ============================================================\n")
        f.write("SELECT 'rollback_total' AS metric, COUNT(*) AS count\n")
        f.write("  FROM patient_fixed_visits\n")
        f.write(" WHERE mode = 'normal'\n")
        f.write("   AND slot_index = 0\n")
        f.write("   AND (patient_id, weekday, start_time) IN (\n")
        for i, c in enumerate(rollback_targets):
            wd_int = c["weekday_int"]
            start_sql = _start_sql(c["start"])
            sep = "," if i < len(rollback_targets) - 1 else ""
            f.write(
                f"      ('{c['patient_id']}', {wd_int}, '{start_sql}'){sep}"
                f"  -- {c['patient_code']} {c['patient_name']} {c['weekday']}\n"
            )
        f.write("   );\n\n")

        f.write("-- 問題なければ COMMIT, 問題あれば ROLLBACK\n")
        f.write("-- COMMIT;\n")
        f.write("ROLLBACK;\n")

    # ===== コンソール サマリ =====
    print()
    print("=== 出力 ===")
    print(f"  {OUTPUT_MD}")
    print(f"  {OUTPUT_SQL}")
    print(f"  {OUTPUT_ROLLBACK_SQL}")
    print()
    print(f"User キー総数: {total}")
    print(f"  一致 (変更不要): {match_count}")
    print(f"  変更要: {len(changes)}")
    print(f"    INAGE→TSUGA cross-office: {cross_office_inage_to_tsuga}")
    print(f"    TSUGA→INAGE cross-office (想定外): {cross_office_tsuga_to_inage}")
    print(f"    TSUGA-B/C/D/E → TSUGA-A 集約: {tsuga_fallback_count}")
    print(f"  User のみ (DB に無い): {len(skipped_no_db)}")
    print(f"  DB のみ (User に無い): {len(skipped_no_user)}")


if __name__ == "__main__":
    main()
