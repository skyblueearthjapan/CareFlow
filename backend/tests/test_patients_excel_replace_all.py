"""Tests for /api/v1/patients/import-export/replace-all (完全置換).

Coverage:
  1. RBAC: admin 200 / manager 403 / staff 403
  2. dry_run=True: DB 変更なし、削除対象一覧が返る
  3. apply: Excel に無い既存患者が soft delete される
  4. apply: 空セルが NULL で上書きされる
  5. apply: 既存 PFV が全件物理削除されて Excel から再投入される
  6. apply: error 1 件で全 rollback (transaction_applied=False, 422)
  7. apply: 新規患者 + 既存更新 + 削除 + 復活 mix
  8. summary 件数の正確性
"""

from __future__ import annotations

from datetime import UTC, datetime, time
from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.core.security import create_access_token, hash_password
from app.models import (
    CourseTemplate,
    Office,
    Patient,
    PatientFixedVisit,
    User,
)
from app.services.patient_excel.schema import (
    PATIENT_COL_INDEX,
    PATIENT_COLUMNS,
    PFV_COL_INDEX,
    PFV_COLUMNS,
    SHEET_PATIENTS,
    SHEET_PFV,
    SHEET_WEEKLY,
    WEEKLY_COL_INDEX,
    WEEKLY_COLUMNS,
)


async def _make_user(db, email: str, role: str) -> User:
    user = User(email=email, password_hash=hash_password("pw"), role=role)
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


def _bearer(user: User) -> dict[str, str]:
    token = create_access_token(subject=user.id, role=user.role, staff_id=user.staff_id)
    return {"Authorization": f"Bearer {token}"}


async def _make_office(db, code: str, name: str | None = None) -> Office:
    o = Office(code=code, name=name or code)
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return o


async def _make_patient(db, **overrides) -> Patient:
    defaults = {
        "code": "P-DEFAULT",
        "name": "デフォルト患者",
        "sex": "male",
        "status": "active",
        "address": "千葉市稲毛区test",
    }
    defaults.update(overrides)
    p = Patient(**defaults)
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


async def _make_pfv(
    db,
    *,
    patient_id,
    weekday: int = 0,
    slot_index: int = 0,
    course_template_id=None,
    start_time=time(9, 0),
    duration_min: int = 30,
):
    pfv = PatientFixedVisit(
        patient_id=patient_id,
        mode="normal",
        weekday=weekday,
        slot_index=slot_index,
        start_time=start_time,
        duration_min=duration_min,
        course_template_id=course_template_id,
    )
    db.add(pfv)
    await db.commit()
    await db.refresh(pfv)
    return pfv


def _patient_headers() -> list[str]:
    return [str(c["header"]) for c in PATIENT_COLUMNS]


def _pfv_headers() -> list[str]:
    return [str(c["header"]) for c in PFV_COLUMNS]


def _build_workbook_bytes(
    *,
    patient_rows: list[dict] | None = None,
    pfv_rows: list[dict] | None = None,
) -> bytes:
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    for col_idx, header in enumerate(_patient_headers(), start=1):
        ws_p.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(patient_rows or [], start=2):
        for col_key, idx in PATIENT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_p.cell(row=r_idx, column=idx + 1, value=v)

    ws_f = wb.create_sheet(title=SHEET_PFV)
    for col_idx, header in enumerate(_pfv_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(pfv_rows or [], start=2):
        for col_key, idx in PFV_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload(client, user, *, content: bytes, dry_run: bool = True):
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    return await client.post(
        f"/api/v1/patients/import-export/replace-all?dry_run={'true' if dry_run else 'false'}",
        headers=_bearer(user),
        files=files,
    )


# ---------------------------------------------------------------------------
# 1) RBAC
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_admin_ok(client, db) -> None:
    admin = await _make_user(db, "ra-admin@example.com", "admin")
    content = _build_workbook_bytes(patient_rows=[])
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text


@pytest.mark.asyncio
async def test_replace_all_manager_forbidden(client, db) -> None:
    manager = await _make_user(db, "ra-mgr@example.com", "manager")
    content = _build_workbook_bytes(patient_rows=[])
    res = await _upload(client, manager, content=content, dry_run=True)
    assert res.status_code == 403


@pytest.mark.asyncio
async def test_replace_all_staff_forbidden(client, db) -> None:
    staff_user = await _make_user(db, "ra-staff@example.com", "staff")
    content = _build_workbook_bytes(patient_rows=[])
    res = await _upload(client, staff_user, content=content, dry_run=True)
    assert res.status_code == 403


# ---------------------------------------------------------------------------
# 2) dry_run
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_dry_run_returns_soft_delete_list_no_db_change(client, db) -> None:
    admin = await _make_user(db, "ra-dry@example.com", "admin")
    # 既存 alive 2 名
    p_keep = await _make_patient(db, code="P-RA-KEEP", name="残す")
    p_keep_id = p_keep.id
    p_drop = await _make_patient(db, code="P-RA-DROP", name="消える")
    p_drop_id = p_drop.id

    # Excel には KEEP のみ
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(p_keep_id),
                "patient_code": "P-RA-KEEP",
                "name": "残す",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is False
    assert body["summary"]["patients_to_soft_delete"] == 1
    assert body["summary"]["patients_to_update"] >= 0  # name 同じなら noop
    assert len(body["patients_to_soft_delete_preview"]) == 1
    assert body["patients_to_soft_delete_preview"][0]["patient_code"] == "P-RA-DROP"

    # DB 変更なし
    db.expire_all()
    p_drop_after = await db.get(Patient, p_drop_id)
    assert p_drop_after is not None
    assert p_drop_after.deleted_at is None


# ---------------------------------------------------------------------------
# 3) apply: 削除
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_apply_soft_deletes_missing_patients(client, db) -> None:
    admin = await _make_user(db, "ra-del@example.com", "admin")
    p_keep = await _make_patient(db, code="P-RA-K1", name="残す")
    p_keep_id = p_keep.id
    p_drop = await _make_patient(db, code="P-RA-D1", name="消える")
    p_drop_id = p_drop.id
    # PFV 付き
    pfv_drop = await _make_pfv(db, patient_id=p_drop_id, weekday=0)
    pfv_drop_id = pfv_drop.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(p_keep_id),
                "patient_code": "P-RA-K1",
                "name": "残す",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_to_soft_delete"] == 1

    db.expire_all()
    p_keep_after = await db.get(Patient, p_keep_id)
    assert p_keep_after is not None and p_keep_after.deleted_at is None
    p_drop_after = await db.get(Patient, p_drop_id)
    assert p_drop_after is not None and p_drop_after.deleted_at is not None  # soft deleted
    # PFV は物理削除
    pfv_after = await db.get(PatientFixedVisit, pfv_drop_id)
    assert pfv_after is None


# ---------------------------------------------------------------------------
# 4) apply: 空セル → NULL 上書き
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_apply_blank_cells_overwrite_to_null(client, db) -> None:
    admin = await _make_user(db, "ra-null@example.com", "admin")
    patient = await _make_patient(
        db,
        code="P-RA-NUL",
        name="保持",
        sex="male",
        status="active",
        address="千葉市稲毛区test",
        kana="ホジ",
        note="残ってるはずの備考",
    )
    patient_id = patient.id

    # Excel では kana/note を空セルに (= NULL に上書きされるはず)
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-RA-NUL",
                "name": "保持",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test",
                # kana / note 空セル
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True

    db.expire_all()
    p_after = await db.get(Patient, patient_id)
    assert p_after is not None
    assert p_after.kana is None  # NULL で上書きされた
    assert p_after.note is None


# ---------------------------------------------------------------------------
# 5) apply: PFV 全件 replace
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_apply_pfv_replaces_all(client, db) -> None:
    admin = await _make_user(db, "ra-pfv@example.com", "admin")
    patient = await _make_patient(db, code="P-RA-PFV", name="PFV テスト")
    patient_id = patient.id
    # 既存 PFV 2 件
    pfv1 = await _make_pfv(db, patient_id=patient_id, weekday=0)
    pfv2 = await _make_pfv(db, patient_id=patient_id, weekday=1)
    pfv1_id = pfv1.id
    pfv2_id = pfv2.id

    # Excel では 1 件のみ (異なる weekday)
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-RA-PFV",
                "name": "PFV テスト",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test",
            }
        ],
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "水",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "10:00",
                "duration_min": 45,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_to_create"] == 1
    # 既存 PFV 2 件 + (削除対象 patient はゼロ) → pfv_to_replace=2
    assert body["summary"]["pfv_to_replace"] == 2

    db.expire_all()
    # 旧 PFV は両方消えてる
    assert await db.get(PatientFixedVisit, pfv1_id) is None
    assert await db.get(PatientFixedVisit, pfv2_id) is None
    # 新 PFV が 1 件入ってる
    remaining = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(remaining) == 1
    assert remaining[0].weekday == 2  # 水
    assert remaining[0].duration_min == 45


# ---------------------------------------------------------------------------
# 6) apply: error → 全 rollback (422)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_apply_error_rolls_back_everything(client, db) -> None:
    """完全置換は atomic. error 1 件あれば全 rollback (422 返却)."""
    admin = await _make_user(db, "ra-rollback@example.com", "admin")
    patient = await _make_patient(
        db, code="P-RA-RB", name="ロールバック前", sex="male", status="active", address="addr"
    )
    patient_id = patient.id

    # 正常 update 1 件 + error 1 件 (enum 違反) → atomic で何も適用されないこと
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-RA-RB",
                "name": "更新後",
                "sex": "male",
                "status": "active",
                "address": "addr-new",
            },
            {
                "patient_code": "P-RA-BAD",
                "name": "エラー行",
                "sex": "INVALID",  # enum 違反
                "status": "active",
                "address": "addr",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 422, res.text
    # 既存患者は更新されていないこと
    db.expire_all()
    p_after = await db.get(Patient, patient_id)
    assert p_after is not None
    assert p_after.name == "ロールバック前"  # 更新されていない


# ---------------------------------------------------------------------------
# 7) apply mix: new + update + delete
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_apply_mix_new_update_delete(client, db) -> None:
    admin = await _make_user(db, "ra-mix@example.com", "admin")
    p_upd = await _make_patient(db, code="P-RA-U", name="旧", sex="male", status="active")
    p_upd_id = p_upd.id
    p_del = await _make_patient(db, code="P-RA-X", name="消える", sex="male", status="active")
    p_del_id = p_del.id

    content = _build_workbook_bytes(
        patient_rows=[
            # update 行
            {
                "patient_id": str(p_upd_id),
                "patient_code": "P-RA-U",
                "name": "新",
                "sex": "male",
                "status": "active",
                "address": "addr-new",
            },
            # new 行
            {
                "patient_code": "P-RA-N",
                "name": "新規",
                "sex": "female",
                "status": "active",
                "address": "addr-n",
            },
            # P-RA-X は Excel に無い → soft delete
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    s = body["summary"]
    assert s["patients_to_create"] == 1
    assert s["patients_to_update"] == 1
    assert s["patients_to_soft_delete"] == 1
    assert s["patients_error"] == 0

    db.expire_all()
    new_p = await db.scalar(select(Patient).where(Patient.code == "P-RA-N"))
    assert new_p is not None
    assert new_p.deleted_at is None
    upd_p = await db.get(Patient, p_upd_id)
    assert upd_p.name == "新"
    del_p = await db.get(Patient, p_del_id)
    assert del_p.deleted_at is not None


# ---------------------------------------------------------------------------
# 8) 復活: soft-deleted 同じ code を Excel で再登録
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_resurrects_soft_deleted_patient(client, db) -> None:
    admin = await _make_user(db, "ra-res@example.com", "admin")
    patient = await _make_patient(db, code="P-RA-RES", name="削除前")
    patient.deleted_at = datetime.now(UTC)
    await db.commit()
    original_id = patient.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-RA-RES",
                "name": "復活後",
                "sex": "male",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True

    db.expire_all()
    p_after = await db.get(Patient, original_id)
    assert p_after is not None
    assert p_after.deleted_at is None
    assert p_after.name == "復活後"


# ---------------------------------------------------------------------------
# CRITICAL 回帰: update 行の NOT NULL カラムが空セル → error (rollback)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_required_field_null_in_update_row_errors(client, db) -> None:
    """update パスで NOT NULL カラム (name/status) が空セルの場合、
    error として扱われ全件 rollback (422) される."""
    admin = await _make_user(db, "ra-null-update@example.com", "admin")
    patient = await _make_patient(
        db, code="P-RA-NUL-UPD", name="既存患者", sex="male", status="active", address="addr"
    )
    patient_id = patient.id

    # patient_id 明示 (update パス) + name を空セルに
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-RA-NUL-UPD",
                # name 欠落 (空セル = None)
                "sex": "male",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    # error 行があるので 422
    assert res.status_code == 422, res.text

    # DB は変更されていないこと
    db.expire_all()
    p_after = await db.get(Patient, patient_id)
    assert p_after is not None
    assert p_after.name == "既存患者"  # rollback 確認


# ---------------------------------------------------------------------------
# E-4 parity: replace_all.py も importer.py と同じく以下を吸収する.
#   - time_type が空セル → DEFAULT_TIME_TYPE で吸収 (error にしない)
#   - course_template_code が拠点不在 → course_template_id=None で保存
#   - round-trip (export → そのまま replace-all import) で error 0 件
# 並列構成: test_patients_excel_api.py の対応する E-4 テストを参照.
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_accepts_empty_time_type_with_default(client, db) -> None:
    """time_type 空セルは default で吸収され error にならない (E-4 parity)."""
    admin = await _make_user(db, "ra-e4-tt@example.com", "admin")
    patient = await _make_patient(db, code="P-RA-E4-TT", name="time_type空でも OK")
    patient_id = patient.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-RA-E4-TT",
                "name": "time_type空でも OK",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test",
            }
        ],
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                # time_type を意図的に省略.
                "start_time": "09:00",
                "duration_min": 30,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_to_create"] == 1
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1


@pytest.mark.asyncio
async def test_replace_all_fallback_for_unknown_course_template_code(client, db) -> None:
    """course_template_code が患者拠点に存在しない場合、error にならず
    course_template_id=None で取り込む (E-4 parity)."""
    admin = await _make_user(db, "ra-e4-ct@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    patient = await _make_patient(
        db,
        code="P-RA-E4-CTU",
        name="unknown ct",
        primary_office_id=office.id,
    )
    patient_id = patient.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-RA-E4-CTU",
                "name": "unknown ct",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test",
                "office_code": "INAGE",
            }
        ],
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
                "duration_min": 30,
                "course_template_code": "D",  # 拠点に存在しない
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_to_create"] == 1
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1
    assert rows[0].course_template_id is None  # 拠点不在 code は剥がされる


@pytest.mark.asyncio
async def test_replace_all_export_then_import_round_trip_completes_without_errors(
    client, db
) -> None:
    """export → そのまま replace-all import で error 0 件 (E-4 parity).

    replace_all は import (差分パス) と違い「全 PFV を物理削除して Excel から
    再投入」する仕様なので、round-trip 結果は全 PFV が pfv_to_create に倒れる.
    cross-office template は同じく空欄に倒される (= silent loss だが運用上許容).
    """
    admin = await _make_user(db, "ra-e4-roundtrip@example.com", "admin")
    office_a = await _make_office(db, code="INAGE", name="稲毛")
    office_b = await _make_office(db, code="TSUGA", name="都賀")
    template_a = CourseTemplate(office_id=office_a.id, label="M")
    template_b = CourseTemplate(office_id=office_b.id, label="C")
    db.add_all([template_a, template_b])
    await db.commit()
    await db.refresh(template_a)
    await db.refresh(template_b)

    patient_a = await _make_patient(
        db, code="P-RA-RT-A", name="患者A", primary_office_id=office_a.id
    )
    patient_b = await _make_patient(
        db, code="P-RA-RT-B", name="患者B", primary_office_id=office_a.id
    )
    # 患者 A: weekly_pattern エントリ無し (time_type 解決不能だが PFV には影響しない).
    patient_a.weekly_pattern = None
    await db.commit()
    # PFV-A: same-office template
    await _make_pfv(db, patient_id=patient_a.id, weekday=0, course_template_id=template_a.id)
    # PFV-B: 患者 B (拠点 INAGE) に対する PFV だが template_b (拠点 TSUGA) を指す
    # クロスオフィス参照. export で course_template_code が空欄に倒れ、replace-all
    # で再投入時に course_template_id=None になる (silent loss).
    await _make_pfv(db, patient_id=patient_b.id, weekday=2, course_template_id=template_b.id)

    # export
    export_res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert export_res.status_code == 200
    exported_bytes = export_res.content

    # そのまま replace-all (dry_run=True で計上のみ).
    res = await _upload(client, admin, content=exported_bytes, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    summary = body["summary"]
    # round-trip では error 0 件 (E-4 で吸収済).
    assert summary["patients_error"] == 0, body
    assert summary["pfv_error"] == 0, body
    # Excel に居る 2 名は削除対象にならない.
    assert summary["patients_to_soft_delete"] == 0, body
    # PFV は全件再投入される. 既存 2 件 / 再投入 2 件.
    assert summary["pfv_to_replace"] == 2, body
    assert summary["pfv_to_create"] == 2, body


# ---------------------------------------------------------------------------
# Phase E-7 (gap P0-1): requires_multiple_staff replace-all 往復
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_roundtrip_requires_multiple_staff(client, db) -> None:
    """replace-all セマンティクスでも requires_multiple_staff が保持される."""
    admin = await _make_user(db, "ra-rms@example.com", "admin")
    p = await _make_patient(
        db,
        code="P-RA-RMS",
        name="複数必須",
        status="active",
        address="千葉市稲毛区",
        requires_multiple_staff=True,
    )
    pid = p.id
    # export → そのまま replace-all で再 import
    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/patients/import-export/replace-all?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    assert p_after.requires_multiple_staff is True


# ---------------------------------------------------------------------------
# Phase E-7 (gap P2): クロスオフィス course_template warning
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_crossoffice_course_template_emits_warning(client, db, caplog) -> None:
    """異なる拠点の course_template を持つ PFV を export すると logger.warning が出る."""
    import logging as _logging

    admin = await _make_user(db, "ex-co-warn@example.com", "admin")
    office_inage = await _make_office(db, code="INAGE", name="稲毛")
    office_tsuga = await _make_office(db, code="TSUGA", name="都賀")
    # patient は INAGE 所属
    p = await _make_patient(
        db,
        code="P-CO-WARN",
        name="クロスオフィス対象",
        status="active",
        address="千葉市稲毛区",
        primary_office_id=office_inage.id,
    )
    # course_template は TSUGA 拠点 → クロスオフィス参照
    ct_tsuga = CourseTemplate(office_id=office_tsuga.id, label="X")
    db.add(ct_tsuga)
    await db.commit()
    await db.refresh(ct_tsuga)
    await _make_pfv(db, patient_id=p.id, weekday=0, course_template_id=ct_tsuga.id)

    # caplog で warning を捕捉
    with caplog.at_level(_logging.WARNING, logger="app.services.patient_excel.exporter"):
        export_res = await client.get(
            "/api/v1/patients/import-export/export", headers=_bearer(admin)
        )
    assert export_res.status_code == 200
    # warning が 1 件以上出ていること.
    warning_msgs = [r.message for r in caplog.records if r.levelno == _logging.WARNING]
    assert any("クロスオフィス" in m for m in warning_msgs), warning_msgs
    # Phase E-7 (gap P2): response header にクロスオフィス warning 件数が出る.
    # operator が export 結果から気付けるように常に header を返す.
    assert "X-Excel-Crossoffice-Warnings-Count" in export_res.headers
    assert int(export_res.headers["X-Excel-Crossoffice-Warnings-Count"]) >= 1


@pytest.mark.asyncio
async def test_export_crossoffice_roundtrip_zero_error(client, db) -> None:
    """クロスオフィス参照があっても round-trip 0 エラー仕様は維持される (warning のみ)."""
    admin = await _make_user(db, "ex-co-rt@example.com", "admin")
    office_inage = await _make_office(db, code="INAGE", name="稲毛")
    office_tsuga = await _make_office(db, code="TSUGA", name="都賀")
    p = await _make_patient(
        db,
        code="P-CO-RT",
        name="クロスオフィス往復",
        status="active",
        address="千葉市稲毛区",
        primary_office_id=office_inage.id,
    )
    ct_tsuga = CourseTemplate(office_id=office_tsuga.id, label="Y")
    db.add(ct_tsuga)
    await db.commit()
    await db.refresh(ct_tsuga)
    await _make_pfv(db, patient_id=p.id, weekday=0, course_template_id=ct_tsuga.id)

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    # 通常 import で round-trip — error が 0 件であることを確認.
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    assert body["summary"]["patients_error"] == 0, body
    assert body["summary"]["pfv_error"] == 0, body


# ---------------------------------------------------------------------------
# Phase G-48: 完全置換でも 日本語化 + code 突合 + weekly 統合シートを扱える
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_replace_all_japanese_enum_new_patient(client, db) -> None:
    """完全置換でも日本語 enum ラベルで新規患者を登録できる (DB は英語値)."""
    admin = await _make_user(db, "ra-g48-ja@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-RA-G48-JA",
                "name": "日本語完全置換",
                "sex": "女性",
                "status": "稼働",
                "insurance": "医療保険",
                "sex_restriction": "女性のみ",
                "address": "千葉市稲毛区test",
                "requires_multiple_staff": "はい",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p = await db.scalar(select(Patient).where(Patient.code == "P-RA-G48-JA"))
    assert p is not None
    assert p.sex == "female"
    assert p.status == "active"
    assert p.insurance == "medical"
    assert p.sex_restriction == "female_only"
    assert p.requires_multiple_staff is True


@pytest.mark.asyncio
async def test_replace_all_update_via_code_without_id(client, db) -> None:
    """完全置換でも patient_id 空 + code で既存 alive 患者を突合 UPDATE する."""
    admin = await _make_user(db, "ra-g48-code@example.com", "admin")
    p = await _make_patient(
        db, code="P-RA-G48-CODE", name="旧", sex="male", status="active", address="addr"
    )
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-RA-G48-CODE",
                "name": "新",
                "sex": "男性",
                "status": "稼働",
                "address": "addr-new",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    # code 突合で update. 削除対象 0 (Excel に居る).
    assert body["summary"]["patients_to_update"] == 1
    assert body["summary"]["patients_to_create"] == 0
    assert body["summary"]["patients_to_soft_delete"] == 0
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.name == "新"
    assert p_after.deleted_at is None


@pytest.mark.asyncio
async def test_replace_all_weekly_integrated_sheet(client, db) -> None:
    """完全置換でも統合シートの weekly_pattern を反映する."""
    admin = await _make_user(db, "ra-g48-wk@example.com", "admin")
    p = await _make_patient(db, code="P-RA-G48-WK", name="weekly統合")
    p.weekly_pattern = None
    await db.commit()
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-RA-G48-WK",
                "name": "weekly統合",
                "sex": "男性",
                "status": "稼働",
                "address": "千葉市稲毛区test",
                "weekly_time_type": "時間帯",
                "preferred_start": "09:00",
                "preferred_end": "11:00",
                "preferred_weekdays": "火,木",
                "visit_frequency": "毎週",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern["time_type"] == "時間帯"
    assert set(p_after.weekly_pattern["preferred_weekdays"]) == {"Tue", "Thu"}
    assert p_after.weekly_pattern["visit_frequency"] == "every"


@pytest.mark.asyncio
async def test_replace_all_full_japanese_roundtrip_zero_change(client, db) -> None:
    """完全置換 round-trip: 日本語 export → 無編集 import で患者 0 変更.

    (PFV は仕様上全件再投入されるが、患者本体と weekly は noop になる.)
    """
    admin = await _make_user(db, "ra-g48-rt@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    p = await _make_patient(
        db,
        code="P-RA-G48-RT",
        name="往復",
        kana="オウフク",
        sex="female",
        status="active",
        insurance="care",
        address="千葉市稲毛区test",
        sex_restriction="male_only",
        requires_multiple_staff=True,
        primary_office_id=office.id,
    )
    p.weekly_pattern = {
        "frequency_per_week": 2,
        "visit_frequency": "biweekly",
        "preferred_weekdays": ["Mon", "Fri"],
        "service_minutes": 30,
        "time_type": "固定",
        "preferred_start": "08:30",
    }
    await db.commit()

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    res = await _upload(client, admin, content=export_res.content, dry_run=True)
    assert res.status_code == 200, res.text
    summary = res.json()["summary"]
    assert summary["patients_error"] == 0, res.json()
    assert summary["patients_to_update"] == 0, res.json()
    assert summary["patients_to_soft_delete"] == 0, res.json()


@pytest.mark.asyncio
async def test_replace_all_roundtrip_preserves_unmanaged_weekly_keys(client, db) -> None:
    """CRITICAL 回帰防止 (Phase G-48 hotfix, replace-all 経路):

    バックアップ復元経路 (replace-all) でも weekly_pattern の管理外キー
    (entries / staff_count) を持つ患者を export → 無編集 import したとき、

      (a) patients_to_update == 0 (患者本体・weekly は noop)
      (b) DB 上の entries / staff_count が **保持** される

    ことを担保する. replace-all は丸ごと置換セマンティクスだが、weekly_pattern
    だけは merge し管理外キーを破壊しない. (PFV は仕様上全件再投入されるが患者
    本体・weekly は noop.)

    visit_frequency は canonical 値 (biweekly) を入れ round-trip が noop に
    なる (= legacy-JP 正規化が canonical をそのまま通す) ことも確認する.
    """
    admin = await _make_user(db, "ra-g48-rt-unmanaged@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    p = await _make_patient(
        db,
        code="P-RA-G48-ENTRIES",
        name="管理外キー保持",
        sex="female",
        status="active",
        address="千葉市稲毛区test",
        primary_office_id=office.id,
    )
    original_entries = [
        {
            "weekday": "Mon",
            "time_type": "時間帯",
            "preferred_start": "10:00",
            "preferred_end": "11:30",
            "service_minutes": 45,
            "staff_count": 1,
        },
        {
            "weekday": "Wed",
            "time_type": "固定",
            "preferred_start": "14:00",
            "preferred_end": "15:00",
            "service_minutes": 60,
            "staff_count": 2,  # 2 名体制エントリ
        },
    ]
    p.weekly_pattern = {
        # 管理 8 キー (round-trip で復元される)
        "frequency_per_week": 2,
        "visit_frequency": "biweekly",  # canonical → JP 正規化が noop で通すこと
        "preferred_weekdays": ["Mon", "Wed"],
        "service_minutes": 45,
        "time_type": "時間帯",
        "preferred_start": "10:00",
        "preferred_end": "11:30",
        # 管理外キー (merge で保持されねばならない)
        "staff_count": 1,
        "entries": original_entries,
    }
    await db.commit()
    pid = p.id

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    exported = export_res.content

    # (a) dry_run で患者本体・weekly が noop であることを確認
    dry_res = await _upload(client, admin, content=exported, dry_run=True)
    assert dry_res.status_code == 200, dry_res.text
    summary = dry_res.json()["summary"]
    assert summary["patients_error"] == 0, dry_res.json()
    assert summary["patients_to_update"] == 0, dry_res.json()
    assert summary["patients_to_soft_delete"] == 0, dry_res.json()

    # apply して DB 上の entries / staff_count が保持されることを確認
    apply_res = await _upload(client, admin, content=exported, dry_run=False)
    assert apply_res.status_code == 200, apply_res.text
    assert apply_res.json()["summary"]["patients_to_update"] == 0, apply_res.json()

    db.expire_all()
    p_after = await db.get(Patient, pid)
    # (b) 管理外キーが保持される
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern.get("staff_count") == 1
    assert p_after.weekly_pattern.get("entries") == original_entries
    wed = next(e for e in p_after.weekly_pattern["entries"] if e["weekday"] == "Wed")
    assert wed["staff_count"] == 2
    # canonical visit_frequency も維持
    assert p_after.weekly_pattern["visit_frequency"] == "biweekly"
    assert p_after.weekly_pattern["frequency_per_week"] == 2


@pytest.mark.asyncio
async def test_replace_all_legacy_weekly_sheet_delete_clears(client, db) -> None:
    """後方互換 / HIGH 回帰防止 (Phase G-48 hotfix-2, replace-all 経路):

    旧 3 シート構成の「希望訪問パターン」シートに ``<DELETE>`` 行を含むファイルを
    replace-all import すると、(a) 例外 (TypeError) を出さず 200 で完了し、
    (b) 当該患者の weekly_pattern が None にクリアされること.
    """
    admin = await _make_user(db, "ra-g48-legdel@example.com", "admin")
    p = await _make_patient(db, code="P-RA-LEGDEL", name="旧形式削除")
    p.weekly_pattern = {
        "time_type": "固定",
        "preferred_start": "10:00",
        "preferred_weekdays": ["Tue", "Thu"],
        "staff_count": 2,
        "entries": [{"weekday": "Tue", "start": "10:00"}],
    }
    await db.commit()
    pid = p.id

    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    ws_p.append([str(c["header"]) for c in PATIENT_COLUMNS])
    prow = [None] * len(PATIENT_COLUMNS)
    prow[PATIENT_COL_INDEX["patient_id"]] = str(pid)
    prow[PATIENT_COL_INDEX["patient_code"]] = "P-RA-LEGDEL"
    # replace-all は全行を「最終状態」として扱うため必須列を埋める.
    prow[PATIENT_COL_INDEX["name"]] = "旧形式削除"
    prow[PATIENT_COL_INDEX["sex"]] = "男性"
    prow[PATIENT_COL_INDEX["status"]] = "稼働"
    prow[PATIENT_COL_INDEX["address"]] = "千葉市稲毛区test"
    ws_p.append(prow)
    ws_f = wb.create_sheet(title=SHEET_PFV)
    ws_f.append([str(c["header"]) for c in PFV_COLUMNS])
    # 旧 独立「希望訪問パターン」シートに <DELETE> 行.
    ws_w = wb.create_sheet(SHEET_WEEKLY)
    ws_w.append([str(c["header"]) for c in WEEKLY_COLUMNS])
    wrow = [None] * len(WEEKLY_COLUMNS)
    wrow[WEEKLY_COL_INDEX["patient_id"]] = str(pid)
    wrow[WEEKLY_COL_INDEX["patient_code"]] = "P-RA-LEGDEL"
    wrow[WEEKLY_COL_INDEX["delete_flag"]] = "<DELETE>"
    ws_w.append(wrow)
    buf = BytesIO()
    wb.save(buf)

    # (a) 例外を出さず 200 で完了する.
    res = await _upload(client, admin, content=buf.getvalue(), dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    # (b) weekly_pattern が None にクリアされる.
    assert p_after.weekly_pattern is None
