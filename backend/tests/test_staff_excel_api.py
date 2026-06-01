"""Tests for /api/v1/staff/import-export/* (Excel import / export).

Coverage map (spec §7):
  1.  export: 0 名のとき空テンプレート、N 名のとき N+1 行
  2.  export: shift が staff_id でリンクされている
  3.  import dry_run: 新規行が検出される
  4.  import dry_run: 更新行で changes が diff される
  5.  import dry_run: <DELETE> フラグで delete 候補になる
  6.  import dry_run: <CLEAR> で NULL 上書き候補になる
  7.  import dry_run: 空セルは「変更しない」 (noop)
  8.  import dry_run: バリデーションエラー (UUID 不正、enum 違反、必須欠落)
  9.  import dry_run: DB は変更されない
  10. import apply: 新規 + 更新 + 削除が 1 transaction で反映
  11. import apply: partial commit — error 行は skip、有効行のみ反映
  12. import apply: staff soft-delete で関連 shift も物理削除
  13. import apply: 拠点コード → primary_office_id 解決
  14. import apply: staff_code 重複で error (skip、有効行のみ反映)
  15. RBAC: staff は 403 (3 endpoints すべて)
  16. RBAC: manager OK
  17. template ダウンロード: 1 行のみ
  18. magic word: case-insensitive
  19. shift update via composite key (staff_id, weekday)
  20. shift: 新規スタッフ + その shifts を staff_code リンクで登録 (UUID 不要)
  21. shift: is_on=FALSE のとき start/end 不要で noop / delete
  27. import apply: 全件 error → transaction_applied=False、DB 変更なし
  28. import apply: shift_error 1 件 + staff_new 2 件 → staff は反映、shift は正常分のみ反映
"""

from __future__ import annotations

from datetime import UTC, datetime, time
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.core.security import create_access_token, hash_password
from app.models import (
    Office,
    Staff,
    StaffShift,
    User,
)
from app.services.staff_excel.schema import (
    SHEET_SHIFT,
    SHEET_STAFF,
    SHIFT_COL_INDEX,
    SHIFT_COLUMNS,
    STAFF_COL_INDEX,
    STAFF_COLUMNS,
)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


async def _make_user(db, email: str, role: str) -> User:
    user = User(email=email, password_hash=hash_password("pw"), role=role)
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


def _bearer(user: User) -> dict[str, str]:
    token = create_access_token(subject=user.id, role=user.role, staff_id=user.staff_id)
    return {"Authorization": f"Bearer {token}"}


async def _make_office(db, code: str, name: str | None = None) -> Office:
    o = Office(code=code, name=name or code)
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return o


async def _make_staff(db, **overrides) -> Staff:
    defaults = {
        "code": "S-DEFAULT",
        "name": "デフォルト",
        "status": "active",
        "role": "staff",
    }
    defaults.update(overrides)
    s = Staff(**defaults)
    db.add(s)
    await db.commit()
    await db.refresh(s)
    return s


async def _make_shift(
    db,
    *,
    staff_id,
    weekday: int = 0,
    is_on: bool = True,
    start_time=time(9, 0),
    end_time=time(18, 0),
) -> StaffShift:
    sh = StaffShift(
        staff_id=staff_id,
        weekday=weekday,
        is_on=is_on,
        start_time=start_time,
        end_time=end_time,
    )
    db.add(sh)
    await db.commit()
    await db.refresh(sh)
    return sh


def _staff_headers() -> list[str]:
    return [str(c["header"]) for c in STAFF_COLUMNS]


def _shift_headers() -> list[str]:
    return [str(c["header"]) for c in SHIFT_COLUMNS]


def _build_workbook_bytes(
    *,
    staff_rows: list[dict] | None = None,
    shift_rows: list[dict] | None = None,
) -> bytes:
    """テスト用ヘルパー: 各 dict は { col_key: value } を持つ."""
    wb = Workbook()
    ws_s = wb.active
    ws_s.title = SHEET_STAFF
    for col_idx, header in enumerate(_staff_headers(), start=1):
        ws_s.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(staff_rows or [], start=2):
        for col_key, idx in STAFF_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_s.cell(row=r_idx, column=idx + 1, value=v)

    ws_f = wb.create_sheet(title=SHEET_SHIFT)
    for col_idx, header in enumerate(_shift_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(shift_rows or [], start=2):
        for col_key, idx in SHIFT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload(client, admin, *, content: bytes, dry_run: bool = True):
    """multipart helper. fastapi UploadFile は ``file`` という名前で受ける."""
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    return await client.post(
        f"/api/v1/staff/import-export/import?dry_run={'true' if dry_run else 'false'}",
        headers=_bearer(admin),
        files=files,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


# 1) export: 0 名のとき empty + N 名のとき N+1 行
@pytest.mark.asyncio
async def test_export_empty_db_returns_template_only(client, db) -> None:
    admin = await _make_user(db, "stx-empty@example.com", "admin")
    res = await client.get(
        "/api/v1/staff/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200, res.text
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_STAFF in wb.sheetnames
    ws = wb[SHEET_STAFF]
    assert ws.max_row == 1  # ヘッダー 1 行のみ


@pytest.mark.asyncio
async def test_export_n_staff_writes_n_plus_1_rows(client, db) -> None:
    admin = await _make_user(db, "stx-n@example.com", "admin")
    for code in ("S-EX-001", "S-EX-002", "S-EX-003"):
        await _make_staff(db, code=code, name=f"name-{code}")

    res = await client.get(
        "/api/v1/staff/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_STAFF]
    assert ws.max_row == 4
    codes = [ws.cell(row=r, column=STAFF_COL_INDEX["staff_code"] + 1).value for r in (2, 3, 4)]
    assert set(codes) == {"S-EX-001", "S-EX-002", "S-EX-003"}


# 2) export: shift が staff_id でリンクされている (Phase G-56: 編集用シート 1 行)
@pytest.mark.asyncio
async def test_export_shift_links_to_staff_code(client, db) -> None:
    from app.services.staff_excel.schema import SHEET_SHIFT_EDIT, SHIFT_EDIT_COL_INDEX

    admin = await _make_user(db, "stx-sh@example.com", "admin")
    staff = await _make_staff(db, code="S-SH-LINK", name="シフト持ち")
    await _make_shift(db, staff_id=staff.id, weekday=0)

    res = await client.get(
        "/api/v1/staff/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws_f = wb[SHEET_SHIFT_EDIT]
    assert ws_f.max_row == 2  # 1 header + 1 staff row
    code_cell = ws_f.cell(row=2, column=SHIFT_EDIT_COL_INDEX["staff_code"] + 1).value
    name_cell = ws_f.cell(row=2, column=SHIFT_EDIT_COL_INDEX["staff_name"] + 1).value
    assert code_cell == "S-SH-LINK"
    assert name_cell == "シフト持ち"


# 3) dry_run: 新規行検出
@pytest.mark.asyncio
async def test_import_dry_run_detects_new_staff(client, db) -> None:
    admin = await _make_user(db, "stx-new@example.com", "admin")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-NEW-001",
                "name": "新規 太郎",
                "sex": "male",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is False
    assert body["summary"]["staff_new"] == 1
    assert body["summary"]["staff_error"] == 0
    row = body["staff_rows"][0]
    assert row["operation"] == "new"
    assert row["staff_code"] == "S-NEW-001"


# 4) dry_run: 更新行で changes が diff される
@pytest.mark.asyncio
async def test_import_dry_run_diffs_updates(client, db) -> None:
    admin = await _make_user(db, "stx-upd@example.com", "admin")
    staff = await _make_staff(
        db, code="S-UPD-001", name="旧名前", kana="フルナマエ", status="active", role="staff"
    )
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(staff.id),
                "staff_code": "S-UPD-001",
                "name": "新名前",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["staff_update"] == 1
    row = body["staff_rows"][0]
    assert row["operation"] == "update"
    fields = {c["field"]: c for c in row["changes"]}
    assert "name" in fields
    assert fields["name"]["old_value"] == "旧名前"
    assert fields["name"]["new_value"] == "新名前"


# 5) dry_run: <DELETE> フラグ
@pytest.mark.asyncio
async def test_import_dry_run_detects_delete_flag(client, db) -> None:
    admin = await _make_user(db, "stx-del@example.com", "admin")
    staff = await _make_staff(db, code="S-DEL-001", name="削除予定")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(staff.id),
                "staff_code": "S-DEL-001",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_delete"] == 1
    assert body["staff_rows"][0]["operation"] == "delete"


# 6) dry_run: <CLEAR> で NULL 上書き
@pytest.mark.asyncio
async def test_import_dry_run_clear_to_null(client, db) -> None:
    admin = await _make_user(db, "stx-clr@example.com", "admin")
    staff = await _make_staff(db, code="S-CLR-001", name="クリア対象", note="残ってる備考")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(staff.id),
                "staff_code": "S-CLR-001",
                "name": "クリア対象",
                "status": "active",
                "role": "staff",
                "note": "<CLEAR>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    row = body["staff_rows"][0]
    assert row["operation"] == "update"
    notes = [c for c in row["changes"] if c["field"] == "note"]
    assert len(notes) == 1
    assert notes[0]["new_value"] is None
    assert notes[0]["old_value"] == "残ってる備考"


# 7) dry_run: 空セルは noop
@pytest.mark.asyncio
async def test_import_dry_run_blank_cells_are_noop(client, db) -> None:
    admin = await _make_user(db, "stx-noop@example.com", "admin")
    staff = await _make_staff(
        db,
        code="S-NOOP-001",
        name="変更なし",
        status="active",
        role="staff",
        kana="ヘンコウナシ",
    )
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(staff.id),
                "staff_code": "S-NOOP-001",
                # 他全部空 → 「触らない」
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    row = body["staff_rows"][0]
    assert row["operation"] == "noop"
    assert body["summary"]["staff_noop"] == 1


# 8) dry_run: バリデーションエラー多種
@pytest.mark.asyncio
async def test_import_dry_run_validation_errors(client, db) -> None:
    admin = await _make_user(db, "stx-err@example.com", "admin")
    content = _build_workbook_bytes(
        staff_rows=[
            # 行 2: UUID 不正
            {
                "staff_id": "not-a-uuid",
                "staff_code": "S-BAD-1",
                "name": "uuid不正",
            },
            # 行 3: enum 違反
            {
                "staff_code": "S-BAD-2",
                "name": "enum違反",
                "sex": "INVALID",
                "status": "active",
                "role": "staff",
            },
            # 行 4: 必須欠落 (name)
            {
                "staff_code": "S-BAD-3",
                "status": "active",
                "role": "staff",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_error"] == 3
    ops = {r["row_number"]: r for r in body["staff_rows"]}
    assert ops[2]["operation"] == "error"
    assert ops[3]["operation"] == "error"
    assert ops[4]["operation"] == "error"
    assert "UUID" in ops[2]["error_message"] or "uuid" in ops[2]["error_message"].lower()
    assert "sex" in ops[3]["error_message"].lower() or "性別" in ops[3]["error_message"]
    assert "name" in ops[4]["error_message"].lower() or "新規作成" in ops[4]["error_message"]


# 9) dry_run: DB は変更されない
@pytest.mark.asyncio
async def test_import_dry_run_does_not_modify_db(client, db) -> None:
    admin = await _make_user(db, "stx-noch@example.com", "admin")
    before_count = len((await db.scalars(select(Staff))).all())

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-DRY-001",
                "name": "新規",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200
    assert res.json()["transaction_applied"] is False

    db.expire_all()
    after_count = len((await db.scalars(select(Staff))).all())
    assert after_count == before_count


# 10) apply: 新規 + 更新 + 削除 が 1 TX で反映
@pytest.mark.asyncio
async def test_import_apply_mixed_operations(client, db) -> None:
    admin = await _make_user(db, "stx-apply@example.com", "admin")
    s_upd = await _make_staff(db, code="S-AP-UPD", name="旧", status="active", role="staff")
    s_del = await _make_staff(db, code="S-AP-DEL", name="消す", status="active", role="staff")

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-AP-NEW",
                "name": "新規",
                "sex": "female",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_id": str(s_upd.id),
                "staff_code": "S-AP-UPD",
                "name": "新",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_id": str(s_del.id),
                "staff_code": "S-AP-DEL",
                "delete_flag": "<DELETE>",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"] == {
        "staff_new": 1,
        "staff_update": 1,
        "staff_delete": 1,
        "staff_error": 0,
        "staff_noop": 0,
        "shift_new": 0,
        "shift_update": 0,
        "shift_delete": 0,
        "shift_error": 0,
        "shift_noop": 0,
        # Phase E-7 (gap P1): 勤務例外シート集計フィールド.
        "override_new": 0,
        "override_update": 0,
        "override_delete": 0,
        "override_error": 0,
        "override_noop": 0,
    }

    db.expire_all()
    rows = (
        await db.scalars(select(Staff).where(Staff.code.in_(["S-AP-NEW", "S-AP-UPD", "S-AP-DEL"])))
    ).all()
    by_code = {s.code: s for s in rows}
    assert by_code["S-AP-NEW"].deleted_at is None
    assert by_code["S-AP-UPD"].name == "新"
    assert by_code["S-AP-DEL"].deleted_at is not None  # soft delete


# 11) apply: partial commit — error 行は skip、有効行のみ反映
@pytest.mark.asyncio
async def test_import_apply_partial_commit_skips_error_rows(client, db) -> None:
    """error 1 件 + 正常 update 1 件 → 正常 update は commit、error 行は skip."""
    admin = await _make_user(db, "stx-rb@example.com", "admin")
    s_upd = await _make_staff(db, code="S-RB-UPD", name="旧")
    s_upd_id = s_upd.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(s_upd_id),
                "staff_code": "S-RB-UPD",
                "name": "正常な更新",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-RB-ERR",
                "name": "エラー行",
                "sex": "INVALID",
                "status": "active",
                "role": "staff",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    # 有効な op (update 1) があるので transaction は commit される.
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_error"] == 1
    assert body["summary"]["staff_update"] == 1

    # S-RB-UPD は更新されているはず (partial commit).
    db.expire_all()
    refreshed = await db.get(Staff, s_upd_id)
    assert refreshed.name == "正常な更新"
    # S-RB-ERR は skip されているので DB に存在しない.
    err_row = await db.scalar(select(Staff).where(Staff.code == "S-RB-ERR"))
    assert err_row is None


# 12) apply: staff soft-delete で関連 shift も物理削除
@pytest.mark.asyncio
async def test_staff_delete_cascades_shift_cleanup(client, db) -> None:
    admin = await _make_user(db, "stx-delcasc@example.com", "admin")
    staff = await _make_staff(db, code="S-DELCASC", name="orphan source")
    staff_id = staff.id
    shift = await _make_shift(db, staff_id=staff_id, weekday=0)
    # 復合主キーなので id 列はない; (staff_id, weekday) で識別.

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(staff_id),
                "staff_code": "S-DELCASC",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_delete"] == 1

    db.expire_all()
    staff_after = await db.get(Staff, staff_id)
    assert staff_after is not None
    assert staff_after.deleted_at is not None  # soft delete
    # 関連 shift は物理削除されているはず.
    remaining = (await db.scalars(select(StaffShift).where(StaffShift.staff_id == staff_id))).all()
    assert remaining == []
    # shift fixture も解放しておく
    del shift


# 13) apply: 拠点コード解決
@pytest.mark.asyncio
async def test_import_apply_resolves_office_code(client, db) -> None:
    admin = await _make_user(db, "stx-off@example.com", "admin")
    office = await _make_office(db, code="TSUGA", name="都賀")
    office_id = office.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-OFF-001",
                "name": "拠点指定",
                "status": "active",
                "role": "staff",
                "office_code": "TSUGA",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True

    db.expire_all()
    rows = (await db.scalars(select(Staff).where(Staff.code == "S-OFF-001"))).all()
    assert len(rows) == 1
    assert rows[0].primary_office_id == office_id


# 14) apply: staff_code 重複 → error 行は skip (有効 op が他に無ければ TX 非実行)
@pytest.mark.asyncio
async def test_import_apply_duplicate_code_errors(client, db) -> None:
    admin = await _make_user(db, "stx-dup@example.com", "admin")
    await _make_staff(db, code="S-DUP-001", name="既存")

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-DUP-001",  # 重複!
                "name": "重複新規",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    # 唯一の行が error なので有効 op = 0 → transaction_applied=False.
    assert body["transaction_applied"] is False
    assert body["summary"]["staff_error"] == 1
    assert "S-DUP-001" in body["staff_rows"][0]["error_message"]


# 14b) 同一ファイル内で staff_code が重複していたら 2 行目を error にする
@pytest.mark.asyncio
async def test_import_in_file_duplicate_code_errors(client, db) -> None:
    admin = await _make_user(db, "stx-infile@example.com", "admin")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-INFILE-DUP",
                "name": "1st",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-INFILE-DUP",
                "name": "2nd",
                "status": "active",
                "role": "staff",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_new"] == 1
    assert body["summary"]["staff_error"] == 1


# 15) RBAC: staff は 403
@pytest.mark.asyncio
async def test_import_export_requires_admin_or_manager(client, db) -> None:
    staff_user = await _make_user(db, "stx-staff@example.com", "staff")
    res = await client.get(
        "/api/v1/staff/import-export/export",
        headers=_bearer(staff_user),
    )
    assert res.status_code == 403

    res = await client.get(
        "/api/v1/staff/import-export/template",
        headers=_bearer(staff_user),
    )
    assert res.status_code == 403

    # POST /import も同様に staff は 403.
    content = _build_workbook_bytes(staff_rows=[])
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=true",
        headers=_bearer(staff_user),
        files=files,
    )
    assert res.status_code == 403


# 16) RBAC: manager は import endpoint に到達できる (200).
@pytest.mark.asyncio
async def test_import_allows_manager_role(client, db) -> None:
    manager = await _make_user(db, "stx-mgr@example.com", "manager")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-MGR-OK",
                "name": "manager OK",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, manager, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["staff_new"] == 1


# 17) template ダウンロード: 1 行 (ヘッダーのみ)
@pytest.mark.asyncio
async def test_template_download_has_header_only(client, db) -> None:
    admin = await _make_user(db, "stx-tpl@example.com", "admin")
    res = await client.get(
        "/api/v1/staff/import-export/template",
        headers=_bearer(admin),
    )
    from app.services.staff_excel.schema import SHEET_SHIFT_EDIT

    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_STAFF in wb.sheetnames
    assert SHEET_SHIFT_EDIT in wb.sheetnames
    assert wb[SHEET_STAFF].max_row == 1
    assert wb[SHEET_SHIFT_EDIT].max_row == 1


# 18) magic word: case-insensitive
@pytest.mark.asyncio
async def test_magic_word_case_insensitive(client, db) -> None:
    """<delete> (lower-case) も認識される."""
    admin = await _make_user(db, "stx-magic@example.com", "admin")
    staff = await _make_staff(db, code="S-MAGIC-001", name="ケース不問")

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(staff.id),
                "staff_code": "S-MAGIC-001",
                "delete_flag": "<delete>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_delete"] == 1
    assert body["staff_rows"][0]["operation"] == "delete"


# 19) shift update via composite key (staff_id, weekday)
@pytest.mark.asyncio
async def test_shift_update_via_composite_key(client, db) -> None:
    admin = await _make_user(db, "stx-shup@example.com", "admin")
    staff = await _make_staff(db, code="S-SH-UP", name="SH更新")
    await _make_shift(db, staff_id=staff.id, weekday=2, start_time=time(9, 0), end_time=time(18, 0))
    staff_id = staff.id

    content = _build_workbook_bytes(
        shift_rows=[
            {
                "staff_id": str(staff_id),
                "weekday": "水",
                "is_on": "TRUE",
                "start_time": "10:30",
                "end_time": "19:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["shift_update"] == 1

    db.expire_all()
    refreshed = await db.scalar(
        select(StaffShift).where(StaffShift.staff_id == staff_id, StaffShift.weekday == 2)
    )
    assert refreshed.start_time == time(10, 30)
    assert refreshed.end_time == time(19, 0)


# 20) shift: 新規 staff + その shifts を staff_code リンクで登録 (UUID 不要)
@pytest.mark.asyncio
async def test_new_staff_with_shifts_via_code_link(client, db) -> None:
    admin = await _make_user(db, "stx-link@example.com", "admin")

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-LINK-001",
                "name": "リンク新規",
                "status": "active",
                "role": "staff",
            }
        ],
        shift_rows=[
            {
                # staff_id 空, staff_code で参照
                "staff_code": "S-LINK-001",
                "weekday": "月",
                "is_on": "TRUE",
                "start_time": "09:00",
                "end_time": "18:00",
            },
            {
                "staff_code": "S-LINK-001",
                "weekday": "火",
                "is_on": "TRUE",
                "start_time": "09:00",
                "end_time": "18:00",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_new"] == 1
    assert body["summary"]["shift_new"] == 2

    db.expire_all()
    new_staff = await db.scalar(select(Staff).where(Staff.code == "S-LINK-001"))
    assert new_staff is not None
    shifts = (await db.scalars(select(StaffShift).where(StaffShift.staff_id == new_staff.id))).all()
    assert len(shifts) == 2
    weekdays = sorted(s.weekday for s in shifts)
    assert weekdays == [0, 1]


# 21) shift: is_on=FALSE のとき start/end が空でも error にならない
@pytest.mark.asyncio
async def test_shift_is_on_false_allows_empty_times(client, db) -> None:
    admin = await _make_user(db, "stx-off-day@example.com", "admin")
    staff = await _make_staff(db, code="S-OFFDAY", name="休み日あり")
    staff_id = staff.id  # expire_all 前に id を確保

    content = _build_workbook_bytes(
        shift_rows=[
            {
                "staff_id": str(staff_id),
                "weekday": "日",
                "is_on": "FALSE",
                # start_time / end_time 空 → OK
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["shift_new"] == 1

    db.expire_all()
    sh = await db.scalar(
        select(StaffShift).where(StaffShift.staff_id == staff_id, StaffShift.weekday == 6)
    )
    assert sh is not None
    assert sh.is_on is False


# 22) ファイル size 10MB 超で 413
@pytest.mark.asyncio
async def test_upload_too_large_413(client, db) -> None:
    admin = await _make_user(db, "stx-big@example.com", "admin")
    # 10MB + 1 byte の dummy bytes.
    big = b"x" * (10 * 1024 * 1024 + 1)
    files = {
        "file": (
            "huge.xlsx",
            big,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 413


# 23) 存在しない staff_id を指定したらエラー
@pytest.mark.asyncio
async def test_unknown_staff_id_errors(client, db) -> None:
    admin = await _make_user(db, "stx-unk@example.com", "admin")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(uuid4()),  # 存在しない
                "staff_code": "S-UNK-001",
                "name": "存在しない",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_error"] == 1
    assert "存在" in body["staff_rows"][0]["error_message"]


# 24) HIGH#2 regression: 既存 shift に staff_id+weekday のみ記入 → noop
@pytest.mark.asyncio
async def test_shift_blank_cells_preserve_existing(client, db) -> None:
    """blank cell = 既存値継承の契約. is_on=TRUE, start=09:00, end=18:00 な既存
    shift に対して staff_id + weekday のみの行を import すると、本来 noop に
    なるべき (現状は『勤務=TRUE のとき開始/終了時刻必須』エラーになる回帰防止)."""
    admin = await _make_user(db, "stx-blank@example.com", "admin")
    staff = await _make_staff(db, code="S-BLANK", name="空セル継承")
    await _make_shift(
        db, staff_id=staff.id, weekday=0, is_on=True, start_time=time(9, 0), end_time=time(18, 0)
    )
    staff_id = staff.id

    content = _build_workbook_bytes(
        shift_rows=[
            {
                "staff_id": str(staff_id),
                "weekday": "月",
                # is_on / start_time / end_time すべて空 → 既存値継承で noop
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["shift_error"] == 0, body
    assert body["summary"]["shift_noop"] == 1
    assert body["shift_rows"][0]["operation"] == "noop"


# 25) HIGH#2 regression: 既存 shift で end_time だけ変更 → single-field update
@pytest.mark.asyncio
async def test_shift_partial_update_only_end_time(client, db) -> None:
    admin = await _make_user(db, "stx-partial@example.com", "admin")
    staff = await _make_staff(db, code="S-PART", name="部分更新")
    await _make_shift(
        db, staff_id=staff.id, weekday=1, is_on=True, start_time=time(9, 0), end_time=time(18, 0)
    )
    staff_id = staff.id

    content = _build_workbook_bytes(
        shift_rows=[
            {
                "staff_id": str(staff_id),
                "weekday": "火",
                # is_on / start_time は空 = 既存値継承
                "end_time": "20:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["shift_update"] == 1
    row = body["shift_rows"][0]
    changes = {c["field"]: c for c in row["changes"]}
    assert "end_time" in changes
    assert changes["end_time"]["new_value"] == "20:00"
    # start_time / is_on は変更されていない (= changes に含まれない)
    assert "start_time" not in changes
    assert "is_on" not in changes

    db.expire_all()
    refreshed = await db.scalar(
        select(StaffShift).where(StaffShift.staff_id == staff_id, StaffShift.weekday == 1)
    )
    assert refreshed.start_time == time(9, 0)  # 既存値維持
    assert refreshed.end_time == time(20, 0)  # 更新
    assert refreshed.is_on is True


# 26) HIGH#3 regression: DB レベルの UNIQUE 制約 (partial UNIQUE INDEX)
@pytest.mark.asyncio
async def test_db_unique_constraint_on_staff_code(db) -> None:
    """staff.code に対して deleted_at IS NULL AND code IS NOT NULL の partial
    UNIQUE INDEX が効いていること (concurrent import race condition 防止).

    SQLite 3.8+ で同条件の partial UNIQUE INDEX をサポートするためテストで再現可."""
    from sqlalchemy.exc import IntegrityError

    await _make_staff(db, code="S-UNIQUE-001", name="既存")

    duplicate = Staff(code="S-UNIQUE-001", name="重複", status="active", role="staff")
    db.add(duplicate)
    with pytest.raises(IntegrityError):
        await db.commit()
    await db.rollback()


# 27) partial commit: error 1 件 + new 2 件 → new 2 件のみ DB に反映
@pytest.mark.asyncio
async def test_import_apply_partial_commit_new_rows_with_one_error(client, db) -> None:
    """error 1 件 + new 2 件 で apply → DB に new 2 件のみ反映、error 1 件は skip."""
    admin = await _make_user(db, "stx-partial-mix@example.com", "admin")

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-PC-OK-1",
                "name": "正常 1",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-PC-ERR",
                "name": "エラー (enum 違反)",
                "sex": "INVALID",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-PC-OK-2",
                "name": "正常 2",
                "status": "active",
                "role": "staff",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_new"] == 2
    assert body["summary"]["staff_error"] == 1

    # DB: 正常 2 件のみ反映 (error 行は skip)
    db.expire_all()
    rows = (
        await db.scalars(
            select(Staff).where(Staff.code.in_(["S-PC-OK-1", "S-PC-ERR", "S-PC-OK-2"]))
        )
    ).all()
    codes = sorted(s.code for s in rows)
    assert codes == ["S-PC-OK-1", "S-PC-OK-2"]


# 28) partial commit: 全件 error → transaction_applied=False、DB 変更なし
@pytest.mark.asyncio
async def test_import_apply_all_errors_no_commit(client, db) -> None:
    admin = await _make_user(db, "stx-partial-allerr@example.com", "admin")
    before_count = len((await db.scalars(select(Staff))).all())

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-ALL-ERR-1",
                "name": "エラー 1",
                "sex": "INVALID",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-ALL-ERR-2",
                "name": "エラー 2",
                "status": "INVALID_STATUS",
                "role": "staff",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    # 有効 op = 0 件 → 非 commit.
    assert body["transaction_applied"] is False
    assert body["summary"]["staff_error"] == 2

    # DB は変更なし.
    db.expire_all()
    after_count = len((await db.scalars(select(Staff))).all())
    assert after_count == before_count


# 29) partial commit: shift_error 1 件 + staff_new 2 件 → staff 反映、shift は正常分のみ反映
@pytest.mark.asyncio
async def test_import_apply_partial_commit_mixed_staff_and_shift_errors(client, db) -> None:
    """staff シートに正常 new 2 件、shift シートに error 1 件 + 正常 new 1 件.
    → staff 2 件 + shift 1 件 が commit される (shift error 1 件は skip)."""
    admin = await _make_user(db, "stx-partial-shift@example.com", "admin")

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-MIX-A",
                "name": "スタッフ A",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-MIX-B",
                "name": "スタッフ B",
                "status": "active",
                "role": "staff",
            },
        ],
        shift_rows=[
            {
                # 正常 shift 行 (スタッフ A 紐付け via code)
                "staff_code": "S-MIX-A",
                "weekday": "月",
                "is_on": "TRUE",
                "start_time": "09:00",
                "end_time": "18:00",
            },
            {
                # error 行: weekday の値が候補外
                "staff_code": "S-MIX-B",
                "weekday": "INVALID_DAY",
                "is_on": "TRUE",
                "start_time": "10:00",
                "end_time": "19:00",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_new"] == 2
    assert body["summary"]["shift_new"] == 1
    assert body["summary"]["shift_error"] == 1

    db.expire_all()
    # スタッフ 2 件はどちらも DB にいる.
    staff_rows = (
        await db.scalars(select(Staff).where(Staff.code.in_(["S-MIX-A", "S-MIX-B"])))
    ).all()
    assert {s.code for s in staff_rows} == {"S-MIX-A", "S-MIX-B"}
    by_code = {s.code: s for s in staff_rows}
    # スタッフ A の shift のみ DB に存在 (スタッフ B の error 行は skip).
    shifts_a = (
        await db.scalars(select(StaffShift).where(StaffShift.staff_id == by_code["S-MIX-A"].id))
    ).all()
    assert len(shifts_a) == 1
    shifts_b = (
        await db.scalars(select(StaffShift).where(StaffShift.staff_id == by_code["S-MIX-B"].id))
    ).all()
    assert shifts_b == []


# 30) bug 2: staff_id 空 + staff_code = 既存スタッフ + <DELETE> で削除成功
@pytest.mark.asyncio
async def test_import_apply_delete_via_staff_code_only(client, db) -> None:
    """ユーザーが手で「削除したい code + <DELETE>」だけ書いた Excel でも削除できる."""
    admin = await _make_user(db, "stx-delcode@example.com", "admin")
    staff = await _make_staff(db, code="S-DELCODE-001", name="code削除")
    staff_id = staff.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                # staff_id 空 + staff_code のみ + <DELETE>
                "staff_code": "S-DELCODE-001",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_delete"] == 1
    row = body["staff_rows"][0]
    assert row["operation"] == "delete"
    assert row["staff_code"] == "S-DELCODE-001"

    db.expire_all()
    refreshed = await db.get(Staff, staff_id)
    assert refreshed is not None
    assert refreshed.deleted_at is not None  # soft delete されている


# 31) bug 2: staff_id 空 + 存在しない staff_code + <DELETE> は idempotent noop
@pytest.mark.asyncio
async def test_import_delete_via_unknown_staff_code_is_noop(client, db) -> None:
    """既に削除済み (or そもそも居ない) code を <DELETE> しても error にせず noop."""
    admin = await _make_user(db, "stx-delnoop@example.com", "admin")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-DELNOOP-XYZ",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_noop"] == 1
    assert body["summary"]["staff_error"] == 0
    assert body["staff_rows"][0]["operation"] == "noop"


# 32) bug 2: staff_id 空 + staff_code 空 + <DELETE> は error
@pytest.mark.asyncio
async def test_import_delete_with_both_id_and_code_blank_errors(client, db) -> None:
    admin = await _make_user(db, "stx-delblank@example.com", "admin")
    content = _build_workbook_bytes(
        staff_rows=[
            {
                # 両方空
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["staff_error"] == 1
    msg = body["staff_rows"][0]["error_message"]
    assert "staff_id" in msg and "staff_code" in msg


# 33) resurrection: soft-deleted staff_code を Excel で再 import すると復活する
@pytest.mark.asyncio
async def test_resurrection_via_code_for_soft_deleted_staff(client, db) -> None:
    """UI で削除 (soft delete) → 同じ code を Excel で再 import.

    soft-deleted 行があるまま INSERT すると UNIQUE 制約違反で 409 になるため、
    importer は INSERT ではなく既存行の deleted_at=NULL + 内容更新で復活させる.
    """
    admin = await _make_user(db, "stx-resurrect@example.com", "admin")
    staff = await _make_staff(
        db, code="S-RES-001", name="削除前の名前", status="active", role="staff"
    )
    staff.deleted_at = datetime.now(UTC)
    await db.commit()

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-RES-001",
                "name": "復活後の名前",
                "status": "active",
                "role": "manager",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_update"] == 1
    assert body["summary"]["staff_new"] == 0
    assert body["summary"]["staff_error"] == 0
    row = body["staff_rows"][0]
    assert row["operation"] == "update"
    assert row["staff_code"] == "S-RES-001"
    fields = {c["field"]: c for c in row["changes"]}
    assert "deleted_at" in fields, f"changes に deleted_at が無い: {row['changes']}"
    assert fields["deleted_at"]["new_value"] is None
    assert fields["deleted_at"]["old_value"] is not None
    assert "name" in fields and fields["name"]["new_value"] == "復活後の名前"
    assert "role" in fields and fields["role"]["new_value"] == "manager"

    await db.refresh(staff)
    assert staff.deleted_at is None
    assert staff.name == "復活後の名前"
    assert staff.role == "manager"


# 34) resurrection: 復活時に staff の id が再発番されないこと (UUID 継続性)
@pytest.mark.asyncio
async def test_resurrection_preserves_id(client, db) -> None:
    """復活時に id は元のまま (新しい UUID は発番されない)."""
    admin = await _make_user(db, "stx-resurrect-id@example.com", "admin")
    staff = await _make_staff(db, code="S-RES-ID", name="UUID 継続テスト")
    original_id = staff.id
    staff.deleted_at = datetime.now(UTC)
    await db.commit()

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-RES-ID",
                "name": "復活した",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    row = body["staff_rows"][0]
    assert row["operation"] == "update"
    assert row["staff_id"] == str(original_id)

    await db.refresh(staff)
    assert staff.id == original_id
    assert staff.deleted_at is None
    assert staff.code == "S-RES-ID"
    rows = (await db.scalars(select(Staff).where(Staff.code == "S-RES-ID"))).all()
    assert len(rows) == 1
    assert rows[0].id == original_id


# ---------------------------------------------------------------------------
# Phase E-7 (gap P0-2): Staff.secondary_offices Excel 往復
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_writes_secondary_office_codes(client, db) -> None:
    """secondary_offices を持つ staff を export すると comma-separated に出力される."""
    from app.models import StaffSecondaryOffice

    admin = await _make_user(db, "ex-sec-w@example.com", "admin")
    o_inage = await _make_office(db, code="INAGE")
    o_tsuga = await _make_office(db, code="TSUGA")
    s = await _make_staff(db, code="S-SEC-W", name="サブ持ち", primary_office_id=o_inage.id)
    db.add(StaffSecondaryOffice(staff_id=s.id, office_id=o_tsuga.id))
    await db.commit()

    res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_STAFF]
    val = ws.cell(row=2, column=STAFF_COL_INDEX["secondary_office_codes"] + 1).value
    assert val == "TSUGA"


@pytest.mark.asyncio
async def test_import_new_staff_with_secondary_offices(client, db) -> None:
    """新規 staff の secondary_office_codes をカンマ区切りで投入 → relationship が作られる."""
    from app.models import StaffSecondaryOffice

    admin = await _make_user(db, "im-sec-new@example.com", "admin")
    o_inage = await _make_office(db, code="INAGE")
    o_tsuga = await _make_office(db, code="TSUGA")
    inage_id = o_inage.id
    tsuga_id = o_tsuga.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-NEW-SEC",
                "name": "新規サブ持ち",
                "status": "active",
                "role": "staff",
                "office_code": "INAGE",
                "secondary_office_codes": "TSUGA",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    s = (await db.scalars(select(Staff).where(Staff.code == "S-NEW-SEC"))).first()
    assert s is not None
    sec_rows = (
        await db.scalars(select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == s.id))
    ).all()
    assert {r.office_id for r in sec_rows} == {tsuga_id}
    _ = inage_id


@pytest.mark.asyncio
async def test_import_unknown_secondary_office_code_warning(client, db, caplog) -> None:
    """不明な secondary_office_codes は warning + skip (error にしない)."""
    import logging as _logging

    from app.models import StaffSecondaryOffice

    admin = await _make_user(db, "im-sec-warn@example.com", "admin")
    o_inage = await _make_office(db, code="INAGE")
    o_tsuga = await _make_office(db, code="TSUGA")
    inage_id = o_inage.id
    tsuga_id = o_tsuga.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-SEC-WARN",
                "name": "不明 code 混在",
                "status": "active",
                "role": "staff",
                "office_code": "INAGE",
                # NOTEXIST は不明 code → warning + skip, TSUGA だけ relation が作られる.
                "secondary_office_codes": "NOTEXIST,TSUGA",
            }
        ],
    )
    with caplog.at_level(_logging.WARNING, logger="app.services.staff_excel.importer"):
        res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    # error にはなっていない
    body = res.json()
    assert body["summary"]["staff_error"] == 0
    # warning は出ている
    warnings = [r.message for r in caplog.records if r.levelno == _logging.WARNING]
    assert any("NOTEXIST" in m for m in warnings)
    # DB には TSUGA だけ残る
    db.expire_all()
    s = (await db.scalars(select(Staff).where(Staff.code == "S-SEC-WARN"))).first()
    sec_rows = (
        await db.scalars(select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == s.id))
    ).all()
    assert {r.office_id for r in sec_rows} == {tsuga_id}
    _ = inage_id


@pytest.mark.asyncio
async def test_roundtrip_secondary_offices(client, db) -> None:
    """export → import で secondary_offices が保持される."""
    from app.models import StaffSecondaryOffice

    admin = await _make_user(db, "rt-sec@example.com", "admin")
    o_inage = await _make_office(db, code="INAGE")
    o_tsuga = await _make_office(db, code="TSUGA")
    tsuga_id = o_tsuga.id
    s = await _make_staff(db, code="S-RT-SEC", name="往復", primary_office_id=o_inage.id)
    sid = s.id
    db.add(StaffSecondaryOffice(staff_id=sid, office_id=tsuga_id))
    await db.commit()

    export_res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    assert body["summary"]["staff_error"] == 0
    db.expire_all()
    sec_rows = (
        await db.scalars(select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == sid))
    ).all()
    assert {r.office_id for r in sec_rows} == {tsuga_id}


# ---------------------------------------------------------------------------
# Phase E-7 (gap P1): StaffWeeklyOverride Excel シート往復
# ---------------------------------------------------------------------------


def _override_headers() -> list[str]:
    from app.services.staff_excel.schema import OVERRIDE_COLUMNS

    return [str(c["header"]) for c in OVERRIDE_COLUMNS]


def _build_workbook_with_overrides(
    *,
    staff_rows: list[dict] | None = None,
    shift_rows: list[dict] | None = None,
    override_rows: list[dict] | None = None,
) -> bytes:
    from app.services.staff_excel.schema import OVERRIDE_COL_INDEX, SHEET_OVERRIDE

    wb = Workbook()
    ws_s = wb.active
    ws_s.title = SHEET_STAFF
    for col_idx, header in enumerate(_staff_headers(), start=1):
        ws_s.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(staff_rows or [], start=2):
        for col_key, idx in STAFF_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_s.cell(row=r_idx, column=idx + 1, value=v)
    ws_f = wb.create_sheet(title=SHEET_SHIFT)
    for col_idx, header in enumerate(_shift_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(shift_rows or [], start=2):
        for col_key, idx in SHIFT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)
    ws_o = wb.create_sheet(title=SHEET_OVERRIDE)
    for col_idx, header in enumerate(_override_headers(), start=1):
        ws_o.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(override_rows or [], start=2):
        for col_key, idx in OVERRIDE_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_o.cell(row=r_idx, column=idx + 1, value=v)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_export_writes_overrides_sheet(client, db) -> None:
    """StaffWeeklyOverride を export すると勤務例外（編集用）シートに 1 週 1 行で書き出される."""
    from app.models import StaffWeeklyOverride
    from app.services.staff_excel.schema import OVERRIDE_EDIT_COL_INDEX, SHEET_OVERRIDE_EDIT

    admin = await _make_user(db, "ex-ov@example.com", "admin")
    s = await _make_staff(db, code="S-OV-EX", name="例外持ち")
    db.add(
        StaffWeeklyOverride(
            staff_id=s.id,
            iso_year=2026,
            iso_week=20,
            weekday=2,
            override_type="off",
            reason="私用",
        )
    )
    await db.commit()

    res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_OVERRIDE_EDIT in wb.sheetnames
    ws_o = wb[SHEET_OVERRIDE_EDIT]
    # Phase G-58.1: header(1) + 記入例(2) + 実データ(3) = max_row 3.
    assert ws_o.max_row == 3
    val_year = ws_o.cell(row=3, column=OVERRIDE_EDIT_COL_INDEX["iso_year"] + 1).value
    val_week = ws_o.cell(row=3, column=OVERRIDE_EDIT_COL_INDEX["iso_week"] + 1).value
    # 水曜セル = "休"
    val_wed = ws_o.cell(row=3, column=OVERRIDE_EDIT_COL_INDEX["wed"] + 1).value
    assert val_year == 2026
    assert val_week == 20
    assert val_wed == "休"


@pytest.mark.asyncio
async def test_import_create_override(client, db) -> None:
    """勤務例外シートから新規 override が INSERT される."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "im-ov-new@example.com", "admin")
    s = await _make_staff(db, code="S-OV-NEW", name="新規 override")
    sid = s.id

    content = _build_workbook_with_overrides(
        override_rows=[
            {
                "staff_id": str(sid),
                "iso_year": 2026,
                "iso_week": 21,
                "weekday": "金",
                "override_type": "custom_time",
                "start_time": "10:00",
                "end_time": "16:00",
                "reason": "通院",
            }
        ],
    )
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["override_new"] == 1
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert len(ovs) == 1
    assert ovs[0].iso_year == 2026
    assert ovs[0].iso_week == 21
    assert ovs[0].weekday == 4  # 金
    assert ovs[0].override_type == "custom_time"
    assert ovs[0].start_time == time(10, 0)
    assert ovs[0].end_time == time(16, 0)
    assert ovs[0].reason == "通院"


@pytest.mark.asyncio
async def test_import_update_override(client, db) -> None:
    """既存 override の reason を更新 (UNIQUE key で UPDATE が行われる)."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "im-ov-upd@example.com", "admin")
    s = await _make_staff(db, code="S-OV-UPD", name="更新")
    sid = s.id
    ov = StaffWeeklyOverride(
        staff_id=sid,
        iso_year=2026,
        iso_week=22,
        weekday=0,
        override_type="off",
        reason="旧理由",
    )
    db.add(ov)
    await db.commit()

    content = _build_workbook_with_overrides(
        override_rows=[
            {
                "staff_id": str(sid),
                "iso_year": 2026,
                "iso_week": 22,
                "weekday": "月",
                "override_type": "off",
                "reason": "新理由",
            }
        ],
    )
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["override_update"] == 1
    db.expire_all()
    ov_after = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert len(ov_after) == 1
    assert ov_after[0].reason == "新理由"


@pytest.mark.asyncio
async def test_import_delete_override(client, db) -> None:
    """<DELETE> フラグで既存 override が物理削除される."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "im-ov-del@example.com", "admin")
    s = await _make_staff(db, code="S-OV-DEL", name="削除")
    sid = s.id
    ov = StaffWeeklyOverride(
        staff_id=sid,
        iso_year=2026,
        iso_week=23,
        weekday=1,
        override_type="off",
    )
    db.add(ov)
    await db.commit()

    content = _build_workbook_with_overrides(
        override_rows=[
            {
                "staff_id": str(sid),
                "iso_year": 2026,
                "iso_week": 23,
                "weekday": "火",
                "override_type": "off",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["override_delete"] == 1
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert ovs == []


@pytest.mark.asyncio
async def test_roundtrip_overrides(client, db) -> None:
    """export → import で override が完全保持される (round-trip 0 エラー)."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "rt-ov@example.com", "admin")
    s = await _make_staff(db, code="S-RT-OV", name="往復")
    sid = s.id
    db.add(
        StaffWeeklyOverride(
            staff_id=sid,
            iso_year=2026,
            iso_week=24,
            weekday=3,
            override_type="custom_time",
            start_time=time(8, 30),
            end_time=time(17, 30),
            reason="短縮勤務",
        )
    )
    await db.commit()

    export_res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    # round-trip では override は noop のみで error / new / update なし
    assert body["summary"]["override_error"] == 0
    assert body["summary"]["override_new"] == 0
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert len(ovs) == 1
    assert ovs[0].override_type == "custom_time"
    assert ovs[0].start_time == time(8, 30)
    assert ovs[0].reason == "短縮勤務"


# ---------------------------------------------------------------------------
# Phase E-7 cross-review cleanup (HIGH/MEDIUM/LOW)
# ---------------------------------------------------------------------------


# #3 [MEDIUM]: 旧 2 シート Excel (勤務例外シートなし) との互換性
@pytest.mark.asyncio
async def test_import_legacy_excel_without_override_sheet(client, db) -> None:
    """旧 Excel ファイル (Staff + Shift 2 シートのみ、勤務例外シートなし) を import.

    importer は ``SHEET_OVERRIDE not in wb.sheetnames`` のとき skip するため、
    200 OK + override_* 全 0 + その他は通常通り処理される.
    """
    admin = await _make_user(db, "stx-legacy@example.com", "admin")
    # _build_workbook_bytes は SHEET_STAFF + SHEET_SHIFT の 2 シートのみ作る
    # (= 旧 Excel 相当). _build_workbook_with_overrides との対比で互換性を担保する.
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-LEGACY-001",
                "name": "旧 Excel 互換",
                "status": "active",
                "role": "staff",
            }
        ],
        shift_rows=[
            {
                "staff_code": "S-LEGACY-001",
                "weekday": "月",
                "is_on": "TRUE",
                "start_time": "09:00",
                "end_time": "18:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_new"] == 1
    assert body["summary"]["shift_new"] == 1
    # 勤務例外シートが無いため override_* は全て 0.
    assert body["summary"]["override_new"] == 0
    assert body["summary"]["override_update"] == 0
    assert body["summary"]["override_delete"] == 0
    assert body["summary"]["override_error"] == 0
    assert body["summary"]["override_noop"] == 0


# #4 [MEDIUM]: secondary_offices を <CLEAR> で解除
@pytest.mark.asyncio
async def test_import_secondary_offices_clear_via_marker(client, db) -> None:
    """通常 import で secondary_office_codes 列に <CLEAR> を入れて relationship を解除."""
    from app.models import StaffSecondaryOffice

    admin = await _make_user(db, "im-sec-clr@example.com", "admin")
    o_inage = await _make_office(db, code="INAGE")
    o_tsuga = await _make_office(db, code="TSUGA")
    s = await _make_staff(db, code="S-SEC-CLR", name="サブ解除", primary_office_id=o_inage.id)
    sid = s.id
    db.add(StaffSecondaryOffice(staff_id=sid, office_id=o_tsuga.id))
    await db.commit()
    # 事前確認: 1 件登録されている.
    before = (
        await db.scalars(select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == sid))
    ).all()
    assert len(before) == 1

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-SEC-CLR",
                "secondary_office_codes": "<CLEAR>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_update"] == 1
    assert body["summary"]["staff_error"] == 0

    db.expire_all()
    after = (
        await db.scalars(select(StaffSecondaryOffice).where(StaffSecondaryOffice.staff_id == sid))
    ).all()
    assert after == []


# #6 [LOW]: 1 ファイル内で override の delete / update / new を同時実行
@pytest.mark.asyncio
async def test_import_override_combined_delete_update_new_in_single_file(client, db) -> None:
    """1 ファイル内で既存 override の delete + update + 新規 add を同時実行."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "im-ov-combo@example.com", "admin")
    s = await _make_staff(db, code="S-OV-COMBO", name="複合 op")
    sid = s.id
    # 既存 override 2 件: 1 件は delete 対象、もう 1 件は update 対象.
    db.add(
        StaffWeeklyOverride(
            staff_id=sid,
            iso_year=2026,
            iso_week=30,
            weekday=0,  # 月: delete 予定
            override_type="off",
            reason="削除対象",
        )
    )
    db.add(
        StaffWeeklyOverride(
            staff_id=sid,
            iso_year=2026,
            iso_week=30,
            weekday=1,  # 火: update 予定
            override_type="custom_time",
            start_time=time(9, 0),
            end_time=time(17, 0),
            reason="更新前",
        )
    )
    await db.commit()

    content = _build_workbook_with_overrides(
        override_rows=[
            # delete: 月
            {
                "staff_id": str(sid),
                "iso_year": 2026,
                "iso_week": 30,
                "weekday": "月",
                "override_type": "off",
                "delete_flag": "<DELETE>",
            },
            # update: 火 (時刻変更)
            {
                "staff_id": str(sid),
                "iso_year": 2026,
                "iso_week": 30,
                "weekday": "火",
                "override_type": "custom_time",
                "start_time": "10:00",
                "end_time": "18:00",
                "reason": "更新後",
            },
            # new: 水
            {
                "staff_id": str(sid),
                "iso_year": 2026,
                "iso_week": 30,
                "weekday": "水",
                "override_type": "off",
                "reason": "新規",
            },
        ],
    )
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["override_delete"] == 1
    assert body["summary"]["override_update"] == 1
    assert body["summary"]["override_new"] == 1
    assert body["summary"]["override_error"] == 0

    db.expire_all()
    ovs = (
        await db.scalars(
            select(StaffWeeklyOverride)
            .where(StaffWeeklyOverride.staff_id == sid)
            .order_by(StaffWeeklyOverride.weekday)
        )
    ).all()
    # 月 (delete) は消えており、火 (update) + 水 (new) が残る.
    assert len(ovs) == 2
    weekdays = [ov.weekday for ov in ovs]
    assert weekdays == [1, 2]
    # 火 の中身が更新されている.
    upd = next(ov for ov in ovs if ov.weekday == 1)
    assert upd.start_time == time(10, 0)
    assert upd.end_time == time(18, 0)
    assert upd.reason == "更新後"


# ---------------------------------------------------------------------------
# Phase G-56: 「勤務シフト（編集用）」/「勤務例外（編集用）」 1 スタッフ 1 行
# ---------------------------------------------------------------------------


def _shift_edit_headers() -> list[str]:
    from app.services.staff_excel.schema import SHIFT_EDIT_COLUMNS

    return [str(c["header"]) for c in SHIFT_EDIT_COLUMNS]


def _override_edit_headers() -> list[str]:
    from app.services.staff_excel.schema import OVERRIDE_EDIT_COLUMNS

    return [str(c["header"]) for c in OVERRIDE_EDIT_COLUMNS]


def _build_edit_workbook_bytes(
    *,
    staff_rows: list[dict] | None = None,
    shift_edit_rows: list[dict] | None = None,
    override_edit_rows: list[dict] | None = None,
) -> bytes:
    """Phase G-56: 編集用シート (1 スタッフ 1 行) を持つワークブックを作る.

    各 dict は { col_key: value } を持つ.
    """
    from app.services.staff_excel.schema import (
        OVERRIDE_EDIT_COL_INDEX,
        SHEET_OVERRIDE_EDIT,
        SHEET_SHIFT_EDIT,
        SHIFT_EDIT_COL_INDEX,
    )

    wb = Workbook()
    ws_s = wb.active
    ws_s.title = SHEET_STAFF
    for col_idx, header in enumerate(_staff_headers(), start=1):
        ws_s.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(staff_rows or [], start=2):
        for col_key, idx in STAFF_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_s.cell(row=r_idx, column=idx + 1, value=v)

    ws_f = wb.create_sheet(title=SHEET_SHIFT_EDIT)
    for col_idx, header in enumerate(_shift_edit_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(shift_edit_rows or [], start=2):
        for col_key, idx in SHIFT_EDIT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)

    ws_o = wb.create_sheet(title=SHEET_OVERRIDE_EDIT)
    for col_idx, header in enumerate(_override_edit_headers(), start=1):
        ws_o.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(override_edit_rows or [], start=2):
        for col_key, idx in OVERRIDE_EDIT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_o.cell(row=r_idx, column=idx + 1, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_export_writes_shift_edit_one_row_per_staff(client, db) -> None:
    """export: 1 スタッフ = シフト編集用シートの 1 行. 各曜日が開始/終了列に展開される."""
    from app.services.staff_excel.schema import SHEET_SHIFT_EDIT, SHIFT_EDIT_COL_INDEX

    admin = await _make_user(db, "g56-ex-shift@example.com", "admin")
    s = await _make_staff(db, code="S-G56-SH", name="編集シフト")
    # 月: 出勤 9-18, 火: 休 (is_on=False)
    await _make_shift(
        db, staff_id=s.id, weekday=0, is_on=True, start_time=time(9, 0), end_time=time(18, 0)
    )
    await _make_shift(db, staff_id=s.id, weekday=1, is_on=False, start_time=None, end_time=None)

    res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200, res.text
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_SHIFT_EDIT in wb.sheetnames
    ws = wb[SHEET_SHIFT_EDIT]
    assert ws.max_row == 2  # header + 1 staff row
    assert ws.cell(row=2, column=SHIFT_EDIT_COL_INDEX["staff_code"] + 1).value == "S-G56-SH"
    assert ws.cell(row=2, column=SHIFT_EDIT_COL_INDEX["mon_start"] + 1).value == "09:00"
    assert ws.cell(row=2, column=SHIFT_EDIT_COL_INDEX["mon_end"] + 1).value == "18:00"
    # 火 (休) は空欄
    assert ws.cell(row=2, column=SHIFT_EDIT_COL_INDEX["tue_start"] + 1).value is None
    assert ws.cell(row=2, column=SHIFT_EDIT_COL_INDEX["tue_end"] + 1).value is None


@pytest.mark.asyncio
async def test_shift_edit_roundtrip_no_changes(client, db) -> None:
    """export → 無編集 import で全 noop (round-trip 安定)."""
    admin = await _make_user(db, "g56-rt-shift@example.com", "admin")
    s = await _make_staff(db, code="S-G56-RT", name="往復")
    sid = s.id
    for wd in range(7):
        if wd < 5:
            await _make_shift(
                db,
                staff_id=sid,
                weekday=wd,
                is_on=True,
                start_time=time(9, 0),
                end_time=time(18, 0),
            )
        else:
            await _make_shift(
                db, staff_id=sid, weekday=wd, is_on=False, start_time=None, end_time=None
            )

    export_res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    assert body["summary"]["shift_error"] == 0
    assert body["summary"]["shift_new"] == 0
    assert body["summary"]["shift_update"] == 0
    assert body["summary"]["shift_delete"] == 0
    # 全て noop
    assert all(r["operation"] == "noop" for r in body["shift_rows"])


@pytest.mark.asyncio
async def test_shift_edit_per_staff_replace_turns_weekday_off(client, db) -> None:
    """per-staff replace: ある曜日を空欄にすると is_on=False (休) に置換される."""
    admin = await _make_user(db, "g56-off@example.com", "admin")
    s = await _make_staff(db, code="S-G56-OFF", name="休化")
    sid = s.id
    await _make_shift(
        db, staff_id=sid, weekday=0, is_on=True, start_time=time(9, 0), end_time=time(18, 0)
    )

    # 月の開始/終了を空欄にする (= 休)
    content = _build_edit_workbook_bytes(
        shift_edit_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-G56-OFF",
                # mon_start / mon_end 空欄 → is_on=False
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    db.expire_all()
    sh = await db.scalar(
        select(StaffShift).where(StaffShift.staff_id == sid, StaffShift.weekday == 0)
    )
    assert sh is not None
    assert sh.is_on is False
    assert sh.start_time is None


@pytest.mark.asyncio
async def test_shift_edit_does_not_touch_other_staff(client, db) -> None:
    """per-staff replace は対象 staff の shift のみ. 行に出ない staff は不変."""
    admin = await _make_user(db, "g56-other@example.com", "admin")
    s1 = await _make_staff(db, code="S-G56-A", name="A")
    s2 = await _make_staff(db, code="S-G56-B", name="B")
    s2_id = s2.id
    await _make_shift(
        db, staff_id=s2_id, weekday=0, is_on=True, start_time=time(8, 0), end_time=time(17, 0)
    )

    content = _build_edit_workbook_bytes(
        shift_edit_rows=[
            {
                "staff_id": str(s1.id),
                "staff_code": "S-G56-A",
                "mon_start": "09:00",
                "mon_end": "18:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    # s2 の shift は不変
    sh2 = await db.scalar(
        select(StaffShift).where(StaffShift.staff_id == s2_id, StaffShift.weekday == 0)
    )
    assert sh2 is not None
    assert sh2.start_time == time(8, 0)


@pytest.mark.asyncio
async def test_shift_edit_new_staff_via_code_link(client, db) -> None:
    """新規 staff + 編集用シフト行を staff_code リンクで登録."""
    admin = await _make_user(db, "g56-newlink@example.com", "admin")
    content = _build_edit_workbook_bytes(
        staff_rows=[
            {
                "staff_code": "S-G56-NEW",
                "name": "新規",
                "status": "active",
                "role": "staff",
            }
        ],
        shift_edit_rows=[
            {
                "staff_code": "S-G56-NEW",
                "mon_start": "09:00",
                "mon_end": "18:00",
                "tue_start": "10:00",
                "tue_end": "19:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["staff_new"] == 1
    db.expire_all()
    new_staff = await db.scalar(select(Staff).where(Staff.code == "S-G56-NEW"))
    assert new_staff is not None
    shifts = (await db.scalars(select(StaffShift).where(StaffShift.staff_id == new_staff.id))).all()
    # 7 曜日分 (月火 出勤, 残り 休) が登録される
    assert len(shifts) == 7
    by_wd = {sh.weekday: sh for sh in shifts}
    assert by_wd[0].is_on is True and by_wd[0].start_time == time(9, 0)
    assert by_wd[2].is_on is False


@pytest.mark.asyncio
async def test_override_edit_export_empty_sheet(client, db) -> None:
    """override が 0 件でも勤務例外（編集用）シートはヘッダー + 記入例行で出力される.

    Phase G-58.1: 記入方法が分かるよう、データが無くても常に記入例 (サンプル) 行を
    ヘッダー直下に置く (header(1) + 記入例(2) = max_row 2).
    """
    from app.services.staff_excel.schema import (
        OVERRIDE_EDIT_COL_INDEX,
        SAMPLE_ROW_MARKER,
        SHEET_OVERRIDE_EDIT,
    )

    admin = await _make_user(db, "g56-ov-empty@example.com", "admin")
    await _make_staff(db, code="S-G56-OVE", name="例外なし")
    res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_OVERRIDE_EDIT in wb.sheetnames
    ws_o = wb[SHEET_OVERRIDE_EDIT]
    assert ws_o.max_row == 2  # header + 記入例行
    sample_code = ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["staff_code"] + 1).value
    assert sample_code == SAMPLE_ROW_MARKER


@pytest.mark.asyncio
async def test_override_edit_roundtrip(client, db) -> None:
    """override export → 無編集 import で round-trip 安定 (人工データ)."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "g56-ov-rt@example.com", "admin")
    s = await _make_staff(db, code="S-G56-OVRT", name="例外往復")
    sid = s.id
    db.add(
        StaffWeeklyOverride(
            staff_id=sid, iso_year=2026, iso_week=20, weekday=2, override_type="off", reason="私用"
        )
    )
    db.add(
        StaffWeeklyOverride(
            staff_id=sid,
            iso_year=2026,
            iso_week=20,
            weekday=4,
            override_type="custom_time",
            start_time=time(10, 0),
            end_time=time(16, 0),
            reason="私用",
        )
    )
    await db.commit()

    export_res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    assert body["summary"]["override_error"] == 0
    assert body["summary"]["override_new"] == 0
    assert body["summary"]["override_update"] == 0
    assert body["summary"]["override_delete"] == 0
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert len(ovs) == 2


@pytest.mark.asyncio
async def test_override_edit_per_week_replace(client, db) -> None:
    """per-(staff, week) replace: 空欄曜日=削除, 休=off, HH:MM-HH:MM=custom_time."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "g56-ov-rep@example.com", "admin")
    s = await _make_staff(db, code="S-G56-OVREP", name="週置換")
    sid = s.id
    # 既存: 月=off, 火=custom. import で 月削除, 火更新, 水追加.
    db.add(
        StaffWeeklyOverride(
            staff_id=sid, iso_year=2026, iso_week=25, weekday=0, override_type="off"
        )
    )
    db.add(
        StaffWeeklyOverride(
            staff_id=sid,
            iso_year=2026,
            iso_week=25,
            weekday=1,
            override_type="custom_time",
            start_time=time(9, 0),
            end_time=time(17, 0),
        )
    )
    await db.commit()

    content = _build_edit_workbook_bytes(
        override_edit_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-G56-OVREP",
                "iso_year": 2026,
                "iso_week": 25,
                # mon 空欄 → 削除
                "tue": "10:00-18:00",  # 火 更新
                "wed": "休",  # 水 追加
                "reason": "調整",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["override_delete"] == 1
    assert body["summary"]["override_update"] == 1
    assert body["summary"]["override_new"] == 1
    db.expire_all()
    ovs = (
        await db.scalars(
            select(StaffWeeklyOverride)
            .where(StaffWeeklyOverride.staff_id == sid)
            .order_by(StaffWeeklyOverride.weekday)
        )
    ).all()
    assert [ov.weekday for ov in ovs] == [1, 2]
    tue = next(ov for ov in ovs if ov.weekday == 1)
    assert tue.override_type == "custom_time"
    assert tue.start_time == time(10, 0)
    wed = next(ov for ov in ovs if ov.weekday == 2)
    assert wed.override_type == "off"


@pytest.mark.asyncio
async def test_override_edit_duplicate_week_errors(client, db) -> None:
    """同一 (staff, iso_year, iso_week) が複数行 → 2 行目が error."""
    admin = await _make_user(db, "g56-ov-dup@example.com", "admin")
    s = await _make_staff(db, code="S-G56-OVDUP", name="週重複")
    sid = s.id
    content = _build_edit_workbook_bytes(
        override_edit_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-G56-OVDUP",
                "iso_year": 2026,
                "iso_week": 26,
                "mon": "休",
            },
            {
                "staff_id": str(sid),
                "staff_code": "S-G56-OVDUP",
                "iso_year": 2026,
                "iso_week": 26,
                "tue": "休",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["override_error"] == 1


@pytest.mark.asyncio
async def test_legacy_shift_sheet_fallback(client, db) -> None:
    """旧形式 (per-row 勤務シフトシート) でもクラッシュせず処理される (後方互換)."""
    admin = await _make_user(db, "g56-legacy@example.com", "admin")
    staff = await _make_staff(db, code="S-G56-LEG", name="旧形式")
    staff_id = staff.id
    # _build_workbook_bytes は旧 SHEET_SHIFT を作る.
    content = _build_workbook_bytes(
        shift_rows=[
            {
                "staff_id": str(staff_id),
                "weekday": "月",
                "is_on": "TRUE",
                "start_time": "09:00",
                "end_time": "18:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["shift_new"] == 1
    db.expire_all()
    sh = await db.scalar(
        select(StaffShift).where(StaffShift.staff_id == staff_id, StaffShift.weekday == 0)
    )
    assert sh is not None
    assert sh.start_time == time(9, 0)


# ---------------------------------------------------------------------------
# Phase G-58.1: 勤務例外（編集用）記入例 (サンプル) 行
# ---------------------------------------------------------------------------


def test_is_sample_row_nfkc_unit() -> None:
    """is_sample_row は NFKC 正規化で全角/半角揺れを吸収する (純粋関数)."""
    from app.services.staff_excel.schema import SAMPLE_ROW_MARKER, is_sample_row

    # staff_code 一致.
    assert is_sample_row(SAMPLE_ROW_MARKER, None) is True
    # staff_id 一致.
    assert is_sample_row(None, SAMPLE_ROW_MARKER) is True
    # NFKC で半角括弧表記も一致 (全角→半角に正規化されて等価).
    assert is_sample_row("(記入例)", None) is True
    # 前後空白を吸収.
    assert is_sample_row("  （記入例）  ", None) is True
    # 実データは一致しない.
    assert is_sample_row("S-001", "11111111-1111-1111-1111-111111111111") is False
    assert is_sample_row(None, None) is False
    assert is_sample_row("記入例", None) is False  # 括弧なしは別物


@pytest.mark.asyncio
async def test_override_sample_row_exported_with_content(client, db) -> None:
    """export の勤務例外（編集用）シートに記入例行が常に row 2 に書かれる."""
    from app.services.staff_excel.schema import (
        OVERRIDE_EDIT_COL_INDEX,
        SAMPLE_ROW_MARKER,
        SHEET_OVERRIDE_EDIT,
    )

    admin = await _make_user(db, "g58-sample-ex@example.com", "admin")
    await _make_staff(db, code="S-G58-SMP", name="サンプル確認")
    res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws_o = wb[SHEET_OVERRIDE_EDIT]
    # row 2 = 記入例行.
    assert ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["staff_code"] + 1).value == (
        SAMPLE_ROW_MARKER
    )
    assert ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["iso_year"] + 1).value == 2026
    assert ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["iso_week"] + 1).value == 23
    assert ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["mon"] + 1).value == "休"
    assert ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["tue"] + 1).value == "10:00-15:00"
    reason = ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["reason"] + 1).value
    assert reason is not None and "（例）" in reason  # reason に例文あり
    # 記入例行はグレー背景 + 斜体で装飾されている.
    cell = ws_o.cell(row=2, column=OVERRIDE_EDIT_COL_INDEX["staff_code"] + 1)
    assert cell.font.italic is True


@pytest.mark.asyncio
async def test_override_sample_row_skipped_on_import(client, db) -> None:
    """記入例行は import で完全に無視される (DB 変更ゼロ・error/noop どちらにもならない)."""
    from app.models import StaffWeeklyOverride
    from app.services.staff_excel.schema import SAMPLE_ROW_MARKER

    admin = await _make_user(db, "g58-sample-im@example.com", "admin")
    s = await _make_staff(db, code="S-G58-SKIP", name="スキップ確認")
    sid = s.id

    # 記入例行のみ (export と同じ内容) を含む override-edit シートを import.
    content = _build_edit_workbook_bytes(
        override_edit_rows=[
            {
                "staff_code": SAMPLE_ROW_MARKER,
                "staff_name": f"↓この行は取り込まれません{SAMPLE_ROW_MARKER}",
                "iso_year": 2026,
                "iso_week": 23,
                "mon": "休",
                "tue": "10:00-15:00",
                "reason": "（例）火曜は通院のため時間変更、月曜は休み",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    # 記入例行はどの operation にも計上されない.
    assert body["summary"]["override_new"] == 0
    assert body["summary"]["override_update"] == 0
    assert body["summary"]["override_delete"] == 0
    assert body["summary"]["override_error"] == 0
    assert body["summary"]["override_noop"] == 0
    assert body["override_rows"] == []
    # DB は完全に変更なし.
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert ovs == []


@pytest.mark.asyncio
async def test_override_sample_row_roundtrip_all_noop(client, db) -> None:
    """export(記入例行含む) → 無編集 import が全 noop (記入例 skip + 実データ noop)."""
    from app.models import StaffWeeklyOverride

    admin = await _make_user(db, "g58-sample-rt@example.com", "admin")
    s = await _make_staff(db, code="S-G58-RT", name="往復確認")
    sid = s.id
    db.add(
        StaffWeeklyOverride(
            staff_id=sid,
            iso_year=2026,
            iso_week=24,
            weekday=3,
            override_type="custom_time",
            start_time=time(8, 30),
            end_time=time(17, 30),
            reason="短縮勤務",
        )
    )
    await db.commit()

    export_res = await client.get("/api/v1/staff/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/staff/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    # 記入例行は skip, 実データは noop → new/update/delete/error すべて 0.
    assert body["summary"]["override_new"] == 0
    assert body["summary"]["override_update"] == 0
    assert body["summary"]["override_delete"] == 0
    assert body["summary"]["override_error"] == 0
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert len(ovs) == 1
    assert ovs[0].override_type == "custom_time"
    assert ovs[0].reason == "短縮勤務"


@pytest.mark.asyncio
async def test_override_sample_row_skipped_on_replace_all(client, db) -> None:
    """replace_all でも記入例行は skip され DB 変更ゼロ (実データのみ反映)."""
    from app.models import StaffWeeklyOverride
    from app.services.staff_excel.schema import SAMPLE_ROW_MARKER

    admin = await _make_user(db, "g58-sample-ra@example.com", "admin")
    s = await _make_staff(db, code="S-G58-RA", name="完全置換確認")
    sid = s.id

    content = _build_edit_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-G58-RA",
                "name": "完全置換確認",
                "status": "active",
                "role": "staff",
            }
        ],
        shift_edit_rows=[{"staff_id": str(sid), "staff_code": "S-G58-RA"}],
        override_edit_rows=[
            {
                "staff_code": SAMPLE_ROW_MARKER,
                "iso_year": 2026,
                "iso_week": 23,
                "mon": "休",
                "tue": "10:00-15:00",
            },
        ],
    )
    files = {
        "file": (
            "replace.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/staff/import-export/replace-all?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    # 記入例行は new/error に計上されない.
    assert body["summary"]["override_to_create"] == 0
    assert body["summary"]["override_error"] == 0
    db.expire_all()
    ovs = (
        await db.scalars(select(StaffWeeklyOverride).where(StaffWeeklyOverride.staff_id == sid))
    ).all()
    assert ovs == []
