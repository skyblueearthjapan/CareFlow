"""Tests for 患者個別カルテ (A4・1 患者 1 ファイル) backend.

Coverage:
  * カルテ export 構造 (各セル値マップ)
  * 全プルダウン (DataValidation) 存在
  * 時間タイプ 午前/午後/終日 のグレーアウト条件付き書式
  * round-trip 0 差分 (DB 実患者 → build → parse → diff が全 noop)
  * 空白テンプレからの新規作成
  * 〇× 往復
  * PFV 患者単位 replace (空欄曜日削除 / 記入 upsert / special 保持)
  * 希望時刻分割往復 (B13/C13/D13)
  * サービス時間「分」除去
  * 拠点「（自動）」除去
  * クロス拠点 (稲B / 津A) トークン往復
  * API 3 本 (karte-template / karte-export / karte-import) のパス・RBAC
"""

from __future__ import annotations

from datetime import time
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.core.security import create_access_token, hash_password
from app.models import (
    CourseTemplate,
    Office,
    Patient,
    PatientFixedVisit,
    User,
)
from app.services.patient_excel.karte import (
    OFFICE_AUTO_SUFFIX,
    REST_MARK,
    SHEET_KARTE,
    build_blank_karte_template,
    build_karte_workbook,
    parse_karte_workbook,
)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


async def _make_user(db, email: str, role: str) -> User:
    user = User(email=email, password_hash=hash_password("pw"), role=role)
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


def _bearer(user: User) -> dict[str, str]:
    token = create_access_token(subject=user.id, role=user.role, staff_id=user.staff_id)
    return {"Authorization": f"Bearer {token}"}


async def _make_office(db, code: str, name: str | None = None) -> Office:
    o = Office(code=code, name=name or code)
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return o


async def _make_course_template(db, *, office_id, label: str) -> CourseTemplate:
    ct = CourseTemplate(office_id=office_id, label=label)
    db.add(ct)
    await db.commit()
    await db.refresh(ct)
    return ct


async def _make_patient(db, **overrides) -> Patient:
    defaults = {
        "code": "P-KARTE",
        "name": "カルテ患者",
        "sex": "male",
        "status": "active",
        "address": "千葉市稲毛区test1",
    }
    defaults.update(overrides)
    p = Patient(**defaults)
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


async def _make_pfv(
    db,
    *,
    patient_id,
    weekday: int = 0,
    slot_index: int = 0,
    mode: str = "normal",
    start_time=time(9, 30),
    duration_min: int = 60,
    course_template_id=None,
) -> PatientFixedVisit:
    pfv = PatientFixedVisit(
        patient_id=patient_id,
        mode=mode,
        weekday=weekday,
        slot_index=slot_index,
        start_time=start_time,
        duration_min=duration_min,
        course_template_id=course_template_id,
    )
    db.add(pfv)
    await db.commit()
    await db.refresh(pfv)
    return pfv


def _dv_cells(ws) -> set[str]:
    """シート上で DataValidation が設定されているセル参照の集合を返す."""
    cells: set[str] = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            cells.add(str(rng.coord))
    return cells


async def _upload_karte(client, user, *, content: bytes, dry_run: bool = True):
    files = {
        "file": (
            "karte.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    return await client.post(
        f"/api/v1/patients/import-export/karte-import?dry_run={'true' if dry_run else 'false'}",
        headers=_bearer(user),
        files=files,
    )


def _karte_bytes_from_cells(*, base: bytes, overrides: dict[str, object]) -> bytes:
    """既存カルテ bytes の特定セルを上書きした bytes を返す (編集シミュレーション)."""
    wb = load_workbook(BytesIO(base))
    ws = wb[SHEET_KARTE] if SHEET_KARTE in wb.sheetnames else wb[wb.sheetnames[0]]
    for ref, value in overrides.items():
        ws[ref] = value
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 1) export 構造: 各セル値マップ
# ---------------------------------------------------------------------------


def test_build_karte_cell_map() -> None:
    oid = uuid4()
    office = Office(id=oid, code="INAGE", name="稲毛")
    ct = CourseTemplate(id=uuid4(), office_id=oid, label="A")
    p = Patient(
        id=uuid4(),
        code="P001",
        name="山田 太郎",
        kana="ヤマダ タロウ",
        sex="female",
        status="active",
        insurance="care",
        address="千葉市稲毛区xxx",
        primary_office_id=oid,
        sex_restriction="female_only",
        requires_multiple_staff=True,
        note="ヒアリングメモ",
        weekly_pattern={
            "frequency_per_week": 3,
            "visit_frequency": "every",
            "preferred_weekdays": ["Mon", "Wed", "Fri"],
            "service_minutes": 60,
            "time_type": "時間帯",
            "preferred_start": "09:30",
            "preferred_end": "10:30",
        },
    )
    pfvs = [
        PatientFixedVisit(
            id=uuid4(),
            patient_id=p.id,
            mode="normal",
            weekday=0,
            slot_index=0,
            start_time=time(9, 30),
            duration_min=60,
            course_template_id=ct.id,
        )
    ]
    wb = build_karte_workbook(patient=p, fixed_visits=pfvs, offices=[office], course_templates=[ct])
    ws = wb[SHEET_KARTE]

    assert ws["A1"].value == "訪問看護スケジューリング用カルテ"
    assert ws["F1"].value == "コード:"
    assert ws["G1"].value == "P001"
    assert ws["B2"].value == "山田 太郎"
    assert ws["F2"].value == "ヤマダ タロウ"
    assert ws["B5"].value == "女性"
    assert ws["D5"].value == "稼働"
    assert ws["F5"].value == "介護"  # 短縮表記
    assert ws["B6"].value == f"稲毛{OFFICE_AUTO_SUFFIX}"
    assert ws["D6"].value == "千葉市稲毛区xxx"
    assert ws["B9"].value == 3
    assert ws["D9"].value == "毎週"
    # 希望曜日 月水金 = 〇, 他 ×.
    assert [ws[f"{c}11"].value for c in "BCDEFGH"] == ["〇", "×", "〇", "×", "〇", "×", "×"]
    assert ws["B12"].value == "60分"
    assert ws["D12"].value == "時間帯"
    assert ws["B13"].value == "09:30"
    assert ws["C13"].value == "〜"
    assert ws["D13"].value == "10:30"
    assert ws["B14"].value == "女性のみ"
    assert ws["D14"].value == "はい"
    # 固定訪問: 月のみ 09:30 / 稲A. 他は休 (―).
    assert ws["B18"].value == "09:30"
    assert ws["B19"].value == "稲A"
    assert ws["C18"].value == REST_MARK
    assert ws["A22"].value == "ヒアリングメモ"


# 2) 全プルダウン存在
def test_build_karte_dropdowns_present() -> None:
    oid = uuid4()
    office = Office(id=oid, code="INAGE", name="稲毛")
    ct = CourseTemplate(id=uuid4(), office_id=oid, label="A")
    p = Patient(id=uuid4(), code="P002", name="t", sex="male", status="active", address="a")
    wb = build_karte_workbook(patient=p, fixed_visits=[], offices=[office], course_templates=[ct])
    ws = wb[SHEET_KARTE]
    cells = _dv_cells(ws)
    expected = {
        "B5",
        "D5",
        "F5",
        "B6",
        "B9",
        "D9",
        "B11",
        "C11",
        "D11",
        "E11",
        "F11",
        "G11",
        "H11",
        "B12",
        "D12",
        "B13",
        "D13",
        "B14",
        "D14",
        "B18",
        "C18",
        "D18",
        "E18",
        "F18",
        "G18",
        "B19",
        "C19",
        "D19",
        "E19",
        "F19",
        "G19",
    }
    assert expected.issubset(cells), f"missing dropdowns: {expected - cells}"
    # 固定訪問の日曜セル (H18/H19) は取込対象外のため dropdown を付けない.
    assert "H18" not in cells
    assert "H19" not in cells


# 3) グレーアウト条件付き書式
def test_build_karte_time_greyout_conditional_format() -> None:
    oid = uuid4()
    office = Office(id=oid, code="INAGE", name="稲毛")
    p = Patient(id=uuid4(), code="P003", name="t", sex="male", status="active", address="a")
    wb = build_karte_workbook(patient=p, fixed_visits=[], offices=[office], course_templates=[])
    ws = wb[SHEET_KARTE]
    cf_ranges = {str(cf.sqref) for cf in ws.conditional_formatting}
    assert "B13" in cf_ranges
    assert "D13" in cf_ranges
    # 条件式に 午前/午後/終日 が含まれる.
    found = False
    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            if rule.formula and "午前" in rule.formula[0] and "終日" in rule.formula[0]:
                found = True
    assert found


# 4) blank template: 値なし・dropdown あり
def test_blank_template_has_dropdowns_no_values() -> None:
    oid = uuid4()
    office = Office(id=oid, code="INAGE", name="稲毛")
    ct = CourseTemplate(id=uuid4(), office_id=oid, label="A")
    wb = build_blank_karte_template(offices=[office], course_templates=[ct])
    ws = wb[SHEET_KARTE]
    # 値セル (G1 / B2 / B5 ...) は空.
    assert ws["G1"].value is None
    assert ws["B2"].value is None
    assert ws["B5"].value is None
    # ラベル/見出しは存在.
    assert ws["A1"].value == "訪問看護スケジューリング用カルテ"
    # dropdown あり.
    assert "B5" in _dv_cells(ws)


# ---------------------------------------------------------------------------
# round-trip 0 差分 (DB)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_roundtrip_zero_diff(client, db) -> None:
    """DB 実患者 → build → parse → diff が全 noop."""
    admin = await _make_user(db, "k-rt@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    ct = await _make_course_template(db, office_id=office.id, label="A")
    p = await _make_patient(
        db,
        code="P-RT-001",
        name="往復 太郎",
        kana="オウフク タロウ",
        sex="female",
        status="active",
        insurance="care",
        address="千葉市稲毛区rt",
        primary_office_id=office.id,
        requires_multiple_staff=True,
        note="メモ",
        weekly_pattern={
            "frequency_per_week": 2,
            "visit_frequency": "every",
            "preferred_weekdays": ["Mon", "Wed"],
            "service_minutes": 60,
            "time_type": "時間帯",
            "preferred_start": "09:30",
            "preferred_end": "10:30",
        },
    )
    await _make_pfv(
        db, patient_id=p.id, weekday=0, start_time=time(9, 30), course_template_id=ct.id
    )
    await _make_pfv(
        db, patient_id=p.id, weekday=2, start_time=time(9, 30), course_template_id=ct.id
    )

    # export カルテを取得.
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    assert res.status_code == 200, res.text
    karte_bytes = res.content

    # そのまま import (dry_run) → 全 noop.
    res = await _upload_karte(client, admin, content=karte_bytes, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    s = body["summary"]
    assert s["patients_new"] == 0
    assert s["patients_update"] == 0
    assert s["patients_delete"] == 0
    assert s["patients_error"] == 0
    assert s["pfv_new"] == 0
    assert s["pfv_update"] == 0
    assert s["pfv_delete"] == 0
    assert s["pfv_error"] == 0


@pytest.mark.asyncio
async def test_roundtrip_zero_diff_preserves_unmanaged_weekly_keys(client, db) -> None:
    """weekly_pattern の管理外キー (entries/staff_count) を持つ患者も round-trip noop."""
    admin = await _make_user(db, "k-rt2@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    ct = await _make_course_template(db, office_id=office.id, label="A")
    p = await _make_patient(
        db,
        code="P-RT-002",
        name="管理外キー患者",
        sex="male",
        status="active",
        address="千葉市稲毛区rt2",
        primary_office_id=office.id,
        weekly_pattern={
            "frequency_per_week": 1,
            "preferred_weekdays": ["Tue"],
            "service_minutes": 45,
            "time_type": "固定",
            "preferred_start": "10:00",
            "entries": [{"weekday": "Tue", "staff_count": 2}],
            "staff_count": 2,
        },
    )
    await _make_pfv(
        db,
        patient_id=p.id,
        weekday=1,
        start_time=time(10, 0),
        duration_min=45,
        course_template_id=ct.id,
    )
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    res = await _upload_karte(client, admin, content=res.content, dry_run=True)
    body = res.json()
    s = body["summary"]
    assert s["patients_update"] == 0
    assert s["patients_new"] == 0
    assert s["pfv_new"] == 0
    assert s["pfv_update"] == 0
    assert s["pfv_delete"] == 0


# ---------------------------------------------------------------------------
# 空白テンプレからの新規作成
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_blank_template_creates_new_patient(client, db) -> None:
    admin = await _make_user(db, "k-new@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    ct = await _make_course_template(db, office_id=office.id, label="A")
    office_id = office.id
    ct_id = ct.id

    wb = build_blank_karte_template(offices=[office], course_templates=[ct])
    ws = wb[SHEET_KARTE]
    # 新規記入.
    ws["G1"] = "P-NEW-KARTE"
    ws["B2"] = "新規 花子"
    ws["F2"] = "シンキ ハナコ"
    ws["B5"] = "女性"
    ws["D5"] = "稼働"
    ws["F5"] = "医療"
    ws["B6"] = "稲毛"
    ws["D6"] = "千葉市稲毛区new"
    ws["B12"] = "60分"
    ws["D12"] = "固定"
    ws["B18"] = "09:30"
    ws["B19"] = "稲A"
    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    res = await _upload_karte(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_new"] == 1
    assert body["summary"]["patients_error"] == 0
    assert body["summary"]["pfv_new"] == 1

    db.expire_all()
    from sqlalchemy import select

    rows = (await db.scalars(select(Patient).where(Patient.code == "P-NEW-KARTE"))).all()
    assert len(rows) == 1
    new_p = rows[0]
    assert new_p.name == "新規 花子"
    assert new_p.sex == "female"
    assert new_p.insurance == "medical"
    assert new_p.primary_office_id == office_id
    pfvs = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == new_p.id))
    ).all()
    assert len(pfvs) == 1
    assert pfvs[0].weekday == 0
    assert pfvs[0].course_template_id == ct_id


# ---------------------------------------------------------------------------
# PFV 患者単位 replace (空欄曜日削除 / special 保持)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_replace_blank_weekday_deletes(client, db) -> None:
    """カルテで曜日を「―」にすると normal slot0 PFV が削除される. special は保持."""
    admin = await _make_user(db, "k-pfvrep@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    ct = await _make_course_template(db, office_id=office.id, label="A")
    p = await _make_patient(
        db, code="P-PFVREP", name="PFV置換", address="千葉市稲毛区", primary_office_id=office.id
    )
    # 月・水の normal PFV + special PFV.
    await _make_pfv(
        db, patient_id=p.id, weekday=0, start_time=time(9, 30), course_template_id=ct.id
    )
    wed = await _make_pfv(
        db, patient_id=p.id, weekday=2, start_time=time(9, 30), course_template_id=ct.id
    )
    special = await _make_pfv(
        db, patient_id=p.id, weekday=0, mode="special", start_time=time(14, 0)
    )
    wed_id = wed.id
    special_id = special.id

    # export → 水曜を「―」に上書き (= 水曜削除) して import.
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    edited = _karte_bytes_from_cells(
        base=res.content, overrides={"D18": REST_MARK, "D19": REST_MARK}
    )
    res = await _upload_karte(client, admin, content=edited, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["pfv_delete"] == 1

    db.expire_all()
    # 水曜 normal は削除.
    assert await db.get(PatientFixedVisit, wed_id) is None
    # special は保持.
    assert await db.get(PatientFixedVisit, special_id) is not None


@pytest.mark.asyncio
async def test_pfv_replace_upsert_new_weekday(client, db) -> None:
    """カルテで新たな曜日に時刻を入れると normal slot0 PFV が upsert される."""
    admin = await _make_user(db, "k-pfvup@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    ct = await _make_course_template(db, office_id=office.id, label="A")
    p = await _make_patient(
        db, code="P-PFVUP", name="PFV追加", address="千葉市稲毛区", primary_office_id=office.id
    )
    p_id = p.id
    await _make_pfv(
        db, patient_id=p_id, weekday=0, start_time=time(9, 30), course_template_id=ct.id
    )

    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p_id}",
        headers=_bearer(admin),
    )
    # 火曜 (C18/C19) に新規記入.
    edited = _karte_bytes_from_cells(base=res.content, overrides={"C18": "11:00", "C19": "稲A"})
    res = await _upload_karte(client, admin, content=edited, dry_run=False)
    body = res.json()
    assert body["summary"]["pfv_new"] == 1

    db.expire_all()
    from sqlalchemy import select

    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(
                PatientFixedVisit.patient_id == p_id,
                PatientFixedVisit.weekday == 1,
            )
        )
    ).all()
    assert len(rows) == 1
    assert rows[0].start_time == time(11, 0)


# ---------------------------------------------------------------------------
# 〇× 往復 / 希望時刻分割 / サービス時間「分」除去 / 拠点「（自動）」除去
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_weekday_marks_and_split_time_roundtrip(client, db) -> None:
    admin = await _make_user(db, "k-marks@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    p = await _make_patient(
        db,
        code="P-MARKS",
        name="〇×往復",
        address="千葉市稲毛区",
        primary_office_id=office.id,
        weekly_pattern={
            "preferred_weekdays": ["Mon", "Fri"],
            "service_minutes": 90,
            "time_type": "時間帯",
            "preferred_start": "08:00",
            "preferred_end": "09:00",
        },
    )
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_KARTE]
    # 〇× が月・金 のみ 〇.
    assert ws["B11"].value == "〇"  # 月
    assert ws["F11"].value == "〇"  # 金
    assert ws["C11"].value == "×"  # 火
    # サービス時間「90分」表記.
    assert ws["B12"].value == "90分"
    # 希望時刻 split.
    assert ws["B13"].value == "08:00"
    assert ws["D13"].value == "09:00"
    # 拠点（自動）.
    assert ws["B6"].value.endswith(OFFICE_AUTO_SUFFIX)

    # round-trip noop.
    res = await _upload_karte(client, admin, content=res.content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_update"] == 0
    assert body["summary"]["patients_new"] == 0


@pytest.mark.asyncio
async def test_office_auto_suffix_stripped_on_import(client, db) -> None:
    """拠点セルが「稲毛（自動）」でも INAGE に正規化され noop (round-trip 安定)."""
    admin = await _make_user(db, "k-office@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    p = await _make_patient(
        db, code="P-OFFSUF", name="拠点suffix", address="千葉市稲毛区", primary_office_id=office.id
    )
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    # build した B6 は「稲毛（自動）」のはず.
    wb = load_workbook(BytesIO(res.content))
    assert wb[SHEET_KARTE]["B6"].value == f"稲毛{OFFICE_AUTO_SUFFIX}"
    # parse → 標準シートで office_code=INAGE.
    std = parse_karte_workbook(res.content)
    from app.services.patient_excel.schema import PATIENT_COL_INDEX, SHEET_PATIENTS

    sw = load_workbook(BytesIO(std))
    code = sw[SHEET_PATIENTS].cell(row=2, column=PATIENT_COL_INDEX["office_code"] + 1).value
    assert code == "INAGE"


# ---------------------------------------------------------------------------
# クロス拠点 (稲B / 津A) トークン往復
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_cross_office_course_token_roundtrip(client, db) -> None:
    admin = await _make_user(db, "k-cross@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    ct_inage_b = await _make_course_template(db, office_id=inage.id, label="B")
    ct_tsuga_a = await _make_course_template(db, office_id=tsuga.id, label="A")
    p = await _make_patient(
        db,
        code="P-CROSS",
        name="クロス拠点",
        address="千葉市稲毛区",
        primary_office_id=inage.id,
        weekly_pattern={"service_minutes": 60},
    )
    p_id = p.id
    # 月 = 稲B, 木 = 津A (クロス拠点). PFV duration は weekly service_minutes と一致.
    await _make_pfv(
        db, patient_id=p_id, weekday=0, start_time=time(9, 30), course_template_id=ct_inage_b.id
    )
    await _make_pfv(
        db, patient_id=p_id, weekday=3, start_time=time(13, 0), course_template_id=ct_tsuga_a.id
    )

    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p_id}",
        headers=_bearer(admin),
    )
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_KARTE]
    assert ws["B19"].value == "稲B"  # 月
    assert ws["E19"].value == "津A"  # 木

    # round-trip noop (クロス拠点コースが course_template に正しく解決される).
    res = await _upload_karte(client, admin, content=res.content, dry_run=True)
    body = res.json()
    assert body["summary"]["pfv_new"] == 0
    assert body["summary"]["pfv_update"] == 0
    assert body["summary"]["pfv_delete"] == 0
    assert body["summary"]["pfv_error"] == 0


# ---------------------------------------------------------------------------
# 日曜固定訪問 非対応 (グレー + dropdown 無し) / 日曜 PFV 温存 (CRITICAL)
# ---------------------------------------------------------------------------


def test_karte_sunday_fixed_cells_greyed_no_dropdown() -> None:
    """固定訪問の日曜セル (H18/H19) がグレー塗り + 斜体・dropdown 無し + 注記あり."""
    from app.services.patient_excel.schema import ID_COLUMN_FILL_COLOR

    oid = uuid4()
    office = Office(id=oid, code="INAGE", name="稲毛")
    ct = CourseTemplate(id=uuid4(), office_id=oid, label="A")
    p = Patient(id=uuid4(), code="P-SUN", name="t", sex="male", status="active", address="a")
    wb = build_karte_workbook(patient=p, fixed_visits=[], offices=[office], course_templates=[ct])
    ws = wb[SHEET_KARTE]

    # dropdown 無し.
    cells = _dv_cells(ws)
    assert "H18" not in cells
    assert "H19" not in cells
    # グレー塗り + 斜体.
    for ref in ("H18", "H19"):
        assert ws[ref].font.italic is True
        assert ws[ref].fill.fgColor.rgb == ID_COLUMN_FILL_COLOR
    # 日曜見出し H17「日」は残る.
    assert ws["H17"].value == "日"
    # 注記 (A16 見出しに併記) に「日曜は取込対象外」明記.
    assert ws["A16"].value is not None and "日曜は取込対象外" in ws["A16"].value
    # blank template でも同様.
    wb_blank = build_blank_karte_template(offices=[office], course_templates=[ct])
    ws_blank = wb_blank[SHEET_KARTE]
    blank_cells = _dv_cells(ws_blank)
    assert "H18" not in blank_cells
    assert "H19" not in blank_cells
    assert ws_blank["A16"].value is not None and "日曜は取込対象外" in ws_blank["A16"].value


@pytest.mark.asyncio
async def test_sunday_pfv_preserved_on_karte_roundtrip(client, db) -> None:
    """CRITICAL: 患者の「日曜 normal slot0 PFV」がカルテ無編集往復で削除されない.

    日曜 (weekday=6) は固定訪問パイプライン (月..土) の管理対象外のため、
    カルテ build→parse→import で pfv_delete=0、かつ日曜 PFV が DB に残存する.
    """
    admin = await _make_user(db, "k-sun@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    ct = await _make_course_template(db, office_id=office.id, label="A")
    p = await _make_patient(
        db,
        code="P-SUN-RT",
        name="日曜固定",
        address="千葉市稲毛区sun",
        primary_office_id=office.id,
        weekly_pattern={"service_minutes": 60},
    )
    # 月 normal PFV + 日曜 (weekday=6) normal slot0 PFV.
    await _make_pfv(
        db, patient_id=p.id, weekday=0, start_time=time(9, 30), course_template_id=ct.id
    )
    sunday = await _make_pfv(
        db, patient_id=p.id, weekday=6, start_time=time(10, 0), course_template_id=ct.id
    )
    sunday_id = sunday.id

    # export → 無編集で import.
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    res = await _upload_karte(client, admin, content=res.content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    # 日曜 PFV は削除カウントに含まれない.
    assert body["summary"]["pfv_delete"] == 0

    db.expire_all()
    # 日曜 PFV は DB に残存する.
    assert await db.get(PatientFixedVisit, sunday_id) is not None


@pytest.mark.asyncio
async def test_insurance_medical_roundtrip_noop(client, db) -> None:
    """insurance=medical (短縮「医療」) の患者で export→import が patients_update=0 (noop).

    医療 ⇄ medical の往復対称性を検証 (従来は care のみ検証されていた).
    """
    admin = await _make_user(db, "k-med@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    p = await _make_patient(
        db,
        code="P-MED",
        name="医療保険",
        insurance="medical",
        address="千葉市稲毛区med",
        primary_office_id=office.id,
    )
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    # build した F5 は短縮「医療」.
    wb = load_workbook(BytesIO(res.content))
    assert wb[SHEET_KARTE]["F5"].value == "医療"

    # round-trip noop.
    res = await _upload_karte(client, admin, content=res.content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_update"] == 0
    assert body["summary"]["patients_new"] == 0
    assert body["summary"]["patients_error"] == 0


# ---------------------------------------------------------------------------
# API パス / RBAC
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_karte_template_endpoint(client, db) -> None:
    admin = await _make_user(db, "k-tpl@example.com", "admin")
    res = await client.get(
        "/api/v1/patients/import-export/karte-template",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_KARTE in wb.sheetnames


@pytest.mark.asyncio
async def test_karte_export_filename_header(client, db) -> None:
    admin = await _make_user(db, "k-fn@example.com", "admin")
    p = await _make_patient(db, code="P-FN-001", name="ファイル名")
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={p.id}",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    cd = res.headers.get("content-disposition", "")
    assert "P-FN-001" in cd
    assert "filename*=UTF-8" in cd


@pytest.mark.asyncio
async def test_karte_export_404_for_unknown(client, db) -> None:
    admin = await _make_user(db, "k-404@example.com", "admin")
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={uuid4()}",
        headers=_bearer(admin),
    )
    assert res.status_code == 404


@pytest.mark.asyncio
async def test_karte_rbac_staff_forbidden(client, db) -> None:
    staff = await _make_user(db, "k-staff@example.com", "staff")
    res = await client.get(
        "/api/v1/patients/import-export/karte-template",
        headers=_bearer(staff),
    )
    assert res.status_code == 403
    res = await client.get(
        f"/api/v1/patients/import-export/karte-export?patient_id={uuid4()}",
        headers=_bearer(staff),
    )
    assert res.status_code == 403

    wb = build_blank_karte_template(offices=[], course_templates=[])
    buf = BytesIO()
    wb.save(buf)
    res = await _upload_karte(client, staff, content=buf.getvalue(), dry_run=True)
    assert res.status_code == 403


@pytest.mark.asyncio
async def test_karte_import_dry_run_does_not_modify_db(client, db) -> None:
    admin = await _make_user(db, "k-dry@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    from sqlalchemy import select

    before = len((await db.scalars(select(Patient))).all())

    wb = build_blank_karte_template(offices=[office], course_templates=[])
    ws = wb[SHEET_KARTE]
    ws["G1"] = "P-DRY-KARTE"
    ws["B2"] = "ドライ"
    ws["B5"] = "男性"
    ws["D5"] = "稼働"
    ws["D6"] = "千葉市稲毛区dry"
    buf = BytesIO()
    wb.save(buf)
    res = await _upload_karte(client, admin, content=buf.getvalue(), dry_run=True)
    assert res.status_code == 200
    assert res.json()["transaction_applied"] is False

    db.expire_all()
    after = len((await db.scalars(select(Patient))).all())
    assert after == before
