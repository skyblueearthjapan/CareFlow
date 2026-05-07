"""Tests for scripts/import_acceptance_calendar.py (W15 Phase 2 B-2).

検証観点:
  1. パース正常系: フィクスチャ xlsx から正しい dict が組み立つ
  2. 「定休日」セルがスキップされる (日曜列)
  3. 拠点認識: name LIKE で正しく解決される
  4. 拠点未登録時に SystemExit (ERROR 終了)
  5. dry-run モード: DB に書き込まない
  6. apply モード: acceptance_calendar へ upsert される
  7. 冪等性: 2 回 apply しても同じ件数
  8. ブロック 2 (都賀拠点) のパース正常系
"""

from __future__ import annotations

import sys
from datetime import time
from pathlib import Path

import openpyxl
import pytest

# ── sys.path: mirror the import pattern used by the script itself ──
_BACKEND_ROOT = Path(__file__).resolve().parent.parent.parent
_SCRIPTS_DIR = _BACKEND_ROOT / "scripts"
for _p in (_SCRIPTS_DIR, _BACKEND_ROOT):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from import_acceptance_calendar import (  # noqa: E402  # type: ignore[import]
    _parse_status,
    _parse_time_slot,
    _resolve_office,
    _upsert_block,
    import_acceptance_calendar,
    parse_sheet,
)

# ---------------------------------------------------------------------------
# Helpers: build a minimal fixture xlsx in memory (tmp_path)
# ---------------------------------------------------------------------------

_WEEKDAY_COLS = [3, 5, 7, 9, 11, 13, 15]
_WEEKDAY_LABELS = ["月", "火", "水", "木", "金", "土", "日"]


def _build_fixture_xlsx(tmp_path: Path) -> Path:
    """Create a minimal「受け入れカレンダー」sheet with the real cell layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受け入れカレンダー"

    # ── Block 1: 稲毛拠点 ──
    ws.cell(row=2, column=2, value="稲毛区・花見川区・美浜区にお住まいの方")
    for col, label in zip(_WEEKDAY_COLS, _WEEKDAY_LABELS, strict=True):
        ws.cell(row=3, column=col, value=label)
    # Row 4: 10:00- — one of each status + 定休日 in Sunday col
    ws.cell(row=4, column=2, value="10:00-")
    ws.cell(row=4, column=3, value="×")  # 月 → unavailable
    ws.cell(row=4, column=5, value="○")  # 火 → available
    ws.cell(row=4, column=7, value="△")  # 水 → consult
    ws.cell(row=4, column=9, value="△")  # 木 → consult
    ws.cell(row=4, column=11, value="×")  # 金 → unavailable
    ws.cell(row=4, column=13, value="△")  # 土 → consult
    ws.cell(row=4, column=15, value="定休日")  # 日 → skip
    # Row 6: 11:00-  all ○ (Sunday blank → unavailable)
    ws.cell(row=6, column=2, value="11:00-")
    for col in [3, 5, 7, 9, 11, 13]:
        ws.cell(row=6, column=col, value="○")
    ws.cell(row=6, column=15, value=None)  # Sunday blank → unavailable
    # Rows 8-18: fill with ×
    for row_num in [8, 10, 12, 14, 16, 18]:
        hour = 12 + (row_num - 8) // 2
        ws.cell(row=row_num, column=2, value=f"{hour}:00-")
        for col in [3, 5, 7, 9, 11, 13]:
            ws.cell(row=row_num, column=col, value="×")
    ws.cell(row=20, column=2, value="表示の意味合い：○受け入れ可能　△相談ください　×受け入れ枠なし")

    # ── Block 2: 都賀拠点 ──
    ws.cell(row=22, column=2, value="中央区・若葉区にお住まいの方")
    for col, label in zip(_WEEKDAY_COLS, _WEEKDAY_LABELS, strict=True):
        ws.cell(row=23, column=col, value=label)
    ws.cell(row=24, column=2, value="10:00-")
    for col in [3, 5, 7, 9, 11, 13]:
        ws.cell(row=24, column=col, value="○")
    ws.cell(row=24, column=15, value="定休日")  # 日 → skip
    for row_num in [26, 28, 30, 32, 34, 36, 38]:
        hour = 11 + (row_num - 26) // 2
        ws.cell(row=row_num, column=2, value=f"{hour}:00-")
        for col in [3, 5, 7, 9, 11, 13]:
            ws.cell(row=row_num, column=col, value="△")
    ws.cell(row=40, column=2, value="表示の意味合い：○受け入れ可能　△相談ください　×受け入れ枠なし")

    out = tmp_path / "test_acceptance.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# 1. パース正常系
# ---------------------------------------------------------------------------


def test_parse_sheet_returns_two_blocks(tmp_path: Path) -> None:
    """parse_sheet returns exactly 2 BlockResult objects (one per 拠点)."""
    xlsx = _build_fixture_xlsx(tmp_path)
    blocks = parse_sheet(xlsx)
    assert len(blocks) == 2
    assert blocks[0].office_like == "稲毛"
    assert blocks[0].area_label == "稲毛区・花見川区・美浜区"
    assert blocks[1].office_like == "都賀"
    assert blocks[1].area_label == "中央区・若葉区"


def test_parse_sheet_block1_entry_structure(tmp_path: Path) -> None:
    """Block 1 entries have correct weekday / time_slot / status fields."""
    xlsx = _build_fixture_xlsx(tmp_path)
    blocks = parse_sheet(xlsx)
    b1 = blocks[0]

    # The 10:00 row has 定休日 in Sunday (col O) → 6 entries (Mon-Sat only)
    row_10 = [e for e in b1.entries if e.time_slot == time(10, 0)]
    assert len(row_10) == 6, f"Expected 6 entries for 10:00 (日=定休日 skipped), got {len(row_10)}"

    # Verify specific statuses from fixture data
    mon_10 = next(e for e in row_10 if e.weekday == 0)
    assert mon_10.status == "unavailable"  # ×

    tue_10 = next(e for e in row_10 if e.weekday == 1)
    assert tue_10.status == "available"  # ○

    wed_10 = next(e for e in row_10 if e.weekday == 2)
    assert wed_10.status == "consult"  # △


# ---------------------------------------------------------------------------
# 2. 「定休日」セルがスキップされる
# ---------------------------------------------------------------------------


def test_teikyu_cell_is_skipped(tmp_path: Path) -> None:
    """「定休日」value generates no entry (Sunday / weekday=6 absent in 10:00 row)."""
    xlsx = _build_fixture_xlsx(tmp_path)
    blocks = parse_sheet(xlsx)
    b1 = blocks[0]

    # Sunday (weekday=6) at 10:00 must NOT exist (defined as 定休日 in fixture)
    sunday_10 = [e for e in b1.entries if e.weekday == 6 and e.time_slot == time(10, 0)]
    assert sunday_10 == [], "定休日 cell must not create an entry"


def test_parse_status_teikyu_returns_none() -> None:
    """_parse_status returns None for '定休日' (sentinel for skip)."""
    assert _parse_status("定休日") is None


def test_parse_status_maps_symbols() -> None:
    assert _parse_status("○") == "available"
    assert _parse_status("△") == "consult"
    assert _parse_status("×") == "unavailable"
    assert _parse_status(None) == "unavailable"  # blank → unavailable
    assert _parse_status("") == "unavailable"  # empty string → unavailable


# ---------------------------------------------------------------------------
# 3. _parse_time_slot helper
# ---------------------------------------------------------------------------


def test_parse_time_slot_valid() -> None:
    assert _parse_time_slot("10:00-") == time(10, 0, 0)
    assert _parse_time_slot("17:00-") == time(17, 0, 0)
    assert _parse_time_slot("9:30-") == time(9, 30, 0)


def test_parse_time_slot_invalid_returns_none() -> None:
    assert _parse_time_slot(None) is None
    assert _parse_time_slot("定休日") is None
    assert _parse_time_slot("") is None
    assert _parse_time_slot("表示の意味合い") is None


# ---------------------------------------------------------------------------
# 4. 拠点未登録時 SystemExit (ERROR 終了)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_resolve_office_raises_on_missing(db) -> None:
    """_resolve_office raises SystemExit when no matching office exists."""
    with pytest.raises(SystemExit) as exc_info:
        await _resolve_office(db, "稲毛", "稲毛区・花見川区・美浜区")
    assert "ERROR" in str(exc_info.value)


# ---------------------------------------------------------------------------
# 5. dry-run: DB に書き込まない
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_dry_run_does_not_write_db(tmp_path: Path, db, capsys) -> None:
    """dry-run mode prints summary but writes nothing to acceptance_calendar."""
    from import_acceptance_calendar import import_acceptance_calendar
    from sqlalchemy import func, select

    from app.models.acceptance_calendar import AcceptanceCalendar

    xlsx = _build_fixture_xlsx(tmp_path)
    await import_acceptance_calendar(xlsx, dry_run=True)

    # DB must remain empty
    count = await db.scalar(select(func.count()).select_from(AcceptanceCalendar))
    assert count == 0

    captured = capsys.readouterr()
    assert "受け入れカレンダー" in captured.out
    assert "dry-run" in captured.out


# ---------------------------------------------------------------------------
# 6. apply モード: _upsert_block が DB に行を書く
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_upsert_block_inserts_entries(tmp_path: Path, db) -> None:
    """_upsert_block inserts acceptance_calendar rows via the test DB session."""
    from import_acceptance_calendar import _upsert_block
    from sqlalchemy import func, select

    from app.models.acceptance_calendar import AcceptanceCalendar
    from app.models.office import Office

    # Seed one office
    office = Office(name="稲毛ステーション")
    db.add(office)
    await db.commit()
    await db.refresh(office)

    # Build a minimal BlockResult from parse_sheet
    xlsx = _build_fixture_xlsx(tmp_path)
    blocks = parse_sheet(xlsx)
    inamo_block = blocks[0]  # 稲毛拠点

    counters: dict[str, int] = {"new": 0, "update": 0}
    await _upsert_block(db, office, inamo_block, counters)
    await db.commit()

    count = await db.scalar(select(func.count()).select_from(AcceptanceCalendar))
    assert count == len(inamo_block.entries), "All parsed entries must be inserted"
    assert counters["new"] == len(inamo_block.entries)


# ---------------------------------------------------------------------------
# 7. ブロック 2 (都賀拠点) のパース正常系
# ---------------------------------------------------------------------------


def test_parse_sheet_block2_all_available_at_10(tmp_path: Path) -> None:
    """Block 2 at 10:00 has 6 'available' entries (Mon-Sat; 定休日 on Sun)."""
    xlsx = _build_fixture_xlsx(tmp_path)
    blocks = parse_sheet(xlsx)
    b2 = blocks[1]

    row_10 = [e for e in b2.entries if e.time_slot == time(10, 0)]
    assert len(row_10) == 6  # Sunday skipped
    assert all(e.status == "available" for e in row_10)


# ---------------------------------------------------------------------------
# 8. エラーケース
# ---------------------------------------------------------------------------


def test_parse_sheet_file_not_found(tmp_path: Path) -> None:
    """parse_sheet raises FileNotFoundError for missing file."""
    with pytest.raises(FileNotFoundError):
        parse_sheet(tmp_path / "nonexistent.xlsx")


def test_parse_sheet_missing_sheet(tmp_path: Path) -> None:
    """parse_sheet raises KeyError when sheet name is not in workbook."""
    wb = openpyxl.Workbook()
    wb.active.title = "wrong_sheet"
    bad_xlsx = tmp_path / "bad.xlsx"
    wb.save(bad_xlsx)
    with pytest.raises(KeyError):
        parse_sheet(bad_xlsx)


# ---------------------------------------------------------------------------
# 9. 冪等性: 2 回 upsert しても件数が変わらない
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_upsert_block_idempotent(tmp_path: Path, db) -> None:
    """同一 payload を 2 回 upsert しても acceptance_calendar の件数が変わらない."""
    from sqlalchemy import func, select

    from app.models.acceptance_calendar import AcceptanceCalendar
    from app.models.office import Office

    # Seed one office
    office = Office(name="稲毛ステーション")
    db.add(office)
    await db.commit()
    await db.refresh(office)

    xlsx = _build_fixture_xlsx(tmp_path)
    blocks = parse_sheet(xlsx)
    inamo_block = blocks[0]

    # 1 回目
    counters1: dict[str, int] = {"new": 0, "update": 0}
    await _upsert_block(db, office, inamo_block, counters1)
    await db.commit()

    count_after_first = await db.scalar(select(func.count()).select_from(AcceptanceCalendar))

    # 2 回目 (同一 payload)
    counters2: dict[str, int] = {"new": 0, "update": 0}
    await _upsert_block(db, office, inamo_block, counters2)
    await db.commit()

    count_after_second = await db.scalar(select(func.count()).select_from(AcceptanceCalendar))

    assert count_after_first == count_after_second, (
        f"Idempotency violated: {count_after_first} → {count_after_second}"
    )


# ---------------------------------------------------------------------------
# 10. 不正 status: 未知シンボルは None を返す
# ---------------------------------------------------------------------------


def test_parse_status_unknown_symbol() -> None:
    """未知のシンボルは None を返す (skip 扱い)."""
    assert _parse_status("UNKNOWN") is None
    assert _parse_status("?") is None


# ---------------------------------------------------------------------------
# 11. apply 統合テスト: xlsx → import_acceptance_calendar で期待件数 INSERT
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_acceptance_calendar_apply_full(tmp_path: Path, db) -> None:
    """xlsx → import_acceptance_calendar(apply=True) で acceptance_calendar 行が期待件数 INSERT."""
    from sqlalchemy import func, select

    from app.models.acceptance_calendar import AcceptanceCalendar
    from app.models.office import Office

    # Seed both offices
    inamo_office = Office(name="稲毛ステーション")
    tsuga_office = Office(name="都賀ステーション")
    db.add(inamo_office)
    db.add(tsuga_office)
    await db.commit()

    # Patch get_session_factory to return a factory that yields our test `db`
    from contextlib import asynccontextmanager
    from unittest.mock import MagicMock, patch

    import app.db.session as _db_session

    @asynccontextmanager
    async def _fake_factory():
        yield db

    fake_factory_callable = MagicMock(return_value=_fake_factory())

    xlsx = _build_fixture_xlsx(tmp_path)

    # Parse expected counts before apply so we can assert exact numbers
    blocks = parse_sheet(xlsx)
    expected_inamo = len(blocks[0].entries)
    expected_tsuga = len(blocks[1].entries)
    expected_total = expected_inamo + expected_tsuga

    with patch.object(_db_session, "get_session_factory", return_value=fake_factory_callable):
        await import_acceptance_calendar(xlsx, dry_run=False)

    total = await db.scalar(select(func.count()).select_from(AcceptanceCalendar))
    assert total == expected_total, (
        f"Expected {expected_total} acceptance_calendar rows, got {total}"
    )
