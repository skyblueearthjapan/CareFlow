"""Tests for scripts/import_schedule_template.py (W15 Phase2 B-1).

カバー範囲:
1. パース正常系: フィクスチャ xlsx から正しい CourseBlock / VisitSlot が構築される
2. 拠点認識: 住所文字列から正しく office_name を推論できる
3. 患者突合: name 完全一致 / 住所部分一致 / 突合失敗 各シナリオ
4. dry-run mode: DB 書き込み 0 件
5. apply mode: 期待件数 INSERT (mock DB)
6. パース補助関数
"""

from __future__ import annotations

import sys
import uuid
from datetime import time
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest

# sys.path にスクリプトディレクトリを追加
_BACKEND_ROOT = Path(__file__).resolve().parent.parent.parent
_SCRIPTS_DIR = _BACKEND_ROOT / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))

from import_schedule_template import (  # noqa: E402
    OfficeSection,
    _parse_cap,
    _parse_time,
    infer_office_name,
    parse_sheet,
    run_import,
)

# ---------------------------------------------------------------------------
# フィクスチャ: 最小 xlsx を in-memory で生成
# ---------------------------------------------------------------------------


def _make_fixture_xlsx(tmp_path: Path) -> Path:
    """最小構成の「スケジュール枠組み（仮）」シートを含む xlsx を生成する。

    構造:
      Row 1:  曜日ヘッダ (B=月曜日, G=火曜日)
      Row 2:  コースラベル A (cap: Mon=3, Tue=2)
      Row 3:  フィールドヘッダ
      Row 4:  データ行 (Mon 9:30 / 田中太郎 / 千葉県千葉市稲毛区園生町1-1)
      Row 5:  データ行 (Tue 10:00 / 空 / -)
      Row 6:  データ行 (Mon 10:30 / 鈴木花子 / 千葉県千葉市稲毛区天台1-1)
      Row 7:  空行 (ブロック終端)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "スケジュール枠組み（仮）"

    # Row 1: 曜日ヘッダ
    ws.cell(row=1, column=2).value = "月曜日"
    ws.cell(row=1, column=7).value = "火曜日"
    ws.cell(row=1, column=12).value = "水曜日"
    ws.cell(row=1, column=17).value = "木曜日"
    ws.cell(row=1, column=22).value = "金曜日"
    ws.cell(row=1, column=27).value = "土曜日"

    # Row 2: 件数 + コースラベル A
    ws.cell(row=2, column=1).value = 3  # cap Mon
    ws.cell(row=2, column=2).value = "A"
    ws.cell(row=2, column=6).value = 2  # cap Tue
    ws.cell(row=2, column=7).value = "A"
    ws.cell(row=2, column=11).value = "稲毛"  # cap Wed = 不可
    ws.cell(row=2, column=12).value = "A"
    ws.cell(row=2, column=16).value = 1  # cap Thu
    ws.cell(row=2, column=17).value = "A"
    ws.cell(row=2, column=21).value = 2  # cap Fri
    ws.cell(row=2, column=22).value = "A"
    ws.cell(row=2, column=26).value = "稲毛"  # cap Sat = 不可
    ws.cell(row=2, column=27).value = "A"

    # Row 3: フィールドヘッダ
    for start_col in [1, 6, 11, 16, 21, 26]:
        ws.cell(row=3, column=start_col).value = "時間帯"
        ws.cell(row=3, column=start_col + 1).value = "氏名"
        ws.cell(row=3, column=start_col + 2).value = "住所"
        ws.cell(row=3, column=start_col + 3).value = "複数"
        ws.cell(row=3, column=start_col + 4).value = "条件"

    # Row 4: データ行 (Mon: 田中太郎, Tue: 前川双子)
    ws.cell(row=4, column=1).value = time(9, 30)
    ws.cell(row=4, column=2).value = "田中太郎"
    ws.cell(row=4, column=3).value = "千葉県千葉市稲毛区園生町1-1"
    ws.cell(row=4, column=6).value = time(9, 30)
    ws.cell(row=4, column=7).value = "前川双子"
    ws.cell(row=4, column=8).value = "千葉県千葉市美浜区磯辺5-1"
    ws.cell(row=4, column=11).value = time(9, 30)  # 水曜 (稲毛不可) → データあっても無視

    # Row 5: Mon 10:00 空, Tue 10:00 空
    ws.cell(row=5, column=1).value = time(10, 0)
    ws.cell(row=5, column=6).value = time(10, 0)

    # Row 6: Mon 10:30 鈴木花子 (条件付き)
    ws.cell(row=6, column=1).value = time(10, 30)
    ws.cell(row=6, column=2).value = "鈴木花子"
    ws.cell(row=6, column=3).value = "千葉県千葉市稲毛区天台1-1"
    ws.cell(row=6, column=5).value = "女性のみ"

    # Row 7: 完全空行 (ブロック終端)

    xlsx_path = tmp_path / "fixture_schedule.xlsx"
    wb.save(str(xlsx_path))
    return xlsx_path


# ---------------------------------------------------------------------------
# テスト 1: パース補助関数
# ---------------------------------------------------------------------------


class TestParseHelpers:
    """_parse_cap / _parse_time の単体テスト."""

    def test_parse_cap_number(self) -> None:
        assert _parse_cap(6.0) == 6

    def test_parse_cap_inamoto_marker(self) -> None:
        assert _parse_cap("稲毛") == 0

    def test_parse_cap_toga_marker(self) -> None:
        assert _parse_cap("都賀") == 0

    def test_parse_cap_none(self) -> None:
        assert _parse_cap(None) == 0

    def test_parse_time_time_obj(self) -> None:
        assert _parse_time(time(9, 30)) == time(9, 30)

    def test_parse_time_string(self) -> None:
        assert _parse_time("09:30:00") == time(9, 30)

    def test_parse_time_none(self) -> None:
        assert _parse_time(None) is None


# ---------------------------------------------------------------------------
# テスト 2: Excel パース正常系
# ---------------------------------------------------------------------------


class TestParseSheet:
    """parse_sheet() の正常系テスト."""

    def test_returns_one_section(self, tmp_path: Path) -> None:
        xlsx_path = _make_fixture_xlsx(tmp_path)
        sections = parse_sheet(xlsx_path)
        assert len(sections) == 1

    def test_course_label(self, tmp_path: Path) -> None:
        xlsx_path = _make_fixture_xlsx(tmp_path)
        sections = parse_sheet(xlsx_path)
        block = sections[0].course_blocks[0]
        assert block.label == "A"

    def test_capacities(self, tmp_path: Path) -> None:
        xlsx_path = _make_fixture_xlsx(tmp_path)
        sections = parse_sheet(xlsx_path)
        block = sections[0].course_blocks[0]
        assert block.capacities["mon"] == 3
        assert block.capacities["tue"] == 2
        assert block.capacities["wed"] == 0  # 稲毛 = 不可
        assert block.capacities["sat"] == 0  # 稲毛 = 不可

    def test_visits_count(self, tmp_path: Path) -> None:
        xlsx_path = _make_fixture_xlsx(tmp_path)
        sections = parse_sheet(xlsx_path)
        block = sections[0].course_blocks[0]
        # Mon: 田中太郎, 鈴木花子 / Tue: 前川双子
        names = {v.name for v in block.visits}
        assert "田中太郎" in names
        assert "鈴木花子" in names
        assert "前川双子" in names

    def test_condition_captured(self, tmp_path: Path) -> None:
        xlsx_path = _make_fixture_xlsx(tmp_path)
        sections = parse_sheet(xlsx_path)
        block = sections[0].course_blocks[0]
        assert "女性のみ" in block.notes_set

    def test_empty_slot_excluded(self, tmp_path: Path) -> None:
        """「空」スロットは VisitSlot に含まれない。"""
        xlsx_path = _make_fixture_xlsx(tmp_path)
        sections = parse_sheet(xlsx_path)
        block = sections[0].course_blocks[0]
        names = [v.name for v in block.visits]
        assert "空" not in names


# ---------------------------------------------------------------------------
# テスト 3: 拠点認識
# ---------------------------------------------------------------------------


class TestInferOfficeName:
    """infer_office_name() の単体テスト."""

    def test_infer_office1_by_address(self) -> None:
        section = OfficeSection(
            office_hint=None,
            course_blocks=[],
            addresses=[
                "千葉県千葉市稲毛区園生町1-1",
                "千葉県千葉市花見川区さつきが丘1-1",
                "千葉県千葉市美浜区磯辺5-1",
            ],
        )
        assert infer_office_name(section) == "宮野木"

    def test_infer_office2_by_hint(self) -> None:
        section = OfficeSection(
            office_hint="都賀",
            course_blocks=[],
            addresses=[
                "千葉県千葉市若葉区都賀1-1",
                "千葉県千葉市中央区都町1-1",
            ],
        )
        assert infer_office_name(section) == "都賀"

    def test_infer_office2_by_address(self) -> None:
        section = OfficeSection(
            office_hint=None,
            course_blocks=[],
            addresses=[
                "千葉県千葉市中央区都町1-1",
                "千葉県千葉市若葉区若松町1-1",
            ],
        )
        assert infer_office_name(section) == "都賀"


# ---------------------------------------------------------------------------
# テスト 4: dry-run モード (DB 書き込みなし)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_dry_run_no_db_writes(tmp_path: Path) -> None:
    """dry-run では session.commit() が呼ばれない。"""
    xlsx_path = _make_fixture_xlsx(tmp_path)

    # DB をモック
    mock_session = AsyncMock()
    mock_session.__aenter__ = AsyncMock(return_value=mock_session)
    mock_session.__aexit__ = AsyncMock(return_value=False)

    # scalars().first() → None (患者なし、拠点なし)
    mock_scalars = AsyncMock()
    mock_scalars.first = MagicMock(return_value=None)
    mock_scalars.all = MagicMock(return_value=[])
    mock_session.scalars = AsyncMock(return_value=mock_scalars)
    mock_session.flush = AsyncMock()
    mock_session.commit = AsyncMock()
    mock_session.rollback = AsyncMock()
    mock_session.add = MagicMock()

    mock_factory = MagicMock(return_value=mock_session)

    with (
        patch("import_schedule_template.get_session_factory", return_value=mock_factory),
        patch("import_schedule_template.dispose_engine", new_callable=AsyncMock),
    ):
        await run_import(xlsx_path, apply=False)

    # dry-run: commit() は呼ばれない
    mock_session.commit.assert_not_called()
    # rollback() は呼ばれる
    mock_session.rollback.assert_called_once()


# ---------------------------------------------------------------------------
# テスト 5: apply モード (期待件数 INSERT)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_apply_mode_inserts(tmp_path: Path) -> None:
    """apply モードで拠点/CT は INSERT され、患者突合成功なら PFV も INSERT される。"""
    xlsx_path = _make_fixture_xlsx(tmp_path)

    # 拠点 mock
    mock_office = MagicMock()
    mock_office.id = uuid.uuid4()

    # 患者 mock (名前に関係なく返す)
    mock_patient = MagicMock()
    mock_patient.id = uuid.uuid4()
    mock_patient.name = "田中太郎"
    mock_patient.address = "千葉県千葉市稲毛区園生町1-1"

    call_count = 0

    async def mock_scalars(stmt):  # noqa: ANN001
        nonlocal call_count
        call_count += 1
        mock_result = MagicMock()
        # 1回目: office検索 → office返す
        # 2回目以降: CT検索 → None (新規INSERT), PFV検索 → None (新規INSERT)
        # 患者検索 → patient返す
        stmt_str = str(stmt)
        if "offices" in stmt_str:
            mock_result.first = MagicMock(return_value=mock_office)
            mock_result.all = MagicMock(return_value=[mock_office])
        elif "course_templates" in stmt_str:
            mock_result.first = MagicMock(return_value=None)
            mock_result.all = MagicMock(return_value=[])
        elif "patients" in stmt_str:
            mock_result.first = MagicMock(return_value=mock_patient)
            mock_result.all = MagicMock(return_value=[mock_patient])
        else:
            mock_result.first = MagicMock(return_value=None)
            mock_result.all = MagicMock(return_value=[])
        return mock_result

    mock_session = AsyncMock()
    mock_session.__aenter__ = AsyncMock(return_value=mock_session)
    mock_session.__aexit__ = AsyncMock(return_value=False)
    mock_session.scalars = AsyncMock(side_effect=mock_scalars)
    mock_session.flush = AsyncMock()
    mock_session.commit = AsyncMock()
    mock_session.rollback = AsyncMock()
    mock_session.add = MagicMock()

    mock_factory = MagicMock(return_value=mock_session)

    with (
        patch("import_schedule_template.get_session_factory", return_value=mock_factory),
        patch("import_schedule_template.dispose_engine", new_callable=AsyncMock),
    ):
        summary = await run_import(xlsx_path, apply=True)

    # apply: commit() が呼ばれる
    mock_session.commit.assert_called_once()
    # CourseTemplate は 1 件 NEW
    assert summary.ct_new == 1
    assert summary.ct_update == 0


# ---------------------------------------------------------------------------
# テスト 6: 患者突合
# ---------------------------------------------------------------------------


class TestFindPatient:
    """_find_patient() の突合ロジックを間接的に検証。"""

    @pytest.mark.asyncio
    async def test_name_exact_match(self, tmp_path: Path) -> None:
        """name 完全一致で患者が見つかる場合、突合成功になる。"""
        xlsx_path = _make_fixture_xlsx(tmp_path)

        mock_patient = MagicMock()
        mock_patient.id = uuid.uuid4()
        mock_patient.name = "田中太郎"
        mock_patient.address = "千葉県千葉市稲毛区園生町1-1"

        async def mock_scalars(stmt):  # noqa: ANN001
            mock_result = MagicMock()
            stmt_str = str(stmt)
            if "offices" in stmt_str:
                mock_office = MagicMock()
                mock_office.id = uuid.uuid4()
                mock_result.first = MagicMock(return_value=mock_office)
                mock_result.all = MagicMock(return_value=[mock_office])
            elif "course_templates" in stmt_str:
                mock_result.first = MagicMock(return_value=None)
                mock_result.all = MagicMock(return_value=[])
            elif "patients" in stmt_str:
                mock_result.first = MagicMock(return_value=mock_patient)
                mock_result.all = MagicMock(return_value=[mock_patient])
            else:
                mock_result.first = MagicMock(return_value=None)
                mock_result.all = MagicMock(return_value=[])
            return mock_result

        mock_session = AsyncMock()
        mock_session.__aenter__ = AsyncMock(return_value=mock_session)
        mock_session.__aexit__ = AsyncMock(return_value=False)
        mock_session.scalars = AsyncMock(side_effect=mock_scalars)
        mock_session.flush = AsyncMock()
        mock_session.commit = AsyncMock()
        mock_session.rollback = AsyncMock()
        mock_session.add = MagicMock()

        mock_factory = MagicMock(return_value=mock_session)

        with (
            patch("import_schedule_template.get_session_factory", return_value=mock_factory),
            patch("import_schedule_template.dispose_engine", new_callable=AsyncMock),
        ):
            summary = await run_import(xlsx_path, apply=False)

        # 少なくとも 1 件は突合成功
        assert len(summary.matched) >= 1

    @pytest.mark.asyncio
    async def test_no_match_unmatched_list(self, tmp_path: Path) -> None:
        """患者が見つからない場合、unmatched リストに追加される。"""
        xlsx_path = _make_fixture_xlsx(tmp_path)

        async def mock_scalars(stmt):  # noqa: ANN001
            mock_result = MagicMock()
            stmt_str = str(stmt)
            if "offices" in stmt_str:
                mock_office = MagicMock()
                mock_office.id = uuid.uuid4()
                mock_result.first = MagicMock(return_value=mock_office)
                mock_result.all = MagicMock(return_value=[mock_office])
            elif "course_templates" in stmt_str:
                mock_result.first = MagicMock(return_value=None)
                mock_result.all = MagicMock(return_value=[])
            else:
                # 患者: 全て None
                mock_result.first = MagicMock(return_value=None)
                mock_result.all = MagicMock(return_value=[])
            return mock_result

        mock_session = AsyncMock()
        mock_session.__aenter__ = AsyncMock(return_value=mock_session)
        mock_session.__aexit__ = AsyncMock(return_value=False)
        mock_session.scalars = AsyncMock(side_effect=mock_scalars)
        mock_session.flush = AsyncMock()
        mock_session.commit = AsyncMock()
        mock_session.rollback = AsyncMock()
        mock_session.add = MagicMock()

        mock_factory = MagicMock(return_value=mock_session)

        with (
            patch("import_schedule_template.get_session_factory", return_value=mock_factory),
            patch("import_schedule_template.dispose_engine", new_callable=AsyncMock),
        ):
            summary = await run_import(xlsx_path, apply=False)

        assert len(summary.unmatched) >= 1
        assert len(summary.matched) == 0


# ---------------------------------------------------------------------------
# テスト 7: 都賀拠点未登録時 SystemExit(1)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_apply_raises_system_exit_when_office_missing(tmp_path: Path) -> None:
    """--apply 時に拠点が DB に未登録なら SystemExit(1) を raise する。"""
    xlsx_path = _make_fixture_xlsx(tmp_path)

    async def mock_scalars_no_office(stmt):  # noqa: ANN001
        mock_result = MagicMock()
        # offices / patients / course_templates すべて None を返す
        mock_result.first = MagicMock(return_value=None)
        mock_result.all = MagicMock(return_value=[])
        return mock_result

    mock_session = AsyncMock()
    mock_session.__aenter__ = AsyncMock(return_value=mock_session)
    mock_session.__aexit__ = AsyncMock(return_value=False)
    mock_session.scalars = AsyncMock(side_effect=mock_scalars_no_office)
    mock_session.flush = AsyncMock()
    mock_session.commit = AsyncMock()
    mock_session.rollback = AsyncMock()
    mock_session.add = MagicMock()

    mock_factory = MagicMock(return_value=mock_session)

    with (
        patch("import_schedule_template.get_session_factory", return_value=mock_factory),
        patch("import_schedule_template.dispose_engine", new_callable=AsyncMock),
        pytest.raises(SystemExit) as exc_info,
    ):
        await run_import(xlsx_path, apply=True)

    assert exc_info.value.code == 1


# ---------------------------------------------------------------------------
# テスト 8: PFV upsert で論理削除列なし → deleted_at フィルタ不要を確認
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_upsert_without_deleted_at_filter(tmp_path: Path) -> None:
    """PatientFixedVisit に deleted_at 列が存在しないため、
    upsert SELECT は deleted_at IS NULL フィルタなしで正常に動作する。"""
    xlsx_path = _make_fixture_xlsx(tmp_path)

    mock_office = MagicMock()
    mock_office.id = uuid.uuid4()

    mock_patient = MagicMock()
    mock_patient.id = uuid.uuid4()
    mock_patient.name = "田中太郎"
    mock_patient.address = "千葉県千葉市稲毛区園生町1-1"

    # 既存 PFV (active) を返す mock → "update" パスに入る
    mock_pfv = MagicMock()
    mock_pfv.start_time = None  # 差異を作って changed=True にする

    async def mock_scalars(stmt):  # noqa: ANN001
        mock_result = MagicMock()
        stmt_str = str(stmt)
        if "offices" in stmt_str:
            mock_result.first = MagicMock(return_value=mock_office)
            mock_result.all = MagicMock(return_value=[mock_office])
        elif "course_templates" in stmt_str:
            mock_result.first = MagicMock(return_value=None)
            mock_result.all = MagicMock(return_value=[])
        elif "patients" in stmt_str:
            mock_result.first = MagicMock(return_value=mock_patient)
            mock_result.all = MagicMock(return_value=[mock_patient])
        elif "patient_fixed_visits" in stmt_str:
            # 既存 PFV を返す → "update" パス
            mock_result.first = MagicMock(return_value=mock_pfv)
            mock_result.all = MagicMock(return_value=[mock_pfv])
        else:
            mock_result.first = MagicMock(return_value=None)
            mock_result.all = MagicMock(return_value=[])
        return mock_result

    mock_session = AsyncMock()
    mock_session.__aenter__ = AsyncMock(return_value=mock_session)
    mock_session.__aexit__ = AsyncMock(return_value=False)
    mock_session.scalars = AsyncMock(side_effect=mock_scalars)
    mock_session.flush = AsyncMock()
    mock_session.commit = AsyncMock()
    mock_session.rollback = AsyncMock()
    mock_session.add = MagicMock()

    mock_factory = MagicMock(return_value=mock_session)

    with (
        patch("import_schedule_template.get_session_factory", return_value=mock_factory),
        patch("import_schedule_template.dispose_engine", new_callable=AsyncMock),
    ):
        summary = await run_import(xlsx_path, apply=False)

    # dry-run でもマッチが起きていること (PFV は update パス)
    assert summary.pfv_update >= 1
