"""Tests for /api/v1/patients/import-export/* (Excel import / export).

Coverage map (spec §6):
  1.  export: 0 名のとき空テンプレート、N 名のとき N+1 行
  2.  export: PFV が patient_code でリンクされている
  3.  import dry_run: 新規行が検出される
  4.  import dry_run: 更新行で changes が diff される
  5.  import dry_run: <DELETE> フラグで delete 候補になる
  6.  import dry_run: <CLEAR> で NULL 上書き候補になる
  7.  import dry_run: 空セルは「変更しない」 (noop)
  8.  import dry_run: バリデーションエラー (UUID 不正、enum 違反、必須欠落)
  9.  import dry_run: DB は変更されない
  10. import apply: 新規 + 更新 + 削除が 1 transaction で反映
  11. import apply: partial commit — error 行は skip、有効行のみ反映
  12. import apply: PFV の物理削除
  13. import apply: course_template_code → course_template_id 解決
  14. import apply: 拠点コード → primary_office_id 解決
  15. import apply: patient_code 重複で error (skip され、有効行のみ反映)
  16. RBAC: staff は 403
  17. template ダウンロード: 1 行のみ (ヘッダー)
  27. import apply: 全件 error → transaction_applied=False、DB 変更なし
  28. import apply: pfv_error 1 件 + patient_new 2 件 → patient は反映、pfv は skip
"""

from __future__ import annotations

from datetime import UTC, datetime, time
from io import BytesIO
from uuid import UUID, uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_, select

from app.core.security import create_access_token, hash_password
from app.models import (
    City,
    CourseTemplate,
    Office,
    OfficeCity,
    Patient,
    PatientFixedVisit,
    User,
)
from app.models.patient_ng_staff import PatientNgStaff
from app.models.patient_same_address_link import PatientSameAddressLink
from app.models.staff import Staff
from app.services.patient_excel.schema import (
    PATIENT_COL_INDEX,
    PATIENT_COLUMNS,
    PFV_COL_INDEX,
    PFV_COLUMNS,
    PFV_EDIT_COL_INDEX,
    PFV_EDIT_COLUMNS,
    SHEET_PATIENTS,
    SHEET_PFV,
    SHEET_PFV_EDIT,
    SHEET_PFV_GRID,
    SHEET_WEEKLY,
    WEEKLY_COL_INDEX,
)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


async def _make_user(db, email: str, role: str) -> User:
    user = User(email=email, password_hash=hash_password("pw"), role=role)
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


def _bearer(user: User) -> dict[str, str]:
    token = create_access_token(subject=user.id, role=user.role, staff_id=user.staff_id)
    return {"Authorization": f"Bearer {token}"}


async def _make_office(db, code: str, name: str | None = None) -> Office:
    o = Office(code=code, name=name or code)
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return o


async def _make_patient(db, **overrides) -> Patient:
    defaults = {
        "code": "P-DEFAULT",
        "name": "デフォルト患者",
        "sex": "male",
        "status": "active",
        "address": "千葉市稲毛区test1",
    }
    defaults.update(overrides)
    p = Patient(**defaults)
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


async def _make_pfv(
    db,
    *,
    patient_id,
    weekday: int = 0,
    slot_index: int = 0,
    mode: str = "normal",
    start_time=time(9, 0),
    duration_min: int = 30,
    course_template_id=None,
) -> PatientFixedVisit:
    pfv = PatientFixedVisit(
        patient_id=patient_id,
        mode=mode,
        weekday=weekday,
        slot_index=slot_index,
        start_time=start_time,
        duration_min=duration_min,
        course_template_id=course_template_id,
    )
    db.add(pfv)
    await db.commit()
    await db.refresh(pfv)
    return pfv


def _patient_headers() -> list[str]:
    return [str(c["header"]) for c in PATIENT_COLUMNS]


def _pfv_headers() -> list[str]:
    return [str(c["header"]) for c in PFV_COLUMNS]


def _build_workbook_bytes(
    *,
    patient_rows: list[dict] | None = None,
    pfv_rows: list[dict] | None = None,
) -> bytes:
    """テスト用ヘルパー: 各 dict は { col_key: value } を持つ.

    値は str / int / float / None / time など何でもよい (openpyxl 任せ).
    """
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    for col_idx, header in enumerate(_patient_headers(), start=1):
        ws_p.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(patient_rows or [], start=2):
        for col_key, idx in PATIENT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_p.cell(row=r_idx, column=idx + 1, value=v)

    ws_f = wb.create_sheet(title=SHEET_PFV)
    for col_idx, header in enumerate(_pfv_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(pfv_rows or [], start=2):
        for col_key, idx in PFV_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _edit_pfv_headers() -> list[str]:
    return [str(c["header"]) for c in PFV_EDIT_COLUMNS]


def _build_edit_workbook_bytes(
    *,
    patient_rows: list[dict] | None = None,
    pfv_edit_rows: list[dict] | None = None,
) -> bytes:
    """Phase G-51: 新「固定訪問パターン（編集用）」(患者 1 行) シート付き workbook.

    pfv_edit_rows の各 dict は PFV_EDIT_COL_INDEX のキー (patient_code, mon_time,
    mon_course, service_minutes, ...) を持つ.
    """
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    for col_idx, header in enumerate(_patient_headers(), start=1):
        ws_p.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(patient_rows or [], start=2):
        for col_key, idx in PATIENT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_p.cell(row=r_idx, column=idx + 1, value=v)

    ws_f = wb.create_sheet(title=SHEET_PFV_EDIT)
    for col_idx, header in enumerate(_edit_pfv_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(pfv_edit_rows or [], start=2):
        for col_key, idx in PFV_EDIT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload(client, admin, *, content: bytes, dry_run: bool = True):
    """multipart helper. fastapi UploadFile は ``file`` という名前で受ける."""
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    return await client.post(
        f"/api/v1/patients/import-export/import?dry_run={'true' if dry_run else 'false'}",
        headers=_bearer(admin),
        files=files,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


# 1) export: 0 名のとき empty + 1 名のとき 2 行
@pytest.mark.asyncio
async def test_export_empty_db_returns_template_only(client, db) -> None:
    admin = await _make_user(db, "ex-empty@example.com", "admin")
    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200, res.text
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_PATIENTS in wb.sheetnames
    ws = wb[SHEET_PATIENTS]
    # ヘッダー 1 行のみ (max_row が 1 になることを許容)
    assert ws.max_row == 1
    # Phase E-7 (gap P2): クロスオフィス warning が 0 件でも response header を返す.
    assert res.headers.get("X-Excel-Crossoffice-Warnings-Count") == "0"


@pytest.mark.asyncio
async def test_export_n_patients_writes_n_plus_1_rows(client, db) -> None:
    admin = await _make_user(db, "ex-n@example.com", "admin")
    for code in ("P-EX-001", "P-EX-002", "P-EX-003"):
        await _make_patient(db, code=code, name=f"名前{code}")

    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PATIENTS]
    assert ws.max_row == 4  # 1 header + 3 patients
    codes = [ws.cell(row=r, column=PATIENT_COL_INDEX["patient_code"] + 1).value for r in (2, 3, 4)]
    assert set(codes) == {"P-EX-001", "P-EX-002", "P-EX-003"}


# 2) export: 「編集用」シートが患者 1 行で patient_code / 曜日時刻にリンクされている
@pytest.mark.asyncio
async def test_export_pfv_links_to_patient_code(client, db) -> None:
    admin = await _make_user(db, "ex-pfv@example.com", "admin")
    patient = await _make_patient(db, code="P-PFV-LINK", name="リンク患者")
    await _make_pfv(db, patient_id=patient.id, weekday=0, start_time=time(9, 0))

    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    # Phase G-51: PFV は「編集用」(患者 1 行) + 静的グリッドの 2 シート構成.
    assert SHEET_PFV_EDIT in wb.sheetnames
    assert SHEET_PFV_GRID in wb.sheetnames
    ws_f = wb[SHEET_PFV_EDIT]
    assert ws_f.max_row == 2  # 1 header + 1 患者行
    code_cell = ws_f.cell(row=2, column=PFV_EDIT_COL_INDEX["patient_code"] + 1).value
    name_cell = ws_f.cell(row=2, column=PFV_EDIT_COL_INDEX["patient_name"] + 1).value
    mon_time = ws_f.cell(row=2, column=PFV_EDIT_COL_INDEX["mon_time"] + 1).value
    assert code_cell == "P-PFV-LINK"
    assert name_cell == "リンク患者"
    assert mon_time == "09:00"


# 3) dry_run: 新規行検出
@pytest.mark.asyncio
async def test_import_dry_run_detects_new_patient(client, db) -> None:
    admin = await _make_user(db, "im-new@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-NEW-001",
                "name": "新規 太郎",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区xxx",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is False
    assert body["summary"]["patients_new"] == 1
    assert body["summary"]["patients_error"] == 0
    row = body["patient_rows"][0]
    assert row["operation"] == "new"
    assert row["patient_code"] == "P-NEW-001"


# 4) dry_run: 更新行で changes が diff される
@pytest.mark.asyncio
async def test_import_dry_run_diffs_updates(client, db) -> None:
    admin = await _make_user(db, "im-upd@example.com", "admin")
    patient = await _make_patient(
        db, code="P-UPD-001", name="旧名前", address="旧住所", sex="male", status="active"
    )
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient.id),
                "patient_code": "P-UPD-001",
                "name": "新名前",
                "sex": "male",
                "status": "active",
                "address": "新住所",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_update"] == 1
    row = body["patient_rows"][0]
    assert row["operation"] == "update"
    fields = {c["field"]: c for c in row["changes"]}
    assert "name" in fields
    assert fields["name"]["old_value"] == "旧名前"
    assert fields["name"]["new_value"] == "新名前"
    assert "address" in fields


# 5) dry_run: <DELETE> フラグ
@pytest.mark.asyncio
async def test_import_dry_run_detects_delete_flag(client, db) -> None:
    admin = await _make_user(db, "im-del@example.com", "admin")
    patient = await _make_patient(db, code="P-DEL-001", name="削除予定")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient.id),
                "patient_code": "P-DEL-001",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_delete"] == 1
    assert body["patient_rows"][0]["operation"] == "delete"


# 6) dry_run: <CLEAR> で NULL 上書き
@pytest.mark.asyncio
async def test_import_dry_run_clear_to_null(client, db) -> None:
    admin = await _make_user(db, "im-clr@example.com", "admin")
    patient = await _make_patient(db, code="P-CLR-001", name="クリア対象", note="残ってる備考")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient.id),
                "patient_code": "P-CLR-001",
                "name": "クリア対象",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区test1",
                "note": "<CLEAR>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    row = body["patient_rows"][0]
    assert row["operation"] == "update"
    notes = [c for c in row["changes"] if c["field"] == "note"]
    assert len(notes) == 1
    assert notes[0]["new_value"] is None
    assert notes[0]["old_value"] == "残ってる備考"


# 7) dry_run: 空セルは noop
@pytest.mark.asyncio
async def test_import_dry_run_blank_cells_are_noop(client, db) -> None:
    admin = await _make_user(db, "im-noop@example.com", "admin")
    patient = await _make_patient(
        db,
        code="P-NOOP-001",
        name="変更なし",
        sex="male",
        status="active",
        address="千葉市稲毛区test1",
        kana="ヘンコウナシ",
    )
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient.id),
                "patient_code": "P-NOOP-001",
                # 他全部空 → 「触らない」
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    row = body["patient_rows"][0]
    # 既存値と一致 (= 触らない) なので noop. patient_code は明示指定だが既存値と同じ.
    assert row["operation"] == "noop"
    assert body["summary"]["patients_noop"] == 1


# 8) dry_run: バリデーションエラー多種
@pytest.mark.asyncio
async def test_import_dry_run_validation_errors(client, db) -> None:
    admin = await _make_user(db, "im-err@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            # 行 2: UUID 不正
            {
                "patient_id": "not-a-uuid",
                "patient_code": "P-BAD-1",
                "name": "uuid不正",
            },
            # 行 3: enum 違反
            {
                "patient_code": "P-BAD-2",
                "name": "enum違反",
                "sex": "INVALID",
                "status": "active",
                "address": "addr",
            },
            # 行 4: 必須欠落 (name)
            {
                "patient_code": "P-BAD-3",
                "sex": "male",
                "status": "active",
                "address": "addr",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_error"] == 3
    ops = {r["row_number"]: r for r in body["patient_rows"]}
    assert ops[2]["operation"] == "error"
    assert ops[3]["operation"] == "error"
    assert ops[4]["operation"] == "error"
    assert "UUID" in ops[2]["error_message"] or "uuid" in ops[2]["error_message"].lower()
    assert "sex" in ops[3]["error_message"].lower() or "性別" in ops[3]["error_message"]
    assert (
        "name" in ops[4]["error_message"].lower()
        or "患者名" in ops[4]["error_message"]
        or "新規作成" in ops[4]["error_message"]
    )


# 9) dry_run: DB は変更されない
@pytest.mark.asyncio
async def test_import_dry_run_does_not_modify_db(client, db) -> None:
    admin = await _make_user(db, "im-noch@example.com", "admin")
    before_count = len((await db.scalars(select(Patient))).all())

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-DRY-001",
                "name": "新規",
                "sex": "male",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200
    assert res.json()["transaction_applied"] is False

    # 再 SELECT
    db.expire_all()
    after_count = len((await db.scalars(select(Patient))).all())
    assert after_count == before_count


# 10) apply: 新規 + 更新 + 削除 が 1 TX で反映
@pytest.mark.asyncio
async def test_import_apply_mixed_operations(client, db) -> None:
    admin = await _make_user(db, "im-apply@example.com", "admin")
    p_upd = await _make_patient(db, code="P-AP-UPD", name="旧", sex="male", status="active")
    p_del = await _make_patient(db, code="P-AP-DEL", name="消す", sex="male", status="active")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-AP-NEW",
                "name": "新規",
                "sex": "female",
                "status": "active",
                "address": "addr-new",
            },
            {
                "patient_id": str(p_upd.id),
                "patient_code": "P-AP-UPD",
                "name": "新",
                "sex": "male",
                "status": "active",
                "address": "addr-new",
            },
            {
                "patient_id": str(p_del.id),
                "patient_code": "P-AP-DEL",
                "delete_flag": "<DELETE>",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"] == {
        "patients_new": 1,
        "patients_update": 1,
        "patients_delete": 1,
        "patients_error": 0,
        "patients_noop": 0,
        "pfv_new": 0,
        "pfv_update": 0,
        "pfv_delete": 0,
        "pfv_error": 0,
        "pfv_noop": 0,
    }

    # 確認 (API 側で commit 済みなので、テスト用 session の identity map を expire してから再取得)
    db.expire_all()
    rows = (
        await db.scalars(
            select(Patient).where(Patient.code.in_(["P-AP-NEW", "P-AP-UPD", "P-AP-DEL"]))
        )
    ).all()
    by_code = {p.code: p for p in rows}
    assert by_code["P-AP-NEW"].deleted_at is None
    assert by_code["P-AP-UPD"].name == "新"
    assert by_code["P-AP-DEL"].deleted_at is not None  # soft delete


# 11) apply: partial commit — error 行は skip、有効行のみ反映
@pytest.mark.asyncio
async def test_import_apply_partial_commit_skips_error_rows(client, db) -> None:
    """error 1 件 + 正常 update 1 件 → 正常 update は commit、error 行は skip."""
    admin = await _make_user(db, "im-rb@example.com", "admin")
    p_upd = await _make_patient(db, code="P-RB-UPD", name="旧")
    p_upd_id = p_upd.id  # expire_all 前に id を確保 (lazy-load 回避)

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(p_upd_id),
                "patient_code": "P-RB-UPD",
                "name": "正常な更新",
                "sex": "male",
                "status": "active",
                "address": "addr",
            },
            {
                "patient_code": "P-RB-ERR",
                "name": "エラー行",
                "sex": "INVALID",
                "status": "active",
                "address": "addr",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    # 有効な op (update 1) があるので transaction は commit される.
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_error"] == 1
    assert body["summary"]["patients_update"] == 1

    # P-RB-UPD は更新されているはず (partial commit).
    db.expire_all()
    refreshed = await db.get(Patient, p_upd_id)
    assert refreshed.name == "正常な更新"
    # P-RB-ERR は skip されているので DB に存在しない.
    err_row = await db.scalar(select(Patient).where(Patient.code == "P-RB-ERR"))
    assert err_row is None


# 12) apply: PFV の物理削除
@pytest.mark.asyncio
async def test_import_apply_pfv_hard_delete(client, db) -> None:
    admin = await _make_user(db, "im-pfvdel@example.com", "admin")
    patient = await _make_patient(db, code="P-PFV-DEL", name="PFV削除")
    pfv = await _make_pfv(db, patient_id=patient.id, weekday=0)
    pfv_id = pfv.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(patient.id),
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_delete"] == 1

    db.expire_all()
    refreshed = await db.get(PatientFixedVisit, pfv_id)
    assert refreshed is None  # 物理削除


# 13) apply: course_template_code 解決
@pytest.mark.asyncio
async def test_import_apply_resolves_course_template(client, db) -> None:
    admin = await _make_user(db, "im-ct@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    template = CourseTemplate(office_id=office.id, label="A")
    db.add(template)
    await db.commit()
    await db.refresh(template)

    patient = await _make_patient(db, code="P-CT-001", name="CT患者", primary_office_id=office.id)
    patient_id = patient.id
    template_id = template.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "火",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "10:00",
                "course_template_code": "A",
                "duration_min": 45,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    assert res.status_code == 200, body
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1

    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1
    assert rows[0].course_template_id == template_id


# 14) apply: 拠点コード解決
@pytest.mark.asyncio
async def test_import_apply_resolves_office_code(client, db) -> None:
    admin = await _make_user(db, "im-off@example.com", "admin")
    office = await _make_office(db, code="TSUGA", name="都賀")
    office_id = office.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-OFF-001",
                "name": "拠点指定",
                "sex": "male",
                "status": "active",
                "address": "千葉市若葉区xxx",
                "office_code": "TSUGA",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True

    db.expire_all()
    rows = (await db.scalars(select(Patient).where(Patient.code == "P-OFF-001"))).all()
    assert len(rows) == 1
    assert rows[0].primary_office_id == office_id


# 15) Phase G-48: patient_id 空 + 既存 code → 新規 error ではなく UPDATE 突合.
# ユーザーが export (code 入り) を直接編集して再アップするだけで更新できる.
@pytest.mark.asyncio
async def test_import_existing_code_without_id_updates(client, db) -> None:
    admin = await _make_user(db, "im-dup@example.com", "admin")
    p = await _make_patient(db, code="P-DUP-001", name="既存", sex="male", status="active")
    pid = p.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                # patient_id 空 + 既存 code → その患者を UPDATE する (旧来は error).
                "patient_code": "P-DUP-001",
                "name": "更新後の名前",
                "sex": "female",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_error"] == 0
    assert body["summary"]["patients_update"] == 1
    assert body["summary"]["patients_new"] == 0
    row = body["patient_rows"][0]
    assert row["operation"] == "update"
    assert row["patient_id"] == str(pid)

    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.name == "更新後の名前"
    assert p_after.sex == "female"
    # 新規行は作られていない (同一 code が 1 件だけ).
    rows = (await db.scalars(select(Patient).where(Patient.code == "P-DUP-001"))).all()
    assert len(rows) == 1


# 15b) 同一ファイル内 patient_code 重複は従来どおり error.
@pytest.mark.asyncio
async def test_import_in_file_duplicate_code_errors(client, db) -> None:
    admin = await _make_user(db, "im-dup-infile@example.com", "admin")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-INFILE-DUP",
                "name": "1行目",
                "sex": "male",
                "status": "active",
                "address": "addr",
            },
            {
                "patient_code": "P-INFILE-DUP",  # 同ファイル内重複!
                "name": "2行目",
                "sex": "female",
                "status": "active",
                "address": "addr",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    # 1 件目は new, 2 件目は同ファイル内重複で error.
    assert body["summary"]["patients_error"] == 1
    assert body["summary"]["patients_new"] == 1
    err_row = next(r for r in body["patient_rows"] if r["operation"] == "error")
    assert "P-INFILE-DUP" in err_row["error_message"]


# 16) RBAC: staff は 403
@pytest.mark.asyncio
async def test_import_export_requires_admin_or_manager(client, db) -> None:
    staff_user = await _make_user(db, "im-staff@example.com", "staff")
    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(staff_user),
    )
    assert res.status_code == 403

    res = await client.get(
        "/api/v1/patients/import-export/template",
        headers=_bearer(staff_user),
    )
    assert res.status_code == 403

    # POST /import も同様に staff は 403 にならねばならない (危険 endpoint).
    content = _build_workbook_bytes(patient_rows=[])
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(staff_user),
        files=files,
    )
    assert res.status_code == 403


# 16b) RBAC: manager は import endpoint に到達できる (200).
@pytest.mark.asyncio
async def test_import_allows_manager_role(client, db) -> None:
    manager = await _make_user(db, "im-mgr@example.com", "manager")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-MGR-OK",
                "name": "manager OK",
                "sex": "male",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, manager, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_new"] == 1


# 17) template ダウンロード: 1 行 (ヘッダーのみ)
@pytest.mark.asyncio
async def test_template_download_has_header_only(client, db) -> None:
    admin = await _make_user(db, "im-tpl@example.com", "admin")
    res = await client.get(
        "/api/v1/patients/import-export/template",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert SHEET_PATIENTS in wb.sheetnames
    assert SHEET_PFV_EDIT in wb.sheetnames
    assert SHEET_PFV_GRID in wb.sheetnames
    assert wb[SHEET_PATIENTS].max_row == 1
    # 編集用シートはヘッダーのみ (患者行 0).
    assert wb[SHEET_PFV_EDIT].max_row == 1


# 18) magic word: case-insensitive
@pytest.mark.asyncio
async def test_magic_word_case_insensitive(client, db) -> None:
    """<delete> (lower-case) も認識される."""
    admin = await _make_user(db, "im-magic@example.com", "admin")
    patient = await _make_patient(db, code="P-MAGIC-001", name="ケース不問")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient.id),
                "patient_code": "P-MAGIC-001",
                "delete_flag": "<delete>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_delete"] == 1
    assert body["patient_rows"][0]["operation"] == "delete"


# 19) PFV update via key (patient_id, mode, weekday, slot_index)
@pytest.mark.asyncio
async def test_pfv_update_via_composite_key(client, db) -> None:
    admin = await _make_user(db, "im-pfvup@example.com", "admin")
    patient = await _make_patient(db, code="P-PFV-UP", name="PFV更新")
    pfv = await _make_pfv(
        db, patient_id=patient.id, weekday=2, slot_index=0, start_time=time(9, 0), duration_min=30
    )
    patient_id = patient.id
    pfv_id = pfv.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "水",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "10:30",
                "duration_min": 60,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_update"] == 1

    db.expire_all()
    refreshed = await db.get(PatientFixedVisit, pfv_id)
    assert refreshed.start_time == time(10, 30)
    assert refreshed.duration_min == 60


# 19b) regression: 患者を <DELETE> すると関連 PFV も物理削除される (orphan 防止)
@pytest.mark.asyncio
async def test_patient_delete_cascades_pfv_cleanup(client, db) -> None:
    admin = await _make_user(db, "im-delcasc@example.com", "admin")
    patient = await _make_patient(db, code="P-DELCASC", name="orphan source")
    patient_id = patient.id
    pfv = await _make_pfv(db, patient_id=patient_id, weekday=0)
    pfv_id = pfv.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(patient_id),
                "patient_code": "P-DELCASC",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_delete"] == 1

    db.expire_all()
    patient_after = await db.get(Patient, patient_id)
    assert patient_after is not None
    assert patient_after.deleted_at is not None  # soft delete
    # 関連 PFV は物理削除されているはず.
    pfv_after = await db.get(PatientFixedVisit, pfv_id)
    assert pfv_after is None
    remaining = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert remaining == []


# 20) patient_id 不正 (DB に無い UUID)
@pytest.mark.asyncio
async def test_unknown_patient_id_errors(client, db) -> None:
    admin = await _make_user(db, "im-unk@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(uuid4()),  # 存在しない
                "patient_code": "P-UNK-001",
                "name": "存在しない",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_error"] == 1
    assert "存在" in body["patient_rows"][0]["error_message"]


# 21) PFV patient_code リンク (新規患者 + その PFV を 1 回の import で登録)
@pytest.mark.asyncio
async def test_pfv_links_to_new_patient_via_code(client, db) -> None:
    """新規患者 (patient_id 空) + 同 import 内の PFV を patient_code でリンクできる."""
    admin = await _make_user(db, "im-pfvlink-new@example.com", "admin")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-LINK-NEW",
                "name": "リンク新規",
                "sex": "male",
                "status": "active",
                "address": "千葉市稲毛区xxx",
            }
        ],
        pfv_rows=[
            {
                # patient_id 空 + patient_code で参照
                "patient_code": "P-LINK-NEW",
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
                "duration_min": 30,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_new"] == 1
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["patients_error"] == 0
    assert body["summary"]["pfv_error"] == 0

    # DB に新規患者 + 紐付いた PFV が 1 件あること
    db.expire_all()
    patient_rows = (await db.scalars(select(Patient).where(Patient.code == "P-LINK-NEW"))).all()
    assert len(patient_rows) == 1
    patient = patient_rows[0]
    pfv_rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient.id)
        )
    ).all()
    assert len(pfv_rows) == 1
    assert pfv_rows[0].weekday == 0
    assert pfv_rows[0].duration_min == 30


# 22) PFV patient_code が import 内にも DB にも無い → error
@pytest.mark.asyncio
async def test_pfv_code_lookup_unknown_errors(client, db) -> None:
    admin = await _make_user(db, "im-pfvlink-miss@example.com", "admin")
    content = _build_workbook_bytes(
        pfv_rows=[
            {
                # patient_id 空 + 同 import 内にも DB にも無い code
                "patient_code": "P-MISS-XYZ",
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["pfv_error"] == 1
    row = body["pfv_rows"][0]
    assert row["operation"] == "error"
    assert "P-MISS-XYZ" in row["error_message"]


# 23) 既存患者の PFV を patient_id 空 + patient_code で参照しても OK
@pytest.mark.asyncio
async def test_pfv_code_lookup_existing_patient_ok(client, db) -> None:
    admin = await _make_user(db, "im-pfvlink-exist@example.com", "admin")
    patient = await _make_patient(db, code="P-LINK-EXIST", name="既存患者")
    patient_id = patient.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                # patient_id 空 + 既存 code で参照
                "patient_code": "P-LINK-EXIST",
                "weekday": "火",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "10:00",
                "duration_min": 45,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1
    assert rows[0].weekday == 1


# 24) PFV の patient_id 空 + patient_code も空 → error
@pytest.mark.asyncio
async def test_pfv_both_id_and_code_blank_errors(client, db) -> None:
    admin = await _make_user(db, "im-pfvlink-blank@example.com", "admin")
    content = _build_workbook_bytes(
        pfv_rows=[
            {
                # patient_id も patient_code も空
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["pfv_error"] == 1
    row = body["pfv_rows"][0]
    assert row["operation"] == "error"
    msg = row["error_message"]
    assert "patient_id" in msg and "patient_code" in msg and "必須" in msg


# 25) HIGH#1 regression: PFV で patient_id と patient_code が異なる患者を指している場合 error
@pytest.mark.asyncio
async def test_pfv_id_and_code_mismatch_errors(client, db) -> None:
    """両者が異なる患者を指していても silent に id を優先して code を無視するのは
    データ破壊リスクなので明示的に error にする."""
    admin = await _make_user(db, "im-pfv-mismatch@example.com", "admin")
    patient_a = await _make_patient(db, code="P-MIS-A", name="患者A")
    patient_b = await _make_patient(db, code="P-MIS-B", name="患者B")
    patient_a_id = patient_a.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                # id=A, code=B (異なる患者) → error
                "patient_id": str(patient_a_id),
                "patient_code": patient_b.code,
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["pfv_error"] == 1
    row = body["pfv_rows"][0]
    assert row["operation"] == "error"
    msg = row["error_message"]
    assert "patient_id" in msg and "patient_code" in msg
    assert "P-MIS-B" in msg


# 26) HIGH#1 regression: PFV で patient_id と patient_code が一致していれば正常処理
@pytest.mark.asyncio
async def test_pfv_id_and_code_consistent_ok(client, db) -> None:
    admin = await _make_user(db, "im-pfv-consistent@example.com", "admin")
    patient = await _make_patient(db, code="P-CONS-OK", name="一致患者")
    patient_id = patient.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                # id=A, code=A (同じ患者) → 正常
                "patient_id": str(patient_id),
                "patient_code": "P-CONS-OK",
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1


# 27) partial commit: error 1 件 + new 2 件 → new 2 件のみ DB に反映
@pytest.mark.asyncio
async def test_import_apply_partial_commit_new_rows_with_one_error(client, db) -> None:
    """error 1 件 + new 2 件 で apply → DB に new 2 件のみ反映、error 1 件は skip."""
    admin = await _make_user(db, "im-partial-mix@example.com", "admin")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-PC-OK-1",
                "name": "正常 1",
                "sex": "male",
                "status": "active",
                "address": "addr1",
            },
            {
                "patient_code": "P-PC-ERR",
                "name": "エラー (enum 違反)",
                "sex": "INVALID",
                "status": "active",
                "address": "addr-err",
            },
            {
                "patient_code": "P-PC-OK-2",
                "name": "正常 2",
                "sex": "female",
                "status": "active",
                "address": "addr2",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_new"] == 2
    assert body["summary"]["patients_error"] == 1

    # DB: 正常 2 件のみ反映 (error 行は skip)
    db.expire_all()
    rows = (
        await db.scalars(
            select(Patient).where(Patient.code.in_(["P-PC-OK-1", "P-PC-ERR", "P-PC-OK-2"]))
        )
    ).all()
    codes = sorted(p.code for p in rows)
    assert codes == ["P-PC-OK-1", "P-PC-OK-2"]


# 28) partial commit: 全件 error → transaction_applied=False、DB 変更なし
@pytest.mark.asyncio
async def test_import_apply_all_errors_no_commit(client, db) -> None:
    admin = await _make_user(db, "im-partial-allerr@example.com", "admin")
    before_count = len((await db.scalars(select(Patient))).all())

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-ALL-ERR-1",
                "name": "エラー 1",
                "sex": "INVALID",
                "status": "active",
                "address": "addr",
            },
            {
                "patient_code": "P-ALL-ERR-2",
                "name": "エラー 2",
                "sex": "male",
                "status": "INVALID_STATUS",
                "address": "addr",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    # 有効 op = 0 件 → 非 commit.
    assert body["transaction_applied"] is False
    assert body["summary"]["patients_error"] == 2

    # DB は変更なし.
    db.expire_all()
    after_count = len((await db.scalars(select(Patient))).all())
    assert after_count == before_count


# 29) partial commit: pfv_error 1 件 + patient_new 2 件 → patient 反映、pfv は正常分のみ反映
@pytest.mark.asyncio
async def test_import_apply_partial_commit_mixed_patient_and_pfv_errors(client, db) -> None:
    """patient シートに正常 new 2 件、pfv シートに error 1 件 + 正常 new 1 件.
    → patient 2 件 + pfv 1 件 が commit される (pfv error 1 件は skip)."""
    admin = await _make_user(db, "im-partial-pfv@example.com", "admin")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-MIX-A",
                "name": "患者 A",
                "sex": "male",
                "status": "active",
                "address": "addr-a",
            },
            {
                "patient_code": "P-MIX-B",
                "name": "患者 B",
                "sex": "female",
                "status": "active",
                "address": "addr-b",
            },
        ],
        pfv_rows=[
            {
                # 正常 pfv 行 (患者 A 紐付け via code)
                "patient_code": "P-MIX-A",
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
                "duration_min": 30,
            },
            {
                # error 行: weekday の値が候補外
                "patient_code": "P-MIX-B",
                "weekday": "INVALID_DAY",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "10:00",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_new"] == 2
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 1

    db.expire_all()
    # 患者 2 件はどちらも DB にいる.
    patient_rows = (
        await db.scalars(select(Patient).where(Patient.code.in_(["P-MIX-A", "P-MIX-B"])))
    ).all()
    assert {p.code for p in patient_rows} == {"P-MIX-A", "P-MIX-B"}
    by_code = {p.code: p for p in patient_rows}
    # 患者 A の PFV のみ DB に存在 (患者 B の error 行は skip).
    pfvs_a = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == by_code["P-MIX-A"].id)
        )
    ).all()
    assert len(pfvs_a) == 1
    pfvs_b = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == by_code["P-MIX-B"].id)
        )
    ).all()
    assert pfvs_b == []


# 30) regression (bug 1): 新規 patient INSERT 時に special_week_active が NULL ではなく
#     [] になり、GET /api/v1/patients の Pydantic レスポンス schema を通る.
@pytest.mark.asyncio
async def test_import_apply_new_patient_special_week_active_defaults_to_empty_list(
    client, db
) -> None:
    """Excel import で新規追加した patient の special_week_active が NULL のままだと
    /api/v1/patients が ``Input should be a valid list`` で 500 になる回帰を防ぐ.

    importer は通常の POST /patients ルートを通らず ORM INSERT を直接行うため、
    Pydantic の default_factory に頼れない. importer 側で必ず [] を埋める必要がある.
    """
    admin = await _make_user(db, "im-swa@example.com", "admin")
    # bearer headers は expire_all 前に capture しておく (expire_all 後に
    # admin.id を触ると async greenlet 越しの lazy load が走って fail する).
    headers = _bearer(admin)
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-SWA-001",
                "name": "swa新規",
                "sex": "male",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert res.json()["transaction_applied"] is True

    # DB から直接取得して default を確認.
    db.expire_all()
    patient = await db.scalar(select(Patient).where(Patient.code == "P-SWA-001"))
    assert patient is not None
    assert patient.special_week_active == []  # NULL ではなく []

    # API ラウンドトリップ (回帰の本丸 — Pydantic validation を通ること)
    res = await client.get("/api/v1/patients", headers=headers)
    assert res.status_code == 200, res.text
    items = res.json()
    by_code = {p["code"]: p for p in items}
    assert "P-SWA-001" in by_code
    assert by_code["P-SWA-001"]["special_week_active"] == []


# 31) bug 2: patient_id 空 + patient_code = 既存患者 + <DELETE> で削除成功
@pytest.mark.asyncio
async def test_import_apply_delete_via_patient_code_only(client, db) -> None:
    """ユーザーが手で「削除したい code + <DELETE>」だけ書いた Excel でも削除できる."""
    admin = await _make_user(db, "im-delcode@example.com", "admin")
    patient = await _make_patient(db, code="P-DELCODE-001", name="code削除")
    patient_id = patient.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                # patient_id 空 + patient_code のみ + <DELETE>
                "patient_code": "P-DELCODE-001",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["patients_delete"] == 1
    row = body["patient_rows"][0]
    assert row["operation"] == "delete"
    assert row["patient_code"] == "P-DELCODE-001"

    db.expire_all()
    refreshed = await db.get(Patient, patient_id)
    assert refreshed is not None
    assert refreshed.deleted_at is not None  # soft delete されている


# 32) bug 2: patient_id 空 + 存在しない patient_code + <DELETE> は idempotent noop
@pytest.mark.asyncio
async def test_import_delete_via_unknown_patient_code_is_noop(client, db) -> None:
    """既に削除済み (or そもそも居ない) code を <DELETE> しても error にせず noop."""
    admin = await _make_user(db, "im-delnoop@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-DELNOOP-XYZ",
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_noop"] == 1
    assert body["summary"]["patients_error"] == 0
    assert body["patient_rows"][0]["operation"] == "noop"


# 33) bug 2: patient_id 空 + patient_code 空 + <DELETE> は error
@pytest.mark.asyncio
async def test_import_delete_with_both_id_and_code_blank_errors(client, db) -> None:
    admin = await _make_user(db, "im-delblank@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                # 両方空
                "delete_flag": "<DELETE>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    body = res.json()
    assert body["summary"]["patients_error"] == 1
    msg = body["patient_rows"][0]["error_message"]
    assert "patient_id" in msg and "patient_code" in msg


# 34) resurrection: soft-deleted patient_code を Excel で再 import すると復活する
@pytest.mark.asyncio
async def test_resurrection_via_code_for_soft_deleted_patient(client, db) -> None:
    """ユーザー運用フロー: UI で削除 (soft delete) → 同じ code を Excel で再 import.

    soft-deleted 行があるまま INSERT すると UNIQUE 制約違反で 409 になるため、
    importer は INSERT ではなく既存行の deleted_at=NULL + 内容更新で復活させる.
    """
    admin = await _make_user(db, "im-resurrect@example.com", "admin")
    patient = await _make_patient(
        db,
        code="P-RES-001",
        name="削除前の名前",
        sex="male",
        status="active",
        address="千葉市稲毛区old",
    )
    # 直接 soft delete (UI からの削除を再現)
    patient.deleted_at = datetime.now(UTC)
    await db.commit()

    content = _build_workbook_bytes(
        patient_rows=[
            {
                # patient_id 空 + 同じ code で再 import (新規行のつもりで書く)
                "patient_code": "P-RES-001",
                "name": "復活後の名前",
                "sex": "female",
                "status": "active",
                "address": "千葉市稲毛区new",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    # 設計判断: 復活も UI 上は update として表示 (operation="update").
    assert body["summary"]["patients_update"] == 1
    assert body["summary"]["patients_new"] == 0
    assert body["summary"]["patients_error"] == 0
    row = body["patient_rows"][0]
    assert row["operation"] == "update"
    assert row["patient_code"] == "P-RES-001"
    # changes に deleted_at: <old> → null が出ている (= 復活の証跡).
    fields = {c["field"]: c for c in row["changes"]}
    assert "deleted_at" in fields, f"changes に deleted_at が無い: {row['changes']}"
    assert fields["deleted_at"]["new_value"] is None
    assert fields["deleted_at"]["old_value"] is not None
    # 内容更新も diff されている.
    assert "name" in fields and fields["name"]["new_value"] == "復活後の名前"
    assert "sex" in fields and fields["sex"]["new_value"] == "female"
    assert "address" in fields and fields["address"]["new_value"] == "千葉市稲毛区new"

    # DB: deleted_at=NULL に戻り、内容も更新されている.
    await db.refresh(patient)
    assert patient.deleted_at is None
    assert patient.name == "復活後の名前"
    assert patient.sex == "female"
    assert patient.address == "千葉市稲毛区new"


# 35) resurrection: 復活時に patient の id が再発番されないこと (UUID 継続性)
@pytest.mark.asyncio
async def test_resurrection_preserves_id(client, db) -> None:
    """復活時に id は元のまま (新しい UUID は発番されない).

    これにより audit log 等の外部参照が断絶しない.
    """
    admin = await _make_user(db, "im-resurrect-id@example.com", "admin")
    patient = await _make_patient(db, code="P-RES-ID", name="UUID 継続テスト")
    original_id = patient.id
    patient.deleted_at = datetime.now(UTC)
    await db.commit()

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-RES-ID",
                "name": "復活した",
                "sex": "male",
                "status": "active",
                "address": "addr",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    row = body["patient_rows"][0]
    assert row["operation"] == "update"
    # 返却された patient_id が元の id と同じ.
    assert row["patient_id"] == str(original_id)

    # DB 確認: 元の id がそのまま生きている (= 同じ行が復活した).
    await db.refresh(patient)
    assert patient.id == original_id
    assert patient.deleted_at is None
    assert patient.code == "P-RES-ID"
    # 同じ code を持つ他レコードが新規作成されていないこと.
    rows = (await db.scalars(select(Patient).where(Patient.code == "P-RES-ID"))).all()
    assert len(rows) == 1


# ---------------------------------------------------------------------------
# E-4: バックアップ運用 (export → そのまま import) で 0 エラーを担保
# ---------------------------------------------------------------------------


# E-4-1) Phase G-51: export 時に patient.weekly_pattern が無くても「編集用」シートの
# time_type が default ("時間帯") で書き出され、import 側で空欄エラーにならない.
@pytest.mark.asyncio
async def test_patient_export_fills_default_time_type_when_empty(client, db) -> None:
    admin = await _make_user(db, "e4-export-tt@example.com", "admin")
    patient = await _make_patient(db, code="P-E4-TT-1", name="time_type欠落")
    # weekly_pattern は明示的に None (該当エントリ無し).
    patient.weekly_pattern = None
    await db.commit()
    await _make_pfv(db, patient_id=patient.id, weekday=0)

    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws_f = wb[SHEET_PFV_EDIT]
    # 患者行 (row 2) の time_type セルが空ではなく default "時間帯".
    tt_value = ws_f.cell(row=2, column=PFV_EDIT_COL_INDEX["time_type"] + 1).value
    assert tt_value == "時間帯"


# E-4-2) Phase G-51: クロス拠点 course (患者 primary≠template office) は
# **拠点付きトークン** (例 "津B") で書き出される (1 行形式でクロス拠点を表現可能).
@pytest.mark.asyncio
async def test_patient_export_omits_cross_office_course_template_code(client, db) -> None:
    admin = await _make_user(db, "e4-export-ct@example.com", "admin")
    office_a = await _make_office(db, code="INAGE", name="稲毛")
    office_b = await _make_office(db, code="TSUGA", name="都賀")
    # 患者の拠点 = INAGE
    patient = await _make_patient(
        db,
        code="P-E4-CT-1",
        name="cross-office",
        primary_office_id=office_a.id,
    )
    # PFV は TSUGA 側の template を指す (クロス拠点 = 月=津B).
    template_b = CourseTemplate(office_id=office_b.id, label="B")
    db.add(template_b)
    await db.commit()
    await db.refresh(template_b)
    await _make_pfv(db, patient_id=patient.id, weekday=0, course_template_id=template_b.id)

    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws_f = wb[SHEET_PFV_EDIT]
    course_value = ws_f.cell(row=2, column=PFV_EDIT_COL_INDEX["mon_course"] + 1).value
    # 拠点付きトークン (TSUGA→津) で書き出される.
    assert course_value == "津B"


# E-4-3) Phase G-51: 同一拠点 course は拠点付きトークン (例 "稲M") で書き出される.
@pytest.mark.asyncio
async def test_patient_export_writes_course_template_code_for_same_office(client, db) -> None:
    admin = await _make_user(db, "e4-export-ct-ok@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    template = CourseTemplate(office_id=office.id, label="M")
    db.add(template)
    await db.commit()
    await db.refresh(template)
    patient = await _make_patient(
        db,
        code="P-E4-CT-OK",
        name="same-office",
        primary_office_id=office.id,
    )
    await _make_pfv(db, patient_id=patient.id, weekday=0, course_template_id=template.id)

    res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws_f = wb[SHEET_PFV_EDIT]
    course_value = ws_f.cell(row=2, column=PFV_EDIT_COL_INDEX["mon_course"] + 1).value
    assert course_value == "稲M"


# E-4-4) import 時に time_type が空セルでも error にならず、default で吸収される.
@pytest.mark.asyncio
async def test_patient_import_accepts_empty_time_type_with_default(client, db) -> None:
    admin = await _make_user(db, "e4-import-tt@example.com", "admin")
    patient = await _make_patient(db, code="P-E4-TT-IN", name="time_type空でも OK")
    patient_id = patient.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                # time_type を意図的に省略.
                "start_time": "09:00",
                "duration_min": 30,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 0
    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1


# E-4-5) import 時に course_template_code が患者拠点に存在しない場合でも
# error にならず、course_template_id=None で PFV を保存する.
@pytest.mark.asyncio
async def test_patient_import_fallback_for_unknown_course_template_code(client, db) -> None:
    admin = await _make_user(db, "e4-import-ct@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    # 患者拠点 = INAGE. ただし INAGE には 'D' template は存在しない.
    patient = await _make_patient(
        db,
        code="P-E4-CTU",
        name="unknown ct",
        primary_office_id=office.id,
    )
    patient_id = patient.id

    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(patient_id),
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
                "duration_min": 30,
                "course_template_code": "D",  # 拠点に存在しない
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(
            select(PatientFixedVisit).where(PatientFixedVisit.patient_id == patient_id)
        )
    ).all()
    assert len(rows) == 1
    # course_template_id は剥がされて None で保存.
    assert rows[0].course_template_id is None


# E-4-6) round-trip: export → そのまま import で 0 エラーを担保.
# 複数患者 + 複数 PFV (うち 1 件は weekly_pattern エントリ無し、
# 1 件はクロスオフィス template 参照) を入れた状態でも、
# export してから何も編集せず import すると error が 0 件になる.
#
# 注: 現状の round-trip 挙動は「データロス vs UX」のトレードオフで
# **意図的** に silent update を許容している:
#   - PFV-A (same-office template, weekly_pattern 未設定) → noop
#     (time_type は表示専用で PFV テーブルに保存先が無いため変更検出されない)
#   - PFV-B (cross-office template) → silent update
#     (export 時に course_template_code が空欄に倒され、import 時に
#      course_template_id=None で再解決されるため、既存 template_b.id との
#      差分として update が記録される — データロスではあるが運用上許容)
#
# クリーンデータでの完全 noop round-trip は
# ``test_patient_export_then_import_clean_data_full_noop_round_trip`` を参照.
@pytest.mark.asyncio
async def test_patient_export_then_import_round_trip_completes_without_errors(client, db) -> None:
    admin = await _make_user(db, "e4-roundtrip@example.com", "admin")
    office_a = await _make_office(db, code="INAGE", name="稲毛")
    office_b = await _make_office(db, code="TSUGA", name="都賀")
    template_a = CourseTemplate(office_id=office_a.id, label="M")
    template_b = CourseTemplate(office_id=office_b.id, label="C")  # 患者 A の拠点に C は無い
    db.add_all([template_a, template_b])
    await db.commit()
    await db.refresh(template_a)
    await db.refresh(template_b)

    patient_a = await _make_patient(db, code="P-RT-A", name="患者A", primary_office_id=office_a.id)
    patient_b = await _make_patient(db, code="P-RT-B", name="患者B", primary_office_id=office_a.id)
    # 1) PFV-A: weekly_pattern エントリ無し (time_type 解決不能だが PFV テーブルには
    #    time_type カラムが無いので diff には影響しない)
    patient_a.weekly_pattern = None
    # 2) PFV-B: 患者 B (拠点 INAGE) に対する PFV だが template_b (拠点 TSUGA) を
    #    指している (クロスオフィス参照). export は course_template_code を空欄に倒し、
    #    import 時に course_template_id=None として再投入される → silent update.
    await db.commit()
    await _make_pfv(db, patient_id=patient_a.id, weekday=0, course_template_id=template_a.id)
    await _make_pfv(db, patient_id=patient_b.id, weekday=2, course_template_id=template_b.id)

    # export
    export_res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert export_res.status_code == 200
    exported_bytes = export_res.content

    # そのまま import (dry_run=True で error 計上を確認)
    files = {
        "file": (
            "exported.xlsx",
            exported_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    summary = body["summary"]
    # round-trip では「変更なし」(noop) が主体になる. error は 0 件.
    assert summary["patients_error"] == 0, body
    assert summary["pfv_error"] == 0, body
    # Phase G-51: クロス拠点コースは拠点付きトークン (例 "津C") で表現されるため、
    # 旧来の silent update (course_template_id が剥がれる) は解消され、PFV-A/PFV-B とも
    # 完全 noop で round-trip する (データロスゼロ).
    assert summary["pfv_update"] == 0, body
    assert summary["pfv_noop"] == 2, body


# E-4-7) round-trip (クリーンデータ): 同じ拠点の template + weekly_pattern 設定済 + 全列適切な
# patient で export → import すると、patient と PFV が共に完全 noop になる
# (= データロスや silent update が一切無いことを担保).
@pytest.mark.asyncio
async def test_patient_export_then_import_clean_data_full_noop_round_trip(client, db) -> None:
    admin = await _make_user(db, "e4-roundtrip-clean@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    template = CourseTemplate(office_id=office.id, label="M")
    db.add(template)
    await db.commit()
    await db.refresh(template)

    patient = await _make_patient(
        db,
        code="P-RT-CLEAN",
        name="クリーン患者",
        primary_office_id=office.id,
    )
    # weekly_pattern を設定 (time_type を export で resolvable にする).
    patient.weekly_pattern = {"time_type": "固定"}
    await db.commit()
    await _make_pfv(
        db,
        patient_id=patient.id,
        weekday=0,
        course_template_id=template.id,
        start_time=time(9, 0),
        duration_min=30,
    )

    # export → そのまま import.
    export_res = await client.get(
        "/api/v1/patients/import-export/export",
        headers=_bearer(admin),
    )
    assert export_res.status_code == 200
    files = {
        "file": (
            "exported.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    summary = body["summary"]
    # クリーンデータでは patient / PFV ともに完全 noop. update / error は 0 件.
    assert summary["patients_error"] == 0, body
    assert summary["patients_update"] == 0, body
    assert summary["patients_noop"] == 1, body
    assert summary["pfv_error"] == 0, body
    assert summary["pfv_update"] == 0, body
    assert summary["pfv_noop"] == 1, body


# ---------------------------------------------------------------------------
# Phase E-7 (gap P0-1): Patient.requires_multiple_staff Excel 往復
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_writes_requires_multiple_staff_true(client, db) -> None:
    """requires_multiple_staff=True の患者を export すると列に "TRUE" が入る."""
    admin = await _make_user(db, "ex-rms-t@example.com", "admin")
    await _make_patient(
        db, code="P-RMS-T", name="複数必須", status="active", requires_multiple_staff=True
    )
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PATIENTS]
    val = ws.cell(row=2, column=PATIENT_COL_INDEX["requires_multiple_staff"] + 1).value
    # Phase G-48: 日本語ラベル「はい/いいえ」で書き出す.
    assert val == "はい"


@pytest.mark.asyncio
async def test_export_writes_requires_multiple_staff_false(client, db) -> None:
    """requires_multiple_staff=False の患者を export すると列に "FALSE" が入る."""
    admin = await _make_user(db, "ex-rms-f@example.com", "admin")
    await _make_patient(
        db, code="P-RMS-F", name="単独可", status="active", requires_multiple_staff=False
    )
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PATIENTS]
    val = ws.cell(row=2, column=PATIENT_COL_INDEX["requires_multiple_staff"] + 1).value
    # Phase G-48: 日本語ラベル「はい/いいえ」で書き出す.
    assert val == "いいえ"


@pytest.mark.asyncio
async def test_import_new_patient_with_requires_multiple_staff_true(client, db) -> None:
    """新規患者の requires_multiple_staff=TRUE が DB に保存される."""
    admin = await _make_user(db, "im-rms-new@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-NEW-RMS",
                "name": "新規複数必須",
                "sex": "female",
                "status": "active",
                "address": "千葉市稲毛区",
                "requires_multiple_staff": "TRUE",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p = (await db.scalars(select(Patient).where(Patient.code == "P-NEW-RMS"))).first()
    assert p is not None
    assert p.requires_multiple_staff is True


@pytest.mark.asyncio
async def test_import_update_existing_requires_multiple_staff(client, db) -> None:
    """既存患者の requires_multiple_staff を TRUE → FALSE に変更."""
    admin = await _make_user(db, "im-rms-upd@example.com", "admin")
    p = await _make_patient(
        db, code="P-RMS-UPD", name="切替対象", status="active", requires_multiple_staff=True
    )
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-RMS-UPD",
                "requires_multiple_staff": "FALSE",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_update"] == 1
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    assert p_after.requires_multiple_staff is False


@pytest.mark.asyncio
async def test_roundtrip_requires_multiple_staff(client, db) -> None:
    """export → import で requires_multiple_staff が保持される."""
    admin = await _make_user(db, "rt-rms@example.com", "admin")
    p = await _make_patient(
        db, code="P-RT-RMS", name="往復対象", status="active", requires_multiple_staff=True
    )
    pid = p.id
    # export
    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "export.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    # そのまま再 import
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    assert body["summary"]["patients_error"] == 0
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    assert p_after.requires_multiple_staff is True  # 保持


@pytest.mark.asyncio
async def test_import_blank_requires_multiple_staff_preserves_existing(client, db) -> None:
    """既存患者の更新で requires_multiple_staff を空セルにすると値が維持される (通常 import)."""
    admin = await _make_user(db, "im-rms-blank@example.com", "admin")
    p = await _make_patient(
        db, code="P-RMS-BLANK", name="維持対象", status="active", requires_multiple_staff=True
    )
    pid = p.id
    # name を変更するが requires_multiple_staff は空セル
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-RMS-BLANK",
                "name": "新名前",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    assert p_after.name == "新名前"
    assert p_after.requires_multiple_staff is True  # 維持された


# ---------------------------------------------------------------------------
# Phase E-8: 希望訪問パターン (patient.weekly_pattern) Excel 対応
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_includes_weekly_pattern_in_patient_sheet(client, db) -> None:
    """Phase G-48: weekly_pattern は患者マスタシートに統合して書き出される.

    旧 Phase E-8 の独立「希望訪問パターン」シートは export では作られない.
    """
    admin = await _make_user(db, "e8-export@example.com", "admin")
    p = await _make_patient(db, code="P-E8-1", name="希望あり患者")
    p.weekly_pattern = {
        "time_type": "時間帯",
        "preferred_start": "09:00",
        "preferred_end": "12:00",
        "preferred_weekdays": ["Mon", "Wed", "Fri"],
        "service_minutes": 35,
        "frequency_per_week": 3,
        "visit_frequency": "every",
    }
    await db.commit()
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    # 独立シートは作られない (2 シート構成).
    assert SHEET_WEEKLY not in wb.sheetnames
    ws = wb[SHEET_PATIENTS]
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[PATIENT_COL_INDEX["patient_code"]] == "P-E8-1"
    assert row[PATIENT_COL_INDEX["weekly_time_type"]] == "時間帯"
    assert row[PATIENT_COL_INDEX["preferred_start"]] == "09:00"
    assert row[PATIENT_COL_INDEX["preferred_end"]] == "12:00"
    assert row[PATIENT_COL_INDEX["frequency_per_week"]] == 3
    assert row[PATIENT_COL_INDEX["service_minutes"]] == 35
    # 訪問頻度: DB "every" → 日本語 "毎週".
    assert row[PATIENT_COL_INDEX["visit_frequency"]] == "毎週"
    # Phase G-50: 希望曜日は 7 列 (月..日) の 〇/×. Mon/Wed/Fri が「〇」.
    assert row[PATIENT_COL_INDEX["pref_wd_mon"]] == "〇"
    assert row[PATIENT_COL_INDEX["pref_wd_tue"]] == "×"
    assert row[PATIENT_COL_INDEX["pref_wd_wed"]] == "〇"
    assert row[PATIENT_COL_INDEX["pref_wd_thu"]] == "×"
    assert row[PATIENT_COL_INDEX["pref_wd_fri"]] == "〇"
    assert row[PATIENT_COL_INDEX["pref_wd_sat"]] == "×"
    assert row[PATIENT_COL_INDEX["pref_wd_sun"]] == "×"


@pytest.mark.asyncio
async def test_import_updates_weekly_pattern(client, db) -> None:
    """Phase G-48: 統合シートで patient.weekly_pattern を更新できる."""
    admin = await _make_user(db, "e8-import@example.com", "admin")
    p = await _make_patient(db, code="P-E8-2", name="更新対象")
    p.weekly_pattern = None
    await db.commit()
    pid = p.id

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-E8-2",
                "name": "更新対象",
                "sex": "男性",
                "status": "稼働",
                "address": "千葉市稲毛区test1",
                "weekly_time_type": "固定",
                "preferred_start": "10:00",
                "service_minutes": 45,
                # Phase G-49: 希望曜日は 7 列 はい/いいえ. 月・水を「はい」.
                "pref_wd_mon": "はい",
                "pref_wd_wed": "はい",
                "visit_frequency": "隔週",
                "frequency_per_week": 2,
            }
        ],
    )

    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern["time_type"] == "固定"
    assert p_after.weekly_pattern["preferred_start"] == "10:00"
    assert p_after.weekly_pattern["service_minutes"] == 45
    assert set(p_after.weekly_pattern.get("preferred_weekdays", [])) == {"Mon", "Wed"}
    # 訪問頻度は DB 英語正準値へ正規化される.
    assert p_after.weekly_pattern["visit_frequency"] == "biweekly"
    assert p_after.weekly_pattern["frequency_per_week"] == 2


@pytest.mark.asyncio
async def test_weekly_pattern_roundtrip(client, db) -> None:
    """Phase G-48: export → 編集なし import で weekly_pattern が完全 noop."""
    admin = await _make_user(db, "e8-roundtrip@example.com", "admin")
    p = await _make_patient(db, code="P-E8-RT", name="round-trip")
    p.weekly_pattern = {
        "time_type": "時間帯",
        "preferred_start": "13:00",
        "preferred_end": "15:00",
        "preferred_weekdays": ["Tue", "Thu"],
        "service_minutes": 30,
        "visit_frequency": "monthly",
    }
    await db.commit()
    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "exported.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    body = import_res.json()
    summary = body["summary"]
    # round-trip: clean data → 完全 noop (weekly_pattern も含む)
    assert summary["patients_error"] == 0, body
    assert summary["patients_update"] == 0, body
    assert summary["patients_noop"] == 1, body


@pytest.mark.asyncio
async def test_weekly_pattern_roundtrip_preserves_unmanaged_keys(client, db) -> None:
    """CRITICAL 回帰防止 (Phase G-48 hotfix): weekly_pattern に管理外キー
    (entries / staff_count) を持つ患者を export → 無編集 import したとき、

      (a) patients_update == 0 (真の noop)
      (b) DB 上の entries / staff_count が **保持** される

    ことを担保する. exporter は管理 8 キーのみ書き出し importer も 8 キーのみ
    再構築するため、丸ごと置換だと無編集でも entries が消えて update 扱いになる
    (= スケジューラの訪問時刻 / 2 名体制が静かに壊れる). merge 化でこれを防ぐ.

    既存の round-trip テスト (test_weekly_pattern_roundtrip) は 8 キーのみで
    管理外キーが盲点だったため、本テストを管理外キー入りで追加する.

    visit_frequency は canonical 値 (biweekly) を入れて round-trip が noop に
    なる (= legacy-JP 正規化が canonical をそのまま通す) ことも確認する.
    """
    admin = await _make_user(db, "g48-rt-unmanaged@example.com", "admin")
    p = await _make_patient(db, code="P-G48-RT-ENTRIES", name="管理外キー保持")
    original_wp = {
        # ---- 管理 8 キー (exporter/import が扱う、round-trip で復元される) ----
        "frequency_per_week": 3,
        "visit_frequency": "biweekly",  # canonical 値 (JP 正規化が noop で通すこと)
        "preferred_weekdays": ["Mon", "Wed", "Fri"],
        "service_minutes": 45,
        "time_type": "時間帯",
        "preferred_start": "10:00",
        "preferred_end": "11:30",
        # ---- 管理外キー (本パイプライン非管理. merge で保持されねばならない) ----
        "staff_count": 1,
        "entries": [
            {
                "weekday": "Mon",
                "time_type": "時間帯",
                "preferred_start": "10:00",
                "preferred_end": "11:30",
                "service_minutes": 45,
                "staff_count": 1,
            },
            {
                "weekday": "Wed",
                "time_type": "固定",
                "preferred_start": "14:00",
                "preferred_end": "15:00",
                "service_minutes": 60,
                "staff_count": 2,  # 2 名体制エントリ — 消えると静かに壊れる
            },
            {
                "weekday": "Fri",
                "time_type": "時間帯",
                "preferred_start": "10:00",
                "preferred_end": "11:30",
                "service_minutes": 45,
                "staff_count": 1,
            },
        ],
    }
    p.weekly_pattern = dict(original_wp)
    await db.commit()
    pid = p.id

    # export → 無編集 import (dry_run でまず noop を確認)
    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    exported = export_res.content
    dry_files = {
        "file": (
            "exported.xlsx",
            exported,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    dry_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=dry_files,
    )
    assert dry_res.status_code == 200, dry_res.text
    summary = dry_res.json()["summary"]
    # (a) 真の noop — entries 持ち患者でも update にならない
    assert summary["patients_error"] == 0, dry_res.json()
    assert summary["patients_update"] == 0, dry_res.json()
    assert summary["patients_noop"] == 1, dry_res.json()

    # apply して DB 上の entries / staff_count が保持されることを確認
    apply_files = {
        "file": (
            "exported.xlsx",
            exported,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    apply_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=false",
        headers=_bearer(admin),
        files=apply_files,
    )
    assert apply_res.status_code == 200, apply_res.text
    assert apply_res.json()["summary"]["patients_update"] == 0, apply_res.json()

    db.expire_all()
    p_after = await db.get(Patient, pid)
    # (b) 管理外キーが保持される
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern.get("staff_count") == 1
    assert p_after.weekly_pattern.get("entries") == original_wp["entries"]
    # 2 名体制エントリも保持
    wed = next(e for e in p_after.weekly_pattern["entries"] if e["weekday"] == "Wed")
    assert wed["staff_count"] == 2
    # 管理 8 キーも維持 (canonical visit_frequency が round-trip 安定)
    assert p_after.weekly_pattern["visit_frequency"] == "biweekly"
    assert p_after.weekly_pattern["frequency_per_week"] == 3


@pytest.mark.asyncio
async def test_weekly_pattern_partial_edit_preserves_unmanaged_keys(client, db) -> None:
    """Phase G-48 hotfix: 管理 8 キーの一部だけ編集して import した場合でも、
    管理外キー (entries/staff_count) は保持され、編集した管理キーのみ反映される.

    blank=keep + 管理外キー保持 の両立を確認する (merge セマンティクス)."""
    admin = await _make_user(db, "g48-partial-unmanaged@example.com", "admin")
    p = await _make_patient(db, code="P-G48-PARTIAL", name="部分編集")
    p.weekly_pattern = {
        "time_type": "時間帯",
        "preferred_start": "09:00",
        "preferred_end": "10:00",
        "service_minutes": 30,
        "staff_count": 2,
        "entries": [
            {"weekday": "Tue", "time_type": "時間帯", "staff_count": 2},
        ],
    }
    await db.commit()
    pid = p.id
    # service_minutes のみ 30 → 60 に変更、他 weekly 列は export 相当で埋める.
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G48-PARTIAL",
                "name": "部分編集",
                "weekly_time_type": "時間帯",
                "preferred_start": "09:00",
                "preferred_end": "10:00",
                "service_minutes": 60,  # ここだけ変更
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    # 編集した管理キーは反映
    assert p_after.weekly_pattern["service_minutes"] == 60
    # 管理外キーは保持
    assert p_after.weekly_pattern.get("staff_count") == 2
    assert p_after.weekly_pattern.get("entries") == [
        {"weekday": "Tue", "time_type": "時間帯", "staff_count": 2}
    ]


# ---------------------------------------------------------------------------
# Phase G-48: 日本語化 + 英↔日マッピング + code 突合 (UUID 不要化) + 統合シート
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_writes_japanese_enum_labels(client, db) -> None:
    """Phase G-48: enum 系が日本語ラベルで書き出される (DB 英語値 → 日本語)."""
    admin = await _make_user(db, "g48-ex-ja@example.com", "admin")
    await _make_patient(
        db,
        code="P-G48-EX",
        name="日本語出力",
        sex="female",
        status="suspended",
        insurance="medical",
        sex_restriction="female_only",
        address="千葉市稲毛区test1",
    )
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PATIENTS]
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[PATIENT_COL_INDEX["sex"]] == "女性"
    assert row[PATIENT_COL_INDEX["status"]] == "休止"
    assert row[PATIENT_COL_INDEX["insurance"]] == "医療保険"
    assert row[PATIENT_COL_INDEX["sex_restriction"]] == "女性のみ"


@pytest.mark.asyncio
async def test_export_sex_restriction_none_writes_nashi(client, db) -> None:
    """sex_restriction が DB NULL のとき「なし」を明示出力する."""
    admin = await _make_user(db, "g48-ex-nashi@example.com", "admin")
    await _make_patient(db, code="P-G48-NASHI", name="制限なし", sex_restriction=None)
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PATIENTS]
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[PATIENT_COL_INDEX["sex_restriction"]] == "なし"


@pytest.mark.asyncio
async def test_import_japanese_enum_labels_new_patient(client, db) -> None:
    """Phase G-48: 日本語ラベルで新規患者を登録すると DB に英語 enum が格納される."""
    admin = await _make_user(db, "g48-im-ja@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-G48-JA-NEW",
                "name": "日本語入力",
                "sex": "女性",
                "status": "入院",
                "insurance": "介護保険",
                "sex_restriction": "男性のみ",
                "address": "千葉市稲毛区xxx",
                "requires_multiple_staff": "はい",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p = await db.scalar(select(Patient).where(Patient.code == "P-G48-JA-NEW"))
    assert p is not None
    assert p.sex == "female"
    assert p.status == "admitted"
    assert p.insurance == "care"
    assert p.sex_restriction == "male_only"
    assert p.requires_multiple_staff is True


@pytest.mark.asyncio
async def test_import_english_enum_values_still_accepted(client, db) -> None:
    """後方互換: 英語 enum 値 (旧 export) でも import が壊れない."""
    admin = await _make_user(db, "g48-im-en@example.com", "admin")
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-G48-EN-NEW",
                "name": "英語値互換",
                "sex": "male",
                "status": "active",
                "insurance": "medical",
                "sex_restriction": "female_only",
                "address": "千葉市稲毛区xxx",
                "requires_multiple_staff": "TRUE",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p = await db.scalar(select(Patient).where(Patient.code == "P-G48-EN-NEW"))
    assert p is not None
    assert p.sex == "male"
    assert p.status == "active"
    assert p.insurance == "medical"
    assert p.sex_restriction == "female_only"
    assert p.requires_multiple_staff is True


@pytest.mark.asyncio
async def test_import_sex_restriction_nashi_clears_to_null(client, db) -> None:
    """「なし」を import すると sex_restriction が NULL になる (制限解除)."""
    admin = await _make_user(db, "g48-im-nashi@example.com", "admin")
    p = await _make_patient(db, code="P-G48-CLR", name="制限解除", sex_restriction="female_only")
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G48-CLR",
                "sex_restriction": "なし",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_update"] == 1
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.sex_restriction is None


@pytest.mark.asyncio
async def test_import_update_via_code_without_id(client, db) -> None:
    """Phase G-48 (最重要): patient_id 空 + patient_code で既存患者を UPDATE."""
    admin = await _make_user(db, "g48-code-upd@example.com", "admin")
    p = await _make_patient(
        db, code="P-G48-CODE", name="旧名前", sex="male", status="active", address="旧住所"
    )
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                # patient_id 列なし (UUID 不要) — code だけで突合.
                "patient_code": "P-G48-CODE",
                "name": "新名前",
                "address": "新住所",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_update"] == 1
    assert body["summary"]["patients_new"] == 0
    assert body["patient_rows"][0]["patient_id"] == str(pid)
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.name == "新名前"
    assert p_after.address == "新住所"


@pytest.mark.asyncio
async def test_pfv_mode_japanese_label(client, db) -> None:
    """Phase G-48: PFV モードを日本語「特別」で指定すると DB に "special" が入る."""
    admin = await _make_user(db, "g48-pfv-mode@example.com", "admin")
    p = await _make_patient(db, code="P-G48-PFV", name="モード日本語")
    pid = p.id
    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(pid),
                "weekday": "月",
                "slot_index": 0,
                "mode": "特別",
                "time_type": "固定",
                "start_time": "09:00",
                "duration_min": 30,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    assert len(rows) == 1
    assert rows[0].mode == "special"


@pytest.mark.asyncio
async def test_export_pfv_special_mode_excluded_from_edit_sheet(client, db) -> None:
    """Phase G-51: special mode PFV のみの患者は「編集用」シートに出力されない.

    「編集用」シートは normal mode の確定週次固定枠の正本のみを 1 患者 1 行で扱う.
    special mode は本シート対象外で DB に保持される (誤削除防止).
    """
    admin = await _make_user(db, "g48-ex-pfv-mode@example.com", "admin")
    p = await _make_patient(db, code="P-G48-PFVEX", name="モード出力")
    await _make_pfv(db, patient_id=p.id, weekday=0, mode="special")
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    wb = load_workbook(BytesIO(res.content))
    ws_f = wb[SHEET_PFV_EDIT]
    # special のみ → 編集用シートに患者行は無い (ヘッダーのみ).
    assert ws_f.max_row == 1


@pytest.mark.asyncio
async def test_weekday_7columns_yesno_input(client, db) -> None:
    """Phase G-49: 希望曜日 7 列 (月..日) の「はい」を読み preferred_weekdays に反映."""
    admin = await _make_user(db, "g49-wd-7col@example.com", "admin")
    p = await _make_patient(db, code="P-G49-WD", name="曜日7列")
    p.weekly_pattern = None
    await db.commit()
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G49-WD",
                # 月・火・水・金 を「はい」、他は空 (= いいえ相当).
                "pref_wd_mon": "はい",
                "pref_wd_tue": "はい",
                "pref_wd_wed": "はい",
                "pref_wd_fri": "はい",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.weekly_pattern is not None
    # Mon..Sun 正準順で格納される.
    assert p_after.weekly_pattern["preferred_weekdays"] == ["Mon", "Tue", "Wed", "Fri"]


@pytest.mark.asyncio
async def test_weekday_7columns_mark_input(client, db) -> None:
    """Phase G-50: 希望曜日 7 列の「〇」を選択扱いで読み preferred_weekdays に反映.
    「×」は非選択. 〇/× が新しい既定の選択肢."""
    admin = await _make_user(db, "g50-wd-mark@example.com", "admin")
    p = await _make_patient(db, code="P-G50-WD", name="曜日マーク")
    p.weekly_pattern = None
    await db.commit()
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G50-WD",
                # 月・水・金 を「〇」、火・木・土・日 を「×」.
                "pref_wd_mon": "〇",
                "pref_wd_tue": "×",
                "pref_wd_wed": "〇",
                "pref_wd_thu": "×",
                "pref_wd_fri": "〇",
                "pref_wd_sat": "×",
                "pref_wd_sun": "×",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern["preferred_weekdays"] == ["Mon", "Wed", "Fri"]


@pytest.mark.asyncio
async def test_weekly_blank_keeps_existing(client, db) -> None:
    """weekly フィールド全空の行は既存 weekly_pattern を維持 (clear しない)."""
    admin = await _make_user(db, "g48-wk-keep@example.com", "admin")
    p = await _make_patient(db, code="P-G48-KEEP", name="維持")
    existing_wp = {
        "time_type": "時間帯",
        "preferred_start": "09:00",
        "preferred_end": "12:00",
        "preferred_weekdays": ["Mon"],
    }
    p.weekly_pattern = dict(existing_wp)
    await db.commit()
    pid = p.id
    # name だけ変更、weekly 列は全空.
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G48-KEEP",
                "name": "新名前",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.name == "新名前"
    # weekly_pattern は破壊されず維持される.
    assert p_after.weekly_pattern == existing_wp


@pytest.mark.asyncio
async def test_full_japanese_roundtrip_noop(client, db) -> None:
    """Phase G-48: 日本語 enum + weekly 統合 + 全列適切な患者で export → 無編集
    import が完全 noop (0 変更) になる (round-trip 安定性、最重要)."""
    admin = await _make_user(db, "g48-rt-noop@example.com", "admin")
    office = await _make_office(db, code="INAGE", name="稲毛")
    p = await _make_patient(
        db,
        code="P-G48-RT",
        name="往復患者",
        kana="オウフクカンジャ",
        sex="female",
        status="active",
        insurance="care",
        address="千葉市稲毛区test1",
        sex_restriction="female_only",
        requires_multiple_staff=True,
        primary_office_id=office.id,
    )
    p.weekly_pattern = {
        "frequency_per_week": 3,
        "visit_frequency": "every",
        "visit_weeks": "1,3",
        "preferred_weekdays": ["Mon", "Wed", "Fri"],
        "service_minutes": 40,
        "time_type": "時間帯",
        "preferred_start": "09:00",
        "preferred_end": "12:00",
    }
    await db.commit()

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "exported.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    summary = import_res.json()["summary"]
    assert summary["patients_error"] == 0, import_res.json()
    assert summary["patients_update"] == 0, import_res.json()
    assert summary["patients_noop"] == 1, import_res.json()


@pytest.mark.asyncio
async def test_backward_compat_legacy_three_sheet_english(client, db) -> None:
    """後方互換: 旧 3 シート構成 (英語値 + 曜日 7 列 TRUE/FALSE) も import 可能.

    旧 export ファイル (Phase E-8 形式) を手で組み立てて import する.
    """
    from app.services.patient_excel.schema import WEEKLY_COLUMNS

    admin = await _make_user(db, "g48-bc-legacy@example.com", "admin")
    p = await _make_patient(db, code="P-G48-LEGACY", name="旧形式")
    p.weekly_pattern = None
    await db.commit()
    pid = p.id

    wb = Workbook()
    # 患者シート: 旧形式は patient_id を持つだけの update なし行でよい.
    # ただし旧 patient sheet には weekly 列が無い前提なので、新ヘッダーで空行を作る
    # と weekly は空 = 維持. weekly は独立シートから来る.
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    ws_p.append([str(c["header"]) for c in PATIENT_COLUMNS])
    prow = [None] * len(PATIENT_COLUMNS)
    prow[PATIENT_COL_INDEX["patient_id"]] = str(pid)
    prow[PATIENT_COL_INDEX["patient_code"]] = "P-G48-LEGACY"
    ws_p.append(prow)
    # PFV シート (空).
    pfv_ws = wb.create_sheet(SHEET_PFV)
    pfv_ws.append([str(c["header"]) for c in PFV_COLUMNS])
    # 旧 独立「希望訪問パターン」シート (曜日 7 列 TRUE/FALSE).
    ws_w = wb.create_sheet(SHEET_WEEKLY)
    ws_w.append([str(c["header"]) for c in WEEKLY_COLUMNS])
    wrow = [None] * len(WEEKLY_COLUMNS)
    wrow[WEEKLY_COL_INDEX["patient_id"]] = str(pid)
    wrow[WEEKLY_COL_INDEX["patient_code"]] = "P-G48-LEGACY"
    wrow[WEEKLY_COL_INDEX["time_type"]] = "固定"
    wrow[WEEKLY_COL_INDEX["preferred_start"]] = "10:00"
    wrow[WEEKLY_COL_INDEX["wd_tue"]] = "TRUE"
    wrow[WEEKLY_COL_INDEX["wd_thu"]] = "TRUE"
    # 旧形式では visit_frequency に日本語が入っていた可能性 → 英語へ正規化されること.
    wrow[WEEKLY_COL_INDEX["visit_frequency"]] = "毎週"
    ws_w.append(wrow)
    buf = BytesIO()
    wb.save(buf)

    res = await _upload(client, admin, content=buf.getvalue(), dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern["time_type"] == "固定"
    assert p_after.weekly_pattern["preferred_start"] == "10:00"
    assert set(p_after.weekly_pattern["preferred_weekdays"]) == {"Tue", "Thu"}
    # 旧シートの日本語 visit_frequency も英語正準値へ正規化される.
    assert p_after.weekly_pattern["visit_frequency"] == "every"


@pytest.mark.asyncio
async def test_backward_compat_legacy_weekly_sheet_delete_clears(client, db) -> None:
    """後方互換 / HIGH 回帰防止 (Phase G-48 hotfix-2):

    旧 3 シート構成の「希望訪問パターン」シートに ``<DELETE>`` 行を含むファイルを
    import すると、(a) 例外 (TypeError) を出さず 200 で完了し、(b) 当該患者の
    weekly_pattern が None にクリアされること.

    旧経路は ``<DELETE>`` 行で merge 関数へ ``wp=None`` を渡すため、None ガードが
    無いと ``key in None`` で TypeError → import が 500 / rollback していた.
    """
    from app.services.patient_excel.schema import WEEKLY_COLUMNS

    admin = await _make_user(db, "g48-bc-legacy-del@example.com", "admin")
    p = await _make_patient(db, code="P-G48-LEGDEL", name="旧形式削除")
    # 既存 weekly_pattern (管理外キー込み) を設定 → クリア対象.
    p.weekly_pattern = {
        "time_type": "固定",
        "preferred_start": "10:00",
        "preferred_weekdays": ["Tue", "Thu"],
        "staff_count": 2,
        "entries": [{"weekday": "Tue", "start": "10:00"}],
    }
    await db.commit()
    pid = p.id

    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    ws_p.append([str(c["header"]) for c in PATIENT_COLUMNS])
    prow = [None] * len(PATIENT_COLUMNS)
    prow[PATIENT_COL_INDEX["patient_id"]] = str(pid)
    prow[PATIENT_COL_INDEX["patient_code"]] = "P-G48-LEGDEL"
    ws_p.append(prow)
    pfv_ws = wb.create_sheet(SHEET_PFV)
    pfv_ws.append([str(c["header"]) for c in PFV_COLUMNS])
    # 旧 独立「希望訪問パターン」シートに <DELETE> 行.
    ws_w = wb.create_sheet(SHEET_WEEKLY)
    ws_w.append([str(c["header"]) for c in WEEKLY_COLUMNS])
    wrow = [None] * len(WEEKLY_COLUMNS)
    wrow[WEEKLY_COL_INDEX["patient_id"]] = str(pid)
    wrow[WEEKLY_COL_INDEX["patient_code"]] = "P-G48-LEGDEL"
    wrow[WEEKLY_COL_INDEX["delete_flag"]] = "<DELETE>"
    ws_w.append(wrow)
    buf = BytesIO()
    wb.save(buf)

    # (a) 例外を出さず 200 で完了する.
    res = await _upload(client, admin, content=buf.getvalue(), dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    # (b) weekly_pattern が None にクリアされる (明示削除なので entries 等も消える).
    assert p_after.weekly_pattern is None


# ---------------------------------------------------------------------------
# Phase G-49: 拠点コード自動割当 (住所ベース) / dropdown / 7 列曜日 / 条件付き書式
# ---------------------------------------------------------------------------


async def _seed_inage_city_office(db) -> Office:
    """稲毛区 → INAGE の city / office_cities を seed して INAGE office を返す."""
    inage_city = City(prefecture="千葉県", name="千葉市稲毛区", jis_code="12103")
    db.add(inage_city)
    await db.commit()
    await db.refresh(inage_city)
    office = await _make_office(db, "INAGE", "稲毛")
    db.add(OfficeCity(office_id=office.id, city_id=inage_city.id))
    await db.commit()
    return office


@pytest.mark.asyncio
async def test_import_new_patient_office_auto_assigned_from_address(client, db) -> None:
    """Phase G-49: 拠点コード空欄の新規患者は住所から primary_office_id を自動割当."""
    admin = await _make_user(db, "g49-auto-new@example.com", "admin")
    office = await _seed_inage_city_office(db)
    office_id = office.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                # patient_id / office_code は空欄. 住所だけから拠点を解決させる.
                "patient_code": "P-G49-AUTO",
                "name": "自動割当",
                "sex": "男性",
                "status": "稼働",
                "address": "千葉県千葉市稲毛区稲毛東1-1-1",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.code == "P-G49-AUTO"))).first()
    assert p_after is not None
    assert p_after.primary_office_id == office_id


@pytest.mark.asyncio
async def test_import_explicit_office_code_respected_over_auto(client, db) -> None:
    """Phase G-49: 拠点コードが入っていれば住所自動割当より明示コードを尊重."""
    admin = await _make_user(db, "g49-explicit@example.com", "admin")
    inage = await _seed_inage_city_office(db)
    inage_id = inage.id
    # 別拠点 TSUGA も用意 (住所は稲毛区だが TSUGA を明示指定).
    tsuga = await _make_office(db, "TSUGA", "都賀")
    tsuga_id = tsuga.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-G49-EXP",
                "name": "明示尊重",
                "sex": "男性",
                "status": "稼働",
                "address": "千葉県千葉市稲毛区稲毛東1-1-1",
                "office_code": "TSUGA",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.code == "P-G49-EXP"))).first()
    assert p_after is not None
    # 明示コード (TSUGA) が尊重され、住所自動割当 (INAGE) では上書きされない.
    assert p_after.primary_office_id == tsuga_id
    assert p_after.primary_office_id != inage_id


@pytest.mark.asyncio
async def test_import_existing_patient_blank_office_not_auto_assigned(client, db) -> None:
    """Phase G-49: 既存患者の拠点コード空欄は「維持」(住所自動割当しない)."""
    admin = await _make_user(db, "g49-existing@example.com", "admin")
    await _seed_inage_city_office(db)
    # 既存患者 (primary_office 未設定). 空欄 import で None のまま維持される.
    p = await _make_patient(
        db,
        code="P-G49-EXIST",
        name="既存維持",
        address="千葉県千葉市稲毛区test",
        primary_office_id=None,
    )
    pid = p.id
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G49-EXIST",
                # office_code 空欄 → 既存患者は触らない.
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    # 既存患者・空欄は自動割当しないので noop (= primary_office_id 変更なし).
    assert body["patient_rows"][0]["operation"] == "noop"


@pytest.mark.asyncio
async def test_export_office_code_roundtrip_noop(client, db) -> None:
    """Phase G-49: 拠点コードあり export → 無編集 import = noop (round-trip 安定)."""
    admin = await _make_user(db, "g49-rt-office@example.com", "admin")
    office = await _seed_inage_city_office(db)
    await _make_patient(
        db,
        code="P-G49-RT",
        name="拠点RT",
        address="千葉県千葉市稲毛区test",
        primary_office_id=office.id,
    )
    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "exported.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    summary = import_res.json()["summary"]
    assert summary["patients_error"] == 0
    assert summary["patients_update"] == 0
    assert summary["patients_noop"] == 1


@pytest.mark.asyncio
async def test_template_has_dropdowns_and_conditional_format(client, db) -> None:
    """Phase G-49: テンプレートに dropdown (1-7 / 5 分刻み / HH:MM / 7 列はい・いいえ)
    と時刻グレーアウトの条件付き書式が設定される."""
    admin = await _make_user(db, "g49-tmpl@example.com", "admin")
    res = await client.get("/api/v1/patients/import-export/template", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PATIENTS]
    # dropdown formula を集約.
    dv_formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    joined = " ".join(dv_formulas)
    # 週訪問回数 1〜7.
    assert '"1,2,3,4,5,6,7"' in dv_formulas
    # サービス時間 5 分刻み (15 / 180 が含まれる).
    assert any("15," in f and "180" in f for f in dv_formulas)
    # HH:MM (06:00 / 20:00 が含まれる).
    assert any("06:00" in f and "20:00" in f for f in dv_formulas)
    # Phase G-50: 希望曜日 7 列は 〇/× dropdown (7 列分).
    assert joined.count('"〇,×"') >= 7
    # 複数スタッフ必須は引き続き はい/いいえ.
    assert '"はい,いいえ"' in joined
    # 時間タイプ列を参照する条件付き書式が時刻 2 列に設定されている.
    cf_ranges = [str(r.sqref) for r in ws.conditional_formatting]
    assert len(cf_ranges) >= 2
    # 数式に 午前/午後/終日 が含まれる.
    all_rules_have_greyout = False
    for rng in ws.conditional_formatting:
        for rule in ws.conditional_formatting[rng]:
            if rule.formula and any("午前" in f for f in rule.formula):
                all_rules_have_greyout = True
    assert all_rules_have_greyout


@pytest.mark.asyncio
async def test_weekday_7columns_roundtrip_noop(client, db) -> None:
    """Phase G-49: 7 列曜日 export → 無編集 import = noop (round-trip 安定)."""
    admin = await _make_user(db, "g49-wd-rt@example.com", "admin")
    p = await _make_patient(db, code="P-G49-WDRT", name="曜日RT")
    p.weekly_pattern = {
        "preferred_weekdays": ["Mon", "Wed", "Fri"],
        "time_type": "時間帯",
        "preferred_start": "09:00",
        "preferred_end": "10:00",
    }
    await db.commit()
    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "exported.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    summary = import_res.json()["summary"]
    assert summary["patients_error"] == 0
    assert summary["patients_update"] == 0
    assert summary["patients_noop"] == 1


@pytest.mark.asyncio
async def test_weekday_7columns_all_no_keeps_existing(client, db) -> None:
    """Phase G-49/G-50: 7 列全「×」/空でも preferred_weekdays は維持 (blank=keep)."""
    admin = await _make_user(db, "g49-wd-keep@example.com", "admin")
    p = await _make_patient(db, code="P-G49-WDKEEP", name="曜日維持")
    p.weekly_pattern = {"preferred_weekdays": ["Mon"], "time_type": "時間帯"}
    await db.commit()
    pid = p.id
    # 曜日 7 列を全て「×」、他 weekly も空 → weekly 全空扱い = 維持.
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G49-WDKEEP",
                "pref_wd_mon": "×",
                "pref_wd_tue": "×",
                "pref_wd_wed": "×",
                "pref_wd_thu": "×",
                "pref_wd_fri": "×",
                "pref_wd_sat": "×",
                "pref_wd_sun": "×",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = await db.get(Patient, pid)
    # 全いいえ + 他空 → weekly 全空 = 既存維持 (Mon 保持).
    assert p_after.weekly_pattern is not None
    assert p_after.weekly_pattern["preferred_weekdays"] == ["Mon"]


# ---------------------------------------------------------------------------
# Phase G-51: 「固定訪問パターン（編集用）」(患者 1 行) + 静的グリッド集計
# ---------------------------------------------------------------------------


async def _setup_two_office_courses(db):
    """INAGE/TSUGA 拠点 + course_templates (稲B, 稲C, 津A) を作る helper."""
    office_inage = await _make_office(db, code="INAGE", name="稲毛")
    office_tsuga = await _make_office(db, code="TSUGA", name="都賀")
    cts = {
        ("INAGE", "B"): CourseTemplate(office_id=office_inage.id, label="B"),
        ("INAGE", "C"): CourseTemplate(office_id=office_inage.id, label="C"),
        ("TSUGA", "A"): CourseTemplate(office_id=office_tsuga.id, label="A"),
    }
    for ct in cts.values():
        db.add(ct)
    await db.commit()
    for ct in cts.values():
        await db.refresh(ct)
    return office_inage, office_tsuga, cts


# G-51-1) round-trip: export → 無編集 import が全 noop (0 更新 0 エラー).
@pytest.mark.asyncio
async def test_pfv_edit_export_import_round_trip_noop(client, db) -> None:
    admin = await _make_user(db, "g51-rt@example.com", "admin")
    office_inage, office_tsuga, cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-RT", name="往復患者", primary_office_id=office_inage.id
    )
    patient.weekly_pattern = {"time_type": "固定"}
    await db.commit()
    # 月=稲B, 木=津A (クロス拠点), duration 35 で統一.
    await _make_pfv(
        db,
        patient_id=patient.id,
        weekday=0,
        start_time=time(9, 0),
        duration_min=35,
        course_template_id=cts[("INAGE", "B")].id,
    )
    await _make_pfv(
        db,
        patient_id=patient.id,
        weekday=3,
        start_time=time(10, 30),
        duration_min=35,
        course_template_id=cts[("TSUGA", "A")].id,
    )

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    files = {
        "file": (
            "exported.xlsx",
            export_res.content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    summary = import_res.json()["summary"]
    assert summary["pfv_error"] == 0, summary
    assert summary["pfv_new"] == 0, summary
    assert summary["pfv_update"] == 0, summary
    assert summary["pfv_delete"] == 0, summary
    assert summary["pfv_noop"] == 2, summary
    assert summary["patients_error"] == 0, summary


# G-51-2) per-patient replace: 行に無い曜日の既存 normal PFV は削除、ある曜日は upsert.
@pytest.mark.asyncio
async def test_pfv_edit_per_patient_replace_deletes_absent_weekday(client, db) -> None:
    admin = await _make_user(db, "g51-replace@example.com", "admin")
    office_inage, _office_tsuga, cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-REP", name="置換患者", primary_office_id=office_inage.id
    )
    pid = patient.id
    expected_ct_id = cts[("INAGE", "B")].id  # expire_all 前に capture.
    # 既存: 月 + 水 の 2 枠.
    await _make_pfv(db, patient_id=pid, weekday=0, start_time=time(9, 0), duration_min=30)
    await _make_pfv(db, patient_id=pid, weekday=2, start_time=time(9, 0), duration_min=30)

    # 編集用シート: 月だけ残す (水は時刻空欄 = 削除), 月の時刻を 10:00 に変更.
    content = _build_edit_workbook_bytes(
        pfv_edit_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G51-REP",
                "service_minutes": 30,
                "time_type": "固定",
                "mon_time": "10:00",
                "mon_course": "稲B",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_update"] == 1  # 月: 時刻 + course 更新
    assert body["summary"]["pfv_delete"] == 1  # 水: 削除
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    assert len(rows) == 1
    assert rows[0].weekday == 0
    assert rows[0].start_time == time(10, 0)
    assert rows[0].course_template_id == expected_ct_id


# G-51-2b) per-patient replace: 管理対象外曜日 (日曜=weekday6) の既存 normal PFV は
# 削除しない (CRITICAL 回帰防止). 編集用シートは月..土のみ管理するため、行に無くても
# 日曜 PFV は温存され、月..土の削除/upsert 挙動は不変であること.
@pytest.mark.asyncio
async def test_pfv_edit_per_patient_replace_preserves_sunday(client, db) -> None:
    admin = await _make_user(db, "g51-sun@example.com", "admin")
    office_inage, _office_tsuga, _cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-SUN", name="日曜温存", primary_office_id=office_inage.id
    )
    pid = patient.id
    # 既存: 月 (管理対象) + 日曜 (weekday=6, 管理対象外) の 2 枠.
    await _make_pfv(db, patient_id=pid, weekday=0, start_time=time(9, 0), duration_min=30)
    sunday = await _make_pfv(db, patient_id=pid, weekday=6, start_time=time(10, 0), duration_min=30)
    sunday_id = sunday.id

    # 編集用シート: 月だけ記入 (日曜は管理対象外でシートに列が無い).
    content = _build_edit_workbook_bytes(
        pfv_edit_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G51-SUN",
                "service_minutes": 30,
                "time_type": "固定",
                "mon_time": "09:00",
                "mon_course": "稲B",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    # 日曜 PFV は削除されない.
    assert body["summary"]["pfv_delete"] == 0
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    # 日曜 PFV は DB に残存.
    assert await db.get(PatientFixedVisit, sunday_id) is not None
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    weekdays = {r.weekday for r in rows}
    assert weekdays == {0, 6}  # 月 + 日曜 (両方残る).


# G-51-3) 拠点付きコース解決: クロス拠点 "津A" が course_template_id + sub_office_id に解決.
@pytest.mark.asyncio
async def test_pfv_edit_cross_office_course_resolves(client, db) -> None:
    admin = await _make_user(db, "g51-cross@example.com", "admin")
    office_inage, office_tsuga, cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-X", name="クロス患者", primary_office_id=office_inage.id
    )
    pid = patient.id
    expected_ct_id = cts[("TSUGA", "A")].id  # expire_all 前に capture.
    content = _build_edit_workbook_bytes(
        pfv_edit_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G51-X",
                "service_minutes": 40,
                "time_type": "固定",
                "thu_time": "11:00",
                "thu_course": "津A",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 0

    db.expire_all()
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    assert len(rows) == 1
    assert rows[0].weekday == 3
    assert rows[0].duration_min == 40
    # クロス拠点 course はトークンに拠点が含まれるため course_template_id で直接表現される.
    assert rows[0].course_template_id == expected_ct_id


# G-51-4) 存在しないコース表記 → best-effort (course_template_id=None で保存、error にしない).
@pytest.mark.asyncio
async def test_pfv_edit_unknown_course_token_best_effort(client, db) -> None:
    admin = await _make_user(db, "g51-unkct@example.com", "admin")
    office_inage, _office_tsuga, _cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-UNK", name="未知コース", primary_office_id=office_inage.id
    )
    pid = patient.id
    content = _build_edit_workbook_bytes(
        pfv_edit_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G51-UNK",
                "service_minutes": 30,
                "time_type": "固定",
                "mon_time": "09:00",
                "mon_course": "稲Z",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["pfv_error"] == 0
    assert body["summary"]["pfv_new"] == 1
    db.expire_all()
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    assert len(rows) == 1
    assert rows[0].course_template_id is None


# G-51-5) 静的グリッド集計シートは importer に無視される (編集しても取り込まれない).
@pytest.mark.asyncio
async def test_static_grid_sheet_is_ignored_by_importer(client, db) -> None:
    admin = await _make_user(db, "g51-grid-ignore@example.com", "admin")
    office_inage, _office_tsuga, _cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-GRID", name="グリッド患者", primary_office_id=office_inage.id
    )
    pid = patient.id
    await _make_pfv(db, patient_id=pid, weekday=0, start_time=time(9, 0), duration_min=30)

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    wb = load_workbook(BytesIO(export_res.content))
    assert SHEET_PFV_GRID in wb.sheetnames
    wb[SHEET_PFV_GRID].cell(row=99, column=99, value="改ざんデータ")
    buf = BytesIO()
    wb.save(buf)
    files = {
        "file": (
            "tampered.xlsx",
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    import_res = await client.post(
        "/api/v1/patients/import-export/import?dry_run=true",
        headers=_bearer(admin),
        files=files,
    )
    assert import_res.status_code == 200, import_res.text
    summary = import_res.json()["summary"]
    assert summary["pfv_error"] == 0, summary
    assert summary["pfv_new"] == 0, summary
    assert summary["pfv_noop"] == 1, summary


# ---------------------------------------------------------------------------
# Phase G-55: 静的グリッド集計シートを「スケジュール枠組み（仮）」形式に再構築
# (稲毛先の縦ブロック / A1=月の日合計 / コース件数 / 30 分スロット配置)
# ---------------------------------------------------------------------------


def _grid_day_start_col(day_idx: int) -> int:
    """グリッドの日ブロック開始列 (1-indexed): col 1,6,11,16,21,26."""
    return day_idx * 5 + 1


# G-55-1) グリッドは稲毛(INAGE)→都賀(TSUGA) の順でブロックを並べ、各拠点内は A,B,..
@pytest.mark.asyncio
async def test_grid_blocks_inage_before_tsuga(client, db) -> None:
    admin = await _make_user(db, "g55-order@example.com", "admin")
    office_inage, office_tsuga, cts = await _setup_two_office_courses(db)
    # 稲毛: 月=稲B, 都賀: 木=津A.
    p_inage = await _make_patient(
        db, code="P-G55-IN", name="稲毛患者", primary_office_id=office_inage.id
    )
    p_tsuga = await _make_patient(
        db, code="P-G55-TS", name="都賀患者", primary_office_id=office_tsuga.id
    )
    await _make_pfv(
        db,
        patient_id=p_inage.id,
        weekday=0,
        start_time=time(11, 0),
        course_template_id=cts[("INAGE", "B")].id,
    )
    await _make_pfv(
        db,
        patient_id=p_tsuga.id,
        weekday=3,
        start_time=time(11, 0),
        course_template_id=cts[("TSUGA", "A")].id,
    )

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PFV_GRID]

    # 各ブロックの「コース文字」(row2 col2) を上から順に集める.
    # ブロック高 21 + 空き 1 = 22 行刻み. 拠点グループ間はさらに +1 行.
    # row1 col2 == "月曜日" のブロックを検出し、その次行 (row2) col2 = コース文字.
    course_chars: list[str] = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "月曜日":
            course_chars.append(ws.cell(row=r + 1, column=2).value)
    # course_templates は 稲B, 稲C, 津A. 稲毛先なので B, C, A の順.
    assert course_chars == ["B", "C", "A"], course_chars


# G-55-2) A1 = 月曜の日合計 (全コース合計). A1 にコメント (閲覧専用) が付く.
@pytest.mark.asyncio
async def test_grid_a1_is_monday_day_total_with_comment(client, db) -> None:
    admin = await _make_user(db, "g55-a1@example.com", "admin")
    office_inage, _office_tsuga, cts = await _setup_two_office_courses(db)
    # 月曜に 2 名 (稲B / 稲C), 火曜に 1 名 → A1 (月) = 2.
    for i, (lbl, t) in enumerate([("B", time(11, 0)), ("C", time(13, 0))]):
        p = await _make_patient(
            db, code=f"P-G55-MON{i}", name=f"月患者{i}", primary_office_id=office_inage.id
        )
        await _make_pfv(
            db,
            patient_id=p.id,
            weekday=0,
            start_time=t,
            course_template_id=cts[("INAGE", lbl)].id,
        )
    p_tue = await _make_patient(
        db, code="P-G55-TUE", name="火患者", primary_office_id=office_inage.id
    )
    await _make_pfv(
        db,
        patient_id=p_tue.id,
        weekday=1,
        start_time=time(11, 0),
        course_template_id=cts[("INAGE", "B")].id,
    )

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PFV_GRID]
    # A1 = 月曜の全コース合計 = 2.
    assert ws["A1"].value == 2
    # row1 各日 col1 = 日合計 (稲毛先頭ブロックのみ). 月=2, 火=1.
    assert ws.cell(row=1, column=_grid_day_start_col(0)).value == 2  # 月
    assert ws.cell(row=1, column=_grid_day_start_col(1)).value == 1  # 火
    # row1 各日 col2 = 曜日名.
    assert ws.cell(row=1, column=_grid_day_start_col(0) + 1).value == "月曜日"
    # A1 コメント (ホバー) が付く.
    assert ws["A1"].comment is not None
    assert "閲覧専用" in ws["A1"].comment.text


# G-55-3) コース件数 (row2 col1) = その (office×course×曜日) の患者数.
@pytest.mark.asyncio
async def test_grid_course_count_and_subheaders(client, db) -> None:
    admin = await _make_user(db, "g55-count@example.com", "admin")
    office_inage, _office_tsuga, cts = await _setup_two_office_courses(db)
    # 稲B 月曜に 2 名.
    for i in range(2):
        p = await _make_patient(
            db, code=f"P-G55-B{i}", name=f"B患者{i}", primary_office_id=office_inage.id
        )
        await _make_pfv(
            db,
            patient_id=p.id,
            weekday=0,
            start_time=time(11, 0) if i == 0 else time(13, 0),
            course_template_id=cts[("INAGE", "B")].id,
        )

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PFV_GRID]
    # 先頭ブロック = 稲B. row2 col1 (月のコース件数) = 2.
    assert ws.cell(row=2, column=_grid_day_start_col(0)).value == 2
    assert ws.cell(row=2, column=_grid_day_start_col(0) + 1).value == "B"
    # Phase G-57: row2 の住所列 (col+2) に拠点名 (稲毛/都賀) を表示し稲毛/都賀を判別可能に.
    assert ws.cell(row=2, column=_grid_day_start_col(0) + 2).value == "稲毛"
    # row3 = サブ見出し [時間帯, 氏名, 住所, 複数, 条件].
    assert [ws.cell(row=3, column=_grid_day_start_col(0) + k).value for k in range(5)] == [
        "時間帯",
        "氏名",
        "住所",
        "複数",
        "条件",
    ]


# G-57) 編集用シートのコース dropdown は稲毛(稲)先 → 都賀(津) 順.
@pytest.mark.asyncio
async def test_edit_sheet_course_dropdown_inage_first(client, db) -> None:
    admin = await _make_user(db, "g57-dropdown@example.com", "admin")
    await _setup_two_office_courses(db)
    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PFV_EDIT]
    # コース dropdown (拠点付きトークンを含む DataValidation) を探す.
    course_dv = None
    for dv in ws.data_validations.dataValidation:
        if dv.formula1 and "稲" in dv.formula1 and "津" in dv.formula1:
            course_dv = dv
            break
    assert course_dv is not None, "コース dropdown が見つからない"
    # 稲 (INAGE: 稲B,稲C) が 津 (TSUGA: 津A) より前に並ぶ.
    assert course_dv.formula1.index("稲B") < course_dv.formula1.index("津A")


# G-55-4) 患者は start_time の 30 分スロット行に配置され、住所/複数/条件 列も描画.
@pytest.mark.asyncio
async def test_grid_patients_placed_in_30min_slots(client, db) -> None:
    admin = await _make_user(db, "g55-slot@example.com", "admin")
    office_inage, _office_tsuga, cts = await _setup_two_office_courses(db)
    # 稲B 月曜 11:00 に 複数 + 女性のみ 患者.
    p = await _make_patient(
        db,
        code="P-G55-SLOT",
        name="スロット患者",
        primary_office_id=office_inage.id,
        address="千葉県千葉市稲毛区test99",
        requires_multiple_staff=True,
        sex_restriction="female_only",
    )
    await _make_pfv(
        db,
        patient_id=p.id,
        weekday=0,
        start_time=time(11, 0),
        course_template_id=cts[("INAGE", "B")].id,
    )

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb[SHEET_PFV_GRID]
    # スロット先頭行 = block row1(1) + 3 = row4 (09:30). 11:00 はスロット index 3 → row7.
    c0 = _grid_day_start_col(0)
    assert ws.cell(row=4, column=c0).value == "09:30"
    assert ws.cell(row=7, column=c0).value == "11:00"
    assert ws.cell(row=7, column=c0 + 1).value == "スロット患者"  # 氏名
    assert ws.cell(row=7, column=c0 + 2).value == "千葉県千葉市稲毛区test99"  # 住所
    assert ws.cell(row=7, column=c0 + 3).value == "複数"  # 複数
    assert ws.cell(row=7, column=c0 + 4).value == "女性のみ"  # 条件


# G-51-6) 旧 per-visit 形式「固定訪問パターン」シートは後方互換 fallback で取り込める.
@pytest.mark.asyncio
async def test_legacy_pervisit_pfv_sheet_fallback(client, db) -> None:
    admin = await _make_user(db, "g51-legacy@example.com", "admin")
    patient = await _make_patient(db, code="P-G51-LEG", name="旧形式患者")
    pid = patient.id
    content = _build_workbook_bytes(
        pfv_rows=[
            {
                "patient_id": str(pid),
                "weekday": "月",
                "slot_index": 0,
                "mode": "normal",
                "time_type": "固定",
                "start_time": "09:00",
                "duration_min": 30,
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["pfv_new"] == 1
    assert body["summary"]["pfv_error"] == 0
    db.expire_all()
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    assert len(rows) == 1
    assert rows[0].weekday == 0


# G-51-7) per-patient replace は他 mode (special) を壊さない.
@pytest.mark.asyncio
async def test_pfv_edit_replace_preserves_special_mode(client, db) -> None:
    admin = await _make_user(db, "g51-special@example.com", "admin")
    office_inage, _office_tsuga, _cts = await _setup_two_office_courses(db)
    patient = await _make_patient(
        db, code="P-G51-SP", name="特別保持", primary_office_id=office_inage.id
    )
    pid = patient.id
    await _make_pfv(db, patient_id=pid, weekday=0, start_time=time(9, 0), mode="normal")
    await _make_pfv(db, patient_id=pid, weekday=1, start_time=time(9, 0), mode="special")

    content = _build_edit_workbook_bytes(
        pfv_edit_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-G51-SP",
                "service_minutes": 30,
                "time_type": "固定",
                "mon_time": "09:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["pfv_error"] == 0
    db.expire_all()
    rows = (
        await db.scalars(select(PatientFixedVisit).where(PatientFixedVisit.patient_id == pid))
    ).all()
    modes = sorted((r.mode, r.weekday) for r in rows)
    assert modes == [("normal", 0), ("special", 1)]  # special 保持


# ---------------------------------------------------------------------------
# NG スタッフ (patient_ng_staff) Excel 往復
#
# セマンティクス (staff_excel の secondary_office_codes 列と同じ):
#   空セル = 維持 / <CLEAR> = 全解除 / カンマ区切り = その集合に一致させる.
#   不明コード・退職済みスタッフの新規指定は error にせず warning + skip.
# ---------------------------------------------------------------------------


async def _make_staff_for_ng(db, *, code: str, name: str, deleted: bool = False):
    """テスト用スタッフを作り **staff_id (UUID) を返す**.

    ORM オブジェクトではなく id を返すのは、後段で ``db.expire_all()`` を呼んだ
    あとに ``staff.id`` へ触れると同期 lazy load (MissingGreenlet) になるため.
    """
    s = Staff(code=code, name=name, status="active")
    if deleted:
        s.deleted_at = datetime.now(UTC)
    db.add(s)
    await db.commit()
    await db.refresh(s)
    return s.id


async def _add_ng(db, *, patient_id, staff_id, note: str | None = None) -> None:
    db.add(PatientNgStaff(patient_id=patient_id, staff_id=staff_id, note=note))
    await db.commit()


async def _ng_staff_ids(db, patient_id) -> set:
    """patient_ng_staff を DB から直接引く (column select なので identity map を介さない).

    ここで ``db.expire_all()`` は呼ばない — 呼び出し側が保持している ORM
    オブジェクト (Staff など) まで expire され、後続の属性アクセスが
    同期 lazy load となって MissingGreenlet になるため.
    """
    rows = (
        await db.scalars(
            select(PatientNgStaff.staff_id).where(PatientNgStaff.patient_id == patient_id)
        )
    ).all()
    return set(rows)


def _ng_cell(wb_bytes: bytes, row: int = 2):
    wb = load_workbook(BytesIO(wb_bytes))
    ws = wb[SHEET_PATIENTS]
    return ws.cell(row=row, column=PATIENT_COL_INDEX["ng_staff_codes"] + 1).value


# --- export ---------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_writes_ng_staff_codes_comma_joined_sorted(client, db) -> None:
    """NG スタッフありの患者 → コードを sorted して カンマ区切りで 1 セルに書く."""
    admin = await _make_user(db, "ng-ex-1@example.com", "admin")
    p = await _make_patient(db, code="P-NG-EX1", name="NGあり")
    # わざと辞書順と逆順に登録して、export が sorted することを担保する.
    s_b = await _make_staff_for_ng(db, code="NG-B", name="B さん")
    s_a = await _make_staff_for_ng(db, code="NG-A", name="A さん")
    await _add_ng(db, patient_id=p.id, staff_id=s_b)
    await _add_ng(db, patient_id=p.id, staff_id=s_a)

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    assert _ng_cell(res.content) == "NG-A,NG-B"


@pytest.mark.asyncio
async def test_export_writes_blank_ng_staff_codes_when_none(client, db) -> None:
    """NG スタッフ無しの患者 → 空セル (= 再取込で「維持」)."""
    admin = await _make_user(db, "ng-ex-2@example.com", "admin")
    await _make_patient(db, code="P-NG-EX2", name="NGなし")

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    assert _ng_cell(res.content) is None


@pytest.mark.asyncio
async def test_export_includes_retired_staff_in_ng_column(client, db) -> None:
    """退職済みスタッフの既存 NG 行も export される (落とすと round-trip で暗黙解除される)."""
    admin = await _make_user(db, "ng-ex-3@example.com", "admin")
    p = await _make_patient(db, code="P-NG-EX3", name="退職者NG")
    s_alive = await _make_staff_for_ng(db, code="NG-ALIVE", name="現役")
    s_gone = await _make_staff_for_ng(db, code="NG-GONE", name="退職", deleted=True)
    await _add_ng(db, patient_id=p.id, staff_id=s_alive)
    await _add_ng(db, patient_id=p.id, staff_id=s_gone)

    res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert res.status_code == 200
    assert _ng_cell(res.content) == "NG-ALIVE,NG-GONE"


# --- import (差分) --------------------------------------------------------


@pytest.mark.asyncio
async def test_import_sets_ng_staff_on_existing_patient(client, db) -> None:
    """既存患者に NG スタッフを新規設定 (noop ではなく update として検出される)."""
    admin = await _make_user(db, "ng-im-1@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM1", name="設定対象")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    s2 = await _make_staff_for_ng(db, code="NG-S2", name="S2")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM1",
                "ng_staff_codes": "NG-S1,NG-S2",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_update"] == 1, body
    # diff は field="ng_staff" でコードリスト表示.
    change = next(c for c in body["patient_rows"][0]["changes"] if c["field"] == "ng_staff")
    assert change["old_value"] == []
    assert change["new_value"] == ["NG-S1", "NG-S2"]
    assert await _ng_staff_ids(db, pid) == {s1, s2}


@pytest.mark.asyncio
async def test_import_ng_staff_matches_set_add_and_remove(client, db) -> None:
    """カンマ区切りの集合に一致させる = 追加と削除が同時に起きる."""
    admin = await _make_user(db, "ng-im-2@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM2", name="集合一致")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    s2 = await _make_staff_for_ng(db, code="NG-S2", name="S2")
    s3 = await _make_staff_for_ng(db, code="NG-S3", name="S3")
    await _add_ng(db, patient_id=pid, staff_id=s1)
    await _add_ng(db, patient_id=pid, staff_id=s2)

    # S1 を残し / S2 を消し / S3 を足す.
    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM2",
                "ng_staff_codes": "NG-S1,NG-S3",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert await _ng_staff_ids(db, pid) == {s1, s3}  # s2 は消える


@pytest.mark.asyncio
async def test_import_ng_staff_clear_marker_removes_all(client, db) -> None:
    """<CLEAR> で NG スタッフを全解除."""
    admin = await _make_user(db, "ng-im-3@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM3", name="全解除")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    await _add_ng(db, patient_id=pid, staff_id=s1)

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM3",
                "ng_staff_codes": "<CLEAR>",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert res.json()["summary"]["patients_update"] == 1
    assert await _ng_staff_ids(db, pid) == set()


@pytest.mark.asyncio
async def test_import_blank_ng_staff_preserves_existing(client, db) -> None:
    """空セル = 維持. 他列だけ更新しても NG は消えない."""
    admin = await _make_user(db, "ng-im-4@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM4", name="維持対象")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    await _add_ng(db, patient_id=pid, staff_id=s1)

    content = _build_workbook_bytes(
        patient_rows=[
            {"patient_id": str(pid), "patient_code": "P-NG-IM4", "name": "新名前"},
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    assert p_after.name == "新名前"
    assert await _ng_staff_ids(db, pid) == {s1}  # 維持


@pytest.mark.asyncio
async def test_import_ng_staff_unknown_code_warns_and_skips(client, db) -> None:
    """不明コードは error にせず skip. 既知コードだけが反映される."""
    admin = await _make_user(db, "ng-im-5@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM5", name="不明コード")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM5",
                "ng_staff_codes": "NG-S1,NG-NOPE",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert res.json()["summary"]["patients_error"] == 0
    assert await _ng_staff_ids(db, pid) == {s1}


@pytest.mark.asyncio
async def test_import_ng_staff_retired_new_assignment_skipped(client, db) -> None:
    """退職済みスタッフの **新規** 指定は warning + skip (CRUD API の 422 と整合)."""
    admin = await _make_user(db, "ng-im-6@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM6", name="退職者指定")
    pid = p.id
    s_alive = await _make_staff_for_ng(db, code="NG-S1", name="現役")
    await _make_staff_for_ng(db, code="NG-GONE", name="退職", deleted=True)

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM6",
                "ng_staff_codes": "NG-S1,NG-GONE",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert res.json()["summary"]["patients_error"] == 0
    assert await _ng_staff_ids(db, pid) == {s_alive}  # 退職者は入らない


@pytest.mark.asyncio
async def test_import_ng_staff_retired_existing_row_is_kept(client, db) -> None:
    """退職済みでも **既存** NG 行はそのまま維持される (round-trip で暗黙解除しない)."""
    admin = await _make_user(db, "ng-im-7@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM7", name="退職者維持")
    pid = p.id
    s_gone = await _make_staff_for_ng(db, code="NG-GONE", name="退職", deleted=True)
    await _add_ng(db, patient_id=pid, staff_id=s_gone)

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM7",
                "ng_staff_codes": "NG-GONE",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    # 集合が既に一致しているので noop.
    assert res.json()["summary"]["patients_noop"] == 1, res.text
    assert await _ng_staff_ids(db, pid) == {s_gone}


@pytest.mark.asyncio
async def test_import_ng_staff_keeps_note_on_surviving_row(client, db) -> None:
    """Excel は note を扱わない: 集合に残るスタッフの note は潰さない."""
    admin = await _make_user(db, "ng-im-8@example.com", "admin")
    p = await _make_patient(db, code="P-NG-IM8", name="note維持")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    s2 = await _make_staff_for_ng(db, code="NG-S2", name="S2")
    await _add_ng(db, patient_id=pid, staff_id=s1, note="相性不良")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-IM8",
                "ng_staff_codes": "NG-S1,NG-S2",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    rows = (await db.scalars(select(PatientNgStaff).where(PatientNgStaff.patient_id == pid))).all()
    by_staff = {r.staff_id: r for r in rows}
    assert by_staff[s1].note == "相性不良"  # 既存行は維持
    assert by_staff[s2].note is None  # 新規行は note=NULL
    assert by_staff[s2].decided_by_user_id is None  # Excel 経由は設定者を記録しない


@pytest.mark.asyncio
async def test_import_new_patient_with_ng_staff(client, db) -> None:
    """新規患者行でも NG スタッフを設定できる (仮 UUID で INSERT 後に紐付け)."""
    admin = await _make_user(db, "ng-im-9@example.com", "admin")
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")

    content = _build_workbook_bytes(
        patient_rows=[
            {
                "patient_code": "P-NG-NEW",
                "name": "新規NGあり",
                "sex": "female",
                "status": "active",
                "address": "千葉市稲毛区",
                "ng_staff_codes": "NG-S1",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    db.expire_all()
    p = (await db.scalars(select(Patient).where(Patient.code == "P-NG-NEW"))).first()
    assert p is not None
    assert await _ng_staff_ids(db, p.id) == {s1}


@pytest.mark.asyncio
async def test_import_dry_run_does_not_touch_ng_staff(client, db) -> None:
    """dry_run では NG 行が変わらない."""
    admin = await _make_user(db, "ng-im-10@example.com", "admin")
    p = await _make_patient(db, code="P-NG-DRY", name="dry")
    pid = p.id
    await _make_staff_for_ng(db, code="NG-S1", name="S1")

    content = _build_workbook_bytes(
        patient_rows=[
            {"patient_id": str(pid), "patient_code": "P-NG-DRY", "ng_staff_codes": "NG-S1"},
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    assert res.json()["summary"]["patients_update"] == 1
    assert await _ng_staff_ids(db, pid) == set()  # DB は無変更


# --- round-trip -----------------------------------------------------------


@pytest.mark.asyncio
async def test_roundtrip_ng_staff_is_noop(client, db) -> None:
    """export → そのまま import で NG スタッフは完全 noop (変更検出されない)."""
    admin = await _make_user(db, "ng-rt-1@example.com", "admin")
    p = await _make_patient(db, code="P-NG-RT", name="往復")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    s2 = await _make_staff_for_ng(db, code="NG-S2", name="S2")
    await _add_ng(db, patient_id=pid, staff_id=s1)
    await _add_ng(db, patient_id=pid, staff_id=s2)

    export_res = await client.get("/api/v1/patients/import-export/export", headers=_bearer(admin))
    assert export_res.status_code == 200
    res = await _upload(client, admin, content=export_res.content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["summary"]["patients_error"] == 0, body
    assert body["summary"]["patients_update"] == 0, body
    assert body["summary"]["patients_noop"] == 1, body
    assert await _ng_staff_ids(db, pid) == {s1, s2}


# --- 削除経路の中間テーブル掃除 -------------------------------------------


@pytest.mark.asyncio
async def test_import_delete_flag_cleans_ng_staff_and_same_address_links(client, db) -> None:
    """<DELETE> 経由の soft delete でも NG 行 / 同住所リンクが物理削除される.

    soft delete では FK ON DELETE CASCADE が発火しないため、アプリ層で明示 DELETE
    しないと「同じ code で復活したときに古い紐付けが蘇る」既知罠にハマる.
    """
    admin = await _make_user(db, "ng-del-1@example.com", "admin")
    # patient_same_address_links には CHECK (patient_a_id < patient_b_id) があるため、
    # 「target が a 側」「target が b 側」の 2 本を張るには UUID の大小を固定する必要が
    # ある. Patient.id は Python 側 default (uuid4) なので明示指定できる.
    # 数字のみの UUID は sqlite の NUMERIC affinity で float に化けるため英字を混ぜる.
    low_id = UUID("aaaaaaaa-0000-4000-8000-00000000000a")
    target_id = UUID("bbbbbbbb-0000-4000-8000-00000000000b")
    high_id = UUID("cccccccc-0000-4000-8000-00000000000c")
    await _make_patient(db, id=low_id, code="P-NG-LOW", name="相方低")
    await _make_patient(db, id=target_id, code="P-NG-DEL", name="削除対象")
    await _make_patient(db, id=high_id, code="P-NG-HIGH", name="相方高")
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    await _add_ng(db, patient_id=target_id, staff_id=s1)
    await _add_ng(db, patient_id=low_id, staff_id=s1)
    # 同住所リンクを a 側 / b 側の両方で 1 本ずつ張る.
    db.add(PatientSameAddressLink(patient_a_id=low_id, patient_b_id=target_id, pair_mode="blocked"))
    db.add(
        PatientSameAddressLink(patient_a_id=target_id, patient_b_id=high_id, pair_mode="preferred")
    )
    await db.commit()

    content = _build_workbook_bytes(
        patient_rows=[
            {"patient_id": str(target_id), "patient_code": "P-NG-DEL", "delete_flag": "<DELETE>"},
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert res.json()["summary"]["patients_delete"] == 1

    db.expire_all()
    assert await _ng_staff_ids(db, target_id) == set()  # 掃除された
    assert await _ng_staff_ids(db, low_id) == {s1}  # 巻き添えにしない
    links = (
        await db.scalars(
            select(PatientSameAddressLink).where(
                or_(
                    PatientSameAddressLink.patient_a_id == target_id,
                    PatientSameAddressLink.patient_b_id == target_id,
                )
            )
        )
    ).all()
    assert list(links) == []


# --- 後方互換 (新列を持たない旧 export ファイル) ---------------------------


def _build_legacy_workbook_bytes_without_ng_column(*, patient_rows: list[dict]) -> bytes:
    """NG スタッフ列を持たない旧形式 (delete_flag までの列数) の workbook を作る."""
    legacy_columns = [c for c in PATIENT_COLUMNS if c["key"] != "ng_staff_codes"]
    assert str(legacy_columns[-1]["key"]) == "delete_flag"  # 新列は必ず末尾である前提
    legacy_index = {str(c["key"]): i for i, c in enumerate(legacy_columns)}
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    for col_idx, col in enumerate(legacy_columns, start=1):
        ws_p.cell(row=1, column=col_idx, value=str(col["header"]))
    for r_idx, row_dict in enumerate(patient_rows, start=2):
        for col_key, idx in legacy_index.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_p.cell(row=r_idx, column=idx + 1, value=v)
    ws_f = wb.create_sheet(title=SHEET_PFV)
    for col_idx, header in enumerate(_pfv_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_import_legacy_file_without_ng_column_still_works(client, db) -> None:
    """新列を持たない旧 export ファイル: 従来どおり更新でき、NG は維持される.

    列位置ベースで読むため、新列を末尾以外に足すと旧ファイルが silent 破壊される.
    このテストは「末尾追加」の不変条件を守るためのガード.
    """
    admin = await _make_user(db, "ng-legacy-1@example.com", "admin")
    p = await _make_patient(db, code="P-NG-LEGACY", name="旧形式", note="元の備考")
    pid = p.id
    s1 = await _make_staff_for_ng(db, code="NG-S1", name="S1")
    await _add_ng(db, patient_id=pid, staff_id=s1)

    content = _build_legacy_workbook_bytes_without_ng_column(
        patient_rows=[
            {
                "patient_id": str(pid),
                "patient_code": "P-NG-LEGACY",
                "name": "旧形式-改",
                "note": "新しい備考",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    assert res.json()["summary"]["patients_error"] == 0, res.text
    db.expire_all()
    p_after = (await db.scalars(select(Patient).where(Patient.id == pid))).first()
    assert p_after is not None
    # 列ズレが起きていれば name / note が壊れる.
    assert p_after.name == "旧形式-改"
    assert p_after.note == "新しい備考"
    assert await _ng_staff_ids(db, pid) == {s1}  # 列が無い = 維持
