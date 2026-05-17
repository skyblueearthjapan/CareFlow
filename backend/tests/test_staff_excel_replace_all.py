"""Tests for /api/v1/staff/import-export/replace-all (完全置換).

Coverage:
  1. RBAC: admin 200 / manager 403 / staff 403
  2. dry_run: DB 変更なし、削除対象一覧が返る
  3. apply: Excel に無い既存スタッフが soft delete される (関連 shift は物理削除)
  4. apply: 空セルが NULL で上書きされる
  5. apply: 既存 shift が全件物理削除されて Excel から再投入される
  6. apply: error 1 件で全 rollback (422)
"""

from __future__ import annotations

from datetime import time
from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.core.security import create_access_token, hash_password
from app.models import (
    Office,
    Staff,
    StaffShift,
    User,
)
from app.services.staff_excel.schema import (
    SHEET_SHIFT,
    SHEET_STAFF,
    SHIFT_COL_INDEX,
    SHIFT_COLUMNS,
    STAFF_COL_INDEX,
    STAFF_COLUMNS,
)


async def _make_user(db, email: str, role: str) -> User:
    user = User(email=email, password_hash=hash_password("pw"), role=role)
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


def _bearer(user: User) -> dict[str, str]:
    token = create_access_token(subject=user.id, role=user.role, staff_id=user.staff_id)
    return {"Authorization": f"Bearer {token}"}


async def _make_office(db, code: str, name: str | None = None) -> Office:
    o = Office(code=code, name=name or code)
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return o


async def _make_staff(db, **overrides) -> Staff:
    defaults = {
        "code": "S-DEFAULT",
        "name": "デフォルト",
        "status": "active",
        "role": "staff",
    }
    defaults.update(overrides)
    s = Staff(**defaults)
    db.add(s)
    await db.commit()
    await db.refresh(s)
    return s


async def _make_shift(db, *, staff_id, weekday: int = 0):
    sh = StaffShift(
        staff_id=staff_id,
        weekday=weekday,
        is_on=True,
        start_time=time(9, 0),
        end_time=time(18, 0),
    )
    db.add(sh)
    await db.commit()
    await db.refresh(sh)
    return sh


def _staff_headers() -> list[str]:
    return [str(c["header"]) for c in STAFF_COLUMNS]


def _shift_headers() -> list[str]:
    return [str(c["header"]) for c in SHIFT_COLUMNS]


def _build_workbook_bytes(
    *,
    staff_rows: list[dict] | None = None,
    shift_rows: list[dict] | None = None,
) -> bytes:
    wb = Workbook()
    ws_s = wb.active
    ws_s.title = SHEET_STAFF
    for col_idx, header in enumerate(_staff_headers(), start=1):
        ws_s.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(staff_rows or [], start=2):
        for col_key, idx in STAFF_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_s.cell(row=r_idx, column=idx + 1, value=v)
    ws_f = wb.create_sheet(title=SHEET_SHIFT)
    for col_idx, header in enumerate(_shift_headers(), start=1):
        ws_f.cell(row=1, column=col_idx, value=header)
    for r_idx, row_dict in enumerate(shift_rows or [], start=2):
        for col_key, idx in SHIFT_COL_INDEX.items():
            v = row_dict.get(col_key)
            if v is not None:
                ws_f.cell(row=r_idx, column=idx + 1, value=v)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload(client, user, *, content: bytes, dry_run: bool = True):
    files = {
        "file": (
            "test.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    return await client.post(
        f"/api/v1/staff/import-export/replace-all?dry_run={'true' if dry_run else 'false'}",
        headers=_bearer(user),
        files=files,
    )


# ---------------------------------------------------------------------------
# RBAC
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_admin_ok(client, db) -> None:
    admin = await _make_user(db, "sra-admin@example.com", "admin")
    content = _build_workbook_bytes(staff_rows=[])
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text


@pytest.mark.asyncio
async def test_staff_replace_all_manager_forbidden(client, db) -> None:
    manager = await _make_user(db, "sra-mgr@example.com", "manager")
    content = _build_workbook_bytes(staff_rows=[])
    res = await _upload(client, manager, content=content, dry_run=True)
    assert res.status_code == 403


@pytest.mark.asyncio
async def test_staff_replace_all_staff_forbidden(client, db) -> None:
    staff_user = await _make_user(db, "sra-staff@example.com", "staff")
    content = _build_workbook_bytes(staff_rows=[])
    res = await _upload(client, staff_user, content=content, dry_run=True)
    assert res.status_code == 403


# ---------------------------------------------------------------------------
# dry_run
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_dry_run_returns_soft_delete_list(client, db) -> None:
    admin = await _make_user(db, "sra-dry@example.com", "admin")
    s_keep = await _make_staff(db, code="S-RA-K", name="残す")
    s_keep_id = s_keep.id
    s_drop = await _make_staff(db, code="S-RA-D", name="消える")
    s_drop_id = s_drop.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(s_keep_id),
                "staff_code": "S-RA-K",
                "name": "残す",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=True)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is False
    assert body["summary"]["staff_to_soft_delete"] == 1
    assert len(body["staff_to_soft_delete_preview"]) == 1
    assert body["staff_to_soft_delete_preview"][0]["staff_code"] == "S-RA-D"

    db.expire_all()
    s_drop_after = await db.get(Staff, s_drop_id)
    assert s_drop_after is not None
    assert s_drop_after.deleted_at is None  # DB 変更なし


# ---------------------------------------------------------------------------
# apply: soft delete + shift hard delete
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_apply_soft_deletes_missing_staff(client, db) -> None:
    admin = await _make_user(db, "sra-del@example.com", "admin")
    s_keep = await _make_staff(db, code="S-RA-K1", name="残す")
    s_keep_id = s_keep.id
    s_drop = await _make_staff(db, code="S-RA-X1", name="消える")
    s_drop_id = s_drop.id
    await _make_shift(db, staff_id=s_drop_id, weekday=0)

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(s_keep_id),
                "staff_code": "S-RA-K1",
                "name": "残す",
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True

    db.expire_all()
    s_drop_after = await db.get(Staff, s_drop_id)
    assert s_drop_after.deleted_at is not None
    # 関連 shift は物理削除
    remaining = (await db.scalars(select(StaffShift).where(StaffShift.staff_id == s_drop_id))).all()
    assert remaining == []


# ---------------------------------------------------------------------------
# apply: 空セル → NULL
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_apply_blank_cells_overwrite_to_null(client, db) -> None:
    admin = await _make_user(db, "sra-null@example.com", "admin")
    s = await _make_staff(
        db,
        code="S-RA-NUL",
        name="保持",
        kana="ホジ",
        note="残ってる備考",
        status="active",
        role="staff",
    )
    sid = s.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-RA-NUL",
                "name": "保持",
                "status": "active",
                "role": "staff",
                # kana / note 空セル
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text

    db.expire_all()
    after = await db.get(Staff, sid)
    assert after.kana is None
    assert after.note is None


# ---------------------------------------------------------------------------
# apply: shift 全件 replace
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_apply_shifts_replace_all(client, db) -> None:
    admin = await _make_user(db, "sra-sh@example.com", "admin")
    s = await _make_staff(db, code="S-RA-SH", name="シフトテスト")
    sid = s.id
    sh1 = await _make_shift(db, staff_id=sid, weekday=0)
    sh1_key = (sid, sh1.weekday)
    sh2 = await _make_shift(db, staff_id=sid, weekday=1)

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-RA-SH",
                "name": "シフトテスト",
                "status": "active",
                "role": "staff",
            }
        ],
        shift_rows=[
            {
                "staff_id": str(sid),
                "weekday": "水",
                "is_on": "TRUE",
                "start_time": "10:00",
                "end_time": "19:00",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["transaction_applied"] is True
    assert body["summary"]["shift_to_create"] == 1
    assert body["summary"]["shift_to_replace"] == 2

    db.expire_all()
    rows = (await db.scalars(select(StaffShift).where(StaffShift.staff_id == sid))).all()
    assert len(rows) == 1
    assert rows[0].weekday == 2  # 水
    assert rows[0].start_time == time(10, 0)
    # 旧 shift (月/火) は無いこと
    _ = sh1_key, sh2  # silence unused warnings


# ---------------------------------------------------------------------------
# apply: error → rollback (422)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_apply_error_rolls_back(client, db) -> None:
    admin = await _make_user(db, "sra-rb@example.com", "admin")
    s = await _make_staff(db, code="S-RA-RB", name="ロールバック前", status="active", role="staff")
    sid = s.id

    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-RA-RB",
                "name": "更新後",
                "status": "active",
                "role": "staff",
            },
            {
                "staff_code": "S-RA-BAD",
                "name": "エラー",
                "sex": "INVALID",  # enum 違反
                "status": "active",
                "role": "staff",
            },
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    assert res.status_code == 422, res.text
    db.expire_all()
    after = await db.get(Staff, sid)
    assert after.name == "ロールバック前"  # 変更なし


# ---------------------------------------------------------------------------
# CRITICAL 回帰: update 行の NOT NULL カラムが空セル → error (rollback)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_staff_replace_all_required_field_null_in_update_row_errors(client, db) -> None:
    """update パスで NOT NULL カラム (name/status/role) が空セルの場合、
    error として扱われ全件 rollback (422) される."""
    admin = await _make_user(db, "sra-null-update@example.com", "admin")
    s = await _make_staff(
        db, code="S-RA-NUL-UPD", name="既存スタッフ", status="active", role="staff"
    )
    sid = s.id

    # staff_id 明示 (update パス) + name を空セルに
    content = _build_workbook_bytes(
        staff_rows=[
            {
                "staff_id": str(sid),
                "staff_code": "S-RA-NUL-UPD",
                # name 欠落 (空セル = None)
                "status": "active",
                "role": "staff",
            }
        ],
    )
    res = await _upload(client, admin, content=content, dry_run=False)
    # error 行があるので 422
    assert res.status_code == 422, res.text

    # DB は変更されていないこと
    db.expire_all()
    after = await db.get(Staff, sid)
    assert after is not None
    assert after.name == "既存スタッフ"  # rollback 確認
