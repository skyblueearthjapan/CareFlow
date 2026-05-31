"""Phase E-5 (項目 ⑥B): patient_fixed_visits.sub_office_id tests.

検証観点:
  1. PFV 作成時に sub_office_id を受理.
  2. 既存 PFV に sub_office を後付け設定 (PUT).
  3. sub_office_id と course_template の office 一致検証 (422).
  4. diff_add pool 展開で sub_office 経由患者が候補化される.
  5. full_optimize は sub_office を見ない (= 既存挙動維持).
  6. Excel export に sub_office_code 列を含む.
  7. Excel import で sub_office_code を解決して sub_office_id を埋める.
  8. Excel round-trip (export → import) で 0 エラー.
"""

from __future__ import annotations

from datetime import date, time
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.security import create_access_token, hash_password
from app.models import (
    CourseTemplate,
    Office,
    Patient,
    PatientFixedVisit,
    User,
)
from app.services.patient_excel.exporter import build_workbook, workbook_to_bytes
from app.services.patient_excel.importer import parse_and_diff
from app.services.patient_excel.schema import PFV_COL_INDEX, SHEET_PFV
from app.services.scheduling.auto_allocator_v2 import run_v2_pipeline

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


async def _make_user(db, email: str, role: str) -> User:
    u = User(email=email, password_hash=hash_password("pw"), role=role)
    db.add(u)
    await db.commit()
    await db.refresh(u)
    return u


def _bearer(user: User) -> dict[str, str]:
    token = create_access_token(subject=user.id, role=user.role, staff_id=user.staff_id)
    return {"Authorization": f"Bearer {token}"}


async def _make_office(db, *, code: str, name: str | None = None) -> Office:
    o = Office(code=code, name=name or code)
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return o


async def _make_patient(
    db,
    *,
    code: str,
    primary_office_id=None,
    lat: float | None = None,
    lng: float | None = None,
) -> Patient:
    p = Patient(
        code=code,
        name=f"患者{code}",
        special_week_active=[],
        primary_office_id=primary_office_id,
        sex="male",
        status="active",
        lat=lat,
        lng=lng,
    )
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


async def _make_course_template(db, *, office_id, label: str) -> CourseTemplate:
    ct = CourseTemplate(office_id=office_id, label=label)
    db.add(ct)
    await db.commit()
    await db.refresh(ct)
    return ct


# ---------------------------------------------------------------------------
# 1. PFV create with sub_office_id (PUT)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_create_with_sub_office_id(client, db) -> None:
    """PUT /patients/{id}/fixed-visits で sub_office_id を受理して保存できる."""
    admin = await _make_user(db, "pfv-subof-create@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    patient = await _make_patient(db, code="P-SUB-1", primary_office_id=inage.id)

    body = {
        "mode": "normal",
        "items": [
            {
                "weekday": 0,
                "start_time": "09:00",
                "duration_min": 30,
                "sub_office_id": str(tsuga.id),
            }
        ],
    }
    res = await client.put(
        f"/api/v1/patients/{patient.id}/fixed-visits",
        headers=_bearer(admin),
        json=body,
    )
    assert res.status_code == 200, res.text
    data = res.json()
    assert len(data) == 1
    assert data[0]["sub_office_id"] == str(tsuga.id)


# ---------------------------------------------------------------------------
# 2. Update existing PFV to add sub_office
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_update_set_sub_office_id(client, db) -> None:
    """既存 PFV (sub_office_id=NULL) を PUT で更新して sub_office_id をセットできる."""
    admin = await _make_user(db, "pfv-subof-update@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    patient = await _make_patient(db, code="P-SUB-2", primary_office_id=inage.id)

    # 1 回目: sub_office なし
    res = await client.put(
        f"/api/v1/patients/{patient.id}/fixed-visits",
        headers=_bearer(admin),
        json={
            "mode": "normal",
            "items": [{"weekday": 1, "start_time": "10:00", "duration_min": 30}],
        },
    )
    assert res.status_code == 200, res.text
    assert res.json()[0]["sub_office_id"] is None

    # 2 回目: sub_office_id を追加
    res = await client.put(
        f"/api/v1/patients/{patient.id}/fixed-visits",
        headers=_bearer(admin),
        json={
            "mode": "normal",
            "items": [
                {
                    "weekday": 1,
                    "start_time": "10:00",
                    "duration_min": 30,
                    "sub_office_id": str(tsuga.id),
                }
            ],
        },
    )
    assert res.status_code == 200, res.text
    assert res.json()[0]["sub_office_id"] == str(tsuga.id)


# ---------------------------------------------------------------------------
# 3. sub_office_id × course_template_id の office 一致 検証
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_sub_office_validates_course_template_office(client, db) -> None:
    """sub_office_id 指定時に course_template が別 office を指していたら 422."""
    admin = await _make_user(db, "pfv-subof-valid@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    patient = await _make_patient(db, code="P-SUB-3", primary_office_id=inage.id)
    # 稲毛の course template
    ct_inage_a = await _make_course_template(db, office_id=inage.id, label="A")

    # sub_office_id=tsuga なのに course_template が稲毛 → 422
    res = await client.put(
        f"/api/v1/patients/{patient.id}/fixed-visits",
        headers=_bearer(admin),
        json={
            "mode": "normal",
            "items": [
                {
                    "weekday": 0,
                    "start_time": "09:00",
                    "duration_min": 30,
                    "sub_office_id": str(tsuga.id),
                    "course_template_id": str(ct_inage_a.id),
                }
            ],
        },
    )
    assert res.status_code == 422, res.text

    # 同じ tsuga の course なら OK
    ct_tsuga_a = await _make_course_template(db, office_id=tsuga.id, label="A")
    res = await client.put(
        f"/api/v1/patients/{patient.id}/fixed-visits",
        headers=_bearer(admin),
        json={
            "mode": "normal",
            "items": [
                {
                    "weekday": 0,
                    "start_time": "09:00",
                    "duration_min": 30,
                    "sub_office_id": str(tsuga.id),
                    "course_template_id": str(ct_tsuga_a.id),
                }
            ],
        },
    )
    assert res.status_code == 200, res.text


@pytest.mark.asyncio
async def test_pfv_sub_office_id_unknown_office_422(client, db) -> None:
    """存在しない sub_office_id を渡すと 422."""
    import uuid as _uuid

    admin = await _make_user(db, "pfv-subof-unknown@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    patient = await _make_patient(db, code="P-SUB-UNK", primary_office_id=inage.id)

    res = await client.put(
        f"/api/v1/patients/{patient.id}/fixed-visits",
        headers=_bearer(admin),
        json={
            "mode": "normal",
            "items": [
                {
                    "weekday": 0,
                    "start_time": "09:00",
                    "duration_min": 30,
                    "sub_office_id": str(_uuid.uuid4()),
                }
            ],
        },
    )
    assert res.status_code == 422, res.text


# ---------------------------------------------------------------------------
# 4. Auto-allocator: pool 展開で sub_office 経由患者を候補化 (diff_add)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pool_expansion_includes_sub_office_patients(client, db) -> None:
    """diff_add で sub_office_id 経由のフォロー患者が pool に登場する.

    稲毛が主担当の患者 P が PFV.sub_office_id=都賀 を持つ場合、
    都賀を対象に run_v2_pipeline(diff_add) を呼ぶと P の visit が pool_visits に
    現れる (office_id=都賀 で配置候補化).
    """
    _admin = await _make_user(db, "pool-sub-test@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    # 稲毛主担当の患者
    patient = await _make_patient(
        db,
        code="P-POOL-SUB",
        primary_office_id=inage.id,
        lat=35.65,
        lng=140.10,
    )
    # PFV: 月曜 09:00, sub_office_id=都賀
    pfv = PatientFixedVisit(
        patient_id=patient.id,
        mode="normal",
        weekday=0,
        slot_index=0,
        start_time=time(9, 0),
        duration_min=30,
        sub_office_id=tsuga.id,
    )
    db.add(pfv)
    await db.commit()

    # 都賀のみ scope で diff_add 実行
    iso_year, iso_week, _ = date(2026, 6, 1).isocalendar()
    result = await run_v2_pipeline(
        db,
        iso_year=iso_year,
        iso_week=iso_week,
        office_ids=[tsuga.id],
        mode="diff_add",
    )
    pool_visits = result["pool_visits"]
    matching = [v for v in pool_visits if v.patient_id == patient.id]
    assert len(matching) >= 1, (
        f"sub_office 経由患者の pool visit が出ない (pool_visits={len(pool_visits)})"
    )
    # office_id は tsuga にリルートされているはず
    assert all(v.office_id == tsuga.id for v in matching), (
        f"matching visit の office_id が tsuga でない: {[v.office_id for v in matching]}"
    )


# ---------------------------------------------------------------------------
# 5. full_optimize は sub_office を見ない (既存挙動維持)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_auto_allocator_v2_does_not_use_sub_office_for_full_optimize(client, db) -> None:
    """full_optimize モードでは sub_office_id 経由の patient を別 office に流さない.

    都賀のみ scope で full_optimize を呼んだ場合、稲毛主担当の患者は scope 外
    なので含まれない (sub_office_id があっても fallback 経路は通らない).
    """
    _admin = await _make_user(db, "full-opt-sub@example.com", "admin")
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    patient = await _make_patient(
        db,
        code="P-FOPT-SUB",
        primary_office_id=inage.id,
        lat=35.65,
        lng=140.10,
    )
    pfv = PatientFixedVisit(
        patient_id=patient.id,
        mode="normal",
        weekday=0,
        slot_index=0,
        start_time=time(9, 0),
        duration_min=30,
        sub_office_id=tsuga.id,
    )
    db.add(pfv)
    await db.commit()

    iso_year, iso_week, _ = date(2026, 6, 1).isocalendar()
    result = await run_v2_pipeline(
        db,
        iso_year=iso_year,
        iso_week=iso_week,
        office_ids=[tsuga.id],
        mode="full_optimize",
    )
    # full_optimize は sub_office を読まないので、稲毛主担当の患者は visits に出ない
    after_visits = result["after_visits"]
    matching = [v for v in after_visits if v.patient_id == patient.id]
    assert len(matching) == 0, (
        f"full_optimize で sub_office 経由患者が流入してはいけない: {len(matching)} 件"
    )


# ---------------------------------------------------------------------------
# 6. Phase G-51: 「編集用」シート (患者 1 行) — sub_office_id 単体は列で表現しないが、
#    クロス拠点コースはトークンで表現される.
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_excel_export_includes_sub_office(client, db) -> None:
    """Phase G-51: export はクロス拠点 course を拠点付きトークン (例 "津A") で表現する.

    旧 per-visit シートの ``sub_office_code`` 列は廃止. course_template が cross-office
    の場合、その course は token に拠点が含まれる形 ("津A") で「編集用」シートに書かれる.
    """
    from app.services.patient_excel.schema import PFV_EDIT_COL_INDEX, SHEET_PFV_EDIT

    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    # TSUGA の course_template (クロス拠点 course を token で表現するため).
    ct_tsuga = CourseTemplate(office_id=tsuga.id, label="A")
    db.add(ct_tsuga)
    await db.commit()
    await db.refresh(ct_tsuga)
    patient = await _make_patient(
        db,
        code="P-EXP-SUB",
        primary_office_id=inage.id,
        lat=35.65,
        lng=140.10,
    )
    pfv = PatientFixedVisit(
        patient_id=patient.id,
        mode="normal",
        weekday=0,
        slot_index=0,
        start_time=time(9, 0),
        duration_min=30,
        course_template_id=ct_tsuga.id,
        sub_office_id=tsuga.id,
    )
    db.add(pfv)
    await db.commit()

    patients_rows = (await db.scalars(select(Patient))).all()
    pfv_rows = (await db.scalars(select(PatientFixedVisit))).all()
    office_rows = (await db.scalars(select(Office))).all()
    ct_rows = (await db.scalars(select(CourseTemplate))).all()

    wb = build_workbook(
        patients=patients_rows,
        pfvs=pfv_rows,
        offices=office_rows,
        course_templates=ct_rows,
    )
    raw = workbook_to_bytes(wb)

    # 「編集用」シート (患者 1 行) に月曜の course が "津A" (クロス拠点トークン) で書かれる.
    wb2 = load_workbook(BytesIO(raw), data_only=True)
    ws = wb2[SHEET_PFV_EDIT]
    mon_course = ws.cell(row=2, column=PFV_EDIT_COL_INDEX["mon_course"] + 1).value
    assert mon_course == "津A"


# ---------------------------------------------------------------------------
# 7. Excel import が sub_office_code を解決して sub_office_id を埋める
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_excel_import_resolves_sub_office_code(client, db) -> None:
    """sub_office_code=TSUGA の行を import すると PFV.sub_office_id=都賀.id に解決される."""
    from openpyxl import Workbook

    from app.services.patient_excel.schema import (
        PATIENT_COL_INDEX,
        PATIENT_COLUMNS,
        PFV_COLUMNS,
        SHEET_PATIENTS,
    )

    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    patient = await _make_patient(db, code="P-IMP-SUB", primary_office_id=inage.id)

    # Workbook を組み立てる: PFV シートに sub_office_code=TSUGA を入れる
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PATIENTS
    for col_idx, col_def in enumerate(PATIENT_COLUMNS, start=1):
        ws_p.cell(row=1, column=col_idx, value=str(col_def["header"]))
    # 既存 patient を書く
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["patient_id"] + 1, value=str(patient.id))
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["patient_code"] + 1, value=patient.code)
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["name"] + 1, value=patient.name)
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["sex"] + 1, value="male")
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["status"] + 1, value="active")
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["address"] + 1, value="千葉市")
    ws_p.cell(row=2, column=PATIENT_COL_INDEX["office_code"] + 1, value="INAGE")

    ws_f = wb.create_sheet(title=SHEET_PFV)
    for col_idx, col_def in enumerate(PFV_COLUMNS, start=1):
        ws_f.cell(row=1, column=col_idx, value=str(col_def["header"]))
    ws_f.cell(row=2, column=PFV_COL_INDEX["patient_id"] + 1, value=str(patient.id))
    ws_f.cell(row=2, column=PFV_COL_INDEX["weekday"] + 1, value="月")
    ws_f.cell(row=2, column=PFV_COL_INDEX["slot_index"] + 1, value=0)
    ws_f.cell(row=2, column=PFV_COL_INDEX["mode"] + 1, value="normal")
    ws_f.cell(row=2, column=PFV_COL_INDEX["time_type"] + 1, value="固定")
    ws_f.cell(row=2, column=PFV_COL_INDEX["start_time"] + 1, value="09:00")
    ws_f.cell(row=2, column=PFV_COL_INDEX["duration_min"] + 1, value=30)
    ws_f.cell(row=2, column=PFV_COL_INDEX["sub_office_code"] + 1, value="TSUGA")

    buf = BytesIO()
    wb.save(buf)

    _patient_rows, pfv_rows_diff, _summary, _patient_ops, pfv_ops = await parse_and_diff(
        db, file_bytes=buf.getvalue()
    )
    # PFV op が 1 件 (new) で sub_office_id=tsuga.id
    new_ops = [op for op in pfv_ops if op.get("_op") == "new"]
    assert len(new_ops) == 1
    assert new_ops[0]["sub_office_id"] == tsuga.id


# ---------------------------------------------------------------------------
# 8. Excel round-trip (export → import) で 0 エラー
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_pfv_excel_round_trip_with_sub_office_no_errors(client, db) -> None:
    """sub_office_id 入り PFV を export → import して error 0 件."""
    inage = await _make_office(db, code="INAGE", name="稲毛")
    tsuga = await _make_office(db, code="TSUGA", name="都賀")
    patient = await _make_patient(
        db,
        code="P-RT-SUB",
        primary_office_id=inage.id,
        lat=35.65,
        lng=140.10,
    )
    pfv = PatientFixedVisit(
        patient_id=patient.id,
        mode="normal",
        weekday=0,
        slot_index=0,
        start_time=time(9, 0),
        duration_min=30,
        sub_office_id=tsuga.id,
    )
    db.add(pfv)
    await db.commit()

    patients_rows = (await db.scalars(select(Patient))).all()
    pfv_rows = (await db.scalars(select(PatientFixedVisit))).all()
    office_rows = (await db.scalars(select(Office))).all()
    ct_rows = (await db.scalars(select(CourseTemplate))).all()

    wb = build_workbook(
        patients=patients_rows,
        pfvs=pfv_rows,
        offices=office_rows,
        course_templates=ct_rows,
    )
    raw = workbook_to_bytes(wb)

    _patient_rows_diff, pfv_rows_diff, summary, _patient_ops, _pfv_ops = await parse_and_diff(
        db, file_bytes=raw
    )
    assert summary.patients_error == 0, "patient 行に error が出てはいけない"
    assert summary.pfv_error == 0, "PFV 行に error が出てはいけない"
    # 再 import なので noop または unchanged になるはず
    new_or_update = [r for r in pfv_rows_diff if r.operation in ("new", "update")]
    assert len(new_or_update) == 0, (
        f"round-trip で new/update 操作が発生してはいけない: {[r.operation for r in pfv_rows_diff]}"
    )
